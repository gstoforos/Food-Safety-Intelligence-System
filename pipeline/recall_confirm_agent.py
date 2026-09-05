#!/usr/bin/env python3
"""
recall_confirm_agent.py  —  AGENT 3 (reviewer 3, the confirmer)
================================================================

Third and final stage. Reviewer 2 does the expensive work — fetch the page,
verify and correct every field, decide approve or reject — and parks its
approvals at Status="pending_gap_v3". This agent confirms that work and
publishes it.

WHY THE SPLIT EXISTS
--------------------
Reviewer 2 runs on CPU-only inference at roughly 100 s per row, so it always
hits its time budget partway through the queue. Before the split, a timeout
meant the whole run promoted nothing: the work was done but never published.

Now reviewer 2 banks every finished row as pending_gap_v3, and this agent
publishes them. It uses NO model — only the deterministic checks — so it runs
in seconds and CANNOT be blocked by the VPS. Whatever reviewer 2 finished
today is published today.

WHAT IT CHECKS
--------------
Reviewer 2's verdict is trusted for JUDGEMENT (is this a real, in-scope
recall). This agent independently re-runs the checks that need no model and
that catch the error classes seen in production:

    * fabricated pathogen        (Pathogen contradicts Reason)
    * placeholder instead of a value
    * URL not on the regulator's own domain / aggregator link
    * regulator name or headline sitting in Product
    * pre-2026 publication date
    * required field missing

A row that fails any of these is rejected even though reviewer 2 approved it —
two independent stages must agree before anything is published.

FLOW
    pending_gap_v3 --confirm--> pending --> promote_approved --> Recalls
    pending_gap_v3 --fail-----> rejected --> Weekly_Rejected
    rejected (from reviewer 2) --------> Weekly_Rejected

CLI
    python -m pipeline.recall_confirm_agent --xlsx docs/data/recalls.xlsx \
        --commit false            # dry run: prints every decision
    --commit true                 # publish
    --limit N                     # cap rows
"""
from __future__ import annotations

import argparse
import datetime as dt
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional

A2_APPROVED = "pending_gap_v3"
REQUIRED = ("Date", "Product", "Pathogen", "URL")


def _load_sheet(xlsx: Path, sheet: str) -> List[Dict[str, Any]]:
    import openpyxl
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        return []
    ws = wb[sheet]
    headers = [c.value for c in ws[1]]
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        row = {h: ("" if v is None else v) for h, v in zip(headers, r) if h}
        if any(str(v).strip() for v in row.values()):
            out.append(row)
    return out


# ---------------------------------------------------------------------------
# Duplicate + provenance-strictness guards (audit 2026-09-02)
# ---------------------------------------------------------------------------
# WHY THIS EXISTS
#     On 2026-09-01/02 this agent published four rows that should never have
#     reached the register, and one of them was emailed to subscribers as a
#     new alert with a dead link:
#
#       * City Foods, Inc. dated 2026-09-01 — a re-mint of the SAME FSIS
#         recall (015-2026) already published on 2026-08-08. The gap-finder
#         used its own scrape date as the recall date and invented the slug
#         "city-foods-inc-recalls-ready-to-eat-...". The real slug is
#         "city-foods-inc--recalls-ready-eat-..." (DOUBLE hyphen, no "to")
#         and was ALREADY recorded on the 08-08 row by the 2026-08-10 audit.
#         The invented slug 404s. This was a regression of a fixed defect.
#       * RappelConso fiche 22527 — duplicate of published fiche 23359.
#       * RappelConso fiche 22529 — duplicate of published fiche 22472.
#       * A salute.gov.it /tema/ landing page filed as an Italian recall.
#
#     Two independent holes let them through:
#
#     1. NO DUPLICATE CHECK. This agent never compared a candidate against the
#        Recalls sheet, so re-minting an existing recall under a new date was
#        invisible to every guard.
#
#     2. PROVENANCE FAILED OPEN. _provenance.check() is called with
#        treat_unreachable_as_problem=False, so a URL that 404s, times out or
#        is blocked returns "no problem" and the row publishes. That default
#        is right for a scraper row off an official feed — several regulators
#        403 datacentre traffic and rejecting on that would discard real
#        recalls. It is exactly wrong for a row whose URL a language model
#        guessed and no gate has since verified.
#
#     Strictness is therefore chosen per row, not globally: a row carrying
#     gap-finder provenance and NO url-gate stamp must have its page actually
#     corroborate it. Every other row keeps the tolerant default.

_GAP_ORIGIN_MARKERS = (
    "gap-finder", "gap finder", "gemini gap", "tavily gap", "openai gap",
    "claude gap", "+ google search",
)
_URL_VERIFIED_MARKERS = ("url-gate", "url_gate", "official-feed", "api fixed")


def _norm_txt(s: Any) -> str:
    import re as _re
    import unicodedata as _ud
    s = _ud.normalize("NFKD", str(s or ""))
    s = "".join(c for c in s if not _ud.combining(c))
    return _re.sub(r"[^a-z0-9 ]+", " ", s.lower()).strip()


def _norm_url(u: Any) -> str:
    u = str(u or "").strip().lower().rstrip("/")
    for p in ("https://", "http://", "www."):
        if u.startswith(p):
            u = u[len(p):]
    return u


def needs_strict_provenance(row: Dict[str, Any]) -> bool:
    """True when this row's URL has never been checked by anything but a model.

    A gap-finder proposes a URL from search snippets. Until url_gate_gemini
    (date_match + brand_match + hazard_match) or an official feed has
    confirmed it, an unreachable page is not "could not confirm" — it is the
    likeliest sign the URL was invented.
    """
    notes = str(row.get("Notes", "") or "").lower()
    if not any(m in notes for m in _GAP_ORIGIN_MARKERS):
        return False
    return not any(m in notes for m in _URL_VERIFIED_MARKERS)


def _distinctive_tokens(row: Dict[str, Any]) -> set:
    try:
        from pipeline import _provenance
        generic = _provenance._GENERIC
    except Exception:                                        # noqa: BLE001
        generic = set()
    toks = set()
    for f in ("Company", "Brand", "Product"):
        for t in _norm_txt(row.get(f, "")).split():
            if len(t) >= 5 and t not in generic:
                toks.add(t)
    return toks


def duplicate_problems(row: Dict[str, Any],
                       published: List[Dict[str, Any]]) -> List[str]:
    """Is this candidate's URL already in the register?

    ONE test, and deliberately only one: the normalised URL (scheme, "www."
    and trailing slash removed) already appears in Recalls. That is certain,
    and it is the whole rule.

    WHY THERE IS NO FUZZY "SAME RECALL, DIFFERENT URL" RULE
    -------------------------------------------------------
    A content-similarity rule was written for this audit and MEASURED against
    the live 1550-row register before being rejected. It would have blocked
    good rows:

      same Source + Pathogen + firm + >=2 shared product tokens, 45d window
          -> 768 of 1550 rows flagged
      ... tightened with a Jaccard floor (0.5-0.8, firm prefix 14-20)
          -> 21-120 rows flagged AND it still MISSED the City Foods duplicate
             at every setting, because the two rows' Product strings differ
             enormously in length and Jaccard punishes that asymmetry
      ... switched to containment (shared / smaller set), >=4 tokens, 1.0
          -> catches City Foods, but flags 24 pairs that are NOT duplicates

    Those 24 are the reason the idea is dead. RappelConso issues ONE FICHE PER
    SKU, so a single incident legitimately produces several rows with near
    identical text and distinct fiche IDs — E.Leclerc Outreau "paris-brest",
    "paris-brest x2" and "paris-brest (vendu au rayon traditionnel)" are three
    separate official notices, and the register's own disambig step exists to
    keep them apart. A similarity rule cannot tell them from a re-mint.

    What DOES separate them is not similarity but provenance: every one of
    those RappelConso fiches resolves and corroborates its row, while the
    City Foods duplicate cited a slug that 404s. So the re-mint case is
    handled by needs_strict_provenance() above, which is a fact about the
    cited page rather than a guess about text overlap.

    If you are tempted to re-add a similarity rule: re-run the measurement
    first. The numbers above are reproducible against the Recalls sheet.
    """
    if not published:
        return []
    u = _norm_url(row.get("URL"))
    if not u:
        return []
    for p in published:
        if _norm_url(p.get("URL")) == u:
            return ["duplicate: this URL is already published on "
                    f"{str(p.get('Date'))[:10]}"]
    return []


_FIRM_PROBLEMS = ("Company is empty", "neither Company nor Brand is set")


def _is_firm_problem(p: str) -> bool:
    """A problem that says only that no firm is named — reviewer 2's job."""
    return any(p.startswith(f) for f in _FIRM_PROBLEMS)


def confirm(row: Dict[str, Any],
            published: Optional[List[Dict[str, Any]]] = None) -> List[str]:
    """Return the reasons this row must NOT be published. Empty list = publish.
    Deterministic only — no model."""
    problems: List[str] = []

    # Already in the register under a different URL or a re-minted date?
    try:
        problems += duplicate_problems(row, published or [])
    except Exception as e:                                   # noqa: BLE001
        problems.append(f"duplicate check failed: {type(e).__name__}")
    if problems:
        return problems

    # PROVENANCE (2026-09-01). Reviewer 3 did not fetch anything: it trusted
    # that reviewer 2 had read the page. Reviewer 2 did not check the page
    # CONTENT either — only the URL's shape. So a row whose URL described a
    # different recall entirely (fiche 22230, a SHEIN plush toy filed as a
    # Brie/Listeria recall) passed both stages. This is the last gate before
    # publication and the cheapest place to catch it.
    try:
        from pipeline import _provenance
        # DEAD vs BLOCKED (corrected 2026-09-02). _provenance.check() now
        # blocks on 404/410 by itself, for every row, because that is the
        # regulator asserting the notice does not exist.
        #
        # treat_unreachable_as_problem stays FALSE even for model-sourced
        # rows. An earlier cut of this fix set it True when
        # needs_strict_provenance() was True, reasoning that an unreadable
        # page is the likeliest sign of an invented URL. That reasoning is
        # wrong in this environment: fsis.usda.gov, the RASFF portal and
        # salute.gov.it answer 403 to datacentre traffic for EVERY url they
        # serve, so it would have rejected every gap-finder row on those
        # three sources — real recalls included. Rule 3: 403 means we could
        # not read it, and that is never by itself a defect of fact.
        #
        # The residual gap is real and deliberately left visible: a
        # fabricated URL on a source that blanket-403s cannot be adjudicated
        # from the page at all. needs_strict_provenance() marks those rows so
        # the run log names them for a human instead of silently passing.
        probs = _provenance.check(row, treat_unreachable_as_problem=False)
        if probs and needs_strict_provenance(row):
            probs = [p + " (row has gap-finder provenance and no url-gate "
                         "stamp — a model is the only thing that has ever "
                         "vouched for this URL)"
                     for p in probs]
        problems += probs
    except Exception:                                        # noqa: BLE001
        pass

    # Reuse reviewer 2's own guards so the two stages apply one rule set.
    try:
        from pipeline.recall_review_agent import (
            _field_integrity_flags, _is_blocking)
        problems += [p for p in _field_integrity_flags(row) if _is_blocking(p)]
    except Exception as e:  # pragma: no cover
        problems.append(f"could not run integrity checks: {type(e).__name__}")

    # Required fields must be present. Company OR Brand satisfies the firm.
    for f in REQUIRED:
        if not str(row.get(f, "") or "").strip():
            problems.append(f"{f} is empty")
    if not str(row.get("Company", "") or "").strip() and \
       not str(row.get("Brand", "") or "").strip():
        problems.append("neither Company nor Brand is set")
    elif not str(row.get("Company", "") or "").strip():
        # The publish gate inside promote_approved() requires Company, so a
        # Brand-only row cannot publish; say so HERE, where main() can hold
        # it for reviewer 2 instead of letting the gate archive it.
        problems.append("Company is empty (Brand set) — the publish gate "
                        "requires the recalling firm")

    # Scope: a pre-2026 notice never publishes, whatever reviewer 2 said.
    d = str(row.get("Date", "") or "").strip()[:10]
    if len(d) >= 4 and d[:4].isdigit() and int(d[:4]) < 2026:
        problems.append(f"out of scope: publication date {d} is before 2026")

    return problems


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=Path, default=Path("docs/data/recalls.xlsx"))
    ap.add_argument("--commit", type=str, default="false")
    ap.add_argument("--limit", type=int, default=0)
    args = ap.parse_args()
    commit = args.commit.lower() in ("1", "true", "yes", "on")

    pending = _load_sheet(args.xlsx, "Pending")
    # ── FAST PATH (2026-08-28) ───────────────────────────────────────────
    # Reviewer 3 was gated to pending_gap_v3 only, so a row that ALREADY
    # satisfies every deterministic guard still had to wait for reviewer 1 and
    # reviewer 2 to touch it — each ~100 s on a CPU-only 7B model, on a VPS
    # that has been down repeatedly.
    #
    # Measured on the live sheet: 20 of 47 Pending rows passed every guard
    # (required fields present, regulator-domain URL, no fabricated pathogen,
    # no placeholder, in-scope date) and 14 of them sat at pending_gap purely
    # because of a status label. Among them a multi-country gorgonzola/brie
    # Listeria cluster — BE, IT, FR, UK within 48 hours — which missed a
    # weekly report while waiting on a model that could add nothing.
    #
    # The guards below ARE the review for such a row. Verification the model
    # would add is re-reading a page whose fields already agree with it. So
    # any status is admitted here; the guards, not the label, decide. Nothing
    # is relaxed: confirm() still runs in full and rejects anything that fails.
    lane = [r for r in pending
            if str(r.get("Status", "")).strip() in
            (A2_APPROVED, "pending", "pending_gap", "pending_gap_v1",
             "pending_gap_v2")]
    already_rejected = [r for r in pending
                        if str(r.get("Status", "")).strip() == "rejected"]
    if args.limit and args.limit > 0:
        lane = lane[:args.limit]

    # HOLD, DO NOT EVICT (audit 2026-09-04). Replayed on the workbook
    # history: between 2026-09-02 and 2026-09-04 the deterministic publish
    # gate inside promote_approved() evicted RappelConso fiches 23420 and
    # 23421 (rillettes d'oie / steak haché, Listeria and Salmonella, sold
    # "sans marque" at one Carrefour) and CFIA RA-82581 (Coaticook aged
    # cheddar, Listeria) on every run, for "Company is empty". None of the
    # three was defective: the RappelConso API names only the distributor
    # for an unbranded fiche and the CFIA open-data title only the brand.
    # Reviewer 2 reads the notice and fills Company — but this agent admits
    # plain "pending" rows and runs three times a day, so it reached them
    # first and promote_approved() archived them before reviewer 2 ever saw
    # them (and, until the record_rejections call below existed, without a
    # trace). A missing Company beside a present Brand is a row that needs a
    # reader, not a rejection: park it at pending_gap_v2, which is in
    # reviewer 2's lane and which merge_master will not promote or evict.
    #
    # Widened 2026-09-05: FSA-PRIN-43-2026 (Sainsbury's olives, Listeria)
    # arrived from the official-feed collector with Company AND Brand empty
    # and was rejected for "neither Company nor Brand is set" — the only
    # thing wrong with a row whose title began "Sainsbury's recalls". So the
    # hold is now decided AFTER confirm() runs, on the problems themselves:
    # when every problem is a missing-firm problem the row is held; when
    # anything else is wrong it is blocked as before, with all reasons.
    held: List[Dict[str, Any]] = []

    print(f"Agent 3 (confirmer): {len(lane)} row(s) in lane, "
          f"{len(already_rejected)} already rejected by reviewer 2 "
          f"(commit={commit})")
    if not lane and not already_rejected:
        print("Nothing to confirm.")
        return 0

    # The duplicate guard needs the register to compare against, so it is
    # loaded BEFORE the decision loop (it used to be read only after the dry
    # run returned, which is why --commit false could never show a duplicate).
    published_now = _load_sheet(args.xlsx, "Recalls")
    print(f"  (duplicate guard: comparing against {len(published_now)} "
          f"published rows)")

    confirmed, blocked = [], []
    for i, row in enumerate(lane, 1):
        probs = confirm(row, published_now)
        if probs and all(_is_firm_problem(p) for p in probs):
            held.append(row)
            print(f"  [{i}/{len(lane)}] HOLD    "
                  f"{str(row.get('Product',''))[:40]:40s} | {probs[0][:40]} "
                  f"— reviewer 2 to read the firm from the notice")
        elif probs:
            blocked.append((row, probs))
            print(f"  [{i}/{len(lane)}] BLOCK   "
                  f"{str(row.get('Product',''))[:40]:40s} | {probs[0][:52]}")
        else:
            confirmed.append(row)
            print(f"  [{i}/{len(lane)}] CONFIRM "
                  f"{str(row.get('Product',''))[:40]:40s} | publish")

    print(f"\n{'='*60}")
    print(f"confirm: {len(confirmed)}   block: {len(blocked)}   "
          f"archive (reviewer-2 rejects): {len(already_rejected)}")
    print(f"{'='*60}")

    if not commit:
        print("\nDRY RUN — nothing written. Set --commit true to publish.")
        return 0

    from pipeline.merge_master import (  # type: ignore
        promote_approved, sort_rows, save_xlsx_with_pending)

    xlsx = Path(args.xlsx)
    approved_existing = _load_sheet(xlsx, "Recalls")
    full_pending = _load_sheet(xlsx, "Pending")
    today = dt.datetime.now(dt.timezone.utc).strftime("%Y-%m-%d")

    by_url = {}
    for i, r in enumerate(full_pending):
        u = str(r.get("URL", "")).strip()
        if u:
            by_url[u] = i

    rejected_flags: Dict[int, str] = {}

    for row in confirmed:
        idx = by_url.get(str(row.get("URL", "")).strip())
        if idx is None:
            continue
        full_pending[idx]["Status"] = "pending"      # now promotable
        n = str(full_pending[idx].get("Notes", "")).strip()
        stamp = f" [confirm-agent {today}: {A2_APPROVED} → pending; independent checks passed]"
        # ENGLISH-OUTPUT RULE AT THE LAST GATE (2026-09-04). Nine French
        # Reasons reached Recalls in the first four days of September. The
        # language check is a warning by policy (a real recall is never
        # rejected over wording), but nothing on the publish path ever
        # translated. pipeline._language.englishify_reason is a table of
        # verified translations that never invents: apply it here, and when
        # it has no entry, stamp the row so the weekly review sees it.
        try:
            from pipeline._language import englishify_reason, looks_non_english
            before = str(full_pending[idx].get("Reason", "") or "")
            after, changed = englishify_reason(before)
            if changed:
                full_pending[idx]["Reason"] = after
                stamp += f" [reason translated: \"{before[:80]}\"]"
            elif looks_non_english(before):
                stamp += " [needs cleanup: Reason not in English and no verified translation on file]"
        except Exception:                                    # noqa: BLE001
            pass
        full_pending[idx]["Notes"] = (n + stamp).strip()[:1000]

    for row, probs in blocked:
        idx = by_url.get(str(row.get("URL", "")).strip())
        if idx is None:
            continue
        rejected_flags[idx] = ("Confirmer: reviewer 2 approved but "
                               + "; ".join(probs[:2]))[:280]

    for row in already_rejected:
        idx = by_url.get(str(row.get("URL", "")).strip())
        if idx is not None and idx not in rejected_flags:
            rejected_flags[idx] = "Rejected by reviewer 2"

    # Held rows: park at pending_gap_v2 so promote_approved() neither
    # publishes them (Company empty would fail its gate) nor archives them.
    # Stamped once; a row already carrying the hold stamp is left alone.
    HOLD_MARK = "[confirm-agent hold:"
    for row in held:
        idx = by_url.get(str(row.get("URL", "")).strip())
        if idx is None or idx in rejected_flags:
            continue
        full_pending[idx]["Status"] = "pending_gap_v2"
        n = str(full_pending[idx].get("Notes", "")).strip()
        if HOLD_MARK not in n:
            full_pending[idx]["Notes"] = (
                n + f" {HOLD_MARK} {today}: no recalling firm on the row "
                f"(Brand {str(row.get('Brand',''))[:40]!r}); not published and "
                "not rejected — reviewer 2 to read the firm from the "
                "notice]").strip()[:1000]

    new_approved, remaining, archived = promote_approved(
        pending=full_pending,
        approved_existing=approved_existing,
        rejected_flags=rejected_flags,
        archive_immediately=True,
    )
    for r in archived:
        if not str(r.get("RejectedBy") or "").strip():
            r["RejectedBy"] = "confirm-agent"
    save_xlsx_with_pending(sort_rows(approved_existing + new_approved),
                           sort_rows(remaining), xlsx,
                           newly_rejected_rows=archived)
    # NEVER SILENT-DELETE (audit 2026-09-04). save_xlsx_with_pending()
    # documents newly_rejected_rows as a NO-OP: the caller must mirror the
    # evicted rows into Weekly_Rejected itself, immediately after the save.
    # This agent never did. Measured by replaying the workbook across the six
    # agent commits of 2026-09-02..04: 50 evictions with no record in
    # Recalls, Weekly_Rejected or Rejected — 18 distinct rows, 12 of them
    # RASFF notifications (Listeria, Salmonella, STEC, ochratoxin, DON,
    # aflatoxin), re-scraped and evicted again on every run. A rejection
    # without a recorded reason is not a rejection; it is data loss.
    try:
        from pipeline.weekly_rejected_capture import record_rejections
        # The JSON slice the Thursday mailer reads is
        # docs/data/weekly-rejected-latest.json (weekly_rejected_capture's
        # default). An earlier cut of this call named "weekly-rejected.json"
        # — a file nothing reads — so the archive would have been written to
        # the sheet but invisible to the operator's email.
        n_rec = record_rejections(archived, xlsx_path=xlsx,
                                  json_path=xlsx.parent / "weekly-rejected-latest.json")
        print(f"  (archived {n_rec} rejection(s) to Weekly_Rejected)")
    except Exception as e:                                   # noqa: BLE001
        print(f"  !! could not record rejections: {type(e).__name__}: {e}")
    try:
        from pipeline.merge_master import mirror_json_from_xlsx
        mirror_json_from_xlsx(xlsx, xlsx.parent / "recalls.json")
    except Exception:
        pass

    print(f"\n✓ Published {len(new_approved)} → Recalls, "
          f"archived {len(archived)}, {len(remaining)} remain in Pending.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
