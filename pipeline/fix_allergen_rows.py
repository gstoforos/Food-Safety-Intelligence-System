#!/usr/bin/env python3
"""
fix_allergen_rows.py  —  one-time corrective (2026-07-29)
=========================================================

Removes the two confirmed OUT-OF-SCOPE undeclared-allergen recalls that were
wrongly promoted to Recalls with a false "Listeria monocytogenes" pathogen.
Matched by EXACT URL (not pattern) so nothing else can be touched.

Per policy (2026-07-29): undeclared-allergen recalls are OUT OF SCOPE →
they belong in Weekly_Rejected, not Recalls.

  1. https://www.foodstandards.gov.au/food-recalls/recall-alert/auxico-perth-pty-ltd-lgm-hot-chilli-oil-275g
     (LGM HOT CHILLI OIL 275G — undeclared peanuts)
  2. https://www.foodstandards.gov.au/food-recalls/recall-alert/viet-meatballs-chinese-sausage-500g
     (Chinese Sausage 500g — undeclared gluten)

Usage:
    python fix_allergen_rows.py --xlsx docs/data/recalls.xlsx --commit false
    python fix_allergen_rows.py --xlsx docs/data/recalls.xlsx --commit true
"""
from __future__ import annotations
import argparse
from pathlib import Path

TARGET_URLS = {
    "https://www.foodstandards.gov.au/food-recalls/recall-alert/auxico-perth-pty-ltd-lgm-hot-chilli-oil-275g",
    "https://www.foodstandards.gov.au/food-recalls/recall-alert/viet-meatballs-chinese-sausage-500g",
}
REJECT_REASON = "Out of scope: undeclared allergen (not a microbial pathogen) — policy 2026-07-29"


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=Path, default=Path("docs/data/recalls.xlsx"))
    ap.add_argument("--commit", type=str, default="false")
    args = ap.parse_args()
    commit = args.commit.lower() in ("1", "true", "yes", "on")

    import openpyxl
    wb = openpyxl.load_workbook(args.xlsx)
    ws = wb["Recalls"]
    headers = [c.value for c in ws[1]]
    url_i = headers.index("URL")

    # Find the exact rows (1-based sheet row numbers)
    to_remove = []
    for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        u = str(row[url_i]).strip() if row[url_i] else ""
        if u in TARGET_URLS:
            to_remove.append((ridx, dict(zip(headers, row))))

    print(f"Matched {len(to_remove)} row(s) in Recalls by exact URL:")
    for ridx, rowd in to_remove:
        print(f"  row {ridx}: {str(rowd.get('Product',''))[:45]} | "
              f"Pathogen={rowd.get('Pathogen','')} | {rowd.get('URL','')}")

    if len(to_remove) != 2:
        print(f"\n⚠ Expected 2 rows, found {len(to_remove)}. "
              f"Aborting so nothing wrong is touched.")
        return 1

    if not commit:
        print("\nDRY RUN — nothing changed. Re-run with --commit true to apply.")
        return 0

    # Append to Weekly_Rejected with the correct reason, then delete from Recalls.
    wr = wb["Weekly_Rejected"] if "Weekly_Rejected" in wb.sheetnames else None
    if wr is None:
        print("⚠ Weekly_Rejected sheet not found; aborting.")
        return 1
    wr_headers = [c.value for c in wr[1]]

    def col(name):
        return wr_headers.index(name) if name in wr_headers else None

    for _ridx, rowd in to_remove:
        new = [""] * len(wr_headers)
        for k, v in rowd.items():
            if k in wr_headers:
                new[wr_headers.index(k)] = v
        # Clear the wrong pathogen; set rejection reason fields if present
        if col("Pathogen") is not None:
            new[col("Pathogen")] = ""
        for rc in ("RejectionReason", "Reason"):
            if col(rc) is not None:
                new[col(rc)] = REJECT_REASON
                break
        if col("RejectedBy") is not None:
            new[col("RejectedBy")] = "policy-fix-2026-07-29"
        wr.append(new)

    # Delete from Recalls bottom-up so row indices stay valid
    for ridx, _rowd in sorted(to_remove, key=lambda x: -x[0]):
        ws.delete_rows(ridx, 1)

    wb.save(args.xlsx)
    print(f"\n✓ Moved {len(to_remove)} allergen rows Recalls → Weekly_Rejected.")

    # Mirror JSON if the helper is available
    try:
        import sys
        sys.path.insert(0, ".")
        from pipeline.merge_master import mirror_json_from_xlsx
        json_path = args.xlsx.parent / "recalls.json"
        mirror_json_from_xlsx(args.xlsx, json_path)
        print(f"✓ recalls.json mirrored ({json_path}).")
    except Exception as e:
        print(f"  (JSON mirror skipped: {e}; run your normal mirror step.)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
