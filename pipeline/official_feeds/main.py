"""
Official-feed collector orchestrator.

Usage:
  python -m pipeline.official_feeds.main --source uk --dry-run
  python -m pipeline.official_feeds.main --source ireland --max-age-days 14

Flow per source:
  1. fetch()        — pull structured records from authority API/RSS/HTML
  2. age filter     — keep records published within --max-age-days
  3. classify       — tier via gap_finder rules on the clean hazard text
  4. dedup          — exact match on source_id (authority notation)
  5. xlsx append    — Pending rows in docs/data/recalls.xlsx
"""

from __future__ import annotations

import argparse
from datetime import datetime, timezone

from .base import get, all_codes
from .classify import classify_record


def run_source(code: str, max_age_days: int = 14, dry_run: bool = False,
               xlsx_path: str = "docs/data/recalls.xlsx") -> dict:
    src = get(code)
    print(f"=== AFTS Official-Feed Collector ({code}/{src.name_en}) "
          f"started {datetime.now(timezone.utc).isoformat()} ===")
    print(f"  authority: {src.authority_short}")
    print(f"  xlsx:      {xlsx_path}")
    print(f"  dry-run:   {dry_run}")

    print("\n=== Stage 1: Fetch official feed ===")
    records = src.fetcher()
    print(f"  official records: {len(records)}")
    for rec in records:
        if not rec.region:
            rec.region = src.region

    # Stage 1b: Google News supplement (insurance against flaky official feeds)
    if src.gnews_authority:
        print("\n=== Stage 1b: Google News supplement ===")
        from .gnews import fetch_gnews
        gn = fetch_gnews(
            authority=src.gnews_authority,
            country_code=records[0].country_code if records else src.code[:2],
            country_name=src.name_en,
            authority_short=src.authority_short,
            pathogen_terms=list(src.gnews_terms),
            hl=src.gnews_hl, gl=src.gnews_gl, ceid=src.gnews_ceid,
            days_back=src.gnews_days_back,
            country_keywords=src.gnews_country_keywords,
            country_domains=src.gnews_country_domains,
            block_title_keywords=src.gnews_block_title_keywords,
            use_description=src.gnews_use_description,
            authority_aliases=src.gnews_authority_aliases,
        )
        records = _merge_dedup(records, gn)
        for rec in records:
            if not rec.region:
                rec.region = src.region
        print(f"  combined after merge+dedup: {len(records)}")

    # Stage 1c: Dated web search (numeric-date queries via Searx). Catches
    # zero-day recalls the official API hasn't batched yet, and pulls authority
    # URLs straight from results. Self-gates: no-op unless Searx is configured.
    from .agents import searx_search as _sx
    if src.authority_domain and _sx.is_configured():
        print("\n=== Stage 1c: Dated web search ===")
        from .dated_search import fetch_dated_search
        ds = fetch_dated_search(
            authority_short=src.authority_short,
            authority_domain=src.authority_domain,
            country_code=records[0].country_code if records else src.code[:2],
            country_name=src.name_en,
            region=src.region,
            days_back=max(src.gnews_days_back, 3),
        )
        if ds:
            records = _merge_dedup(records, ds)
            for rec in records:
                if not rec.region:
                    rec.region = src.region
            print(f"  combined after dated-search merge: {len(records)}")

    print("\n=== Stage 2: Age filter ===")
    now = datetime.now(timezone.utc)
    kept = []
    too_old = no_date = 0
    for rec in records:
        age = rec.age_days(now)
        if age is None:
            # Undated record: no parseable publication date. Previously kept
            # on the assumption the listing was recent, but that put blank-date
            # rows into recalls.xlsx and the alert email. Drop them — a real
            # recall always has a date, and per-source fetchers now supply a
            # listing-level fallback before giving up.
            no_date += 1
            continue
        elif age <= max_age_days:
            kept.append(rec)
        else:
            too_old += 1
    print(f"  kept: {len(kept)}  (too old: {too_old}, dropped no-date: {no_date})")

    print("\n=== Stage 3: Classify ===")
    accepted = []
    rejected = 0
    for rec in kept:
        c = classify_record(rec)
        tag = f"{c['verdict']}/{c['category']}"
        tier = c["tier"]
        short = (rec.title or "")[:60]
        print(f"  [{tag} tier={tier} m={c['matched']!r}] {short}")
        if c["verdict"] == "accept":
            rec.raw["_tier"] = tier
            rec.raw["_category"] = c["category"]
            rec.raw["_matched"] = c["matched"]
            accepted.append(rec)
        else:
            rejected += 1
    print(f"  accepted: {len(accepted)}, rejected: {rejected}")

    # ── Stage 3b: Resolve authority URLs for GNews-surfaced accepts ──
    # For every accepted record whose source_id starts with "GN-" (came
    # from Google News, not the official feed), follow the news article
    # and try to extract the regulator's own URL from the article body.
    # If found, rec.url is REPLACED with the authority URL — that's the
    # one stored in Pending. If not found, rec.url keeps the news URL
    # but is flagged so the operator can resolve manually.
    if accepted and src.authority_domain:
        print("\n=== Stage 3b: Resolve authority URLs ===")
        import re as _re
        pat = _re.compile(src.authority_url_pattern) if src.authority_url_pattern else None

        # Route to the market-specific AI agent if this source is wired to one
        use_agent = bool(src.market_agent and src.regulator_code)
        if use_agent:
            if src.market_agent == "north_america":
                from .agents.north_america import find_url as agent_find_url
                print(f"  [Agent] AFTS North America Recall Agent "
                      f"(regulator: {src.regulator_code})")
            else:
                # Unknown market agent — fall back to legacy path
                use_agent = False
                print(f"  [WARN] unknown market_agent {src.market_agent!r}, "
                      f"falling back to legacy resolver")

        if not use_agent:
            from .url_resolver import resolve_authority_url

        resolved = unresolved = 0
        for rec in accepted:
            if not rec.source_id.startswith("GN-"):
                continue   # official-feed records already have authority URL
            news_url = rec.url
            pub_iso = (rec.published.strftime("%Y-%m-%d")
                       if rec.published else "")

            if use_agent:
                auth_url = agent_find_url(rec.title, src.regulator_code)
            else:
                auth_url = resolve_authority_url(
                    news_url, rec.title, src.authority_domain, pat,
                    bulk_index_queries=src.bulk_index_queries,
                    authority_short=src.authority_short,
                    published_iso=pub_iso)

            if auth_url:
                rec.raw["_news_url"] = news_url   # keep original for audit
                rec.url = auth_url
                resolved += 1
                print(f"  ✓ {auth_url[:90]}")
            else:
                rec.raw["_news_url"] = news_url
                rec.raw["_pending_no_auth_url"] = True
                unresolved += 1
                print(f"  ✗ no {src.authority_domain} URL — "
                      f"{rec.title[:60]}")
        print(f"  resolved: {resolved}, unresolved (kept news URL + flag): {unresolved}")

    # ── Stage 3c: Authority / news-mirror / recency guard ───────────────
    # Run every accepted record through the shared allowlist guard so
    # non-regulator URLs (facebook.com and other news mirrors) are dropped
    # before Stage 4 writes them to xlsx. stdlib-only import; wrapped so a
    # guard failure can never crash the collector.
    try:
        from pipeline._gap_finder_guards import check_gap_finder_row
    except Exception as _e:      # pragma: no cover
        check_gap_finder_row = None
        print(f"  [WARN] gap-finder guard unavailable ({_e}); skipping Stage 3c")
    if check_gap_finder_row is not None and accepted:
        print("\n=== Stage 3c: Authority / news / recency guard ===")
        guarded, dropped = [], 0
        for rec in accepted:
            guard_row = {
                "URL": rec.url,
                "Date": (rec.published.strftime("%Y-%m-%d")
                         if rec.published else ""),
                "Company": rec.company,
                "Product": rec.product or rec.title,
            }
            ok, reason, note = check_gap_finder_row(guard_row)
            if note:
                print(f"  [note] {note}")
            if ok:
                guarded.append(rec)
            else:
                dropped += 1
                print(f"  ✗ dropped [{rec.source_id}] {reason} — "
                      f"{(rec.company or rec.title)[:50]}")
        print(f"  kept: {len(guarded)}, dropped: {dropped}")
        accepted = guarded

    print("\n=== Stage 4: Write to xlsx ===")
    appended = skipped = 0
    if dry_run:
        print(f"  DRY RUN — would append {len(accepted)} pending")
        for rec in accepted:
            print(f"    + [{rec.source_id}] {rec.company} | {rec.product or rec.title[:40]} "
                  f"| tier {rec.raw.get('_tier')}")
    else:
        appended, skipped = _append_xlsx(accepted, xlsx_path)
        print(f"  appended: {appended}, skipped (dupe): {skipped}")

    print("\n" + "=" * 60)
    print("RUN SUMMARY")
    print("=" * 60)
    print(f"  source:            {code} ({src.name_en})")
    print(f"  total candidates:  {len(records)}")
    print(f"  after age filter:  {len(kept)}")
    print(f"  accepted:          {len(accepted)}")
    print(f"  rejected:          {rejected}")
    print(f"  → appended:        {appended}")
    print(f"  → skipped (dupe):  {skipped}")
    print("=" * 60)
    print(f"✓ {code} completed\n")
    return {"candidates": len(records), "accepted": len(accepted),
            "appended": appended, "skipped": skipped}


def _merge_dedup(official: list, gnews: list) -> list:
    """
    Merge official + Google-News (and dated-search) records. Official records
    win. A non-official record is dropped if it duplicates an official record by
    EITHER:
      (1) near-duplicate title (Jaccard >= 0.5 on word sets), OR
      (2) same pathogen AND >= 2 shared distinctive tokens (audit 2026-06-22).

    Path (2) folds news rows into the matching openFDA enforcement record even
    when the headline wording differs — e.g. the "Alfredo sauce / Salmonella"
    headlines collapse into the openFDA "Coffee Connexion / Alfredo Sauce" row
    (FDA issued NO press release for that recall, so openFDA is its only home).
    The >=2-token requirement prevents false merges of distinct same-pathogen
    recalls that share only a generic food word (Clover Hill vs Ambriola both
    being "cheese / Listeria").
    """
    def words(s: str) -> set:
        return {w for w in "".join(
            c.lower() if c.isalnum() else " " for c in (s or "")
        ).split() if len(w) > 2}

    _PATHO = ("salmonella", "listeria", "botulism", "botulinum", "cronobacter",
              "hepatitis", "norovirus", "cyclospora", "campylobacter",
              "shigella", "vibrio", "cereulide", "cereus")

    def pathogen(s: str) -> str:
        t = (s or "").lower()
        if "coli" in t or "e. coli" in t or "escherichia" in t:
            return "ecoli"
        for p in _PATHO:
            if p in t:
                return "botulism" if p in ("botulism", "botulinum") else p
        return ""

    # Words too generic to distinguish one recall from another.
    _STOP = {
        "recall", "recalls", "recalled", "recalling", "fda", "usda", "fsis",
        "cfia", "due", "over", "sold", "states", "state", "class", "highest",
        "risk", "level", "food", "foods", "products", "product", "brand",
        "company", "inc", "llc", "due", "because", "potential", "possible",
        "contamination", "contaminated", "alert", "alerts", "warning", "warns",
        "outbreak", "issues", "announces", "expands", "expand", "upgrades",
        "upgrade", "linked", "health", "public", "consumers", "nationwide",
        "voluntary", "voluntarily", "market", "withdrawal", "safety", "risk",
        "after", "amid", "more", "than", "this", "that", "with", "from",
        "may", "due", "the", "and", "for", "are", "new",
    } | set(_PATHO) | {"ecoli", "coli"}

    def distinctive(*parts) -> set:
        toks = set()
        for s in parts:
            toks |= words(s)
        return {w for w in toks if w not in _STOP and len(w) >= 4}

    official_word_sets = [words(r.title) for r in official]
    official_sig = [
        (pathogen(f"{r.title} {r.hazard} {r.product}"),
         distinctive(r.title, r.product, r.company, r.hazard))
        for r in official
    ]

    merged = list(official)
    for g in gnews:
        gw = words(g.title)
        if not gw:
            continue
        dup = False
        # (1) title near-duplicate
        for ow in official_word_sets:
            if ow and len(gw & ow) / len(gw | ow) >= 0.5:
                dup = True
                break
        # (2) same pathogen + >= 2 shared distinctive tokens
        if not dup:
            gp = pathogen(f"{g.title} {g.hazard}")
            gt = distinctive(g.title, g.hazard)
            if gp and len(gt) >= 2:
                for op, ot in official_sig:
                    if op == gp and len(gt & ot) >= 2:
                        dup = True
                        break
        if not dup:
            merged.append(g)
    return merged


def _append_xlsx(records, xlsx_path):
    """
    Append accepted records as rows on the 'Pending' sheet, matching the
    exact AFTS schema. Dedup on URL against BOTH Pending and Recalls sheets
    (don't re-add something already pending or already promoted).
    """
    import openpyxl
    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except FileNotFoundError:
        print(f"  [WARN] {xlsx_path} not found — skipping write")
        return 0, 0

    if "Pending" not in wb.sheetnames:
        print("  [WARN] no 'Pending' sheet — skipping write")
        return 0, 0
    ws = wb["Pending"]
    header = [c.value for c in ws[1]]

    # Collect existing URLs from Pending + Recalls for dedup
    existing_urls = set()
    for sheet_name in ("Pending", "Recalls"):
        if sheet_name not in wb.sheetnames:
            continue
        sh = wb[sheet_name]
        sh_header = [c.value for c in sh[1]]
        try:
            url_idx = sh_header.index("URL")
        except ValueError:
            continue
        for row in sh.iter_rows(min_row=2):
            v = row[url_idx].value
            if v:
                existing_urls.add(str(v).strip())

    appended = skipped = 0
    for rec in records:
        if rec.url and rec.url.strip() in existing_urls:
            skipped += 1
            continue
        ws.append(_row_for(rec, header))
        existing_urls.add(rec.url.strip() if rec.url else "")
        appended += 1
    if appended:
        wb.save(xlsx_path)
    return appended, skipped


_CLASS_MAP = {"recall": "Recall", "allergy": "Alert", "action": "Alert"}

_PATHOGEN_DISPLAY = {
    "salmonella": "Salmonella", "salmonell": "Salmonella",
    "listeria": "Listeria", "listeria monocytogenes": "Listeria monocytogenes",
    "stec": "STEC", "e.coli": "E. coli", "e. coli": "E. coli",
    "cereulide": "Cereulide (Bacillus cereus)",
    "botulism": "Clostridium botulinum", "clostridium": "Clostridium botulinum",
    "hepatitis": "Hepatitis", "norovirus": "Norovirus",
    "campylobacter": "Campylobacter",
    "aflatoxin": "Aflatoxin", "aflatoksiini": "Aflatoxin",
    "ochratoxin": "Ochratoxin", "okratoksiini": "Ochratoxin",
    "hometoksiini": "Mycotoxin", "mycotoxin": "Mycotoxin",
}


def _pathogen_name(rec) -> str:
    """Human-readable pathogen for the Pathogen column."""
    matched = (rec.raw.get("_matched") or "").lower().strip()
    if matched in _PATHOGEN_DISPLAY:
        return _PATHOGEN_DISPLAY[matched]
    if matched:
        return matched.title()
    cat = rec.raw.get("_category", "")
    return "Mycotoxin" if cat == "microbial_toxin" else "Pathogen"


def _row_for(rec, header):
    """Map a Record to a row in the exact order of the Pending sheet header."""
    now_iso = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    prov = []
    if rec.source_id:
        prov.append(f"source_id={rec.source_id}")
    if rec.raw.get("gnews_query"):
        prov.append(f"[via Google News: {rec.raw['gnews_query']}]")
    else:
        prov.append("[via official-feed collector]")
    # Audit trail: when URL resolution rewrote the URL from news →
    # authority, keep the original news URL in Notes so the operator can
    # cross-check during review.
    if rec.raw.get("_news_url") and rec.raw["_news_url"] != rec.url:
        prov.append(f"[news_src={rec.raw['_news_url']}]")
    notes = " ".join(prov)

    # Status flag: if URL resolution found no authority URL for a GNews
    # accept, mark the row so the operator can manually find the regulator
    # link before promotion. Pipeline correctly identified the recall but
    # couldn't resolve the authoritative URL — needs human assist.
    status = ("pending_no_auth_url" if rec.raw.get("_pending_no_auth_url")
              else "pending_gap")

    mapping = {
        "Date": rec.published.strftime("%Y-%m-%d") if rec.published else "",
        "Source": rec.authority,
        "Company": rec.company,
        "Brand": rec.company,
        "Product": rec.product or rec.title,
        "Pathogen": _pathogen_name(rec),
        "Reason": rec.hazard or rec.title,
        "Class": rec.recall_class or _CLASS_MAP.get(rec.alert_type, "Recall"),
        "Country": rec.country_name,
        "Region": rec.region,
        "Tier": rec.raw.get("_tier"),
        "Outbreak": rec.outbreak or 0,
        "URL": rec.url,
        "Notes": notes,
        "ScrapedAt": now_iso,
        "Status": status,
        "RejectedBy": None,
    }
    return [mapping.get(col, "") for col in header]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--source", required=True,
                    help=f"one of: {', '.join(all_codes())}")
    ap.add_argument("--max-age-days", type=int, default=14)
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--xlsx", default="docs/data/recalls.xlsx")
    args = ap.parse_args()
    run_source(args.source, max_age_days=args.max_age_days,
               dry_run=args.dry_run, xlsx_path=args.xlsx)


if __name__ == "__main__":
    main()
