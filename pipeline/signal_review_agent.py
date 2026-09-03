#!/usr/bin/env python3
"""
signal_review_agent.py  —  SIGNAL REVIEW AGENT (deterministic, no model)
=========================================================================

The fourth agent. The three recall agents move rows from Pending to Recalls;
this one reads the register the way an epidemiologist reads a surveillance
board and writes down what the aberration detector cannot say on its own.

WHY IT EXISTS — THE CYCLOSPORA CASE (2026-07-18)
-------------------------------------------------
The register holds ONE Cyclospora record, ever: Taylor Fresh Foods iceberg
lettuce, FDA, Outbreak = 1, Tier 1, published in the week of 13–19 July.
By late August that outbreak was the largest Cyclospora outbreak on record
(11,000+ patients) and the NEWS sheet carried seven headlines about it in
six days. pipeline/signal_detector.py never alarmed and never could:

    * the stratum's 7-week baseline is [0,0,0,0,0,0,0] — mean 0.0, zero
      non-zero weeks — so the sparsity ladder suppresses it before any test
      runs (MIN_BASELINE_MEAN = 1.0, MIN_BASELINE_NONZERO = 3);
    * one record is below MIN_ABSOLUTE_COUNT = 4;
    * with π̂ = 0 the share-channel binomial is degenerate — P[X ≥ 1] under
      Binomial(N, 0) is exactly 0, which is "infinitely significant" and
      therefore meaningless. A p-value needs a baseline to be measured
      against; a pathogen the register has never seen has none.

That is not a bug in the detector. It measures recall-NOTICE counts, and one
notice can carry 11,000 cases. It is the "no denominator" limit stated in
the detector's own docstring, showing up as a NOVELTY BLIND SPOT: the more
unprecedented an event, the less the count channel can see it.

WHAT THIS AGENT ADDS — four things the count channel does not test
------------------------------------------------------------------
  1. NOVELTY CHANNEL.  A pathogen (global stratum) whose guarded baseline is
     entirely zero and which appears this week — reported as NOVEL, with an
     explicit "untestable" label instead of a fabricated p-value. Severity is
     read from the records themselves: Outbreak = 1 and/or Tier 1. A first
     appearance in the register's whole history is marked as such.

  2. OUTBREAK CHANNEL.  Outbreak = 1 is a regulator's declaration, not a
     count. Weekly Outbreak = 1 records vs their own guarded 7-week baseline,
     exact Poisson upper tail. Small numbers, so the tail is reported and
     the reader decides; nothing here goes through FDR.

  3. NEWS CORROBORATION.  The NEWS sheet is a second, independent stream the
     repository already collects (and purges after 7 days). Headlines per
     pathogen for the scanned week are counted and PERSISTED to
     docs/data/news-pulse.jsonl, so a baseline accumulates week by week.
     Until it has one, the count is reported as "no baseline yet".

  4. ALARM REVIEW.  The detector's ledger has a `verdict` column that its
     own docstring says to "fill by hand: genuine | artifact | unclear".
     This agent fills a verdict HINT with a stated reason, from checks a
     reviewer would run first:
        publication cluster   this week's records in the stratum come
                              mostly from ONE firm (one incident, many
                              notices — RappelConso issues one fiche per
                              SKU, so Leclerc Dinan produced 20 fiches)
        single publisher      ≥ 70% of the stratum's records from one source
        volume artefact       count-only channel while corpus C2 ≥ 2
        corroborated          Outbreak = 1 records in the stratum, or a
                              news burst on the pathogen
     A hint is a hint. It is written next to the alarm, never in place of
     the alarm, and the page shows both.

WHAT IT DELIBERATELY DOES NOT DO
--------------------------------
  * It does not alarm. It annotates. The detector's alarm list is unchanged.
  * It does not predict. Every string it writes describes this week.
  * It does not fetch anything. It reads the workbook and the board.

OUTPUT
    docs/data/signals-review.json      this week's review (page reads it)
    docs/data/signals-review.jsonl     append-only history
    docs/data/news-pulse.jsonl         per-week news counts per pathogen

CLI
    python -m pipeline.signal_review_agent                 # latest week
    python -m pipeline.signal_review_agent --asof 2026-07-18
    python -m pipeline.signal_review_agent --dry-run       # print, write nothing
"""
from __future__ import annotations

import argparse
import json
import math
import os
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import pandas as pd  # noqa: E402

from pipeline import signal_detector as S  # noqa: E402

XLSX = ROOT / "docs" / "data" / "recalls.xlsx"
BOARD = ROOT / "docs" / "data" / "signals-board.json"
OUT = ROOT / "docs" / "data" / "signals-review.json"
OUT_HIST = ROOT / "docs" / "data" / "signals-review.jsonl"
NEWS_PULSE = ROOT / "docs" / "data" / "news-pulse.jsonl"

NEWS_BURST = 3            # headlines on one pathogen in the week = corroboration
CLUSTER_SHARE = 0.60      # one firm supplies this share of a stratum's week
PUBLISHER_SHARE = S.PUBLISHER_DOMINANCE


# ─────────────────────────────────────────────────────────────────────────────
# helpers
# ─────────────────────────────────────────────────────────────────────────────

def _truthy(v: Any) -> bool:
    return str(v).strip().lower() in ("1", "1.0", "true", "yes", "y")


def poisson_sf(k: int, lam: float) -> float:
    """P[X >= k] for X ~ Poisson(lam), exact, summed from the lower tail."""
    if k <= 0:
        return 1.0
    if lam <= 0:
        return 0.0
    cdf = 0.0
    term = math.exp(-lam)
    for i in range(0, k):
        cdf += term
        term *= lam / (i + 1)
    return max(0.0, 1.0 - cdf)


def _week_for(asof: Optional[str], corpus: S.Corpus) -> pd.Period:
    if asof:
        w = pd.Timestamp(asof).to_period("W-SUN")
        if w not in corpus.weeks:
            raise SystemExit(f"signal_review_agent: week {w} not in corpus "
                             f"(latest complete is {corpus.weeks[-1]})")
        return w
    return corpus.weeks[-1]


def _records(corpus: S.Corpus, week: pd.Period) -> pd.DataFrame:
    return corpus.frame[corpus.frame["week"] == week]


# ─────────────────────────────────────────────────────────────────────────────
# 1. novelty channel
# ─────────────────────────────────────────────────────────────────────────────

def novelty(corpus: S.Corpus, strata: Dict[str, S.Stratum],
            week: pd.Period) -> List[Dict]:
    weeks = corpus.weeks
    idx = weeks.index(week)
    df = _records(corpus, week)
    out: List[Dict] = []
    for key, s in strata.items():
        if s.level != "global":
            continue
        observed = int(s.series.get(week, 0))
        if observed < 1:
            continue
        b2 = S._window(s.series, weeks, idx, S.BASELINE_WEEKS, S.GUARD_WEEKS)
        if not b2 or sum(b2) > 0:
            continue                     # has a baseline → the detector's job
        prior = [int(s.series.get(w, 0)) for w in weeks[:idx]]
        first_ever = sum(prior) == 0
        recs = df[df["pathogen_c"] == s.pathogen]
        outbreak = int(sum(_truthy(v) for v in recs["Outbreak"]))
        tier1 = int(sum(str(v).strip() in ("1", "1.0") for v in recs["Tier"]))
        sev = ("outbreak-flagged" if outbreak else
               "tier-1" if tier1 else "no severity flag")
        out.append({
            "pathogen": s.pathogen, "stratum_key": key,
            "observed": observed,
            "baseline": [int(x) for x in b2],
            "first_appearance_ever": first_ever,
            "outbreak_records": outbreak, "tier1_records": tier1,
            "severity": sev,
            "records": [{
                "date": str(r.get("Date"))[:10], "source": str(r.get("Source")),
                "company": str(r.get("Company"))[:80],
                "product": str(r.get("Product"))[:100],
                "country": str(r.get("Country")),
                "outbreak": _truthy(r.get("Outbreak")),
                "tier": str(r.get("Tier")),
            } for _, r in recs.head(6).iterrows()],
            "testable": False,
            "why_untestable": (
                "guarded 7-week baseline is entirely zero: the sparsity "
                "ladder suppresses the stratum (mean < 1/week, no non-zero "
                "weeks) and with pi_hat = 0 the share-channel binomial is "
                "degenerate. A p-value cannot be measured against a baseline "
                "that does not exist; this is reported as a fact, not scored."),
        })
    sev_rank = {"outbreak-flagged": 0, "tier-1": 1, "no severity flag": 2}
    out.sort(key=lambda r: (sev_rank[r["severity"]], -r["observed"]))
    return out


# ─────────────────────────────────────────────────────────────────────────────
# 2. outbreak channel
# ─────────────────────────────────────────────────────────────────────────────

def outbreak_channel(corpus: S.Corpus, week: pd.Period) -> Dict:
    weeks = corpus.weeks
    idx = weeks.index(week)
    df = corpus.frame
    series = {w: int(sum(_truthy(v) for v in df[df["week"] == w]["Outbreak"]))
              for w in weeks}
    b2 = S._window(series, weeks, idx, S.BASELINE_WEEKS, S.GUARD_WEEKS)
    obs = series[week]
    lam = (sum(b2) / len(b2)) if b2 else 0.0
    p = poisson_sf(obs, lam) if lam > 0 else (0.0 if obs > 0 else 1.0)
    recs = _records(corpus, week)
    recs = recs[[_truthy(v) for v in recs["Outbreak"]]]
    return {
        "observed": obs,
        "baseline": [int(x) for x in b2],
        "baseline_mean": round(lam, 2),
        "poisson_p_upper": round(p, 6) if lam > 0 else None,
        "note": ("baseline has no outbreak-flagged weeks; a Poisson tail "
                 "cannot be formed and the count is reported as is"
                 if lam == 0 else
                 "exact Poisson upper tail vs the guarded 7-week mean; "
                 "small counts — read the tail, do not threshold it"),
        "series_12w": [{"week": str(w), "n": series[w]}
                       for w in weeks[max(0, idx - 11):idx + 1]],
        "records": [{
            "date": str(r.get("Date"))[:10], "source": str(r.get("Source")),
            "pathogen": str(r.get("pathogen_c")),
            "company": str(r.get("Company"))[:80],
            "country": str(r.get("Country")),
        } for _, r in recs.head(12).iterrows()],
    }


# ─────────────────────────────────────────────────────────────────────────────
# 3. news corroboration
# ─────────────────────────────────────────────────────────────────────────────

def _load_news(xlsx: Path) -> List[Dict]:
    import openpyxl
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    if "NEWS" not in wb.sheetnames:
        return []
    ws = wb["NEWS"]
    hdr = [c.value for c in ws[1]]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        d = {h: v for h, v in zip(hdr, r) if h}
        if any(str(v).strip() for v in d.values() if v is not None):
            rows.append(d)
    return rows


def news_corroboration(xlsx: Path, strata: Dict[str, S.Stratum],
                       week: pd.Period, dry_run: bool,
                       is_latest: bool = True,
                       week_records: Optional[Counter] = None) -> Dict:
    """The NEWS sheet is a rolling 7-day window at RUN time. It describes the
    latest complete week and nothing else; for a replayed past week the counts
    are stale and are flagged as such, never used for corroboration."""
    news = _load_news(xlsx)
    pathogens = sorted({s.pathogen for s in strata.values() if s.pathogen})
    # a headline mentions a pathogen if any distinctive token of the canonical
    # name appears in it — "Cyclospora", "Listeria", "Salmonella", "botulinum"
    # Only the organism / hazard NAME may match a headline. Generic words in a
    # canonical label ("contamination", "physical", "hazard") matched every
    # headline about anything — on 2026-09-03 "Mouse contamination", "Rodent
    # contamination" and "Physical/foreign-body contamination" all counted a
    # Listeria-in-cheese headline as corroboration. A label with no
    # distinctive token never matches.
    STOP = {"contamination", "contaminant", "physical", "biological",
            "microbial", "microbiological", "hazard", "hazards", "foreign",
            "generic", "producing", "toxin", "toxins", "heavy", "metal",
            "metals", "other", "unknown", "various", "related", "product",
            "products", "matter", "material", "residue", "residues",
            "chemical", "agent", "agents", "group", "serotype", "strain"}
    def tokens(p: str) -> List[str]:
        words = [w.strip(".,;:").lower()
                 for w in p.replace("(", " ").replace(")", " ").replace("/", " ").split()]
        return [w for w in words if len(w) >= 5 and w not in STOP]
    counts: Counter = Counter()
    hits: Dict[str, List[str]] = defaultdict(list)
    for n in news:
        title = str(n.get("Title", "") or "").lower()
        pcol = str(n.get("Pathogen", "") or "").lower()
        text = title + " " + pcol
        for p in pathogens:
            if any(t in text for t in tokens(p)):
                counts[p] += 1
                if len(hits[p]) < 5:
                    hits[p].append(str(n.get("Title", ""))[:110])
    # persist this week's counts so a baseline accumulates
    history: List[Dict] = []
    if NEWS_PULSE.exists():
        for line in NEWS_PULSE.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line:
                try:
                    history.append(json.loads(line))
                except Exception:      # noqa: BLE001
                    pass
    entry = {"week": str(week), "counts": dict(counts), "n_headlines": len(news),
             "captured_utc": datetime.now(timezone.utc).isoformat(timespec="seconds")}
    if is_latest and not dry_run and not any(h.get("week") == str(week) for h in history):
        with open(NEWS_PULSE, "a", encoding="utf-8") as fh:
            fh.write(json.dumps(entry, ensure_ascii=False) + "\n")
    prior = [h for h in history if h.get("week") != str(week)]
    out = []
    for p, c in counts.most_common():
        base = [h.get("counts", {}).get(p, 0) for h in prior[-S.BASELINE_WEEKS:]]
        out.append({
            "pathogen": p, "headlines": c, "burst": c >= NEWS_BURST,
            "baseline_weeks": len(base),
            "baseline_mean": round(sum(base) / len(base), 2) if base else None,
            "samples": hits[p],
        })
    # a burst with NO register activity this week is its own finding: the
    # count detector is blind to it by construction, so it is said out loud
    wr = week_records or Counter()
    news_only = [{"pathogen": o["pathogen"], "headlines": o["headlines"],
                  "register_records_this_week": int(wr.get(o["pathogen"], 0)),
                  "samples": o["samples"]}
                 for o in out if o["burst"] and wr.get(o["pathogen"], 0) == 0]
    return {"n_headlines": len(news), "burst_threshold": NEWS_BURST,
            "applies_to_week": bool(is_latest),
            "note": ("rolling 7-day NEWS sheet read at run time; describes the "
                     "latest complete week" if is_latest else
                     "STALE — the NEWS sheet holds the last 7 days at run time, "
                     "not this replayed week; counts shown for transparency, "
                     "not used for corroboration"),
            "baseline_weeks_available": len(prior), "pathogens": out,
            "news_only": news_only}


# ─────────────────────────────────────────────────────────────────────────────
# 4. alarm review
# ─────────────────────────────────────────────────────────────────────────────

def _stratum_records(corpus: S.Corpus, s: S.Stratum, week: pd.Period) -> pd.DataFrame:
    df = _records(corpus, week)
    m = pd.Series(True, index=df.index)
    if s.pathogen is not None:
        m &= df["pathogen_c"] == s.pathogen
    if s.country is not None:
        m &= df["Country"] == s.country
    elif s.region is not None:
        m &= df["Region"] == s.region
    if s.tier is not None:
        m &= df["Tier"].astype(str).str.replace(".0", "", regex=False) == str(s.tier)
    return df[m]


def review_alarms(board: Dict, corpus: S.Corpus, strata: Dict[str, S.Stratum],
                  week: pd.Period, news: Dict) -> List[Dict]:
    vt = (board.get("context") or {}).get("volume_test") or {}
    corpus_c2 = float(vt.get("c2") or 0.0)
    news_by = ({n["pathogen"]: n["headlines"] for n in news.get("pathogens", [])}
               if news.get("applies_to_week") else {})
    out = []
    for sig in board.get("signals", []):
        s = strata.get(sig["stratum_key"])
        recs = _stratum_records(corpus, s, week) if s else pd.DataFrame()
        n = len(recs)
        firms = Counter(str(c).strip().lower()[:40] for c in recs["Company"]) if n else Counter()
        top_firm, top_n = (firms.most_common(1)[0] if firms else ("", 0))
        cluster = n >= 3 and top_n / n >= CLUSTER_SHARE
        single_pub = float(sig.get("dominant_share") or 0) >= PUBLISHER_SHARE
        outbreak_n = int(sum(_truthy(v) for v in recs["Outbreak"])) if n else 0
        headlines = news_by.get(sig.get("pathogen"), 0) if sig.get("pathogen") else 0
        volume_art = sig.get("channel") == "count-only" and corpus_c2 >= 2.0
        reasons = []
        if outbreak_n:
            reasons.append(f"{outbreak_n} outbreak-flagged record(s) in the stratum this week")
        if headlines >= NEWS_BURST:
            reasons.append(f"{headlines} headlines on {sig.get('pathogen')} this week")
        if cluster:
            reasons.append(f"{top_n} of {n} records from one firm ({top_firm}) — one incident, many notices")
        if single_pub:
            reasons.append(f"{int(round(float(sig.get('dominant_share')) * 100))}% of records from {sig.get('dominant_source')}")
        if volume_art:
            reasons.append(f"count-only signal while corpus C2 = {corpus_c2:+.2f} — publisher volume swing")
        if outbreak_n or headlines >= NEWS_BURST:
            hint = "corroborated"
        elif volume_art:
            hint = "volume-artefact"
        elif cluster:
            hint = "publication-cluster"
        elif single_pub:
            hint = "single-publisher"
        else:
            hint = "unclear"
        out.append({
            "label": sig["label"], "stratum_key": sig["stratum_key"],
            "channel": sig["channel"], "effect": sig["effect"],
            "p_value": sig["p_value"], "fdr_pass": sig["fdr_pass"],
            "records": n, "distinct_firms": len(firms),
            "top_firm": top_firm, "top_firm_share": round(top_n / n, 2) if n else None,
            "outbreak_records": outbreak_n, "headlines": headlines,
            "verdict_hint": hint,
            "reason": "; ".join(reasons) or "no corroboration and no artefact signature found",
        })
    return out


# ─────────────────────────────────────────────────────────────────────────────
# main
# ─────────────────────────────────────────────────────────────────────────────

def render(review: Dict) -> str:
    m = review["meta"]
    L = [f"Signal review — week {m['week']} ({m['week_start']} to {m['week_end']})",
         f"corpus {m['corpus_week_total']} · alarms reviewed {len(review['alarm_reviews'])} "
         f"· novel {len(review['novelty'])} · outbreak-flagged {review['outbreak_channel']['observed']} "
         f"· headlines {review['news']['n_headlines']}"]
    for nv in review["novelty"]:
        L.append(f"  NOVEL   {nv['pathogen']:<34} {nv['observed']} record(s) · {nv['severity']}"
                 f"{' · FIRST APPEARANCE EVER' if nv['first_appearance_ever'] else ''} · untestable")
    oc = review["outbreak_channel"]
    L.append(f"  OUTBREAK channel: {oc['observed']} vs baseline {oc['baseline']} (mean {oc['baseline_mean']})"
             + (f" · Poisson P[X≥obs] = {oc['poisson_p_upper']}" if oc['poisson_p_upper'] is not None else " · no baseline"))
    if not review["news"]["applies_to_week"]:
        L.append("  NEWS    (stale for a replayed week — shown, not used)")
    for no in review["news"].get("news_only", []):
        L.append(f"  NEWS-ONLY {no['pathogen']:<32} {no['headlines']} headline(s), "
                 f"0 register records this week — the count detector is blind to this by construction")
    for nw in review["news"]["pathogens"][:6]:
        base = "" if nw["baseline_mean"] is None else f" · baseline {nw['baseline_mean']}"
        burst = " · BURST" if nw["burst"] else ""
        L.append(f"  NEWS    {nw['pathogen']:<34} {nw['headlines']} headline(s){burst}{base}")
    for a in review["alarm_reviews"]:
        L.append(f"  ALARM   {a['label']:<34} {a['channel']:<11} ×{a['effect']:<5} → {a['verdict_hint']:<19} {a['reason']}")
    if not review["alarm_reviews"]:
        L.append("  (no alarms this week to review)")
    return "\n".join(L)


def main(argv: Optional[List[str]] = None) -> int:
    ap = argparse.ArgumentParser(description="AFTS signal review agent")
    ap.add_argument("--xlsx", default=str(XLSX))
    ap.add_argument("--board", default=str(BOARD))
    ap.add_argument("--asof", default=None, help="ISO date inside the week to review")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args(argv)

    corpus = S.load_corpus(args.xlsx)
    strata = S.build_strata(corpus)
    week = _week_for(args.asof, corpus)

    board: Dict = {}
    if Path(args.board).exists():
        board = json.loads(Path(args.board).read_text(encoding="utf-8"))
        if board.get("meta", {}).get("week") != str(week):
            # the board is for another week: rebuild the alarm list for THIS
            # week from the detector so the review is internally consistent
            sigs, meta = S.detect(corpus, strata, asof=week)
            board = {"meta": meta, "signals": [S.asdict(x) for x in sigs],
                     "context": S.build_context(corpus, week)}
    else:
        sigs, meta = S.detect(corpus, strata, asof=week)
        board = {"meta": meta, "signals": [S.asdict(x) for x in sigs],
                 "context": S.build_context(corpus, week)}

    is_latest = week == corpus.weeks[-1]
    wk_recs = Counter(str(p) for p in _records(corpus, week)["pathogen_c"])
    news = news_corroboration(Path(args.xlsx), strata, week, args.dry_run,
                              is_latest=is_latest, week_records=wk_recs)
    review = {
        "meta": {
            "week": str(week),
            "week_start": str(week.start_time.date()),
            "week_end": str(week.end_time.date()),
            "corpus_week_total": int(corpus.totals.get(week, 0)),
            "generated_utc": datetime.now(timezone.utc).isoformat(timespec="seconds"),
            "agent": "signal_review_agent", "model": None,
            "advisory_only": True,
            "channels": ["novelty", "outbreak", "news_corroboration", "alarm_review"],
        },
        "novelty": novelty(corpus, strata, week),
        "outbreak_channel": outbreak_channel(corpus, week),
        "news": news,
        "alarm_reviews": review_alarms(board, corpus, strata, week, news),
    }
    print(render(review))
    if args.dry_run:
        print("\nDRY RUN — nothing written.")
        return 0
    OUT.write_text(json.dumps(review, ensure_ascii=False, indent=1), encoding="utf-8")
    with open(OUT_HIST, "a", encoding="utf-8") as fh:
        fh.write(json.dumps(review, ensure_ascii=False) + "\n")
    print(f"\nwrote {OUT}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
