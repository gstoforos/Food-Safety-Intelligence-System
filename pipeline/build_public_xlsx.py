"""
pipeline/build_public_xlsx.py
=================================

Builds the public-facing formatted Excel file shipped to:
  - the Thursday 17:00 Athens review email (info@advfood.tech)
  - the public dashboard's "Download XLSX" button

Reads:  docs/data/recalls.xlsx  (Recalls sheet — operator file, raw)
Writes: docs/data/afts-recalls-public.xlsx  (formatted, public-ready)

Triggered by:  .github/workflows/public-xlsx-build.yml
               Thursday 16:30 Athens — runs BEFORE the 17:00 mailer
               (the mailer's fetchRecallsXlsxBlob_() picks up this file
               from GitHub Pages).

Formatting decisions locked in audit 2026-05-07 (no green, black text):
  - Title row: black bold on darker grey (9CA3AF)
  - Header row: black bold on lighter grey (E5E7EB)
  - Tier 1 cell: white bold on red (DC2626)        — severity warning, kept
  - Tier 2 cell: black bold on light orange (FED7AA)
  - Tier 3 cell: black on very light grey (F3F4F6)
  - Outbreak: empty when 0, "YES" in bold orange when 1
  - URL column: blue underlined hyperlink, label = domain+path truncated 55 chars
  - Product + Reason: wrap text, top-aligned
  - Freeze header (row 2 frozen)
  - AutoFilter on header so subscribers can sort/filter
  - Landscape print, fit-to-1-page-wide

This script is read-only with respect to the operator file —
it only WRITES afts-recalls-public.xlsx, never modifies recalls.xlsx.

CLI usage:
    python -m pipeline.build_public_xlsx \\
        --src docs/data/recalls.xlsx \\
        --dst docs/data/afts-recalls-public.xlsx
"""
from __future__ import annotations

import argparse
import logging
import sys
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)

# ── Column layout (must match Recalls sheet order in recalls.xlsx) ─────
# Operator-only columns (DateAdded, LastUpdated, LastChecked, Status,
# ApprovedBy, etc.) are STRIPPED — subscribers see only the public columns.
PUBLIC_COLUMNS = [
    "Date", "Source", "Company", "Brand", "Product", "Pathogen",
    "Reason", "Class", "Country", "Region", "Tier", "Outbreak", "URL",
]

# Column-letter widths — keys are letters of the OUTPUT sheet (1-indexed
# A..M for the 13 public columns).
WIDTHS = {
    "A": 12,   # Date
    "B": 20,   # Source
    "C": 28,   # Company
    "D": 22,   # Brand
    "E": 60,   # Product (long, wraps)
    "F": 26,   # Pathogen
    "G": 35,   # Reason (wraps)
    "H": 11,   # Class
    "I": 14,   # Country
    "J": 11,   # Region
    "K": 7,    # Tier
    "L": 10,   # Outbreak
    "M": 36,   # URL
}

# Style palette (audit 2026-05-07 — no green, black text default)
TITLE_FILL    = PatternFill("solid", fgColor="9CA3AF")  # darker grey
HEADER_FILL   = PatternFill("solid", fgColor="E5E7EB")  # lighter grey
TIER1_FILL    = PatternFill("solid", fgColor="DC2626")  # red
TIER2_FILL    = PatternFill("solid", fgColor="FED7AA")  # orange
TIER3_FILL    = PatternFill("solid", fgColor="F3F4F6")  # very light grey

TITLE_FONT    = Font(name="Calibri", size=14, bold=True, color="000000")
HEADER_FONT   = Font(name="Calibri", size=11, bold=True, color="000000")
DATA_FONT     = Font(name="Calibri", size=10, color="000000")
DATA_BOLD     = Font(name="Calibri", size=10, color="000000", bold=True)
URL_FONT      = Font(name="Calibri", size=10, color="1D4ED8", underline="single")
TIER1_FONT    = Font(name="Calibri", size=10, color="FFFFFF", bold=True)
OUTBREAK_FONT = Font(name="Calibri", size=10, color="C2410C", bold=True)

THIN_GREY = Side(style="thin", color="E5E7EB")
BORDER    = Border(left=THIN_GREY, right=THIN_GREY, top=THIN_GREY, bottom=THIN_GREY)

ALIGN_LEFT_TOP   = Alignment(horizontal="left",   vertical="top",    wrap_text=True, indent=1)
ALIGN_LEFT_MID   = Alignment(horizontal="left",   vertical="center", indent=1)
ALIGN_CENTER_MID = Alignment(horizontal="center", vertical="center")

WRAP_COL_INDICES = {5, 7}  # Product, Reason (1-indexed: E, G)
URL_COL_INDEX    = 13      # M
TIER_COL_INDEX   = 11      # K
OB_COL_INDEX     = 12      # L


def _truncate_url_label(url: str, max_len: int = 55) -> str:
    """Format a URL for display: domain + path, truncated. Cosmetic only —
    the cell's hyperlink target stays the full original URL."""
    try:
        p = urlparse(url)
        label = (p.netloc + p.path)
        if len(label) > max_len:
            return label[:max_len] + "…"
        return label
    except Exception:
        return url[:max_len] + ("…" if len(url) > max_len else "")


def build_public_xlsx(src: Path, dst: Path) -> int:
    """Read recalls.xlsx Recalls sheet, write a formatted public xlsx.
    Returns the number of recall rows written (excluding headers)."""
    if not src.exists():
        raise FileNotFoundError(f"Source workbook not found: {src}")

    src_wb = load_workbook(src, data_only=True)
    if "Recalls" not in src_wb.sheetnames:
        raise ValueError(f"'Recalls' sheet missing from {src}; "
                         f"sheets present: {src_wb.sheetnames}")
    src_ws = src_wb["Recalls"]

    # Header → column-index map so we can pull only the public columns
    # in their canonical order, regardless of operator-file column order.
    src_headers = {c.value: i + 1 for i, c in enumerate(src_ws[1])}
    missing = [h for h in PUBLIC_COLUMNS if h not in src_headers]
    if missing:
        raise ValueError(f"Source recalls.xlsx missing required columns: "
                         f"{missing} (have {list(src_headers.keys())})")

    # ── Build the destination workbook ────────────────────────────────
    from openpyxl import Workbook
    dst_wb = Workbook()
    dst_ws = dst_wb.active
    dst_ws.title = "Recalls"

    # Title row (row 1, merged)
    n_rows = src_ws.max_row - 1  # exclude header
    title_text = (
        "AFTS · Food Safety Intelligence System — "
        "Recalls Export · "
        f"{n_rows} pathogen recalls"
    )
    dst_ws.cell(row=1, column=1).value = title_text
    dst_ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=len(PUBLIC_COLUMNS))
    title_cell = dst_ws.cell(row=1, column=1)
    title_cell.font      = TITLE_FONT
    title_cell.fill      = TITLE_FILL
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    dst_ws.row_dimensions[1].height = 32

    # Header row (row 2)
    for col_idx, header in enumerate(PUBLIC_COLUMNS, start=1):
        c = dst_ws.cell(row=2, column=col_idx, value=header)
        c.font      = HEADER_FONT
        c.fill      = HEADER_FILL
        c.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=True, indent=1)
    dst_ws.row_dimensions[2].height = 28

    # Column widths
    for col_letter, width in WIDTHS.items():
        dst_ws.column_dimensions[col_letter].width = width

    # ── Data rows: copy public columns, apply formatting ──────────────
    out_row = 3
    for src_row in src_ws.iter_rows(min_row=2, max_row=src_ws.max_row,
                                    values_only=False):
        # Build dict from this source row keyed by source header name.
        src_values = {h: src_row[i - 1].value
                      for h, i in src_headers.items()
                      if i - 1 < len(src_row)}

        for col_idx, public_header in enumerate(PUBLIC_COLUMNS, start=1):
            value = src_values.get(public_header, "")
            cell  = dst_ws.cell(row=out_row, column=col_idx, value=value)
            cell.font   = DATA_FONT
            cell.border = BORDER

            if col_idx in WRAP_COL_INDICES:
                cell.alignment = ALIGN_LEFT_TOP
            elif col_idx in (TIER_COL_INDEX, OB_COL_INDEX):
                cell.alignment = ALIGN_CENTER_MID
            else:
                cell.alignment = ALIGN_LEFT_MID

            # URL → hyperlink with truncated label
            if col_idx == URL_COL_INDEX and value:
                url_str = str(value).strip()
                cell.hyperlink = url_str
                cell.font      = URL_FONT
                cell.value     = _truncate_url_label(url_str)

        # Tier color coding (column K)
        tier_cell = dst_ws.cell(row=out_row, column=TIER_COL_INDEX)
        try:
            tier_int = int(tier_cell.value) if tier_cell.value is not None else 0
        except (TypeError, ValueError):
            tier_int = 0
        if tier_int == 1:
            tier_cell.fill = TIER1_FILL
            tier_cell.font = TIER1_FONT
        elif tier_int == 2:
            tier_cell.fill = TIER2_FILL
            tier_cell.font = DATA_BOLD
        else:
            tier_cell.fill = TIER3_FILL

        # Outbreak: empty cell when 0, bold orange "YES" when 1
        ob_cell = dst_ws.cell(row=out_row, column=OB_COL_INDEX)
        try:
            ob_int = int(ob_cell.value) if ob_cell.value is not None else 0
        except (TypeError, ValueError):
            ob_int = 0
        if ob_int == 1:
            ob_cell.font  = OUTBREAK_FONT
            ob_cell.value = "YES"
        else:
            ob_cell.value = ""

        dst_ws.row_dimensions[out_row].height = 38
        out_row += 1

    # Freeze header (row 2 frozen → row 3 onward scrolls)
    dst_ws.freeze_panes = "A3"

    # AutoFilter on header row across all data rows
    last_col = get_column_letter(len(PUBLIC_COLUMNS))
    dst_ws.auto_filter.ref = f"A2:{last_col}{dst_ws.max_row}"

    # Print setup — landscape, fit to one page wide
    dst_ws.page_setup.orientation = dst_ws.ORIENTATION_LANDSCAPE
    dst_ws.page_setup.fitToWidth  = 1
    dst_ws.page_setup.fitToHeight = 0
    dst_ws.print_options.gridLines = False

    dst.parent.mkdir(parents=True, exist_ok=True)
    dst_wb.save(dst)
    return n_rows


def _parse_args(argv: list[str]) -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Build the public-facing formatted recalls xlsx."
    )
    p.add_argument("--src", type=Path,
                   default=Path("docs/data/recalls.xlsx"),
                   help="Source operator workbook (default: docs/data/recalls.xlsx)")
    p.add_argument("--dst", type=Path,
                   default=Path("docs/data/afts-recalls-public.xlsx"),
                   help="Output public workbook "
                        "(default: docs/data/afts-recalls-public.xlsx)")
    p.add_argument("-v", "--verbose", action="store_true")
    return p.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = _parse_args(argv if argv is not None else sys.argv[1:])
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    )
    log.info("Building public xlsx: %s → %s", args.src, args.dst)
    n = build_public_xlsx(args.src, args.dst)
    log.info("Wrote %d recall rows to %s (%.1f KB)",
             n, args.dst, args.dst.stat().st_size / 1024)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
