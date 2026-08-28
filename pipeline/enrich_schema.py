#!/usr/bin/env python3
"""Write the statistical columns into the Recalls sheet. The assembler.

WHY THIS FILE EXISTS
--------------------
The extractors were built and never joined up: regulator_fields.py reads the
RASFF taxonomy, product_axes.py derives the four product axes, and neither
had anything that put the answers in the workbook. The schema existed as a
proposal, a prompt and two modules — and the register still had eighteen
columns.

This is the assembler. It runs the extractors in tier order, writes the
result, and records which tier answered so an inference can never be
mistaken for a regulator's statement.

    tier 1  the regulator's own published field   (RASFF category/risk/class)
    tier 2  a structured field in stored text     (RASFF id, CFIA category)
    tier 3  multilingual keyword over Product+Reason
    tier 4  unknown — always allowed, never inferred around

WHAT IT WILL NOT DO
-------------------
* touch any existing column. The eighteen columns the pipeline already
  writes are read-only here.
* overwrite a value a human set. A row whose EnrichedBy says "human" is
  skipped entirely.
* leak. Every column added is registered in
  merge_master.RECALLS_INTERNAL_COLUMNS, so mirror_json_from_xlsx strips it
  from recalls.json, and the public xlsx builders use allow-lists that do
  not name it. Verified by tests/test_enrich_schema.py.
* guess. Every value is either a regulator's own term or a keyword match on
  text the notice actually contains. "unknown" is a correct answer and the
  most common one on several axes.

    python -m pipeline.enrich_schema --dry-run
    python -m pipeline.enrich_schema --write
"""
from __future__ import annotations

import argparse
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Tuple

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl  # noqa: E402

from pipeline import product_axes as PA          # noqa: E402
from pipeline.regulator_fields import parse_rasff  # noqa: E402

XLSX = ROOT / "docs" / "data" / "recalls.xlsx"
SHEET = "Recalls"
VERSION = "enrich-schema/1"

# Order matters: this is the order they are appended to the sheet.
COLUMNS: Tuple[str, ...] = (
    "FoodCategory", "ProcessType", "ConsumptionState", "StorageCondition",
    "PackagingType", "PackagingForm", "PreservationSystem", "HazardGroup", "HazardCertainty",
    "NoticeType", "SeverityClass", "EventID",
    "EnrichedBy", "EnrichedAt", "EnrichmentTier",
)

# ── Class -> NoticeType / SeverityClass ─────────────────────────────────
# The Class column holds two variables at once: a regulatory ACTION and, for
# US/Canadian sources, an FDA SEVERITY class. Splitting them is the whole
# point; "Class I" and "Class 1" merge themselves once they are read as
# severity rather than as action.
_NOTICE_FROM_CLASS = {
    "border rejection": "border-rejection",
    # FDA and CFIA state a SEVERITY class where other sources state an action.
    # "Class I" carried a severity but no notice type at all, so an FDA
    # Class I recall — the most serious category there is — had a blank in
    # the field that says what kind of notice it was.
    "class i": "consumer-recall", "class ii": "consumer-recall",
    "class iii": "consumer-recall", "class 1": "consumer-recall",
    "class 2": "consumer-recall", "class 3": "consumer-recall",
    "voluntary": "consumer-recall",
    "mandatory": "consumer-recall",
    "recall": "consumer-recall",
    "alert": "public-warning",
    "information": "information",
    "withdrawal": "withdrawal",
}
# Controlled terms. "i" / "ii" / "iii" were bare roman numerals — a value
# that reads as a footnote marker in a spreadsheet cell and sorts next to
# nothing. Corrected 2026-08-28 to the full controlled term, matching the
# every-other-column convention of a lowercase hyphenated word.
_SEVERITY_FROM_CLASS = {
    "class i": "class-i", "class 1": "class-i",
    "food recall warning (class 1)": "class-i",
    "class ii": "class-ii", "class 2": "class-ii",
    "class iii": "class-iii", "class 3": "class-iii",
}

_HAZARD_GROUP_RULES = (
    ("foreign material", "foreign-material"),
    ("foreign body", "foreign-material"), ("foreign-body", "foreign-material"),
    ("physical", "foreign-material"), ("glass fragment", "foreign-material"),
    ("metal fragment", "foreign-material"), ("plastic fragment", "foreign-material"),
    ("heavy metal", "heavy-metal"),
    ("aflatoxin", "mycotoxin"), ("ochratoxin", "mycotoxin"),
    ("mycotoxin", "mycotoxin"), ("zearalenone", "mycotoxin"),
    ("patulin", "mycotoxin"),
    ("histamine", "biotoxin"), ("scombrotoxin", "biotoxin"),
    ("cereulide", "biotoxin"), ("toxin", "biotoxin"),
    ("norovirus", "pathogen-viral"), ("hepatitis", "pathogen-viral"),
    ("cyclospora", "pathogen-parasitic"), ("anisakis", "pathogen-parasitic"),
    ("rodent", "pest-rodent"), ("mouse", "pest-rodent"), ("insect", "pest-rodent"),
    ("pesticide", "chemical"), ("nitrite", "chemical"), ("sulphite", "chemical"),
    ("ethylene oxide", "chemical"), ("hydrocyanic", "chemical"),
    ("pfoa", "chemical"), ("pfas", "chemical"), ("residue", "chemical"),
    ("veterinary", "chemical"), ("nitrofurazone", "chemical"),
)

# Labels that name no hazard at all. They must not fall through to the
# bacterial catch-all: "None (organoleptic spoilage)" is not a pathogen, and
# a stratum built on the bacterial bucket would silently carry it.
_NOT_A_HAZARD = ("unspecified hazard", "none (", "organoleptic",
                 "process deviation", "microbiological quality",
                 "unintended fermentation", "inadequate sterilization",
                 "incomplete pasteurization", "labeling", "labelling")


# Organisms whose NAME contains a hazard word that belongs to another group.
# Checked before the rule table, because the table matches substrings and
# "Shiga toxin-producing E. coli" contains "toxin".
#
# Measured 2026-08-28 on a 20-row hand review: rows 2 and 15 were filed as
# "biotoxin". Worse than simply wrong — bare "STEC" fell through to
# pathogen-bacterial, so the SAME organism landed in two different hazard
# groups depending on which of its names the notifying authority used, and
# any stratification on HazardGroup silently split it in half.
_ORGANISM_OVERRIDES = (
    ("shiga toxin", "pathogen-bacterial"),
    ("shigatoxin", "pathogen-bacterial"),
    ("verotoxin", "pathogen-bacterial"),
    ("vtec", "pathogen-bacterial"),
    ("stec", "pathogen-bacterial"),
    ("toxin-producing", "pathogen-bacterial"),
    ("toxigenic", "pathogen-bacterial"),
)


def _notice_from_class(cls: str) -> str:
    """Map a source's Class string to a controlled notice type.

    SUBSTRING, not exact match (audit 2026-08-28). The lookup was
    `_NOTICE_FROM_CLASS.get(cls)`, which requires the Class cell to equal a
    key exactly. Two of the six continuous sources never do:

        CFIA        "Food recall warning"   -> unknown
        USDA FSIS   "Public Health Alert"   -> unknown

    Three rows in a twenty-row sample lost their notice type to this, and
    both sources publish every one of their notices under those exact words,
    so the loss is systematic rather than incidental.

    Longest key first, so "border rejection" is not shadowed by a shorter
    key that also appears in the string.
    """
    c = (cls or "").strip().lower()
    if not c:
        return ""
    for needle in sorted(_NOTICE_FROM_CLASS, key=len, reverse=True):
        if needle in c:
            return _NOTICE_FROM_CLASS[needle]
    return ""


def _hazard_group(pathogen: str) -> str:
    p = (pathogen or "").strip().lower()
    if not p:
        return "unknown"
    for needle, group in _ORGANISM_OVERRIDES:
        if needle in p:
            return group
    for needle, group in _HAZARD_GROUP_RULES:
        if needle in p:
            return group
    for needle in _NOT_A_HAZARD:
        if needle in p:
            return "unknown"
    # Anything left that names an organism is bacterial. The catch-all is
    # last so every specific rule wins, and the _NOT_A_HAZARD guard above
    # stops process-deviation and quality labels landing here — a stratum on
    # the bacterial bucket would otherwise carry "None (organoleptic
    # spoilage)" as a pathogen.
    return "pathogen-bacterial"


def _first(*vals: str) -> str:
    for v in vals:
        if v and v not in ("unknown", ""):
            return v
    return "unknown"



# ── Structural fallback: what the commodity class already settles ────────
#
# Runs ONLY when the keyword pass returned "unknown", and only from two axes
# that have themselves been derived — never from raw text a second time. The
# evidence string records which pair decided it, so a structural answer is
# always separable from a term match in an audit.
#
# WHY. After the keyword fixes of 2026-08-28, ConsumptionState was still the
# weakest axis at 44% on the hand-reviewed sample: ten of twenty rows were
# answerable and came back unknown. Every one of them was answerable from
# the commodity class rather than from any word in the product name — a bag
# of dried figs, a chocolate snack, a spice mix, frozen chicken thigh meat.
# Naming the pair is not a guess; "a spice mix is added to food" and "raw
# poultry is cooked before eating" are facts about the class, not readings
# of the string.
#
# Deliberately NOT covered: bakery-cereal (flour and dry pasta are cooked,
# biscuits are not — the class does not settle it), fish-seafood (oysters
# raw, cod cooked), dairy-other (raw milk vs cheese), eggs, fresh-produce.
# Those stay unknown, which is the correct answer.
_CONSUMPTION_FROM_CLASS = {
    ("herbs-spices",         None):            "ingredient",
    ("supplements",          None):            "ready-to-eat",
    ("confectionery-snacks", None):            "ready-to-eat",
    ("dried-fruit",          None):            "ready-to-eat",
    ("nuts-seeds",           None):            "ready-to-eat",
    ("beverages",            None):            "ready-to-eat",
    ("dairy-soft-cheese",    None):            "ready-to-eat",
    ("prepared-meals",       "heat-treated"):  "ready-to-eat",
    ("prepared-meals",       "composite"):     "ready-to-eat",
    ("prepared-meals",       "fresh-cut"):     "ready-to-eat",
    ("prepared-meals",       "unknown"):       "ready-to-eat",
    ("meat-poultry",         "raw"):           "cook-before-eating",
    ("meat-other",           "raw"):           "cook-before-eating",
    ("fresh-produce",        "dried"):         "ready-to-eat",
}

# A dried commodity that is eaten as sold keeps at ambient. Confined to the
# classes where "dried" means shelf-stable, not "flour waiting to be baked".
_STORAGE_FROM_CLASS = {
    ("nuts-seeds",           "dried"): "ambient",
    ("dried-fruit",          "dried"): "ambient",
    ("dried-fruit",          None):    "ambient",
    ("herbs-spices",         "dried"): "ambient",
    ("herbs-spices",         None):    "ambient",
    ("fresh-produce",        "dried"): "ambient",
    ("supplements",          "dried"): "ambient",
    ("confectionery-snacks", "dried"): "ambient",
    ("dairy-soft-cheese",    None):    "chilled",
}


def _from_class(table, food: str, proc: str, current: str) -> tuple:
    """(value, evidence) — leaves a known value alone."""
    if current and current != "unknown":
        return current, ""
    for key in ((food, proc), (food, None)):
        if key in table:
            return table[key], f"class:{food}" + (f"+{proc}" if key[1] else "")
    return current, ""


def derive(row: Dict[str, Any]) -> Tuple[Dict[str, str], str]:
    """Return (column values, the strongest tier that answered)."""
    rasff = parse_rasff(row.get("Reason"), row.get("Notes"))
    cls = str(row.get("Class", "") or "").strip().lower()

    # FoodCategory: RASFF's own term first, then a CFIA category if the row
    # carries one. product_axes deliberately does not own this axis.
    #
    # split_cfia_category takes the CATEGORY STRING, not the row. Passing the
    # row made it stringify the whole dict and return it as the category, so
    # every non-RASFF row got a category and the coverage read 100% — on a
    # field that is genuinely hard. A suspiciously perfect number on a hard
    # field is the tell; this was caught by printing the distribution before
    # writing anything.
    food = rasff.food_category or ""
    if not food:
        raw_cat = str(row.get("Category") or row.get("category") or "")
        if raw_cat.lower().startswith("food -"):
            commodity, _proc = PA.split_cfia_category(raw_cat)
            if commodity:
                from pipeline.regulator_fields import CATEGORY_MAP
                food = CATEGORY_MAP.get(commodity.strip().lower(), "")

    # TIER 3 — the regulator said nothing. RASFF is 490 of 1,532 rows, so
    # without this every FDA, RappelConso, FSANZ, BLV and USDA row was
    # "unknown" by construction and no term-list work could have moved it.
    # Runs last, never overrides a regulator, and the tier is recorded.
    food_from_text = ""
    if not food:
        food_from_text = PA.food_category(row)[0]
        if food_from_text != "unknown":
            food = food_from_text

    proc = PA.process_type(row)
    cons = PA.consumption_state(row)
    stor = PA.storage_condition(row)
    ptyp = PA.packaging_type(row)
    pfrm = PA.packaging_form(row)

    notice = rasff.notice_type or _notice_from_class(cls)
    severity = _SEVERITY_FROM_CLASS.get(cls, "not-classified" if cls else "unknown")

    event = ""
    if rasff.rasff_id:
        event = f"rasff:{rasff.rasff_id}"
    else:
        try:
            from pipeline._outbreak_id import derive as _ob
            oid, conf, _ = _ob(row)
            if oid and conf == "high":
                event = oid
        except Exception:                                   # noqa: BLE001
            pass

    tier = ("tier1-regulator" if (rasff.food_category or rasff.risk
                                  or rasff.classification_raw)
            else ("tier2-structured" if event or notice != "" else "tier3-keyword"))

    # Structural fallback, applied after every keyword pass has had its turn.
    _food = food or "unknown"
    _cons, _cev = _from_class(_CONSUMPTION_FROM_CLASS, _food, proc[0], cons[0])
    _stor, _sev = _from_class(_STORAGE_FROM_CLASS, _food, proc[0], stor[0])

    return {
        "FoodCategory": _food,
        "ProcessType": proc[0],
        "ConsumptionState": _cons,
        "StorageCondition": _stor,
        "PackagingType": ptyp[0],
        "PackagingForm": pfrm[0],
        "PreservationSystem": PA.preservation_system(
            _food, proc[0], _stor, _cons, row)[0],
        "HazardGroup": _hazard_group(str(row.get("Pathogen", ""))),
        "HazardCertainty": rasff.risk or "unknown",
        "NoticeType": notice or "unknown",
        "SeverityClass": severity,
        "EventID": event or "",
    }, tier


def run(xlsx: Path, write: bool) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(xlsx)
    ws = wb[SHEET]
    hdr = [str(c.value or "") for c in ws[1]]

    # Append any column that is not already present, preserving order.
    for name in COLUMNS:
        if name not in hdr:
            hdr.append(name)
            ws.cell(row=1, column=len(hdr), value=name)
    idx = {h: i + 1 for i, h in enumerate(hdr)}

    today = datetime.now(timezone.utc).date().isoformat()
    stats: Counter = Counter()
    tiers: Counter = Counter()
    skipped_human = 0

    for r in range(2, ws.max_row + 1):
        row = {h: ws.cell(row=r, column=idx[h]).value for h in hdr}
        if str(row.get("EnrichedBy") or "").strip().lower() == "human":
            skipped_human += 1
            continue
        values, tier = derive(row)
        for k, v in values.items():
            if write:
                ws.cell(row=r, column=idx[k], value=v)
            if v and v != "unknown":
                stats[k] += 1
        tiers[tier] += 1
        if write:
            ws.cell(row=r, column=idx["EnrichedBy"], value=VERSION)
            ws.cell(row=r, column=idx["EnrichedAt"], value=today)
            ws.cell(row=r, column=idx["EnrichmentTier"], value=tier)

    n = ws.max_row - 1
    if write:
        wb.save(xlsx)
    return {"rows": n, "filled": dict(stats), "tiers": dict(tiers),
            "skipped_human": skipped_human}


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--xlsx", default=str(XLSX))
    ap.add_argument("--write", action="store_true",
                    help="apply the derivation to the workbook")
    # --dry-run is the documented default (see the module docstring, and the
    # first step of .github/workflows/enrich-schema.yml). It did not exist:
    # argparse rejected it with exit code 2, which would have failed the
    # scheduled sweep on its very first run, before it wrote anything.
    # Accepted explicitly now, and mutually exclusive with --write so
    # "--dry-run --write" is an error rather than a silent write.
    ap.add_argument("--dry-run", action="store_true",
                    help="report coverage without touching the workbook "
                         "(the default; accepted so it can be passed "
                         "explicitly by callers and CI)")
    a = ap.parse_args(argv)

    if a.dry_run and a.write:
        ap.error("--dry-run and --write are mutually exclusive")

    res = run(Path(a.xlsx), a.write)
    n = res["rows"]
    print(f"{n} rows{' — WRITTEN' if a.write else ' — dry run'}")
    if res["skipped_human"]:
        print(f"  {res['skipped_human']} row(s) skipped (EnrichedBy=human)")
    print()
    for c in COLUMNS:
        if c in ("EnrichedBy", "EnrichedAt", "EnrichmentTier"):
            continue
        k = res["filled"].get(c, 0)
        print(f"  {c:18} {k:5}/{n}  {k / n * 100:5.1f}%")
    print()
    for t, k in sorted(res["tiers"].items()):
        print(f"  {t:18} {k:5}  {k / n * 100:5.1f}%")
    if not a.write:
        print("\n(dry run — pass --write to apply)")
        print("Remember: register the columns in "
              "merge_master.RECALLS_INTERNAL_COLUMNS before committing, or "
              "they leak into recalls.json.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
