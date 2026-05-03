"""
pipeline/weekly_review_capture.py

Mirrors every Pending → Recalls promotion into a new "Weekly_Review"
sheet of docs/data/recalls.xlsx, tagged with the Thursday end-date of
the review window in which George will see it (Thu 17:00 Athens cutoff).

Also emits docs/data/weekly-review-latest.json — a slice of the
upcoming/current Thu→Thu window, fetched by the Apps Script Thursday
mailer (sendThursdayManualReview).

Sheet schema (Weekly_Review):
    Recalls columns + Week_Added (ISO date of the Thursday on which
    the row gets reviewed) + Reviewed (Y/N — manual stamp; not used
    by the mailer in v1, reserved for future filtering).

Cutoff rule (matches the Thursday 17:00 Athens email):
    A row promoted strictly BEFORE Thursday 17:00 Athens belongs to
    that Thursday's review. A row promoted at or after that boundary
    rolls over to the following Thursday. Each promotion lands in
    exactly one Thursday email.

Author: AFTS / G. Stoforos
"""
from __future__ import annotations

import json
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional
from zoneinfo import ZoneInfo

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook import Workbook

ROOT = Path(__file__).resolve().parent.parent
XLSX_DEFAULT = ROOT / "docs" / "data" / "recalls.xlsx"
JSON_DEFAULT = ROOT / "docs" / "data" / "weekly-review-latest.json"

SHEET_NAME = "Weekly_Review"
ATHENS = ZoneInfo("Europe/Athens")
REVIEW_HOUR_LOCAL = 17  # Thursday 17:00 Athens cutoff

# Mirror the Recalls schema (must stay in sync with merge_master.RECALLS_SCHEMA).
RECALLS_COLS: List[str] = [
    "Date", "Source", "Company", "Brand", "Product", "Pathogen", "Reason",
    "Class", "Country", "Region", "Tier", "Outbreak", "URL", "Notes",
    "DateAdded", "LastUpdated", "LastChecked",
]
EXTRA_COLS: List[str] = ["Week_Added", "Reviewed"]
SHEET_COLS: List[str] = RECALLS_COLS + EXTRA_COLS


# ---------------------------------------------------------------------------
# Date math
# ---------------------------------------------------------------------------
def review_thursday_for(now_utc: Optional[datetime] = None) -> date:
    """
    The next Thursday review date — i.e. the Thursday email a row
    promoted RIGHT NOW will appear in. Rows promoted at/after Thursday
    17:00 Athens roll over to the following Thursday.
    """
    if now_utc is None:
        now_utc = datetime.now(timezone.utc)
    local = now_utc.astimezone(ATHENS)
    d = local.date()
    days_until_thu = (3 - d.weekday()) % 7  # Mon=0 .. Thu=3 .. Sun=6
    if days_until_thu == 0 and local.hour >= REVIEW_HOUR_LOCAL:
        days_until_thu = 7
    return d + timedelta(days=days_until_thu)


# ---------------------------------------------------------------------------
# Sheet I/O helpers
# ---------------------------------------------------------------------------
def _ensure_sheet(wb: Workbook):
    if SHEET_NAME in wb.sheetnames:
        return wb[SHEET_NAME]
    ws = wb.create_sheet(SHEET_NAME)
    # Header row with bold + light fill so the tab is visually distinct
    fill = PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid")
    for i, h in enumerate(SHEET_COLS, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True)
        c.fill = fill
    ws.freeze_panes = "A2"
    return ws


def _existing_keys(ws) -> set:
    """URL+Date dedup against rows already in Weekly_Review."""
    if ws.max_row < 2:
        return set()
    headers = [str(c.value or "") for c in ws[1]]
    try:
        url_idx = headers.index("URL")
        date_idx = headers.index("Date")
    except ValueError:
        return set()
    keys: set = set()
    for r in ws.iter_rows(min_row=2, values_only=True):
        url = (str(r[url_idx]) if r[url_idx] else "").strip().lower()
        d = str(r[date_idx])[:10] if r[date_idx] else ""
        if url:
            keys.add((url, d))
    return keys


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------
def record_promotions(
    promoted_rows: List[Dict[str, Any]],
    xlsx_path: Path = XLSX_DEFAULT,
    json_path: Path = JSON_DEFAULT,
) -> int:
    """
    Append each just-promoted row to the Weekly_Review sheet, tagged
    with the upcoming Thursday review date. Then refresh the JSON slice
    so the Apps Script Thursday mailer always reads current data.

    Idempotent — calling twice with the same rows is a no-op (URL+Date
    dedup against existing Weekly_Review entries).
    """
    if not xlsx_path.exists():
        return 0

    week_end = review_thursday_for().isoformat()

    wb = openpyxl.load_workbook(xlsx_path)
    ws = _ensure_sheet(wb)
    seen = _existing_keys(ws)

    appended = 0
    for r in promoted_rows or []:
        url = (str(r.get("URL", "") or "")).strip().lower()
        d = str(r.get("Date", ""))[:10]
        if not url or (url, d) in seen:
            continue
        seen.add((url, d))
        row_out = [r.get(c, "") for c in RECALLS_COLS] + [week_end, "N"]
        ws.append(row_out)
        appended += 1

    if appended:
        # Preserve sheet order: Recalls, Pending, others, NEWS last
        # (matches save_xlsx_with_pending logic in merge_master.py).
        ordered = []
        for s in ("Recalls", "Pending"):
            if s in wb.sheetnames:
                ordered.append(s)
        for s in wb.sheetnames:
            if s in ordered or s == "NEWS":
                continue
            ordered.append(s)
        if "NEWS" in wb.sheetnames:
            ordered.append("NEWS")
        wb._sheets = [wb[s] for s in ordered]
        wb.save(xlsx_path)

    # Always refresh the JSON slice so the mailer sees the latest state
    # even if 0 rows appended this run (idempotent timestamp bump).
    export_week_slice(xlsx_path=xlsx_path, json_path=json_path,
                      week_end=week_end)
    return appended


def export_week_slice(
    xlsx_path: Path = XLSX_DEFAULT,
    json_path: Path = JSON_DEFAULT,
    week_end: Optional[str] = None,
) -> Dict[str, Any]:
    """
    Write a JSON slice of the upcoming/current Thu→Thu review window
    for the Apps Script Thursday-17:00 mailer to fetch.

    week_end: ISO date of the closing Thursday. Default = the Thursday
              the next promotion would land in (i.e. the upcoming review).
    """
    if week_end is None:
        week_end = review_thursday_for().isoformat()

    rows: List[Dict[str, Any]] = []
    if xlsx_path.exists():
        wb = openpyxl.load_workbook(xlsx_path, read_only=True)
        if SHEET_NAME in wb.sheetnames:
            ws = wb[SHEET_NAME]
            headers = [str(c.value or "") for c in ws[1]]
            try:
                we_idx = headers.index("Week_Added")
            except ValueError:
                we_idx = -1
            if we_idx >= 0:
                for r in ws.iter_rows(min_row=2, values_only=True):
                    if str(r[we_idx] or "") != week_end:
                        continue
                    obj = {}
                    for h, v in zip(headers, r):
                        if v is None:
                            obj[h] = ""
                        elif isinstance(v, (datetime, date)):
                            obj[h] = v.isoformat()
                        else:
                            obj[h] = v
                    rows.append(obj)

    # Sort by Tier (1 first), then Date desc, so the most-urgent rows
    # appear at the top of the email. Python's sort is stable, so we
    # do this in two passes: date desc, then tier asc.
    rows.sort(key=lambda r: str(r.get("Date", ""))[:10], reverse=True)
    def _tier_key(r):
        try:
            return int(r.get("Tier", 0) or 0) or 9
        except (TypeError, ValueError):
            return 9
    rows.sort(key=_tier_key)

    payload = {
        "week_end": week_end,
        "generated_utc": datetime.now(timezone.utc)
            .isoformat(timespec="seconds"),
        "row_count": len(rows),
        "tier1_count": sum(1 for r in rows if str(r.get("Tier", "")) == "1"),
        "outbreak_count": sum(1 for r in rows if str(r.get("Outbreak", "0"))
                                                 not in ("", "0", "0.0", "False", "false")),
        "dissent_count": sum(1 for r in rows
                             if "[dissent" in str(r.get("Notes", ""))),
        "rows": rows,
    }
    json_path.parent.mkdir(parents=True, exist_ok=True)
    json_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )
    return payload


def _neg_date(s: str) -> str:  # kept for back-compat; no longer used internally
    if not s or len(s) < 10:
        return "0000-00-00"
    return s


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import sys
    cmd = sys.argv[1] if len(sys.argv) > 1 else "export"
    if cmd == "export":
        result = export_week_slice()
        print(f"weekly-review-latest.json: {result['row_count']} rows "
              f"(Tier1={result['tier1_count']}, "
              f"Outbreak={result['outbreak_count']}, "
              f"Dissent={result['dissent_count']}) "
              f"for week ending {result['week_end']}")
    elif cmd == "thursday":
        print(review_thursday_for().isoformat())
    else:
        print("Usage: python -m pipeline.weekly_review_capture [export|thursday]",
              file=sys.stderr)
        sys.exit(2)
