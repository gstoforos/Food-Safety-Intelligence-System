"""
AFTS Food Safety Intelligence — Gap Finder
Orchestrator (parametric — supports any country via --country flag).

Wires the 4 pipeline stages end-to-end:
  Stage 1: news_scraper.collect_rss + collect_google_news → candidates.jsonl
  Stage 2: search_verifier.build_index + match_in_index   → verified.jsonl
  Stage 3: extractor.extract_one (rules + LLM)            → pending.jsonl + rejected.jsonl
  Stage 4: append rows to docs/data/recalls.xlsx          (idempotent, dedupe by URL)

All paths come from the CountryConfig. The xlsx file is shared across countries
(single AFTS knowledge base); the per-country JSONL outputs are isolated.

CLI:
    python -m pipeline.gap_finder.main --country gr
    python -m pipeline.gap_finder.main --country it --verbose
    python -m pipeline.gap_finder.main --country it --dry-run     # no xlsx writes
    python -m pipeline.gap_finder.main --country it --skip-news   # reuse existing JSONL
"""

from __future__ import annotations
import argparse
import json
import re
import sys
import time
import traceback
import unicodedata
from dataclasses import dataclass, field, asdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

try:
    from . import news_scraper, extractor
    from .countries import get as get_country
    from .countries.base import CountryConfig
except ImportError:
    from gap_finder import news_scraper, extractor                     # type: ignore
    from gap_finder.countries import get as get_country                # type: ignore
    from gap_finder.countries.base import CountryConfig                # type: ignore


# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────────────────────────

DEFAULT_XLSX = "docs/data/recalls.xlsx"
PENDING_SHEET = "Pending"
RECALLS_SHEET = "Recalls"
REJECTED_SHEET = "Weekly_Rejected"


# ─────────────────────────────────────────────────────────────────────────────
# RUN STATE — tallies for the final summary
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class RunState:
    country_code: str = ""
    country_name: str = ""
    started_at: str = ""
    candidates_found: int = 0
    verified_count: int = 0
    unmatched_count: int = 0
    extracted_accepted: int = 0
    extracted_rejected: int = 0
    appended_pending: int = 0
    appended_rejected: int = 0
    skipped_dupe_pending: int = 0
    skipped_dupe_recalls: int = 0
    skipped_dupe_rejected: int = 0
    errors: list[dict] = field(default_factory=list)

    def add_error(self, stage: str, exc: Exception) -> None:
        self.errors.append({
            "stage": stage, "type": type(exc).__name__, "msg": str(exc),
            "traceback": traceback.format_exc()[:500],
        })


# ─────────────────────────────────────────────────────────────────────────────
# XLSX I/O (idempotent append, dedupe by URL)
# ─────────────────────────────────────────────────────────────────────────────

def _load_existing_urls(xlsx_path: str, sheet_name: str) -> set[str]:
    """Read every URL already in the given sheet."""
    if not Path(xlsx_path).exists():
        return set()
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("  [WARN] openpyxl missing — cannot dedupe", file=sys.stderr)
        return set()
    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            return set()
        ws = wb[sheet_name]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        try:
            url_col = headers.index("URL")
        except ValueError:
            return set()
        urls: set[str] = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if url_col < len(row) and row[url_col]:
                urls.add(str(row[url_col]).strip())
        wb.close()
        return urls
    except Exception as e:
        print(f"  [WARN] dedupe read failed for {sheet_name}: {e}", file=sys.stderr)
        return set()


def _append_rows(
    xlsx_path: str, sheet_name: str, rows: list[dict], dedupe_against: set[str],
) -> tuple[int, int]:
    """Append rows to xlsx sheet, deduplicating by URL. Returns (appended, skipped)."""
    if not rows:
        return 0, 0
    try:
        from openpyxl import load_workbook, Workbook
    except ImportError:
        print("  [ERROR] openpyxl missing — cannot write xlsx", file=sys.stderr)
        return 0, 0

    p = Path(xlsx_path)
    p.parent.mkdir(parents=True, exist_ok=True)
    if p.exists():
        wb = load_workbook(xlsx_path)
    else:
        wb = Workbook()
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        # Headers = union of all row keys, with stable ordering
        headers = list(rows[0].keys())
        ws.append(headers)
    else:
        ws = wb[sheet_name]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    appended = skipped = 0
    for row in rows:
        url = (row.get("URL") or "").strip()
        if url and url in dedupe_against:
            skipped += 1
            continue
        ws.append([row.get(h, "") for h in headers])
        if url:
            dedupe_against.add(url)
        appended += 1

    wb.save(xlsx_path)
    return appended, skipped


def _recall_identity(row: dict) -> str:
    """Content fingerprint for one accepted recall, independent of which news
    outlet surfaced it. Multiple outlets cover the same EFET recall with
    different URLs/titles; without this they each become a separate Pending
    row a reviewer must merge by hand (run 27459015513 produced 7 rows for
    2 real recalls).

    Priority key: the authority recall ID embedded in Reason
    (e.g. 'Recall ID 842632') — that's the regulator's own unique handle.
    Fallback: normalized company + pathogen + first 3 product words.
    """
    import re as _re

    def _norm(x: str) -> str:
        x = (x or "").lower().strip()
        x = unicodedata.normalize("NFD", x)
        x = "".join(c for c in x if unicodedata.category(c) != "Mn")
        return _re.sub(r"\s+", " ", x)

    def _squash(x: str) -> str:
        # Aggressive: strip ALL non-alphanumeric so company legal-form
        # variants collapse — "ΜΑΒΕΕ" == "Μ.Α.Β.Ε.Ε." == "Μ.Α.Β.Ε.Ε" — and
        # outlet-to-outlet spelling/punctuation noise disappears.
        return _re.sub(r"[^a-z0-9α-ω]", "", _norm(x))

    company_sq = _squash(row.get("Company", ""))
    pathogen = _norm(row.get("Pathogen", ""))

    # The EFET recall ID is the regulator's own unique handle. When present
    # AND we have a company, combine them: same company + same bulletin ID =
    # same recall, regardless of how each outlet worded the product. This
    # collapses the multi-outlet duplicates that vary company punctuation and
    # product wording, WITHOUT over-merging two different firms that happen to
    # appear in one combined bulletin (company is part of the key).
    reason = row.get("Reason", "") or ""
    m = _re.search(r"(?:recall\s*id|id|allerta|pratica|αρ\.?\s*παρτιδας)"
                   r"\s*[:#]?\s*([0-9]{3,})", reason, _re.IGNORECASE)
    recall_id = m.group(1) if m else ""

    if company_sq and recall_id:
        return f"cid:{company_sq}|{recall_id}"
    if company_sq:
        # No ID — fall back to company + first 2 product words.
        product = " ".join(_norm(row.get("Product", "")).split()[:2])
        return f"cp:{company_sq}|{product}"
    # Empty extraction (LLM-failed rows): keep distinct by URL so a reviewer
    # sees each one for manual fixing rather than silently merging them.
    return f"url:{(row.get('URL') or '').strip()}"


def _richness(row: dict) -> int:
    """Score how complete an extraction is, so when collapsing duplicates we
    keep the best one. LLM-extraction-failed rows (empty company/brand) score
    lowest and get dropped in favour of a fully-extracted sibling."""
    score = 0
    for field in ("Company", "Brand", "Product", "Pathogen", "Region"):
        v = (row.get(field) or "").strip()
        if v and "extraction failed" not in v.lower():
            score += len(v)
    return score


def dedupe_by_recall_identity(rows: list[dict]) -> tuple[list[dict], int]:
    """Collapse rows that describe the same recall, keeping the richest.
    Returns (deduped_rows, n_collapsed)."""
    best: dict[str, dict] = {}
    order: list[str] = []
    for row in rows:
        key = _recall_identity(row)
        if key not in best:
            best[key] = row
            order.append(key)
        elif _richness(row) > _richness(best[key]):
            best[key] = row
    deduped = [best[k] for k in order]
    return deduped, len(rows) - len(deduped)


def write_pending_rows(rows: list[dict], xlsx_path: str) -> tuple[int, int, int]:
    """Write to Pending sheet. Skip URLs already in Pending OR Recalls.
    Returns (appended, skipped_pending_dupe, skipped_recalls_dupe)."""
    # Collapse same-recall duplicates from multiple news outlets first.
    rows, collapsed = dedupe_by_recall_identity(rows)
    if collapsed:
        print(f"  [dedup] collapsed {collapsed} duplicate-recall row(s) "
              f"→ {len(rows)} unique", file=sys.stderr)
    pending_urls = _load_existing_urls(xlsx_path, PENDING_SHEET)
    recalls_urls = _load_existing_urls(xlsx_path, RECALLS_SHEET)

    pre_recalls_skip = sum(1 for r in rows if (r.get("URL") or "").strip() in recalls_urls)
    rows_not_in_recalls = [r for r in rows if (r.get("URL") or "").strip() not in recalls_urls]

    appended, pending_skipped = _append_rows(
        xlsx_path, PENDING_SHEET, rows_not_in_recalls, pending_urls
    )
    return appended, pending_skipped, pre_recalls_skip


def write_rejected_rows(rows: list[dict], xlsx_path: str) -> tuple[int, int]:
    """Write to Weekly_Rejected sheet. Dedupe by URL."""
    rejected_urls = _load_existing_urls(xlsx_path, REJECTED_SHEET)
    return _append_rows(xlsx_path, REJECTED_SHEET, rows, rejected_urls)


# ─────────────────────────────────────────────────────────────────────────────
# JSONL I/O
# ─────────────────────────────────────────────────────────────────────────────

def read_jsonl(path: str) -> list[dict]:
    p = Path(path)
    if not p.exists():
        return []
    out: list[dict] = []
    with p.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                out.append(json.loads(line))
            except json.JSONDecodeError:
                continue
    return out


def append_jsonl(record: dict, path: str) -> None:
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    with p.open("a", encoding="utf-8") as f:
        f.write(json.dumps(record, ensure_ascii=False) + "\n")


# ─────────────────────────────────────────────────────────────────────────────
# PIPELINE STAGES
# ─────────────────────────────────────────────────────────────────────────────

def stage_news_scraper(cfg: CountryConfig, verbose: bool = False) -> list[dict]:
    print("\n=== Stage 1: News Scraper ===", file=sys.stderr)
    rss = news_scraper.collect_rss(cfg, verbose=verbose)
    gn = news_scraper.collect_google_news(cfg, verbose=verbose)
    all_cands = rss + gn
    deduped = news_scraper.deduplicate(all_cands)
    print(f"  RSS: {len(rss)}, Google News: {len(gn)}, deduped: {len(deduped)}",
          file=sys.stderr)
    news_scraper.write_jsonl(deduped, cfg.candidates_path)
    return [c.to_dict() for c in deduped]


def stage_article_fetch(
    candidates: list[dict], cfg: CountryConfig, verbose: bool = False,
    max_age_days: int = 14,
) -> tuple[list[dict], list[dict]]:
    """Stage 2: enrich candidates with body=title+description, filter recent, dedupe.

    Pragmatic design:
      - Google News URLs are JS-redirect proxies — can't fetch real article.
      - Direct-RSS feeds (ilfattoalimentare.it) provide rich <description> in RSS.
      - So body = title + description. No HTTP fetch needed.

    Critical filters (else Stage 3 runs 40+ min on historical noise):
      1. **Date filter**: drop candidates older than max_age_days. Gap finder
         only cares about NEW recalls — old ones already flowed through FSIS.
      2. **Title-fingerprint dedupe**: multiple news outlets cover the same
         recall. Keep ONE representative per story (the first encountered).

    Schema-compatible with old verified.jsonl so extractor.py works unchanged.
    """
    print(f"\n=== Stage 2: Candidate Enrichment ===", file=sys.stderr)
    if not candidates:
        return [], []

    now_utc = datetime.now(timezone.utc)
    cutoff = now_utc.timestamp() - (max_age_days * 86400)

    def parse_pub(s: str) -> Optional[float]:
        """Best-effort RFC822/ISO → UNIX timestamp."""
        if not s:
            return None
        for fmt in ["%a, %d %b %Y %H:%M:%S %Z",
                    "%a, %d %b %Y %H:%M:%S %z",
                    "%Y-%m-%dT%H:%M:%S%z",
                    "%Y-%m-%dT%H:%M:%SZ", "%Y-%m-%d"]:
            try:
                dt = datetime.strptime(s, fmt)
                if dt.tzinfo is None:
                    dt = dt.replace(tzinfo=timezone.utc)
                return dt.timestamp()
            except ValueError:
                continue
        return None

    def title_tokens(t: str) -> set[str]:
        """Distinctive tokens for similarity comparison."""
        import unicodedata
        n = unicodedata.normalize("NFD", t.lower())
        n = "".join(c for c in n if unicodedata.category(c) != "Mn")
        return set(re.findall(r"[a-z0-9]{4,}", n))

    # Italian stopwords that shouldn't count toward title similarity
    STOP_IT = {"richiamo", "richiamato", "richiamata", "richiamati", "ritiro",
               "ritirato", "ritirata", "ritirati", "lotto", "lotti", "mercato",
               "supermercati", "supermercato", "marca", "marchio", "prodotto",
               "prodotti", "ministero", "salute", "rischio", "alimentare",
               "presenza", "possibile"}

    def is_duplicate(tokens: set[str], seen_token_sets: list[set[str]]) -> bool:
        """Return True if this title shares ≥40% meaningful tokens with any prior."""
        sig = tokens - STOP_IT
        if len(sig) < 2:
            return False
        for prev in seen_token_sets:
            prev_sig = prev - STOP_IT
            if not prev_sig:
                continue
            overlap = len(sig & prev_sig) / max(len(sig | prev_sig), 1)
            if overlap >= 0.4:
                return True
        return False

    # Hard cap: even after date + dedup, never process more than this many
    # records through Stage 3 (LLM extraction). Each LLM call is 5-10s; at 50
    # records that's 4-8 min Stage 3, comfortably inside workflow timeout.
    MAX_RECORDS = 50

    too_old = 0
    too_old_no_date = 0
    seen_token_sets: list[set[str]] = []
    dup_skipped = 0
    capped = 0
    verified: list[dict] = []
    now_iso = now_utc.isoformat()

    # Sort candidates by published date desc (newest first) so the cap keeps
    # the most recent records when there are too many.
    def sort_key(c):
        ts = parse_pub(c.get("published", ""))
        return ts if ts is not None else 0
    sorted_candidates = sorted(candidates, key=sort_key, reverse=True)

    # ── A. Cheap pre-filter: date + dedupe BEFORE we spend HTTP fetches ─────
    pre_filtered: list[dict] = []
    for cand in sorted_candidates:
        if len(pre_filtered) >= MAX_RECORDS:
            capped += 1
            continue

        pub_ts = parse_pub(cand.get("published", ""))
        if pub_ts is None:
            too_old_no_date += 1
        elif pub_ts < cutoff:
            too_old += 1
            continue

        title = cand.get("title", "")
        tokens = title_tokens(title)
        if is_duplicate(tokens, seen_token_sets):
            dup_skipped += 1
            continue
        seen_token_sets.append(tokens)
        pre_filtered.append(cand)

    # ── B. Parallel HTTP fetch via article_fetcher (the REAL enrichment) ────
    # Google News redirector URLs get followed; real publisher HTML gets
    # extracted via BeautifulSoup heuristics. Bodies typically 2000-5000
    # chars — long enough to contain the hazard term that titles abbreviate.
    print(f"  fetching {len(pre_filtered)} article bodies in parallel…",
          file=sys.stderr)
    try:
        from . import article_fetcher
    except ImportError:
        import pipeline.gap_finder.article_fetcher as article_fetcher  # type: ignore
    enriched_recs, failed_recs = article_fetcher.enrich_all(
        pre_filtered, cfg, verbose=verbose)
    fetched_ok = len(enriched_recs)
    fetched_fail = len(failed_recs)

    # ── C. Convert enriched records to dict (efet_body is real article text)
    for rec in enriched_recs:
        d = rec.to_dict()
        d["discovered_via"] = "article_fetch"
        verified.append(d)

    # ── D. Fallback: keep failed-fetch candidates with RSS-description body
    # so we don't lose visibility on them. Title+description is better than
    # nothing — the classifier still has a shot at catching the hazard.
    for cand in failed_recs:
        title = cand.get("title", "")
        description = cand.get("description", "")
        body_parts = [title]
        if description and description.lower() != title.lower():
            body_parts.append(description)
        body = "\n\n".join(body_parts)
        pub_ts = parse_pub(cand.get("published", ""))
        date_iso = ""
        if pub_ts is not None:
            date_iso = datetime.fromtimestamp(pub_ts, tz=timezone.utc).strftime("%Y-%m-%d")
        record = {
            "news_url": cand.get("url", ""),
            "news_title": title,
            "news_published": cand.get("published", ""),
            "news_source_domain": cand.get("source_domain", ""),
            "efet_url": cand.get("url", ""),
            "efet_title": title,
            "efet_date_iso": date_iso,
            "efet_body": body,
            "match_score": 1.0,
            "matched_at": now_iso,
            "discovered_via": "rss_fallback",
            "fetch_status": cand.get("fetch_status", "failed"),
        }
        verified.append(record)

    body_lens = [len(r["efet_body"]) for r in verified]
    avg_len = sum(body_lens) // len(body_lens) if body_lens else 0
    rich = sum(1 for l in body_lens if l > 200)
    print(f"  candidates in:      {len(candidates)}", file=sys.stderr)
    print(f"  too old (>{max_age_days}d):     {too_old}", file=sys.stderr)
    print(f"  no date (kept):     {too_old_no_date}", file=sys.stderr)
    print(f"  dup-title skipped:  {dup_skipped}", file=sys.stderr)
    print(f"  capped (>{MAX_RECORDS}):       {capped}", file=sys.stderr)
    print(f"  HTTP fetched ok:    {fetched_ok}", file=sys.stderr)
    print(f"  HTTP fetched fail:  {fetched_fail} (fallback to RSS desc)",
          file=sys.stderr)
    print(f"  enriched out:       {len(verified)} (avg body {avg_len} chars, "
          f"{rich} with >200 chars)", file=sys.stderr)

    p = Path(cfg.verified_path)
    p.parent.mkdir(parents=True, exist_ok=True)
    with p.open("w", encoding="utf-8") as f:
        for r in verified:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")

    return verified, []


def stage_extractor(
    verified: list[dict], cfg: CountryConfig, verbose: bool = False,
) -> tuple[list[dict], list[dict]]:
    print("\n=== Stage 3: Extractor ===", file=sys.stderr)
    if not verified:
        return [], []

    client = extractor.LlamaClient()
    print(f"  LlamaClient: {client.base_url}  model={client.model}",
          file=sys.stderr)
    if not client.health():
        print(f"  ERROR: cannot reach llama-server at {client.base_url}",
              file=sys.stderr)
        return [], []

    pending_rows: list[dict] = []
    rejected_rows: list[dict] = []

    for i, v in enumerate(verified, 1):
        if verbose:
            print(f"  [{i}/{len(verified)}] {v.get('efet_title', '')[:60]}",
                  file=sys.stderr)
        try:
            pending, rejected = extractor.extract_one(v, client, cfg, verbose=verbose)
        except Exception as e:
            print(f"  ERROR record {i}: {e}", file=sys.stderr)
            continue
        if pending:
            pending_rows.append(pending)
        if rejected:
            rejected_rows.append(rejected)

    extractor.write_jsonl(pending_rows, cfg.pending_path)
    extractor.write_jsonl(rejected_rows, cfg.rejected_path)

    print(f"  accepted: {len(pending_rows)}, rejected: {len(rejected_rows)}",
          file=sys.stderr)
    return pending_rows, rejected_rows


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main() -> int:
    parser = argparse.ArgumentParser(
        description="AFTS Gap Finder — Orchestrator (parametric)"
    )
    parser.add_argument("--country", required=True,
                        help="ISO2 country code: gr, it, ...")
    parser.add_argument("--xlsx", default=DEFAULT_XLSX,
                        help=f"Path to recalls.xlsx (default: {DEFAULT_XLSX})")
    parser.add_argument("--dry-run", action="store_true",
                        help="Run pipeline but skip xlsx writes")
    parser.add_argument("--skip-news", action="store_true",
                        help="Skip Stage 1 — reuse existing candidates.jsonl")
    parser.add_argument("--skip-verify", action="store_true",
                        help="Skip Stage 2 — reuse existing verified.jsonl")
    parser.add_argument("--max-age-days", type=int, default=14,
                        help="Drop candidates older than N days (default: 14). "
                             "Gap finder only cares about NEW recalls.")
    parser.add_argument("--verbose", "-v", action="store_true")
    args = parser.parse_args()

    cfg = get_country(args.country)
    state = RunState(
        country_code=cfg.code, country_name=cfg.name_en,
        started_at=datetime.now(timezone.utc).isoformat(),
    )

    print(f"=== AFTS Gap Finder ({cfg.code}/{cfg.name_en}) "
          f"started {state.started_at} ===", file=sys.stderr)
    print(f"  authority: {cfg.authority_short} ({cfg.authority_domain})",
          file=sys.stderr)
    print(f"  xlsx:      {args.xlsx}", file=sys.stderr)
    print(f"  dry-run:   {args.dry_run}", file=sys.stderr)

    # ── Stage 1: News ───────────────────────────────────────────────────────
    if args.skip_news:
        candidates = read_jsonl(cfg.candidates_path)
        print(f"\n[Stage 1] SKIPPED — loaded {len(candidates)} candidates",
              file=sys.stderr)
    else:
        try:
            candidates = stage_news_scraper(cfg, verbose=args.verbose)
        except Exception as e:
            state.add_error("news_scraper", e)
            print(f"  ERROR: {e}", file=sys.stderr)
            candidates = []
    state.candidates_found = len(candidates)

    # ── Stage 2: Article Fetch ──────────────────────────────────────────────
    if args.skip_verify:
        verified = read_jsonl(cfg.verified_path)
        unmatched = read_jsonl(cfg.unmatched_path)
        print(f"\n[Stage 2] SKIPPED — loaded {len(verified)} verified, "
              f"{len(unmatched)} unmatched", file=sys.stderr)
    else:
        try:
            verified, unmatched = stage_article_fetch(
                candidates, cfg,
                verbose=args.verbose,
                max_age_days=args.max_age_days,
            )
        except Exception as e:
            state.add_error("article_fetcher", e)
            print(f"  ERROR: {e}", file=sys.stderr)
            verified, unmatched = [], []
    state.verified_count = len(verified)
    state.unmatched_count = len(unmatched)

    # ── Stage 3: Extract ────────────────────────────────────────────────────
    try:
        pending_rows, rejected_rows = stage_extractor(verified, cfg, verbose=args.verbose)
    except Exception as e:
        state.add_error("extractor", e)
        print(f"  ERROR: {e}", file=sys.stderr)
        pending_rows, rejected_rows = [], []
    state.extracted_accepted = len(pending_rows)
    state.extracted_rejected = len(rejected_rows)

    # ── Stage 4: Write xlsx ─────────────────────────────────────────────────
    print(f"\n=== Stage 4: Write to {args.xlsx} ===", file=sys.stderr)
    if args.dry_run:
        print(f"  DRY RUN — no xlsx writes performed", file=sys.stderr)
        print(f"  would append: {len(pending_rows)} pending, "
              f"{len(rejected_rows)} rejected", file=sys.stderr)
    else:
        try:
            app_p, skip_pending, skip_recalls = write_pending_rows(
                pending_rows, args.xlsx
            )
            state.appended_pending = app_p
            state.skipped_dupe_pending = skip_pending
            state.skipped_dupe_recalls = skip_recalls

            app_r, skip_rej = write_rejected_rows(rejected_rows, args.xlsx)
            state.appended_rejected = app_r
            state.skipped_dupe_rejected = skip_rej
        except Exception as e:
            state.add_error("write_xlsx", e)
            print(f"  ERROR writing xlsx: {e}", file=sys.stderr)

    # ── Summary ─────────────────────────────────────────────────────────────
    print("\n" + "=" * 70, file=sys.stderr)
    print("RUN SUMMARY", file=sys.stderr)
    print("=" * 70, file=sys.stderr)
    print(f"  country:                 {cfg.code} ({cfg.name_en})", file=sys.stderr)
    print(f"  candidates found:        {state.candidates_found}", file=sys.stderr)
    print(f"  verified (matched):      {state.verified_count}", file=sys.stderr)
    print(f"  unmatched:               {state.unmatched_count}", file=sys.stderr)
    print(f"  extracted accepted:      {state.extracted_accepted}", file=sys.stderr)
    print(f"  extracted rejected:      {state.extracted_rejected}", file=sys.stderr)
    print(f"  → appended Pending:      {state.appended_pending}", file=sys.stderr)
    print(f"  → appended Rejected:     {state.appended_rejected}", file=sys.stderr)
    print(f"  skipped (dupe Pending):  {state.skipped_dupe_pending}", file=sys.stderr)
    print(f"  skipped (dupe Recalls):  {state.skipped_dupe_recalls}", file=sys.stderr)
    print(f"  skipped (dupe Rejected): {state.skipped_dupe_rejected}", file=sys.stderr)
    print(f"  errors:                  {len(state.errors)}", file=sys.stderr)
    print("=" * 70, file=sys.stderr)

    # ── Run log ─────────────────────────────────────────────────────────────
    log_record = asdict(state)
    log_record["finished_at"] = datetime.now(timezone.utc).isoformat()
    append_jsonl(log_record, cfg.run_log_path)

    return 1 if state.errors else 0


if __name__ == "__main__":
    sys.exit(main())
