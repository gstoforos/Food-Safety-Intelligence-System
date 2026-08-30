"""
Master merge logic (Pending-sheet architecture).

recalls.xlsx holds THREE sheets:
  - Recalls  : approved, published data (consumed by the weekly report)
  - Pending  : freshly scraped rows awaiting validation + review
  - NEWS     : unrelated news-feed sheet, preserved as-is

Daily pipeline flow:
  1. Scrapers write to Pending (via append_to_pending)
  2. Enrichment + URL validation + AI review run against Pending
  3. promote_approved() moves rows that pass all checks into Recalls
  4. Rejected rows stay in Pending with a rejection reason stored in Notes
     (prefixed "REJECTED: <reason> | <original notes>") so a human can triage.

Dedup:
  - Primary key: URL (lowercased, stripped)
  - Fallback:    date + company + pathogen
  - Dedup applies within Pending and across Pending->Recalls promotion.
"""
from __future__ import annotations
import json
import logging
import re
import unicodedata
from datetime import datetime, timezone
from pathlib import Path
from typing import List, Dict, Any, Tuple, Optional
from urllib.parse import urlparse
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill

from scrapers._models import Recall

log = logging.getLogger(__name__)

SCHEMA = ["Date", "Source", "Company", "Brand", "Product", "Pathogen", "Reason",
          "Class", "Country", "Region", "Tier", "Outbreak", "URL", "Notes"]

# ── Internal-only tracking columns (Recalls sheet only) ──────────────────
# These columns are appended to the Recalls sheet for internal bookkeeping
# and are EXCLUDED from the public-facing recalls.json that feeds the
# dashboard. See mirror_json_from_xlsx() for the filtering.
#
# DateAdded   — date the row was first promoted to Recalls (set once,
#               never changed). Used to distinguish original publication
#               date (Date) from when FSIS captured it.
# LastUpdated — date the row was last modified (any field changed). Set
#               by promote_approved on insert and by audit/fix code paths
#               that touch existing rows.
# LastChecked — date a URL gate (Gemini grounded check or url_guardian
#               reachability check) last validated this row's URL. Used
#               by url_guardian to skip rows checked recently and avoid
#               redundant Gemini grounded calls.
# report_week — sticky stamp identifying which weekly report this row
#               belongs to (e.g. "W19", "W20"). Set ONCE at promote time
#               from the row's Date using compute_report_week() and never
#               overwritten — even if the row is later edited or the URL
#               re-validated. The weekly builder filters by this stamp
#               instead of doing date-window math, so late-arriving rows
#               (e.g. scraped Wednesday for an event published the prior
#               Thursday) correctly land in the right week's report.
#               See compute_report_week() below for the rule.
#               Introduced 2026-05-10 per operator instruction.
# Analytical columns written by pipeline/enrich_schema.py. They exist for
# statistical stratification and are NOT part of what FSIS publishes:
# mirror_json_from_xlsx strips every name in this list from recalls.json,
# and the two public-xlsx builders use allow-lists that do not name them.
# Adding a column here is the ONLY thing that keeps it out of the dashboard.
_ENRICHMENT_COLUMNS = [
    "FoodCategory", "ProcessType", "ConsumptionState", "StorageCondition",
    "PackagingType", "PackagingForm", "PreservationSystem",
    "HazardGroup", "HazardCertainty",
    "NoticeType", "SeverityClass", "EventID",
    "EnrichedBy", "EnrichedAt", "EnrichmentTier",
]

RECALLS_INTERNAL_COLUMNS = ["DateAdded", "LastUpdated", "LastChecked",
                            "report_week"] + _ENRICHMENT_COLUMNS
RECALLS_SCHEMA = SCHEMA + RECALLS_INTERNAL_COLUMNS

# Pending sheet has the same columns plus three tracking columns.
# Audit 2026-05-05 — added RejectedBy column to track which reviewers have
# rejected this row. Stored as a comma-separated set of reviewer names
# (e.g. "claude-check,gemini-url-gate"). Used by mark_rejected_with_counter
# to physically delete rows once 2+ DIFFERENT reviewers have rejected.
PENDING_SCHEMA = SCHEMA + ["ScrapedAt", "Status", "RejectedBy"]

# ── Rejected sheet — DEPRECATED 2026-05-11 ─────────────────────────────
# The standalone "Rejected" sheet was removed (per operator decision)
# in favor of Weekly_Rejected as the single source of rejection truth.
# REJECTED_SCHEMA is kept here only because external scripts may still
# import the name. New code should reference WEEKLY_REJECTED_HEADERS
# in pipeline/weekly_rejected_capture.py instead.
REJECTED_SCHEMA = SCHEMA + ["ScrapedAt", "Status", "RejectedBy", "RejectedAt"]

NEWS_HEADERS = ["Published (UTC)", "Pathogen", "Event", "Source", "Title",
                "Link", "Retrieved (UTC)"]

# Status values used in the Pending sheet
STATUS_PENDING = "pending"
STATUS_APPROVED = "approved"   # transient — promoted rows are removed from Pending
STATUS_REJECTED = "rejected"

# ── Gap-finder gating state machine (audit 2026-04-29) ──────────────────
# Gap-finder rows (Tavily/Exa/Gemini/Claude/OpenAI search-based recall
# discovery) are LESS trustworthy than scraper rows because they come from
# search-engine indexes, not from authoritative regulator pages. Per
# operator policy, they must NOT be auto-promoted to Recalls. Instead,
# they sit in Pending under one of these gating states until they pass:
#   1. Two independent Gemini URL grounding checks (different runs,
#      different model invocations — non-determinism is the point), then
#   2. One Claude page-content verification.
# State transitions (see promote_gap_rows.py / url_gate_gemini.py /
# claude_check.py for the implementations):
#   pending_gap     -- written by gap-finders; first url_gate run will
#                      either advance to pending_gap_v1 or reject.
#   pending_gap_v1  -- one Gemini URL pass. Second url_gate run advances
#                      to pending_gap_v2 or rejects.
#   pending_gap_v2  -- two Gemini URL passes. claude_check advances to
#                      pending (eligible for normal merge) or rejects.
# promote_approved skips ANY pending_gap* row — they never reach Recalls
# until claude_check has flipped them back to plain "pending".
STATUS_PENDING_GAP    = "pending_gap"
STATUS_PENDING_GAP_V1 = "pending_gap_v1"
STATUS_PENDING_GAP_V2 = "pending_gap_v2"

# ── Enrichment gating (audit 2026-05-08) ───────────────────────────────
# Scraper rows that arrive with an empty Pathogen field — typically from
# HTML-listing fallback paths (FSAI, INVIMA, RappelConso "sans marque",
# CFIA L1 placeholder rows) — used to be hard-rejected at the pathogen
# scope gate. That meant signal got dumped on the floor whenever the
# orchestrator ran without claude_check (`ai=False, review=False`).
# Now they're admitted to Pending under this status. claude_check and
# the Gemini enrichers look for STATUS_PENDING_ENRICHMENT rows, fill in
# Pathogen from the page content, and flip the status back to plain
# "pending" (or rejected if the pathogen turns out to be out of scope).
# promote_approved skips this status — same structural guarantee that
# protects pending_gap* rows.
STATUS_PENDING_ENRICHMENT = "pending_enrichment"

# Sentinel reason returned by validate_pending_row for rows that pass
# every other gate but have an empty Pathogen field. append_to_pending
# converts this into Status=STATUS_PENDING_ENRICHMENT instead of the
# default Status=STATUS_PENDING.
OK_PENDING_ENRICHMENT = "OK_PENDING_ENRICHMENT"

# ── Transient-failure parking (audit 2026-07-28) ───────────────────────
# A row whose content check could not COMPLETE — the page fetch failed, the
# reviewer API errored, the response didn't parse — must not auto-promote
# (nothing verified it), but it is not a gap-finder row either.
#
# Before this status existed, claude_check parked such rows in
# STATUS_PENDING_GAP_V2. That was a DEMOTION: a normal scraper row got
# pushed into the gap-finder state machine, whose only exit is a
# successful Claude pass. When the fetch failure was permanent rather
# than transient the row could never escape:
#
#     fetch fails -> SKIP -> parked in pending_gap_v2 -> blocked from
#     promotion -> exit requires a Claude pass -> which requires the
#     fetch -> which still fails.
#
# 2026-07-24..28: rappel.conso.gouv.fr began serving an incomplete TLS
# chain. 24 rows locked solid — 16 Listeria, 3 STEC, 2 Salmonella,
# Norovirus, Ochratoxin, Aflatoxin — while every run reported a green
# "+0 promoted". Publication stopped for four days and nothing alarmed.
#
# STATUS_PENDING_RETRY keeps the fail-closed guarantee (still
# non-promotable) while preserving the row's identity as an ordinary
# pending row: claude_check flips it straight back to "pending" on the
# next successful pass, in the same run, so it promotes immediately.
STATUS_PENDING_RETRY = "pending_retry"

# ── Transient-failure parking (audit 2026-07-28) ───────────────────────
# A row whose verification could not COMPLETE (fetch failed, reviewer API
# errored) must not auto-promote, but it is not a gap-finder row either.
# Parking it in STATUS_PENDING_GAP_V2 was a DEMOTION whose only exit is a
# successful Claude pass — a closed loop when the fetch failure is
# permanent. This status is equally non-promotable but keeps the row's
# identity, so claude_check flips it straight back to "pending" on the
# next successful pass and it promotes in the same run.
STATUS_PENDING_RETRY = "pending_retry"

GAP_GATING_STATUSES = frozenset({
    STATUS_PENDING_GAP, STATUS_PENDING_GAP_V1, STATUS_PENDING_GAP_V2,
})

# Any non-promotable Pending status. promote_approved uses this set as
# the structural lock against auto-promotion. Add new gating statuses
# here, NOT to GAP_GATING_STATUSES (which retains its narrow gap-finder
# semantics for code that introspects it).
NON_PROMOTABLE_STATUSES = GAP_GATING_STATUSES | frozenset({
    STATUS_PENDING_ENRICHMENT,
    STATUS_PENDING_RETRY,
})


# ---------------------------------------------------------------------------
# Weekly-report stamping (audit 2026-05-10)
# ---------------------------------------------------------------------------
# Each row promoted to Recalls is stamped with `report_week` so the weekly
# builder can filter by the stamp instead of computing a date window. This
# matters for late-arriving rows: a row scraped Wednesday for an event
# published the prior Thursday must end up in LAST week's report — not this
# week's. Stamping at promote time (from the row's Date, not the scrape
# time) makes that automatic.
#
# RULE: report_week = "W{nn}" where nn = ISO week number of the SMALLEST
# Friday F such that F > row_date (strictly greater, not >=).
#
# Why "strictly greater": each weekly report ships Friday morning. A row
# dated Friday F itself isn't yet captured by F's AM scrape, so it lands
# in the NEXT week's report. Verifications (with date → next-Fri-strict
# → ISO week of that Friday):
#   2026-05-01  Fri  →  2026-05-08  →  W19
#   2026-05-02  Sat  →  2026-05-08  →  W19
#   2026-05-07  Thu  →  2026-05-08  →  W19
#   2026-05-08  Fri  →  2026-05-15  →  W20
#   2026-05-14  Thu  →  2026-05-15  →  W20
#
# The stamp is STICKY — once set, never recomputed. URL re-validation,
# pathogen enrichment, and any other audit-fix code paths that touch an
# existing Recalls row must NEVER overwrite report_week.
#
# All Recalls insertions go through promote_approved() — there's no other
# code path that writes to Recalls — so stamping there covers every row.
# That includes the audit-2026-05-11 pending_enrichment unlock path: when
# stranded rows get their Status flipped back to "pending" and Claude's
# clean-row shortcut approves them, they pass through promote_approved
# and get stamped from their own Date (not the scrape time), so a
# previously-stuck row published 2026-05-05 stamps as W19 even if the
# unlock happens in week 21.
def compute_report_week(date_str: str) -> str:
    """Return the report_week stamp for a given row date.

    Args:
        date_str: ISO date string (YYYY-MM-DD) or anything coerce-able.

    Returns:
        "W{nn}" where nn is zero-padded ISO week of the next-Friday-after.
        Empty string if the date can't be parsed.
    """
    if not date_str:
        return ""
    try:
        d = datetime.strptime(str(date_str)[:10], "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return ""
    from datetime import timedelta as _td
    # Mon=0 .. Fri=4 .. Sun=6
    days_until_next_friday = (4 - d.weekday()) % 7
    if days_until_next_friday == 0:
        # d itself is Friday — strict next is +7
        days_until_next_friday = 7
    next_friday = d + _td(days=days_until_next_friday)
    return f"W{next_friday.isocalendar()[1]:02d}"


# ---------------------------------------------------------------------------
# Dedup
# ---------------------------------------------------------------------------
# Content-identity dedup for agencies with no stable alert id in the URL
# (audit 2026-07-09). Host-scoped: every other regulator is unaffected.
try:  # pragma: no cover
    from pipeline._url_identity import (
        has_stable_id as _identity_has_stable_id,
        content_key as _identity_content_key,
    )
except Exception:  # pragma: no cover
    def _identity_has_stable_id(url: str) -> bool:
        return True

    def _identity_content_key(row: Dict[str, Any]) -> str:
        return ""


def _normalize_url_for_dedup(url: str) -> str:
    """Normalize a URL for dedup comparison.

    Audit 2026-05-06: previously the dedup key was just url.strip().lower().
    That missed http:// vs https:// duplicates — production showed a Tavily-
    sourced http://www.fsis.usda.gov/... slipping past the gate even though
    the https:// version was already in Recalls.

    Normalizations applied (all lowercase):
      • strip protocol (http:// → '', https:// → '')
      • strip leading 'www.'
      • strip trailing '/'
      • strip URL fragment '#...'
      • strip non-canonical query strings (utm_*, ref, etc.) but PRESERVE
        identifier-style query params (permalink, id, fiche, ref, recall_id,
        search_api_fulltext)

    Audit 2026-07-26 — TWO OVER-COLLAPSES FOUND, each discarding the only
    thing distinguishing one recall from another on the same host:

      1. FDA serves individual recalls off ONE search endpoint, with the
         recall number in `search_api_fulltext` (H-0700-2026 / H-0699 /
         H-0698 — three separate California Dairies recalls). That param
         was not a keeper, so all three collapsed to the bare search URL
         and two could be silently dropped as "duplicates" on a promote.
      2. Italy's Salute publishes several recalls inside ONE shared PDF,
         addressing each by fragment (flax seeds / hydrocyanic acid vs
         sheep cheese / Listeria). Stripping the fragment collapsed them.

    Fragments are normally presentation-only and must still be stripped
    (that is what makes ...page#section a duplicate of ...page). The narrow
    exception is a document URL: on a .pdf path the host serves a shared
    file and the fragment is the per-recall anchor, so it is identity-
    bearing and is preserved.
    """
    if not url:
        return ""
    s = url.strip().lower()
    base, sep, frag = s.partition("#")
    keep_frag = bool(sep and frag and base.split("?", 1)[0].endswith(".pdf"))
    s = base
    if s.startswith("https://"):
        s = s[8:]
    elif s.startswith("http://"):
        s = s[7:]
    if s.startswith("www."):
        s = s[4:]
    if "?" in s:
        path, _, query = s.partition("?")
        keepers = []
        for kv in query.split("&"):
            k = kv.split("=", 1)[0]
            if k in ("permalink", "id", "fiche", "ref", "recall_id",
                     "search_api_fulltext"):
                keepers.append(kv)
        s = path + (("?" + "&".join(keepers)) if keepers else "")
    if s.endswith("/"):
        s = s[:-1]

    # ── FSANZ amended-alert republication (audit 2026-08-02) ───────────────
    # Food Standards Australia New Zealand does not update a recall alert in
    # place. When an alert is amended it is republished at a NEW slug carrying
    # a status prefix, and BOTH pages stay live:
    #
    #   .../recall-alert/auxico-perth-pty-ltd-lgm-hot-chilli-oil-275g
    #   .../recall-alert/updated-300726-auxico-perth-pty-ltd-lgm-hot-chilli-oil-275g
    #
    # Same company, same product, same recall — one hazard, two addresses. The
    # URL-keyed dedup saw a second address and minted a second row, which then
    # took an independent trip through the reviewers and came back carrying an
    # invented Pathogen at Tier 1 while the correct original sat two rows away
    # in the same sheet. It went out in a subscriber alert.
    #
    # Stripping the status segment collapses the two. Scoped to the FSANZ host
    # and the /recall-alert/ path so no other regulator's slugs are touched,
    # and anchored at the START of the slug so a product legitimately named
    # "...updated..." mid-slug is unaffected.
    if s.startswith("foodstandards.gov.au/food-recalls/recall-alert/"):
        head, _, slug = s.rpartition("/")
        stripped = re.sub(r"^(?:updated?|update|revised|corrected|amended)"
                          r"-\d{4,8}-", "", slug)
        if stripped and stripped != slug:
            s = head + "/" + stripped

    if keep_frag:
        s = s + "#" + frag
    return s


def _dedup_key(row: Dict[str, Any]) -> str:
    """
    URL primary (normalized), fallback to date+company+pathogen.

    Exception (audit 2026-07-09): hosts that carry NO stable alert identifier
    in their URLs are keyed on row CONTENT instead. For those agencies the same
    alert is reachable at several distinct URLs, so a URL key mints a fresh row
    on every rediscovery. fsai.ie is the only such host today; every other
    regulator keeps the exact URL-keyed behaviour it had before.
    """
    raw_url = str(row.get("URL") or "").strip()
    if raw_url and not _identity_has_stable_id(raw_url):
        return _identity_content_key(row)
    url = _normalize_url_for_dedup(raw_url)
    if url:
        return url
    co = unicodedata.normalize("NFD", row.get("Company") or "").encode("ascii", "ignore").decode().lower()
    co = re.sub(r"[^a-z0-9]", "", co)[:30]
    return f"{row.get('Date','')}|{co}|{(row.get('Pathogen','') or '')[:30]}"


# ---------------------------------------------------------------------------
# Date-consistency check (audit 2026-05-06: defense-in-depth)
# ---------------------------------------------------------------------------
# Production failure 2026-05-06: two USDA FSIS recalls from 2018 ("Oct. 19,
# 2018" Envolve Foods Listeria + "March 29, 2018" Target Corp Listeria)
# were promoted from Pending to Recalls and surfaced on the public dashboard
# with Date=2026-05-06 (today's date stamped by the Tavily date-extractor
# fallback). Three independent gates failed:
#   1. Tavily date-extractor regex didn't accept "Oct." (period after abbrev)
#   2. URL gate Gemini returned prose; parse-failure path defaulted to PASS
#   3. Pending → Recalls promotion had NO date-sanity check at all
#
# This function is the third-line defense. It compares the Date field
# against any date pattern found in Notes. If Notes mentions a year that's
# more than 1 year older than Date's year, the promotion is rejected and
# the row stays in Pending for manual review.

# Date-extraction regex used at the promotion gate. Matches:
#   "Oct. 19, 2018"   "March 29, 2018"   "2018-10-19"   "19 March 2018"
_PROMOTION_OLD_DATE_RX = re.compile(
    r"\b(?:"
    r"(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec|"
    r"january|february|march|april|june|july|august|"
    r"september|october|november|december)\.?\s+\d{1,2}[,\s]+(\d{4})"
    r"|"
    r"(\d{4})-(?:0[1-9]|1[0-2])-(?:0[1-9]|[12]\d|3[01])"
    r"|"
    r"\d{1,2}\s+(?:jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec|"
    r"january|february|march|april|june|july|august|"
    r"september|october|november|december)\.?\s+(\d{4})"
    r")\b",
    re.IGNORECASE,
)


def _check_date_consistency(row: Dict[str, Any]) -> Optional[str]:
    """Reject promotion when Notes mentions a year far older than Date field.

    Returns None if the row is consistent (or no date found in Notes).
    Returns a rejection-reason string if inconsistent.

    Heuristic: scan Notes for date patterns; if the OLDEST year mentioned
    is more than 1 calendar year older than the Date field's year, the
    Date field is almost certainly a date-extractor fallback (today's
    date stamped on an old archived page).

    Conservative — only rejects when the gap is unambiguous (>1 year).
    Rows where Notes mentions e.g. "in 2025" while Date is 2026 will pass.
    """
    date_field = str(row.get("Date") or "")[:10]
    notes = str(row.get("Notes") or "")
    if not date_field or not notes:
        return None
    try:
        date_year = int(date_field[:4])
    except ValueError:
        return None

    # Find all 4-digit years mentioned in date contexts in Notes
    years_found = []
    for m in _PROMOTION_OLD_DATE_RX.finditer(notes):
        for g in m.groups():
            if g and g.isdigit() and len(g) == 4:
                yr = int(g)
                if 2000 <= yr <= date_year + 1:
                    years_found.append(yr)
                break

    if not years_found:
        return None

    oldest = min(years_found)
    if oldest <= date_year - 2:
        return (f"date_inconsistent: Date field is {date_field} but Notes "
                f"mentions {oldest} (likely date-extractor fallback on an "
                f"archived page)")
    return None


# ---------------------------------------------------------------------------
# Near-duplicate detection (catches same-recall-different-URL cases)
# ---------------------------------------------------------------------------
# Why this exists: regulators sometimes publish the same recall under two
# URL formats (FDA's canonical company-slug URL vs. their share-link
# "?permalink=<hash>" wrapper). The OpenAI/search-based gap finders find
# the wrapper URL while the direct scraper finds the canonical URL —
# string dedup can't catch this because the URLs are entirely different
# paths. Even the date+company+pathogen fallback fails when the gap-finder
# discovers the recall N days after the scraper did (different Date field).
#
# The near-dup index keys on (source, normalized_company, pathogen) and
# stores a list of dates. A new row is rejected if there's already an
# entry with the same key dated within NEAR_DUP_WINDOW_DAYS days. This
# blocks rediscovery duplicates without blocking legitimate same-company
# recurring recalls (which are usually months apart).
NEAR_DUP_WINDOW_DAYS = 30


def _near_dup_key(row: Dict[str, Any]) -> str:
    """Normalized (source, company, pathogen) tuple — date-independent."""
    src = (row.get("Source") or "").strip().lower()
    co = unicodedata.normalize("NFD", row.get("Company") or "").encode("ascii", "ignore").decode().lower()
    co = re.sub(r"[^a-z0-9]", "", co)[:30]
    pa = (row.get("Pathogen") or "").strip().lower()[:50]
    if not (src and co and pa):
        return ""  # missing any of the three — can't make a meaningful match
    return f"{src}|{co}|{pa}"


def _build_near_dup_index(rows: List[Dict[str, Any]]) -> Dict[str, List[str]]:
    """Return {near_dup_key: [date_str, ...]} for all rows with valid keys."""
    idx: Dict[str, List[str]] = {}
    for r in rows:
        k = _near_dup_key(r)
        d = str(r.get("Date") or "")[:10]
        if k and d:
            idx.setdefault(k, []).append(d)
    return idx


def _is_near_duplicate(
    row: Dict[str, Any], near_dup_index: Dict[str, List[str]],
) -> Tuple[bool, str]:
    """Check if `row` is a near-dup of anything in the index. Returns (is_dup, match_date)."""
    k = _near_dup_key(row)
    new_date_str = str(row.get("Date") or "")[:10]
    if not (k and new_date_str and k in near_dup_index):
        return False, ""
    try:
        new_date = datetime.strptime(new_date_str, "%Y-%m-%d").date()
    except ValueError:
        return False, ""
    for old_date_str in near_dup_index[k]:
        try:
            old_date = datetime.strptime(old_date_str, "%Y-%m-%d").date()
            if abs((new_date - old_date).days) <= NEAR_DUP_WINDOW_DAYS:
                return True, old_date_str
        except ValueError:
            continue
    return False, ""


# ---------------------------------------------------------------------------
# Pending-row validation gate
# ---------------------------------------------------------------------------
# This is the SINGLE chokepoint that blocks garbage from EVERY source
# (scrapers, gap finders, manual injects). Every row added to Pending must
# pass validate_pending_row() — see PIPELINE_FIX_SPEC.md.

# Generic / non-detail URL patterns we never want in Pending. These are
# regulator landing/listing/transparency pages, not specific recall fiches.
_GENERIC_URL_PATTERNS = (
    r"vertexaisearch\.cloud\.google",        # Gemini grounding redirect
    r"fsai\.ie/news-alerts/food\?page=",      # FSAI paginated listing
    r"rasff-window/screen/list\?",            # RASFF list page
    r"quebec\.ca/.*/listeriosis",             # Quebec disease info
    r"quebec\.ca/.*/animal-disease",          # Quebec animal disease info
    r"quebec\.ca/.*/food-recalls$",           # Quebec generic recalls page
    r"regulatory-transparency-and-openness",  # CFIA transparency pages
    r"food-safety-investigations/$",          # CFIA investigation index
    r"/categorie/[\d/]+/?$",                  # RappelConso category index
    r"/rubrik/[^/]+/?$",                      # produktwarnung.eu rubrik
    r"/news-and-alerts/food-alerts/?$",       # FSAI alerts root
    r"/safety/recalls-market-withdrawals-safety-alerts/?$",  # FDA root
    r"/animal-veterinary/news-events/outbreaks-and-advisories/?$",  # FDA pet root
    # CFIA recalls landing page (any locale path or bare host). The CFIA
    # scraper finds specific recall slugs at recalls-rappels.canada.ca/<lang>/<slug>;
    # the bare /fr or /en URL is the listing page itself, never a recall.
    # Triggered by the audit 2026-04-28 leak where the French landing page
    # entered Recalls with the page H1 ("Trouvez des rappels...") as Company.
    r"recalls-rappels\.canada\.ca/(?:fr|en)/?$",
    r"recalls-rappels\.canada\.ca/?$",
    # FDA share-link wrapper format used by their "voluntary-recall" template.
    # Functionally a SPA route — same recall is also published at the
    # canonical /safety/recalls-market-withdrawals-safety-alerts/<slug> URL,
    # which the FDA scraper always finds. Reject the wrapper to prevent
    # duplicates from search-based gap finders (Tavily/Exa/OpenAI) that
    # return whichever URL Google indexed first.
    r"/safety/recalls-market-withdrawals-safety-alerts/voluntary-recall\?permalink=",
    # ── 2026-05-05 audit additions (gap-finder garbage patterns) ───────────
    # Generic full-text search query pages — these aren't recalls, they're
    # search-result lists. CFIA, RappelConso, EFSA, FSANZ all have these.
    r"/search/site",                                # CFIA "advanced search"
    r"/search\?",                                   # generic search query
    r"/recherche\?",                                # French generic search
    r"/recherche/",                                 # French search path
    r"/buscador",                                   # Spanish search
    r"/suche\?",                                    # German search
    # Pagination pages — never recall-specific
    r"/page/\d+/?(?:$|\?)",                         # /page/50/, /page/2/?...
    r"\bpage=\d+",                                  # ?page=50
    # Notification circulars / regulatory bulletins — these are notices ABOUT
    # things to come, not recalls themselves. FSANZ, ANSES, USDA-FSIS all have
    # circular indexes that should never appear in our recall feed.
    r"/notification-circulars?/?$",
    r"/notification-circulars?/index",
    r"/circulars?/notification-circular-",          # FSANZ specific circular ID
    r"/bulletins?/?$",
    r"/news-circulars?/?$",
)

# Company-field strings that the scraper has clearly bungled (they're page
# titles, section headers, or page text — never legitimate company names).
#
# IMPORTANT — what does NOT belong in this set:
#   • "Various brands" / "Various producers" / "Multiple brands"  → legit
#     descriptor when one recall covers many SKUs from different producers
#     (e.g. BLV Salmonellen-Weichkäse, RASFF multi-country alerts).
#   • "Unbranded" / "—" / "sans marque" / "No brand"              → legit
#     descriptor for RappelConso "sans marque" entries, generic raw products,
#     bulk commodity recalls.
#   • "Consult Food …", "Various Foods Ltd", etc.                 → real
#     company names that happen to start with normally-suspect words.
# Company-field cleanup beyond clear scraper bugs is the URL gate's job +
# downstream Claude review's job — not this gate's job.
_GARBAGE_COMPANIES = {
    "list of",                                       # FSAI/CFIA listing-page H1
    "food alerts",                                   # FSAI navigation
    "food alert",                                    # FSAI navigation
    "listeriosis",                                   # disease name as company
    "animals can catch and transmit salmonellosis",  # CFIA page text
    "food safety investigation:",                    # CFIA section header
    "timeline of events:",                           # CFIA section header
    "recall of",                                     # FSAI page-title leak (prefix)
}

# Hard cutoff: nothing dated before this enters Pending.
_MIN_VALID_DATE = "2026-01-01"


# News-outlet hosts. Any URL whose host (or parent domain) matches lands in
# the NEWS sheet via scrapers/news.py — never in Recalls. Gap-finders
# (Tavily/Exa/Gemini) sometimes surface news-article URLs while searching for
# recall content; without this blocklist they would slip into Pending and
# get promoted to Recalls with the article <title> tag scraped as Company.
# Triggered by the audit 2026-04-28 leak (foodsafetynews.com articles
# appearing in Recalls).
# ── RASFF (EU) URL pattern (audit 2026-04-29) ──────────────────────────────
# RASFF rows are accepted only when the URL is a specific notification
# detail page. The Window app at /screen/search and /screen/consumers is
# a Vue/Angular SPA shell with no recall content. The notification deep-
# link route /screen/notification/<id> IS rendered server-side enough to
# carry the recall record. validate_pending_row enforces this constraint
# only when Source starts with "RASFF". See _missing_required in
# run_all.py for the equivalent check during scraper run.
_RASFF_NOTIFICATION_URL_RE = re.compile(
    r"^https://webgate\.ec\.europa\.eu/rasff-window/screen/notification/\d+/?$",
    re.IGNORECASE,
)

_NEWS_HOSTS = frozenset({
    "foodsafetynews.com",
    "foodpoisonjournal.com",
    "foodpoisoningbulletin.com",
    "outbreaknewstoday.com",
    "cidrap.umn.edu",
    "food-safety.com",
    "barfblog.com",
    "foodbusinessnews.net",
    "foodnavigator.com",
    "foodnavigator-usa.com",
    "just-food.com",
    "foodmanufacture.co.uk",
    "foodprocessing.com",
    "foodengineeringmag.com",
    "fooddive.com",
    "reuters.com",
    "apnews.com",
    "bbc.com",
    "bbc.co.uk",
    "bloomberg.com",
    "theguardian.com",
    "nytimes.com",
    "washingtonpost.com",
    "medicalxpress.com",
    "sciencedaily.com",
    "yahoo.com",
    "msn.com",
    "news.google.com",
})


def _host_is_news_outlet(url: str) -> bool:
    """True if the URL host (or any parent of it) is in _NEWS_HOSTS."""
    if not url:
        return False
    try:
        host = urlparse(url).netloc.lower()
    except Exception:
        return False
    if host.startswith("www."):
        host = host[4:]
    if host in _NEWS_HOSTS:
        return True
    # Subdomain match (e.g. recalls.reuters.com endswith .reuters.com)
    for h in _NEWS_HOSTS:
        if host.endswith("." + h):
            return True
    return False


def validate_pending_row(
    row: Dict[str, Any],
    existing_urls: set,
    gap_finder: bool = False,
) -> Tuple[bool, str]:
    """
    Return (is_valid, rejection_reason). Reject garbage before it enters
    Pending. Called by append_to_pending() for every candidate row, no
    matter the source (scrapers, gap finders, manual injects).

    Rules — see PIPELINE_FIX_SPEC.md for rationale:
      • REJECT vertexaisearch redirect URLs (Gemini grounding artifacts)
      • REJECT generic / category / paginated-listing pages
      • REJECT URLs that aren't http/https
      • REJECT garbage Company fields ("List of", "Food Alerts", etc.)
      • REJECT duplicate URLs already in Recalls or Pending
      • REJECT dates before 2026-01-01

    `existing_urls` is a set of already-seen lowercased URLs (Recalls +
    current Pending). Pass an empty set to skip the dedup check.

    NOTE on near-duplicate detection: helpers _near_dup_key /
    _build_near_dup_index / _is_near_duplicate exist above for use by
    standalone audit scripts. They are NOT wired into this gate because
    European regulators (especially RappelConso) routinely publish one
    fiche per SKU — same company, same pathogen, same day, different
    fiche IDs. A naive same-source/company/pathogen+30d gate would
    falsely reject those legitimate per-SKU rows. Use the helpers only
    when you've verified the URL host context makes near-dup detection
    safe (e.g. for FDA where wrapper URLs duplicate canonical URLs).
    """
    url = str(row.get("URL", "") or "").strip()
    company = str(row.get("Company", "") or "").strip()
    date_str_raw = str(row.get("Date", "") or "")
    # ── Defensive Date normalization (audit 2026-05-21) ─────────────────
    # Backstop for openFDA-style YYYYMMDD compact dates that bypass the
    # scrapers._base._new_recall normalizer (e.g. rows constructed via
    # Recall(...) directly, manual injects, or gap-finder rows that
    # carry a date stamp from the source page). Mutate the row dict
    # in place so all downstream consumers see the canonical YYYY-MM-DD.
    # See scrapers/_base._normalize_date_string for the source-level
    # normalizer that catches this at construction time.
    if re.fullmatch(r"20\d{6}", date_str_raw):
        y, mo, da = date_str_raw[:4], date_str_raw[4:6], date_str_raw[6:8]
        if 1 <= int(mo) <= 12 and 1 <= int(da) <= 31:
            normalized = f"{y}-{mo}-{da}"
            row["Date"] = normalized
            date_str_raw = normalized
            log.info(
                "validate_pending_row: normalized YYYYMMDD Date %s → %s "
                "(url=%s)",
                date_str_raw, normalized, url[:60],
            )
    date_str = date_str_raw[:10]
    source = str(row.get("Source", "") or "").strip()

    # ── RASFF (EU) schema awareness (audit 2026-04-29) ──────────────────
    # RASFF rows don't publish company names — they publish origin and
    # distributed countries instead. The Company field on a valid RASFF
    # row holds the formatted string "Origin: <X> | Distributed: <Y>".
    # That string contains a pipe, which would trip the news-article-
    # title leak check below. Detect RASFF up front so we can skip
    # company-garbage heuristics for these rows AND impose the inverse
    # constraint: RASFF URL must point to a /screen/notification/<id>
    # page, not a search/landing/consumer shell.
    is_rasff = source.upper().startswith("RASFF")

    # ── Final gate (locked 2026-04-30) — modules imported at function ──
    # Local imports to avoid circular deps (merge_master is imported by
    # both url_gate_gemini and claude_check, which import these too).
    try:
        from pipeline._url_year import is_year_mismatch
        from pipeline._pathogen_scope import (
            is_in_scope as _is_tier1_pathogen,
            is_empty_pathogen as _is_empty_pathogen,
            is_pet_food_product as _is_pet_food,
            is_pet_food_url as _is_pet_food_url,
        )
        from pipeline._news_mirror_blocklist import is_news_mirror as _is_news_mirror
        from pipeline._cfs_aggregator_guard import is_foreign_cfs_repost as _is_foreign_cfs
    except ImportError:
        is_year_mismatch = None
        _is_tier1_pathogen = None
        _is_empty_pathogen = None
        _is_news_mirror = None
        _is_pet_food = None
        _is_pet_food_url = None
        _is_foreign_cfs = None

    if _is_news_mirror is not None and _is_news_mirror(url):
        return False, "news_mirror_domain (locked 2026-04-30)"

    # ── CFS aggregator guard (added 2026-07-14) ─────────────────────────
    # cfs.gov.hk "Food Incident Post" PDFs re-post other jurisdictions'
    # recalls. When the origin Country is NOT Hong Kong, the row is a
    # cross-source duplicate of the upstream regulator (RappelConso/FDA/
    # FSA/MPI/FSANZ/...) that FSIS already ingests directly. Reject it.
    # HK-origin CFS rows (CFS is the primary/sole source) are kept.
    if _is_foreign_cfs is not None and _is_foreign_cfs(url, row.get("Country", "")):
        return False, ("cfs_foreign_repost: cfs.gov.hk aggregator of a "
                       f"non-HK recall (Country={row.get('Country', '')!r}) "
                       "— upstream regulator row is the primary source")

    # ── Gap-finder-only guards (audit 2026-06-25) ───────────────────────
    # Recency (reject months-old / mis-dated rows), authority-allowlist
    # (reject news + non-government URLs the LLM returned), and product
    # hygiene (reject raw lot/date-code Products, de-dupe doubled Company).
    # SCOPED to gap-finder rows via the flag — scraper rows and manual
    # injects (gap_finder=False, the default) are never affected.
    if gap_finder:
        try:
            from pipeline._gap_finder_guards import check_gap_finder_row
        except ImportError:
            check_gap_finder_row = None
        if check_gap_finder_row is not None:
            gf_ok, gf_reason, gf_note = check_gap_finder_row(row)
            if gf_note:
                log.info("gap-finder guard note: %s (url=%s)", gf_note, url[:70])
            if not gf_ok:
                return False, f"gap_finder_guard: {gf_reason}"

    # ── Pathogen scope check (audit 2026-05-08) ─────────────────────────
    # This check runs AFTER all URL/garbage/dedup checks (further below)
    # so that an empty-pathogen row from a junk URL still gets rejected
    # by the URL gate first — we don't want claude_check wasting AI calls
    # trying to enrich pathogen on a vertexaisearch redirect or a news-
    # outlet URL.
    #
    # Outcome of this gate has three branches:
    #   1. Pathogen is empty/sentinel  → defer until end, accept into
    #      Pending with Status=pending_enrichment for AI enrichment.
    #   2. Pathogen is present but not Tier-1  → reject NOW. We already
    #      know what it is, no enrichment will change that.
    #   3. Pathogen is Tier-1  → continue to remaining gates.
    pathogen_str = str(row.get("Pathogen", "") or "")
    pathogen_empty = (
        _is_empty_pathogen is not None and _is_empty_pathogen(pathogen_str)
    )
    if (_is_tier1_pathogen is not None
            and not pathogen_empty
            and not _is_tier1_pathogen(pathogen_str)):
        return False, f"pathogen_out_of_scope: {pathogen_str!r}"

    # ── Pet / animal food gate (added 2026-05-23) ──────────────────────
    # AFTS-FSIS monitors HUMAN food only. Pet food / dog & cat treats /
    # animal feed / livestock feed are rejected regardless of pathogen.
    # Checks Product, Company, Brand, Reason in any language.
    if _is_pet_food is not None and _is_pet_food(
        row.get("Product", ""),
        row.get("Company", ""),
        row.get("Brand", ""),
        row.get("Reason", ""),
    ):
        return False, "pet_food_out_of_scope: pet / animal food not in AFTS-FSIS human-food scope"

    # The four fields above are what the SCRAPER extracted. When it extracts
    # only "Chicken Recipe" from a page titled "Two Raw Pet Food Products",
    # every one of them is silent and the regulator's own URL is the only
    # place the truth survives. Read it too. (audit 2026-08-30)
    if _is_pet_food_url is not None and _is_pet_food_url(url):
        return False, "pet_food_out_of_scope: regulator URL identifies this as pet / animal food"

    if is_year_mismatch is not None:
        try:
            row_d = (datetime.fromisoformat(date_str).date()
                     if date_str else None)
        except (TypeError, ValueError):
            row_d = None
        ym_reason = is_year_mismatch(row_d, url)
        if ym_reason:
            return False, f"url_year_mismatch: {ym_reason}"

    # Extraction garbage
    company_lc = company.lower()
    brand_lc = str(row.get("Brand", "") or "").strip().lower()
    GARBAGE = {"home","index","page","recalls","alerts","alert","recall","welcome","main"}
    if company_lc and company_lc == brand_lc and company_lc in GARBAGE:
        return False, f"extraction_garbage: Company=Brand={company_lc!r}"
    if re.search(r"/(home|index|main|welcome)/?$", url.lower()):
        return False, "extraction_garbage: URL is landing page"

    # ── REJECT: vertexaisearch redirect URLs (Gemini grounding artifacts) ──
    if "vertexaisearch.cloud.google" in url:
        return False, "Gemini grounding redirect URL, not a real recall"

    # ── REJECT: URL host is a news outlet, not a regulator ──────────────
    # News articles belong in the NEWS sheet (populated by scrapers/news.py).
    # Gap-finders sometimes surface news-article URLs while searching for
    # recall content; reject them at the gate before they reach Recalls.
    if _host_is_news_outlet(url):
        return False, f"News outlet URL — belongs in NEWS sheet, not Recalls: {url[:60]}"

    # ── RASFF URL gate (audit 2026-04-29) ───────────────────────────────
    # RASFF rows must point to a specific notification page. Anything
    # else (the search SPA shell at /screen/search, the consumer portal
    # at /screen/consumers, or a bare /rasff-window root) is rejected
    # here. The notification page is the only URL that contains the
    # actual recall record. The gap-finder URL filter rejects the same
    # landing pages, but this is the structural gate at promotion time.
    if is_rasff:
        if not _RASFF_NOTIFICATION_URL_RE.match(url):
            return False, (f"RASFF row URL must be /screen/notification/<id>, "
                           f"got: {url[:80]}")

    # ── REJECT: generic / informational / listing pages ─────────────────
    for pat in _GENERIC_URL_PATTERNS:
        if re.search(pat, url, re.IGNORECASE):
            return False, f"Generic/info page URL: matches {pat}"

    # ── REJECT: URL is not http/https ───────────────────────────────────
    if url and not url.lower().startswith(("http://", "https://")):
        return False, f"Invalid URL scheme: {url[:30]}"

    # ── REJECT: company field is clearly a page title, not a company ────
    # Skipped for RASFF rows — their Company field legitimately contains
    # the "Origin: <X> | Distributed: <Y>" pattern.
    if not is_rasff:
        co_low = company.lower()
        if co_low in _GARBAGE_COMPANIES:
            return False, f'Company field is not a company: "{company}"'
        # Substring check for "Recall of …" / "List of …" leakage
        for bad in _GARBAGE_COMPANIES:
            if co_low.startswith(bad + " ") or co_low.startswith(bad + ":"):
                return False, f'Company field starts with garbage prefix "{bad}"'
        # Article-title leak: scraped <title> tags from news pages contain a
        # pipe + outlet name (e.g. "Salmonella outbreak ... | Food Safety News")
        # or HTML/JS fragments. Real company names never contain these.
        if " | " in company and re.search(
            r"\|\s*(food\s*safety\s*news|food\s*poison|outbreak\s*news|cidrap|"
            r"reuters|bbc|bloomberg|guardian)\b", company, re.I):
            return False, f'Company field is a news article <title> tag: "{company[:60]}"'
        if re.search(r"window\.\w+|document\.querySelector|<\s*script\b|"
                     r"\{socials\b|addEventListener\(", company, re.I):
            return False, f'Company field contains HTML/JS fragment: "{company[:60]}"'

    # ── REJECT: duplicate URL already in Recalls or Pending ─────────────
    url_norm = url.rstrip("/").lower()
    if url_norm and url_norm in existing_urls:
        return False, "Duplicate URL already exists"

    # ── REJECT: date is before 2026-01-01 ───────────────────────────────
    if date_str and date_str < _MIN_VALID_DATE:
        return False, f"Date before {_MIN_VALID_DATE}: {date_str}"

    # All other gates passed. If pathogen was empty, signal to caller
    # that this row should land in Pending with Status=pending_enrichment
    # so an AI enrichment step can fill in the Pathogen field. Otherwise
    # this is a fully-validated row ready for normal pending flow.
    if pathogen_empty:
        return True, OK_PENDING_ENRICHMENT
    return True, "OK"


# ---------------------------------------------------------------------------
# Load
# ---------------------------------------------------------------------------
def _load_sheet(xlsx_path: Path, sheet: str, schema: List[str]) -> List[Dict[str, Any]]:
    if not xlsx_path.exists():
        return []
    wb = load_workbook(xlsx_path)
    if sheet not in wb.sheetnames:
        return []
    ws = wb[sheet]
    headers = [c.value for c in ws[1]]
    out = []
    # Defensive: openpyxl may return Date cells as datetime objects when a
    # cell was manually edited and Excel auto-typed it. Every downstream
    # consumer (gate, sort, dedup, JSON mirror) expects YYYY-MM-DD strings,
    # so coerce here at the single source of truth.
    from datetime import datetime as _dt, date as _dt_date
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v in (None, "") for v in row):
            continue
        rec = {h: (v if v is not None else "") for h, v in zip(headers, row)}
        # Normalise Date column
        d = rec.get("Date")
        if isinstance(d, (_dt, _dt_date)):
            try:
                rec["Date"] = d.strftime("%Y-%m-%d")
            except (TypeError, ValueError):
                rec["Date"] = ""
        elif d not in (None, "") and not isinstance(d, str):
            # Excel serial number or other unexpected type
            rec["Date"] = str(d)[:10]
        # Backfill any schema cols missing from the sheet (schema evolution safety)
        for col in schema:
            rec.setdefault(col, "" if col not in ("Tier", "Outbreak") else 0)
        out.append(rec)
    return out


def load_existing(xlsx_path: Path) -> List[Dict[str, Any]]:
    """Read approved Recalls sheet -> list of dicts."""
    out = _load_sheet(xlsx_path, "Recalls", RECALLS_SCHEMA)
    log.info("Loaded %d approved rows from Recalls", len(out))
    return out


def load_pending(xlsx_path: Path) -> List[Dict[str, Any]]:
    """Read Pending sheet -> list of dicts. Empty if sheet doesn't exist yet."""
    out = _load_sheet(xlsx_path, "Pending", PENDING_SCHEMA)
    log.info("Loaded %d rows from Pending", len(out))
    return out


# ---------------------------------------------------------------------------
# Merge scraped rows into Pending
# ---------------------------------------------------------------------------
def append_to_pending(
    existing_pending: List[Dict[str, Any]],
    approved: List[Dict[str, Any]],
    new_recalls: List[Recall],
    scraped_at: str,
    gap_finder: bool = False,
) -> List[Dict[str, Any]]:
    """
    Take new scraped+enriched Recall objects and append them to the pending list.

    Dedup rules:
      - If the key is already approved in Recalls  -> skip silently
      - If the key is currently in Pending with Status='pending'    -> skip (waiting)
      - If the key is currently in Pending with Status='rejected'   -> DELETE the
        old rejected row and insert the freshly scraped row for re-validation.
        This gives the source a chance to fix broken links / fill missing fields
        before the next run, and prevents rejected rows from being silently
        re-skipped forever.
      - Otherwise (brand new key) -> insert as Status='pending'.
    """
    keys_in_approved = {_dedup_key(r) for r in approved}

    # ── Near-duplicate index (audit 2026-04-29) ──────────────────────────
    # Catches hallucinated-URL gap-finder duplicates that string dedup
    # can't see. Example: scraper has Listeria/LES ATELIERS DE SEBASTIEN/
    # 2026-04-28 at /fiche-rappel/22142/Interne (real). Gemini gap-finder
    # then hallucinates the same recall at /fiche-rappel/22185/Interne
    # (fake). _dedup_key() compares URLs → no match → both land in
    # Pending → URL gate runs once a day at 07:00 → in the meantime
    # merge_master promotes the hallucinated one to Recalls.
    #
    # _is_near_duplicate keys on (source, normalized_company, pathogen)
    # within a 30-day window — so the same (source, company, pathogen)
    # appearing at a different URL within 30 days is rejected as a near-
    # dup before it ever lands in Pending. This catches the leak at
    # ingest time, not promotion time.
    near_dup_index = _build_near_dup_index(approved + existing_pending)

    # Build set of all URLs already present (Recalls + current Pending) for
    # the validation gate's dedup check. Lowercased + trailing-slash-stripped
    # to match validate_pending_row()'s normalisation.
    #
    # IMPORTANT (audit 2026-04-29): exclude URLs of rows currently in
    # Status="rejected" from existing_urls. The retry path further down
    # promises that a freshly-scraped row matching a rejected row will
    # DELETE the old row and re-queue the new one. That promise was dead
    # because validate_pending_row's dup-URL check fired first (the URL
    # was in existing_urls regardless of Status), bouncing the retry
    # before the retry logic ever ran.
    existing_urls: set = set()
    for r in approved:
        u = str(r.get("URL", "") or "").strip().rstrip("/").lower()
        if u:
            existing_urls.add(u)
    for r in existing_pending:
        if (r.get("Status") or "").lower() == STATUS_REJECTED:
            continue  # let the retry path handle this URL
        u = str(r.get("URL", "") or "").strip().rstrip("/").lower()
        if u:
            existing_urls.add(u)

    # Index existing pending by key so we can drop rejected duplicates in place.
    # Multiple rows with the same key shouldn't happen, but if they do keep them
    # all (one will match; the others are untouched).
    pending_by_key: Dict[str, List[int]] = {}
    for i, r in enumerate(existing_pending):
        pending_by_key.setdefault(_dedup_key(r), []).append(i)

    # Decide which existing-pending rows to drop (rejected rows being re-scraped).
    indices_to_drop: set = set()
    fresh_rows: List[Dict[str, Any]] = []
    retried = 0
    appended = 0
    appended_enrichment = 0
    already_pending = 0
    already_approved = 0
    rejected_by_gate = 0

    for r in new_recalls:
        d = r.to_dict() if isinstance(r, Recall) else dict(r)
        for col in SCHEMA:
            d.setdefault(col, "" if col not in ("Tier", "Outbreak") else 0)

        # ── HARD GATE: validate before any other logic. Blocks garbage from
        # ── ALL sources (scrapers, gap-finders, manual injects).
        ok, why = validate_pending_row(d, existing_urls, gap_finder=gap_finder)
        if not ok:
            log.warning(
                "Pending gate REJECT: %s | url=%s | company=%s",
                why, str(d.get("URL", ""))[:100], str(d.get("Company", ""))[:50],
            )
            rejected_by_gate += 1
            continue

        # Sentinel from validate_pending_row: row passed every other gate
        # but Pathogen is empty. Land it in Pending under enrichment
        # status so claude_check / Gemini can fill the field in the next
        # AI pass. promote_approved skips this status — see
        # NON_PROMOTABLE_STATUSES.
        needs_enrichment = (why == OK_PENDING_ENRICHMENT)

        k = _dedup_key(d)

        if k in keys_in_approved:
            already_approved += 1
            continue

        # ── Near-duplicate check (audit 2026-04-29) ─────────────────────
        # Reject if a recall with the same (source, normalized_company,
        # pathogen) was already approved or pended within the last 30
        # days. This is the gate that stops gap-finder URL hallucinations
        # from ever entering Pending — even when their hallucinated URL
        # is structurally valid (5-digit fiche ID in the right range).
        is_near, match_date = _is_near_duplicate(d, near_dup_index)
        if is_near:
            log.warning(
                "Pending near-dup REJECT: same (source, company, pathogen) "
                "already exists dated %s | new url=%s | company=%s",
                match_date, str(d.get("URL", ""))[:100],
                str(d.get("Company", ""))[:50],
            )
            rejected_by_gate += 1
            continue

        new_status = (STATUS_PENDING_ENRICHMENT
                      if needs_enrichment else STATUS_PENDING)

        if k in pending_by_key:
            # Look at the FIRST matching row's status (practically there's only one)
            existing_idx = pending_by_key[k][0]
            existing_status = (existing_pending[existing_idx].get("Status") or "").lower()
            if existing_status == STATUS_REJECTED:
                # Drop the old rejected row, re-queue the fresh scrape
                indices_to_drop.add(existing_idx)
                d["ScrapedAt"] = scraped_at
                d["Status"] = new_status
                fresh_rows.append(d)
                retried += 1
            else:
                # Still pending from a prior run — leave it alone
                already_pending += 1
            continue

        # Brand new key
        d["ScrapedAt"] = scraped_at
        d["Status"] = new_status
        fresh_rows.append(d)
        if needs_enrichment:
            appended_enrichment += 1
        else:
            appended += 1

    # Assemble output: existing pending minus dropped + new/retried
    kept = [r for i, r in enumerate(existing_pending) if i not in indices_to_drop]
    out = kept + fresh_rows

    log.info(
        "Pending: kept %d (dropped %d rejected for retry), +%d new, "
        "+%d new awaiting enrichment, +%d retried "
        "(skipped: %d already-pending, %d already-approved, %d gate-rejected) = %d total",
        len(kept), len(indices_to_drop), appended, appended_enrichment, retried,
        already_pending, already_approved, rejected_by_gate, len(out),
    )
    return out


# ---------------------------------------------------------------------------
# Promotion: Pending -> Recalls
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# Rejection counter (audit 2026-05-05)
# ---------------------------------------------------------------------------
# Two-strikes-and-out: a Pending row is physically deleted once 2 DIFFERENT
# reviewers have rejected it. Single-reviewer rejections sit in Pending with
# Status=rejected, available for retry on the next scrape (the existing
# append_to_pending re-queue path).
#
# Reviewer name is parsed from the rejection reason via a `<reviewer>: <text>`
# convention. claude_check stamps "Claude check: <reason>", url_gate_gemini
# stamps "Gemini gate: <reason>", openrouter_check stamps "OpenRouter:
# <reason>", and merge_master itself uses "Pending gate: <reason>" for
# structural rejects (broken URLs, missing fields).
#
# Returns (should_delete, updated_row). Caller decides whether to drop the
# row from kept_in_pending based on should_delete.

def _reviewer_from_reason(reason: str) -> str:
    """Extract canonical reviewer name from rejection reason string.

    Maps the leading 'reviewer: ...' prefix to one of the canonical names:
      claude-check, gemini-url-gate, openrouter-check, pending-gate, manual.
    Returns 'unknown' if no recognised prefix is found — these are still
    counted (so the counter advances) but logged for investigation.
    """
    r = (reason or "").strip().lower()
    if r.startswith("claude check") or r.startswith("claude-check"):
        return "claude-check"
    if r.startswith("gemini gate") or r.startswith("gemini-url-gate") or \
            r.startswith("gemini url gate"):
        return "gemini-url-gate"
    if r.startswith("openrouter") or r.startswith("openrouter-check"):
        return "openrouter-check"
    if r.startswith("pending gate") or r.startswith("pending-gate"):
        return "pending-gate"
    if r.startswith("url validator") or r.startswith("url-validator"):
        return "url-validator"
    if r.startswith("manual"):
        return "manual"
    return "unknown"


def mark_rejected_with_counter(row: Dict[str, Any], reason: str
                               ) -> Tuple[bool, Dict[str, Any]]:
    """Apply a rejection to a Pending row, tracking which reviewers reject it.

    Args:
        row    : the Pending row dict (mutated in place)
        reason : human-readable rejection reason, ideally prefixed with
                 reviewer name (e.g. "Claude check: company mismatch")

    Returns:
        (should_delete, updated_row)

        should_delete is True when 2+ DIFFERENT reviewers have rejected this
        row — caller is expected to drop the row from the Pending list.

        Repeat rejections by the SAME reviewer are idempotent: Notes is
        updated with the latest reason but the counter doesn't double-count.
    """
    reviewer = _reviewer_from_reason(reason)

    # Parse existing reviewer set from the RejectedBy column
    raw = (row.get("RejectedBy") or "").strip()
    rejected_by = set(filter(None, (s.strip() for s in raw.split(","))))

    # Same-reviewer repeat is idempotent (still update Notes for visibility)
    is_new_reviewer = reviewer not in rejected_by
    rejected_by.add(reviewer)

    row["RejectedBy"] = ",".join(sorted(rejected_by))
    row["Status"] = STATUS_REJECTED

    # Stamp reason into Notes (preserve any prior REJECTED: prefix history
    # by appending instead of overwriting if a previous reviewer already
    # tagged it).
    orig_notes = (row.get("Notes") or "").strip()
    new_stamp = f"REJECTED: {reason}"
    if orig_notes.startswith("REJECTED:"):
        # Already has a rejection stamp — append the new reviewer's reason
        # so the audit trail is preserved.
        if reason not in orig_notes:
            row["Notes"] = f"{orig_notes} || {reviewer}: {reason}"
    else:
        row["Notes"] = new_stamp + (f" | {orig_notes}" if orig_notes else "")

    # Two-strikes-and-out: only DIFFERENT reviewers count toward delete.
    should_delete = len(rejected_by) >= 2

    if should_delete:
        log.info(
            "Pending DELETE (2+ reviewers rejected): RejectedBy=%s | url=%s",
            row["RejectedBy"], str(row.get("URL", ""))[:100],
        )

    return should_delete, row


def cleanup_orphan_rejected(pending: List[Dict[str, Any]],
                            min_age_hours: int = 24
                            ) -> Tuple[List[Dict[str, Any]], int]:
    """Physically delete already-rejected Pending rows older than min_age_hours.

    Catches orphan rejections from before the RejectedBy counter was added —
    rows that have Status=rejected and a REJECTED: prefix in Notes but no
    RejectedBy column. After 24h these are unlikely to be re-validated, so
    we delete them to keep the Pending sheet clean.

    Returns (filtered_pending, n_deleted).
    """
    from datetime import datetime, timezone, timedelta
    cutoff = datetime.now(timezone.utc) - timedelta(hours=min_age_hours)

    filtered: List[Dict[str, Any]] = []
    n_deleted = 0
    for r in pending:
        if (r.get("Status") or "").lower() != STATUS_REJECTED:
            filtered.append(r)
            continue

        # If RejectedBy already tracks 2+ reviewers, mark_rejected_with_counter
        # would have deleted this row already — keep it as-is for the caller.
        rejected_by = set(filter(None, (
            s.strip() for s in (r.get("RejectedBy") or "").split(","))))
        if len(rejected_by) >= 2:
            n_deleted += 1
            continue  # delete (already past the counter threshold)

        # Otherwise, age-based delete: anything rejected >24h ago goes away.
        scraped_at = (r.get("ScrapedAt") or "").strip()
        try:
            sa = datetime.fromisoformat(scraped_at.replace("Z", "+00:00"))
            if sa.tzinfo is None:
                sa = sa.replace(tzinfo=timezone.utc)
        except (ValueError, TypeError):
            sa = None

        if sa is None or sa < cutoff:
            n_deleted += 1
            log.info(
                "Pending DELETE (orphan rejected, age>=%dh): url=%s",
                min_age_hours, str(r.get("URL", ""))[:100],
            )
            continue

        filtered.append(r)

    return filtered, n_deleted


# ---------------------------------------------------------------------------
# Re-promotion guard (audit 2026-08-02)
# ---------------------------------------------------------------------------
# A rejection was not sticky. Reconstructed from the commit history of
# docs/data/recalls.xlsx on 2026-08-01, for one FSANZ row:
#
#   07-31 21:10  Weekly_Rejected   claude-check: "fail; pathogen mismatch"
#   08-01 02:16  Pending           restored by an "Add files via upload" commit
#   08-01 03:10  Pending           gemini-enrich fills Brand
#   08-01 04:11  (evicted again by the automated run)
#   08-01 07:52  Pending           restored by the next manual upload
#   08-01 13:40  Pending           still there
#   08-01 14:16  Recalls           Qwen review agent APPROVED and promoted it
#   08-01 13:25  → subscriber alert email
#
# Two things are wrong in that trace and this guard fixes the second, which
# is the one that decides whether the first can hurt anyone:
#
#   1. Manually re-uploaded workbook snapshots resurrect rows the pipeline
#      already evicted. That is an operational hazard outside this module.
#   2. NOTHING consulted Weekly_Rejected before promoting. Every reviewer got
#      a clean slate on a row the binding reviewer had already killed, so a
#      resurrected row only had to find one reviewer willing to say yes — and
#      with several reviewers of differing strength in rotation, it did.
#
# Weekly_Rejected is the single source of rejection truth (see REJECTED_SCHEMA
# above). Once a URL is in it, promoting the same URL again is a decision that
# needs a human, not a second opinion from a weaker model. The row is archived
# straight back with the original verdict quoted, so the trail stays intact
# and nothing is silently deleted.
#
# Deliberately keyed on the NORMALISED URL only. Content keys drift as
# enrichment rewrites Company and Brand — that drift is exactly how the row
# escaped in the first place — whereas the address the regulator serves the
# notice at does not.
_REJECTED_URL_CACHE: Dict[str, Tuple[float, Dict[str, str]]] = {}


# Sheets consulted, in order of precedence. "Rejected" is the PERMANENT
# archive written by tools/wipe_weekly_rejected.py; "Weekly_Rejected" is the
# rolling sheet the Thursday wipe empties.
#
# TWO DEFECTS FIXED HERE, 2026-08-18
# ---------------------------------
# 1. THE REASON COLUMN WAS NEVER READ. This function looked for a header
#    called "RejectionReason". The sheet's actual header — see
#    REJECTED_SCHEMA — is "RejectReason". `hdr.index` was guarded by an
#    `in hdr` test, so the mismatch did not raise; i_why was simply None on
#    every call and every entry came out as "a reviewer:" or
#    "gap_finder/za/rules.py:" with the explanation dropped. The guard still
#    blocked the URL, so it looked like it worked, but a human asking "why
#    was this rejected before?" got a colon. Both spellings are accepted now
#    rather than one being renamed, because the sheet is written by several
#    tools and is on disk in the operator's workbook.
#
# 2. ONLY THE ROLLING SHEET WAS READ. The Thursday wipe empties
#    Weekly_Rejected, so every rejection older than one cycle was forgotten
#    and the gap-finders re-found it. wipe_weekly_rejected.py was changed to
#    MOVE rows into a permanent "Rejected" sheet instead of deleting them —
#    but nothing ever read that sheet, so the archive was write-only and the
#    re-ingestion continued. Measured 2026-08-18: five Pending rows were
#    decisions a human had already made, two of them for the fifth time.
_REJECT_SHEETS = ("Rejected", "Weekly_Rejected")
_REJECT_REASON_HEADERS = ("RejectReason", "RejectionReason", "RejectedReason")


def load_rejected_urls(xlsx_path: Optional[Path] = None) -> Dict[str, str]:
    """Map normalised URL -> short description of the recorded rejection.

    Reads the permanent Rejected sheet AND the rolling Weekly_Rejected
    sheet. Returns {} on any problem: this guard must never be the reason a
    pipeline run dies.
    """
    if xlsx_path is None:
        xlsx_path = Path(__file__).resolve().parent.parent / "docs" / "data" / "recalls.xlsx"
    xlsx_path = Path(xlsx_path)
    try:
        stamp = xlsx_path.stat().st_mtime
    except OSError:
        return {}
    key = str(xlsx_path)
    cached = _REJECTED_URL_CACHE.get(key)
    if cached and cached[0] == stamp:
        return cached[1]
    out: Dict[str, str] = {}
    no_url = 0
    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        for sheet in _REJECT_SHEETS:
            if sheet not in wb.sheetnames:
                continue
            rows = list(wb[sheet].values)
            if not rows:
                continue
            hdr = [str(h) for h in rows[0]]
            if "URL" not in hdr:
                log.warning("%s has no URL column; re-promotion guard cannot "
                            "use it", sheet)
                continue
            i_url = hdr.index("URL")
            i_by = hdr.index("RejectedBy") if "RejectedBy" in hdr else None
            i_why = next((hdr.index(h) for h in _REJECT_REASON_HEADERS
                          if h in hdr), None)
            if i_why is None:
                log.warning("%s carries none of %s — rejection reasons will "
                            "not be shown to the next reviewer",
                            sheet, ", ".join(_REJECT_REASON_HEADERS))
            for r in rows[1:]:
                if not r or i_url >= len(r):
                    continue
                u = _normalize_url_for_dedup(str(r[i_url] or ""))
                if not u:
                    # A rejection with no URL cannot be matched against a
                    # future row. Counted and logged rather than dropped in
                    # silence — silent drops in this exact place are why
                    # Weekly_Rejected lost rows for months.
                    no_url += 1
                    continue
                by = str(r[i_by] or "") if i_by is not None and i_by < len(r) else ""
                why = str(r[i_why] or "") if i_why is not None and i_why < len(r) else ""
                desc = f"{by or 'a reviewer'}: {why[:160]}".strip().rstrip(":").strip()
                # Rejected (permanent) is read first and wins: a row that was
                # archived permanently carries the older, binding verdict.
                out.setdefault(u, desc or f"{sheet} (no reason recorded)")
        wb.close()
    except Exception as exc:  # pragma: no cover - defensive
        log.warning("Rejection re-promotion guard unavailable (%s: %s)",
                    type(exc).__name__, str(exc)[:80])
        return {}
    if no_url:
        log.warning("re-promotion guard: %d archived rejection(s) carry no "
                    "URL and cannot block a re-ingestion", no_url)
    _REJECTED_URL_CACHE[key] = (stamp, out)
    return out


def promote_approved(
    pending: List[Dict[str, Any]],
    approved_existing: List[Dict[str, Any]],
    rejected_flags: Dict[int, str],
    *,
    archive_immediately: bool = False,
    previously_rejected: Optional[Dict[str, str]] = None,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]]]:
    """
    Split the Pending list into
        (new_approved_rows_for_Recalls,
         rows_to_keep_in_Pending,
         rows_to_archive_to_Rejected).

    Two-reviewer architecture (locked 2026-05-09):
        Reviewer 1 = Gemini URL gate (pipeline/url_gate_gemini.py)
            • Enriches missing fields (Company / Brand / Pathogen)
            • Validates URLs, sets pass/fail tags
            • NEVER promotes to Recalls (verify-and-tag-only rule)
            • Calls this function with archive_immediately=False (default)

        Reviewer 2 = Claude (morning) OR OpenRouter (evening) — FINAL verdict
            • Per-row content verification, the binding pass/reject
            • Calls this function with archive_immediately=True
            • Eviction is final: row → Weekly_Rejected, removed from Pending
              (caller invokes weekly_rejected_capture.record_rejections
              before/at the moment this returns, so the row is captured
              before its Pending slot disappears)

    Eviction policy controlled by `archive_immediately`:
        False (default — first reviewer / janitor / orchestrator):
            Stamp Status='rejected' in Pending, append reason to Notes.
            Counter-driven 2-reviewer threshold still applies as a safety
            net for hypothetical first-reviewer-only call paths (backfill,
            MERGE_MASTER_PROMOTE=1) — protects against Gemini-only
            eviction if a future config change ever routes its
            rejected_flags here.
        True (second reviewer — Claude or OpenRouter):
            Single-pass eviction. Every row in rejected_flags is archived
            and removed from Pending immediately. The first reviewer has
            already had its say; this is the binding verdict.

    - `rejected_flags` maps pending-row-index -> rejection reason string.
    - Rows NOT in rejected_flags (and whose current Status is 'pending') are
      treated as approved and moved to Recalls, deduped against approved_existing.
    - Rows already marked 'rejected' in a prior run get archived
      out of Pending on the next call (legacy migration after the
      2026-05-09 single-reviewer-eviction policy change).

    Audit 2026-05-08: third return value (archive list) added per operator
    decision — previously second-rejector deletes were silent and untraceable.
    Now they accumulate in a Rejected sheet for human audit.

    Audit 2026-05-09: `archive_immediately` kw-only param added. Makes the
    "Claude/OR is the final reviewer" semantic explicit at the call site
    rather than implicit in caller knowledge.
    """
    approved_keys = {_dedup_key(r) for r in approved_existing}

    # Rejection registry for the re-promotion guard. Loaded once per call;
    # callers that already hold the workbook can pass it in. `{}` explicitly
    # disables the guard (used by the tests and by a deliberate operator
    # override), `None` means "look it up".
    if previously_rejected is None:
        previously_rejected = load_rejected_urls()

    # ── Secondary dedup axis: normalized URL ────────────────────────────
    #
    # AUDIT 2026-07-26 — the FSAI "Butchers Selection" duplicate, promoted
    # for the 5th time. Proven against the live rows:
    #
    #   Dunnes Stores       -> content|2026-06-30|dunnes|salmonella
    #   Butchers Selection  -> content|2026-06-30|butchers-selection|salmonella
    #
    # Both carry the SAME alert URL, differing only in case, so
    # _normalize_url_for_dedup collapses them cleanly. But fsai.ie is
    # registered in _url_identity as a host with no stable alert identifier,
    # so _dedup_key deliberately keys it on CONTENT — and the two rows
    # disagree on Company (retailer vs own-brand). Different content key =>
    # not a duplicate => promoted again.
    #
    # The 2026-07-09 content-key rule is correct and stays: it stops
    # fabricated FSAI URLs (the Horgans case) landing as new rows every pass.
    # The gap is that it REPLACED the URL axis instead of adding to it.
    #
    # Fix: check BOTH axes, reject on either. An identical normalized URL is
    # stronger evidence of sameness than a Company-string mismatch is of
    # difference — two rows cannot be different alerts if the regulator
    # serves them at the same address.
    approved_url_norms = {
        u for u in (_normalize_url_for_dedup(str(r.get("URL", "") or ""))
                    for r in approved_existing) if u
    }

    new_approved: List[Dict[str, Any]] = []
    kept_in_pending: List[Dict[str, Any]] = []
    archived_rejected: List[Dict[str, Any]] = []

    archive_ts = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    def _archive(clean_row: Dict[str, Any]) -> None:
        """Stamp RejectedAt and append to archive list."""
        clean_row["RejectedAt"] = archive_ts
        archived_rejected.append(clean_row)

    for idx, row in enumerate(pending):
        # Strip runtime-only fields (e.g. _url_check) before persisting
        clean = {k: v for k, v in row.items() if not k.startswith("_")}

        # Previously-rejected rows in Pending should not exist after the
        # 2026-05-09 single-reviewer-eviction policy took effect — every
        # second-reviewer rejection archives immediately. This guard
        # handles legacy Status=rejected rows from before the policy
        # change (e.g. xlsx loaded from an older snapshot). When the
        # SECOND reviewer (Claude/OR) runs, migrate them out; when the
        # first reviewer / janitor runs, leave alone (they don't have
        # the authority to evict).
        if clean.get("Status") == STATUS_REJECTED and idx not in rejected_flags:
            if archive_immediately:
                # Second reviewer present — finalize the legacy rejection.
                # One-time migration; subsequent runs see an empty stream.
                _archive(clean)
                continue
            else:
                # First reviewer / janitor — preserve as-is.
                kept_in_pending.append(clean)
                continue

        # ── Non-promotable Pending statuses (gating lock) ─────────────
        # pending_gap / pending_gap_v1 / pending_gap_v2 — gap-finder rows
        #   awaiting Gemini URL grounding + Claude content verification.
        # pending_enrichment — scraper rows missing Pathogen, awaiting AI
        #   enrichment (claude_check / Gemini fills the field, then flips
        #   the status back to "pending").
        # All of these must NOT reach Recalls until the responsible AI
        # step has flipped them to plain "pending". Failures are still
        # routed through rejected_flags above by the gate scripts. This
        # is the structural guarantee that the FSANZ "Australian food"
        # landing-page incident cannot recur, and now also that empty-
        # pathogen rows can't accidentally promote.
        if clean.get("Status") in NON_PROMOTABLE_STATUSES and idx not in rejected_flags:
            kept_in_pending.append(clean)
            continue

        if idx in rejected_flags:
            reason = rejected_flags[idx]
            if archive_immediately:
                # ── SECOND REVIEWER (Claude / OpenRouter) — FINAL verdict ─
                # Single-pass eviction per the operator spec (audit
                # 2026-05-09). Mirror to Weekly_Rejected happens in the
                # CALLER (claude_check / openrouter_check), which invokes
                # weekly_rejected_capture.record_rejections immediately
                # after this function returns — capture-before-delete.
                #
                # Counter is still bumped (audit trail in Notes) so the
                # operator's Thursday review can spot rows rejected
                # multiple times across runs.
                mark_rejected_with_counter(clean, reason)
                if not clean.get("RejectedBy"):
                    rb_match = re.search(
                        r"(claude-check|openrouter-check)",
                        str(reason), re.IGNORECASE,
                    )
                    clean["RejectedBy"] = (
                        rb_match.group(1).lower() if rb_match else "unknown"
                    )
                _archive(clean)
                continue
            else:
                # ── FIRST REVIEWER (Gemini) or janitor / orchestrator ────
                # Don't evict — let the second reviewer (Claude/OR) make
                # the binding call on the next session. Stamp Status=
                # rejected so url-gate skips re-checking the URL but
                # keep the row in Pending for Claude/OR to review.
                #
                # 2-reviewer threshold safety net: if the SAME row gets
                # rejected by 2 different reviewers without ever being
                # finalised by Claude/OR (e.g. Gemini twice in a row,
                # claude-check workflow broken), the counter trips and
                # the row is archived. Belt-and-suspenders for the case
                # where the Claude/OR step is failing silently.
                should_delete, _ = mark_rejected_with_counter(clean, reason)
                if should_delete:
                    _archive(clean)
                    continue
                kept_in_pending.append(clean)
                continue

        # ── Date-consistency gate (audit 2026-05-06: defense-in-depth) ──
        # Last-line defense against date-extractor fallback bugs. If
        # Notes mentions a year far older than the Date field, the
        # extractor probably stamped today's date on an archived page.
        # Honors archive_immediately like rejected_flags above —
        # second reviewer archives, first reviewer parks in Pending.
        date_problem = _check_date_consistency(clean)
        if date_problem:
            if archive_immediately:
                mark_rejected_with_counter(clean, date_problem)
                if not clean.get("RejectedBy"):
                    clean["RejectedBy"] = "date-consistency-gate"
                _archive(clean)
                continue
            should_delete, _ = mark_rejected_with_counter(clean, date_problem)
            if should_delete:
                _archive(clean)
                continue
            kept_in_pending.append(clean)
            continue

        # ── Re-promotion guard (audit 2026-08-02) ───────────────────────
        # A URL already recorded in Weekly_Rejected does not get a second
        # verdict from a different reviewer. See load_rejected_urls() above
        # for the trace that made this necessary.
        _u_now = _normalize_url_for_dedup(str(clean.get("URL", "") or ""))
        if _u_now and _u_now in previously_rejected:
            _prior = previously_rejected[_u_now]
            log.warning("re-promotion BLOCKED %s — already in Weekly_Rejected "
                        "(%s)", str(clean.get("URL", ""))[:90], _prior[:120])
            clean["Notes"] = (
                str(clean.get("Notes") or "").strip()
                + f" [re-promotion blocked 2026-08-02: this URL is already in "
                  f"Weekly_Rejected — {_prior[:200]}]"
            ).strip()
            if not clean.get("RejectedBy"):
                clean["RejectedBy"] = "repromotion-guard"
            _archive(clean)
            continue

        # Approved row: dedup against existing Recalls on BOTH axes —
        # content/dedup key AND normalized URL. See approved_url_norms above
        # for why one axis alone is not enough (FSAI, audit 2026-07-26).
        k = _dedup_key(clean)
        u = _normalize_url_for_dedup(str(clean.get("URL", "") or ""))
        if k in approved_keys or (u and u in approved_url_norms):
            # Already published — drop silently from Pending
            continue
        approved_keys.add(k)
        if u:
            approved_url_norms.add(u)

        # Strip pending-only tracking columns before inserting into Recalls.
        # Fill RECALLS_SCHEMA, including the internal tracking columns:
        #   DateAdded   = today (when row first promoted to Recalls)
        #   LastUpdated = today (initial insert counts as an update)
        #   LastChecked = "" (no URL re-validation has happened yet —
        #                     url_guardian/url_gate will fill this later)
        #   report_week = stamp from row Date (sticky — see compute_report_week)
        from datetime import date as _today_fn
        _today = _today_fn.today().isoformat()
        approved_row = {col: clean.get(col, "" if col not in ("Tier", "Outbreak") else 0)
                        for col in SCHEMA}

        # ── Deterministic publish gate (audit 2026-08-01) ────────────────
        # Runs BEFORE any language-model reviewer and needs neither tokens nor
        # network, so it keeps working when the Gemini quota is exhausted, when
        # the API is down, and when a regulator refuses a TLS handshake. It
        # exists because a 2024 South African recall was published as
        # 2026-07-27 and emailed to subscribers while carrying six defects that
        # no model was needed to see, and because a passenger car, a bath toy,
        # lamp oil, a plastic soup ladle and a sports bottle were all sitting in
        # a pathogen database. See pipeline/_publish_gate.py.
        try:
            from pipeline._publish_gate import publish_blockers  # noqa: WPS433
            _blockers = publish_blockers(clean)
            if _blockers:
                log.warning("publish gate BLOCKED %s: %s",
                            str(clean.get("URL", "<no-url>"))[:90],
                            "; ".join(_blockers))
                clean["Notes"] = (
                    str(clean.get("Notes") or "").strip()
                    + " [publish-gate 2026-08-01: " + "; ".join(_blockers) + "]"
                ).strip()
                _archive(clean)
                continue
        except ImportError:
            pass  # gate module absent — fall through to the old behaviour

        # ── Class language normalisation (audit 2026-05-12) ──────────────
        # Recall.__post_init__() runs _normalize_class_language() at
        # scraper time, but rows that enter Pending via gap-finders or
        # URL-gate enrichment paths bypass that constructor entirely —
        # they're built as plain dicts and travel through validate_pending_row
        # / claude_check / promote_approved without ever being instantiated
        # as a Recall dataclass. Consequence: French/Italian/Spanish/Portuguese/
        # German Class strings ("volontaire (sans arrêté préfectoral)",
        # "richiamo volontario", "rückruf", etc.) leaked all the way into
        # the public Recalls sheet, breaking the English-only convention
        # operators rely on for weekly-report rendering and the language
        # convention "only Company/Brand may stay in source language".
        # Normalising here — at the LAST gate before write — guarantees
        # every Class value in Recalls is the canonical English short form
        # ("Voluntary", "Mandatory", "Recall", "Alert", "Class I", etc.)
        # regardless of which scraper or gap-finder path admitted the row.
        # Local import avoids the merge_master ↔ scrapers cycle.
        try:
            from scrapers._models import _normalize_class_language  # noqa: WPS433
            approved_row["Class"] = _normalize_class_language(
                approved_row.get("Class") or "")
        except Exception as exc:
            # Never let normalisation crash promotion — log and fall through
            # with the original value. Worst case: an un-normalised Class
            # leaks (same as the pre-2026-05-12 behaviour).
            log.warning("Class normalisation skipped for %s: %s: %s",
                        approved_row.get("URL", "<no-url>"),
                        type(exc).__name__, str(exc)[:80])

        # ── Always-Tier-1 enforcement (added 2026-07-14) ────────────────
        # Operator rule: Listeria, Salmonella, E. coli/STEC, C. botulinum,
        # cereulide, Cronobacter, Hepatitis A are ALWAYS Tier 1 — no matter
        # what the source Class or an AI reviewer assigned. Rows were leaking
        # in at Tier 2/3 (RappelConso "voluntary" Listeria recalls tiered
        # from Class; Salmonella at Tier 2). This is the LAST gate before the
        # row is written to Recalls, so enforcing here catches every path
        # (scraper, gap-finder, URL-gate, AI enrichment) in one place.
        try:
            from pipeline._pathogen_scope import enforce_tier1 as _enforce_tier1
            _enforce_tier1(approved_row)
        except Exception as exc:  # never let this crash promotion
            log.warning("Tier-1 enforcement skipped for %s: %s: %s",
                        approved_row.get("URL", "<no-url>"),
                        type(exc).__name__, str(exc)[:80])

        approved_row["DateAdded"] = _today
        approved_row["LastUpdated"] = _today
        approved_row["LastChecked"] = ""
        approved_row["report_week"] = compute_report_week(approved_row.get("Date", ""))

        # ── Analytical schema, filled at promote time (audit 2026-08-28) ──
        # RECALLS_SCHEMA carries fourteen analytical columns, but
        # `approved_row` above is built from SCHEMA alone, so a newly
        # promoted row reached the Recalls sheet with all fourteen blank.
        # Nothing else filled them: pipeline/enrich_schema.py was written,
        # tested and registered, and then not invoked by any workflow. The
        # 1,532 rows enriched by hand on 2026-08-28 would therefore have
        # been the last enriched rows in the register — the corpus would
        # have kept growing while the statistical schema stopped.
        #
        # enrich_schema.derive() is pure: no network, no workbook, no model
        # call. It reads Reason/Notes/Class/Product/Pathogen off the row and
        # returns the RASFF taxonomy fields plus the four product axes. It
        # is cheap enough to run per row at promote time, which is the only
        # place guaranteed to see every row exactly once.
        #
        # The daily sweep (.github/workflows/enrich-schema.yml) is still
        # needed and does NOT duplicate this: it re-derives rows whose text
        # was edited after promotion by the url-gate, the review agents or
        # an operator. It skips any row whose EnrichedBy reads "human".
        try:
            from pipeline.enrich_schema import derive as _derive_schema
            _values, _tier = _derive_schema(approved_row)
            approved_row.update(_values)
            approved_row["EnrichedBy"] = "enrich-schema/1"
            approved_row["EnrichedAt"] = _today
            approved_row["EnrichmentTier"] = _tier
        except Exception as exc:  # never let enrichment block a promotion
            log.warning("Schema enrichment skipped for %s: %s: %s",
                        approved_row.get("URL", "<no-url>"),
                        type(exc).__name__, str(exc)[:80])
            for _col in _ENRICHMENT_COLUMNS:
                approved_row.setdefault(_col, "")

        new_approved.append(approved_row)

    rejected_kept = sum(1 for r in kept_in_pending if r.get("Status") == STATUS_REJECTED)
    log.info("Promotion: %d approved -> Recalls, %d kept in Pending (%d rejected, "
             "%d archived to Weekly_Rejected)",
             len(new_approved), len(kept_in_pending), rejected_kept,
             len(archived_rejected))
    return new_approved, kept_in_pending, archived_rejected


# ---------------------------------------------------------------------------
# Sort / Save
# ---------------------------------------------------------------------------
def sort_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Sort newest first by Date string (YYYY-MM-DD sorts lexically).

    Defensive: rows can land here with Date as a datetime/date object when
    a cell in the xlsx was manually edited and Excel auto-typed it as a
    date. Coerce every key to a YYYY-MM-DD string before comparing so the
    sort never crashes on mixed types.
    """
    def _key(r: Dict[str, Any]) -> str:
        d = r.get("Date")
        if d is None or d == "":
            return ""
        # datetime / date object → ISO string
        if hasattr(d, "strftime"):
            try:
                return d.strftime("%Y-%m-%d")
            except (TypeError, ValueError):
                return ""
        # Anything else → string (truncate to first 10 chars to drop time)
        return str(d)[:10]
    return sorted(rows, key=_key, reverse=True)


# ─── Transcription-artifact normalisation (audit 2026-07-30) ────────────────
# RappelConso's API emits U+00A4 (¤) where its source text had a LINE BREAK:
# as a list-item separator ("marque auchan le charcutier :¤salade museau boeuf
# 250g, ¤salade de museau…"), as a boundary between two run-together items
# ("saucisson sec pur porc 1 kg¤salade de museau 250g"), and as a stray
# trailing character ("Suspicion de contamination bactériologique
# (Listeria)¤"). 34 published rows carried it into Product, Reason and Notes,
# where it renders as a currency sign in the weekly reports and the daily
# briefs.
#
# The codebase already knew the character was an artifact —
# pipeline/_gap_finder_guards.py:161 uses it as a REJECT signal
# (`_ARTIFACT_RE = re.compile(r"[\u00A4\u00D7\uFFFD]")`) — but nothing ever
# stripped it from rows that arrived through the normal scraper path.
#
# Normalised HERE rather than in the RappelConso scraper for the same reason
# the Class guard lives here: rows are updated in place after promotion by
# url-gate and the enrichment passes, so a scraper-side clean can be undone.
# This is the one gate every published row passes through.
_TEXT_ARTIFACT = "\u00a4"


def _strip_text_artifacts(value: Any) -> Any:
    """Replace RappelConso's U+00A4 line-break artifact with real punctuation.

    Leaves any value without the character untouched (identity), so this is
    safe to run over every cell on every write.
    """
    if not isinstance(value, str) or _TEXT_ARTIFACT not in value:
        return value
    t = re.sub(r"\s*\u00a4\s*$", "", value)          # trailing -> drop
    t = re.sub(r"([,:;])\s*\u00a4\s*", r"\1 ", t)     # already punctuated
    t = re.sub(r"\s*\u00a4\s*", "; ", t)             # otherwise a boundary
    return re.sub(r"\s{2,}", " ", t).strip()


def _write_sheet(wb: Workbook,
                 sheet_name: str,
                 schema: List[str],
                 rows: List[Dict[str, Any]],
                 header_fill: PatternFill = None) -> None:
    """(Re)create a sheet with given schema + rows.

    Defensive: Date cells are forced to YYYY-MM-DD strings with General
    number_format so an upstream datetime never gets written back as a
    typed-date cell (which would re-introduce the Excel-serial-leak bug).
    """
    from datetime import datetime as _dt, date as _dt_date
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    for i, h in enumerate(schema, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True)
        if header_fill is not None:
            c.fill = header_fill
    # Absolute-final always-Tier-1 guard (added 2026-07-14). Applies only
    # to the Recalls sheet (never Pending/Weekly_Rejected). Catches any row
    # that reached the writer without passing through promote_approved —
    # e.g. a hand-edited xlsx re-saved by the pipeline, or a direct merge.
    if sheet_name == "Recalls":
        try:
            from pipeline._pathogen_scope import enforce_tier1 as _enforce_tier1
            for _row in rows:
                _enforce_tier1(_row)
        except Exception:
            pass  # writer must never crash on the guard


    # Transcription artifacts — every sheet, every text column.
    _artifact_hits = 0
    for _row in rows:
        for _col in ("Product", "Reason", "Company", "Brand", "Notes", "Pathogen"):
            _v = _row.get(_col)
            _n = _strip_text_artifacts(_v)
            if _n is not _v and _n != _v:
                _row[_col] = _n
                _artifact_hits += 1
    if _artifact_hits:
        log.info("Stripped U+00A4 transcription artifact from %d cell(s) [%s]",
                 _artifact_hits, sheet_name)

    # ── Page status banners folded into Company (audit 2026-08-02) ─────────
    # FSANZ prefixes the <h1> of an amended alert with its own status banner:
    #
    #   "UPDATED 30.07.26 | Auxico (Perth) Pty Ltd - LGM HOT CHILLI OIL 275G"
    #
    # The scraper splits that title on the first " - " to get Company and
    # Product, so the banner lands in Company and the row is published naming
    # a company that does not exist. It also defeats every Company-keyed
    # comparison — content dedup, the FSAI-style identity key, and an
    # operator's eye scanning the sheet.
    #
    # Stripped at the writer, on every sheet, for the same reason the Class
    # and Country guards are: rows are updated in place after promotion by the
    # url-gate and enrichment passes, so a scraper-side clean can be undone,
    # and Pending/Weekly_Review are read by humans too.
    try:
        from pipeline._publish_gate import strip_title_status_prefix  # noqa
        if "Company" in schema:
            for _row in rows:
                _raw = _row.get("Company")
                _clean = strip_title_status_prefix(_raw)
                if isinstance(_raw, str) and _clean != _raw:
                    log.info("Company status banner stripped at writer [%s]: "
                             "%r -> %r", sheet_name, _raw[:60], _clean[:60])
                    _row["Company"] = _clean
    except Exception as exc:
        log.warning("Company banner strip skipped at writer [%s]: %s: %s",
                    sheet_name, type(exc).__name__, str(exc)[:80])

    # ── English-output policy (audit 2026-08-02) ───────────────────────────
    # Operator rule: "everything in English except brand / or product name".
    #
    # 157 published rows carried a non-English Reason. The single largest
    # cause was mechanical and is fixed here: RASFF writes its notification
    # subject BILINGUALLY into both Reason and Product —
    #
    #   "Presencia de Salmonela spp en salchichón procedente de España //
    #    Presence of Salmonella spp. in cured sausage from Spain;
    #    risk: serious; category: meat and meat products"
    #
    # — sometimes native-first, sometimes English-first, with "//", "/",
    # "/////" or ";" between. split_bilingual() scores both halves and keeps
    # the English one, preserving the "; risk: …; category: …" tail.
    #
    # Applied at the writer, on every sheet, for the same reason the Class and
    # Country guards are: rows are updated in place after promotion by the
    # url-gate and enrichment passes, so a scraper-side clean can be undone.
    #
    # ONLY the mechanical split runs here. Translating a whole-sentence French
    # or German motif needs the verified table in pipeline/_language.py and is
    # a deliberate, auditable act — never something a writer does silently. A
    # row that is still non-English after the split is LOGGED, not mangled.
    #
    # Product is included because RASFF puts the same subject string there,
    # and that is a description, not a name. A genuine foreign product name
    # ("brie a l'ail", "Χούμους", "Freshona Bio Beerenmischung") does not
    # split into two languages, so it is never touched — which is exactly the
    # brand/product-name exemption the rule asks for.
    # ── PATHOGEN LABEL NOTATION (audit 2026-08-09) ──────────────────────
    # Operator: "why in some we have listeria and others listeria mono..".
    #
    # Counting the register settles it — bare "Listeria" is not a convention,
    # it is an outlier:
    #     Listeria monocytogenes  494        Salmonella        382
    #     Listeria                  2        Salmonella spp.    11
    #                                        Salmonella spp      2
    #
    # Two different problems hide in that, and only ONE is safe to fix
    # mechanically:
    #
    #   NOTATION — "Salmonella spp." and "Salmonella spp" are the same claim
    #     as "Salmonella": genus named, species unspecified. Three spellings
    #     of one fact split the pathogen table in every weekly and monthly
    #     report — the same ambiguity the W32 reviewer raised about
    #     "Salmonella spp. 15" beside "Salmonella Javiana 1". Collapsed here.
    #
    #   SPECIFICITY — a bare "Listeria" where the source says "Listeria
    #     monocytogenes" is a LOST FACT, not a spelling. It is NOT repaired
    #     here: promoting a genus to a species is a claim about the world and
    #     belongs in an audited row-by-row fix against the source, never in a
    #     silent writer rewrite. Both offending rows were repaired that way on
    #     2026-08-09 — each one's own Reason text already named the species.
    #
    # A trailing "spp"/"spp." is stripped ONLY when it directly follows the
    # genus. "Salmonella Javiana" and "Listeria innocua" name a member of the
    # genus and are never touched — that distinction is the whole point.
    try:
        import re as _re_p                             # noqa: WPS433
        _SPP = _re_p.compile(
            r"^(Listeria|Salmonella|Escherichia|Campylobacter|Vibrio|"
            r"Bacillus|Clostridium|Shigella|Cronobacter|Yersinia)"
            r"\s+spp\.?$", _re_p.IGNORECASE)
        for _row in rows:
            _p = _row.get("Pathogen")
            if not isinstance(_p, str) or not _p.strip():
                continue
            # Multi-pathogen cells are comma-separated; normalise each part.
            _parts = [p.strip() for p in _p.split(",")]
            _fixed = [(_SPP.sub(r"\1", p) if _SPP.match(p) else p)
                      for p in _parts]
            _new = ", ".join(x for x in _fixed if x)
            if _new and _new != _p:
                _row["Pathogen"] = _new
    except Exception:                                  # pragma: no cover
        pass

    # ── HTML ENTITIES (audit 2026-08-09) ────────────────────────────────
    # Two USDA FSIS rows reached Pending as
    #     "City Foods, Inc./Bea&#039;s Best Corned Beef"
    #     "Mary&#039;s Harvest Fresh Foods, Inc."
    # — the scraper read the entity-encoded HTML source and never decoded it.
    # These render literally on the dashboard and in subscriber email, and
    # they break exact-match dedup against the same firm captured correctly
    # elsewhere.
    #
    # Decoded HERE rather than in each scraper because this is the single
    # writer every sheet write passes through, and because the fix must apply
    # to the rows already sitting in Pending, not only to future scrapes.
    # unescape() is idempotent and leaves a bare "&" alone, so re-running it
    # on already-clean text is a no-op.
    try:
        import html as _html_mod                       # noqa: WPS433
        for _row in rows:
            for _col in ("Company", "Brand", "Product", "Reason", "Notes"):
                _val = _row.get(_col)
                if isinstance(_val, str) and "&" in _val and ";" in _val:
                    _dec = _html_mod.unescape(_val)
                    if _dec != _val:
                        _row[_col] = _dec
    except Exception:                                  # pragma: no cover
        pass

    try:
        from pipeline._language import (  # noqa: WPS433
            split_bilingual as _split_bilingual,
            looks_non_english as _looks_non_english,
        )
        _still_foreign = 0
        for _col in ("Reason", "Product"):
            if _col not in schema:
                continue
            for _row in rows:
                # PRODUCT IS SCOPED TO RASFF (audit 2026-08-04, second pass).
                # RASFF is the only source that writes a bilingual notification
                # SUBJECT into Product; everywhere else Product is a NAME, and
                # a name may contain "/" or ";" for entirely innocent reasons.
                # Running the splitter on all sources truncated two real
                # products before this scope was added:
                #   FDA   "Hellas Meze Golden Smoked Whole Herring, vacuum-
                #          packaged, refrigerated; production date 4/12/2025,
                #          best before 4/12/2026, lot L120425F54..."
                #          -> everything after "refrigerated" lost
                #   EFET  'Σαλάτα "ΜΑΡΟΥΛΕΝΙΑ" — μαρούλι romaine... (ΒΙ.ΠΕ. /
                #          Industrial Area, Central Macedonia)'
                #          -> reduced to "Industrial Area, Central Macedonia)."
                # The second is the worse one: the product name disappeared
                # entirely and what survived was a fragment of the address.
                if _col == "Product" and \
                        str(_row.get("Source") or "") != "RASFF (EU)":
                    continue
                _v = _row.get(_col)
                if not isinstance(_v, str) or not _looks_non_english(_v):
                    continue
                _en = _split_bilingual(_v)
                if _en and _en != _v:
                    log.info("Bilingual %s split at writer [%s]: kept the "
                             "English half (%r)", _col, sheet_name, _en[:70])
                    _row[_col] = _en
                elif _col == "Reason":
                    _still_foreign += 1
        if _still_foreign:
            log.warning(
                "%d row(s) still carry a non-English Reason after the writer "
                "split [%s]. Add them to REASON_EN in pipeline/_language.py — "
                "they are NOT machine-translated on purpose.",
                _still_foreign, sheet_name)

        # ── US spelling at the writer (audit 2026-08-14) ────────────────
        # Operator: "mould must be mold to us US english". Applied HERE, on
        # every write, because the alternative is repairing rows after each
        # Australian or British recall arrives — FSANZ and the FSA write
        # British English and always will.
        #
        # It is a SCOPE control as well as a style one. _publish_gate lists
        # "mould"/"mold" among the out-of-scope hazard terms, yet three
        # FSANZ rows with Pathogen "Mould" sat in Recalls until 2026-08-14:
        # the British spelling did not resolve to the quality/spoilage
        # class, the US spelling does. Normalising before the gate sees the
        # row is what makes that check work.
        #
        # Pathogen / Reason / Class ONLY. Product and Brand are exempt
        # under the English-output rule and must match the pack — and
        # "moulded / demoulded" in a Product means shaped in a mould, not
        # fungus (see the false-friend guard in _language.americanize).
        try:
            from pipeline._language import americanize as _us  # noqa: WPS433
            _n_us = 0
            for _row in rows:
                for _col in ("Pathogen", "Reason", "Class"):
                    _v = _row.get(_col)
                    if not isinstance(_v, str) or not _v:
                        continue
                    _a = _us(_v)
                    if _a != _v:
                        _row[_col] = _a
                        _n_us += 1
            if _n_us:
                log.info("US-spelling normalisation applied to %d field(s) "
                         "[%s]", _n_us, sheet_name)
        except Exception as _ue:                              # noqa: BLE001
            log.warning("US-spelling normalisation skipped [%s]: %s",
                        sheet_name, _ue)
    except Exception as exc:
        log.warning("English-output guard skipped at writer [%s]: %s: %s",
                    sheet_name, type(exc).__name__, str(exc)[:80])

    # ── Country-name canonicalisation (audit 2026-08-01) ───────────────────
    # Country is a join key: it drives Region, the country counts in the weekly
    # and monthly reports, and the per-country filters subscribers set on their
    # alert rules. Two spellings of one country silently split all of that.
    #
    # The workbook had 82 rows saying 'United States' and 15 saying 'USA' — and
    # the RASFF rows were self-contradictory, spelling it 'Origin: United
    # States' in their own Company field while Country said 'USA'. Same story
    # for the UK and a few others, so the aliases are canonicalised here at the
    # single writer choke point rather than in each scraper, exactly as the
    # Class guard below.
    _COUNTRY_ALIASES = {
        "usa": "United States", "u.s.a.": "United States", "us": "United States",
        "u.s.": "United States", "united states of america": "United States",
        "uk": "United Kingdom", "u.k.": "United Kingdom",
        "great britain": "United Kingdom", "england": "United Kingdom",
        "holland": "Netherlands", "the netherlands": "Netherlands",
        "czech republic": "Czechia", "turkiye": "Turkey", "türkiye": "Turkey",
        "republic of ireland": "Ireland", "south korea": "Korea, South",
        "russian federation": "Russia",
    }
    # Source labels carry the jurisdiction suffix so two agencies with the same
    # acronym never collide, and so the reports read consistently. 'FSIS' alone
    # is ambiguous; 'USDA FSIS' is the established label on all 9 US rows.
    _SOURCE_ALIASES = {
        "fsis": "USDA FSIS", "usda": "USDA FSIS", "usda-fsis": "USDA FSIS",
        "usda fsis (us)": "USDA FSIS",
        "ncc": "NCC (ZA)", "efet": "EFET (GR)", "aesan": "AESAN (ES)",
        "fsai": "FSAI (IE)", "fsa": "FSA (UK)",
    }
    try:
        if "Country" in schema:
            for _row in rows:
                _raw = str(_row.get("Country") or "").strip()
                _canon = _COUNTRY_ALIASES.get(_raw.lower())
                if _canon and _canon != _raw:
                    log.info("Country canonicalised at writer [%s]: %r -> %r",
                             sheet_name, _raw, _canon)
                    _row["Country"] = _canon
        if "Source" in schema:
            for _row in rows:
                _raw = str(_row.get("Source") or "").strip()
                _canon = _SOURCE_ALIASES.get(_raw.lower())
                if _canon and _canon != _raw:
                    log.info("Source canonicalised at writer [%s]: %r -> %r",
                             sheet_name, _raw, _canon)
                    _row["Source"] = _canon
    except Exception as exc:
        log.warning("Country canonicalisation skipped at writer [%s]: %s: %s",
                    sheet_name, type(exc).__name__, str(exc)[:80])

    # ── Absolute-final Class language normalisation (audit 2026-07-30) ──────
    # Same rationale as the Tier-1 guard above, but deliberately applied to
    # EVERY sheet, not just Recalls.
    #
    # promote_approved normalises Class at its own last gate (audit
    # 2026-05-12) and Recall.__post_init__ normalises at scraper time, yet 41
    # rows dated 2026-05-12..2026-07-28 still reached the published Recalls
    # sheet holding raw French. Distribution over the 535 RappelConso rows
    # when this was found:
    #
    #     394  'Voluntary'
    #      83  'Mandatory'
    #      38  'volontaire (sans arrete prefectoral)'   <- un-normalised
    #      17  'Recall'
    #       3  'impose par arrete prefectoral'          <- un-normalised
    #
    # All 41 carried a [url-gate ...] note: they were updated in place after
    # promotion, so neither earlier gate ran again.
    #
    # WHY ALL SHEETS (audit 2026-07-30, second pass). The first version of
    # this guard sat inside the `sheet_name == "Recalls"` block. Recalls went
    # clean and stayed clean — and four fresh RappelConso rows promptly
    # landed in Pending still holding 'volontaire (sans arrete prefectoral)'.
    # Class is a LANGUAGE normalisation, not a promotion decision: there is
    # no sheet on which raw French is the correct value, operators review
    # Pending and Weekly_Review by eye, and _normalize_class_language is
    # idempotent so running it everywhere costs nothing. Scoping it to one
    # sheet just moved the leak.
    try:
        from scrapers._models import (  # noqa: WPS433
            _normalize_class_language as _norm_cls,
        )
        if "Class" in schema:
            for _row in rows:
                _raw = _row.get("Class") or ""
                _norm = _norm_cls(_raw)
                if _norm and _norm != _raw:
                    log.info("Class normalised at writer [%s]: %r -> %r (%s)",
                             sheet_name, _raw, _norm,
                             str(_row.get("URL", ""))[:80])
                    _row["Class"] = _norm
    except Exception as exc:
        log.warning("Class normalisation skipped at writer [%s]: %s: %s",
                    sheet_name, type(exc).__name__, str(exc)[:80])

    for r_idx, row in enumerate(rows, 2):
        for c_idx, col in enumerate(schema, 1):
            v = row.get(col, "")
            if col in ("Tier", "Outbreak"):
                try:
                    v = int(v) if v not in ("", None) else 0
                except (ValueError, TypeError):
                    v = 0
            elif col == "Date":
                # Force every Date cell to a string + General format
                if isinstance(v, (_dt, _dt_date)):
                    try:
                        v = v.strftime("%Y-%m-%d")
                    except (TypeError, ValueError):
                        v = ""
                elif v not in (None, "") and not isinstance(v, str):
                    v = str(v)[:10]
            cell = ws.cell(row=r_idx, column=c_idx, value=v)
            if col == "Date":
                cell.number_format = "General"
    ws.freeze_panes = "A2"


def save_xlsx_with_pending(
    approved_rows: List[Dict[str, Any]],
    pending_rows: List[Dict[str, Any]],
    xlsx_path: Path,
    newly_rejected_rows: Optional[List[Dict[str, Any]]] = None,
) -> None:
    """
    Save BOTH sheets (Recalls + Pending), preserving NEWS sheet if present.
    Sheet order: Recalls (0), Pending (1), (auxiliary — Weekly_Review,
    Weekly_Rejected, etc.), NEWS (last).

    Audit 2026-05-08: optional `newly_rejected_rows` arg. These rows are
    not written here — callers mirror them into Weekly_Rejected via
    weekly_rejected_capture.record_rejections immediately after this
    returns. The kwarg is kept for API compatibility and is a no-op.

    Two rejection sheets exist and BOTH are preserved untouched by this
    function (see the note in the body, 2026-08-20):
        Weekly_Rejected   rolling; emptied every Thursday after the email
        Rejected          permanent; append-only; never wiped
    An earlier version of this docstring said the "Rejected" sheet "is
    removed". It was, by a line right below — which deleted 138 archived
    rejections the Thursday wipe had just written there.
    """
    if xlsx_path.exists():
        wb = load_workbook(xlsx_path)
    else:
        wb = Workbook()
        if wb.active and wb.active.max_row == 1 and wb.active.max_column == 1:
            wb.remove(wb.active)

    # Write Recalls (approved published data)
    _write_sheet(wb, "Recalls", RECALLS_SCHEMA, approved_rows)

    # Write Pending (amber-ish header fill to make the tab visually distinct)
    pending_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    _write_sheet(wb, "Pending", PENDING_SCHEMA, pending_rows, header_fill=pending_fill)

    # ── Rejected sheet — REMOVED (audit 2026-05-11, per operator decision) ─
    # The standalone "Rejected" sheet was duplicate audit infrastructure.
    # Weekly_Rejected already captures every rejection with the same
    # schema + Thursday-review semantics, and is the single source the
    # Apps Script mailer reads. Keeping both meant rows could drift
    # between them (33 vs 44 dedup gap observed 2026-05-11 17:23 run).
    #
    # newly_rejected_rows is still accepted for backward compatibility
    # with callers that pass it, but it's now passed through to the
    # Weekly_Rejected mirroring done by callers (pipeline.claude_check /
    # pipeline.gemini_check invoke weekly_rejected_capture.record_rejections
    # immediately after this function returns). No sheet write here.
    #
    # ── THE "Rejected" SHEET IS NO LONGER LEGACY — DO NOT DELETE IT ─────
    # REVERSED 2026-08-20. This block used to read:
    #
    #     if "Rejected" in wb.sheetnames:
    #         del wb["Rejected"]
    #
    # a one-shot cleanup added on 2026-05-11 when "Rejected" really was
    # duplicate audit infrastructure. It is not any more, and the two
    # designs had started destroying each other:
    #
    #   * tools/wipe_weekly_rejected.py (2026-08-15) MOVES rows into a
    #     permanent "Rejected" sheet on the Thursday reset instead of
    #     deleting them, because the wipe was erasing the pipeline's only
    #     memory of what had been rejected and the gap-finders re-found
    #     every one of those rows within a cycle.
    #   * load_rejected_urls() (2026-08-18) reads BOTH sheets, so a
    #     rejection stays sticky past the Thursday reset.
    #   * This line then deleted the sheet on the very next save.
    #
    # Measured on the live workbook: the 2026-08-20 wipe archived 138 rows
    # into "Rejected" at 14:51, and the next save_xlsx_with_pending call
    # removed all 138 without a word. The rolling sheet had already been
    # emptied, so that is the whole rejection history gone in one step.
    #
    # The 2026-05-11 concern — two sheets drifting apart — no longer
    # applies: the sheets now have distinct, documented roles.
    #     Weekly_Rejected  rolling; emptied every Thursday after the email
    #     Rejected         permanent; append-only; never wiped
    #
    # The sheet is left exactly as found. This function writes Recalls and
    # Pending; it has no business editing the audit trail.
    if "Rejected" in wb.sheetnames:
        log.debug("Preserving permanent Rejected sheet (%d rows)",
                  wb["Rejected"].max_row - 1)

    # Ensure NEWS sheet exists (empty if it wasn't there before)
    if "NEWS" not in wb.sheetnames:
        news = wb.create_sheet("NEWS")
        for i, h in enumerate(NEWS_HEADERS, 1):
            c = news.cell(row=1, column=i, value=h)
            c.font = Font(bold=True)
        news.freeze_panes = "A2"

    # Reorder: Recalls, Pending, (others — Weekly_Review/Weekly_Rejected/etc.), NEWS last
    # (audit 2026-05-11: standalone "Rejected" sheet removed — rejections
    # are mirrored to Weekly_Rejected by callers, so it's no longer in
    # the explicit order list.)
    ordered = ["Recalls", "Pending"]
    others = [s for s in wb.sheetnames
              if s not in ("Recalls", "Pending", "NEWS")]
    ordered += others
    if "NEWS" in wb.sheetnames:
        ordered.append("NEWS")
    wb._sheets = [wb[s] for s in ordered]

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
    log.info("Saved xlsx: Recalls=%d, Pending=%d -> %s",
             len(approved_rows), len(pending_rows), xlsx_path)


def save_json(rows: List[Dict[str, Any]], json_path: Path) -> None:
    """
    DEPRECATED — writes recalls.json from an in-memory list of rows.

    Using this is an architectural violation: recalls.json MUST mirror what's
    on the Recalls sheet of recalls.xlsx, not an arbitrary in-memory list.
    If the xlsx write fails or gets interrupted, the json would diverge from
    the file that's actually committed.

    Use `mirror_json_from_xlsx(xlsx_path, json_path)` instead. Kept here only
    so legacy callers don't crash outright during the transition.
    """
    log.warning("save_json (in-memory) is deprecated; "
                "use mirror_json_from_xlsx for guaranteed xlsx->json sync")
    json_path.parent.mkdir(parents=True, exist_ok=True)
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(rows, f, ensure_ascii=False, indent=1, default=str)
    log.info("Saved %d approved rows to %s", len(rows), json_path)


def mirror_json_from_xlsx(xlsx_path: Path, json_path: Path) -> int:
    """
    Write recalls.json as a strict mirror of the Recalls sheet in recalls.xlsx.

    This is the ONLY sanctioned way to produce recalls.json. It guarantees
    that json can never drift from xlsx: we read the file that was just
    committed to disk, normalise types (dates to ISO strings), and serialise.

    INTERNAL columns (DateAdded, LastUpdated, LastChecked) are STRIPPED
    here so they don't leak into the public-facing dashboard. Only the
    14 SCHEMA columns make it to recalls.json.

    Returns the number of rows written.
    """
    rows = load_existing(xlsx_path)
    out = []
    for r in rows:
        rec = {}
        for k, v in r.items():
            # Skip internal-only tracking columns — public consumers never see these
            if k in RECALLS_INTERNAL_COLUMNS:
                continue
            if hasattr(v, "isoformat"):
                rec[k] = v.isoformat()[:10]
            elif v is None:
                rec[k] = ""
            else:
                rec[k] = v
        out.append(rec)
    json_path.parent.mkdir(parents=True, exist_ok=True)
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=1, default=str)
    log.info("Mirrored %d rows from xlsx -> %s", len(out), json_path)
    return len(out)


# ---------------------------------------------------------------------------
# Daily-brief rebuild helper (used by url_gate, claude_check, merge_master CLI)
# ---------------------------------------------------------------------------
def rebuild_daily_briefs_for_promoted(
    new_approved: List[Dict[str, Any]],
    full_approved: List[Dict[str, Any]],
) -> Tuple[List[str], List[str]]:
    """
    After Pending → Recalls promotion, rebuild the per-date daily-brief HTML
    for every date that gained at least one new row. Without this, the
    dashboard's DAILY tab and rolling 7-day display stay stale until the
    next scheduled daily-recall-search run.

    Args:
        new_approved : rows just promoted to Recalls this run
        full_approved: the FULL Recalls sheet AFTER promotion (so we render
                       the complete day, not just the newly-added rows)

    Returns:
        (rebuilt_brief_paths, files_to_commit) — caller is responsible for
        adding `files_to_commit` to its git_commit_and_push call. Both lists
        are empty when nothing was promoted or the brief renderer module
        is unavailable.
    """
    files_to_commit: List[str] = []
    rebuilt_briefs: List[str] = []
    if not new_approved:
        return rebuilt_briefs, files_to_commit

    try:
        from pipeline.daily_recall_search import (  # noqa: WPS433
            render_daily_html, update_daily_index,
        )
        from scrapers._models import Recall as _Recall  # noqa: WPS433
    except ImportError as ie:
        log.warning("Cannot import brief renderer (%s) — skipping daily "
                    "brief rebuild", ie)
        return rebuilt_briefs, files_to_commit

    from collections import defaultdict
    from datetime import date as _date

    # ------------------------------------------------------------------
    # Audit 2026-05-13: today (Athens) must NEVER be rebuilt here.
    #
    # The daily-brief design rule (see pipeline.daily_recall_search.main,
    # ~line 1388) is explicit: the dashboard window is [target..target-6]
    # where target = yesterday Athens. Today's brief is rendered TOMORROW
    # at 10:00 Athens, when today becomes the new "target". Rendering it
    # earlier:
    #   • inflates the dashboard to 8 cards (today + yesterday + 6 back)
    #   • pushes the oldest day off the visible feed
    #   • writes a partial brief based on whatever rows happen to have
    #     been promoted before the day ends, instead of the complete day
    #
    # The bug: this helper iterated every Date present in `new_approved`
    # without filtering today/future. When the 17:11 claude_check promoted
    # rows dated 2026-05-13, docs/daily/2026-05-13.html got written and
    # daily-index.json got a 2026-05-13 entry — exactly what the 10:00
    # Thursday workflow is supposed to do, and what George explicitly
    # forbids today's runs from doing.
    #
    # Fix: compute today_athens once, skip any date_str >= today_athens
    # from the rebuild loop. Promoted rows dated today still land in the
    # Recalls sheet (xlsx) and in recalls.json — only the daily-brief HTML
    # + daily-index.json entry are deferred until tomorrow morning.
    # ------------------------------------------------------------------
    try:
        from zoneinfo import ZoneInfo  # noqa: WPS433
        _today_athens = datetime.now(ZoneInfo("Europe/Athens")).date()
    except Exception:  # noqa: BLE001 — fall back to UTC if tzdata missing
        _today_athens = datetime.now(timezone.utc).date()
    _today_iso = _today_athens.isoformat()

    # Group newly-promoted rows by Date, deferring today/future dates.
    # ------------------------------------------------------------------
    # Audit 2026-08-28: the window has a FAR side too.
    #
    # The guard above defers today/future. Nothing guarded the other end,
    # and update_daily_index computes its retention cutoff from the
    # target_date it is handed, not from today:
    #
    #     cutoff = target_date - (KEEP_DAYS - 1)
    #
    # So promoting a back-dated row — a gap-finder catching a June notice
    # in August — called update_daily_index(2026-06-19), which set the
    # cutoff to 2026-06-13 and RETAINED every entry after it. The rolling
    # 7-day feed grew to 9 entries, with stray briefs from June and July
    # on the dashboard beside the current week. Nothing errored; the index
    # was written successfully every time.
    #
    # Back-dated rows still land in Recalls and recalls.json — they are
    # real data. They have no place in a feed whose contract is "the last
    # seven days". Asserted in tests/test_daily_brief_window.py.
    #
    # The feed is anchored to YESTERDAY: the brief for date D is rendered
    # on D+1, so the visible span is [target .. target-6] with
    # target = today-1, making today-7 the oldest date still on screen.
    # Using today-6 here would skip rebuilding the oldest visible day and
    # leave it stale.
    # ------------------------------------------------------------------
    from datetime import timedelta as _timedelta  # noqa: WPS433
    _WINDOW_DAYS = 7
    _oldest_iso = (_today_athens - _timedelta(days=_WINDOW_DAYS)).isoformat()

    by_date: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    _deferred: Dict[str, int] = {}
    _too_old: Dict[str, int] = {}
    for r in new_approved:
        d = str(r.get("Date") or "").strip()[:10]
        if not d:
            continue
        if d >= _today_iso:
            _deferred[d] = _deferred.get(d, 0) + 1
            continue
        if d < _oldest_iso:
            _too_old[d] = _too_old.get(d, 0) + 1
            continue
        by_date[d].append(r)

    if _deferred:
        for d_iso, n in sorted(_deferred.items()):
            log.info("Skip brief rebuild for %s (%d row(s)): today/future "
                     "Athens — will be rendered by tomorrow's 10:00 "
                     "daily-recall-search run (today_athens=%s)",
                     d_iso, n, _today_iso)

    if _too_old:
        for d_iso, n in sorted(_too_old.items()):
            log.info("Skip brief rebuild for %s (%d row(s)): older than the "
                     "%d-day daily window (oldest=%s). The row is in Recalls "
                     "and recalls.json; only the daily feed excludes it.",
                     d_iso, n, _WINDOW_DAYS, _oldest_iso)

    # Fast-lookup of ALL Recalls by date so we render the full day, not
    # only the newly-promoted rows.
    full_by_date: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for r in full_approved:
        d = str(r.get("Date") or "").strip()[:10]
        if d:
            full_by_date[d].append(r)

    for date_str in sorted(by_date.keys(), reverse=True):
        try:
            y, m, d = date_str.split("-")
            target = _date(int(y), int(m), int(d))
        except (ValueError, AttributeError):
            log.warning("Skip brief rebuild — bad date '%s'", date_str)
            continue

        day_rows = full_by_date.get(date_str, [])
        recalls_objs = []
        for row in day_rows:
            try:
                recalls_objs.append(_Recall(**{
                    k: (v if v is not None else "")
                    for k, v in row.items()
                    if k in _Recall.__annotations__
                }))
            except Exception as cce:  # noqa: BLE001
                log.debug("skip row coerce: %s", cce)

        try:
            # Audit 2026-05-12: this previously called render_daily_html()
            # but threw away its return value, never writing the HTML file
            # to disk — the function returns the rendered markup but the
            # caller is responsible for persisting it. Only update_daily_index
            # was actually doing IO, so the dashboard's daily-index.json
            # would reflect the new counts while docs/daily/{date}.html
            # stayed stale (or, for newly-promoted dates with no prior
            # brief, never existed at all). Most visible symptom: dashboard
            # DAILY tab showed total=N for a date but clicking it 404'd
            # because the HTML wasn't there. Fixed by capturing the returned
            # HTML and persisting it before the index update.
            html = render_daily_html(target, recalls_objs)
            # Resolve docs/daily path relative to this module file (ROOT
            # is only a local in main()). Mirrors the convention used in
            # pipeline.daily_recall_search.
            _root = Path(__file__).resolve().parent.parent
            brief_path = _root / "docs" / "daily" / f"{date_str}.html"
            brief_path.parent.mkdir(parents=True, exist_ok=True)
            brief_path.write_text(html, encoding="utf-8")
            update_daily_index(target, recalls_objs)
            rebuilt_briefs.append(str(brief_path))
            files_to_commit.append(f"docs/daily/{date_str}.html")
            log.info("Rebuilt daily brief for %s (%d rows) -> %s",
                     date_str, len(recalls_objs), brief_path)
        except Exception as rerr:  # noqa: BLE001
            log.warning("Brief rebuild failed for %s: %s", date_str, rerr)

    if rebuilt_briefs:
        files_to_commit.append("docs/daily-index.json")

    return rebuilt_briefs, files_to_commit


# ---------------------------------------------------------------------------
# Back-compat shims (kept so legacy call sites don't break)
# ---------------------------------------------------------------------------
def save_xlsx(rows: List[Dict[str, Any]], xlsx_path: Path) -> None:
    """DEPRECATED: single-sheet save. Kept for any legacy caller."""
    log.warning("save_xlsx (single-sheet) is deprecated — use save_xlsx_with_pending")
    existing_pending = load_pending(xlsx_path)
    save_xlsx_with_pending(rows, existing_pending, xlsx_path)


def merge_new(existing: List[Dict[str, Any]], new_recalls: List[Recall]) -> List[Dict[str, Any]]:
    """
    DEPRECATED: direct merge into Recalls (pre-Pending-sheet behavior).
    Kept for any back-compat call; new code should use append_to_pending +
    promote_approved instead.
    """
    existing_keys = {_dedup_key(r) for r in existing}
    merged = list(existing)
    appended = 0
    for r in new_recalls:
        d = r.to_dict() if isinstance(r, Recall) else dict(r)
        for col in SCHEMA:
            d.setdefault(col, "" if col not in ("Tier", "Outbreak") else 0)
        k = _dedup_key(d)
        if k in existing_keys:
            continue
        existing_keys.add(k)
        merged.append(d)
        appended += 1
    log.info("merge_new (legacy): %d existing + %d new = %d total",
             len(existing), appended, len(merged))
    return merged


# ---------------------------------------------------------------------------
# NEWS sheet merge (for RSS news feed scrapers)
# ---------------------------------------------------------------------------
def _norm_news_link(u) -> str:
    """Canonical form of a NEWS link for dedup. See the audit note above."""
    s = str(u or "").strip().lower()
    if not s:
        return ""
    s = s.split("#", 1)[0].rstrip("/")
    for pre in ("https://", "http://"):
        if s.startswith(pre):
            s = s[len(pre):]
            break
    if s.startswith("www."):
        s = s[4:]
    return s


def _news_dedup_key(row: Dict[str, Any]) -> str:
    """Dedup key for a NEWS row: link URL (lowered, stripped)."""
    # AUDIT 2026-08-14 — TRAILING SLASH AND SCHEME ARE NOT IDENTITY.
    # Two Food Safety News items were re-added as duplicates because the
    # feed emitted them once without a trailing slash and once with:
    #   .../2026/08/states-say-cyclospora-outbreak-is-slowing
    #   .../2026/08/states-say-cyclospora-outbreak-is-slowing/
    # Same article, two keys, so the union kept both. The published
    # timestamps also differ in FORMAT ("2026-08-11 04:05 UT" vs
    # "2026-08-11T04:05:06"), which is the tell that they came from two
    # different fetch paths rather than two real postings.
    # Normalise the URL before keying: strip the trailing slash, drop a
    # leading "www.", and treat http/https as the same document.
    link = _norm_news_link(row.get("Link") or row.get("link"))
    if link:
        return link
    title = (row.get("Title") or row.get("title") or "").strip().lower()[:80]
    return f"{row.get('Published (UTC)', '')}|{title}"


def load_news(xlsx_path: Path) -> List[Dict[str, str]]:
    """Load existing NEWS rows from the xlsx."""
    if not xlsx_path.exists():
        return []
    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        if "NEWS" not in wb.sheetnames:
            wb.close()
            return []
        ws = wb["NEWS"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            d = {}
            for i, v in enumerate(row):
                if i < len(headers) and headers[i]:
                    d[headers[i]] = str(v) if v is not None else ""
            if d.get("Title") or d.get("Link"):
                rows.append(d)
        wb.close()
        return rows
    except Exception as e:
        log.warning("Failed to load NEWS sheet: %s", e)
        return []


def append_news_to_xlsx(
    xlsx_path: Path,
    new_items: List[Dict[str, str]],
) -> int:
    """
    Append new NEWS items to the NEWS sheet, deduped against existing rows.
    Returns count of actually-appended items.
    """
    if not new_items:
        return 0

    existing = load_news(xlsx_path)
    seen_keys = {_news_dedup_key(r) for r in existing}

    to_add = []
    for item in new_items:
        k = _news_dedup_key(item)
        if k in seen_keys:
            continue
        seen_keys.add(k)
        to_add.append(item)

    if not to_add:
        log.info("NEWS merge: 0 new items (all duplicates)")
        return 0

    # Open the workbook and append rows to the NEWS sheet
    wb = load_workbook(xlsx_path)
    if "NEWS" not in wb.sheetnames:
        ws = wb.create_sheet("NEWS")
        for i, h in enumerate(NEWS_HEADERS, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = Font(bold=True)
        ws.freeze_panes = "A2"
    else:
        ws = wb["NEWS"]

    for item in to_add:
        row_vals = [item.get(h, "") for h in NEWS_HEADERS]
        ws.append(row_vals)

    wb.save(xlsx_path)
    log.info("NEWS merge: appended %d new items (total now %d)",
             len(to_add), len(existing) + len(to_add))
    return len(to_add)



# =========================================================================
# CLI entry point — run by the hourly merge-master workflow.
#
# AUDIT 2026-04-29 — promotion semantics tightened:
#   The hourly CLI used to call promote_approved() with rejected_flags driven
#   only by review/url_validator (HTTP HEAD/GET reachability). That meant
#   ANY row with a 200-returning URL was promoted to Recalls — including
#   hallucinated RappelConso fiches (e.g. /fiche-rappel/22180/Interne) that
#   render a soft-200 page even when the fiche ID doesn't exist. This let
#   ~7 hallucinated Gemini gap-finder rows leak from the 07:34 Exa run into
#   Recalls before the once-daily 07:00 Gemini URL gate could catch them.
#
#   New rule (per George 2026-04-29):
#     "Only data from URL Gemini followed by Claude check must be allowed
#      1 or two times per day."
#
#   The hourly CLI is now a JANITOR ONLY — it cleans malformed URLs from
#   Pending and removes Pending rows whose URL has been confirmed dead,
#   but it NEVER promotes to Recalls. Promotion happens exclusively via:
#     1. pipeline/url_gate_gemini.py     (07:00 Athens — Gemini URL gate)
#     2. pipeline/claude_check.py        (07:45 Athens — Claude content check)
#
#   To override (e.g. backfill, manual catch-up), set MERGE_MASTER_PROMOTE=1.
# =========================================================================
if __name__ == "__main__":
    import sys, os
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    )
    from pathlib import Path
    ROOT = Path(__file__).resolve().parent.parent
    XLSX = ROOT / "docs" / "data" / "recalls.xlsx"

    if not XLSX.exists():
        log.info("No recalls.xlsx found — nothing to merge.")
        sys.exit(0)

    approved = load_existing(XLSX)
    pending = load_pending(XLSX)
    log.info("State: %d approved, %d pending", len(approved), len(pending))

    if not pending:
        log.info("No Pending rows — nothing to do.")
        sys.exit(0)

    # ── URL reachability check (cheap janitor pass) ─────────────────────
    # We still validate URLs at every hourly run so dead URLs can be
    # marked rejected in Pending. We do NOT use the result to drive
    # promotion — that would re-create the leak this CLI is here to
    # prevent.
    from review.url_validator import validate_all
    log.info("Validating %d Pending URLs (janitor pass — no promotion)...",
             len(pending))
    validated = validate_all(pending, max_workers=5)

    # Mark broken-URL rows as rejected so they don't promote later.
    # Anything that's currently 'pending' AND has a confirmed bad URL
    # gets stamped REJECTED so the next gate pass skips it.
    rejected_flags = {}
    for idx, row in enumerate(validated):
        check = row.get("_url_check", {})
        # Only reject if the URL check says it's BROKEN. Reachable URLs
        # stay 'pending' until the URL gate validates them properly.
        if not check.get("ok", False):
            rejected_flags[idx] = check.get("reason", "URL check failed")

    # Strip the runtime _url_check field before persisting
    clean_pending = [{k: v for k, v in row.items() if k != "_url_check"}
                     for row in validated]

    # ── Cleanup orphan rejected rows (audit 2026-05-05) ─────────────────
    # Physically delete Pending rows that are already rejected and either:
    #   (a) have RejectedBy with 2+ different reviewers (counter triggered)
    #   (b) are older than 24h (orphaned from before the counter existed)
    # Runs every cycle so the Pending sheet doesn't accumulate stale
    # rejections forever.
    clean_pending, n_cleaned = cleanup_orphan_rejected(clean_pending)
    if n_cleaned > 0:
        log.info("Cleanup: physically deleted %d orphan rejected rows", n_cleaned)
        # Re-index rejected_flags after deletion: any flag pointing at a now-
        # deleted index is stale, but since we iterated by index BEFORE
        # the cleanup, the flags still match positions in `validated`. After
        # cleanup, `clean_pending` has fewer entries — rejected_flags must
        # be remapped or cleared. Simplest safe choice: clear it. The next
        # url-validator run will re-flag any remaining broken URLs.
        rejected_flags = {}

    # ── Promotion gate ──────────────────────────────────────────────────
    # OFF by default — only the once-daily Gemini URL gate (07:00) and
    # Claude check (07:45) workflows are permitted to promote rows to
    # Recalls. The hourly CLI just stamps rejections and exits.
    promote_enabled = os.environ.get("MERGE_MASTER_PROMOTE", "").strip() in (
        "1", "true", "yes")

    archived_rejected: List[Dict[str, Any]] = []
    if promote_enabled:
        log.info("MERGE_MASTER_PROMOTE=1 — promotion ENABLED for this run "
                 "(use only for backfill/manual catch-up)")
        new_approved, remaining, archived_rejected = promote_approved(
            clean_pending, approved, rejected_flags,
        )
        if new_approved:
            log.info("Promoted %d rows Pending → Recalls", len(new_approved))
            final_approved = sort_rows(approved + new_approved)
        else:
            log.info("No rows promoted this run.")
            final_approved = sort_rows(approved)
    else:
        log.info("Janitor mode (default): NOT promoting. Only the daily "
                 "Gemini URL gate + Claude check are allowed to advance "
                 "rows from Pending → Recalls.")
        # We still update Pending with rejection stamps for broken URLs.
        # Walk clean_pending; for any idx in rejected_flags, copy the
        # rejection note onto the row so the next gate pass skips it.
        # Audit 2026-05-08: when 2nd reviewer rejection lands here in
        # janitor mode, archive the row to Rejected sheet AND drop it
        # from remaining (was previously left orphaned for cleanup_orphan
        # to find on next run — now handled atomically).
        archive_ts = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
        survivors: List[Dict[str, Any]] = []
        for idx, row in enumerate(clean_pending):
            if idx in rejected_flags and row.get("Status") != STATUS_REJECTED:
                reason = rejected_flags[idx]
                should_delete, _ = mark_rejected_with_counter(row, reason)
                if should_delete:
                    row["RejectedAt"] = archive_ts
                    archived_rejected.append(row)
                    continue
            survivors.append(row)
        new_approved = []
        remaining = survivors
        final_approved = sort_rows(approved)
        if rejected_flags:
            log.info("Marked %d Pending row(s) as rejected; archived %d to "
                     "Rejected sheet (2nd-reviewer)",
                     len(rejected_flags), len(archived_rejected))

    save_xlsx_with_pending(final_approved, sort_rows(remaining), XLSX,
                           newly_rejected_rows=archived_rejected)
    mirror_json_from_xlsx(XLSX, ROOT / "docs" / "data" / "recalls.json")

    # Mirror promotions into the Weekly_Review sheet + refresh the JSON
    # slice consumed by the Apps Script Thursday-17:00 mailer. Captures
    # the (rare) backfill / MERGE_MASTER_PROMOTE=1 path too.
    if new_approved:
        try:
            from pipeline.weekly_review_capture import record_promotions  # noqa: E402
            n_wr = record_promotions(new_approved, xlsx_path=XLSX)
            if n_wr:
                log.info("Weekly_Review: appended %d row(s)", n_wr)
        except Exception as _wr_err:
            log.warning("Weekly_Review capture failed: %s", _wr_err)

    # Rebuild daily briefs for any date that gained newly-promoted rows.
    # Without this, the dashboard's rolling 7-day display stays stale.
    # In janitor mode (no promotion), new_approved is empty — this is a
    # cheap no-op.
    rebuilt_briefs, brief_files = rebuild_daily_briefs_for_promoted(
        new_approved, final_approved,
    )
    if rebuilt_briefs:
        log.info("Rebuilt %d daily brief(s)", len(rebuilt_briefs))

    log.info("Done. Recalls=%d, Pending=%d", len(final_approved), len(remaining))
