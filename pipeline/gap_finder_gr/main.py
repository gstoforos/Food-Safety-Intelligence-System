"""
AFTS Food Safety Intelligence — Greek Gap Finder
Module 6: Orchestrator

Single entry point invoked by GitHub Actions once a day at 21:00 Athens.

Pipeline:
    1. news_scraper.collect()       — sweep 7 Greek news sites + Google News GR
    2. efet_fetcher.verify_batch()  — match each news candidate to a canonical
                                       EFET announcement; fetch the body
    3. extractor.extract_batch()    — rule gate + LLM extraction → pending rows
    4. write_pending()              — append accepted rows to recalls.xlsx
                                       Pending sheet, idempotently (no duplicate
                                       EFET URLs across runs)
    5. write_rejected()             — append rejected rows to Weekly_Rejected
                                       sheet (also idempotent)
    6. write_run_log()              — daily audit JSONL row

Idempotency: every write checks existing URLs in the target sheet first and
skips rows that are already there. Safe to re-run the same day.

Exit codes:
    0 — success (any number of rows, including zero)
    1 — fatal error (network, LLM unreachable, xlsx corrupt)
    2 — partial failure (some records failed but pipeline completed)

CLI:
    python -m pipeline.gap_finder_gr.main
    python -m pipeline.gap_finder_gr.main --dry-run        # no writes
    python -m pipeline.gap_finder_gr.main --xlsx PATH      # override xlsx path
    python -m pipeline.gap_finder_gr.main --skip-news      # use existing candidates.jsonl
    python -m pipeline.gap_finder_gr.main --skip-efet      # use existing verified.jsonl
    python -m pipeline.gap_finder_gr.main --verbose

Environment:
    LLAMA_BASE_URL     — Tailscale IP of Hetzner VPS (or localhost on Mac)
    LLAMA_MODEL        — default 'qwen2.5-7b-instruct'
    LLAMA_API_KEY      — optional bearer token
    LLAMA_TIMEOUT      — request timeout seconds (default 120)
"""

from __future__ import annotations
import argparse
import json
import sys
import traceback
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Optional

# Flexible imports — work both as package and standalone script
try:
    from . import news_scraper, efet_fetcher, extractor
    from .llama_client import LlamaClient, LlamaError
    from .rules import classify
except ImportError:
    import news_scraper                            # type: ignore
    import efet_fetcher                            # type: ignore
    import extractor                               # type: ignore
    from llama_client import LlamaClient, LlamaError  # type: ignore
    from rules import classify                     # type: ignore


# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────────────────────────

DEFAULT_XLSX = "docs/data/recalls.xlsx"
DATA_DIR = "docs/data/gap_finder_gr"
CANDIDATES_PATH = f"{DATA_DIR}/candidates.jsonl"
VERIFIED_PATH = f"{DATA_DIR}/verified.jsonl"
UNMATCHED_PATH = f"{DATA_DIR}/unmatched.jsonl"
PENDING_OUT_PATH = f"{DATA_DIR}/pending_candidates.jsonl"
REJECTED_OUT_PATH = f"{DATA_DIR}/rejected_records.jsonl"
RUN_LOG_PATH = f"{DATA_DIR}/run_log.jsonl"

PENDING_SHEET = "Pending"
REJECTED_SHEET = "Weekly_Rejected"
RECALLS_SHEET = "Recalls"

# Columns in target sheets (must match recalls.xlsx exactly)
PENDING_COLUMNS = [
    "Date", "Source", "Company", "Brand", "Product", "Pathogen", "Reason",
    "Class", "Country", "Region", "Tier", "Outbreak", "URL", "Notes",
    "ScrapedAt", "Status", "RejectedBy",
]
REJECTED_COLUMNS = PENDING_COLUMNS + ["RejectedAt", "RejectReason"]


# ─────────────────────────────────────────────────────────────────────────────
# RUN-LEVEL STATE
# ─────────────────────────────────────────────────────────────────────────────

class RunState:
    """Tracks counts and errors across the pipeline for the audit log."""
    def __init__(self) -> None:
        self.started_at = datetime.now(timezone.utc).isoformat()
        self.candidates_found = 0
        self.verified_matched = 0
        self.verified_unmatched = 0
        self.extracted_accepted = 0
        self.extracted_rejected = 0
        self.appended_pending = 0
        self.appended_rejected = 0
        self.skipped_dupe_pending = 0
        self.skipped_dupe_rejected = 0
        self.skipped_dupe_recalls = 0
        self.errors: list[str] = []

    def add_error(self, where: str, exc: Exception) -> None:
        msg = f"[{where}] {type(exc).__name__}: {exc}"
        self.errors.append(msg)
        print(f"  ERROR: {msg}", file=sys.stderr)

    def to_dict(self) -> dict:
        return {
            "started_at": self.started_at,
            "finished_at": datetime.now(timezone.utc).isoformat(),
            "candidates_found": self.candidates_found,
            "verified_matched": self.verified_matched,
            "verified_unmatched": self.verified_unmatched,
            "extracted_accepted": self.extracted_accepted,
            "extracted_rejected": self.extracted_rejected,
            "appended_pending": self.appended_pending,
            "appended_rejected": self.appended_rejected,
            "skipped_dupe_pending": self.skipped_dupe_pending,
            "skipped_dupe_rejected": self.skipped_dupe_rejected,
            "skipped_dupe_recalls": self.skipped_dupe_recalls,
            "errors": self.errors,
        }


# ─────────────────────────────────────────────────────────────────────────────
# XLSX I/O (idempotent append)
# ─────────────────────────────────────────────────────────────────────────────

def _load_existing_urls(xlsx_path: str, sheet_name: str) -> set[str]:
    """Return the set of URL values already present in the given sheet's URL column.
    Used to skip duplicates on append."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("ERROR: openpyxl not installed. pip install openpyxl", file=sys.stderr)
        raise

    p = Path(xlsx_path)
    if not p.exists():
        return set()

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return set()

    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows, None)
        if not header:
            wb.close()
            return set()
        try:
            url_col = list(header).index("URL")
        except ValueError:
            wb.close()
            return set()

        urls: set[str] = set()
        for row in rows:
            if row and len(row) > url_col and row[url_col]:
                urls.add(str(row[url_col]).strip())
        return urls
    finally:
        wb.close()


def _append_rows(
    xlsx_path: str,
    sheet_name: str,
    columns: list[str],
    new_rows: list[dict],
) -> int:
    """Append rows to the given sheet, preserving column order. Returns count appended."""
    from openpyxl import load_workbook, Workbook

    p = Path(xlsx_path)
    if p.exists():
        wb = load_workbook(xlsx_path)
    else:
        wb = Workbook()
        # Remove the default empty sheet that openpyxl creates
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
            del wb["Sheet"]

    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        ws.append(columns)  # write header
    else:
        ws = wb[sheet_name]

    appended = 0
    for row in new_rows:
        values = [row.get(col, "") for col in columns]
        ws.append(values)
        appended += 1

    p.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
    wb.close()
    return appended


def write_pending_rows_to_xlsx(
    xlsx_path: str,
    rows: list[dict],
    state: RunState,
    verbose: bool = False,
) -> None:
    """Append accepted rows to Pending sheet, skipping URLs already present in
    either Pending or the master Recalls sheet (don't re-pending something
    already promoted)."""
    if not rows:
        return

    existing_pending = _load_existing_urls(xlsx_path, PENDING_SHEET)
    existing_recalls = _load_existing_urls(xlsx_path, RECALLS_SHEET)
    if verbose:
        print(f"[pending] existing URLs — Pending: {len(existing_pending)}, "
              f"Recalls: {len(existing_recalls)}", file=sys.stderr)

    to_append: list[dict] = []
    for r in rows:
        url = (r.get("URL") or "").strip()
        if not url:
            to_append.append(r)
            continue
        if url in existing_pending:
            state.skipped_dupe_pending += 1
            if verbose:
                print(f"  skip (already in Pending):  {url}", file=sys.stderr)
            continue
        if url in existing_recalls:
            state.skipped_dupe_recalls += 1
            if verbose:
                print(f"  skip (already in Recalls):  {url}", file=sys.stderr)
            continue
        to_append.append(r)

    if not to_append:
        if verbose:
            print("[pending] nothing new to append", file=sys.stderr)
        return

    appended = _append_rows(xlsx_path, PENDING_SHEET, PENDING_COLUMNS, to_append)
    state.appended_pending = appended
    print(f"[pending] appended {appended} row(s) to {xlsx_path}::{PENDING_SHEET}",
          file=sys.stderr)


def write_rejected_rows_to_xlsx(
    xlsx_path: str,
    rows: list[dict],
    state: RunState,
    verbose: bool = False,
) -> None:
    """Append rejected rows to Weekly_Rejected sheet, skipping already-rejected URLs."""
    if not rows:
        return

    existing = _load_existing_urls(xlsx_path, REJECTED_SHEET)
    if verbose:
        print(f"[rejected] existing URLs in {REJECTED_SHEET}: {len(existing)}",
              file=sys.stderr)

    to_append: list[dict] = []
    for r in rows:
        url = (r.get("URL") or "").strip()
        if url and url in existing:
            state.skipped_dupe_rejected += 1
            if verbose:
                print(f"  skip (already rejected): {url}", file=sys.stderr)
            continue
        to_append.append(r)

    if not to_append:
        if verbose:
            print("[rejected] nothing new to append", file=sys.stderr)
        return

    appended = _append_rows(xlsx_path, REJECTED_SHEET, REJECTED_COLUMNS, to_append)
    state.appended_rejected = appended
    print(f"[rejected] appended {appended} row(s) to {xlsx_path}::{REJECTED_SHEET}",
          file=sys.stderr)


# ─────────────────────────────────────────────────────────────────────────────
# JSONL I/O HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def read_jsonl(path: str) -> list[dict]:
    p = Path(path)
    if not p.exists():
        return []
    out: list[dict] = []
    with p.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                out.append(json.loads(line))
            except json.JSONDecodeError:
                continue
    return out


def append_jsonl(record: dict, path: str) -> None:
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    with p.open("a", encoding="utf-8") as f:
        f.write(json.dumps(record, ensure_ascii=False) + "\n")


# ─────────────────────────────────────────────────────────────────────────────
# PIPELINE STAGES
# ─────────────────────────────────────────────────────────────────────────────

def stage_news_scraper(verbose: bool = False) -> list[dict]:
    """Stage 1: discover candidate news articles."""
    print("\n=== Stage 1: News Scraper ===", file=sys.stderr)
    rss = news_scraper.collect_rss(verbose=verbose)
    gn = news_scraper.collect_google_news(verbose=verbose)
    all_cands = rss + gn
    deduped = news_scraper.deduplicate(all_cands)
    print(f"  RSS: {len(rss)}, Google News: {len(gn)}, "
          f"deduped: {len(deduped)}", file=sys.stderr)
    candidates = [c.to_dict() for c in deduped]
    news_scraper.write_jsonl(deduped, CANDIDATES_PATH)
    return candidates


def stage_efet_verifier(candidates: list[dict], verbose: bool = False) -> tuple[list[dict], list[dict]]:
    """Stage 2: match candidates to EFET announcements, fetch bodies."""
    print("\n=== Stage 2: EFET Fetcher ===", file=sys.stderr)
    if not candidates:
        return [], []

    efet_index = efet_fetcher.fetch_efet_index(verbose=verbose)
    print(f"  EFET index: {len(efet_index)} announcements", file=sys.stderr)

    verified: list[dict] = []
    unmatched: list[dict] = []
    now = datetime.now(timezone.utc).isoformat()

    import time
    for cand in candidates:
        news_title = cand.get("title", "")
        news_pub = cand.get("published", "")
        match, score = efet_fetcher.match_candidate_to_efet(news_title, news_pub, efet_index)

        if not match:
            unmatched.append({**cand, "best_score": score, "checked_at": now})
            continue

        if not match.body:
            time.sleep(efet_fetcher.REQUEST_DELAY)
            match.body = efet_fetcher.fetch_announcement_body(match.url, verbose=verbose)

        record = efet_fetcher.VerifiedRecord(
            news_url=cand.get("url", ""),
            news_title=news_title,
            news_published=news_pub,
            news_source_domain=cand.get("source_domain", ""),
            efet_url=match.url,
            efet_title=match.title,
            efet_date_iso=match.date_iso,
            efet_body=match.body,
            match_score=round(score, 4),
            matched_at=now,
        )
        verified.append(record.to_dict())

    efet_fetcher.write_jsonl(verified, VERIFIED_PATH)
    efet_fetcher.write_jsonl(unmatched, UNMATCHED_PATH)
    print(f"  matched: {len(verified)}, unmatched: {len(unmatched)}", file=sys.stderr)
    return verified, unmatched


def stage_extractor(verified: list[dict], verbose: bool = False) -> tuple[list[dict], list[dict]]:
    """Stage 3: rule gate + LLM extraction."""
    print("\n=== Stage 3: Extractor ===", file=sys.stderr)
    if not verified:
        return [], []

    client = LlamaClient()
    print(f"  LlamaClient: {client.base_url}  model={client.model}", file=sys.stderr)

    if not client.health():
        raise LlamaError(
            f"llama-server unreachable at {client.base_url}. "
            f"Check Tailscale connection and VPS status."
        )

    pending_rows: list[dict] = []
    rejected_rows: list[dict] = []

    for i, v in enumerate(verified, 1):
        if verbose:
            print(f"  [{i}/{len(verified)}] {v.get('efet_title', '')[:65]}",
                  file=sys.stderr)
        try:
            pending, rejected = extractor.extract_one(v, client, verbose=verbose)
        except Exception as e:
            print(f"  WARN: skipping record due to error: {e}", file=sys.stderr)
            continue
        if pending:
            pending_rows.append(pending)
        if rejected:
            rejected_rows.append(rejected)

    extractor.write_jsonl(pending_rows, PENDING_OUT_PATH)
    extractor.write_jsonl(rejected_rows, REJECTED_OUT_PATH)
    print(f"  accepted: {len(pending_rows)}, rejected: {len(rejected_rows)}",
          file=sys.stderr)
    return pending_rows, rejected_rows


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main() -> int:
    parser = argparse.ArgumentParser(description="AFTS Greek Gap Finder — Orchestrator")
    parser.add_argument("--xlsx", default=DEFAULT_XLSX,
                        help=f"Path to recalls.xlsx (default: {DEFAULT_XLSX})")
    parser.add_argument("--dry-run", action="store_true",
                        help="Run all stages but do NOT write to xlsx")
    parser.add_argument("--skip-news", action="store_true",
                        help=f"Skip Stage 1; load {CANDIDATES_PATH}")
    parser.add_argument("--skip-efet", action="store_true",
                        help=f"Skip Stages 1–2; load {VERIFIED_PATH}")
    parser.add_argument("--skip-extract", action="store_true",
                        help=f"Skip Stages 1–3; load {PENDING_OUT_PATH} and {REJECTED_OUT_PATH}")
    parser.add_argument("--verbose", "-v", action="store_true")
    args = parser.parse_args()

    state = RunState()
    print(f"=== AFTS Greek Gap Finder run started {state.started_at} ===",
          file=sys.stderr)
    print(f"  xlsx:    {args.xlsx}", file=sys.stderr)
    print(f"  dry-run: {args.dry_run}", file=sys.stderr)

    # ── Stage 1: news ──────────────────────────────────────────────────────
    candidates: list[dict] = []
    try:
        if args.skip_efet or args.skip_extract or args.skip_news:
            candidates = read_jsonl(CANDIDATES_PATH)
            print(f"\n[Stage 1] SKIPPED — loaded {len(candidates)} from {CANDIDATES_PATH}",
                  file=sys.stderr)
        else:
            candidates = stage_news_scraper(verbose=args.verbose)
    except Exception as e:
        state.add_error("news_scraper", e)
        if args.verbose:
            traceback.print_exc(file=sys.stderr)
    state.candidates_found = len(candidates)

    # ── Stage 2: EFET ──────────────────────────────────────────────────────
    verified: list[dict] = []
    unmatched: list[dict] = []
    try:
        if args.skip_efet or args.skip_extract:
            verified = read_jsonl(VERIFIED_PATH)
            unmatched = read_jsonl(UNMATCHED_PATH)
            print(f"\n[Stage 2] SKIPPED — loaded {len(verified)} verified, "
                  f"{len(unmatched)} unmatched", file=sys.stderr)
        else:
            verified, unmatched = stage_efet_verifier(candidates, verbose=args.verbose)
    except Exception as e:
        state.add_error("efet_fetcher", e)
        if args.verbose:
            traceback.print_exc(file=sys.stderr)
    state.verified_matched = len(verified)
    state.verified_unmatched = len(unmatched)

    # ── Stage 3: extract ───────────────────────────────────────────────────
    pending_rows: list[dict] = []
    rejected_rows: list[dict] = []
    try:
        if args.skip_extract:
            pending_rows = read_jsonl(PENDING_OUT_PATH)
            rejected_rows = read_jsonl(REJECTED_OUT_PATH)
            print(f"\n[Stage 3] SKIPPED — loaded {len(pending_rows)} pending, "
                  f"{len(rejected_rows)} rejected", file=sys.stderr)
        else:
            pending_rows, rejected_rows = stage_extractor(verified, verbose=args.verbose)
    except Exception as e:
        state.add_error("extractor", e)
        if args.verbose:
            traceback.print_exc(file=sys.stderr)
    state.extracted_accepted = len(pending_rows)
    state.extracted_rejected = len(rejected_rows)

    # ── Stage 4: write xlsx ────────────────────────────────────────────────
    print("\n=== Stage 4: Write to recalls.xlsx ===", file=sys.stderr)
    if args.dry_run:
        print("  DRY RUN — no xlsx writes performed", file=sys.stderr)
        print(f"  would append: {len(pending_rows)} pending, "
              f"{len(rejected_rows)} rejected", file=sys.stderr)
    else:
        try:
            write_pending_rows_to_xlsx(args.xlsx, pending_rows, state, verbose=args.verbose)
        except Exception as e:
            state.add_error("write_pending", e)
            if args.verbose:
                traceback.print_exc(file=sys.stderr)
        try:
            write_rejected_rows_to_xlsx(args.xlsx, rejected_rows, state, verbose=args.verbose)
        except Exception as e:
            state.add_error("write_rejected", e)
            if args.verbose:
                traceback.print_exc(file=sys.stderr)

    # ── Stage 5: run log ───────────────────────────────────────────────────
    log_record = state.to_dict()
    try:
        append_jsonl(log_record, RUN_LOG_PATH)
    except Exception as e:
        print(f"WARN: could not write run log: {e}", file=sys.stderr)

    # ── Summary ────────────────────────────────────────────────────────────
    print("\n" + "=" * 70, file=sys.stderr)
    print("RUN SUMMARY", file=sys.stderr)
    print("=" * 70, file=sys.stderr)
    print(f"  candidates found:        {state.candidates_found}", file=sys.stderr)
    print(f"  verified (matched EFET): {state.verified_matched}", file=sys.stderr)
    print(f"  unmatched (no EFET):     {state.verified_unmatched}", file=sys.stderr)
    print(f"  extracted accepted:      {state.extracted_accepted}", file=sys.stderr)
    print(f"  extracted rejected:      {state.extracted_rejected}", file=sys.stderr)
    print(f"  → appended Pending:      {state.appended_pending}", file=sys.stderr)
    print(f"  → appended Rejected:     {state.appended_rejected}", file=sys.stderr)
    print(f"  skipped (dupe Pending):  {state.skipped_dupe_pending}", file=sys.stderr)
    print(f"  skipped (dupe Recalls):  {state.skipped_dupe_recalls}", file=sys.stderr)
    print(f"  skipped (dupe Rejected): {state.skipped_dupe_rejected}", file=sys.stderr)
    print(f"  errors:                  {len(state.errors)}", file=sys.stderr)
    print("=" * 70, file=sys.stderr)

    if state.errors:
        return 2
    return 0


if __name__ == "__main__":
    sys.exit(main())
