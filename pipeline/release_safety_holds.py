#!/usr/bin/env python3
"""
release_safety_holds.py  —  one-time repair (2026-08-11)
=========================================================

WHAT WENT WRONG
---------------
The review agent's safety guard demoted EVERY row sitting at Status="pending"
that it had not explicitly approved — including rows it never reviewed. When
llama was unreachable the agent approved nothing, so on each run the guard
pushed the whole promotable pool from "pending" down to "pending_gap_v2",
where promote_approved will not touch it.

Net effect: 19 rows stranded, 10 of them RappelConso. Published French recalls
stopped at 2026-08-07 even though the scraper kept collecting daily.

The guard itself is now fixed in recall_review_agent.py — it only applies to
rows a run actually reviewed. This script releases the rows already stranded.

WHAT IT DOES
------------
Finds rows whose Notes contain the guard's own stamp ("safety-hold") and whose
Status is "pending_gap_v2", and returns them to Status="pending" so the normal
promotion path can consider them again. It touches nothing else.

THE TRADE-OFF — READ THIS
-------------------------
Returning a row to "pending" makes it promotable by merge-master WITHOUT the
Qwen agent having verified it. That is the pre-agent behaviour: official-feed
rows (RappelConso, FSA, FSAI, RASFF) are extracted from the regulator's own
API, so they are not gap-finder guesses — but they have not been field-checked
against the page either.

  --sources official   (default) release only rows from regulator feeds
  --sources all                  release every stranded row
  --dry-run first, always.

If you would rather they stay held until the agent can verify them, do not run
this — fix the VPS instead and the agent will clear them properly.

Usage:
    python release_safety_holds.py --xlsx docs/data/recalls.xlsx --commit false
    python release_safety_holds.py --xlsx docs/data/recalls.xlsx --commit true
"""
from __future__ import annotations
import argparse
from pathlib import Path

OFFICIAL = ("rappelconso", "fsa (uk)", "fsai", "rasff", "fsanz", "aesan",
            "bvl", "blv", "favv", "afsca", "nvwa", "ages", "efet", "cfia",
            "fda", "usda fsis", "mfds", "sfa", "cfs", "anvisa",
            "livsmedelsverket", "salute")


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=Path, default=Path("docs/data/recalls.xlsx"))
    ap.add_argument("--commit", type=str, default="false")
    ap.add_argument("--sources", choices=("official", "all"), default="official")
    args = ap.parse_args()
    commit = args.commit.lower() in ("1", "true", "yes", "on")

    import openpyxl
    wb = openpyxl.load_workbook(args.xlsx)
    if "Pending" not in wb.sheetnames:
        print("No Pending sheet.")
        return 1
    ws = wb["Pending"]
    headers = [c.value for c in ws[1]]
    for need in ("Status", "Notes", "Source"):
        if need not in headers:
            print(f"Pending sheet has no {need} column.")
            return 1
    si = headers.index("Status") + 1
    ni = headers.index("Notes") + 1
    src_i = headers.index("Source") + 1
    pi = headers.index("Product") + 1 if "Product" in headers else src_i

    released, skipped = [], []
    for ridx in range(2, ws.max_row + 1):
        notes = str(ws.cell(ridx, ni).value or "")
        status = str(ws.cell(ridx, si).value or "").strip()
        if "safety-hold" not in notes or status != "pending_gap_v2":
            continue
        source = str(ws.cell(ridx, src_i).value or "").lower()
        product = str(ws.cell(ridx, pi).value or "")[:44]
        if args.sources == "official" and not any(o in source for o in OFFICIAL):
            skipped.append((source[:14], product))
            continue
        released.append((ridx, source[:14], product))

    print(f"Stranded rows to release ({args.sources}): {len(released)}")
    for _r, s, p in released:
        print(f"   {s:14s} {p}")
    if skipped:
        print(f"\nSkipped (not an official feed): {len(skipped)}")
        for s, p in skipped[:6]:
            print(f"   {s:14s} {p}")

    if not commit:
        print("\nDRY RUN — nothing written. Re-run with --commit true.")
        return 0
    if not released:
        print("Nothing to release.")
        return 0

    import datetime as dt
    today = dt.datetime.now(dt.timezone.utc).strftime("%Y-%m-%d")
    for ridx, _s, _p in released:
        ws.cell(ridx, si).value = "pending"
        old = str(ws.cell(ridx, ni).value or "")
        ws.cell(ridx, ni).value = (
            old + f" [release {today}: safety-hold lifted — the guard demoted "
            f"rows it had not reviewed; returned to pending]").strip()[:2000]

    wb.save(args.xlsx)
    print(f"\n✓ Released {len(released)} row(s) back to Status='pending'.")
    print("  merge-master can now consider them on its next run.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
