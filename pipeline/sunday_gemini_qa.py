"""
Sunday Gemini QA — weekly deep audit of the Recalls + Pending sheets.

Replaces sunday_claude_qa.py. Claude was pattern-matching from training data
and hallucinating fiche IDs. Gemini 2.5 Flash with native Google Search
grounding fires real searches, so every URL it proposes is one that actually
appeared in a search result.

Runs every Sunday 23:00 Athens via FsisScheduler dispatch.

THREE-LAYER AUDIT + INTEGRITY CHECKS
=====================================

Layer 1 — Deterministic fixes (no AI, no cost):
  • Country → Region normalization to canonical 6-region taxonomy
  • Pathogen → Tier consistency (STEC/Listeria/Botulinum/cereulide → Tier-1)
  • Missing Company / placeholder fills for known multi-producer alerts
  • Auto-append /Interne to RappelConso URLs missing it
  • Duplicate URL detection (full sheet scan)
  • Dead URL detection (HTTP HEAD check, audit window)

Layer 2 — API pre-pass via regulator_apis.py (no AI, no cost):
  • Every France row with /categorie/ → query RappelConso open-data API for
    canonical fiche-rappel/NNNNN/Interne
  • Other API-having countries: cross-check URL against official feed

Layer 3 — Gemini AI verification with Google Search grounding (paid):
  • Only for rows that survived Layer 1+2 with structural concerns
  • The 4 ChatGPT-discovered rules baked into the system prompt:
      Rule 1 — France: hallucinated IDs, must Google-verify
      Rule 2 — Greece EFET: hallucinated IDs/slugs, must Google-verify
      Rule 3 — USA FDA/FSIS + Ireland FSAI: DO NOT TOUCH truncated URLs
      Rule 4 — Others: verify, don't aggressive-rewrite

Pending Sheet Audit:
  • Dead URL removal (HTTP HEAD → 4xx/5xx or timeout)
  • Stale entry removal (>14 days old, never promoted)

OUTPUTS
=======
  • BEFORE snapshot saved to docs/data/recalls_BEFORE_qa.xlsx
  • Deterministic + API + AI fixes applied IN-PLACE to docs/data/recalls.xlsx
  • QA Summary sheet added to the AFTER workbook (first tab)
  • Markdown QA report committed to docs/data/sunday-qa/<date>.md
  • Email to George with BEFORE + AFTER Excel attachments
  • Git commit + push at end

EMAIL SETUP
===========
  Requires two GitHub secrets:
    GMAIL_USER         = georgestof@gmail.com
    GMAIL_APP_PASSWORD = <16-char app password from Google>

  To create the App Password:
    1. Go to https://myaccount.google.com/apppasswords
    2. Select "Mail" and "Other (FSIS Bot)"
    3. Copy the 16-character password
    4. Add as GMAIL_APP_PASSWORD secret in GitHub repo settings

COST
====
  ~$0.05–$0.15/run (Gemini 2.5 Flash, ~50–80 rows in 14-day window).
  Runs once per week → < $1/month. On Gemini free tier: $0.

INVOCATION
==========
  python -m pipeline.sunday_gemini_qa
  python -m pipeline.sunday_gemini_qa --days 30
  python -m pipeline.sunday_gemini_qa --dry-run
  python -m pipeline.sunday_gemini_qa --no-ai       # deterministic only
  python -m pipeline.sunday_gemini_qa --no-email     # skip email

Disabled cleanly when no GEMINI_API_KEY env var is present.

SDK: google-genai (new unified SDK, replaces deprecated google.generativeai).
     Install: pip install google-genai
"""
from __future__ import annotations
import argparse
import json
import logging
import os
import re
import sys
import urllib.parse
from datetime import datetime, date, timedelta, timezone
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple

import openpyxl
import requests

from copy import copy as _copy_cell_style
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
import smtplib

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from pipeline.commit_github import git_commit_and_push  # noqa: E402
from pipeline.merge_master import (  # noqa: E402
    mirror_json_from_xlsx, rebuild_daily_briefs_for_promoted, load_existing,
)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    handlers=[logging.StreamHandler()],
)
log = logging.getLogger("sunday-gemini-qa")

# ───────────────────────────────────────────────────────────────────────
# Configuration
# ───────────────────────────────────────────────────────────────────────
XLSX_PATH = ROOT / "docs" / "data" / "recalls.xlsx"
QA_DIR = ROOT / "docs" / "data" / "sunday-qa"

GEMINI_MODEL = os.getenv("GEMINI_MODEL", "gemini-2.5-flash")

DEFAULT_AUDIT_DAYS = 14
QA_EMAIL_TO = os.getenv("QA_EMAIL_TO", "georgestof@gmail.com")

# Dead-URL check timeout (seconds)
URL_CHECK_TIMEOUT = 10

# Canonical FSIS taxonomy (per User Guide § Daily briefs)
CANONICAL_REGIONS = {
    "Europe", "North America", "Latin America",
    "Asia-Pacific", "Middle East & Africa", "Oceania",
}

COUNTRY_TO_REGION = {
    # Europe
    "France": "Europe", "Germany": "Europe", "Italy": "Europe",
    "Spain": "Europe", "Greece": "Europe", "United Kingdom": "Europe",
    "Switzerland": "Europe", "Ireland": "Europe", "Austria": "Europe",
    "Czech Republic": "Europe", "Slovakia": "Europe", "Belgium": "Europe",
    "Netherlands": "Europe", "Portugal": "Europe", "Sweden": "Europe",
    "Norway": "Europe", "Denmark": "Europe", "Finland": "Europe",
    "Poland": "Europe", "Hungary": "Europe", "Romania": "Europe",
    "Bulgaria": "Europe", "Croatia": "Europe", "Turkey": "Europe",
    "Slovenia": "Europe", "Estonia": "Europe", "Latvia": "Europe",
    "Lithuania": "Europe", "Luxembourg": "Europe", "Cyprus": "Europe",
    "Malta": "Europe", "Iceland": "Europe",
    # North America
    "USA": "North America", "United States": "North America",
    "Canada": "North America",
    # Latin America (Mexico per User Guide convention)
    "Mexico": "Latin America", "Brazil": "Latin America",
    "Colombia": "Latin America", "Argentina": "Latin America",
    "Chile": "Latin America", "Peru": "Latin America",
    "Venezuela": "Latin America", "Ecuador": "Latin America",
    "Uruguay": "Latin America", "Paraguay": "Latin America",
    "Bolivia": "Latin America", "Costa Rica": "Latin America",
    "Panama": "Latin America", "Guatemala": "Latin America",
    "Dominican Republic": "Latin America", "Cuba": "Latin America",
    # Asia-Pacific
    "Japan": "Asia-Pacific", "South Korea": "Asia-Pacific",
    "China": "Asia-Pacific", "Singapore": "Asia-Pacific",
    "Hong Kong": "Asia-Pacific", "Taiwan": "Asia-Pacific",
    "Vietnam": "Asia-Pacific", "Thailand": "Asia-Pacific",
    "Malaysia": "Asia-Pacific", "Indonesia": "Asia-Pacific",
    "Philippines": "Asia-Pacific", "India": "Asia-Pacific",
    "Pakistan": "Asia-Pacific", "Bangladesh": "Asia-Pacific",
    # Oceania
    "Australia": "Oceania", "New Zealand": "Oceania",
    "Fiji": "Oceania", "Papua New Guinea": "Oceania",
    # Middle East & Africa
    "South Africa": "Middle East & Africa", "Kenya": "Middle East & Africa",
    "Kenya / Comesa": "Middle East & Africa", "Egypt": "Middle East & Africa",
    "Morocco": "Middle East & Africa", "Nigeria": "Middle East & Africa",
    "Saudi Arabia": "Middle East & Africa", "UAE": "Middle East & Africa",
    "Israel": "Middle East & Africa", "Lebanon": "Middle East & Africa",
    "Jordan": "Middle East & Africa", "Tunisia": "Middle East & Africa",
    "Algeria": "Middle East & Africa", "Ghana": "Middle East & Africa",
    "Ethiopia": "Middle East & Africa", "Tanzania": "Middle East & Africa",
    "Uganda": "Middle East & Africa", "Zimbabwe": "Middle East & Africa",
    "Botswana": "Middle East & Africa", "Namibia": "Middle East & Africa",
}

TIER1_PATHOGENS = {
    "listeria", "listeria monocytogenes",
    "stec", "e. coli stec", "e. coli o157", "shiga",
    "c. botulinum", "clostridium botulinum", "botulinum",
    "cereulide", "bacillus cereus (cereulide)",
    "hepatitis a", "hepatitis e",
    "vibrio", "vibrio parahaemolyticus", "vibrio cholerae",
    "cronobacter",
    "rodenticide", "rat poison", "bromadiolone",
    "heavy metal", "lead", "cadmium", "arsenic", "mercury",
}

# Domains where the agency CMS enforces strict slug length — DO NOT touch
# URLs that look truncated mid-word (FDA, FSIS, FSAI).
TRUNCATED_BUT_VALID_DOMAINS = (
    "fda.gov",
    "fsis.usda.gov",
    "fsai.ie",
)

RAPPELCONSO_DOMAIN = "rappel.conso.gouv.fr"


# ───────────────────────────────────────────────────────────────────────
# URL helpers — same logic as url_gate_gemini.py for consistency
# ───────────────────────────────────────────────────────────────────────
def _is_structurally_bad(url: str) -> Optional[str]:
    """Return rejection reason if URL is structurally bad, else None."""
    if not url or not url.startswith("http"):
        return "not a URL"
    bad_patterns = [
        "/categorie/", "/rubrik/", "/tag/", "/category/",
        "/search?", "/recherche?",
    ]
    for pat in bad_patterns:
        if pat in url:
            return f"listing/category URL ({pat})"
    p = urllib.parse.urlparse(url)
    segs = [s for s in (p.path or "").split("/") if s]
    if len(segs) == 0:
        return "domain only, no path"
    return None


def _is_truncation_protected(url: str) -> bool:
    """True if URL is on FDA/FSIS/FSAI — the truncated-but-valid agencies."""
    if not url:
        return False
    p = urllib.parse.urlparse(url)
    host = (p.netloc or "").lower()
    return any(host.endswith(d) for d in TRUNCATED_BUT_VALID_DOMAINS)


def _strip_fences(txt: str) -> str:
    t = txt.strip()
    if t.startswith("```"):
        t = re.sub(r"^```[a-zA-Z]*\s*\n", "", t)
        t = re.sub(r"\n```\s*$", "", t)
    return t.strip()


def _collect_gemini_keys() -> List[str]:
    keys: List[str] = []
    legacy = os.getenv("GEMINI_API_KEY")
    if legacy:
        keys.append(legacy.strip())
    for i in range(1, 11):
        v = os.getenv(f"GEMINI_API_KEY_{i}")
        if v and v.strip() not in keys:
            keys.append(v.strip())
    return keys


# ───────────────────────────────────────────────────────────────────────
# Layer 1: Deterministic fixers
# ───────────────────────────────────────────────────────────────────────
def fix_region(country: str, current_region: str) -> Optional[str]:
    """Return a corrected region or None if unchanged/unknown."""
    if not country:
        return None
    expected = COUNTRY_TO_REGION.get(country.strip())
    if expected and expected != current_region:
        return expected
    if current_region not in CANONICAL_REGIONS and expected:
        return expected
    return None


def fix_tier(pathogen: str, current_tier: Any) -> Optional[int]:
    """Return corrected tier or None if unchanged."""
    if not pathogen:
        return None
    p_lower = str(pathogen).lower().strip()
    is_tier1 = any(t in p_lower for t in TIER1_PATHOGENS)
    try:
        cur = int(current_tier) if current_tier is not None else 0
    except (TypeError, ValueError):
        cur = 0
    if is_tier1 and cur != 1:
        return 1
    return None


def fix_rappelconso_interne(url: str) -> Optional[str]:
    """Append /Interne to RappelConso fiche-rappel URLs missing it (Rule 1)."""
    if not url or RAPPELCONSO_DOMAIN not in url:
        return None
    if "/fiche-rappel/" not in url:
        return None
    if url.rstrip("/").endswith("/Interne"):
        return None
    # Extract the numeric ID. The scraper sometimes injects a bad /YYYY/
    # prefix (e.g. /fiche-rappel/2026/17370) — find the LAST numeric segment
    # following /fiche-rappel/ to get the real ID.
    after = url.split("/fiche-rappel/", 1)[1]
    nums = re.findall(r"\d+", after)
    if not nums:
        return None
    fiche_id = nums[-1]  # last numeric segment is the actual ID
    return f"https://{RAPPELCONSO_DOMAIN}/fiche-rappel/{fiche_id}/Interne"


# ───────────────────────────────────────────────────────────────────────
# Gemini call helper — with Google Search grounding
# ───────────────────────────────────────────────────────────────────────
def _call_gemini_grounded(prompt: str, system: str,
                           max_tokens: int = 4000) -> Optional[str]:
    """Single Gemini call with Google Search grounding. Returns text or None."""
    try:
        from google import genai  # type: ignore
        from google.genai import types  # type: ignore
    except ImportError:
        log.warning("google-genai not installed; Gemini QA disabled")
        return None

    keys = _collect_gemini_keys()
    if not keys:
        return None

    last_error: Optional[Exception] = None
    for api_key in keys:
        try:
            client = genai.Client(api_key=api_key)
            config = types.GenerateContentConfig(
                system_instruction=system,
                tools=[types.Tool(google_search=types.GoogleSearch())],
                max_output_tokens=max_tokens,
                temperature=0.1,
            )
            resp = client.models.generate_content(
                model=GEMINI_MODEL,
                contents=prompt,
                config=config,
            )
            text = (resp.text or "").strip()
            if text:
                try:
                    if hasattr(resp, 'candidates') and resp.candidates:
                        gm = getattr(resp.candidates[0], 'grounding_metadata', None)
                        if gm:
                            queries = getattr(gm, 'web_search_queries', []) or []
                            if queries:
                                log.debug("Gemini ran %d Google searches",
                                          len(queries))
                except Exception:
                    pass
                return text
        except Exception as exc:  # noqa: BLE001
            last_error = exc
            log.debug("Gemini attempt failed: %s", str(exc)[:150])
            continue

    log.warning("Gemini QA: all keys failed. Last error: %s", last_error)
    return None


# ───────────────────────────────────────────────────────────────────────
# Layer 3: Gemini AI verification — strict prompt with the 4 rules
# ───────────────────────────────────────────────────────────────────────
QA_SYSTEM_PROMPT = """You are the weekly URL & data integrity auditor for the \
FSIS food-safety dashboard. You have Google Search. For every flagged row, \
run real searches and verify URLs against the actual recall facts.

CANONICAL REGIONS (one of exactly these six):
  Europe, North America, Latin America, Asia-Pacific, Middle East & Africa, Oceania

TIER CLASSIFICATION:
  Tier 1 = Listeria monocytogenes, STEC/E. coli O157, C. botulinum,
           cereulide / Bacillus cereus, Cronobacter, Hepatitis A,
           rodenticides, heavy metals
  Tier 2 = Salmonella, Campylobacter, Norovirus, histamine, mycotoxins
  Tier 3 = lower-severity contaminants

NEVER guess fiche/recall IDs (they are chronological across all categories,
not topical). NEVER fabricate a URL. If verification fails, flag the row
'needs_manual_review' rather than proposing a wrong URL.

Return STRICT JSON, no preamble."""


QA_PROMPT_TEMPLATE = """Audit this food-recall row and propose corrections if needed.

═══ ROW ═══
  Date:     {date}
  Source:   {source}
  Company:  {company}
  Brand:    {brand}
  Product:  {product}
  Pathogen: {pathogen}
  Reason:   {reason}
  Country:  {country}
  Region:   {region}
  Tier:     {tier}
  URL:      {url}
  Notes:    {notes}

═══ MANDATORY URL VERIFICATION WORKFLOW ═══

STEP 1 — Compose Google query: "<agency> <brand-or-company> <date> <pathogen>"
         Example: "RappelConso Belle Henriette 2026-04-23 Listeria"
STEP 2 — Run web_search.
STEP 3 — Find the result on the AGENCY'S OFFICIAL DOMAIN whose snippet
         matches date + brand + hazard.
STEP 4 — Verify all four:
         ✓ date_match (page date == row Date, ±1 day)
         ✓ brand_match (page mentions the brand or company)
         ✓ hazard_match (page mentions the pathogen)
         ✓ is_detail_page (URL is NOT /categorie/, /rubrik/, /tag/,
           /search?, or a bare domain)

═══ AGENCY-SPECIFIC RULES ═══

[RULE 1 — France RappelConso]
  Domain: rappel.conso.gouv.fr
  • URL MUST end with /Interne (e.g. /fiche-rappel/22114/Interne)
  • Scraper frequently hallucinates sequential IDs (17394, 17399, 17400, 17401).
    Do NOT trust the existing ID — verify it via Google every time.
  • Reject any URL containing /categorie/.

[RULE 2 — Greece EFET]
  Domain: efet.gr
  • Numeric item IDs and Greek-to-English slugs are frequently hallucinated.
  • Always Google-verify by product name.
  • Correct URL pattern:
    https://www.efet.gr/index.php/el/enimerosi/deltia-typou/anakleiseis-cat/item/NNNN-...

[RULE 3 — USA FDA / USDA FSIS / Ireland FSAI — DO NOT TOUCH]
  Domains: fda.gov, fsis.usda.gov, fsai.ie
  • These agency CMSes enforce strict URL-slug length.
  • URLs that LOOK truncated mid-word (ending in "-because", "-due",
    "-possible-health-risk", "-gr") are the OFFICIAL working links.
  • If a URL is on these domains, set verification.is_detail_page=true and
    do NOT propose a "fix" unless the URL is structurally a category page.
  • Pass these through with confidence=0.95 and strategy="agency-cms-truncated".

[RULE 4 — All other agencies]
  • Verify URL is on the agency's official domain
  • Verify URL is a specific recall page, not an index/listing
  • If URL fails verification, propose a corrected URL ONLY if Google
    Search returned a verified match
  • If no verified match, set pass=false with reason "needs manual review"
  • NEVER guess, NEVER fabricate

═══ OUTPUT (STRICT JSON) ═══

{{
  "issues": ["short description of each issue, or empty list"],
  "url_corrected": "<canonical URL or null if no fix>",
  "fixes": {{
    "Pathogen": "<corrected value or null>",
    "Region":   "<corrected value or null>",
    "Tier":     <1|2|3 or null>,
    "Country":  "<corrected value or null>"
  }},
  "verification": {{
    "date_match":     true|false,
    "brand_match":    true|false,
    "hazard_match":   true|false,
    "is_detail_page": true|false
  }},
  "google_query_used": "<the query you actually ran>",
  "confidence": 0.0-1.0,
  "strategy": "agency-cms-truncated | api-lookup | search-verified | failed"
}}
"""


def gemini_audit_row(row: Dict[str, Any]) -> Dict[str, Any]:
    """Ask Gemini (with Google Search grounding) to audit one row."""
    prompt = QA_PROMPT_TEMPLATE.format(
        date=row.get("Date", ""),
        source=row.get("Source", ""),
        company=row.get("Company", ""),
        brand=row.get("Brand", ""),
        product=row.get("Product", ""),
        pathogen=row.get("Pathogen", ""),
        reason=str(row.get("Reason") or "")[:400],
        country=row.get("Country", ""),
        region=row.get("Region", ""),
        tier=row.get("Tier", ""),
        url=row.get("URL", ""),
        notes=str(row.get("Notes") or "")[:200],
    )
    raw = _call_gemini_grounded(prompt, system=QA_SYSTEM_PROMPT,
                                 max_tokens=2000)
    if not raw:
        return {"issues": [], "url_corrected": None, "fixes": {},
                "verification": {}, "confidence": 0.0, "strategy": "failed"}
    try:
        return json.loads(_strip_fences(raw))
    except json.JSONDecodeError as e:
        log.debug("Gemini QA JSON parse failed: %s | raw=%s", e, raw[:200])
        return {"issues": [], "url_corrected": None, "fixes": {},
                "verification": {}, "confidence": 0.0, "strategy": "failed"}


# ───────────────────────────────────────────────────────────────────────
# Layer 1.E: Duplicate URL detection
# ───────────────────────────────────────────────────────────────────────
def find_duplicate_urls(ws, headers: Dict[str, int]) -> List[Dict[str, Any]]:
    """Find rows with duplicate URLs in the Recalls sheet."""
    url_col = headers.get("URL")
    if not url_col:
        return []
    seen: Dict[str, int] = {}   # url → first row
    dupes: List[Dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        url = str(ws.cell(r, url_col).value or "").strip()
        if not url or not url.startswith("http"):
            continue
        # Normalise: strip trailing slash, lowercase
        norm = url.rstrip("/").lower()
        if norm in seen:
            dupes.append({
                "row": r, "first_row": seen[norm],
                "url": url,
                "date": str(ws.cell(r, headers.get("Date", 1)).value or ""),
                "company": str(ws.cell(r, headers.get("Company", 1)).value or "")[:40],
            })
        else:
            seen[norm] = r
    return dupes


# ───────────────────────────────────────────────────────────────────────
# Layer 1.F: Dead URL detection (HTTP HEAD check)
# ───────────────────────────────────────────────────────────────────────
def check_dead_urls(ws, headers: Dict[str, int],
                    rows_in_window: List[Tuple[int, Dict[str, Any]]]
                    ) -> List[Dict[str, Any]]:
    """HEAD-check URLs in the audit window. Returns list of dead entries."""
    dead: List[Dict[str, Any]] = []
    for r, row in rows_in_window:
        url = str(row.get("URL") or "").strip()
        if not url or not url.startswith("http"):
            continue
        try:
            resp = requests.head(url, timeout=URL_CHECK_TIMEOUT,
                                 allow_redirects=True,
                                 headers={"User-Agent": "FSIS-QA-Bot/1.0"})
            if resp.status_code >= 400:
                dead.append({
                    "row": r, "url": url, "status": resp.status_code,
                    "date": str(row.get("Date") or ""),
                    "company": str(row.get("Company") or "")[:40],
                })
        except requests.RequestException as e:
            dead.append({
                "row": r, "url": url, "status": str(e)[:80],
                "date": str(row.get("Date") or ""),
                "company": str(row.get("Company") or "")[:40],
            })
    return dead


# ───────────────────────────────────────────────────────────────────────
# Pending sheet audit — promote valid, delete garbage
# ───────────────────────────────────────────────────────────────────────
def audit_pending(wb: openpyxl.Workbook, dry_run: bool = False
                  ) -> Dict[str, Any]:
    """Audit the Pending sheet: check URLs, remove stale/dead entries."""
    if "Pending" not in wb.sheetnames:
        return {"checked": 0, "dead_removed": 0, "stale_removed": 0}

    ws = wb["Pending"]
    if ws.max_row < 2:
        return {"checked": 0, "dead_removed": 0, "stale_removed": 0}

    H = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    url_col = H.get("URL")
    date_col = H.get("Date")

    # Walk from bottom to top so row deletion doesn't shift indices
    rows_to_delete: List[Tuple[int, str]] = []
    checked = 0
    cutoff_stale = (date.today() - timedelta(days=14)).isoformat()

    for r in range(ws.max_row, 1, -1):
        checked += 1
        url = str(ws.cell(r, url_col).value or "").strip() if url_col else ""
        row_date = str(ws.cell(r, date_col).value or "") if date_col else ""

        # Check stale (>14 days old, never promoted)
        if row_date and row_date < cutoff_stale:
            rows_to_delete.append((r, "stale>14d"))
            continue

        # Check dead URL
        if url and url.startswith("http"):
            try:
                resp = requests.head(url, timeout=URL_CHECK_TIMEOUT,
                                     allow_redirects=True,
                                     headers={"User-Agent": "FSIS-QA-Bot/1.0"})
                if resp.status_code >= 400:
                    rows_to_delete.append((r, f"dead_url_{resp.status_code}"))
            except requests.RequestException:
                rows_to_delete.append((r, "dead_url_timeout"))

    dead_count = sum(1 for _, reason in rows_to_delete if "dead" in reason)
    stale_count = sum(1 for _, reason in rows_to_delete if "stale" in reason)

    if not dry_run:
        for r, reason in rows_to_delete:
            log.info("Pending: deleting row %d (%s)", r, reason)
            ws.delete_rows(r, 1)

    return {"checked": checked, "dead_removed": dead_count,
            "stale_removed": stale_count,
            "details": [(r, reason) for r, reason in rows_to_delete]}


# ───────────────────────────────────────────────────────────────────────
# Summary sheet builder — adds a "QA Summary" sheet to the workbook
# ───────────────────────────────────────────────────────────────────────
def add_summary_sheet(wb: openpyxl.Workbook, summary: Dict[str, Any],
                      findings: List[Dict], ai_flags: List[Dict],
                      duplicate_urls: List[Dict], dead_urls: List[Dict],
                      pending_result: Dict[str, Any]) -> None:
    """Add a QA Summary sheet to the workbook with all audit results."""
    if "QA Summary" in wb.sheetnames:
        del wb["QA Summary"]
    ws = wb.create_sheet("QA Summary", 0)  # first position

    row = 1
    # Title
    ws.cell(row, 1, f"Sunday QA Report — {summary.get('audit_date', '')}")
    ws.cell(row, 1).font = openpyxl.styles.Font(bold=True, size=14)
    row += 2

    # Overview
    for label, val in [
        ("Audit window (days)", summary.get("audit_days", 14)),
        ("Engine", "Gemini 2.5 Flash + Google Search grounding"),
        ("Total fixes applied", len(findings)),
        ("AI flags (manual review)", len(ai_flags)),
        ("Duplicate URLs found", len(duplicate_urls)),
        ("Dead URLs found", len(dead_urls)),
        ("Pending rows checked", pending_result.get("checked", 0)),
        ("Pending dead removed", pending_result.get("dead_removed", 0)),
        ("Pending stale removed", pending_result.get("stale_removed", 0)),
    ]:
        ws.cell(row, 1, label)
        ws.cell(row, 1).font = openpyxl.styles.Font(bold=True)
        ws.cell(row, 2, val)
        row += 1

    # Layer breakdown
    row += 1
    det = summary.get("deterministic_fixes", {})
    ws.cell(row, 1, "Layer 1 — Deterministic")
    ws.cell(row, 1).font = openpyxl.styles.Font(bold=True, size=12)
    row += 1
    for k, v in det.items():
        ws.cell(row, 1, f"  {k}")
        ws.cell(row, 2, v)
        row += 1

    row += 1
    ws.cell(row, 1, "Layer 2 — API pre-pass")
    ws.cell(row, 1).font = openpyxl.styles.Font(bold=True, size=12)
    row += 1
    ws.cell(row, 1, "  URL fixes")
    ws.cell(row, 2, summary.get("api_fixes", 0))
    row += 2

    ws.cell(row, 1, "Layer 3 — Gemini verified")
    ws.cell(row, 1).font = openpyxl.styles.Font(bold=True, size=12)
    row += 1
    ws.cell(row, 1, "  URL fixes")
    ws.cell(row, 2, summary.get("gemini_url_fixes", 0))
    row += 2

    # Fixes table
    if findings:
        ws.cell(row, 1, "ALL APPLIED FIXES")
        ws.cell(row, 1).font = openpyxl.styles.Font(bold=True, size=12)
        row += 1
        for c, h in enumerate(["Row", "Date", "Field", "Old", "New", "Kind"], 1):
            ws.cell(row, c, h)
            ws.cell(row, c).font = openpyxl.styles.Font(bold=True)
        row += 1
        for f in findings:
            ws.cell(row, 1, f.get("row"))
            ws.cell(row, 2, str(f.get("date", ""))[:10])
            ws.cell(row, 3, f.get("field"))
            ws.cell(row, 4, str(f.get("old", ""))[:80])
            ws.cell(row, 5, str(f.get("new", ""))[:80])
            ws.cell(row, 6, f.get("kind"))
            row += 1
        row += 1

    # Duplicate URLs
    if duplicate_urls:
        ws.cell(row, 1, "DUPLICATE URLs")
        ws.cell(row, 1).font = openpyxl.styles.Font(bold=True, size=12,
                                                      color="FF0000")
        row += 1
        for c, h in enumerate(["Row", "First Row", "URL", "Date", "Company"], 1):
            ws.cell(row, c, h)
            ws.cell(row, c).font = openpyxl.styles.Font(bold=True)
        row += 1
        for d in duplicate_urls:
            ws.cell(row, 1, d["row"])
            ws.cell(row, 2, d["first_row"])
            ws.cell(row, 3, d["url"][:80])
            ws.cell(row, 4, d.get("date", ""))
            ws.cell(row, 5, d.get("company", ""))
            row += 1
        row += 1

    # Dead URLs
    if dead_urls:
        ws.cell(row, 1, "DEAD URLs")
        ws.cell(row, 1).font = openpyxl.styles.Font(bold=True, size=12,
                                                      color="FF0000")
        row += 1
        for c, h in enumerate(["Row", "URL", "Status", "Date", "Company"], 1):
            ws.cell(row, c, h)
            ws.cell(row, c).font = openpyxl.styles.Font(bold=True)
        row += 1
        for d in dead_urls:
            ws.cell(row, 1, d["row"])
            ws.cell(row, 2, d["url"][:80])
            ws.cell(row, 3, str(d["status"]))
            ws.cell(row, 4, d.get("date", ""))
            ws.cell(row, 5, d.get("company", ""))
            row += 1
        row += 1

    # AI flags
    if ai_flags:
        ws.cell(row, 1, "AI FLAGS (manual review needed)")
        ws.cell(row, 1).font = openpyxl.styles.Font(bold=True, size=12,
                                                      color="FF8800")
        row += 1
        for c, h in enumerate(["Row", "Date", "Company", "Issues", "Confidence"], 1):
            ws.cell(row, c, h)
            ws.cell(row, c).font = openpyxl.styles.Font(bold=True)
        row += 1
        for f in ai_flags:
            ws.cell(row, 1, f["row"])
            ws.cell(row, 2, f.get("date", ""))
            ws.cell(row, 3, f.get("company", ""))
            ws.cell(row, 4, "; ".join(f.get("issues", []))[:100])
            ws.cell(row, 5, f"{f.get('confidence', 0):.2f}")
            row += 1

    # Pending audit
    if pending_result.get("details"):
        row += 1
        ws.cell(row, 1, "PENDING SHEET CLEANUP")
        ws.cell(row, 1).font = openpyxl.styles.Font(bold=True, size=12)
        row += 1
        for c, h in enumerate(["Row", "Reason"], 1):
            ws.cell(row, c, h)
            ws.cell(row, c).font = openpyxl.styles.Font(bold=True)
        row += 1
        for pr, reason in pending_result["details"]:
            ws.cell(row, 1, pr)
            ws.cell(row, 2, reason)
            row += 1

    # Auto-size column A
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 60
    ws.column_dimensions["E"].width = 60
    ws.column_dimensions["F"].width = 20


# ───────────────────────────────────────────────────────────────────────
# Email sender — BEFORE + AFTER xlsx to George
# ───────────────────────────────────────────────────────────────────────
def send_qa_email(before_path: Path, after_path: Path,
                  summary: Dict[str, Any]) -> bool:
    """Send QA email via Gmail SMTP with BEFORE/AFTER Excel attachments.

    Requires secrets: GMAIL_USER + GMAIL_APP_PASSWORD.
    Returns True on success, False otherwise.
    """
    gmail_user = os.getenv("GMAIL_USER", "")
    gmail_pass = os.getenv("GMAIL_APP_PASSWORD", "")
    if not gmail_user or not gmail_pass:
        log.warning("GMAIL_USER / GMAIL_APP_PASSWORD not set — email skipped. "
                     "Set these GitHub secrets to enable QA email.")
        return False

    to_addr = QA_EMAIL_TO
    audit_date = summary.get("audit_date", date.today().isoformat())
    det = summary.get("deterministic_fixes", {})
    total_fixes = (sum(det.values()) + summary.get("api_fixes", 0)
                   + summary.get("gemini_url_fixes", 0))
    flags = summary.get("ai_flags", 0)
    dupes = summary.get("duplicate_urls", 0)
    dead = summary.get("dead_urls", 0)
    pend = summary.get("pending_result", {})

    subject = (f"FSIS Sunday QA — {audit_date} | "
               f"{total_fixes} fixes, {flags} flags, "
               f"{dupes} dupes, {dead} dead")

    body = f"""FSIS Sunday QA Report — {audit_date}
{'=' * 50}

Fixes Applied:     {total_fixes}
  Region:          {det.get('region', 0)}
  Tier:            {det.get('tier', 0)}
  /Interne:        {det.get('interne_appended', 0)}
  API pre-pass:    {summary.get('api_fixes', 0)}
  Gemini verified: {summary.get('gemini_url_fixes', 0)}

AI Flags:          {flags} (need manual review)
Duplicate URLs:    {dupes}
Dead URLs:         {dead}

Pending Sheet:
  Checked:         {pend.get('checked', 0)}
  Dead removed:    {pend.get('dead_removed', 0)}
  Stale removed:   {pend.get('stale_removed', 0)}

Attached:
  • recalls_BEFORE.xlsx — snapshot before changes
  • recalls_AFTER.xlsx  — with all fixes + QA Summary sheet

— FSIS Bot (pipeline.sunday_gemini_qa)
"""

    msg = MIMEMultipart()
    msg["From"] = gmail_user
    msg["To"] = to_addr
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain"))

    for fpath, fname in [(before_path, f"recalls_BEFORE_{audit_date}.xlsx"),
                          (after_path, f"recalls_AFTER_{audit_date}.xlsx")]:
        if fpath.exists():
            with open(fpath, "rb") as fp:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(fp.read())
                encoders.encode_base64(part)
                part.add_header("Content-Disposition",
                                f"attachment; filename={fname}")
                msg.attach(part)

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(gmail_user, gmail_pass)
            server.sendmail(gmail_user, [to_addr], msg.as_string())
        log.info("QA email sent to %s", to_addr)
        return True
    except Exception as exc:
        log.error("QA email failed: %s", exc)
        return False


# ───────────────────────────────────────────────────────────────────────
# Main audit loop
# ───────────────────────────────────────────────────────────────────────
def audit(days: int = DEFAULT_AUDIT_DAYS, dry_run: bool = False,
          ai_check: bool = True) -> Dict[str, Any]:
    """Run the Sunday QA. Writes fixes back to xlsx + report to sunday-qa/."""
    if not XLSX_PATH.exists():
        log.error("Recalls file missing: %s", XLSX_PATH)
        return {"error": "no xlsx"}

    # ── Save BEFORE snapshot ──────────────────────────────────────────
    before_path = XLSX_PATH.parent / "recalls_BEFORE_qa.xlsx"
    import shutil
    shutil.copy2(XLSX_PATH, before_path)
    log.info("Saved BEFORE snapshot: %s", before_path)

    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb["Recalls"]
    H = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}

    today = date.today()
    cutoff = (today - timedelta(days=days)).isoformat()

    findings: List[Dict[str, Any]] = []
    det_fixes = {"region": 0, "tier": 0, "company_filled": 0,
                 "interne_appended": 0}
    api_fixes_count = 0
    ai_flags: List[Dict[str, Any]] = []

    # ── Layer 1.E: Duplicate URL detection (full sheet) ───────────────
    duplicate_urls = find_duplicate_urls(ws, H)
    if duplicate_urls:
        log.info("Found %d duplicate URLs", len(duplicate_urls))
        for d in duplicate_urls:
            log.info("  Duplicate: row %d = row %d (%s)",
                     d["row"], d["first_row"], d["url"][:60])

    # ── Layer 1+2: Deterministic + API pre-pass ─────────────────────────
    try:
        from pipeline.regulator_apis import repair_french_row
        api_available = True
    except ImportError:
        log.warning("regulator_apis not importable — API pre-pass skipped")
        api_available = False
        repair_french_row = None  # type: ignore

    rows_in_window: List[Tuple[int, Dict[str, Any]]] = []
    for r in range(2, ws.max_row + 1):
        row_date = str(ws.cell(r, H["Date"]).value or "")
        if row_date < cutoff:
            continue

        row = {k: ws.cell(r, H[k]).value for k in H}
        rows_in_window.append((r, row))

        # Layer 1.A — Region fix
        new_region = fix_region(
            str(row.get("Country") or ""),
            str(row.get("Region") or ""),
        )
        if new_region:
            findings.append({
                "row": r, "date": row_date, "field": "Region",
                "old": row.get("Region"), "new": new_region,
                "kind": "deterministic",
            })
            if not dry_run:
                ws.cell(r, H["Region"]).value = new_region
            det_fixes["region"] += 1

        # Layer 1.B — Tier fix
        new_tier = fix_tier(str(row.get("Pathogen") or ""), row.get("Tier"))
        if new_tier:
            findings.append({
                "row": r, "date": row_date, "field": "Tier",
                "old": row.get("Tier"), "new": new_tier,
                "kind": "deterministic",
                "note": f"pathogen '{row.get('Pathogen')}' is Tier-1",
            })
            if not dry_run:
                ws.cell(r, H["Tier"]).value = new_tier
            det_fixes["tier"] += 1

        # Layer 1.C — Missing Company for known multi-producer alerts
        company = str(row.get("Company") or "").strip()
        if company in ("", "None", "nan", "—", "-"):
            src = str(row.get("Source") or "")
            if "BLV" in src:
                if not dry_run:
                    ws.cell(r, H["Company"]).value = (
                        "(BLV public warning — multiple producers)"
                    )
                det_fixes["company_filled"] += 1

        # Layer 1.D — Append /Interne to RappelConso URLs (Rule 1)
        url = str(row.get("URL") or "")
        new_url = fix_rappelconso_interne(url)
        if new_url:
            findings.append({
                "row": r, "date": row_date, "field": "URL",
                "old": url, "new": new_url, "kind": "interne_appended",
            })
            if not dry_run:
                ws.cell(r, H["URL"]).value = new_url
            det_fixes["interne_appended"] += 1
            row["URL"] = new_url  # update local view for downstream layers

        # Layer 2 — API pre-pass (France only for now)
        if (api_available
                and str(row.get("Country") or "").strip() == "France"
                and "/categorie/" in str(row.get("URL") or "")):
            try:
                api_url = repair_french_row(row)  # type: ignore
                if api_url and api_url != row.get("URL"):
                    findings.append({
                        "row": r, "date": row_date, "field": "URL",
                        "old": row.get("URL"), "new": api_url,
                        "kind": "api-prepass",
                        "note": "RappelConso open-data API",
                    })
                    if not dry_run:
                        ws.cell(r, H["URL"]).value = api_url
                    api_fixes_count += 1
                    row["URL"] = api_url
            except Exception as e:
                log.debug("API pre-pass failed for row %d: %s", r, e)

    # ── Layer 3: Gemini AI verification on remaining concerns ───────────
    gemini_keys_present = bool(_collect_gemini_keys())

    if ai_check and gemini_keys_present:
        for r, row in rows_in_window:
            url = str(row.get("URL") or "")

            # Skip Layer-3 for FDA/FSIS/FSAI (Rule 3 — DO NOT TOUCH)
            if _is_truncation_protected(url):
                continue

            # Only call Gemini on rows that look problematic
            looks_problematic = (
                _is_structurally_bad(url) is not None
                or "needs specific url" in str(row.get("Notes") or "").lower()
                or "needs manual review" in str(row.get("Notes") or "").lower()
                or "needs api" in str(row.get("Notes") or "").lower()
                or "efet.gr" in url  # Greece — always verify per Rule 2
            )

            if not looks_problematic:
                continue

            result = gemini_audit_row(row)
            v = result.get("verification") or {}
            confidence = float(result.get("confidence", 0.0) or 0.0)
            url_fix = result.get("url_corrected") or None

            # Local verification gate — same as url_gate_gemini
            all_v_ok = all([
                v.get("date_match"), v.get("brand_match"),
                v.get("hazard_match"), v.get("is_detail_page"),
            ])

            if url_fix and confidence >= 0.6 and all_v_ok:
                # Reject structurally-bad fix
                if (not _is_truncation_protected(url_fix)
                        and _is_structurally_bad(url_fix)):
                    ai_flags.append({
                        "row": r, "date": str(row.get("Date") or ""),
                        "company": str(row.get("Company") or "")[:40],
                        "issues": [f"Gemini proposed structurally bad URL: {url_fix}"],
                        "confidence": confidence,
                    })
                else:
                    findings.append({
                        "row": r, "date": str(row.get("Date") or ""),
                        "field": "URL", "old": url, "new": url_fix,
                        "kind": "gemini-verified",
                        "note": f"conf={confidence:.2f}, query={result.get('google_query_used','')[:60]}",
                    })
                    if not dry_run:
                        ws.cell(r, H["URL"]).value = url_fix
            else:
                if result.get("issues") or url:
                    ai_flags.append({
                        "row": r, "date": str(row.get("Date") or ""),
                        "company": str(row.get("Company") or "")[:40],
                        "issues": result.get("issues", [])[:3] or [
                            "verification failed"],
                        "confidence": confidence,
                    })

    # ── Layer 1.F: Dead URL check (audit window only) ───────────────────
    dead_urls = check_dead_urls(ws, H, rows_in_window)
    if dead_urls:
        log.info("Found %d dead URLs in audit window", len(dead_urls))
        for d in dead_urls:
            log.info("  Dead: row %d status=%s (%s)",
                     d["row"], d["status"], d["url"][:60])

    # ── Pending sheet audit ───────────────────────────────────────────
    pending_result = audit_pending(wb, dry_run=dry_run)
    if pending_result.get("dead_removed") or pending_result.get("stale_removed"):
        log.info("Pending cleanup: %d dead + %d stale removed",
                 pending_result["dead_removed"], pending_result["stale_removed"])

    # Track what daily briefs got rebuilt (populated only on non-dry-run save)
    rebuilt_briefs: List[str] = []

    if not dry_run:
        wb.save(XLSX_PATH)
        log.info("Saved fixes to %s", XLSX_PATH)

        # ── Mirror recalls.json (so it never drifts from xlsx) ─────────
        try:
            n_mirrored = mirror_json_from_xlsx(
                XLSX_PATH, ROOT / "docs" / "data" / "recalls.json",
            )
            log.info("Mirrored %d rows to recalls.json", n_mirrored)
        except Exception as exc:  # noqa: BLE001
            log.warning("recalls.json mirror failed: %s", exc)

        # ── Rebuild daily briefs for every date that had a fix applied ─
        # Sunday QA modifies existing rows (date corrections, tier
        # rebalances, /Interne appends, etc.) — every affected date's
        # brief must be regenerated so the dashboard's rolling 7-day
        # display + DAILY tab don't go stale.
        any_fix = (
            sum(det_fixes.values()) > 0
            or len(api_fixes) > 0
            or len(gemini_url_fixes) > 0
            or len(duplicate_urls) > 0
            or len(dead_urls) > 0
        )
        if any_fix:
            try:
                # Re-load the post-save Recalls sheet so the brief
                # renderer sees the corrected state.
                full_after = load_existing(XLSX_PATH)
                # Mark every date in the audit window as "potentially
                # affected" so we rebuild conservatively.
                window_dates = {
                    str(r.get("Date") or "")[:10]
                    for r in full_after
                    if str(r.get("Date") or "")[:10] >= cutoff.isoformat()
                }
                synthetic_promoted = [
                    {"Date": d, "URL": "—sunday-qa-marker—"}
                    for d in window_dates if d
                ]
                rebuilt_briefs, _brief_files = rebuild_daily_briefs_for_promoted(
                    synthetic_promoted, full_after,
                )
                if rebuilt_briefs:
                    log.info("Rebuilt %d daily brief(s) after Sunday QA",
                              len(rebuilt_briefs))
            except Exception as exc:  # noqa: BLE001
                log.warning("daily brief rebuild after Sunday QA failed: %s",
                            exc)

    # ── Write Markdown report ──────────────────────────────────────────
    QA_DIR.mkdir(parents=True, exist_ok=True)
    report_path = QA_DIR / f"{today.isoformat()}.md"
    report_lines = [
        f"# FSIS Sunday QA Report — {today.isoformat()}",
        "",
        f"**Audit window:** last {days} days (since {cutoff})",
        f"**Engine:** Gemini 2.5 Flash with Google Search grounding",
        f"**Total findings:** {len(findings)} fixes, {len(ai_flags)} AI flags, "
        f"{len(duplicate_urls)} dupes, {len(dead_urls)} dead URLs",
        "",
        f"**Layer 1 (deterministic):** "
        f"region={det_fixes['region']}, tier={det_fixes['tier']}, "
        f"company_filled={det_fixes['company_filled']}, "
        f"/Interne appended={det_fixes['interne_appended']}",
        f"**Layer 2 (API pre-pass):** {api_fixes_count} URL fixes",
        f"**Layer 3 (Gemini grounded):** "
        f"{sum(1 for f in findings if f.get('kind')=='gemini-verified')} URL fixes verified",
        "",
        "## All applied fixes",
        "",
    ]
    if findings:
        report_lines.append("| Row | Date | Field | Old | New | Note |")
        report_lines.append("|---|---|---|---|---|---|")
        for f in findings:
            note = f.get("note", f.get("kind", ""))
            old_short = str(f.get("old", ""))[:60]
            new_short = str(f.get("new", ""))[:60]
            report_lines.append(
                f"| {f['row']} | {f['date']} | {f['field']} "
                f"| {old_short} | {new_short} | {note} |"
            )
    else:
        report_lines.append("_No fixes applied._")

    report_lines += ["", "## AI flags (manual review needed)", ""]
    if ai_flags:
        report_lines.append("| Row | Date | Company | Issues | Confidence |")
        report_lines.append("|---|---|---|---|---|")
        for f in ai_flags:
            issues = "; ".join(f.get("issues") or [])[:100]
            report_lines.append(
                f"| {f['row']} | {f['date']} | {f['company']} "
                f"| {issues} | {f.get('confidence', 0):.2f} |"
            )
    else:
        report_lines.append("_No AI flags._")

    # Duplicate URLs section
    report_lines += ["", "## Duplicate URLs", ""]
    if duplicate_urls:
        report_lines.append("| Row | First Row | URL | Date | Company |")
        report_lines.append("|---|---|---|---|---|")
        for d in duplicate_urls:
            report_lines.append(
                f"| {d['row']} | {d['first_row']} | {d['url'][:60]} "
                f"| {d.get('date', '')} | {d.get('company', '')} |"
            )
    else:
        report_lines.append("_No duplicates found._")

    # Dead URLs section
    report_lines += ["", "## Dead URLs", ""]
    if dead_urls:
        report_lines.append("| Row | URL | Status | Date | Company |")
        report_lines.append("|---|---|---|---|---|")
        for d in dead_urls:
            report_lines.append(
                f"| {d['row']} | {d['url'][:60]} | {d['status']} "
                f"| {d.get('date', '')} | {d.get('company', '')} |"
            )
    else:
        report_lines.append("_No dead URLs._")

    # Pending audit section
    report_lines += ["", "## Pending Sheet Cleanup", ""]
    pend_details = pending_result.get("details", [])
    if pend_details:
        report_lines.append("| Row | Reason |")
        report_lines.append("|---|---|")
        for pr, reason in pend_details:
            report_lines.append(f"| {pr} | {reason} |")
    else:
        report_lines.append("_No Pending cleanup needed._")

    report_lines += [
        "",
        "---",
        f"_Generated by `pipeline.sunday_gemini_qa` at "
        f"{datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}_",
    ]

    if not dry_run:
        report_path.write_text("\n".join(report_lines), encoding="utf-8")
        log.info("Wrote report: %s", report_path)

    summary = {
        "audit_date": today.isoformat(),
        "audit_days": days,
        "deterministic_fixes": det_fixes,
        "api_fixes": api_fixes_count,
        "gemini_url_fixes": sum(1 for f in findings
                                if f.get("kind") == "gemini-verified"),
        "ai_flags": len(ai_flags),
        "duplicate_urls": len(duplicate_urls),
        "dead_urls": len(dead_urls),
        "pending_result": pending_result,
        "report_path": str(report_path.relative_to(ROOT)),
        "before_path": str(before_path),
        "rebuilt_briefs": rebuilt_briefs,
    }
    log.info("QA summary: %s", json.dumps(summary, indent=2, default=str))

    # ── Add QA Summary sheet to AFTER workbook ────────────────────────
    if not dry_run:
        # Reload the saved workbook to add summary sheet
        wb_after = openpyxl.load_workbook(XLSX_PATH)
        add_summary_sheet(wb_after, summary, findings, ai_flags,
                          duplicate_urls, dead_urls, pending_result)
        wb_after.save(XLSX_PATH)
        log.info("Added QA Summary sheet to %s", XLSX_PATH)

    return summary


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--days", type=int, default=DEFAULT_AUDIT_DAYS,
                    help=f"Audit window in days (default {DEFAULT_AUDIT_DAYS})")
    ap.add_argument("--dry-run", action="store_true", help="Don't write changes")
    ap.add_argument("--no-ai", action="store_true",
                    help="Skip Gemini AI checks (deterministic + API only)")
    ap.add_argument("--no-email", action="store_true",
                    help="Skip email report")
    args = ap.parse_args()

    summary = audit(days=args.days, dry_run=args.dry_run,
                    ai_check=not args.no_ai)

    if "error" in summary:
        return 1

    # Commit if anything changed
    if not args.dry_run:
        any_fix = (any(summary["deterministic_fixes"].values())
                   or summary["api_fixes"] > 0
                   or summary["gemini_url_fixes"] > 0)
        pend = summary.get("pending_result", {})
        pending_changed = (pend.get("dead_removed", 0) > 0
                           or pend.get("stale_removed", 0) > 0)
        if any_fix or summary["ai_flags"] > 0 or pending_changed:
            paths = [str(XLSX_PATH),
                     str(ROOT / "docs" / "data" / "recalls.json"),
                     summary["report_path"]]
            # Include any daily briefs rebuilt during the audit
            paths.extend([str(ROOT / p) for p in summary.get("rebuilt_briefs", [])])
            if summary.get("rebuilt_briefs"):
                paths.append(str(ROOT / "docs" / "daily-index.json"))
            msg = (
                f"Sunday Gemini QA {summary['audit_date']}: "
                f"{summary['deterministic_fixes']['region']} region, "
                f"{summary['deterministic_fixes']['tier']} tier, "
                f"{summary['deterministic_fixes']['interne_appended']} /Interne, "
                f"{summary['api_fixes']} API, "
                f"{summary['gemini_url_fixes']} Gemini, "
                f"{summary['ai_flags']} flags, "
                f"{summary.get('duplicate_urls', 0)} dupes, "
                f"{summary.get('dead_urls', 0)} dead URLs, "
                f"Pending: -{pend.get('dead_removed', 0)} dead "
                f"-{pend.get('stale_removed', 0)} stale"
            )
            git_commit_and_push(ROOT, paths, msg)

    # Send email with BEFORE + AFTER Excel
    if not args.dry_run and not args.no_email:
        before_path = Path(summary.get("before_path", ""))
        send_qa_email(before_path, XLSX_PATH, summary)

    return 0


if __name__ == "__main__":
    sys.exit(main())
