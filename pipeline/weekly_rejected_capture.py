"""
pipeline/weekly_rejected_capture.py

Mirrors every claude-check / openrouter-check rejection into a new
"Weekly_Rejected" sheet of docs/data/recalls.xlsx, tagged with the
Thursday end-date of the review window in which George will see it
(Thu 17:00 Athens cutoff).

Architectural twin of pipeline/weekly_review_capture.py: same Thursday
cutoff math, same JSON-slice export pattern, same idempotent dedup.
The two modules diverge only in:
  • SHEET_NAME ("Weekly_Rejected" vs "Weekly_Review")
  • EXTRA_COLS schema (rejection adds RejectedBy + RejectionReason)
  • JSON path (weekly-rejected-latest.json vs weekly-review-latest.json)
  • Public function name (record_rejections vs record_promotions)

WHY THIS EXISTS (audit 2026-05-09)
==================================
Pre-this-module, claude-check rejections sat in Pending with
Status=rejected, mixed with active-pending rows, until a SECOND
different reviewer also rejected (the 2-reviewer-threshold gate).
The threshold meant rejections cluttered Pending for days and the
operator had no Thursday-window view of "what got rejected this week,
do I want to dispute any of these?"

This module captures rejections at the moment they happen, into a
Thursday-rolling sheet that gets emailed Thursday 17:00 Athens and
wiped Thursday 17:30 Athens (alongside Weekly_Review). The operator's
review window now sees BOTH passes and rejects — full triage.

Sheet schema (Weekly_Rejected):
    Recalls columns + Week_Added (ISO date of the Thursday on which
    the row gets reviewed) + RejectedBy (which reviewer flagged it)
    + RejectionReason (short verdict text from the reviewer's audit
    stamp) + Reviewed (Y/N — manual operator stamp; reserved for
    future filtering).

Cutoff rule (matches the Thursday 17:00 Athens email):
    A row rejected strictly BEFORE Thursday 17:00 Athens belongs to
    that Thursday's review. A row rejected at or after that boundary
    rolls over to the following Thursday. Each rejection lands in
    exactly one Thursday email.

Author: AFTS / G. Stoforos
"""
from __future__ import annotations

import json
import logging
import re
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from zoneinfo import ZoneInfo

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook import Workbook

log = logging.getLogger("pipeline.weekly_rejected_capture")

ROOT = Path(__file__).resolve().parent.parent
XLSX_DEFAULT = ROOT / "docs" / "data" / "recalls.xlsx"
JSON_DEFAULT = ROOT / "docs" / "data" / "weekly-rejected-latest.json"

SHEET_NAME = "Weekly_Rejected"
ATHENS = ZoneInfo("Europe/Athens")
REVIEW_HOUR_LOCAL = 17  # Thursday 17:00 Athens cutoff (matches Weekly_Review)

# Mirror the Recalls schema (must stay in sync with merge_master.RECALLS_SCHEMA).
RECALLS_COLS: List[str] = [
    "Date", "Source", "Company", "Brand", "Product", "Pathogen", "Reason",
    "Class", "Country", "Region", "Tier", "Outbreak", "URL", "Notes",
    "DateAdded", "LastUpdated", "LastChecked",
]
# RejectedBy / RejectionReason added relative to Weekly_Review schema:
# operator needs to see WHO flagged it and WHY without parsing Notes.
EXTRA_COLS: List[str] = ["Week_Added", "RejectedBy", "RejectionReason", "Reviewed"]
SHEET_COLS: List[str] = RECALLS_COLS + EXTRA_COLS


# ---------------------------------------------------------------------------
# Date math (verbatim from weekly_review_capture.py — same Thursday cutoff)
# ---------------------------------------------------------------------------
def review_thursday_for(now_utc: Optional[datetime] = None) -> date:
    """
    The next Thursday review date — i.e. the Thursday email a row
    rejected RIGHT NOW will appear in. Rows rejected at/after Thursday
    17:00 Athens roll over to the following Thursday.
    """
    if now_utc is None:
        now_utc = datetime.now(timezone.utc)
    local = now_utc.astimezone(ATHENS)
    d = local.date()
    days_until_thu = (3 - d.weekday()) % 7  # Mon=0 .. Thu=3 .. Sun=6
    if days_until_thu == 0 and local.hour >= REVIEW_HOUR_LOCAL:
        days_until_thu = 7
    return d + timedelta(days=days_until_thu)


# ---------------------------------------------------------------------------
# Sheet I/O helpers
# ---------------------------------------------------------------------------
def _sheet_headers(ws) -> List[str]:
    """The header row actually on the sheet, in order."""
    if ws.max_row < 1:
        return []
    return [str(c.value or "").strip() for c in ws[1]]


def _ensure_sheet(wb: Workbook):
    """Return the Weekly_Rejected sheet, creating it or WIDENING it so that
    every column this module writes has a header of its own.

    THE BUG THIS REPLACES (found 2026-08-18)
    ========================================
    This function used to `return wb[SHEET_NAME]` untouched when the sheet
    already existed, and record_rejections then wrote rows with
    `ws.append(row_out)` — which lays values down BY POSITION.

    Three modules define a schema for this one sheet and they do not agree:

        pipeline/extractor.py      REJECTED_COLUMNS      19 cols
        pipeline/merge_master.py   REJECTED_SCHEMA       (legacy)
        this module                SHEET_COLS            21 cols

    The live workbook carries the 19-column extractor header:

        ... Notes | ScrapedAt | Status | RejectedBy | RejectedAt | RejectReason

    while row_out is RECALLS_COLS + [week_end, rejected_by, reason, "N"],
    i.e. 21 values whose 15th onward are DateAdded, LastUpdated,
    LastChecked, Week_Added, RejectedBy, RejectionReason, Reviewed. Appended
    positionally against that header, the next archived rejection would have
    landed as:

        ScrapedAt    <- DateAdded
        Status       <- LastUpdated
        RejectedBy   <- LastChecked        (so: always blank or a date)
        RejectedAt   <- Week_Added         (a FUTURE Thursday, not a stamp)
        RejectReason <- rejected_by        (the reviewer's name, not a reason)
        (no header)  <- the actual reason
        (no header)  <- "N"

    Every one of those is silently plausible — a date under a date header,
    a name under a text header — which is why nothing caught it. The 19
    columns already on disk show the same shift, so the reason a row was
    rejected is not readable anywhere in the archive today.

    Writing by NAME instead of position makes the three schemas irrelevant:
    whatever headers the sheet has, each value goes under its own, and any
    header this module needs but the sheet lacks is appended once.
    """
    fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2",
                       fill_type="solid")
    if SHEET_NAME not in wb.sheetnames:
        ws = wb.create_sheet(SHEET_NAME)
        # Header row — distinct light-red fill so the tab visually
        # contrasts with Weekly_Review's blue header.
        for i, h in enumerate(SHEET_COLS, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = Font(bold=True)
            c.fill = fill
        ws.freeze_panes = "A2"
        return ws

    ws = wb[SHEET_NAME]
    headers = _sheet_headers(ws)
    # Historic SPELLINGS of the same column, so a widened sheet does not end
    # up with two reason columns. Note what is NOT aliased: Week_Added is not
    # RejectedAt. The old positional write put the review Thursday into
    # RejectedAt, which is why archived rows carry a rejection timestamp
    # dated in the future. They are different facts — Week_Added is which
    # Thursday review the row belongs to, RejectedAt is when it was rejected
    # — and each gets its own column.
    ALIAS = {"RejectionReason": ("RejectReason", "RejectedReason")}
    have = set(headers)
    for want in SHEET_COLS:
        if want in have:
            continue
        if any(a in have for a in ALIAS.get(want, ())):
            continue
        col = len(headers) + 1
        c = ws.cell(row=1, column=col, value=want)
        c.font = Font(bold=True)
        c.fill = fill
        headers.append(want)
        have.add(want)
        log.info("Weekly_Rejected: added missing column %r at position %d",
                 want, col)
    return ws


def _content_key(get) -> tuple:
    """Identity for a row that has NO URL. `get` is a field accessor.

    AUDIT 2026-08-14 — record_rejections() used to key on URL alone and
    `continue` on an empty URL, so a rejection with no URL was never
    written to Weekly_Rejected AT ALL. Not a dedup collision: an outright
    silent drop, which is exactly what the archive exists to prevent.

    This is not a rare edge case and it is getting commoner. The
    URL-guardian BLANKS the URL of any row whose link 404s or errors
    ("[URL-guardian 2026-08-14: blanked http_error ...]", then "blanked
    empty"). Every row it blanks becomes unarchivable. In the 2026-08-14
    W33 review, 3 of 21 rejections vanished this way — including a real
    FDA Salmonella-in-eggs outbreak row and a Kettle Cuisine Listeria
    row, i.e. precisely the rows an operator would most want to dispute
    on Thursday.

    Same shape as the fix already made in xlsx_merge._wr_dedup_key for
    the identical bug class; that fix was never propagated here.
    """
    def _f(name, n=None):
        v = get(name)
        if v in (None, ""):
            return ""
        v = str(v).strip().lower()
        return v[:n] if n else v
    return ("", _f("Date", 10), _f("Source"), _f("Company", 60),
            _f("Product", 60), _f("Pathogen", 40), _f("Reason", 60))


def _row_key(get) -> tuple:
    """URL+Date when a URL exists, content signature when it does not."""
    url = _s(get("URL")).strip().lower()
    if url:
        return (url, _s(get("Date"))[:10])
    return _content_key(get)


def _s(v) -> str:
    return "" if v in (None, "") else str(v)


def _existing_keys(ws) -> set:
    """Dedup against rows already in Weekly_Rejected.

    URL+Date where a URL exists; a content signature where it does not,
    so URL-less rows stay distinct instead of collapsing (or vanishing).
    See _content_key for why.
    """
    if ws.max_row < 2:
        return set()
    headers = [str(c.value or "") for c in ws[1]]
    if "URL" not in headers or "Date" not in headers:
        return set()
    idx = {h: i for i, h in enumerate(headers) if h}
    keys: set = set()
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v in (None, "") for v in r):
            continue

        def get(name, _r=r):
            i = idx.get(name)
            return _r[i] if i is not None and i < len(_r) else ""

        keys.add(_row_key(get))
    return keys


# ---------------------------------------------------------------------------
# RejectedBy / Reason extraction — operator sees WHO + WHY without
# digging through Notes blobs.
# ---------------------------------------------------------------------------
_REVIEWER_TAG_RE = re.compile(
    r"\[(claude-check|openrouter-check|gemini-check)\b", re.IGNORECASE,
)
_VERDICT_REASON_RE = re.compile(
    r"\[(?:claude-check|openrouter-check|gemini-check)\s+\d{4}-\d{2}-\d{2}:\s*"
    r"(?:fail|reject)[^;]*;\s*([^\]]+)\]",
    re.IGNORECASE,
)
_FALLBACK_REASON_RE = re.compile(
    r"(?:rejected|failed|out_of_scope|garbage)[:\s]+([^\]]+)",
    re.IGNORECASE,
)


def _extract_rejection_metadata(row: Dict[str, Any]) -> Tuple[str, str]:
    """Pull (rejected_by, reason) from a rejected row.

    Priority order:
      1. row["RejectedBy"] field (set by claude_check / openrouter_check
         when stamping the audit tag) → reviewer name
      2. Notes audit stamp regex → fallback parse of "[claude-check
         2026-05-09: fail; <reason>]" or "[openrouter-check ...]"
      3. Empty strings if nothing parseable
    """
    rejected_by = str(row.get("RejectedBy", "") or "").strip()
    notes = str(row.get("Notes", "") or "")

    if not rejected_by:
        m = _REVIEWER_TAG_RE.search(notes)
        if m:
            rejected_by = m.group(1).lower()

    reason = ""
    rm = _VERDICT_REASON_RE.search(notes)
    if rm:
        reason = rm.group(1).strip()
    else:
        rm = _FALLBACK_REASON_RE.search(notes)
        if rm:
            reason = rm.group(1).strip()[:200]

    return rejected_by[:80], reason[:300]


# ---------------------------------------------------------------------------
# Public API — primary entry point used by claude_check / merge_master
# ---------------------------------------------------------------------------
def record_rejections(
    rejected_rows: List[Dict[str, Any]],
    xlsx_path: Path = XLSX_DEFAULT,
    json_path: Path = JSON_DEFAULT,
) -> int:
    """
    Append each just-rejected row to the Weekly_Rejected sheet, tagged
    with the upcoming Thursday review date. Then refresh the JSON slice
    so the Apps Script Thursday mailer always reads current data.

    Idempotent — calling twice with the same rows is a no-op (URL+Date
    dedup against existing Weekly_Rejected entries).

    Designed to be called from claude_check.py / openrouter_check.py
    immediately after a row is finalized as rejected, BEFORE/AT the
    moment it is evicted from Pending by promote_approved. The row dict
    must have its full Recalls-schema fields populated plus Notes
    containing the audit stamp.
    """
    if not xlsx_path.exists():
        return 0
    if not rejected_rows:
        # Always refresh the JSON snapshot anyway so the mailer sees
        # current state on zero-reject runs (idempotent timestamp bump).
        export_week_slice(xlsx_path=xlsx_path, json_path=json_path)
        return 0

    week_end = review_thursday_for().isoformat()

    wb = openpyxl.load_workbook(xlsx_path)
    ws = _ensure_sheet(wb)
    seen = _existing_keys(ws)

    appended = 0
    dropped_no_identity = 0
    for r in rejected_rows or []:
        # AUDIT 2026-08-14: was `if not url or (url, d) in seen: continue`,
        # which threw away every URL-less rejection instead of archiving it.
        key = _row_key(r.get)
        if key in seen:
            continue
        # A row with neither a URL nor any content to key on is the only
        # thing we still refuse — and we COUNT it rather than pass over it
        # in silence, so "nothing was archived" can never look like
        # "nothing was rejected".
        if not any(part for part in key):
            dropped_no_identity += 1
            continue
        seen.add(key)

        rejected_by, reason = _extract_rejection_metadata(r)
        # US spelling in the archive too (audit 2026-08-14). Weekly_Rejected
        # is not written through merge_master's writer, so it does not pick
        # up the normalisation there — and this sheet is what the operator
        # reads in the Thursday review email, so "Mould" would survive in
        # the one place a human actually looks. Pathogen/Reason/Class only;
        # Notes keep the original wording because they are the audit trail.
        try:
            from pipeline._language import americanize as _us
        except Exception:                                      # noqa: BLE001
            def _us(v):                                        # type: ignore
                return v
        payload = {}
        for c in RECALLS_COLS:
            _v = r.get(c, "")
            if c in ("Pathogen", "Reason", "Class") and isinstance(_v, str):
                _v = _us(_v)
            payload[c] = _v
        payload["Week_Added"] = week_end
        payload["RejectedBy"] = rejected_by
        payload["RejectionReason"] = reason
        payload["Reviewed"] = "N"
        # Fill the columns the older extractor schema defines but this
        # module does not, so a widened sheet has no ragged gaps.
        payload.setdefault("ScrapedAt", r.get("ScrapedAt", ""))
        payload.setdefault("Status", r.get("Status", "") or "rejected")
        # The moment of rejection, not the Thursday it will be reviewed on.
        payload.setdefault("RejectedAt",
                           r.get("RejectedAt")
                           or datetime.now(timezone.utc).date().isoformat())
        payload.setdefault("RejectReason", reason)

        # BY NAME, NOT BY POSITION — see _ensure_sheet for what positional
        # appends did to this sheet.
        headers = _sheet_headers(ws)
        missing = [h for h in ("Date", "URL") if h not in headers]
        if missing:
            # Refuse BEFORE writing anything: an archived rejection whose
            # identity columns have nowhere to go could never be matched
            # against a re-ingestion, and a half-written row is worse than
            # a loud failure.
            log.error("Weekly_Rejected: sheet header has no %s column — "
                      "refusing to archive a rejection that could never be "
                      "matched against a re-ingestion",
                      " or ".join(missing))
            dropped_no_identity += 1
            continue
        row_idx = ws.max_row + 1
        for col, h in enumerate(headers, 1):
            if h and h in payload:
                ws.cell(row=row_idx, column=col, value=payload[h])
        appended += 1

    if dropped_no_identity:
        log.error(
            "Weekly_Rejected: %d rejection(s) had NO URL and no content to "
            "key on, so they could not be archived. The operator's Thursday "
            "review will not show them. This is data loss, not filtering.",
            dropped_no_identity,
        )

    if appended:
        # Preserve sheet order: Recalls, Pending, (auxiliary), NEWS last
        # (matches save_xlsx_with_pending logic in merge_master.py and
        # mirrors weekly_review_capture's ordering).
        ordered = []
        for s in ("Recalls", "Pending"):
            if s in wb.sheetnames:
                ordered.append(s)
        for s in wb.sheetnames:
            if s in ordered or s == "NEWS":
                continue
            ordered.append(s)
        if "NEWS" in wb.sheetnames:
            ordered.append("NEWS")
        wb._sheets = [wb[s] for s in ordered]
        wb.save(xlsx_path)

    # Always refresh the JSON slice so the mailer sees the latest state
    # even if 0 rows appended this run (idempotent timestamp bump).
    export_week_slice(xlsx_path=xlsx_path, json_path=json_path,
                      week_end=week_end)
    return appended


def export_week_slice(
    xlsx_path: Path = XLSX_DEFAULT,
    json_path: Path = JSON_DEFAULT,
    week_end: Optional[str] = None,
) -> Dict[str, Any]:
    """
    Write a JSON slice of the upcoming/current Thu→Thu rejection window
    for the Apps Script Thursday-17:00 mailer to fetch.

    week_end: ISO date of the closing Thursday. Default = the Thursday
              the next rejection would land in (i.e. the upcoming review).
    """
    if week_end is None:
        week_end = review_thursday_for().isoformat()

    rows: List[Dict[str, Any]] = []
    if xlsx_path.exists():
        wb = openpyxl.load_workbook(xlsx_path, read_only=True)
        if SHEET_NAME in wb.sheetnames:
            ws = wb[SHEET_NAME]
            headers = [str(c.value or "") for c in ws[1]]
            try:
                we_idx = headers.index("Week_Added")
            except ValueError:
                we_idx = -1
            if we_idx >= 0:
                for r in ws.iter_rows(min_row=2, values_only=True):
                    if str(r[we_idx] or "") != week_end:
                        continue
                    obj = {}
                    for h, v in zip(headers, r):
                        if v is None:
                            obj[h] = ""
                        elif isinstance(v, (datetime, date)):
                            obj[h] = v.isoformat()
                        else:
                            obj[h] = v
                    rows.append(obj)

    # Sort: Tier 1 first (most-urgent disputes the operator may want to
    # contest), then Date desc within tier. Stable sort, two passes.
    rows.sort(key=lambda r: str(r.get("Date", ""))[:10], reverse=True)
    def _tier_key(r):
        try:
            return int(r.get("Tier", 0) or 0) or 9
        except (TypeError, ValueError):
            return 9
    rows.sort(key=_tier_key)

    payload = {
        "week_end": week_end,
        "generated_utc": datetime.now(timezone.utc)
            .isoformat(timespec="seconds"),
        "row_count": len(rows),
        "tier1_count": sum(1 for r in rows if str(r.get("Tier", "")) == "1"),
        "outbreak_count": sum(1 for r in rows if str(r.get("Outbreak", "0"))
                                                 not in ("", "0", "0.0", "False", "false")),
        # Group by reviewer so the operator can spot a single reviewer
        # over-rejecting (early warning for prompt-tuning regressions).
        "by_reviewer": _count_by_reviewer(rows),
        "rows": rows,
    }
    json_path.parent.mkdir(parents=True, exist_ok=True)
    json_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )
    return payload


def _count_by_reviewer(rows: List[Dict[str, Any]]) -> Dict[str, int]:
    """Histogram of rejections by reviewer for operator situational awareness."""
    counts: Dict[str, int] = {}
    for r in rows:
        rb = str(r.get("RejectedBy", "") or "").strip().lower() or "unknown"
        counts[rb] = counts.get(rb, 0) + 1
    return dict(sorted(counts.items(), key=lambda kv: -kv[1]))


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import sys
    cmd = sys.argv[1] if len(sys.argv) > 1 else "export"
    if cmd == "export":
        result = export_week_slice()
        print(f"weekly-rejected-latest.json: {result['row_count']} rows "
              f"(Tier1={result['tier1_count']}, "
              f"Outbreak={result['outbreak_count']}, "
              f"by_reviewer={result['by_reviewer']}) "
              f"for week ending {result['week_end']}")
    elif cmd == "thursday":
        print(review_thursday_for().isoformat())
    else:
        print("Usage: python -m pipeline.weekly_rejected_capture [export|thursday]",
              file=sys.stderr)
        sys.exit(2)
