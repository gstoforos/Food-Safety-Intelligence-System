#!/usr/bin/env python3
"""
recall_review_agent.py
======================

Self-hosted Qwen 2.5 7B review agent. Takes a candidate recall row (from
any gap-finder), visits its regulator/news page, and verifies + fills +
corrects EVERY field against the page text until it is 100% faithful to
the source — then decides APPROVE (→ Recalls) or REJECT (→ Weekly_Rejected).

Runs entirely on the AFTS Qwen VPS via LlamaClient (Tailscale, no API
key). Nothing here depends on Claude/Gemini. Future-proof: when the VPS
moves to a Mac, only LLAMA_BASE_URL changes.

Design — mirrors the existing official_feeds extractor contract:
  - Same PENDING_COLUMNS schema.
  - Same "be faithful, never invent, empty string if not stated" rules.
  - Same LlamaClient.chat() tool-calling loop.
  - Reuses article_fetcher.fetch_html for the TLS-impersonated page fetch
    and searx_search when a field needs corroboration or a better URL.

The agent's job per row:
  1. Fetch the row's URL. If dead/soft-404, use web_search to find the
     correct official page for THIS recall (company + product + hazard).
  2. Read the page text. For each field (Date, Company, Brand, Product,
     Pathogen, Reason, Country, Region) confirm it matches the page, or
     correct it. Fill any blank the page supports. Never invent.
  3. Verify the recall is real, in-scope (2026+, food, Tier-1 hazard
     universe), and not a duplicate.
  4. Return a corrected row + verdict {approve|reject} (retry only on infra
     per-field provenance note.

Verdicts:
  approve       — every required field verified against the page, in scope
  reject        — not a recall / out of scope / pre-2026 / dead URL / dup
  retry         — INFRA failure only (llama down); row left in Pending

CLI:
  python -m pipeline.recall_review_agent --xlsx docs/data/recalls.xlsx \\
      --commit false            # dry run prints verdicts
  --commit true                 # writes approvals→Recalls, rejects→Weekly_Rejected
  --source-filter "EFET"        # review only rows from one source
  --limit N                     # cap rows this run

Env:
  LLAMA_BASE_URL, LLAMA_MODEL     (from llama_client)
  REVIEW_MAX_PAGE_CHARS  default 12000
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import sys
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple

# ── Reuse the existing pipeline infrastructure ───────────────────────────
# These imports match the official_feeds subsystem. If run outside that
# package context, the try/except keeps the module importable for testing.
try:
    from pipeline.official_feeds.agents import llama_client
except Exception:  # pragma: no cover
    llama_client = None
try:
    from pipeline.official_feeds.agents import searx_search
except Exception:  # pragma: no cover
    searx_search = None
try:
    from pipeline.official_feeds import article_fetcher
except Exception:  # pragma: no cover
    article_fetcher = None


PENDING_COLUMNS = [
    "Date", "Source", "Company", "Brand", "Product", "Pathogen", "Reason",
    "Class", "Country", "Region", "Tier", "Outbreak", "URL", "Notes",
    "ScrapedAt", "Status", "RejectedBy",
]

MAX_PAGE_CHARS = int(os.environ.get("REVIEW_MAX_PAGE_CHARS", "5000"))


# ─── Tool-calling schema for the agent ───────────────────────────────────

def _tool_schema() -> List[Dict[str, Any]]:
    return [
        {
            "type": "function",
            "function": {
                "name": "fetch_page",
                "description": "Fetch the readable text of a web page by URL. "
                               "Use this to read the recall page before "
                               "judging any field.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "url": {"type": "string",
                                "description": "Full http(s) URL to fetch."},
                    },
                    "required": ["url"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "web_search",
                "description": "Search the web for the official regulator "
                               "page for a recall. Use ONLY if the row's URL "
                               "is dead or wrong. Query with company + product "
                               "+ hazard.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "query": {"type": "string"},
                    },
                    "required": ["query"],
                },
            },
        },
    ]


def _fetch_page_text(url: str) -> Tuple[str, str]:
    """Fetch a page's readable text. Self-contained (no CountryConfig).

    Tries curl_cffi with Chrome TLS impersonation first (matches the
    pipeline's article_fetcher / _akamai_fetch approach for sites that
    fingerprint Python's TLS), falls back to stdlib requests. Returns
    (text, status).
    """
    html = ""
    status = "ok"
    headers = {
        "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                       "AppleWebKit/537.36 (KHTML, like Gecko) "
                       "Chrome/131.0 Safari/537.36"),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    }
    try:
        from curl_cffi import requests as cffi_requests  # type: ignore
        resp = cffi_requests.get(url, headers=headers, timeout=30,
                                 impersonate="chrome131",
                                 allow_redirects=True)
        html = resp.text or ""
        if resp.status_code >= 400:
            return "", f"http_{resp.status_code}"
    except Exception:
        try:
            import requests as _requests
            resp = _requests.get(url, headers=headers, timeout=30,
                                 allow_redirects=True)
            html = resp.text or ""
            if resp.status_code >= 400:
                return "", f"http_{resp.status_code}"
        except Exception as e:  # noqa: BLE001
            return "", f"error_{type(e).__name__}"

    # Strip HTML to readable text
    try:
        from bs4 import BeautifulSoup
        soup = BeautifulSoup(html, "html.parser")
        for tag in soup(["script", "style", "nav", "header", "footer"]):
            tag.decompose()
        text = soup.get_text(separator="\n")
        text = "\n".join(ln.strip() for ln in text.splitlines() if ln.strip())
    except Exception:
        text = html
    return text[:MAX_PAGE_CHARS], status


def _make_tool_executor(seen_urls: set) -> Callable[[str, dict], str]:
    def execute(name: str, args: dict) -> str:
        if name == "fetch_page":
            url = (args or {}).get("url", "").strip()
            if not url:
                return json.dumps({"error": "no url"})
            text, status = _fetch_page_text(url)
            seen_urls.add(url)
            return json.dumps({"url": url, "status": status, "text": text})
        if name == "web_search":
            q = (args or {}).get("query", "").strip()
            if searx_search is None:
                return json.dumps({"error": "searx unavailable"})
            try:
                results = searx_search.search(q)
                slim = [{"title": r.get("title", ""), "url": r.get("url", ""),
                         "content": (r.get("content", "") or "")[:200]}
                        for r in (results or [])[:6]]
                return json.dumps({"query": q, "results": slim})
            except Exception as e:  # noqa: BLE001
                return json.dumps({"error": f"{type(e).__name__}: {e}"})
        return json.dumps({"error": f"unknown tool {name}"})
    return execute


# ─── Prompts ─────────────────────────────────────────────────────────────

REVIEW_SYSTEM = (
    "You are a senior food-safety analyst performing FINAL verification of "
    "a recall record before it is published on a public dashboard. You have "
    "two tools: fetch_page (read a URL) and web_search (find the official "
    "page if the given URL is dead). You MUST read the actual regulator page "
    "before judging any field. You never invent facts: if the page does not "
    "state something, leave that field an empty string. You answer ONLY with "
    "the final JSON object described by the user — no prose, no markdown."
)

def build_review_prompt(row: Dict[str, Any]) -> str:
    """COMPACT prompt sized for a 4096-token context.

    Rules already enforced deterministically in code are NOT repeated here —
    they are applied after the model answers, so spending context on them is
    waste:
      * pre-2026 publication date        -> hard guard in review_row()
      * language / placeholder / headline / regulator-as-product
                                         -> _field_integrity_flags()
      * USA -> "United States", USDA FSIS source
                                         -> _normalize_country_source()
    What remains here is only what the model alone can judge: reading the
    page, correcting fields, hazard scope, and outbreak evidence.
    """
    def g(k):
        return row.get(k, "") or ""

    is_rasff = "rasff" in str(g("Source")).lower()
    rasff_note = ("\nRASFF row: its Company/Brand/Country format is already "
                  "correct — keep it, never write 'Unbranded'.\n"
                  if is_rasff else "")

    return (
        "Verify this recall against its source page.\n\n"
        f"Date {g('Date')} | Source {g('Source')} | Country {g('Country')}\n"
        f"Company {g('Company')} | Brand {g('Brand')}\n"
        f"Product {g('Product')}\n"
        f"Pathogen {g('Pathogen')} | Reason {g('Reason')} | "
        f"Outbreak {g('Outbreak')}\n"
        f"URL {g('URL')}\n"
        f"{rasff_note}\n"
        "1. fetch_page the URL. If it is dead or the wrong recall, web_search\n"
        "   for the official regulator page and fetch that instead.\n"
        "2. Correct EVERY field to what the page says. Date = the regulator's\n"
        "   ORIGINAL publication date, not an update. Product, Pathogen,\n"
        "   Reason and Region in ENGLISH; Company and Brand keep their\n"
        "   original language. Product is the food item itself, never the\n"
        "   alert headline and never the agency's name.\n"
        "3. Company and Brand must appear verbatim on the page. If no brand is\n"
        "   named (sold loose / a la coupe / sans marque), use \"Unbranded\".\n"
        "   Never invent one, and never trust a value already on the row\n"
        "   without seeing it on the page.\n"
        "4. SCOPE — approve ONLY a 2026+ food recall whose hazard is a\n"
        "   microbial pathogen (Listeria, Salmonella, E. coli/STEC,\n"
        "   Cronobacter, botulinum, norovirus, hepatitis A). REJECT anything\n"
        "   else and name the real hazard:\n"
        "     \"undeclared allergen (X)\" only for the 14 legal allergens —\n"
        "       SUGAR IS NOT AN ALLERGEN;\n"
        "     \"labelling error\" for wrong or swapped labels, or a wrong sugar\n"
        "       or nutrition declaration;\n"
        "     \"foreign body (X)\", \"chemical contamination (X)\",\n"
        "     \"biotoxin (X)\", \"non-food product\".\n"
        "   Also reject a duplicate, or an UPDATED re-issue of a recall that\n"
        "   is already in the register (put the original URL in duplicate_of).\n"
        "5. OUTBREAK. Ignore standard risk boilerplate (\"may cause severe\n"
        "   illness in pregnant women...\", \"no reported illnesses\") and\n"
        "   RASFF's \"risk: serious\", which is only a severity label present on\n"
        "   every notification. Set 1 ONLY for a stated case count, an\n"
        "   epidemiological investigation opened because people fell ill,\n"
        "   attributed deaths, or a linked CDC / FDA / PHAC / UKHSA outbreak\n"
        "   notice for THIS product. The outbreak page is often separate from\n"
        "   the recall notice, so run ONE web_search for it before settling on\n"
        "   0. Contamination findings and routine or environmental sampling\n"
        "   are 0. Default 0.\n"
        "6. If a required field (Date, Company or Brand, Product, Pathogen,\n"
        "   URL) cannot be confirmed from the page, REJECT — there is no\n"
        "   human to defer to.\n\n"
        "Reply with ONLY this JSON:\n"
        '{"verdict":"approve"|"reject","reason":"<one line>",'
        '"verified_url":"","duplicate_of":"","outbreak":0,'
        '"outbreak_evidence":"","fields":{"Date":"","Company":"","Brand":"",'
        '"Product":"","Pathogen":"","Reason":"","Country":"","Region":""},'
        '"provenance":""}'
    )

# ─── The agent ───────────────────────────────────────────────────────────

def review_row(row: Dict[str, Any]) -> Dict[str, Any]:
    """Run the Qwen review agent on one row. Verdict is one of:
       - "approve" : verified + complete + in-scope 2026+ pathogen recall
       - "reject"  : the model READ the page and it is invalid / out of scope
                     / incomplete / cannot be verified. There is NO human, so
                     an unverifiable row is rejected, not held.
       - "retry"   : INFRASTRUCTURE failure only (llama down / no response /
                     unparseable). NOT the row's fault → leave it untouched in
                     Pending so the next scheduled run tries again.
    """
    infra = {"verdict": "retry", "fields": {},
             "verified_url": row.get("URL", ""), "provenance": "",
             "outbreak": row.get("Outbreak", 0), "outbreak_evidence": ""}
    if llama_client is None or not llama_client.is_configured():
        return {**infra, "reason": "INFRA: llama not configured (retry next run)"}
    if llama_client.is_open():
        return {**infra, "reason": "INFRA: llama circuit breaker open (retry)"}

    seen: set = set()
    messages = [
        {"role": "system", "content": REVIEW_SYSTEM},
        {"role": "user", "content": build_review_prompt(row)},
    ]
    out = llama_client.chat(
        messages=messages,
        tools=_tool_schema(),
        tool_executor=_make_tool_executor(seen),
        temperature=0.0,
        max_tokens=700,
    )
    if not out:
        return {**infra, "reason": "INFRA: no response from llama (retry)"}
    # Parse the JSON (strip any accidental fences)
    txt = out.strip()
    if txt.startswith("```"):
        txt = txt.strip("`")
        if txt.lower().startswith("json"):
            txt = txt[4:]
    try:
        parsed = json.loads(txt)
    except json.JSONDecodeError:
        import re
        m = re.search(r"\{.*\}", txt, re.DOTALL)
        if not m:
            return {**infra, "reason": f"INFRA: unparseable output (retry): {txt[:80]}"}
        try:
            parsed = json.loads(m.group(0))
        except json.JSONDecodeError:
            return {**infra, "reason": "INFRA: json parse failed (retry)"}
    # Valid JSON → the model's verdict is authoritative, but only approve /
    # reject are accepted. Anything else (including a stray 'needs_human')
    # collapses to REJECT: the model read the page and did not verify it.
    v = str(parsed.get("verdict", "")).strip().lower()
    if v not in ("approve", "reject"):
        parsed["verdict"] = "reject"
        parsed["reason"] = ("could not verify as valid in-scope recall: "
                            + str(parsed.get("reason", ""))[:200])
    parsed.setdefault("fields", {})
    parsed.setdefault("verified_url", row.get("URL", ""))
    parsed.setdefault("outbreak", row.get("Outbreak", 0))
    parsed.setdefault("outbreak_evidence", "")
    # ── DETERMINISTIC SCOPE GUARD (does not rely on the model) ──
    # A pre-2026 publication date is out of scope, period. If the model
    # approved a row whose verified Date is before 2026, override to reject.
    # This is the backstop for the stale-recall class (2024 statements and
    # Dec-2025 recalls that arrived stamped with a fresh scrape date).
    if parsed.get("verdict") == "approve":
        # Field-integrity backstop: untranslated text, placeholder strings,
        # headline-as-product and regulator-as-product are all disqualifying.
        _m = dict(row)
        for _k, _v in (parsed.get("fields") or {}).items():
            if _v:
                _m[_k] = _v
        _probs = _field_integrity_flags(_m)
        if _probs:
            parsed["verdict"] = "reject"
            parsed["reason"] = "field integrity: " + "; ".join(_probs[:3])
            return parsed
        d = str((parsed.get("fields") or {}).get("Date")
                or row.get("Date") or "").strip()[:10]
        if len(d) >= 4 and d[:4].isdigit() and int(d[:4]) < 2026:
            parsed["verdict"] = "reject"
            parsed["reason"] = (f"out of scope: original publication date {d} "
                                f"is before 2026")
    return parsed


# Words that betray an untranslated / half-translated field. Deliberately
# short and unambiguous — these are not English and appear in real rows.
_NON_ENGLISH_TOKENS = (
    " dans le ", " du produit", "presencia", "procedente", "procedentes",
    "presence de", "présence", " et de ", " avec ", " sur place",
    " nella ", " nel prodotto", " im produkt", " en el producto",
    " a l'ail", "seche", "sèche", "fabriquees", "fabriquées",
)
# Strings that are explanations, not values.
_PLACEHOLDER_MARKERS = (
    "not specified", "non spécifié", "no especificado", "not stated",
    "unknown", "n/a", "see notice", "see the notice", "aucune information",
)


def _field_integrity_flags(merged: Dict[str, Any]) -> List[str]:
    """Deterministic checks the model cannot skip. Returns a list of problems;
    an approved row with any problem is downgraded to reject.

    RASFF EXEMPTION: RASFF rows are correct by design — the notification
    subject line IS the Product, it stays in the notifier's language, Brand
    is the notifying-country ISO code (or a dash), and Company is the fixed
    "Origin: X | Notifying: Y" string. Applying the language / headline /
    placeholder heuristics to them produces false positives, so RASFF rows
    are checked only for outright emptiness.
    """
    probs = []
    if "rasff" in str(merged.get("Source", "")).lower():
        for fld in ("Product", "URL"):
            if not str(merged.get(fld, "") or "").strip():
                probs.append(f"{fld} is empty")
        return probs
    for fld in ("Product", "Reason", "Region", "Class"):
        v = str(merged.get(fld, "") or "").lower()
        if not v:
            continue
        for tok in _NON_ENGLISH_TOKENS:
            if tok in v:
                probs.append(f"{fld} not in English ({tok.strip()!r})")
                break
    for fld in ("Company", "Brand", "Product", "Reason"):
        v = str(merged.get(fld, "") or "").lower()
        if any(m in v for m in _PLACEHOLDER_MARKERS):
            probs.append(f"{fld} holds a placeholder string, not a value")
    # Product must not be the regulator / an alert headline.
    prod = str(merged.get("Product", "") or "")
    if prod.lstrip().startswith("#") or "\n" in prod:
        probs.append("Product contains formatting debris (# or newline)")
    if len(prod) > 160:
        probs.append("Product looks like a headline, not a product name")
    for agency in ("Agencia Española", "Agencia Espanola", "Food Standards",
                   "Autorité", "Bundesamt", "Ministero della Salute"):
        if agency.lower() in prod.lower():
            probs.append("Product contains the regulator's name")
            break
    # Company and Brand identical AND long => both are the headline.
    c = str(merged.get("Company", "") or "")
    b = str(merged.get("Brand", "") or "")
    if c and c == b and len(c) > 60:
        probs.append("Company and Brand are the same long headline string")
    return probs


def _normalize_country_source(merged: Dict[str, Any]) -> None:
    """Enforce dataset conventions (from the repo's own canonical usage):
       - Country: the United States is written "United States", never "USA"
         (82 rows + gap_finder_tavily/regulator_apis/gap_finder_claude agree).
       - Source: US meat/poultry recalls are "USDA FSIS" (consistent form).
    """
    c = str(merged.get("Country", "")).strip()
    if c.upper() in ("USA", "U.S.A.", "US", "U.S.", "UNITED STATES OF AMERICA",
                     "AMERICA"):
        merged["Country"] = "United States"
    s = str(merged.get("Source", "")).strip()
    # Normalise any FSIS/USDA source label to the canonical "USDA FSIS".
    sl = s.lower()
    if ("fsis" in sl or "usda" in sl) and s != "USDA FSIS":
        merged["Source"] = "USDA FSIS"


def apply_review(row: Dict[str, Any], review: Dict[str, Any]) -> Dict[str, Any]:
    """Merge the agent's verified fields back into the row (corrections win,
    but never blank a field the row already had unless the agent explicitly
    says the page contradicts it — here we take non-empty agent values)."""
    merged = dict(row)
    fields = review.get("fields") or {}
    for k in ("Date", "Company", "Brand", "Product", "Pathogen", "Reason",
              "Country", "Region"):
        v = fields.get(k)
        if v:  # only overwrite with a non-empty verified value
            merged[k] = v
    vu = review.get("verified_url")
    if vu:
        merged["URL"] = vu
    _normalize_country_source(merged)
    return merged


# ─── Sheet I/O ───────────────────────────────────────────────────────────

def load_pending(xlsx: Path) -> Tuple[List[Dict[str, Any]], List[str]]:
    return _load_sheet(xlsx, "Pending"), _sheet_headers(xlsx, "Pending")


def _load_sheet(xlsx: Path, sheet: str) -> List[Dict[str, Any]]:
    import openpyxl
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        return []
    ws = wb[sheet]
    headers = [c.value for c in ws[1]]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        row = {h: ("" if v is None else v) for h, v in zip(headers, r) if h}
        if any(str(v).strip() for v in row.values()):
            rows.append(row)
    return rows


def _sheet_headers(xlsx: Path, sheet: str) -> List[str]:
    import openpyxl
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        return []
    return [c.value for c in wb[sheet][1] if c.value]


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=Path, default=Path("docs/data/recalls.xlsx"))
    ap.add_argument("--commit", type=str, default="false")
    ap.add_argument("--source-filter", type=str, default=None)
    ap.add_argument("--limit", type=int, default=0)
    ap.add_argument("--time-budget-min", type=int, default=20,
                    help="Stop reviewing after N minutes and SAVE what has "
                         "been done, so a workflow timeout cannot discard the "
                         "run. Remaining rows are picked up next run.")
    ap.add_argument("--max-removals", type=int, default=10,
                    help="Audit mode safety cap: abort without writing if the "
                         "model wants to remove more than N rows from Recalls.")
    ap.add_argument("--audit-recalls", type=str, default="false",
                    help="Re-verify rows ALREADY in Recalls (audit mode). "
                         "Corrects fields in place; moves out-of-scope / "
                         "duplicate / unverifiable rows to Weekly_Rejected.")
    args = ap.parse_args()
    commit = args.commit.lower() in ("1", "true", "yes", "on")
    if args.audit_recalls.lower() in ("1", "true", "yes", "on"):
        return audit_recalls(args, commit)

    rows, _ = load_pending(args.xlsx)
    if args.source_filter:
        sf = args.source_filter.lower()
        rows = [r for r in rows if sf in str(r.get("Source", "")).lower()]
    if args.limit and args.limit > 0:
        rows = rows[:args.limit]

    print(f"Reviewing {len(rows)} Pending rows "
          f"(commit={commit}, source_filter={args.source_filter})")
    if not rows:
        print("Nothing to review.")
        return 0

    results = {"approve": [], "reject": [], "retry": []}
    _deadline = (dt.datetime.now(dt.timezone.utc)
                 + dt.timedelta(minutes=max(1, args.time_budget_min)))
    _stopped_early = 0
    for i, row in enumerate(rows, 1):
        if dt.datetime.now(dt.timezone.utc) >= _deadline:
            _stopped_early = len(rows) - i + 1
            print(f"\n  [time budget {args.time_budget_min}m reached] "
                  f"stopping after {i-1} rows; {_stopped_early} left for the "
                  f"next run. Saving progress now.")
            break
        review = review_row(row)
        review["_orig_url"] = str(row.get("URL", "")).strip()
        verdict = review.get("verdict", "retry")
        merged = apply_review(row, review)
        results[verdict if verdict in results else "retry"].append(
            (merged, review))
        print(f"  [{i}/{len(rows)}] {verdict.upper():8s} "
              f"{str(row.get('Source',''))[:14]:14s} "
              f"{str(merged.get('Product',''))[:44]:44s} "
              f"| {review.get('reason','')[:60]}")

    if _stopped_early:
        print(f"NOTE: {_stopped_early} rows not reviewed this run (time budget).")
    print(f"\n{'='*60}")
    print(f"approve: {len(results['approve'])}  "
          f"reject: {len(results['reject'])}  "
          f"retry (infra, left in Pending): {len(results['retry'])}")
    print(f"{'='*60}")

    if not commit:
        print("\nDRY RUN — no writes. Set --commit true to apply:")
        print("  approvals → Recalls (corrected fields), "
              "rejects → Weekly_Rejected, retry → untouched in Pending.")
        return 0

    # ── Write-back: mirror claude_check.py's final-reviewer sequence ──
    # The review agent IS "Reviewer 2 — final verdict", so promote_approved
    # is called with archive_immediately=True, exactly as claude_check does.
    try:
        from pipeline.merge_master import (  # type: ignore
            promote_approved, sort_rows, save_xlsx_with_pending,
            mirror_json_from_xlsx)
    except Exception as e:
        print(f"ERROR importing merge_master helpers: {e}", file=sys.stderr)
        return 1

    XLSX_PATH = Path(args.xlsx)

    # Recalls (approved_existing) + FULL Pending, so indices line up.
    approved_existing = _load_sheet(args.xlsx, "Recalls")
    full_pending = _load_sheet(args.xlsx, "Pending")

    url_to_idx: Dict[str, int] = {}
    for i, pr in enumerate(full_pending):
        u = str(pr.get("URL", "")).strip()
        if u:
            url_to_idx[u] = i

    rejected_flags: Dict[int, str] = {}
    applied_corrections = 0

    # Approvals — write corrected fields IN PLACE; leave out of rejected_flags
    # Also advance gap-gating / enrichment statuses to plain 'pending' so
    # promote_approved will actually promote them (it skips pending_gap* and
    # pending_enrichment). This replicates claude_check.py's documented
    # state-machine advance (audit 2026-04-29 / 2026-05-11) — the review
    # agent is the 2nd/final reviewer, so an approval flips the gate.
    _ADVANCE_FROM = {"pending_gap_v1", "pending_gap_v2", "pending_retry",
                     "pending_enrichment"}
    today_iso = dt.datetime.now(dt.timezone.utc).strftime("%Y-%m-%d")
    gap_advanced = 0
    for merged, review in results["approve"]:
        idx = url_to_idx.get(str(merged.get("URL", "")).strip())
        if idx is None:
            idx = url_to_idx.get(str(review.get("_orig_url", "")).strip())
        if idx is not None:
            for k in ("Date", "Company", "Brand", "Product", "Pathogen",
                      "Reason", "Country", "Region", "URL"):
                if merged.get(k):
                    full_pending[idx][k] = merged[k]
            # Pathogen may be intentionally CLEARED by the agent when the
            # hazard is an allergen / foreign body / chemical (not a pathogen).
            # merged only carries non-empty values, so consult the raw review
            # fields dict: if it explicitly returned Pathogen == "" we honor it.
            rfields = review.get("fields") or {}
            if "Pathogen" in rfields and not str(rfields.get("Pathogen")).strip():
                full_pending[idx]["Pathogen"] = ""
            # Company/Brand: honor an explicit correction to empty (a
            # fabricated value the page doesn't support). "Unbranded" is a
            # normal non-empty value and is written via the loop above.
            for fld in ("Company", "Brand"):
                if fld in rfields and not str(rfields.get(fld)).strip():
                    full_pending[idx][fld] = ""
            # Outbreak is verified explicitly (0 or 1) — always apply it,
            # including a correction from 1→0, since it drives the tier bump.
            ob = review.get("outbreak")
            if ob in (0, 1, "0", "1"):
                full_pending[idx]["Outbreak"] = int(ob)
            applied_corrections += 1
            cur = str(full_pending[idx].get("Status", "")).strip()
            if cur in _ADVANCE_FROM:
                full_pending[idx]["Status"] = "pending"
                notes = str(full_pending[idx].get("Notes", "")).strip()
                tag = (f"[review-agent {today_iso}: {cur} → pending "
                       f"(Qwen verified, final reviewer)]")
                full_pending[idx]["Notes"] = (notes + " " + tag).strip()[:1000]
                gap_advanced += 1

    # Rejects → rejected_flags (index → reason)
    for merged, review in results["reject"]:
        idx = url_to_idx.get(str(merged.get("URL", "")).strip())
        if idx is not None:
            rejected_flags[idx] = f"Review agent: {review.get('reason','')[:280]}"

    # retry (infra failure) → leave the row COMPLETELY untouched in Pending.
    # Do not change status, do not reject. The next scheduled run retries it.

    # ── HARD SAFETY GUARD (no-assumptions rule) ──
    # Only rows the agent EXPLICITLY approved this run may promote. Any row at
    # promotable status 'pending' that is NOT in the approved set is demoted so
    # a server failure / retry can never leak an unverified row into Recalls.
    # (Fixes the incident where rows already at 'pending' promoted during a
    # total llama outage.)
    approved_urls = set()
    for merged, review in results["approve"]:
        for key in ("URL", "_orig_url"):
            u = str((merged.get(key) if key == "URL" else review.get(key)) or "").strip()
            if u:
                approved_urls.add(u)
    for prow in full_pending:
        if str(prow.get("Status", "")).strip() == "pending":
            u = str(prow.get("URL", "")).strip()
            if u not in approved_urls:
                prow["Status"] = "pending_gap_v2"  # hold; do not promote
                note = str(prow.get("Notes", "")).strip()
                prow["Notes"] = (note + " [safety-hold: not approved this "
                                 "run; not promoted]").strip()[:1000]

    new_approved, remaining, archived_rejected = promote_approved(
        pending=full_pending,
        approved_existing=approved_existing,
        rejected_flags=rejected_flags,
        archive_immediately=True,
    )

    print(f"\nApplied {applied_corrections} field-corrected approvals "
          f"({gap_advanced} gap/enrichment rows advanced to 'pending').")
    print(f"Promotion: {len(new_approved)} → Recalls, "
          f"{len(remaining)} remain in Pending, "
          f"{len(archived_rejected)} archived to Rejected.")

    final_approved = sort_rows(approved_existing + new_approved)
    final_pending = sort_rows(remaining)
    save_xlsx_with_pending(final_approved, final_pending, XLSX_PATH,
                           newly_rejected_rows=archived_rejected)
    try:
        from pipeline.merge_master import JSON_PATH  # type: ignore
        mirror_json_from_xlsx(XLSX_PATH, JSON_PATH)
    except Exception:
        pass

    try:
        from pipeline.weekly_review_capture import record_promotions  # noqa
        record_promotions(new_approved, xlsx_path=XLSX_PATH)
    except Exception as e:
        print(f"  (Weekly_Review capture skipped: {e})")
    try:
        from pipeline.weekly_rejected_capture import record_rejections  # noqa
        record_rejections(archived_rejected, xlsx_path=XLSX_PATH)
    except Exception as e:
        print(f"  (Weekly_Rejected capture skipped: {e})")

    print("\n✓ Write-back complete via promote_approved (final-reviewer mode).")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())


# ─── AUDIT MODE: re-verify rows ALREADY promoted to Recalls ──────────────

def audit_recalls(args, commit: bool) -> int:
    """Re-review rows already in Recalls against their source pages.

    Exists because bad rows reach Recalls when an enricher fabricates a field
    and a reviewer passes it on a 'clean-row shortcut' (documented FSANZ
    Listeria default-tagging incident). The scheduled agents only see Pending,
    so nothing ever re-checks a promoted row. This does.

    approve → corrections applied in place (Pathogen, Company, Date, …)
    reject  → row MOVED out of Recalls into Weekly_Rejected with the reason
    retry   → row left exactly as-is (infra failure, not the row's fault)
    """
    import openpyxl
    xlsx = Path(args.xlsx)
    wb = openpyxl.load_workbook(xlsx)
    ws = wb["Recalls"]
    headers = [c.value for c in ws[1]]
    hidx = {h: i for i, h in enumerate(headers) if h}

    rows = []
    for ridx in range(2, ws.max_row + 1):
        row = {h: (ws.cell(ridx, i + 1).value or "") for h, i in hidx.items()}
        if any(str(v).strip() for v in row.values()):
            rows.append((ridx, row))

    if args.source_filter:
        sf = args.source_filter.lower()
        rows = [(i, r) for i, r in rows
                if sf in str(r.get("Source", "")).lower()]
    if args.limit and args.limit > 0:
        rows = rows[:args.limit]

    print(f"AUDIT MODE — re-verifying {len(rows)} rows already in Recalls "
          f"(commit={commit})")
    if not rows:
        print("Nothing to audit.")
        return 0

    fixed, to_remove, retried = [], [], 0
    for n, (ridx, row) in enumerate(rows, 1):
        review = review_row(row)
        v = review.get("verdict", "retry")
        if v == "retry":
            retried += 1
            print(f"  [{n}/{len(rows)}] RETRY    {str(row.get('Product',''))[:40]}")
            continue
        if v == "reject":
            to_remove.append((ridx, row, review.get("reason", "")[:200]))
            print(f"  [{n}/{len(rows)}] REMOVE   "
                  f"{str(row.get('Product',''))[:40]:40s} | {review.get('reason','')[:45]}")
            continue
        merged = apply_review(row, review)
        changes = {k: (row.get(k), merged.get(k))
                   for k in ("Date", "Company", "Brand", "Product", "Pathogen",
                             "Reason", "Country", "Region", "URL")
                   if str(row.get(k, "")) != str(merged.get(k, ""))}
        ob = review.get("outbreak")
        if ob in (0, 1, "0", "1") and str(row.get("Outbreak", "")) != str(int(ob)):
            changes["Outbreak"] = (row.get("Outbreak"), int(ob))
        if changes:
            fixed.append((ridx, changes))
            print(f"  [{n}/{len(rows)}] FIX      "
                  f"{str(row.get('Product',''))[:40]:40s} | "
                  + ", ".join(f"{k}:{a!r}->{b!r}" for k, (a, b) in
                              list(changes.items())[:2])[:60])
        else:
            print(f"  [{n}/{len(rows)}] OK       {str(row.get('Product',''))[:40]}")

    print(f"\n{'='*60}")
    print(f"fix: {len(fixed)}  remove: {len(to_remove)}  "
          f"retry(infra): {retried}  ok: {len(rows)-len(fixed)-len(to_remove)-retried}")
    print(f"{'='*60}")

    # ── SAFETY CAP ──
    # Audit mode deletes from the published register. A misbehaving or
    # context-starved model that rejects everything must not be able to wipe
    # good rows. Abort the whole run if removals exceed the cap.
    cap = max(0, int(getattr(args, "max_removals", 10)))
    if len(to_remove) > cap:
        print(f"\n*** ABORTED — model wants to remove {len(to_remove)} rows, "
              f"cap is {cap}. NOTHING WRITTEN. ***")
        print("    This usually means the model is failing (context exceeded, "
              "bad output) rather than the data being wrong.")
        print(f"    Review the list above. If the removals are genuinely "
              f"correct, re-run with --max-removals {len(to_remove)}.")
        return 2
    if to_remove and len(to_remove) > max(1, len(rows) // 2):
        print(f"\n*** ABORTED — {len(to_remove)} of {len(rows)} reviewed rows "
              f"would be removed (>50%). NOTHING WRITTEN. ***")
        return 2

    if not commit:
        print("\nDRY RUN — nothing written. Re-run with --commit true.")
        return 0

    today = dt.datetime.now(dt.timezone.utc).strftime("%Y-%m-%d")
    # apply field fixes in place
    for ridx, changes in fixed:
        for k, (old, new) in changes.items():
            if k in hidx:
                ws.cell(ridx, hidx[k] + 1).value = new
        if "Notes" in hidx:
            cur = str(ws.cell(ridx, hidx["Notes"] + 1).value or "")
            desc = "; ".join(f"{k} {old!r}->{new!r}"
                             for k, (old, new) in changes.items())
            ws.cell(ridx, hidx["Notes"] + 1).value = (
                cur + f" [audit-agent {today}: {desc}]").strip()[:2000]

    # move rejected rows to Weekly_Rejected, then delete bottom-up
    if to_remove and "Weekly_Rejected" in wb.sheetnames:
        wr = wb["Weekly_Rejected"]
        wrh = [c.value for c in wr[1]]
        for _ridx, row, reason in to_remove:
            new = [""] * len(wrh)
            for k, v in row.items():
                if k in wrh:
                    new[wrh.index(k)] = v
            for rc in ("RejectionReason", "Reason"):
                if rc in wrh:
                    new[wrh.index(rc)] = f"audit-agent {today}: {reason}"
                    break
            if "RejectedBy" in wrh:
                new[wrh.index("RejectedBy")] = "audit-agent"
            wr.append(new)
    for ridx, _row, _reason in sorted(to_remove, key=lambda x: -x[0]):
        ws.delete_rows(ridx, 1)

    wb.save(xlsx)
    print(f"\n✓ Applied {len(fixed)} field corrections, "
          f"removed {len(to_remove)} rows from Recalls.")
    try:
        from pipeline.merge_master import mirror_json_from_xlsx
        mirror_json_from_xlsx(xlsx, xlsx.parent / "recalls.json")
        print("✓ recalls.json mirrored.")
    except Exception as e:
        print(f"  (JSON mirror skipped: {e})")
    return 0
