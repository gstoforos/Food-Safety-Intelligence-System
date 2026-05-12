"""
Gap-finder cascade — single entry point that tries multiple LLM gap-finders
in priority order, stopping at the first one that adds rows.

CASCADE ORDER (audit 2026-04-29):
  Call 1   Gemini FREE tier (gemini-2.5-flash + Google Search grounding)
           — €0/run; 250 req/day daily cap; this is the workhorse.
  Call 2   Gemini PAID tier (Tier 1, same code path, different key)
           — €0.10/run; only if free-tier rate-limited or returned no
             usable response.
  Call 3   Claude (claude-haiku-4-5 + web_search_20250305)
           — ~$0.005/run; only if both Gemini calls failed or returned 0 rows.
  Call 4   OpenAI (gpt-4o-mini-search-preview)
           — ~$0.10/run; final fallback.

Per-step timeout: 90 seconds. A hung call rolls forward to the next provider
so the entire cron slot can't get blocked.

OUTCOME CLASSIFICATION (audit 2026-05-13)
-----------------------------------------
Each step lands in one of three buckets:

  SUCCESS_WITH_ROWS   rc == 0 AND at least 1 row landed in Pending after
                      the push retry / xlsx_merge dance settled.
                      → cascade STOPS and returns 0.

  SUCCESS_EMPTY       rc == 0 but 0 rows landed. Three sub-cases that we
                      DON'T try to disambiguate at this layer:
                        (a) the provider genuinely found nothing new
                        (b) the provider found candidates but every one
                            was de-duplicated by merge_master (already
                            in Pending or Recalls)
                        (c) a candidate WAS staged but lost in xlsx_merge
                            conflict resolution because a parallel writer
                            had just approved the same URL into Recalls
                            (this is correct behaviour — the row is a
                             duplicate of something just promoted)
                      → cascade FALLS THROUGH to the next provider.

  FAILED              Exception, timeout, or rc != 0.
                      → cascade FALLS THROUGH to the next provider.

After the loop:
  • any SUCCESS_WITH_ROWS  → already returned 0 above
  • ≥1 SUCCESS_EMPTY       → return 0 with a "no new recalls today" log
                              (the world has no new recalls right now,
                               or everything found was a duplicate —
                               that is NOT a failure)
  • all FAILED             → return 1 with a loud error so the operator
                              actually investigates a real outage

Pre-2026-05-13 the cascade returned 1 whenever total_added == 0, which
treated quiet days (every provider runs fine but finds nothing new) as
errors and turned the GitHub Actions runs page red. The audit on
2026-05-12's 20:15 run showed this: 4 providers ran, 0 errored, 1 row
was found by OpenAI but then correctly dropped by xlsx_merge because a
parallel workflow had just approved the same URL — cascade exited 1.

CALL-SITE INTEGRATION
---------------------
Single workflow .github/workflows/gap-finder-cascade.yml invokes this
script. Tavily and Exa stay independent on their own dedicated schedule
— they're complementary deterministic-search backstops, not in the
cascade.

CLI
---
    python -m pipeline.gap_finder_cascade               # default
    python -m pipeline.gap_finder_cascade --skip-paid   # never escalate to paid Gemini
    python -m pipeline.gap_finder_cascade --skip-claude
    python -m pipeline.gap_finder_cascade --skip-openai
"""
from __future__ import annotations
import argparse
import logging
import os
import signal
import sys
import time
from contextlib import contextmanager
from datetime import datetime, timezone
from enum import Enum
from pathlib import Path
from typing import Callable, List, Tuple

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    handlers=[logging.StreamHandler()],
)
log = logging.getLogger("gap-finder-cascade")


# ─────────────────────────────────────────────────────────────────────────
# Per-step timeout via SIGALRM. POSIX-only; Windows runners (none of our
# GitHub Actions are Windows) silently degrade to no timeout — that's OK
# because Gemini/Claude/OpenAI clients have their own request timeouts.
# ─────────────────────────────────────────────────────────────────────────
class StepTimeout(Exception):
    pass


@contextmanager
def step_timeout(seconds: int):
    """Best-effort POSIX timeout. No-op on Windows."""
    if not hasattr(signal, "SIGALRM"):
        yield
        return

    def _handler(signum, frame):
        raise StepTimeout(f"step exceeded {seconds}s")

    old_handler = signal.signal(signal.SIGALRM, _handler)
    signal.alarm(seconds)
    try:
        yield
    finally:
        signal.alarm(0)
        signal.signal(signal.SIGALRM, old_handler)


# ─────────────────────────────────────────────────────────────────────────
# Step outcome — three-state classifier replaces the old (ok, added) tuple
# semantics at the cascade-level decision boundary (audit 2026-05-13).
# ─────────────────────────────────────────────────────────────────────────
class StepOutcome(Enum):
    SUCCESS_WITH_ROWS = "rows"     # rc==0 AND added>0 → stop cascade
    SUCCESS_EMPTY     = "empty"    # rc==0 AND added==0 → fall through, NOT a failure
    FAILED            = "fail"     # exception/timeout/rc!=0 → fall through


# ─────────────────────────────────────────────────────────────────────────
# Provider adapters — each function tries one provider and returns
#   (ok, added_unused, msg)
# where ok==True means rc==0 (the provider ran without error). The
# `added_unused` field is kept for log-message compatibility; actual row
# counting is done by the cascade via _pending_count() before/after.
# ─────────────────────────────────────────────────────────────────────────

def _try_gemini(force_paid: bool = False) -> Tuple[bool, int, str]:
    """Call gap_finder_gemini.main(). Per the gap-finder's own logic, it
    rotates through GEMINI_API_KEY_FREE first, then paid keys. We just
    call main() and trust the rotation — the key swap happens inside.

    For the cascade's call 2 (PAID escalation), we set
    GAP_GEMINI_DISABLE_FREE=1 so the gap-finder skips the free-tier key
    entirely. This avoids re-trying the same rate-limited key.
    """
    try:
        from pipeline import gap_finder_gemini
    except ImportError as e:
        return False, 0, f"import failed: {e}"

    if force_paid:
        os.environ["GAP_GEMINI_DISABLE_FREE"] = "1"
        log.info("Gemini retry: disabling FREE-tier key (paid tier only)")
    else:
        os.environ.pop("GAP_GEMINI_DISABLE_FREE", None)

    try:
        rc = gap_finder_gemini.main()
        return (rc == 0), 0, f"main() returned {rc}"
    except Exception as e:  # noqa: BLE001
        return False, 0, f"exception: {type(e).__name__}: {e}"
    finally:
        if force_paid:
            os.environ.pop("GAP_GEMINI_DISABLE_FREE", None)


def _try_claude() -> Tuple[bool, int, str]:
    try:
        from pipeline import gap_finder_claude
    except ImportError as e:
        return False, 0, f"import failed: {e}"
    try:
        rc = gap_finder_claude.main()
        return (rc == 0), 0, f"main() returned {rc}"
    except Exception as e:  # noqa: BLE001
        return False, 0, f"exception: {type(e).__name__}: {e}"


def _try_openai() -> Tuple[bool, int, str]:
    """Call gap_finder_openai.main() — the final fallback. The OpenAI
    finder is OPTIONAL: if the module isn't installed we just skip it
    cleanly rather than crashing the cascade."""
    try:
        from pipeline import gap_finder_openai  # type: ignore
    except ImportError:
        log.info("gap_finder_openai not installed — skipping (OK, optional)")
        return False, 0, "module not present (OPENAI provider optional)"
    if not os.environ.get("OPENAI_API_KEY", "").strip():
        log.info("OPENAI_API_KEY not set — skipping OpenAI fallback")
        return False, 0, "OPENAI_API_KEY not set"
    try:
        rc = gap_finder_openai.main()
        return (rc == 0), 0, f"main() returned {rc}"
    except Exception as e:  # noqa: BLE001
        return False, 0, f"exception: {type(e).__name__}: {e}"


# ─────────────────────────────────────────────────────────────────────────
# Cascade
# ─────────────────────────────────────────────────────────────────────────

CASCADE_TIMEOUT_S = int(os.environ.get("GAP_STEP_TIMEOUT", "180"))

_XLSX_PATH = Path(os.environ.get(
    "FSIS_XLSX_PATH",
    str(Path(__file__).resolve().parent.parent / "docs" / "data" / "recalls.xlsx"),
))


def _pending_count() -> int:
    """Read the current Pending sheet row count. Used to detect whether a
    cascade step actually wrote anything new that survived the push. -1
    on load failure (treated as 'unknown' by the caller — no row credit).
    """
    try:
        if not _XLSX_PATH.exists():
            return 0
        import openpyxl
        wb = openpyxl.load_workbook(_XLSX_PATH, read_only=True, data_only=True)
        if "Pending" not in wb.sheetnames:
            wb.close()
            return 0
        ws = wb["Pending"]
        n = max(0, ws.max_row - 1)  # max_row counts the header row too
        wb.close()
        return n
    except Exception as e:
        log.warning("_pending_count: failed to read xlsx (%s) — "
                    "treating as unknown", e)
        return -1


def run_cascade(skip_paid: bool = False, skip_claude: bool = False,
                skip_openai: bool = False) -> int:
    """Run the cascade. Exit-code policy (audit 2026-05-13):

        any step  SUCCESS_WITH_ROWS                 → 0  (real success, stop)
        ≥1 step   SUCCESS_EMPTY                     → 0  ("nothing new today")
        every step FAILED                           → 1  (real outage)

    Returning 0 on the SUCCESS_EMPTY-only path is the change vs the
    2026-05-06 logic. Empty/duplicate-only results are a legitimate
    outcome — the world doesn't produce a brand-new recall on every
    cron tick, and merge_master + xlsx_merge correctly de-dup against
    rows that have just been approved. Reporting that as a workflow
    failure (red CI) drowns out real failures.
    """
    started = datetime.now(timezone.utc).isoformat()
    log.info("=" * 60)
    log.info("Gap-finder cascade started: %s (per-step timeout=%ds)",
             started, CASCADE_TIMEOUT_S)

    steps: List[Tuple[str, Callable[[], Tuple[bool, int, str]]]] = [
        ("Gemini (free→paid)", lambda: _try_gemini(force_paid=False)),
    ]
    if not skip_paid:
        steps.append(("Gemini (paid only)", lambda: _try_gemini(force_paid=True)))
    if not skip_claude:
        steps.append(("Claude",      _try_claude))
    if not skip_openai:
        steps.append(("OpenAI",      _try_openai))

    initial_pending = _pending_count()
    log.info("Cascade baseline: Pending has %d rows", initial_pending)

    last_msg = ""
    total_added = 0
    outcomes: List[Tuple[str, StepOutcome]] = []
    summary_parts: List[str] = []  # provider-by-provider for final log

    for idx, (name, fn) in enumerate(steps, 1):
        log.info("─" * 50)
        log.info("Cascade call %d/%d: %s", idx, len(steps), name)
        before = _pending_count()
        t0 = time.time()
        try:
            with step_timeout(CASCADE_TIMEOUT_S):
                ok, _added, msg = fn()
        except StepTimeout as e:
            ok, msg = False, f"timeout: {e}"
        elapsed = time.time() - t0
        last_msg = msg
        after = _pending_count()

        if before < 0 or after < 0:
            added_this_step = 0
            log.warning("Cascade row-count metric unreliable (before=%d after=%d)",
                        before, after)
        else:
            added_this_step = max(0, after - before)
        total_added += added_this_step

        if ok and added_this_step > 0:
            outcome = StepOutcome.SUCCESS_WITH_ROWS
            log.info("✓ Cascade success at step %d (%s) in %.1fs — STOP",
                     idx, name, elapsed)
            log.info("  Message: %s", msg)
            log.info("  Added: %d new rows to Pending", added_this_step)
            outcomes.append((name, outcome))
            summary_parts.append(f"{name}=+{added_this_step}")
            log.info("Cascade summary: %s | total +%d rows",
                     " → ".join(summary_parts), total_added)
            return 0
        elif ok:
            outcome = StepOutcome.SUCCESS_EMPTY
            log.warning("⊘ Step %d (%s) ran successfully in %.1fs but added "
                        "ZERO rows — falling through to next provider",
                        idx, name, elapsed)
            log.warning("  Message: %s", msg)
            summary_parts.append(f"{name}=0(empty)")
        else:
            outcome = StepOutcome.FAILED
            log.warning("✗ Step %d (%s) failed in %.1fs — falling through",
                        idx, name, elapsed)
            log.warning("  Reason: %s", msg)
            summary_parts.append(f"{name}=fail")
        outcomes.append((name, outcome))

    # ─────────────────────────────────────────────────────────────────
    # Loop ended without a SUCCESS_WITH_ROWS short-circuit. Classify
    # the overall run.
    # ─────────────────────────────────────────────────────────────────
    has_empty_success = any(o == StepOutcome.SUCCESS_EMPTY for _, o in outcomes)
    all_failed = all(o == StepOutcome.FAILED for _, o in outcomes)
    n_empty = sum(1 for _, o in outcomes if o == StepOutcome.SUCCESS_EMPTY)
    n_failed = sum(1 for _, o in outcomes if o == StepOutcome.FAILED)

    if total_added > 0:
        # Defensive: a step wrote rows but its rc != 0 so we didn't
        # short-circuit. Treat as success.
        log.info("✓ Cascade completed with +%d total rows (across %d steps)",
                 total_added, len(steps))
        log.info("Cascade summary: %s", " → ".join(summary_parts))
        return 0

    if all_failed:
        log.error("All %d cascade providers FAILED — no provider ran "
                  "successfully today.", len(steps))
        log.error("Cascade summary: %s", " → ".join(summary_parts))
        log.error("Last message: %s", last_msg)
        return 1

    # has_empty_success == True: at least one provider ran cleanly and
    # found nothing new (or everything it found was a duplicate of rows
    # already in Pending/Recalls). That's a legitimate quiet result.
    log.info("◌ Cascade complete: no new recalls today "
             "(%d provider(s) ran empty, %d failed). NOT a workflow "
             "failure — see merge_master logs for any duplicates "
             "filtered upstream.", n_empty, n_failed)
    log.info("Cascade summary: %s", " → ".join(summary_parts))
    return 0


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--skip-paid", action="store_true",
                    help="Don't escalate to paid Gemini if free fails "
                         "(useful in cost-sensitive cron slots).")
    ap.add_argument("--skip-claude", action="store_true",
                    help="Skip the Claude fallback step.")
    ap.add_argument("--skip-openai", action="store_true",
                    help="Skip the OpenAI fallback step.")
    args = ap.parse_args()
    return run_cascade(
        skip_paid=args.skip_paid,
        skip_claude=args.skip_claude,
        skip_openai=args.skip_openai,
    )


if __name__ == "__main__":
    sys.exit(main())
