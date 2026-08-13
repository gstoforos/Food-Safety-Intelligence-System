#!/usr/bin/env python3
"""
fix_outbreak_dedupe.py  —  one event, one Outbreak flag
========================================================

Operator rule (2026-08-13): when several rows cover the SAME outbreak, only one
carries Outbreak=1, and FDA is the source of record.

The 2026-08 S. Javiana jalapeño event reached the register as three flagged
rows — an FDA outbreak-investigation page, a CDC investigation page, and a USDA
FSIS public health alert — so a monthly that counts flagged rows reported one
outbreak three times.

Grouping is deliberately conservative (see pipeline/_outbreak_id.py):
  * an agency's own investigation slug IS the event;
  * two slugs from DIFFERENT agencies sharing pathogen+commodity+month are one
    event and merge (CDC and FDA both publish a slug for the same outbreak);
  * two slugs from the SAME agency are never fused — CDC ran two separate
    moringa investigations in 2026-05 and they stay separate;
  * a row that cannot be placed keeps its flag. An unidentifiable outbreak is
    still an outbreak.

Only the Outbreak field changes. Nothing is deleted, no other field is touched.

Usage:
    python fix_outbreak_dedupe.py --xlsx docs/data/recalls.xlsx --commit false
    python fix_outbreak_dedupe.py --xlsx docs/data/recalls.xlsx --commit true
"""
from __future__ import annotations
import argparse
import datetime as dt
import sys
from pathlib import Path


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=Path, default=Path("docs/data/recalls.xlsx"))
    ap.add_argument("--commit", type=str, default="false")
    args = ap.parse_args()
    commit = args.commit.lower() in ("1", "true", "yes", "on")

    sys.path.insert(0, ".")
    from pipeline._outbreak_id import dedupe_outbreak_flags

    import openpyxl
    wb = openpyxl.load_workbook(args.xlsx)
    ws = wb["Recalls"]
    headers = [c.value for c in ws[1]]
    for need in ("Outbreak", "URL"):
        if need not in headers:
            print(f"Recalls has no {need} column.")
            return 1
    oi = headers.index("Outbreak") + 1
    ui = headers.index("URL") + 1
    ni = headers.index("Notes") + 1 if "Notes" in headers else None

    rows = []
    for ridx in range(2, ws.max_row + 1):
        row = {h: (ws.cell(ridx, i + 1).value or "")
               for i, h in enumerate(headers) if h}
        row["_ridx"] = ridx
        rows.append(row)

    losers = dedupe_outbreak_flags(rows)
    print(f"Rows that should lose Outbreak=1: {len(losers)}\n")
    for r in losers:
        print(f"  event {r.get('_outbreak_event')}  (flag stays on "
              f"{r.get('_outbreak_keeper')})")
        print(f"     {str(r.get('Source'))[:12]:12s} "
              f"{str(r.get('Company'))[:44]}")

    if not commit:
        print("\nDRY RUN — nothing written. Re-run with --commit true.")
        return 0
    if not losers:
        print("Nothing to change.")
        return 0

    today = dt.datetime.now(dt.timezone.utc).strftime("%Y-%m-%d")
    for r in losers:
        ridx = r["_ridx"]
        ws.cell(ridx, oi).value = 0
        if ni:
            old = str(ws.cell(ridx, ni).value or "")
            ws.cell(ridx, ni).value = (
                old + f" [outbreak-dedupe {today}: Outbreak 1→0; event "
                f"{r.get('_outbreak_event')} is already flagged on the "
                f"{r.get('_outbreak_keeper')} row — one event, one flag]"
            ).strip()[:2000]

    wb.save(args.xlsx)
    print(f"\n✓ Cleared {len(losers)} duplicate outbreak flag(s).")
    try:
        from pipeline.merge_master import mirror_json_from_xlsx
        mirror_json_from_xlsx(args.xlsx, args.xlsx.parent / "recalls.json")
        print("✓ recalls.json mirrored.")
    except Exception as e:
        print(f"  (JSON mirror skipped: {e})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
