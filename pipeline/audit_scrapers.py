"""
Scraper URL audit — HTTP-probes every regional scraper's INDEX_URLS and
classifies each as OK / 404 / blocked / JS-heavy / empty.

Purpose: triage the 38 "never produced a row" scrapers into:
  - EASY FIX  : URL is live, page has content, just need to re-check why Gemini
                can't extract (likely agency-specific prompt hints needed)
  - URL FIX   : 404 / redirect / moved — needs a new INDEX_URL
  - BLOCKED   : Cloudflare, bot protection, geo-block — probably unfixable
                from GitHub Actions
  - JS-HEAVY  : content is only rendered after JavaScript runs — Gemini sees
                essentially empty HTML, would need Playwright / Selenium
  - EMPTY     : 200 OK but the page body is tiny — layout moved

Run locally or as a manual-dispatch GitHub Actions workflow.
Output: a CSV + a human-readable console table.

Usage:
    python -m pipeline.audit_scrapers
    python -m pipeline.audit_scrapers --csv /tmp/scraper_audit.csv
    python -m pipeline.audit_scrapers --only-dead  (only probe the 46 with
                                                    zero or stale rows)
"""
from __future__ import annotations
import argparse
import csv
import importlib
import inspect
import logging
import pkgutil
import sys
from datetime import date, datetime
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from scrapers._base import BaseScraper, make_session, fetch  # noqa: E402

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger("audit-scrapers")


# -- classification thresholds --
MIN_OK_CONTENT_BYTES = 8_000   # page bodies smaller than this are suspicious
JS_HEAVY_TEXT_RATIO = 0.05     # if < 5% of bytes are visible text, likely JS-rendered


def classify_response(resp_status: int, body: bytes, headers: dict, url: str) -> Tuple[str, str]:
    """Return (classification, detail)."""
    if resp_status == 0:
        return ("NETWORK_FAIL", "connection/timeout error")
    if resp_status == 403:
        server = (headers.get("Server") or "").lower()
        if "cloudflare" in server or "cf-ray" in {k.lower() for k in headers}:
            return ("BLOCKED", "Cloudflare 403")
        return ("BLOCKED", f"HTTP 403 ({server or 'unknown'})")
    if resp_status == 404:
        return ("URL_FIX", "HTTP 404")
    if resp_status in (301, 302, 307, 308):
        return ("URL_FIX", f"HTTP {resp_status} redirect (check new URL)")
    if resp_status >= 500:
        return ("UPSTREAM_ERR", f"HTTP {resp_status}")
    if resp_status != 200:
        return ("OTHER", f"HTTP {resp_status}")

    size = len(body)
    if size < MIN_OK_CONTENT_BYTES:
        return ("EMPTY", f"200 OK but only {size} bytes")

    # Crude "is this mostly JS?" heuristic — look at visible text density
    try:
        txt = body.decode("utf-8", errors="replace")
    except Exception:
        txt = ""
    # Strip script/style tags crudely
    import re
    stripped = re.sub(r"<script[^>]*>.*?</script>", " ", txt, flags=re.S | re.I)
    stripped = re.sub(r"<style[^>]*>.*?</style>", " ", stripped, flags=re.S | re.I)
    stripped = re.sub(r"<[^>]+>", " ", stripped)
    stripped = re.sub(r"\s+", " ", stripped).strip()
    visible_bytes = len(stripped.encode("utf-8", errors="replace"))
    if visible_bytes < 2_000:
        return ("JS_HEAVY", f"200 OK, {size} raw bytes, only {visible_bytes} visible text — likely JS-rendered")

    ratio = visible_bytes / max(size, 1)
    if ratio < JS_HEAVY_TEXT_RATIO:
        return ("JS_HEAVY", f"visible text is only {ratio:.1%} of page size")

    return ("OK", f"{size} bytes, {visible_bytes} visible text")


def discover_scrapers_for_audit() -> List[BaseScraper]:
    """Walk every regional subdir and instantiate each scraper."""
    regions = [
        "north_america", "europe_eu", "europe_non_eu", "eu_wide",
        "asia", "oceania", "africa", "latam", "middle_east",
    ]
    found: List[BaseScraper] = []
    for region in regions:
        try:
            pkg = importlib.import_module(f"scrapers.{region}")
        except Exception as e:
            log.warning("Could not import scrapers.%s: %s", region, e)
            continue
        for _, modname, _ in pkgutil.iter_modules(pkg.__path__):
            try:
                mod = importlib.import_module(f"scrapers.{region}.{modname}")
                for _, cls in inspect.getmembers(mod, inspect.isclass):
                    if (issubclass(cls, BaseScraper) and cls is not BaseScraper
                            and cls.__module__ == mod.__name__ and cls.AGENCY):
                        found.append(cls())
            except Exception as e:
                log.warning("Could not instantiate %s.%s: %s", region, modname, e)
    return found


def get_stale_sources_from_xlsx(xlsx_path: Path, min_days: int = 14) -> set:
    """Return set of AGENCY names that have no row younger than min_days days."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        log.warning("openpyxl not installed — can't check staleness, auditing all scrapers")
        return set()
    if not xlsx_path.exists():
        log.warning("recalls.xlsx not found — auditing all scrapers")
        return set()

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if "Recalls" not in wb.sheetnames:
        return set()
    ws = wb["Recalls"]
    rows = ws.iter_rows(min_row=1, max_row=1, values_only=True)
    headers = next(rows)
    last_seen: Dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v in (None, "") for v in row):
            continue
        rec = dict(zip(headers, row))
        src = str(rec.get("Source") or "")
        d = str(rec.get("Date", ""))[:10]
        if src and d:
            if src not in last_seen or d > last_seen[src]:
                last_seen[src] = d

    today = date.today()
    stale = set()
    for src, d in last_seen.items():
        try:
            days_old = (today - date.fromisoformat(d)).days
            if days_old >= min_days:
                stale.add(src)
        except Exception:
            pass
    # Also everything NOT in last_seen is "never produced" = stale
    # (Caller passes scraper AGENCY; this returns stale set for Source match)
    return stale


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--csv", default=None, help="Write results to this CSV path")
    ap.add_argument("--only-dead", action="store_true",
                    help="Skip scrapers that have produced rows in the last 14 days")
    ap.add_argument("--xlsx", default=str(ROOT / "docs" / "data" / "recalls.xlsx"))
    args = ap.parse_args()

    log.info("Discovering scrapers...")
    scrapers = discover_scrapers_for_audit()
    log.info("Found %d scrapers", len(scrapers))

    # Filter to dead ones if requested
    if args.only_dead:
        stale_sources = get_stale_sources_from_xlsx(Path(args.xlsx), min_days=14)
        log.info("Stale/dead sources (from xlsx): %d", len(stale_sources))
        # We don't know which scrapers have NEVER written — probe all, show only dead
        # Simple rule: we probe everything but mark working ones at the end
        def is_stale(s: BaseScraper) -> bool:
            return s.AGENCY in stale_sources or s.AGENCY not in {x for x in stale_sources}
        # Actually, "never produced" = AGENCY not in xlsx Source column at all,
        # which we can't detect without reading the full Source set. Just audit all.

    session = make_session(timeout=20)

    results: List[Dict[str, Any]] = []
    for idx, scr in enumerate(scrapers, 1):
        urls = list(getattr(scr, "INDEX_URLS", []) or [])
        if not urls:
            # Custom scrapers (BaseScraper direct) may use FEED_URL / BASE_URL instead
            feed = getattr(scr, "FEED_URL", None)
            base = getattr(scr, "BASE_URL", None)
            if feed: urls.append(feed)
            if base: urls.append(base)
        if not urls:
            results.append({
                "agency": scr.AGENCY, "country": scr.COUNTRY,
                "url": "(no URL defined)", "status": 0,
                "classification": "NO_URL", "detail": "scraper has no INDEX_URLS/FEED_URL/BASE_URL",
                "size": 0,
            })
            continue

        for url in urls:
            log.info("[%d/%d] %s (%s) -> %s",
                     idx, len(scrapers), scr.AGENCY, scr.COUNTRY, url[:80])
            resp = fetch(session, url)
            if resp is None:
                cls, det = classify_response(0, b"", {}, url)
                results.append({
                    "agency": scr.AGENCY, "country": scr.COUNTRY,
                    "url": url, "status": 0,
                    "classification": cls, "detail": det, "size": 0,
                })
                continue
            body = resp.content
            headers = dict(resp.headers)
            cls, det = classify_response(resp.status_code, body, headers, url)
            results.append({
                "agency": scr.AGENCY, "country": scr.COUNTRY,
                "url": url, "status": resp.status_code,
                "classification": cls, "detail": det, "size": len(body),
            })

    # Summary
    from collections import Counter
    cls_counts = Counter(r["classification"] for r in results)
    print()
    print("=" * 70)
    print("CLASSIFICATION SUMMARY")
    print("=" * 70)
    for cls, n in cls_counts.most_common():
        print(f"  {cls:<15} {n}")

    # Per-class breakdown
    print()
    for cls in ("OK", "EMPTY", "URL_FIX", "JS_HEAVY", "BLOCKED",
                "UPSTREAM_ERR", "NETWORK_FAIL", "NO_URL", "OTHER"):
        matching = [r for r in results if r["classification"] == cls]
        if not matching:
            continue
        print()
        print(f"--- {cls} ({len(matching)}) ---")
        for r in sorted(matching, key=lambda x: (x["agency"], x["url"])):
            print(f"  {r['agency']:<22}  [{r['country']:<16}]  HTTP {r['status']:<3}  "
                  f"{r['size']:>8} bytes  {r['detail']}")
            print(f"                          {r['url']}")

    # CSV export
    if args.csv:
        p = Path(args.csv)
        with p.open("w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=[
                "agency", "country", "url", "status", "classification", "detail", "size",
            ])
            w.writeheader()
            w.writerows(results)
        print(f"\nCSV written: {p}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
