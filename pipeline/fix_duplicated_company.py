#!/usr/bin/env python3
"""
fix_duplicated_company.py  —  collapse concatenated Company names
==================================================================

RappelConso publishes two names per fiche: the legal entity and the commercial
banner. The scraper concatenates them, so the banner is repeated inside the
legal name:

    "CARREFOUR FRANCE CARREFOUR"          -> "CARREFOUR FRANCE"
    "MONOPRIX HOLDING MONOPRIX"           -> "MONOPRIX HOLDING"
    "STE LIONOR SA Lionor"                -> "STE LIONOR SA"
    "BCT 500 bct 500"                     -> "BCT 500"

THIS IS DELIBERATELY CONSERVATIVE. A naive "a token repeats" rule does real
damage, because several legitimate names repeat tokens:

    "Origin: France | Notifying: France"        <- the CORRECT RASFF format
    "Société ... des Producteurs ... des Vallées ..."   <- French grammar
    "SANULAC NUTRICIÓN MÉXICO S. de R.L. de C.V."      <- legal form

So a row is changed ONLY when the string splits cleanly into A + B where every
token of the suffix B already appears in the prefix A (accent- and
case-insensitive). Anything else is left exactly as it is.

RASFF rows are skipped outright.

Usage:
    python fix_duplicated_company.py --xlsx docs/data/recalls.xlsx --commit false
    python fix_duplicated_company.py --xlsx docs/data/recalls.xlsx --commit true
"""
from __future__ import annotations
import argparse
import datetime as dt
import re
import unicodedata
from pathlib import Path
from typing import Optional

# Tokens too common to prove duplication on their own.
_STOP = {"de", "des", "du", "la", "le", "les", "et", "d", "l", "sa", "sas",
         "sarl", "sasu", "eurl", "snc", "scea", "ste", "societe", "société",
         "the", "of", "and", "s", "inc", "ltd", "llc", "gmbh", "bv", "nv",
         "spa", "srl", "ag", "co", "cv", "rl"}


def _norm(tok: str) -> str:
    t = unicodedata.normalize("NFKD", tok)
    t = "".join(c for c in t if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9]", "", t.lower())


# A collapse that leaves one of these dangling is worse than the original:
#   "ETABLISSEMENTS J. CHENE ET FILS Ets"   "... SALAISONS ... Maison"
_NEVER_TRAILING = _STOP | {"ets", "maison", "centre", "ferme", "cooperative",
                           "coopérative", "distribution", "groupe", "sca",
                           "gie", "scop", "ets."}

# Descriptive phrases are not company names. "Danone infant formula and
# follow-on formula" repeats "formula" legitimately; truncating it corrupts a
# real value. Lower-case connectors are the tell.
_PHRASE = (" and ", " or ", " with ", " containing ", " for ", " from ")


def collapse(value: str) -> Optional[str]:
    """Return the corrected Company, or None to leave the row untouched.

    STRICT RULE: the trailing banner must be an EXACT repeat of the opening
    tokens (accent- and case-insensitive). Nothing else is changed.

        "CARREFOUR FRANCE CARREFOUR"    suffix == first 1 token   -> collapse
        "DES PERCE NEIGES GAEC DES PERCE NEIGES"  == first 3      -> collapse
        "BCT 500 bct 500"               == first 2                -> collapse

    Looser heuristics were tried and rejected: matching any repeated token
    truncated "Danone infant formula and follow-on formula", and matching a
    suffix contained anywhere in the prefix left dangling fragments such as
    "ETABLISSEMENTS J. CHENE ET FILS Ets" and "AUX SAVEURS DE LA FERME SARL
    Aux". A partial collapse is worse than the original, so this rule only
    fires when the duplication is provable.

    Some genuine duplicates are therefore left alone (e.g. "FROMAGERIE JACQUES
    DELIN FROMAGERIE DELIN", where the banner is not a clean repeat). That is
    the intended trade: no corrupted names.
    """
    v = str(value or "").strip()
    if not v or "|" in v or v.lower().startswith("origin:"):
        return None                      # RASFF format — never touch
    toks = v.split()
    if len(toks) < 2:
        return None

    norm = [_norm(x) for x in toks]
    n = len(toks)
    # Longest repeat first, so the whole banner goes in one step.
    for k in range(n // 2, 0, -1):
        head, tail = norm[:k], norm[n - k:]
        if any(not x for x in head + tail):
            continue
        if head != tail:
            continue
        if not any(x not in _STOP for x in tail):
            continue                     # only legal-form particles — ignore
        prefix_toks = toks[:n - k]
        candidate = " ".join(prefix_toks).strip(" -,")
        if not candidate or candidate == v:
            continue
        if _norm(prefix_toks[-1]) in _STOP:
            continue                     # would end on a dangling particle
        # If both halves are the same name, keep the better-typeset one.
        suffix_str = " ".join(toks[n - k:]).strip(" -,")
        if k == n - k and any(ord(c) > 127 for c in suffix_str) \
                and not any(ord(c) > 127 for c in candidate):
            return suffix_str
        return candidate
    return None


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=Path, default=Path("docs/data/recalls.xlsx"))
    ap.add_argument("--commit", type=str, default="false")
    args = ap.parse_args()
    commit = args.commit.lower() in ("1", "true", "yes", "on")

    import openpyxl
    wb = openpyxl.load_workbook(args.xlsx)
    changes = []
    for sheet in ("Recalls", "Pending", "Weekly_Review", "Weekly_Rejected"):
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        headers = [c.value for c in ws[1]]
        if "Company" not in headers:
            continue
        ci = headers.index("Company") + 1
        ni = headers.index("Notes") + 1 if "Notes" in headers else None
        for ridx in range(2, ws.max_row + 1):
            cur = ws.cell(ridx, ci).value
            new = collapse(cur)
            if new:
                changes.append((sheet, ridx, ci, ni, str(cur), new))

    print(f"Company fields to collapse: {len(changes)}\n")
    seen = {}
    for _s, _r, _c, _n, old, new in changes:
        seen.setdefault((old, new), 0)
        seen[(old, new)] += 1
    for (old, new), n in sorted(seen.items(), key=lambda x: -x[1]):
        print(f"  {n:3d}x  {old}")
        print(f"        -> {new}")

    if not commit:
        print("\nDRY RUN — nothing written. Re-run with --commit true.")
        return 0
    if not changes:
        return 0

    today = dt.datetime.now(dt.timezone.utc).strftime("%Y-%m-%d")
    for sheet, ridx, ci, ni, old, new in changes:
        ws = wb[sheet]
        ws.cell(ridx, ci).value = new
        if ni:
            prev = str(ws.cell(ridx, ni).value or "")
            ws.cell(ridx, ni).value = (
                prev + f" [company-fix {today}: {old!r} → {new!r} "
                f"(scraper concatenated legal entity + banner)]"
            ).strip()[:2000]
    wb.save(args.xlsx)
    print(f"\n✓ Corrected {len(changes)} Company field(s).")
    try:
        import sys
        sys.path.insert(0, ".")
        from pipeline.merge_master import mirror_json_from_xlsx
        mirror_json_from_xlsx(args.xlsx, args.xlsx.parent / "recalls.json")
        print("✓ recalls.json mirrored.")
    except Exception as e:
        print(f"  (JSON mirror skipped: {e})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
