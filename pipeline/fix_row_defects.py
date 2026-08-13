#!/usr/bin/env python3
"""
fix_row_defects.py  —  targeted data corrections (2026-08-13)
==============================================================

Two defects, both verified against the register before writing anything.

1. RappelConso fiche 22205 — TRANSPLANTED REASON
   The row is a French cooked-shrimp recall:
       Product : Crevette cuite réfrigérée 40/60 P. Vannamei ... (Lot 180421)
       Company : K.S.P. (Kruz Seafood Production)
       Pathogen: Listeria monocytogenes
   but its Reason reads:
       "Aflatoxins in mini corn wafers from Slovakia, raw material from
        Hungary.; risk: serious; category: cereals..."
   That text is a RASFF notification about a completely different product. It
   is the only RappelConso row in the register carrying RASFF-format Reason
   text, so this is a one-off cross-contamination, not a systematic pattern.

   The Product, Company, Pathogen, Date and URL are internally consistent and
   are LEFT ALONE. Only the alien Reason is replaced, with the hazard the row
   itself states.

2. Brand = a French phrase meaning "no brand" — 12 RappelConso rows
       'Sans marque' / 'SANS MARQUE' / 'Sans Marque' / 'aucune'
   The register's convention for RappelConso is "Unbranded" (117 rows already
   use it). These 12 are the same meaning in the source language.

   NOT TOUCHED: the em dash '—', which is the RASFF convention (186 of its 207
   uses are RASFF rows) and is correct by design, and every non-RappelConso
   source, whose conventions differ.

Usage:
    python fix_row_defects.py --xlsx docs/data/recalls.xlsx --commit false
    python fix_row_defects.py --xlsx docs/data/recalls.xlsx --commit true
"""
from __future__ import annotations
import argparse
import datetime as dt
from pathlib import Path

FICHE_22205 = "https://rappel.conso.gouv.fr/fiche-rappel/22205/Interne"
NEW_REASON_22205 = "Presence of Listeria monocytogenes"

NO_BRAND_FR = {"sans marque", "aucune", "aucun"}


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=Path, default=Path("docs/data/recalls.xlsx"))
    ap.add_argument("--commit", type=str, default="false")
    args = ap.parse_args()
    commit = args.commit.lower() in ("1", "true", "yes", "on")

    import openpyxl
    wb = openpyxl.load_workbook(args.xlsx)
    today = dt.datetime.now(dt.timezone.utc).strftime("%Y-%m-%d")
    reason_fix, brand_fix = [], []

    for sheet in ("Recalls", "Pending", "Weekly_Review", "Weekly_Rejected"):
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        headers = [c.value for c in ws[1]]
        if "URL" not in headers:
            continue
        ui = headers.index("URL") + 1
        ri = headers.index("Reason") + 1 if "Reason" in headers else None
        bi = headers.index("Brand") + 1 if "Brand" in headers else None
        si = headers.index("Source") + 1 if "Source" in headers else None
        ni = headers.index("Notes") + 1 if "Notes" in headers else None

        for ridx in range(2, ws.max_row + 1):
            url = str(ws.cell(ridx, ui).value or "").strip()

            if ri and url == FICHE_22205:
                cur = str(ws.cell(ridx, ri).value or "")
                if "Aflatoxin" in cur:
                    reason_fix.append((sheet, ridx, ri, ni, cur))

            if bi and si:
                brand = str(ws.cell(ridx, bi).value or "").strip()
                src = str(ws.cell(ridx, si).value or "").lower()
                if brand.lower() in NO_BRAND_FR and "rappelconso" in src:
                    brand_fix.append((sheet, ridx, bi, ni, brand))

    print(f"1. Fiche 22205 transplanted Reason : {len(reason_fix)} row(s)")
    for _s, _r, _c, _n, cur in reason_fix:
        print(f"     {cur[:72]}")
        print(f"  -> {NEW_REASON_22205}")
    print(f"\n2. RappelConso Brand meaning 'no brand' : {len(brand_fix)} row(s)")
    seen = {}
    for _s, _r, _c, _n, b in brand_fix:
        seen[b] = seen.get(b, 0) + 1
    for b, n in seen.items():
        print(f"     {n:3d}x  {b!r}  ->  'Unbranded'")

    if not commit:
        print("\nDRY RUN — nothing written. Re-run with --commit true.")
        return 0
    if not reason_fix and not brand_fix:
        print("Nothing to change.")
        return 0

    for sheet, ridx, ci, ni, cur in reason_fix:
        ws = wb[sheet]
        ws.cell(ridx, ci).value = NEW_REASON_22205
        if ni:
            prev = str(ws.cell(ridx, ni).value or "")
            ws.cell(ridx, ni).value = (
                prev + f" [row-fix {today}: Reason replaced — the stored text "
                f"was a RASFF aflatoxin notice about corn wafers, unrelated to "
                f"this shrimp recall; set to the hazard the row itself states]"
            ).strip()[:2000]

    for sheet, ridx, ci, ni, old in brand_fix:
        ws = wb[sheet]
        ws.cell(ridx, ci).value = "Unbranded"
        if ni:
            prev = str(ws.cell(ridx, ni).value or "")
            ws.cell(ridx, ni).value = (
                prev + f" [row-fix {today}: Brand {old!r} → 'Unbranded' "
                f"(register convention for RappelConso)]").strip()[:2000]

    wb.save(args.xlsx)
    print(f"\n✓ Fixed {len(reason_fix)} Reason and {len(brand_fix)} Brand field(s).")
    try:
        import sys
        sys.path.insert(0, ".")
        from pipeline.merge_master import mirror_json_from_xlsx
        mirror_json_from_xlsx(args.xlsx, args.xlsx.parent / "recalls.json")
        print("✓ recalls.json mirrored.")
    except Exception as e:
        print(f"  (JSON mirror skipped: {e})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
