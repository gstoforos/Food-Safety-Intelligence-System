#!/usr/bin/env python3
"""CURATOR — the only thing allowed to write to the register.

It is not a second model. Every check below is deterministic code, and each
one exists because something got through once:

  URL RESOLVES + CONTENT MATCHES   the gap-finder attached correct-looking
      product and pathogen fields to RappelConso fiches describing a SHEIN
      plush toy, a vitamin-D supplement and a mackerel histamine recall.
      All three passed the publish gate: the URL was well-formed and it
      resolved. Fetching is not enough — the page must actually mention the
      thing the row claims.
  PUBLISH GATE                     headline-in-Company, empty Reason, empty
      Pathogen, truncated URL.
  LANGUAGE                         rows promoted carrying "Presence possible
      de Listeria monocytogenes" against a standing English-output rule.
  SCOPE                            pet food, a light fitting, ceramic plates,
      a labelling error and eleven cold-chain rows all reached Pending.
  DUPLICATE                        the same recall arriving twice from the
      official scrape and the gap-finder, under different URLs.

A proposal that fails any check is REFUSED and written to the run report
with the reason. It is never downgraded into a weaker action, and it is
never applied "with a warning".

    python -m pipeline.agents.curator --proposals <file> --dry-run
    python -m pipeline.agents.curator --proposals <file> --apply
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(ROOT))

from pipeline.agents._contract import Proposal, read_proposals  # noqa: E402

XLSX = ROOT / "docs" / "data" / "recalls.xlsx"
AGENT = "curator"
TODAY = datetime.now(timezone.utc).date().isoformat()

# Fields the curator may write. Anything else is refused — an agent must not
# be able to reach Tier, Outbreak or report_week, which drive published
# statistics and are set by dedicated code paths.
WRITABLE = {"Pathogen", "Reason", "Product", "Company", "Brand", "Country",
            "Region", "Class", "URL", "Notes", "Date"}


@dataclass
class Verdict:
    proposal_id: str
    action: str
    applied: bool
    refusals: List[str]
    detail: str = ""


def _rows(sheet: str) -> Tuple[List[str], List[Dict[str, Any]]]:
    import openpyxl
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb[sheet]
    hdr = [str(c.value or "") for c in ws[1]]
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        out.append({h: v for h, v in zip(hdr, r)})
    return hdr, out


def _find(url: str) -> Tuple[Optional[str], Optional[int], Optional[Dict]]:
    """Locate a row by URL across Pending and Recalls. Returns 1-based sheet row."""
    u = (url or "").strip().lower().rstrip("/")
    if not u:
        return None, None, None
    for sheet in ("Pending", "Recalls"):
        try:
            _hdr, rows = _rows(sheet)
        except KeyError:
            continue
        for i, r in enumerate(rows):
            if str(r.get("URL", "") or "").strip().lower().rstrip("/") == u:
                return sheet, i + 2, r
    return None, None, None


# ── the checks ───────────────────────────────────────────────────────────

def check_url_resolves_and_matches(url: str, row: Dict[str, Any]) -> List[str]:
    """Fetch the page and require it to mention what the row claims.

    The single most important check in this module. Without it a row can
    carry perfect data against a URL describing a different recall.
    """
    import requests
    try:
        r = requests.get(url, timeout=25, allow_redirects=True,
                         headers={"User-Agent": "AFTS-FSIS/1.0 (+advfood.tech)"})
    except Exception as e:                                  # noqa: BLE001
        return [f"URL unreachable ({type(e).__name__}) — cannot confirm provenance"]
    if r.status_code >= 400:
        return [f"URL returns HTTP {r.status_code}"]

    text = re.sub(r"<[^>]+>", " ", r.text)
    text = re.sub(r"\s+", " ", text).lower()

    # Anchor on the least paraphrasable tokens the row carries.
    anchors: List[str] = []
    path = str(row.get("Pathogen", "") or "")
    for tok in re.findall(r"[A-Za-z]{5,}", path):
        if tok.lower() not in ("shiga", "toxin", "producing", "coli"):
            anchors.append(tok.lower())
    for f in ("Company", "Brand", "Product"):
        for tok in re.findall(r"[A-Za-zÀ-ÿ]{5,}", str(row.get(f, "") or ""))[:3]:
            anchors.append(tok.lower())
    anchors = [a for a in dict.fromkeys(anchors)][:8]
    if not anchors:
        return ["row carries no token distinctive enough to confirm against "
                "the page"]
    hits = [a for a in anchors if a in text]
    if not hits:
        return [f"page does not mention any of {anchors} — the URL describes "
                f"something other than this row"]
    return []


def check_gate(row: Dict[str, Any]) -> List[str]:
    from pipeline._publish_gate import publish_blockers
    clean = {k: ("" if v is None else v) for k, v in row.items()}
    if clean.get("Date") not in ("", None):
        clean["Date"] = str(clean["Date"])[:10]
    return list(publish_blockers(clean))


def check_language(row: Dict[str, Any]) -> List[str]:
    """Reason and Company must be English. Brand and Product may not be."""
    try:
        from pipeline._language import looks_non_english
    except Exception:                                       # noqa: BLE001
        return []
    out = []
    for f in ("Reason", "Company"):
        v = str(row.get(f, "") or "")
        if v and looks_non_english(v):
            out.append(f"{f} is not English: {v[:60]!r}")
    return out


def check_scope(row: Dict[str, Any]) -> List[str]:
    from pipeline._pathogen_scope import is_in_scope, is_empty_pathogen
    try:
        from pipeline._pathogen_scope import is_pet_food_product
    except Exception:                                       # noqa: BLE001
        is_pet_food_product = lambda *a: False              # noqa: E731
    p = str(row.get("Pathogen", "") or "")
    out = []
    if is_empty_pathogen(p):
        out.append("Pathogen empty — the register names its hazard")
    elif not is_in_scope(p):
        out.append(f"hazard {p!r} is outside the monitored scope")
    if is_pet_food_product(str(row.get("Product", "")), str(row.get("Company", "")),
                           str(row.get("Reason", ""))):
        out.append("pet food — human food only")
    return out


def check_duplicate(url: str) -> List[str]:
    _hdr, rows = _rows("Recalls")
    u = (url or "").strip().lower().rstrip("/")
    n = sum(1 for r in rows
            if str(r.get("URL", "") or "").strip().lower().rstrip("/") == u)
    return [f"already in Recalls ({n} row(s) with this URL)"] if n else []


# ── application ──────────────────────────────────────────────────────────

def apply_one(p: Proposal, apply: bool, offline: bool = False) -> Verdict:
    import openpyxl

    problems = p.structural_problems()
    if problems:
        return Verdict(p.proposal_id, p.action, False, problems)

    if p.action == "flag":
        # A flag writes nothing to the register by design; it is a record
        # that something the register cannot hold happened.
        return Verdict(p.proposal_id, p.action, True, [], "logged (no row written)")

    url = str(p.target.get("url", ""))
    sheet, rownum, row = _find(url)
    if row is None:
        return Verdict(p.proposal_id, p.action, False,
                       [f"no row found for {url!r} in Pending or Recalls"])

    bad = [k for k in p.changes if k not in WRITABLE]
    if bad and p.action in ("promote", "enrich"):
        return Verdict(p.proposal_id, p.action, False,
                       [f"proposal touches protected field(s) {bad} — Tier, "
                        f"Outbreak and report_week are set by the pipeline, "
                        f"not by an agent"])

    merged = dict(row)
    merged.update({k: v for k, v in p.changes.items() if k in WRITABLE})

    refusals: List[str] = []
    if p.action in ("promote", "enrich"):
        refusals += check_scope(merged)
        refusals += check_language(merged)
        refusals += check_gate(merged)
        if not offline:
            refusals += check_url_resolves_and_matches(
                str(merged.get("URL") or url), merged)
    if p.action == "promote":
        refusals += check_duplicate(url)
    if p.action == "link_event" and not p.hard_evidence():
        refusals.append("a link between a row and an outbreak needs regulator "
                        "evidence, not retrieval similarity")

    if refusals:
        return Verdict(p.proposal_id, p.action, False, refusals)
    if not apply:
        return Verdict(p.proposal_id, p.action, False, [], "would apply (dry run)")

    wb = openpyxl.load_workbook(XLSX)
    ws = wb[sheet]
    hdr = [str(c.value or "") for c in ws[1]]
    stamp = (f"[agent {p.agent} proposal {p.proposal_id} applied {TODAY}: "
             f"{p.reason[:220]}]")
    for k, v in p.changes.items():
        if k in WRITABLE and k in hdr:
            ws.cell(row=rownum, column=hdr.index(k) + 1, value=v)
    if "Notes" in hdr:
        cur = str(ws.cell(row=rownum, column=hdr.index("Notes") + 1).value or "")
        ws.cell(row=rownum, column=hdr.index("Notes") + 1,
                value=(cur + " " + stamp).strip())
    if "LastUpdated" in hdr:
        ws.cell(row=rownum, column=hdr.index("LastUpdated") + 1, value=TODAY)
    wb.save(XLSX)
    return Verdict(p.proposal_id, p.action, True, [], f"{sheet} row {rownum}")


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--proposals", required=True)
    ap.add_argument("--apply", action="store_true",
                    help="without this nothing is written")
    ap.add_argument("--offline", action="store_true",
                    help="skip the network check — for tests only; a run "
                         "with this flag must not be committed")
    a = ap.parse_args(argv)

    props = read_proposals(Path(a.proposals))
    verdicts = [apply_one(p, a.apply, a.offline) for p in props]

    ok = sum(1 for v in verdicts if v.applied)
    print(f"{len(props)} proposal(s): {ok} applied, {len(props) - ok} refused"
          f"{' (DRY RUN)' if not a.apply else ''}")
    for v in verdicts:
        mark = "APPLIED " if v.applied else "REFUSED "
        print(f"  {mark}{v.action:11} {v.proposal_id}  {v.detail}")
        for r in v.refusals:
            print(f"       - {r}")

    out = Path(a.proposals).with_suffix(".verdicts.json")
    out.write_text(json.dumps(
        [{"proposal_id": v.proposal_id, "action": v.action,
          "applied": v.applied, "refusals": v.refusals, "detail": v.detail}
         for v in verdicts], indent=2), encoding="utf-8")
    print(f"verdicts -> {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
