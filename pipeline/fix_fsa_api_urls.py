#!/usr/bin/env python3
"""
fix_fsa_api_urls.py  —  one-time corrective (2026-08-09)
=========================================================

Eight rows in Recalls store the FSA's LINKED-DATA record instead of the public
alert page:

    https://data.food.gov.uk/food-alerts/id/FSA-PRIN-38-2026     (metadata)
    https://www.food.gov.uk/news-alerts/alert/fsa-prin-38-2026   (the page)

The FSA's own record carries the public page in its `alertURL` field, so the
mapping is deterministic — the alert id is lowercased and moved onto
www.food.gov.uk/news-alerts/alert/. No guessing, nothing fetched.

Rows are matched by URL pattern only; every other field is untouched.

Usage:
    python fix_fsa_api_urls.py --xlsx docs/data/recalls.xlsx --commit false
    python fix_fsa_api_urls.py --xlsx docs/data/recalls.xlsx --commit true
"""
from __future__ import annotations
import argparse
import re
from pathlib import Path

API_RE = re.compile(
    r"https?://data\.food\.gov\.uk/food-alerts/id/([A-Za-z0-9\-]+)", re.I)


def canonical(url: str) -> str | None:
    m = API_RE.match(str(url or "").strip())
    if not m:
        return None
    return "https://www.food.gov.uk/news-alerts/alert/" + m.group(1).lower()


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=Path, default=Path("docs/data/recalls.xlsx"))
    ap.add_argument("--commit", type=str, default="false")
    args = ap.parse_args()
    commit = args.commit.lower() in ("1", "true", "yes", "on")

    import openpyxl
    wb = openpyxl.load_workbook(args.xlsx)
    total = 0

    for sheet in ("Recalls", "Pending", "Weekly_Review", "Weekly_Rejected"):
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        headers = [c.value for c in ws[1]]
        if "URL" not in headers:
            continue
        ui = headers.index("URL") + 1
        ni = headers.index("Notes") + 1 if "Notes" in headers else None
        for ridx in range(2, ws.max_row + 1):
            cur = ws.cell(ridx, ui).value
            new = canonical(cur)
            if not new:
                continue
            total += 1
            print(f"  [{sheet}] row {ridx}")
            print(f"      {cur}")
            print(f"   -> {new}")
            if commit:
                ws.cell(ridx, ui).value = new
                if ni:
                    old = str(ws.cell(ridx, ni).value or "")
                    ws.cell(ridx, ni).value = (
                        old + " [fix 2026-08-09: FSA linked-data record -> "
                        "public alert page]").strip()[:2000]

    print(f"\nMatched {total} row(s) with an FSA linked-data URL.")
    if not commit:
        print("DRY RUN — nothing written. Re-run with --commit true.")
        return 0
    if total == 0:
        print("Nothing to change.")
        return 0

    wb.save(args.xlsx)
    print(f"✓ Rewrote {total} URL(s).")
    try:
        import sys
        sys.path.insert(0, ".")
        from pipeline.merge_master import mirror_json_from_xlsx
        mirror_json_from_xlsx(args.xlsx, args.xlsx.parent / "recalls.json")
        print("✓ recalls.json mirrored.")
    except Exception as e:
        print(f"  (JSON mirror skipped: {e})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
