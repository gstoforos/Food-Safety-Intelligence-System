"""
AFTS FSIS — 7-day rolling retention for the NEWS sheet.

Opens docs/data/recalls.xlsx, finds the 'NEWS' sheet, and deletes every
row whose 'Published (UTC)' date is more than 7 days old. Keeps the
Recalls + Pending sheets untouched.

George's rule (memory): news feed runs every 1h, NEWS sheet keeps only
the rolling 7-day window. This script is called from the news-feed
workflow immediately after the fetch step so expired rows never sit
around for long.

Usage:
    python -m pipeline.purge_old_news --xlsx docs/data/recalls.xlsx
    python -m pipeline.purge_old_news --xlsx ... --days 7 --dry-run
"""
from __future__ import annotations

import argparse
import logging
import re
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path

import openpyxl   # type: ignore

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
log = logging.getLogger("purge_old_news")

NEWS_SHEET = "NEWS"
PUBLISHED_COL = "Published (UTC)"


def _parse_dt(val) -> datetime | None:
    """NEWS sheet stores ISO-ish strings like '2026-04-17 05:37 UTC' or
    raw datetimes. Accept both."""
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val if val.tzinfo else val.replace(tzinfo=timezone.utc)
    s = str(val).strip()
    # Strip trailing "UTC" if present, keep space before it.
    s = re.sub(r"\s*UTC\s*$", "", s)
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%dT%H:%M:%SZ",
        "%Y-%m-%dT%H:%M:%S.%f",
        "%Y-%m-%d",
    ):
        try:
            dt = datetime.strptime(s, fmt)
            return dt.replace(tzinfo=timezone.utc)
        except ValueError:
            pass
    log.warning("Could not parse 'Published (UTC)' value: %r", val)
    return None


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--days", type=int, default=7,
                    help="Rolling retention window in days (default 7)")
    ap.add_argument("--dry-run", action="store_true",
                    help="Report what would be removed, don't write")
    args = ap.parse_args()

    xlsx = Path(args.xlsx)
    if not xlsx.exists():
        log.error("xlsx not found: %s", xlsx); return 2

    wb = openpyxl.load_workbook(str(xlsx))
    if NEWS_SHEET not in wb.sheetnames:
        log.info("Workbook has no %r sheet — nothing to purge.", NEWS_SHEET)
        return 0

    ws = wb[NEWS_SHEET]
    if ws.max_row < 2:
        log.info("NEWS sheet has no data rows — nothing to purge.")
        return 0

    # Find the Published column by header name.
    headers = [str(c.value or "").strip() for c in ws[1]]
    try:
        col_idx = headers.index(PUBLISHED_COL) + 1   # 1-based
    except ValueError:
        log.error("NEWS sheet missing column %r. Headers: %s",
                  PUBLISHED_COL, headers)
        return 2

    cutoff = datetime.now(timezone.utc) - timedelta(days=args.days)
    log.info("Purging NEWS rows older than %s (%d-day window).",
             cutoff.isoformat(), args.days)

    # Walk rows bottom-up so row deletion doesn't shift indexes under us.
    to_delete = []
    for row_num in range(ws.max_row, 1, -1):
        val = ws.cell(row=row_num, column=col_idx).value
        dt = _parse_dt(val)
        if dt is None:
            # Safer to keep unparseable rows than silently drop them.
            log.warning("Row %d: unparseable date %r — keeping.", row_num, val)
            continue
        if dt < cutoff:
            to_delete.append(row_num)

    log.info("Will purge %d row(s) of %d.", len(to_delete), ws.max_row - 1)

    if args.dry_run:
        for r in to_delete[:10]:
            log.info("  (dry-run) would delete row %d: %s", r,
                     [c.value for c in ws[r]])
        if len(to_delete) > 10:
            log.info("  ... and %d more.", len(to_delete) - 10)
        return 0

    for row_num in to_delete:
        ws.delete_rows(row_num, 1)

    wb.save(str(xlsx))
    log.info("Saved %s — NEWS now has %d data row(s).",
             xlsx, ws.max_row - 1)
    return 0


if __name__ == "__main__":
    sys.exit(main())
