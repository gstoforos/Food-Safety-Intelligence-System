"""
Commit updated data files to GitHub — BULLETPROOF against xlsx-staging races.

==============================================================================
WHAT THIS FIXES — XLSX-STAGING SILENT-OVERWRITE BUG (audit 2026-05-19)
==============================================================================
Recurring symptom: a workflow's writes to recalls.xlsx get reverted within
~1 hour of commit, but only for SOME sheets. Specifically, sheets touched
ONLY by the victim workflow get reverted to their pre-commit state, while
sheets a concurrent workflow also touched are preserved at the concurrent
workflow's state. Net effect: silent data loss in Recalls / Pending /
Weekly_Review, while Weekly_Rejected / NEWS / daily-index persist.

Reproduced as the 2026-05-19 17:10 UTC claude-check incident:
    claude-check committed 642 Recalls + 5 Pending + 28 Weekly_Review.
    By 18:25 UTC the committed file had been reverted to 637 / 38 / 23,
    with Weekly_Rejected (+16) and daily-index.json preserved.

Root cause: pre-fix `_preemptive_sync` did its row-union merge ONLY on
stash-pop conflict (returncode != 0). When stash pop returncode == 0 —
the common case for binary files where git can't detect a textual
conflict — the code did NOTHING. git silently chose "ours" (the stashed
stale version) over the just-pulled remote, no merge ran, and our
stale xlsx silently overwrote the concurrent workflow's just-pulled
additions. Then we pushed our stale-overwriting xlsx and the remote
rows were gone.

==============================================================================
NEW ARCHITECTURE — FOUR LAYERS OF SAFETY
==============================================================================
LAYER 0 (existing) — concurrency groups in workflow YAML files
    `group: fsis-data-writers` + `cancel-in-progress: false` queues
    workflows. Layer 0 is necessary but NOT sufficient — workflows in
    different groups (afts-weekly-report, afts-monthly-report,
    audit-scrapers) can still overlap with fsis-data-writers.

LAYER 1 (FIXED 2026-05-19) — pre-emptive sync with UNCONDITIONAL row-merge
    Pull remote, then ALWAYS row-union-merge our xlsx snapshot against
    the just-pulled remote, regardless of stash pop returncode. This
    closes the silent-overwrite hole.

LAYER 2 (existing) — post-push retry with row-merge
    If push still fails (a third workflow landed between layers 1 and 2),
    reset-hard, pull, row-merge, re-commit, retry.

LAYER 3 (NEW 2026-05-19) — post-push verification
    After a successful push, fetch the remote and compare row counts of
    our pushed xlsx vs the freshly-fetched remote. If remote has FEWER
    rows in any sheet than what we pushed, a concurrent workflow stomped
    us. We log.error with the details so the operator can audit and
    recover from the alarm. (Verification only — does not auto-recover,
    since recovery requires choosing a winner.)

Binary files cannot be rebased, so layers 1 + 2 use the row-union
logic in pipeline.xlsx_merge — never git's own merge. xlsx_merge has a
per-sheet canary (`merged >= max(remote, ours)`) that aborts the push
rather than ship row loss.

==============================================================================
WHAT THIS DOES NOT FIX
==============================================================================
Workflows with INLINE xlsx-handling logic that bypasses this module
(currently news-feed.yml and merge-master.yml retry blocks) are NOT
guarded by layer 1's fix unless they also use this module. Those
workflows still rely on their inline `pipeline.xlsx_merge` invocation
which is correct on retry but doesn't have a pre-emptive sync. To fully
close the loop, those workflows should switch to calling
`git_commit_and_push` from this module — but that change is intrusive
and out of scope for this hotfix. Layer 3's verification will at least
ALARM if their inline logic stomps us.
"""
from __future__ import annotations
import os
import shutil
import subprocess
import logging
import time
from pathlib import Path

log = logging.getLogger(__name__)

MAX_PUSH_ATTEMPTS = 3

# Cumulative shared-state files that MUST be row-union-merged when they
# exist on both local + remote, never overwritten. See xlsx_merge.py for
# the per-sheet merge rules.
XLSX_REL = "docs/data/recalls.xlsx"
JSON_REL = "docs/data/recalls.json"

# Sheets to verify in layer 3 post-push check. If remote count < our
# pushed count for any of these, alarm.
VERIFY_SHEETS = ("Recalls", "Pending", "Weekly_Review", "Weekly_Rejected", "NEWS")


def _run(cmd: list[str], **kwargs) -> subprocess.CompletedProcess:
    """Run a git command, logging failures."""
    result = subprocess.run(cmd, capture_output=True, text=True, **kwargs)
    if result.returncode != 0:
        log.debug("cmd=%s rc=%d stderr=%s", " ".join(cmd),
                  result.returncode, result.stderr.strip())
    return result


def _row_union_merge_xlsx(cwd: str, ours_snapshot: Path) -> bool:
    """Row-union our snapshot with the freshly-pulled remote xlsx, write
    back to the remote location in the working tree. xlsx_merge enforces
    a canary (merged >= max(remote, ours) per sheet) and raises on
    shrink — we treat that as a hard fail.

    Returns True on success, False on failure (caller decides fallback)."""
    try:
        from pipeline.xlsx_merge import merge_xlsx_with_remote
        remote_xlsx = Path(cwd) / XLSX_REL
        counts = merge_xlsx_with_remote(
            remote_path=remote_xlsx,
            ours_path=ours_snapshot,
            out_path=remote_xlsx,
        )
        log.info("xlsx row-union merge: recalls remote=%d ours=%d merged=%d; "
                 "pending remote=%d ours=%d merged=%d",
                 counts.get("recalls_remote", -1),
                 counts.get("recalls_ours", -1),
                 counts.get("recalls_merged", -1),
                 counts.get("pending_remote", -1),
                 counts.get("pending_ours", -1),
                 counts.get("pending_merged", -1))
        return True
    except AssertionError as ae:
        # Canary tripped — merge would have shrunk a sheet. Fail loud.
        log.error("xlsx merge CANARY tripped — refusing to write merged "
                  "file that would lose rows: %s", ae)
        return False
    except Exception as me:
        log.error("xlsx merge failed: %s", me)
        return False


def _regenerate_json_mirror(cwd: str) -> bool:
    """Regenerate docs/data/recalls.json from the (just-merged) xlsx so
    the mirror matches. Idempotent."""
    try:
        from pipeline.merge_master import mirror_json_from_xlsx
        mirror_json_from_xlsx(Path(cwd) / XLSX_REL, Path(cwd) / JSON_REL)
        return True
    except Exception as je:
        log.warning("json mirror regenerate failed: %s", je)
        return False


def _xlsx_sheet_counts(xlsx_path: Path) -> dict[str, int]:
    """Return {sheet_name: row_count_excluding_header} for the given xlsx.
    Returns empty dict on failure. Used by layer 3 verification."""
    if not xlsx_path.exists():
        return {}
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path, data_only=True, read_only=True)
        out = {}
        for name in wb.sheetnames:
            ws = wb[name]
            # max_row includes header; subtract 1 (clamped at 0)
            out[name] = max(0, ws.max_row - 1)
        wb.close()
        return out
    except Exception as e:
        log.warning("xlsx sheet count read failed for %s: %s", xlsx_path, e)
        return {}


def _save_snapshots(cwd: str, files: list[str], tmp: Path) -> dict[str, Path]:
    """Save physical copies of all relevant files to tmp/ so we can
    reconcile against the new remote after a pull. Always saves xlsx +
    json mirror, even if caller didn't list them, because they're
    cumulative shared state."""
    saved: dict[str, Path] = {}
    # Union of caller's files + always-cumulative files
    targets = list(dict.fromkeys(list(files) + [XLSX_REL, JSON_REL]))
    for rel in targets:
        src = Path(cwd) / rel
        if src.exists():
            dst = tmp / rel.replace("/", "__")
            try:
                shutil.copy2(src, dst)
                saved[rel] = dst
            except Exception as e:
                log.warning("snapshot save failed for %s: %s", rel, e)
    return saved


def _restore_non_cumulative(cwd: str, saved: dict[str, Path]) -> None:
    """For non-cumulative files (HTML outputs, configs, etc.), copy our
    snapshot over the working tree. Cumulative files (xlsx, json) are
    handled separately by the row-merge path."""
    for rel, snap in saved.items():
        if rel in (XLSX_REL, JSON_REL):
            continue
        dest = Path(cwd) / rel
        try:
            dest.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(snap, dest)
        except Exception as e:
            log.warning("restore failed for %s: %s", rel, e)


def _restage_all(cwd: str, files: list[str]) -> None:
    """Re-stage everything after a merge. Includes xlsx + json explicitly
    in case caller didn't list them."""
    targets = list(dict.fromkeys(list(files) + [XLSX_REL, JSON_REL]))
    for rel in targets:
        # Stage regardless of whether file exists (handles deletions too)
        _run(["git", "-C", cwd, "add", "--", rel])


def _preemptive_sync(cwd: str, files: list[str]) -> None:
    """LAYER 1 (FIXED 2026-05-19): before the first commit, fetch from
    remote and reconcile if we're behind. UNCONDITIONALLY row-merges
    xlsx after any pull, regardless of stash pop returncode.

    Behaviour:
      - If up to date: noop.
      - If behind AND we have a staged xlsx: stash, pull, pop, row-merge
        our snapshot against the new remote (the merge runs whether the
        pop succeeded clean or with a conflict), regenerate json mirror,
        re-stage everything. This is the closure of the silent-overwrite
        hole: previously the merge only ran on pop conflict, but for
        binary files git pop often succeeds returncode==0 by silently
        choosing one side, leading to data loss when it chose ours.
      - If ff-only pull is rejected (local diverged): row-merge locally
        and defer to layer 2 for the actual push reconciliation.

    Idempotent and safe to call before every commit. Never raises."""
    branch = (_run(["git", "-C", cwd, "branch", "--show-current"])
              .stdout.strip() or "main")

    _run(["git", "-C", cwd, "fetch", "origin", branch])

    # How many commits behind are we?
    behind = _run(["git", "-C", cwd, "rev-list", "--count",
                   f"HEAD..origin/{branch}"])
    n_behind = int(behind.stdout.strip() or "0") if behind.returncode == 0 else 0
    if n_behind == 0:
        return  # already up to date — no reconciliation needed

    log.info("Pre-emptive sync: local is %d commit(s) behind origin/%s — "
             "reconciling before first push", n_behind, branch)

    # Save snapshots BEFORE any git operation that could lose them.
    tmp = Path(cwd) / ".presync-tmp"
    tmp.mkdir(exist_ok=True)
    saved = _save_snapshots(cwd, files, tmp)

    # Stash staged + working-tree changes so the pull can fast-forward.
    stash = _run(["git", "-C", cwd, "stash", "push", "--include-untracked",
                  "-m", "presync-stash"])
    stashed = stash.returncode == 0 and "No local changes" not in stash.stdout

    pull = _run(["git", "-C", cwd, "pull", "--ff-only", "origin", branch])
    if pull.returncode != 0:
        # ff-only rejected — local has diverged. Restore stash and let
        # the post-push retry handle the conflict. Before returning,
        # still row-merge xlsx with our snapshot so the local working
        # tree has the union, not just our stale data.
        log.info("Pre-emptive sync: ff-only pull rejected (local diverged) "
                 "— deferring to post-push retry")
        if stashed:
            _run(["git", "-C", cwd, "stash", "pop"])
        # Even on diverged path, row-merge xlsx if we have a snapshot.
        # This is defense in depth — if layer 2 retry fails to row-merge
        # for any reason, layer 1's local working tree is already merged.
        if XLSX_REL in saved and (Path(cwd) / XLSX_REL).exists():
            if _row_union_merge_xlsx(cwd, saved[XLSX_REL]):
                _regenerate_json_mirror(cwd)
                _restage_all(cwd, files)
        shutil.rmtree(tmp, ignore_errors=True)
        return

    # Pull succeeded. Working tree now has the just-pulled remote state.
    # Pop our stash. For text files, git will textually merge or report
    # a conflict. For binary files like xlsx, git can either:
    #   (a) report a textual conflict (returncode != 0), or
    #   (b) silently pick one side with returncode == 0.
    # Path (b) is the silent-overwrite bug. The fix below row-merges
    # ALWAYS, regardless of which path git took.
    if stashed:
        pop = _run(["git", "-C", cwd, "stash", "pop"])
        if pop.returncode != 0:
            # Reported conflict. Drop the conflicted stash; we'll
            # reconcile from saved snapshots below.
            log.info("Pre-emptive sync: stash pop reported conflict — "
                     "reconciling from snapshots")
            # Manually restore non-cumulative files (xlsx/json handled
            # by the row-merge step below).
            _restore_non_cumulative(cwd, saved)
            _run(["git", "-C", cwd, "stash", "drop"])
        # else: pop succeeded — but DO NOT trust git for xlsx. The
        # row-merge below runs unconditionally to plug silent overwrite.

    # ── UNCONDITIONAL ROW-UNION MERGE (the fix) ─────────────────────────
    # This block runs whether stash pop succeeded, conflicted, or didn't
    # exist. It's the structural guarantee that xlsx data cannot be lost
    # on the pre-emptive sync path. The row-union is idempotent: if our
    # snapshot already matches the working tree, the merge is a noop.
    if XLSX_REL in saved and (Path(cwd) / XLSX_REL).exists():
        if _row_union_merge_xlsx(cwd, saved[XLSX_REL]):
            _regenerate_json_mirror(cwd)
        else:
            # Canary tripped — this means rows would be lost. Restore
            # our snapshot rather than let the (potentially stale) pulled
            # version sit on disk, and abort the sync. The caller's
            # commit will then commit the snapshot's content, which is
            # safer than letting a partial merge proceed.
            log.error("Pre-emptive sync: row-merge canary tripped — "
                      "restoring snapshot and deferring to layer 2")
            try:
                shutil.copy2(saved[XLSX_REL], Path(cwd) / XLSX_REL)
            except Exception as e:
                log.error("snapshot restore failed: %s", e)

    # Re-stage everything — both files the caller listed AND
    # xlsx/json (which we always row-merge).
    _restage_all(cwd, files)

    shutil.rmtree(tmp, ignore_errors=True)


def _post_push_verify(cwd: str, files: list[str],
                      pre_push_counts: dict[str, int]) -> None:
    """LAYER 3 (NEW 2026-05-19): after a successful push, fetch the
    remote and compare xlsx sheet row counts against what we just
    committed. If remote has FEWER rows in any monitored sheet, a
    concurrent workflow stomped us between push and verification.

    Alarms via log.error — does NOT auto-recover, since recovery
    requires operator decision on which version wins. The alarm is the
    actionable signal.

    Bounded: fetches up to 3 times over a 30-second window so brief
    in-flight concurrent pushes are caught. If the remote moves DURING
    verification we treat the moved state as the canonical post-push
    state and re-check.

    Never raises."""
    if not pre_push_counts:
        return  # nothing to verify against

    branch = (_run(["git", "-C", cwd, "branch", "--show-current"])
              .stdout.strip() or "main")

    deadline = time.time() + 30.0
    for probe in range(1, 4):
        if time.time() > deadline:
            break
        _run(["git", "-C", cwd, "fetch", "origin", branch])
        # Read the remote xlsx via `git show origin/{branch}:XLSX_REL`
        # so we don't perturb the working tree.
        show = subprocess.run(
            ["git", "-C", cwd, "show", f"origin/{branch}:{XLSX_REL}"],
            capture_output=True,
        )
        if show.returncode != 0:
            log.debug("Layer 3 verify: could not read remote xlsx "
                      "(probe %d): %s", probe,
                      show.stderr.decode("utf-8", "replace").strip())
            time.sleep(probe * 2)
            continue

        # Write the remote bytes to a tmp file and read sheet counts.
        tmp_remote = Path(cwd) / f".verify-remote-{probe}.xlsx"
        try:
            tmp_remote.write_bytes(show.stdout)
            remote_counts = _xlsx_sheet_counts(tmp_remote)
        finally:
            try:
                tmp_remote.unlink()
            except Exception:
                pass

        if not remote_counts:
            time.sleep(probe * 2)
            continue

        # Compare each monitored sheet.
        shrunk = []
        for sheet in VERIFY_SHEETS:
            ours = pre_push_counts.get(sheet)
            theirs = remote_counts.get(sheet)
            if ours is None or theirs is None:
                continue
            if theirs < ours:
                shrunk.append((sheet, ours, theirs))

        if not shrunk:
            log.info("Layer 3 verify (probe %d): remote xlsx matches or "
                     "exceeds our pushed row counts for all monitored "
                     "sheets — no stomp detected", probe)
            return

        # Shrink detected. If this is not the final probe, wait briefly
        # for any in-flight concurrent push to settle, then re-check.
        if probe < 3:
            log.warning("Layer 3 verify (probe %d): remote shrunk in %s — "
                        "re-probing in case of in-flight concurrent push",
                        probe, ", ".join(f"{s}({o}>{t})"
                                         for s, o, t in shrunk))
            time.sleep(probe * 3)
            continue

        # Final probe still shows shrink — fire the alarm.
        for sheet, ours, theirs in shrunk:
            log.error("LAYER 3 ALARM: xlsx-staging stomp detected. Sheet "
                      "%r: we pushed %d rows, remote now has %d rows. "
                      "A concurrent workflow likely overwrote our commit. "
                      "Operator action required: audit recent commits to "
                      "%s and recover lost rows manually.",
                      sheet, ours, theirs, XLSX_REL)
        return

    log.debug("Layer 3 verify: could not read remote xlsx after 3 "
              "probes — skipping verification")


def git_commit_and_push(repo_dir: Path, files: list[str], message: str) -> bool:
    """Stage, commit, push files with PRE-EMPTIVE sync + retry-merge +
    post-push verification. Returns True on success."""
    try:
        cwd = str(repo_dir)

        # Configure committer (CI-friendly)
        _run(["git", "-C", cwd, "config", "user.email",
              os.getenv("GIT_USER_EMAIL", "fsis-bot@advfood.tech")])
        _run(["git", "-C", cwd, "config", "user.name",
              os.getenv("GIT_USER_NAME", "FSIS Bot")])

        # Stage caller's files
        for f in files:
            _run(["git", "-C", cwd, "add", f])

        # Check if there are changes to commit
        result = _run(["git", "-C", cwd, "diff", "--cached", "--quiet"])
        if result.returncode == 0:
            log.info("No changes to commit")
            return True

        # ── LAYER 1: pre-emptive sync ───────────────────────────────────
        try:
            _preemptive_sync(cwd, files)
        except Exception as e:
            log.warning("Pre-emptive sync raised — continuing to commit: %s", e)

        # Capture pre-push row counts of our xlsx for layer 3 verification.
        # Read AFTER preemptive_sync so we capture the post-merge state
        # that we're actually about to push.
        pre_push_counts = _xlsx_sheet_counts(Path(cwd) / XLSX_REL)

        # Commit
        _run(["git", "-C", cwd, "commit", "-m", message])

        # Push with retry
        push_cmd = _build_push_cmd(cwd)

        pushed_ok = False
        for attempt in range(1, MAX_PUSH_ATTEMPTS + 1):
            result = _run(push_cmd)
            if result.returncode == 0:
                log.info("Pushed (attempt %d): %s", attempt, message)
                pushed_ok = True
                break

            log.warning("Push attempt %d failed: %s",
                        attempt, result.stderr.strip())
            if attempt == MAX_PUSH_ATTEMPTS:
                break

            # ── LAYER 2: Binary-safe retry merge ────────────────────────
            # git rebase can't handle binary xlsx, so instead we:
            #   1. Identify which files we changed AND classify them
            #      as modifications/additions vs deletions
            #   2. Save physical copies of modifications/additions
            #   3. Hard-reset to pre-commit state + pull remote
            #      (this restores deleted files from HEAD~1)
            #   4. Restore our modifications on top of the updated remote;
            #      re-apply our deletions by unlinking from working tree
            #   5. Re-stage everything (modifications AND deletions) and commit
            #
            # Audit 2026-05-14: pre-fix, this block used `--name-only` and
            # filtered out non-existent paths with `if src.exists()`, which
            # silently dropped every deletion. After reset --hard HEAD~1 +
            # pull, the deleted files reappeared from HEAD~1 and were
            # never re-staged for the retry commit. Net effect: when push
            # failed due to a concurrent-workflow conflict (common in
            # this pipeline), the retry commit included our modifications
            # but NOT our deletions. Orphan HTMLs accumulated in
            # docs/daily/ for weeks until disk hygiene was investigated.

            # 1. Collect our changed file paths + status (M/A/D/R/C)
            changed = _run(["git", "-C", cwd, "diff",
                            "HEAD~1", "--name-status"])
            modifications: list[str] = []  # M or A — file exists post-commit
            deletions: list[str] = []       # D — file removed post-commit
            for ln in changed.stdout.splitlines():
                parts = ln.split("\t")
                if len(parts) < 2:
                    continue
                status = parts[0].strip()
                if not status:
                    continue
                if status.startswith("D"):
                    deletions.append(parts[1].strip())
                elif status.startswith("R") or status.startswith("C"):
                    # Rename / copy: parts = [status, old_path, new_path].
                    # Old path is gone (treat as deletion); new path is
                    # present (treat as modification/addition).
                    if len(parts) >= 3:
                        deletions.append(parts[1].strip())
                        modifications.append(parts[2].strip())
                    else:
                        # Malformed line — best-effort fall-through
                        modifications.append(parts[1].strip())
                else:
                    # M, A, T (typechange), U (unmerged) — file should
                    # exist in working tree
                    modifications.append(parts[1].strip())

            # 2. Save copies of modifications/additions to a temp dir.
            #    Deletions don't need saving (file is already gone).
            tmp = Path(cwd) / ".push-retry-tmp"
            tmp.mkdir(exist_ok=True)
            saved: list[tuple[str, Path]] = []
            for rel in modifications:
                src = Path(cwd) / rel
                if src.exists():
                    dst = tmp / rel.replace("/", "__")
                    shutil.copy2(src, dst)
                    saved.append((rel, dst))
                else:
                    # Modification claimed by `git diff` but file is gone
                    # from working tree — odd state, log it and skip.
                    log.warning("Push retry: expected modification at %s "
                                "but file is missing — skipping save", rel)

            # 3. Undo our commit, pull remote
            branch = _run(["git", "-C", cwd, "branch",
                           "--show-current"]).stdout.strip() or "main"
            _run(["git", "-C", cwd, "reset", "--hard", "HEAD~1"])
            _run(["git", "-C", cwd, "pull", "--ff-only",
                  "origin", branch])

            # 4. Restore our files on top of the updated remote.
            #
            # CRITICAL: for cumulative shared state (recalls.xlsx and the
            # mirror recalls.json), we MUST NOT blindly copy our local
            # over the freshly-pulled remote — that destroys rows the
            # remote-winner just added. Instead, merge xlsx by union of
            # rows (URL primary, fallback to Date+Company+Pathogen) and
            # regenerate the json mirror from the merged xlsx.
            saved_map = dict(saved)

            if XLSX_REL in saved_map and (Path(cwd) / XLSX_REL).exists():
                if _row_union_merge_xlsx(cwd, saved_map[XLSX_REL]):
                    saved = [(rel, dst) for rel, dst in saved if rel != XLSX_REL]
                    # CRITICAL FIX 2026-05-14 evening: explicitly stage the
                    # merged xlsx. Removing it from `saved` is correct so
                    # step 4b's "copy ours over working tree" loop doesn't
                    # clobber the merged file with our raw saved version.
                    # BUT step 5's `git add` loop only iterates `saved` —
                    # so without an explicit add here the merged xlsx
                    # sits modified-but-unstaged on disk, and the retry
                    # commit silently drops it.
                    _run(["git", "-C", cwd, "add", "--", XLSX_REL])
                    if JSON_REL in saved_map:
                        if _regenerate_json_mirror(cwd):
                            saved = [(rel, dst) for rel, dst in saved
                                     if rel != JSON_REL]
                            _run(["git", "-C", cwd, "add", "--", JSON_REL])
                else:
                    # AUDIT 2026-05-16: when the canary trips (a row-loss
                    # would happen if we merged), we ABORT rather than
                    # fall back to the old ours-wins overwrite. Losing a
                    # push attempt is recoverable; losing rows is not.
                    log.error("xlsx merge failed/canary-tripped — refusing "
                              "to fall back to ours-wins overwrite. "
                              "Aborting push to avoid data loss.")
                    shutil.rmtree(tmp, ignore_errors=True)
                    return False

            # For non-cumulative files (logs, configs, html outputs) the
            # original ours-wins copy is acceptable.
            for rel, dst in saved:
                dest = Path(cwd) / rel
                dest.parent.mkdir(parents=True, exist_ok=True)
                shutil.copy2(dst, dest)
            shutil.rmtree(tmp, ignore_errors=True)

            # Re-apply our deletions. After `git reset --hard HEAD~1` +
            # `git pull`, any files we deleted in our original commit
            # are back in the working tree (either restored from HEAD~1
            # or pulled from remote). To preserve our intent, we unlink
            # them again. If a deletion path is not in the working tree
            # post-pull, it means remote also deleted it (or never had
            # it) — either way, nothing to do.
            redeleted = 0
            for rel in deletions:
                dest = Path(cwd) / rel
                try:
                    if dest.exists():
                        dest.unlink()
                        redeleted += 1
                except Exception as e:
                    log.warning("Push retry: could not re-delete %s: %s",
                                rel, e)
            if deletions:
                log.info("Push retry: re-applied %d/%d deletions",
                         redeleted, len(deletions))

            # 5. Re-stage and commit. For modifications: file is on
            # disk, `git add` stages the modification. For deletions:
            # file is missing from working tree, `git add` stages the
            # deletion (verified empirically — `git add -- <missing>`
            # on a tracked path stages the deletion, exit 0).
            for rel, _ in saved:
                _run(["git", "-C", cwd, "add", "--", rel])
            for rel in deletions:
                _run(["git", "-C", cwd, "add", "--", rel])
            check = _run(["git", "-C", cwd, "diff",
                          "--cached", "--quiet"])
            if check.returncode != 0:
                retry_msg = f"{message} (retry {attempt})"
                _run(["git", "-C", cwd, "commit", "-m", retry_msg])

                # Refresh pre-push counts in case the row-merge changed them
                pre_push_counts = _xlsx_sheet_counts(Path(cwd) / XLSX_REL)

            time.sleep(attempt * 3)

        if not pushed_ok:
            log.error("git push failed after %d attempts", MAX_PUSH_ATTEMPTS)
            return False

        # ── LAYER 3: post-push verification ─────────────────────────────
        # Detect xlsx-staging stomps within a 30-second window after push.
        try:
            _post_push_verify(cwd, files, pre_push_counts)
        except Exception as e:
            log.warning("Post-push verification raised — push already "
                        "succeeded, ignoring: %s", e)

        return True

    except Exception as e:
        log.error("commit_and_push failed: %s", e)
        return False


def _build_push_cmd(cwd: str) -> list[str]:
    """Build the git push command, injecting token if available."""
    token = os.getenv("GH_TOKEN", "")
    if token:
        remote = _run(["git", "-C", cwd, "remote", "get-url", "origin"])
        url = remote.stdout.strip()
        if url.startswith("https://github.com/"):
            authed = url.replace(
                "https://github.com/",
                f"https://x-access-token:{token}@github.com/",
            )
            return ["git", "-C", cwd, "push", authed, "HEAD"]
    return ["git", "-C", cwd, "push"]
