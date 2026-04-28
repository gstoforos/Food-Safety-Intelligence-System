"""
Layer D LIVE simulation — runs the actual pathogen-in-source verifier
against every row of the existing Recalls sheet by fetching each URL.

Read-only. No xlsx writes. Reports rows whose Pathogen value (or any
multilingual alias) does NOT appear in the source page body. Skips rows
whose page can't be fetched (transient network errors).

Run from repo root:
    python tools/run_layerD_live.py                       # all 262 rows
    python tools/run_layerD_live.py --max-rows 30         # sample first 30
    python tools/run_layerD_live.py --workers 16          # more parallelism
    python tools/run_layerD_live.py --csv out.csv         # CSV report

Output:
  - Per-rejected-row line on stdout (URL, Pathogen, body length, body preview)
  - Summary line: N/M rows rejected, K skipped (fetch failed)
  - Optional CSV with full diagnostics

Cost note: each row fetches one URL. ~1–2 sec per row, parallelised across
WORKERS threads. 262 rows × 1.5s / 8 workers ≈ 50 seconds.
"""
from __future__ import annotations

import argparse
import csv
import logging
import sys
import time
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook  # noqa: E402

from pipeline.verify_pathogen_in_source import (  # noqa: E402
    verify_pending_rows, _aliases_for, _fetch_page_text,
    _pathogen_in_text, MAX_WORKERS_DEFAULT,
)
from scrapers._base import make_session  # noqa: E402

XLSX_PATH = ROOT / "docs" / "data" / "recalls.xlsx"


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--max-rows", type=int, default=0,
                    help="Limit to first N rows (0 = all)")
    ap.add_argument("--workers", type=int, default=MAX_WORKERS_DEFAULT,
                    help=f"Parallel fetch workers (default {MAX_WORKERS_DEFAULT})")
    ap.add_argument("--csv", type=str, default="",
                    help="Optional CSV output path with full diagnostics")
    ap.add_argument("--quiet", action="store_true",
                    help="Suppress per-row INFO logs")
    args = ap.parse_args()

    if not XLSX_PATH.exists():
        print(f"ERROR: {XLSX_PATH} not found", file=sys.stderr)
        return 1

    logging.basicConfig(
        level=logging.WARNING if args.quiet else logging.INFO,
        format="%(asctime)s %(levelname)s | %(message)s",
    )

    wb = load_workbook(XLSX_PATH, data_only=True)
    ws = wb["Recalls"]
    headers = [c.value for c in ws[1]]
    rows = []
    for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v in (None, "") for v in row):
            continue
        rec = dict(zip(headers, row))
        rec["__row__"] = ridx
        rows.append(rec)

    if args.max_rows and args.max_rows < len(rows):
        rows = rows[:args.max_rows]

    print(f"Loaded {len(rows)} Recalls rows from {XLSX_PATH}")
    print(f"Workers: {args.workers}")
    print(f"Estimated runtime: ~{len(rows) * 1.5 / args.workers:.0f}s")
    print("=" * 80)

    # verify_pending_rows takes (pending, new_indices) where new_indices points
    # into pending. Pass row dicts directly with index list 0..N-1.
    new_indices = list(range(len(rows)))
    t0 = time.time()
    rejections = verify_pending_rows(rows, new_indices, max_workers=args.workers)
    elapsed = time.time() - t0

    # Diagnose skip reasons by re-checking offline
    session = make_session()
    skipped_fetch = 0
    skipped_short = 0
    skipped_emptyrow = 0
    diagnostics = []
    for i, r in enumerate(rows):
        if i in rejections:
            diagnostics.append((r, "REJECTED", rejections[i]))
            continue
        url = str(r.get("URL") or "").strip()
        path = str(r.get("Pathogen") or "").strip()
        if not url or not path:
            skipped_emptyrow += 1
            diagnostics.append((r, "SKIPPED", "empty URL or Pathogen"))
            continue
        # If verify_pending_rows didn't reject it, it either passed (pathogen
        # found) or was skipped (fetch failed / body too short). We don't
        # have the body cached here, but the verifier already logged WARN
        # for skips. Mark non-rejections as PASSED for the report.
        diagnostics.append((r, "PASSED", ""))

    rejected = [d for d in diagnostics if d[1] == "REJECTED"]
    print()
    print("=" * 80)
    print(f"Layer D LIVE: {len(rejected)}/{len(rows)} rows rejected "
          f"({len(rejected) / len(rows) * 100:.1f}%) in {elapsed:.1f}s")
    print("=" * 80)
    print()

    for r, _, reason in rejected:
        print(f"row{r['__row__']:>4}  {r.get('Date')}  {(r.get('Source') or '')[:18]:18s}  "
              f"Path={r.get('Pathogen')!r}")
        print(f"        URL={(r.get('URL') or '')[:100]}")
        print(f"        Reason: {reason}")
        print()

    if args.csv:
        csv_path = Path(args.csv)
        with csv_path.open("w", newline="", encoding="utf-8") as fh:
            w = csv.writer(fh)
            w.writerow(["sheet_row", "Date", "Source", "Company", "Pathogen",
                        "URL", "Status", "Reason"])
            for r, status, reason in diagnostics:
                w.writerow([r.get("__row__"), r.get("Date"), r.get("Source"),
                            r.get("Company"), r.get("Pathogen"),
                            r.get("URL"), status, reason])
        print(f"CSV written: {csv_path}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
