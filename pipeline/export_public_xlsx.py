"""
Build the PUBLIC subscriber/visitor xlsx from the master recalls.xlsx.

Audit 2026-04-29 — Notes + internal date columns excluded:
  The master recalls.xlsx Recalls sheet has 17 columns:
    Date, Source, Company, Brand, Product, Pathogen, Reason, Class,
    Country, Region, Tier, Outbreak, URL, Notes,
    DateAdded, LastUpdated, LastChecked
  The free download exposed all 14 dashboard columns (incl. Notes), which
  leaked internal provenance to subscribers in 97 of 271 rows:
    - AI vendor names ("via Gemini gap-finder + Google Search",
      "via OpenAI daily 10:00 Athens search", "Gemini/de from ...")
    - Internal audit stamps ("[Claude-flag 2026-04-16: ...]",
      "[outbreak 2026-04-29: 0→1 — ...]", "[URL-guardian ...]",
      "[url-gate 2026-04-29 ...]", "[audit-applied ...]")
    - Manual-edit traces ("manually added 2026-04-29", "Mannually" typo)
    - URL-repair history ("URL repaired 2026-04-29 — previous URL had year
      value as fiche ID")
  This was a polish/trust issue — subscribers saw the seams of the build
  pipeline. Stripping Notes also drops the legitimate distribution/lot
  info embedded in some rows, but that's acceptable: the URL points to
  the regulator's authoritative page where that detail belongs anyway.

  Internal date columns (DateAdded/LastUpdated/LastChecked) are also
  excluded per spec — subscribers don't need to know when AFTS first
  recorded a row, only the regulator's publication Date.

PUBLIC SCHEMA (13 columns):
  Date, Source, Company, Brand, Product, Pathogen, Reason, Class,
  Country, Region, Tier, Outbreak, URL

USAGE
-----
    python -m pipeline.export_public_xlsx
    # writes docs/data/afts-recalls-public.xlsx

Wire it into your existing dashboard download link:
    (download button) -> /data/afts-recalls-public.xlsx
"""
from __future__ import annotations
import argparse
import logging
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
SRC_PATH = ROOT / "docs" / "data" / "recalls.xlsx"
OUT_PATH = ROOT / "docs" / "data" / "afts-recalls-public.xlsx"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
log = logging.getLogger("export-public-xlsx")

# Public schema — 13 columns. Notes, DateAdded, LastUpdated, LastChecked
# all stay internal-only.
PUBLIC_FIELDS = [
    "Date", "Source", "Company", "Brand", "Product",
    "Pathogen", "Reason", "Class", "Country", "Region",
    "Tier", "Outbreak", "URL",
]

# Optional: subset of fields that benefit from custom column widths
COLUMN_WIDTHS = {
    "Date": 12, "Source": 18, "Company": 28, "Brand": 18,
    "Product": 50, "Pathogen": 24, "Reason": 38, "Class": 12,
    "Country": 14, "Region": 16, "Tier": 6, "Outbreak": 9,
    "URL": 60,
}


def _coerce_date_str(v) -> str:
    if v is None or v == "":
        return ""
    if hasattr(v, "isoformat"):
        try:
            return v.isoformat()[:10]
        except (TypeError, ValueError):
            return ""
    return str(v)[:10]


def build_public_xlsx(src: Path = SRC_PATH, out: Path = OUT_PATH) -> int:
    if not src.exists():
        log.error("Source not found: %s", src)
        return 1

    src_wb = openpyxl.load_workbook(src)
    src_ws = src_wb["Recalls"]
    src_headers = [c.value for c in src_ws[1]]
    ic = {h: i for i, h in enumerate(src_headers)}

    missing = [f for f in PUBLIC_FIELDS if f not in ic]
    if missing:
        log.error("Source xlsx missing public fields: %s", missing)
        return 2

    # Walk source, copy only public fields into the output workbook.
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "Recalls"

    # Header row with light styling
    header_fill = PatternFill(start_color="0A3D2E", end_color="0A3D2E",
                              fill_type="solid")
    header_font = Font(name="Inter", color="FFFFFF", bold=True, size=10)
    body_font = Font(name="Inter", size=10)
    border = Border(
        left=Side(style="thin", color="D4D4D4"),
        right=Side(style="thin", color="D4D4D4"),
        top=Side(style="thin", color="D4D4D4"),
        bottom=Side(style="thin", color="D4D4D4"),
    )

    out_ws.append(PUBLIC_FIELDS)
    for col_idx, name in enumerate(PUBLIC_FIELDS, start=1):
        cell = out_ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = border
        col_letter = get_column_letter(col_idx)
        out_ws.column_dimensions[col_letter].width = COLUMN_WIDTHS.get(name, 16)

    # Data rows
    rows_written = 0
    for row in src_ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        out_row = []
        for f in PUBLIC_FIELDS:
            v = row[ic[f]]
            if f == "Date":
                v = _coerce_date_str(v)
            elif f in ("Tier", "Outbreak"):
                try:
                    v = int(v) if v not in (None, "") else 0
                except (TypeError, ValueError):
                    v = 0
            elif v is None:
                v = ""
            out_row.append(v)
        out_ws.append(out_row)
        rows_written += 1

    # Body styling — apply font + borders to data range
    for row in out_ws.iter_rows(min_row=2, max_row=rows_written + 1,
                                 max_col=len(PUBLIC_FIELDS)):
        for cell in row:
            cell.font = body_font
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="top",
                                       wrap_text=False)

    # Freeze the header row
    out_ws.freeze_panes = "A2"

    # Add an "About" sheet with metadata (no Notes, no provenance)
    about = out_wb.create_sheet(title="About")
    about["A1"] = "AFTS Food Safety Intelligence System"
    about["A1"].font = Font(name="Inter", bold=True, size=14, color="0A3D2E")
    about["A2"] = "Public recall dataset"
    about["A2"].font = Font(name="Inter", size=11, color="555555")
    about["A4"] = f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
    about["A5"] = f"Records: {rows_written}"
    about["A6"] = f"Coverage: pathogen-related recalls from official regulatory agencies"
    about["A7"] = f"Source: 66 agencies across 60+ countries"
    about["A9"] = "Each row links to the original regulatory notice via the URL column."
    about["A10"] = "For full intelligence narrative, predictive models, and engineering"
    about["A11"] = "interpretation, subscribe at advfood.tech/fsis-home"
    about["A13"] = "© 2026 Advanced Food-Tech Solutions · advfood.tech · info@advfood.tech"
    about["A13"].font = Font(name="Inter", size=9, color="888888")
    about.column_dimensions["A"].width = 78

    out.parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(out)
    log.info("Wrote %s (%d records, %d public columns)",
             out, rows_written, len(PUBLIC_FIELDS))
    log.info("Stripped: Notes, DateAdded, LastUpdated, LastChecked "
             "(internal-only)")
    return 0


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--src", type=Path, default=SRC_PATH,
                    help="Source xlsx (default: docs/data/recalls.xlsx)")
    ap.add_argument("--out", type=Path, default=OUT_PATH,
                    help="Output xlsx (default: docs/data/afts-recalls-public.xlsx)")
    args = ap.parse_args()
    return build_public_xlsx(args.src, args.out)


if __name__ == "__main__":
    sys.exit(main())
