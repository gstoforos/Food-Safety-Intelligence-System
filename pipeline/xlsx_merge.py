"""
pipeline/xlsx_merge.py — safe xlsx union-merge for push-retry conflicts.

THE BUG THIS FIXES
==================
news-feed.yml and merge-master.yml both had a "binary-safe push retry" that
copied the loser's stale local xlsx to /tmp, pulled the winner's remote, then
OVERWROTE the freshly-pulled remote with the stale /tmp copy. Result: any
rows the remote-winner had just added (e.g. an operator's manual upload of
recalls.xlsx with +11 new rows) got silently destroyed on the next news
or merge tick.

THE FIX
=======
On push-retry (and now on pre-emptive sync), instead of overwriting, do a
row-level union of:
  - Recalls         : remote ∪ ours; never shrinks. Remote wins on collision.
  - Pending         : remote ∪ ours, minus any row whose dedup_key already
                      appears in the merged Recalls.
  - Weekly_Review   : remote ∪ ours by URL+Date.
  - Weekly_Rejected : remote ∪ ours by URL+Date.
  - NEWS            : remote ∪ ours by NEWS dedup key.

Identity = the same _dedup_key used by merge_master:
URL primary, fallback to Date+Company+Pathogen.

THE CANARY (audit 2026-05-16)
=============================
Every sheet merge now asserts:

    len(merged) >= max(len(remote), len(ours))

If a merge would EVER shrink a sheet relative to either input, that's
mathematically impossible for a true union and indicates a logic bug
(or corrupted data). The assertion raises and the push aborts rather
than silently shipping a smaller file. commit_github.py catches the
assertion and refuses to fall back to ours-wins overwrite.

This is the structural guarantee that 2026-05-16's 6-row stomp class
of bug CANNOT happen again undetected — even if every other safety
layer fails (workflow concurrency, pre-emptive sync), a merge that
would lose rows now fails LOUD instead of silent.

Used by:
  pipeline/commit_github.py  (called from layer 1 pre-sync AND layer 2 retry)
  .github/workflows/news-feed.yml      (via inline Python in retry block)
  .github/workflows/merge-master.yml   (same pattern)
"""
from __future__ import annotations
import logging
import re
import sys
import unicodedata
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook, load_workbook

log = logging.getLogger(__name__)


def _dedup_key(row: Dict[str, Any]) -> str:
    """Same logic as pipeline.merge_master._dedup_key."""
    url = (row.get("URL") or "").strip().lower()
    if url:
        return url
    co = unicodedata.normalize("NFD", str(row.get("Company") or "")) \
        .encode("ascii", "ignore").decode().lower()
    co = re.sub(r"[^a-z0-9]", "", co)[:30]
    return f"{row.get('Date','')}|{co}|{(str(row.get('Pathogen','')) or '')[:30]}"


def _news_dedup_key(row: Dict[str, Any]) -> str:
    """Same logic as pipeline.merge_master._news_dedup_key."""
    link = (row.get("Link") or "").strip().lower()
    if link:
        return link
    title = (str(row.get("Title") or "")).strip().lower()[:120]
    return f"{row.get('Published (UTC)','')}|{title}"


def _wr_dedup_key(row: Dict[str, Any]) -> str:
    """Weekly_Review dedup key — URL+Date, matching the rule used by
    pipeline.weekly_review_capture._existing_keys.

    A row is uniquely identified by its (URL, Date) pair within the
    Weekly_Review sheet. If two pushes both wrote the same recall to
    Weekly_Review (different timestamps, identical URL+Date), they
    collapse to one row in the merge — preserving the earlier-written
    one's Week_Added, Reviewed, and audit-trail fields.
    """
    _u = row.get("URL")
    url = (str(_u).strip().lower() if _u not in (None, "") else "")
    _d = row.get("Date")
    d = (str(_d)[:10] if _d not in (None, "") else "")
    if url:
        return f"{url}|{d}"
    # No URL: do NOT collapse all URL-less rows on the same date into one key
    # (that shrinks the union and trips the no-shrink canary). Build a richer
    # key from the row's distinguishing fields; if those are also empty, fall
    # back to the row's identity (id()) so each URL-less row stays distinct.
    def _f(field, n=None):
        v = row.get(field)
        if v in (None, ""):
            return ""
        v = str(v).strip().lower()
        return v[:n] if n else v
    parts = [
        _f("Pathogen"), _f("Company"), _f("Product", 40),
        _f("Brand"), _f("Reason", 40),
    ]
    sig = "|".join(parts)
    if sig.strip("|"):
        return f"||{d}|{sig}"
    # Completely empty row content — keep distinct by object identity so the
    # union never silently drops it.
    return f"||{d}|__rowid_{id(row)}__"


def _read_sheet(xlsx_path: Path, sheet: str) -> Tuple[List[str], List[Dict[str, Any]]]:
    if not xlsx_path.exists():
        return [], []
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        return [], []
    ws = wb[sheet]
    headers = [c.value for c in ws[1] if c.value is not None]
    rows: List[Dict[str, Any]] = []
    for tup in ws.iter_rows(min_row=2, values_only=True):
        if all(v in (None, "") for v in tup):
            continue
        rec = {h: (v if v is not None else "") for h, v in zip(headers, tup)}
        rows.append(rec)
    wb.close()
    return headers, rows


def _write_sheet(ws, headers: List[str], rows: List[Dict[str, Any]]) -> None:
    for c, h in enumerate(headers, start=1):
        ws.cell(1, c, h)
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, h in enumerate(headers, start=1):
            ws.cell(r_idx, c_idx, row.get(h, ""))


def _merge_unique(
    remote_rows: List[Dict[str, Any]],
    ours_rows: List[Dict[str, Any]],
    key_fn,
) -> List[Dict[str, Any]]:
    """Union by key_fn; remote wins on collision."""
    seen: Dict[str, Dict[str, Any]] = {}
    out: List[Dict[str, Any]] = []
    for r in remote_rows:
        k = key_fn(r)
        if k and k not in seen:
            seen[k] = r
            out.append(r)
    for r in ours_rows:
        k = key_fn(r)
        if k and k not in seen:
            seen[k] = r
            out.append(r)
    return out


def _assert_no_shrink(
    sheet_name: str,
    remote_rows: List[Dict[str, Any]],
    ours_rows: List[Dict[str, Any]],
    merged_rows: List[Dict[str, Any]],
) -> None:
    """CANARY (audit 2026-05-16) — a true union cannot shrink either input.
    If the merged sheet has fewer rows than the larger input, raise so
    commit_github.py refuses the push.

    NOTE: Pending is special-cased by the caller (we filter merged Pending
    against Recalls keys, which can legitimately shrink Pending). For all
    other sheets this canary is unconditional.
    """
    n_remote = len(remote_rows)
    n_ours = len(ours_rows)
    n_merged = len(merged_rows)
    n_max = max(n_remote, n_ours)
    if n_merged < n_max:
        msg = (
            f"xlsx_merge CANARY: {sheet_name} merge would shrink "
            f"({n_merged} < max(remote={n_remote}, ours={n_ours})). "
            f"This is impossible for a true union — refusing to write."
        )
        log.error(msg)
        raise AssertionError(msg)


def merge_xlsx_with_remote(
    remote_path: Path,
    ours_path: Path,
    out_path: Path,
) -> Dict[str, int]:
    """Merge our local xlsx with the just-pulled remote, write to out_path.
    Every sheet merge is checked against the no-shrink canary; the
    function raises AssertionError if any merged sheet is smaller than
    either input.

    Returns counts dict for logging.
    """
    rec_headers, rec_remote = _read_sheet(remote_path, "Recalls")
    _, rec_ours = _read_sheet(ours_path, "Recalls")
    if not rec_headers:
        rec_headers = (
            ["Date", "Source", "Company", "Brand", "Product", "Pathogen",
             "Reason", "Class", "Country", "Region", "Tier", "Outbreak",
             "URL", "Notes"]
        )
    rec_merged = _merge_unique(rec_remote, rec_ours, _dedup_key)
    _assert_no_shrink("Recalls", rec_remote, rec_ours, rec_merged)
    rec_keys = {_dedup_key(r) for r in rec_merged}

    pen_headers, pen_remote = _read_sheet(remote_path, "Pending")
    _, pen_ours = _read_sheet(ours_path, "Pending")
    if not pen_headers:
        # Audit 2026-05-06 — added RejectedBy column (Phase A counter
        # support). Without this, push-retry conflicts would silently
        # strip the RejectedBy column, breaking the 2-reviewer-rejection
        # delete logic in merge_master.cleanup_orphan_rejected.
        pen_headers = rec_headers + ["ScrapedAt", "Status", "RejectedBy"]
    pen_merged_raw = _merge_unique(pen_remote, pen_ours, _dedup_key)
    # SPECIAL CASE: Pending is filtered against the merged Recalls keys,
    # so it CAN legitimately shrink (a row promoted to Recalls by the
    # remote-winner is correctly dropped from our Pending). The no-shrink
    # canary therefore checks the RAW union, not the post-filter result.
    _assert_no_shrink("Pending (pre-filter)", pen_remote, pen_ours,
                      pen_merged_raw)
    pen_merged = [r for r in pen_merged_raw if _dedup_key(r) not in rec_keys]

    news_headers, news_remote = _read_sheet(remote_path, "NEWS")
    _, news_ours = _read_sheet(ours_path, "NEWS")
    if not news_headers:
        news_headers = (
            ["Published (UTC)", "Pathogen", "Event", "Source",
             "Title", "Link", "Retrieved (UTC)"]
        )
    news_merged = _merge_unique(news_remote, news_ours, _news_dedup_key)
    _assert_no_shrink("NEWS", news_remote, news_ours, news_merged)

    # ── Weekly_Review (audit 2026-05-06) ──────────────────────────────
    # Pre-2026-05-06 this module silently dropped the Weekly_Review sheet
    # on every push-retry conflict. The xlsx_merge.merge_xlsx_with_remote
    # path created a fresh workbook with only Recalls/Pending/NEWS, so any
    # operator-built or mailer-state Weekly_Review tab was destroyed on
    # the next news-feed or merge-master tick.
    #
    # Now we union both sides by URL+Date dedup (matches the rule that
    # weekly_review_capture.record_promotions uses). Headers come from
    # remote if present, else from ours, else from the Phase B canonical
    # template.
    wr_headers, wr_remote = _read_sheet(remote_path, "Weekly_Review")
    _, wr_ours = _read_sheet(ours_path, "Weekly_Review")
    if not wr_headers:
        # fall back to ours' headers if available
        wr_headers_local, _ = _read_sheet(ours_path, "Weekly_Review")
        if wr_headers_local:
            wr_headers = wr_headers_local
        else:
            # Canonical schema used by pipeline/weekly_review_capture.py
            wr_headers = rec_headers + [
                "DateAdded", "LastUpdated", "LastChecked",
                "Week_Added", "Reviewed",
            ]
    wr_merged = _merge_unique(wr_remote, wr_ours, _wr_dedup_key)
    _assert_no_shrink("Weekly_Review", wr_remote, wr_ours, wr_merged)

    # ── Weekly_Rejected (audit 2026-05-12) ────────────────────────────
    # Pre-2026-05-12 this module silently dropped the Weekly_Rejected
    # sheet on every push-retry conflict — same bug class as the
    # 2026-05-06 Weekly_Review case, missed because Weekly_Rejected was
    # added later (audit 2026-05-09) and never propagated here.
    # Observed in production: every gap-finder cascade / news-feed run
    # that hit a non-fast-forward and fell back to xlsx_merge ended with
    # a fresh workbook containing only Recalls/Pending/Weekly_Review/NEWS.
    # The 30-row Weekly_Rejected sheet that claude-check had written
    # minutes earlier was gone after the merge, breaking the operator's
    # Thursday review email.
    #
    # Same union logic and dedup_key as Weekly_Review (URL+Date), since
    # passes and rejects share the same identity rule.
    wj_headers, wj_remote = _read_sheet(remote_path, "Weekly_Rejected")
    _, wj_ours = _read_sheet(ours_path, "Weekly_Rejected")
    if not wj_headers:
        wj_headers_local, _ = _read_sheet(ours_path, "Weekly_Rejected")
        if wj_headers_local:
            wj_headers = wj_headers_local
        else:
            # Canonical schema used by pipeline/weekly_rejected_capture.py
            wj_headers = rec_headers + [
                "DateAdded", "LastUpdated", "LastChecked",
                "Week_Added", "RejectedBy", "RejectionReason", "Reviewed",
            ]
    wj_merged = _merge_unique(wj_remote, wj_ours, _wr_dedup_key)
    _assert_no_shrink("Weekly_Rejected", wj_remote, wj_ours, wj_merged)

    wb = Workbook()
    rec_ws = wb.active
    rec_ws.title = "Recalls"
    _write_sheet(rec_ws, rec_headers, rec_merged)
    pen_ws = wb.create_sheet("Pending")
    _write_sheet(pen_ws, pen_headers, pen_merged)
    if wr_merged or wr_remote or wr_ours:
        # Only create Weekly_Review if at least one side had data.
        # If both sides had ZERO Weekly_Review rows AND no Weekly_Review
        # sheet existed, don't create an empty one — record_promotions
        # will create it on the next promotion.
        wr_ws = wb.create_sheet("Weekly_Review")
        _write_sheet(wr_ws, wr_headers, wr_merged)
    if wj_merged or wj_remote or wj_ours:
        # Same gating logic as Weekly_Review — only materialize the sheet
        # if at least one side had a row. Empty Weekly_Rejected sheets
        # are recreated by weekly_rejected_capture.record_rejections.
        wj_ws = wb.create_sheet("Weekly_Rejected")
        _write_sheet(wj_ws, wj_headers, wj_merged)
    news_ws = wb.create_sheet("NEWS")
    _write_sheet(news_ws, news_headers, news_merged)
    wb.save(out_path)

    counts = {
        "recalls_remote": len(rec_remote),
        "recalls_ours":   len(rec_ours),
        "recalls_merged": len(rec_merged),
        "pending_remote": len(pen_remote),
        "pending_ours":   len(pen_ours),
        "pending_merged": len(pen_merged),
        "weekly_review_remote": len(wr_remote),
        "weekly_review_ours":   len(wr_ours),
        "weekly_review_merged": len(wr_merged),
        "weekly_rejected_remote": len(wj_remote),
        "weekly_rejected_ours":   len(wj_ours),
        "weekly_rejected_merged": len(wj_merged),
        "news_remote":    len(news_remote),
        "news_ours":      len(news_ours),
        "news_merged":    len(news_merged),
    }
    log.info("xlsx merge: %s", counts)
    return counts


# ---------------------------------------------------------------------------
# CLI for use from inside YAML retry blocks:
#     python -m pipeline.xlsx_merge <remote_path> <ours_path> <out_path>
# ---------------------------------------------------------------------------
def _cli() -> int:
    logging.basicConfig(level=logging.INFO,
                        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s")
    if len(sys.argv) != 4:
        print("usage: python -m pipeline.xlsx_merge <remote> <ours> <out>",
              file=sys.stderr)
        return 2
    try:
        counts = merge_xlsx_with_remote(
            Path(sys.argv[1]), Path(sys.argv[2]), Path(sys.argv[3]),
        )
    except AssertionError as ae:
        # Canary tripped — exit with a distinctive code so the caller
        # (YAML retry block) can distinguish "no-shrink violation" from
        # other failures.
        print(f"CANARY: {ae}", file=sys.stderr)
        return 3
    print(f"merged: Recalls={counts['recalls_merged']} "
          f"Pending={counts['pending_merged']} NEWS={counts['news_merged']}")
    return 0


if __name__ == "__main__":
    sys.exit(_cli())
