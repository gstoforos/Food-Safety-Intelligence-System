"""
Gap Auditor — the real "what is my automation missing" tool.

Compares what OFFICIAL regulator APIs/RSS show for a given date range
against what currently exists in docs/data/recalls.xlsx, and produces
a gap report per agency with ROOT-CAUSE classification per missed
recall:

   MISS reasons (per row):
     SCRAPER_DEAD     : the agency's scraper returned zero rows but API had data
     SCRAPER_FILTERED : scraper ran but filter excluded this (pathogen keyword mismatch)
     SCHEDULE_GAP     : recall published after last pipeline run — next run will catch it
     URL_GATE_DROP    : scraper did capture it, but URL gate killed it as "dead URL"
     NOT_IN_WHITELIST : pathogen not on FSIS pathogen whitelist (correctly excluded)
     TRULY_MISSING    : none of the above — genuine pipeline bug

Runs standalone — no Anthropic/Gemini/OpenAI calls. Pure HTTP + xlsx compare.

Invocation:
    python -m pipeline.gap_auditor                          # last 7 days
    python -m pipeline.gap_auditor --days 14
    python -m pipeline.gap_auditor --start 2026-04-19 --end 2026-04-22
    python -m pipeline.gap_auditor --output gap_report.xlsx

Outputs two files:
    gap_report_YYYY-MM-DD.xlsx   — multi-sheet Excel with findings
    gap_report_YYYY-MM-DD.csv    — flat CSV for scripting

The xlsx has one sheet per agency plus a Summary sheet that shows:
  - How many recalls each agency published in the window (API truth)
  - How many your pipeline captured (Recalls sheet)
  - Gap = truth - captured
  - Per-miss root cause breakdown
"""
from __future__ import annotations
import argparse
import csv
import json
import logging
import re
import sys
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone, date
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
from xml.etree import ElementTree as ET

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from scrapers._base import make_session, fetch  # noqa: E402
from scrapers._models import PATHOGEN_RULES  # noqa: E402

logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s")
log = logging.getLogger("gap-auditor")


# ---------------------------------------------------------------------------
# Pathogen filter — same whitelist the pipeline uses
# ---------------------------------------------------------------------------
PATHOGEN_KEYWORDS = (
    "listeria", "salmonella", "salmonelle", "e. coli", "e.coli", "escherichia",
    "stec", "o157", "shiga", "botulin", "norovirus", "hepatitis", "hépatit",
    "campylobacter", "cyclospora", "vibrio", "cronobacter", "bacillus cereus",
    "cereulide", "shigella", "yersinia", "biotoxin", "biotoxine",
    "histamine", "scombro", "aflatoxin", "ochratoxin", "patulin",
    "mycotoxin", "alternaria", "fumonisin", "zearalenone", "deoxynivalenol",
)


def is_pathogen_recall(text: str) -> bool:
    """Does this recall text mention an in-whitelist pathogen?"""
    if not text:
        return False
    t = text.lower()
    return any(kw in t for kw in PATHOGEN_KEYWORDS)


# ---------------------------------------------------------------------------
# Regulator API / RSS definitions
# Only agencies with a KNOWN-WORKING public endpoint. Agencies whose sites
# are JS-heavy, blocked, or 404 are not in this list — no point calling them.
# ---------------------------------------------------------------------------
@dataclass
class RegulatorSource:
    agency: str
    country: str
    kind: str                       # 'api' | 'rss'
    url: str
    fetch_params: Dict[str, Any] = field(default_factory=dict)
    # Function name to parse this source's response → list of dicts with
    # 'date' (YYYY-MM-DD), 'company', 'product', 'pathogen', 'url', 'reason'
    parser: str = ""


REGULATORS: List[RegulatorSource] = [
    RegulatorSource(
        agency="RappelConso (FR)", country="France",
        kind="api",
        url="https://data.economie.gouv.fr/api/explore/v2.1/catalog/datasets/rappelconso0/records",
        parser="parse_rappelconso",
    ),
    RegulatorSource(
        agency="FDA", country="USA",
        kind="api",
        url="https://api.fda.gov/food/enforcement.json",
        parser="parse_fda_openfda",
    ),
    RegulatorSource(
        agency="USDA FSIS", country="USA",
        kind="rss",
        url="https://www.fsis.usda.gov/fsis-recalls/rss.xml",
        parser="parse_generic_rss",
    ),
    RegulatorSource(
        agency="CFIA", country="Canada",
        kind="rss",
        url="https://recalls-rappels.canada.ca/en/rss.xml",
        parser="parse_generic_rss",
    ),
    RegulatorSource(
        agency="FSA (UK)", country="United Kingdom",
        kind="api",
        url="https://data.food.gov.uk/food-alerts/id?_sort=-created&_view=published&_pageSize=100",
        parser="parse_fsa_uk",
    ),
    RegulatorSource(
        agency="FSAI (IE)", country="Ireland",
        kind="rss",
        url="https://www.fsai.ie/rss.aspx",
        parser="parse_generic_rss",
    ),
    RegulatorSource(
        agency="FSANZ (AU)", country="Australia",
        kind="rss",
        url="https://www.foodstandards.gov.au/RSS/Recalls/Pages/default.aspx",
        parser="parse_generic_rss",
    ),
    RegulatorSource(
        agency="BVL (DE)", country="Germany",
        kind="rss",
        url="https://www.lebensmittelwarnung.de/bvl-lmw-de/rss/warnungen_aktuell",
        parser="parse_generic_rss",
    ),
    RegulatorSource(
        agency="Livsmedelsverket (SE)", country="Sweden",
        kind="rss",
        url="https://www.livsmedelsverket.se/nyhetsarkiv/?RssFeed=true",
        parser="parse_generic_rss",
    ),
    RegulatorSource(
        agency="FSS (Scotland)", country="United Kingdom",
        kind="rss",
        url="https://www.foodstandards.gov.scot/news-and-alerts/feed",
        parser="parse_generic_rss",
    ),
]


# ---------------------------------------------------------------------------
# Parsers — one per regulator response format
# ---------------------------------------------------------------------------
def parse_rappelconso(body: bytes, url: str, start: date, end: date) -> List[Dict[str, Any]]:
    """Parse data.economie.gouv.fr RappelConso JSON."""
    try:
        data = json.loads(body)
    except Exception as e:
        log.warning("RappelConso JSON parse failed: %s", e)
        return []
    out = []
    for rec in data.get("results", []):
        pub = (rec.get("date_publication") or "")[:10]
        try:
            pub_d = date.fromisoformat(pub)
        except (ValueError, TypeError):
            continue
        if not (start <= pub_d <= end):
            continue
        if (rec.get("categorie_de_produit") or "") != "Alimentation":
            continue
        reason_blob = " ".join(str(rec.get(k, "") or "") for k in [
            "motif_du_rappel", "risques_encourus_par_le_consommateur",
        ])
        fid = rec.get("identifiant_unique_de_l_alerte") or rec.get("reference_fiche") or ""
        link = rec.get("lien_vers_la_fiche_rappel") or (
            f"https://rappel.conso.gouv.fr/fiche-rappel/{fid}/Interne" if fid else ""
        )
        out.append({
            "date": pub,
            "company": rec.get("nom_de_la_societe_responsable_de_la_commercialisation", ""),
            "product": (rec.get("noms_des_modeles_ou_references", "") or
                        rec.get("sous_categorie_de_produit", ""))[:200],
            "pathogen_raw": rec.get("risques_encourus_par_le_consommateur", "")[:150],
            "reason_raw": rec.get("motif_du_rappel", "")[:200],
            "url": link,
            "is_pathogen": is_pathogen_recall(reason_blob),
            "raw_category": rec.get("sous_categorie_de_produit", ""),
        })
    return out


def parse_fda_openfda(body: bytes, url: str, start: date, end: date) -> List[Dict[str, Any]]:
    """Parse openFDA enforcement.json."""
    try:
        data = json.loads(body)
    except Exception:
        return []
    out = []
    for rec in data.get("results", []):
        raw = rec.get("recall_initiation_date") or ""
        if len(raw) == 8 and raw.isdigit():
            pub = f"{raw[:4]}-{raw[4:6]}-{raw[6:8]}"
        else:
            pub = raw[:10]
        try:
            pub_d = date.fromisoformat(pub)
        except (ValueError, TypeError):
            continue
        if not (start <= pub_d <= end):
            continue
        reason = rec.get("reason_for_recall", "") or ""
        recall_num = (rec.get("recall_number") or "").strip()
        # Use the working fda.gov search URL (not the dead accessdata /scripts/ires)
        link = (f"https://www.fda.gov/safety/recalls-market-withdrawals-safety-alerts"
                f"?search_api_fulltext={recall_num}" if recall_num else
                "https://www.fda.gov/safety/recalls-market-withdrawals-safety-alerts")
        out.append({
            "date": pub,
            "company": rec.get("recalling_firm", ""),
            "product": (rec.get("product_description", "") or "")[:200],
            "pathogen_raw": reason[:150],
            "reason_raw": reason[:200],
            "url": link,
            "is_pathogen": is_pathogen_recall(reason),
            "raw_category": rec.get("product_type", ""),
        })
    return out


def parse_fsa_uk(body: bytes, url: str, start: date, end: date) -> List[Dict[str, Any]]:
    """Parse FSA UK food-alerts JSON-LD feed."""
    try:
        data = json.loads(body)
    except Exception:
        return []
    # FSA UK uses JSON-LD @graph / items structure
    items = data.get("items", data.get("@graph", []))
    out = []
    for rec in items:
        pub = (rec.get("created") or rec.get("datePublished") or "")[:10]
        try:
            pub_d = date.fromisoformat(pub)
        except (ValueError, TypeError):
            continue
        if not (start <= pub_d <= end):
            continue
        title = rec.get("title", "") or rec.get("name", "")
        if isinstance(title, dict):
            title = title.get("@value", "") or title.get("en", "")
        desc = rec.get("descriptionOfProblem", "") or rec.get("description", "") or ""
        if isinstance(desc, dict):
            desc = desc.get("@value", "") or desc.get("en", "")
        link = rec.get("@id") or rec.get("webpage") or ""
        blob = f"{title} {desc}"
        out.append({
            "date": pub,
            "company": (rec.get("business", "") or "")[:100],
            "product": str(title)[:200],
            "pathogen_raw": str(desc)[:150],
            "reason_raw": str(desc)[:200],
            "url": link if isinstance(link, str) else "",
            "is_pathogen": is_pathogen_recall(blob),
            "raw_category": "",
        })
    return out


def parse_generic_rss(body: bytes, url: str, start: date, end: date) -> List[Dict[str, Any]]:
    """Parse any RSS 2.0 or Atom feed."""
    try:
        root = ET.fromstring(body)
    except ET.ParseError as e:
        log.warning("RSS parse failed for %s: %s", url, e)
        return []
    out = []
    for item in root.iter("item"):
        title = (item.findtext("title") or "").strip()
        link = (item.findtext("link") or "").strip()
        desc = (item.findtext("description") or "").strip()
        desc = re.sub(r"<[^>]+>", " ", desc).strip()
        pub_raw = (item.findtext("pubDate") or "").strip()
        pub_d = _try_parse_rfc822(pub_raw) or _try_parse_iso(pub_raw)
        if not pub_d:
            continue
        if not (start <= pub_d <= end):
            continue
        blob = f"{title} {desc}"
        out.append({
            "date": pub_d.strftime("%Y-%m-%d"),
            "company": _extract_company(title),
            "product": title[:200],
            "pathogen_raw": title[:150],
            "reason_raw": desc[:200],
            "url": link,
            "is_pathogen": is_pathogen_recall(blob),
            "raw_category": "",
        })
    # Atom fallback
    if not out:
        for entry in root.iter():
            if not entry.tag.endswith("entry"):
                continue
            title = link = pub = desc = ""
            for child in entry:
                tag = child.tag.split("}", 1)[-1]
                if tag == "title":
                    title = (child.text or "").strip()
                elif tag == "link":
                    link = child.get("href") or (child.text or "").strip()
                elif tag in ("summary", "content"):
                    desc = re.sub(r"<[^>]+>", " ", (child.text or "")).strip()
                elif tag in ("published", "updated"):
                    pub = (child.text or "").strip()
            pub_d = _try_parse_iso(pub) or _try_parse_rfc822(pub)
            if not pub_d:
                continue
            if not (start <= pub_d <= end):
                continue
            blob = f"{title} {desc}"
            out.append({
                "date": pub_d.strftime("%Y-%m-%d"),
                "company": _extract_company(title),
                "product": title[:200],
                "pathogen_raw": title[:150],
                "reason_raw": desc[:200],
                "url": link,
                "is_pathogen": is_pathogen_recall(blob),
                "raw_category": "",
            })
    return out


def _try_parse_rfc822(s: str) -> Optional[date]:
    """RFC 822 → date."""
    if not s:
        return None
    try:
        from email.utils import parsedate_to_datetime
        dt = parsedate_to_datetime(s)
        return dt.date()
    except (TypeError, ValueError):
        return None


def _try_parse_iso(s: str) -> Optional[date]:
    """ISO 8601 → date."""
    if not s:
        return None
    for fmt in ("%Y-%m-%dT%H:%M:%S%z", "%Y-%m-%dT%H:%M:%SZ",
                "%Y-%m-%dT%H:%M:%S.%f%z", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _extract_company(title: str) -> str:
    m = re.match(r"^(.+?)\s+(?:recalls?|brand|recalled|may contain|withdraws?)\b",
                 title, re.I)
    if m:
        return m.group(1).strip()[:100]
    for sep in (" - ", " – ", ": "):
        if sep in title:
            return title.split(sep, 1)[0].strip()[:100]
    return title[:100]


# ---------------------------------------------------------------------------
# Recalls.xlsx loader — to compare against
# ---------------------------------------------------------------------------
def load_recalls_xlsx(xlsx_path: Path, start: date, end: date) -> List[Dict[str, Any]]:
    """Load Recalls sheet rows within date range. Returns list of {date, source, url, company, product}."""
    from openpyxl import load_workbook
    if not xlsx_path.exists():
        log.warning("recalls.xlsx not found at %s", xlsx_path)
        return []
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if "Recalls" not in wb.sheetnames:
        return []
    ws = wb["Recalls"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v in (None, "") for v in row):
            continue
        rec = dict(zip(headers, row))
        d_raw = str(rec.get("Date", ""))[:10]
        try:
            d = date.fromisoformat(d_raw)
        except (ValueError, TypeError):
            continue
        if not (start <= d <= end):
            continue
        out.append({
            "date": d_raw,
            "source": str(rec.get("Source") or ""),
            "url": str(rec.get("URL") or ""),
            "company": str(rec.get("Company") or ""),
            "product": str(rec.get("Product") or ""),
            "pathogen": str(rec.get("Pathogen") or ""),
        })
    return out


def load_pending_xlsx(xlsx_path: Path) -> List[Dict[str, Any]]:
    """Load Pending sheet — rows that were captured but not yet promoted."""
    from openpyxl import load_workbook
    if not xlsx_path.exists():
        return []
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if "Pending" not in wb.sheetnames:
        return []
    ws = wb["Pending"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v in (None, "") for v in row):
            continue
        rec = dict(zip(headers, row))
        out.append({
            "date": str(rec.get("Date", ""))[:10],
            "source": str(rec.get("Source") or ""),
            "url": str(rec.get("URL") or ""),
            "company": str(rec.get("Company") or ""),
            "notes": str(rec.get("Notes") or ""),
        })
    return out


# ---------------------------------------------------------------------------
# Matching: is this API-truth recall already in the Recalls sheet?
# ---------------------------------------------------------------------------
def normalize(s: str) -> str:
    s = (s or "").lower().strip()
    s = re.sub(r"\s+", " ", s)
    return s


def match_truth_to_captured(truth: Dict[str, Any],
                            captured: List[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    """Is this regulator-API recall present in the Recalls sheet?"""
    truth_url = normalize(truth["url"])
    truth_co = normalize(truth.get("company", ""))
    truth_prod = normalize(truth.get("product", ""))[:80]

    for row in captured:
        if truth_url and truth_url == normalize(row["url"]):
            return row
        # Fallback: company + first 40 chars of product
        if (truth_co and truth_co == normalize(row["company"]) and
                truth_prod[:40] == normalize(row["product"])[:40]):
            return row
    return None


def classify_miss(truth: Dict[str, Any],
                  pending: List[Dict[str, Any]],
                  all_recalls_same_agency_in_window: int) -> str:
    """Root-cause classification for a missed recall."""
    # In Pending sheet?
    truth_url = normalize(truth["url"])
    for p in pending:
        if truth_url and truth_url == normalize(p["url"]):
            notes = (p.get("notes") or "").lower()
            if "rejected" in notes and "url" in notes:
                return "URL_GATE_DROP"
            return "SCHEDULE_GAP"

    # Not in Pending — was pathogen correctly identified?
    if not truth["is_pathogen"]:
        return "NOT_IN_WHITELIST"

    # Was the scraper for this agency producing ANY rows in the window?
    if all_recalls_same_agency_in_window == 0:
        return "SCRAPER_DEAD"

    # Scraper ran and produced rows, but missed this one → filter mismatch
    return "SCRAPER_FILTERED"


# ---------------------------------------------------------------------------
# Main audit logic
# ---------------------------------------------------------------------------
@dataclass
class AgencyResult:
    agency: str
    country: str
    truth_count: int = 0               # total recalls from API in window
    pathogen_truth_count: int = 0      # subset that are pathogens
    captured_count: int = 0            # in Recalls sheet, same agency + window
    missed: List[Dict[str, Any]] = field(default_factory=list)
    fetch_error: Optional[str] = None


def run_audit(start: date, end: date, xlsx_path: Path) -> List[AgencyResult]:
    """Fetch every regulator, compare to recalls.xlsx, return per-agency results."""
    captured_all = load_recalls_xlsx(xlsx_path, start, end)
    pending = load_pending_xlsx(xlsx_path)
    log.info("Loaded %d rows from Recalls sheet in window %s..%s",
             len(captured_all), start, end)
    log.info("Loaded %d rows from Pending sheet", len(pending))

    # Parsers dispatch map
    parsers = {
        "parse_rappelconso": parse_rappelconso,
        "parse_fda_openfda": parse_fda_openfda,
        "parse_fsa_uk": parse_fsa_uk,
        "parse_generic_rss": parse_generic_rss,
    }

    session = make_session(timeout=30)
    results: List[AgencyResult] = []

    for reg in REGULATORS:
        log.info("%s — %s", reg.agency, reg.url[:80])
        res = AgencyResult(agency=reg.agency, country=reg.country)
        captured_this = [c for c in captured_all if c["source"] == reg.agency]
        res.captured_count = len(captured_this)

        # Build URL with params (for APIs that need filtering)
        fetch_url = reg.url
        fetch_params = {}
        if reg.parser == "parse_rappelconso":
            fetch_params = {
                "where": (f'date_publication >= "{start.isoformat()}" AND '
                          f'date_publication <= "{end.isoformat()}" AND '
                          f'categorie_de_produit = "Alimentation"'),
                "limit": 100,
                "order_by": "date_publication DESC",
            }
        elif reg.parser == "parse_fda_openfda":
            start_fda = start.strftime("%Y%m%d")
            end_fda = end.strftime("%Y%m%d")
            fetch_url = (f"{reg.url}?search=recall_initiation_date:[{start_fda}+TO+{end_fda}]"
                         f"&limit=100")

        try:
            r = fetch(session, fetch_url, params=fetch_params if fetch_params else None)
            if r is None:
                res.fetch_error = "network failure / timeout"
                log.warning("  -> %s FAILED to fetch", reg.agency)
                results.append(res)
                continue
            if r.status_code != 200:
                res.fetch_error = f"HTTP {r.status_code}"
                log.warning("  -> %s HTTP %d", reg.agency, r.status_code)
                results.append(res)
                continue
            truths = parsers[reg.parser](r.content, fetch_url, start, end)
        except Exception as e:
            res.fetch_error = f"{type(e).__name__}: {e}"
            log.warning("  -> %s parser error: %s", reg.agency, e)
            results.append(res)
            continue

        res.truth_count = len(truths)
        res.pathogen_truth_count = sum(1 for t in truths if t["is_pathogen"])

        # Match truth → captured, classify misses
        for truth in truths:
            match = match_truth_to_captured(truth, captured_this)
            if match:
                continue
            reason = classify_miss(truth, pending, res.captured_count)
            miss_row = {
                **truth,
                "agency": reg.agency,
                "country": reg.country,
                "miss_reason": reason,
            }
            res.missed.append(miss_row)

        log.info("  -> API=%d  pathogen=%d  captured=%d  missed=%d",
                 res.truth_count, res.pathogen_truth_count,
                 res.captured_count, len(res.missed))
        results.append(res)

    return results


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------
def write_excel_report(results: List[AgencyResult], out_path: Path,
                       start: date, end: date, xlsx_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ---- Styles (consistent across all sheets) ----
    HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
    BODY_FONT = Font(name="Arial", size=10)
    THIN = Side(style="thin", color="CCCCCC")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    FILL_MISS_DEAD      = PatternFill("solid", fgColor="F8CBAD")  # orange
    FILL_MISS_FILTERED  = PatternFill("solid", fgColor="FFE699")  # yellow
    FILL_MISS_URL_GATE  = PatternFill("solid", fgColor="F4B183")  # darker orange
    FILL_MISS_SCHEDULE  = PatternFill("solid", fgColor="C6E0B4")  # green (expected)
    FILL_OK             = PatternFill("solid", fgColor="E2EFDA")  # light green
    FILL_OFF_WHITELIST  = PatternFill("solid", fgColor="D9D9D9")  # grey
    MISS_FILLS = {
        "SCRAPER_DEAD":     FILL_MISS_DEAD,
        "SCRAPER_FILTERED": FILL_MISS_FILTERED,
        "URL_GATE_DROP":    FILL_MISS_URL_GATE,
        "SCHEDULE_GAP":     FILL_MISS_SCHEDULE,
        "NOT_IN_WHITELIST": FILL_OFF_WHITELIST,
        "TRULY_MISSING":    FILL_MISS_DEAD,
    }

    def style_header(ws, row=1):
        for cell in ws[row]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = BORDER

    # ==== SHEET 1: Summary ====
    s = wb.active
    s.title = "Summary"
    s["A1"] = f"Gap Audit: {start.isoformat()} to {end.isoformat()}"
    s["A1"].font = Font(name="Arial", size=14, bold=True)
    s.merge_cells("A1:H1")
    s["A2"] = f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
    s["A2"].font = Font(name="Arial", size=10, italic=True)
    s.merge_cells("A2:H2")
    s["A3"] = f"Data source: {xlsx_path}"
    s["A3"].font = Font(name="Arial", size=10, italic=True)
    s.merge_cells("A3:H3")

    hdr_row = 5
    headers = [
        "Agency", "Country", "Fetch status",
        "Regulator total", "Regulator pathogen",
        "Pipeline captured", "Gap (# missed)",
        "Capture rate (%)",
    ]
    for i, h in enumerate(headers, 1):
        s.cell(row=hdr_row, column=i, value=h)
    style_header(s, hdr_row)

    for idx, res in enumerate(results, start=hdr_row + 1):
        s.cell(row=idx, column=1, value=res.agency).font = BODY_FONT
        s.cell(row=idx, column=2, value=res.country).font = BODY_FONT
        s.cell(row=idx, column=3, value=(res.fetch_error or "OK")).font = BODY_FONT
        s.cell(row=idx, column=4, value=res.truth_count).font = BODY_FONT
        s.cell(row=idx, column=5, value=res.pathogen_truth_count).font = BODY_FONT
        s.cell(row=idx, column=6, value=res.captured_count).font = BODY_FONT
        s.cell(row=idx, column=7, value=len(res.missed)).font = BODY_FONT
        # Capture-rate formula (Excel does the math)
        formula = (
            f'=IFERROR(IF(E{idx}=0,"-",ROUND(F{idx}/E{idx}*100,1)&"%"),"-")'
        )
        s.cell(row=idx, column=8, value=formula).font = BODY_FONT

        # Highlight gaps
        if res.fetch_error:
            for col in range(1, 9):
                s.cell(row=idx, column=col).fill = FILL_MISS_DEAD
        elif len(res.missed) == 0:
            for col in range(1, 9):
                s.cell(row=idx, column=col).fill = FILL_OK
        elif len(res.missed) > 0:
            s.cell(row=idx, column=7).fill = FILL_MISS_FILTERED

        for col in range(1, 9):
            s.cell(row=idx, column=col).border = BORDER

    # Totals row
    tot_row = hdr_row + 1 + len(results)
    s.cell(row=tot_row, column=1, value="TOTAL").font = Font(name="Arial", size=11, bold=True)
    for col in (4, 5, 6, 7):
        letter = get_column_letter(col)
        s.cell(row=tot_row, column=col,
               value=f"=SUM({letter}{hdr_row+1}:{letter}{tot_row-1})")
        s.cell(row=tot_row, column=col).font = Font(name="Arial", size=11, bold=True)
    s.cell(row=tot_row, column=8,
           value=f'=IFERROR(ROUND(F{tot_row}/E{tot_row}*100,1)&"%","-")')
    s.cell(row=tot_row, column=8).font = Font(name="Arial", size=11, bold=True)
    for col in range(1, 9):
        s.cell(row=tot_row, column=col).fill = PatternFill("solid", fgColor="D9E1F2")
        s.cell(row=tot_row, column=col).border = BORDER

    # Column widths
    col_widths = [22, 18, 25, 15, 17, 17, 14, 17]
    for i, w in enumerate(col_widths, 1):
        s.column_dimensions[get_column_letter(i)].width = w

    # ==== SHEET 2: All Misses (flat) ====
    ms = wb.create_sheet("All Misses")
    ms["A1"] = "Every pathogen recall the regulator published that your pipeline did not capture"
    ms["A1"].font = Font(name="Arial", size=12, bold=True)
    ms.merge_cells("A1:I1")

    hdr2 = ["Date", "Agency", "Country", "Company", "Product",
            "Pathogen (from agency)", "Why missed", "URL", "Reason detail"]
    for i, h in enumerate(hdr2, 1):
        ms.cell(row=3, column=i, value=h)
    style_header(ms, 3)

    r = 4
    for res in results:
        for miss in res.missed:
            if not miss["is_pathogen"] and miss["miss_reason"] == "NOT_IN_WHITELIST":
                continue  # don't pollute this sheet with off-whitelist items
            ms.cell(row=r, column=1, value=miss["date"]).font = BODY_FONT
            ms.cell(row=r, column=2, value=miss["agency"]).font = BODY_FONT
            ms.cell(row=r, column=3, value=miss["country"]).font = BODY_FONT
            ms.cell(row=r, column=4, value=miss["company"]).font = BODY_FONT
            ms.cell(row=r, column=5, value=miss["product"]).font = BODY_FONT
            ms.cell(row=r, column=6, value=miss["pathogen_raw"]).font = BODY_FONT
            ms.cell(row=r, column=7, value=miss["miss_reason"]).font = BODY_FONT
            ms.cell(row=r, column=8, value=miss["url"]).font = BODY_FONT
            ms.cell(row=r, column=9, value=miss["reason_raw"]).font = BODY_FONT
            fill = MISS_FILLS.get(miss["miss_reason"], FILL_MISS_FILTERED)
            for c in range(1, 10):
                ms.cell(row=r, column=c).fill = fill
                ms.cell(row=r, column=c).border = BORDER
            r += 1

    if r == 4:
        ms["A4"] = "No pathogen misses found — every regulator-published recall in this window was captured."
        ms["A4"].font = Font(name="Arial", size=10, italic=True)

    col_widths2 = [12, 22, 16, 25, 40, 25, 18, 45, 35]
    for i, w in enumerate(col_widths2, 1):
        ms.column_dimensions[get_column_letter(i)].width = w
    ms.freeze_panes = "A4"

    # ==== SHEET 3: Root-cause legend ====
    legend = wb.create_sheet("How to read")
    legend["A1"] = "Miss-reason classification guide"
    legend["A1"].font = Font(name="Arial", size=14, bold=True)

    rows = [
        ("SCRAPER_DEAD", "The agency's scraper returned ZERO rows for this window even "
         "though the regulator's API had data. This is the most serious class of miss — "
         "indicates broken scraper code or dead INDEX_URL. Fix: update the scraper."),
        ("SCRAPER_FILTERED", "The scraper DID run and produced rows, but this specific recall "
         "was filtered out. Usually means the pathogen keyword list missed a "
         "language/spelling variant, or the scraper's date filter was too narrow. "
         "Fix: add keyword to PATHOGEN_KEYWORDS in the scraper or relax the filter."),
        ("URL_GATE_DROP", "The scraper captured this row and sent it to Pending, but the "
         "07:30 URL gate rejected it because the URL returned HTTP 404/5xx. "
         "Fix: update the scraper's URL template (e.g. FDA accessdata fix), or run "
         "pipeline/url_resurrect.py to have OpenAI propose a working URL."),
        ("SCHEDULE_GAP", "The recall is in Pending right now but hasn't been promoted yet. "
         "Next URL gate run will pick it up. Not a bug — just timing."),
        ("NOT_IN_WHITELIST", "Agency published this as a recall, but it's an allergen-only, "
         "foreign-object, or chemical (non-pathogen) recall. Your pipeline correctly "
         "excludes these per the FSIS pathogen whitelist. Not a miss."),
        ("TRULY_MISSING", "None of the above explain it — genuine pipeline bug worth "
         "investigating manually."),
    ]
    for i, (code, desc) in enumerate(rows, start=3):
        legend.cell(row=i, column=1, value=code).font = Font(name="Arial", size=10, bold=True)
        legend.cell(row=i, column=1).fill = MISS_FILLS.get(code, FILL_MISS_FILTERED)
        legend.cell(row=i, column=2, value=desc).font = BODY_FONT
        legend.cell(row=i, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    legend.column_dimensions["A"].width = 22
    legend.column_dimensions["B"].width = 110
    for i in range(3, 3 + len(rows)):
        legend.row_dimensions[i].height = 60

    wb.save(out_path)
    log.info("Wrote Excel report: %s", out_path)


def write_csv_report(results: List[AgencyResult], out_path: Path) -> None:
    """Flat CSV with one row per miss."""
    with out_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Date", "Agency", "Country", "Company", "Product",
                    "Pathogen", "Miss reason", "URL", "Reason detail"])
        for res in results:
            for miss in res.missed:
                if not miss["is_pathogen"] and miss["miss_reason"] == "NOT_IN_WHITELIST":
                    continue
                w.writerow([
                    miss["date"], miss["agency"], miss["country"],
                    miss["company"], miss["product"], miss["pathogen_raw"],
                    miss["miss_reason"], miss["url"], miss["reason_raw"],
                ])
    log.info("Wrote CSV report: %s", out_path)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--days", type=int, default=7,
                    help="Look back N days from today")
    ap.add_argument("--start", help="Start date YYYY-MM-DD (overrides --days)")
    ap.add_argument("--end", help="End date YYYY-MM-DD")
    ap.add_argument("--xlsx", default=str(ROOT / "docs" / "data" / "recalls.xlsx"),
                    help="Path to recalls.xlsx")
    ap.add_argument("--output", default=None,
                    help="Output xlsx path (default gap_report_YYYY-MM-DD.xlsx)")
    args = ap.parse_args()

    today = date.today()
    if args.start:
        start = date.fromisoformat(args.start)
        end = date.fromisoformat(args.end) if args.end else today
    else:
        end = today
        start = end - timedelta(days=args.days - 1)

    log.info("Auditing %s to %s (%d days)", start, end, (end - start).days + 1)

    results = run_audit(start, end, Path(args.xlsx))

    out_xlsx = Path(args.output) if args.output else \
        ROOT / f"gap_report_{end.isoformat()}.xlsx"
    out_csv = out_xlsx.with_suffix(".csv")
    write_excel_report(results, out_xlsx, start, end, Path(args.xlsx))
    write_csv_report(results, out_csv)

    # Console summary
    print()
    print("=" * 70)
    print(f"GAP AUDIT — {start} to {end}")
    print("=" * 70)
    print(f"{'Agency':<24} {'Truth':>7} {'Pathogen':>9} {'Captured':>9} {'Missed':>7}")
    print("-" * 70)
    total_truth = total_path = total_cap = total_miss = 0
    for res in results:
        marker = "  ⚠️" if res.fetch_error else ""
        print(f"{res.agency:<24} {res.truth_count:>7} {res.pathogen_truth_count:>9} "
              f"{res.captured_count:>9} {len(res.missed):>7}{marker}")
        total_truth += res.truth_count
        total_path += res.pathogen_truth_count
        total_cap += res.captured_count
        total_miss += len(res.missed)
    print("-" * 70)
    print(f"{'TOTAL':<24} {total_truth:>7} {total_path:>9} {total_cap:>9} {total_miss:>7}")
    print()
    print(f"Reports: {out_xlsx.name} + {out_csv.name}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
