"""
URL Guardian — the every-4-hour integrity pass over docs/data/recalls.xlsx.

What it does, in order:
  1. Load recalls.xlsx
  2. Validate URLs for all rows in the last N days (default 14)
  3. Blank URLs that are truly broken (404/410/5xx) or generic landing pages.
     Keep 403s from known gov/regulatory domains (bot-blocks, not breakage).
  4. Ask OpenAI's find_missing_recalls() what recalls may have been missed
     per major agency over the last M days (default 3). Append novel rows.
  5. Optionally ask Claude Haiku to spot-check Tier-1 rows for pathogen/outbreak
     mis-classification (uses the $5 credit sparingly).
  6. Write back: xlsx + json.
  7. Return a summary dict for logging.

Usage:
    from pipeline.url_guardian import guardian_run
    summary = guardian_run("docs/data/recalls.xlsx")
    print(summary)

Environment variables:
    OPENAI_API_KEY     — required for gap-finder
    ANTHROPIC_API_KEY  — optional, used for Tier-1 spot checks
    GUARDIAN_SINCE_DAYS     — URL-check window (default 14)
    GUARDIAN_GAP_DAYS       — gap-finder window (default 3)
    GUARDIAN_SKIP_AI        — "true" skips gap-finder and Tier-1 review
"""
from __future__ import annotations
import os
import json
import logging
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import List, Dict, Any, Tuple

from openpyxl import load_workbook, Workbook

# Imports from the review package (sibling modules in the user's repo)
try:
    from review.url_validator import validate_all, should_blank_url, is_generic_url
except ImportError:
    # When run as a standalone script from the review/ folder
    from url_validator import validate_all, should_blank_url, is_generic_url  # type: ignore

try:
    from review.openai_client import find_missing_recalls
except ImportError:
    from openai_client import find_missing_recalls  # type: ignore

try:
    from review.claude_client import review_tier1
except ImportError:
    try:
        from claude_client import review_tier1  # type: ignore
    except ImportError:
        review_tier1 = None  # optional

log = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

# Agencies the gap-finder should ask about. Kept small on purpose: one LLM call
# per agency, so cost scales linearly. Expand gradually as the budget allows.
GAP_FINDER_TARGETS: List[Tuple[str, str]] = [
    ("United States", "FDA"),
    ("United States", "USDA FSIS"),
    ("Canada",        "CFIA"),
    ("European Union", "RASFF"),
    ("United Kingdom", "FSA"),
    ("France",        "RappelConso"),
    ("Germany",       "BVL"),
    ("Australia",     "FSANZ"),
]

COLUMNS = ["Date", "Source", "Company", "Brand", "Product", "Pathogen",
           "Reason", "Class", "Country", "Region", "Tier", "Outbreak", "URL", "Notes"]


# --- IO helpers -------------------------------------------------------------
def _load_xlsx(xlsx_path: Path) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[str]]:
    """Returns (recalls, news, headers). News preserved as-is; not touched by guardian."""
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["Recalls"]
    headers = [c.value for c in ws[1]]
    recalls = [dict(zip(headers, r)) for r in ws.iter_rows(min_row=2, values_only=True)]
    news = []
    if "NEWS" in wb.sheetnames:
        nw = wb["NEWS"]
        nh = [c.value for c in nw[1]]
        news = [dict(zip(nh, r)) for r in nw.iter_rows(min_row=2, values_only=True)]
    return recalls, news, headers


def _write_xlsx(xlsx_path: Path, recalls: List[Dict[str, Any]], news: List[Dict[str, Any]],
                headers: List[str]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Recalls"
    ws.append(headers)
    for r in recalls:
        ws.append([r.get(h, "") for h in headers])
    # Recreate NEWS sheet
    nw = wb.create_sheet("NEWS")
    if news:
        news_headers = list(news[0].keys())
        nw.append(news_headers)
        for r in news:
            nw.append([r.get(h, "") for h in news_headers])
    else:
        nw.append(["Published (UTC)", "Pathogen", "Event", "Source", "Title", "Link", "Retrieved (UTC)"])
    wb.save(xlsx_path)


def _write_json(json_path: Path, recalls: List[Dict[str, Any]]):
    """Mirror the xlsx to json for the dashboard's JSON loader (if present)."""
    serializable = []
    for r in recalls:
        rec = {}
        for k, v in r.items():
            if isinstance(v, (date, datetime)):
                rec[k] = v.isoformat()[:10]
            elif v is None:
                rec[k] = ""
            else:
                rec[k] = v
        serializable.append(rec)
    json_path.write_text(json.dumps(serializable, ensure_ascii=False, indent=2), encoding="utf-8")


# --- core logic -------------------------------------------------------------
def _in_window(row: Dict[str, Any], cutoff: date) -> bool:
    d = str(row.get("Date") or "")[:10]
    if not d:
        return False
    try:
        return datetime.strptime(d, "%Y-%m-%d").date() >= cutoff
    except ValueError:
        return False


def _url_health_pass(recalls: List[Dict[str, Any]], since_days: int) -> Dict[str, int]:
    """Validate URLs on recent rows. Blank ones that are truly broken/generic."""
    cutoff = date.today() - timedelta(days=since_days)
    targets = [(i, r) for i, r in enumerate(recalls) if _in_window(r, cutoff)]
    log.info("URL health: checking %d of %d rows (last %d days)",
             len(targets), len(recalls), since_days)

    rows_to_check = [r for _, r in targets]
    validated = validate_all(rows_to_check, max_workers=10)

    stats = {"checked": len(validated), "ok": 0, "bot_blocked": 0,
             "blanked_generic": 0, "blanked_404": 0, "blanked_5xx": 0, "kept_403": 0}
    for (global_idx, _), vrow in zip(targets, validated):
        check = vrow.get("_url_check", {})
        reason = check.get("reason", "")
        if reason == "ok":
            stats["ok"] += 1
            continue
        if reason == "bot_blocked":
            stats["bot_blocked"] += 1
            continue
        if should_blank_url(check):
            original = recalls[global_idx].get("URL", "")
            if reason == "generic":
                stats["blanked_generic"] += 1
            elif 400 <= check.get("status", 0) < 500:
                stats["blanked_404"] += 1
            else:
                stats["blanked_5xx"] += 1
            recalls[global_idx]["URL"] = ""
            # Append audit trail in Notes (truncate if too long)
            notes = str(recalls[global_idx].get("Notes") or "")
            audit = f"[URL-guardian {date.today().isoformat()}: blanked {reason} {original[:60]}]"
            recalls[global_idx]["Notes"] = (notes + " " + audit).strip()[:500]
        else:
            # 403 that isn't a known bot-hostile domain — log but keep
            stats["kept_403"] += 1
    return stats


def _signature(r: Dict[str, Any]) -> str:
    """De-dup key: date | company | pathogen | country, lowercased."""
    d = str(r.get("Date") or "")[:10]
    c = (str(r.get("Company") or "")).strip().lower()
    p = (str(r.get("Pathogen") or "")).strip().lower()
    co = (str(r.get("Country") or "")).strip().lower()
    return f"{d}|{c}|{p}|{co}"


def _gap_finder_pass(recalls: List[Dict[str, Any]], gap_days: int) -> Dict[str, int]:
    """Ask OpenAI what recalls we may have missed. Append novel, non-duplicate rows."""
    if os.getenv("OPENAI_API_KEY", "").strip() == "":
        log.info("Gap-finder: OPENAI_API_KEY missing, skipping")
        return {"suggested": 0, "added": 0, "dupes_rejected": 0}

    existing_sigs = {_signature(r) for r in recalls}
    added = 0
    suggested = 0
    dupes = 0
    today = date.today().isoformat()

    for country, agency in GAP_FINDER_TARGETS:
        try:
            candidates = find_missing_recalls(country, agency, since_days=gap_days) or []
        except Exception as e:
            log.warning("Gap-finder %s/%s failed: %s", country, agency, e)
            continue
        suggested += len(candidates)
        for c in candidates:
            # Skip if URL is obviously generic
            url = (c.get("URL") or "").strip()
            if not url or is_generic_url(url):
                continue
            new_row = {
                "Date":    (c.get("Date") or today)[:10],
                "Source":  f"{agency} (gap-finder)",
                "Company": c.get("Company", ""),
                "Brand":   "",
                "Product": c.get("Product", ""),
                "Pathogen": c.get("Pathogen", ""),
                "Reason":  c.get("Reason", ""),
                "Class":   "",
                "Country": country,
                "Region":  "",
                "Tier":    2,  # default, promotion happens via Claude review
                "Outbreak": 0,
                "URL":     url,
                "Notes":   f"[gap-finder {today} via OpenAI]",
            }
            sig = _signature(new_row)
            if sig in existing_sigs:
                dupes += 1
                continue
            existing_sigs.add(sig)
            recalls.append(new_row)
            added += 1
    log.info("Gap-finder: suggested=%d added=%d dupes_rejected=%d", suggested, added, dupes)
    return {"suggested": suggested, "added": added, "dupes_rejected": dupes}


def _tier1_spot_check(recalls: List[Dict[str, Any]], since_days: int) -> Dict[str, int]:
    """Optional Claude Haiku review of recent Tier-1 rows. Free-credit budget."""
    if review_tier1 is None:
        return {"flags": 0, "reviewed": 0}
    if os.getenv("ANTHROPIC_API_KEY", "").strip() == "":
        return {"flags": 0, "reviewed": 0}
    cutoff = date.today() - timedelta(days=since_days)
    recent_tier1 = [r for r in recalls
                    if int(r.get("Tier") or 2) == 1 and _in_window(r, cutoff)]
    if not recent_tier1:
        return {"flags": 0, "reviewed": 0}
    flags = review_tier1(recent_tier1) or []
    log.info("Claude Tier-1 spot-check: reviewed=%d flags=%d", len(recent_tier1), len(flags))
    # Persist high-severity flags into Notes for human review
    for f in flags:
        if f.get("severity") == "high":
            idx = f.get("row_index", -1)
            if 0 <= idx < len(recent_tier1):
                row = recent_tier1[idx]
                note = f" [Claude-flag {date.today().isoformat()}: {f.get('issue','')[:80]}]"
                row["Notes"] = (str(row.get("Notes") or "") + note).strip()[:500]
    return {"flags": len(flags), "reviewed": len(recent_tier1)}


# --- entry point ------------------------------------------------------------
def guardian_run(xlsx_path: str = "docs/data/recalls.xlsx",
                 json_path: str = "docs/data/recalls.json",
                 since_days: int = None,
                 gap_days: int = None,
                 skip_ai: bool = False) -> Dict[str, Any]:
    """Main 4-hour guardian pass. Returns summary dict."""
    since_days = int(since_days if since_days is not None
                     else os.getenv("GUARDIAN_SINCE_DAYS", "14"))
    gap_days = int(gap_days if gap_days is not None
                   else os.getenv("GUARDIAN_GAP_DAYS", "3"))
    skip_ai = skip_ai or os.getenv("GUARDIAN_SKIP_AI", "false").lower() == "true"

    xp = Path(xlsx_path)
    jp = Path(json_path)
    if not xp.exists():
        log.error("xlsx not found: %s", xp)
        return {"ok": False, "error": "xlsx not found"}

    recalls, news, headers = _load_xlsx(xp)
    before_count = len(recalls)
    log.info("Loaded %d recalls, %d news rows from %s", before_count, len(news), xp)

    # Ensure we always write the canonical column order, even if input is quirky
    headers = COLUMNS if set(headers) >= set(COLUMNS) - {"Notes"} else headers

    # 1. URL health pass — isolated so a network glitch won't kill the pipeline
    try:
        url_stats = _url_health_pass(recalls, since_days)
    except Exception as e:
        log.warning("URL health pass failed (non-fatal): %s", e)
        url_stats = {"checked": 0, "blanked": 0, "kept_bot_block": 0, "error": str(e)}

    # 2. Gap-finder + Tier-1 spot-check (AI passes) — also isolated
    if skip_ai:
        gap_stats = {"suggested": 0, "added": 0, "dupes_rejected": 0, "skipped": True}
        tier1_stats = {"flags": 0, "reviewed": 0, "skipped": True}
    else:
        try:
            gap_stats = _gap_finder_pass(recalls, gap_days)
        except Exception as e:
            log.warning("Gap-finder pass failed (non-fatal): %s", e)
            gap_stats = {"suggested": 0, "added": 0, "dupes_rejected": 0, "error": str(e)}
        try:
            tier1_stats = _tier1_spot_check(recalls, since_days)
        except Exception as e:
            log.warning("Tier-1 spot-check failed (non-fatal): %s", e)
            tier1_stats = {"flags": 0, "reviewed": 0, "error": str(e)}

    # 3. Sort newest-first for writing (matches dashboard expectations)
    recalls.sort(key=lambda r: str(r.get("Date") or ""), reverse=True)

    # 4. Write back
    _write_xlsx(xp, recalls, news, headers)
    try:
        _write_json(jp, recalls)
    except Exception as e:
        log.warning("json write failed (non-fatal): %s", e)

    summary = {
        "ok": True,
        "timestamp": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
        "before_rows": before_count,
        "after_rows": len(recalls),
        "delta_rows": len(recalls) - before_count,
        "url_health": url_stats,
        "gap_finder": gap_stats,
        "tier1_review": tier1_stats,
    }
    log.info("Guardian summary: %s", json.dumps(summary, indent=2))
    return summary


if __name__ == "__main__":
    import sys
    import traceback
    import argparse
    ap = argparse.ArgumentParser(description="FSIS URL Guardian — 4-hour integrity pass")
    ap.add_argument("--xlsx", default="docs/data/recalls.xlsx")
    ap.add_argument("--json", default="docs/data/recalls.json")
    ap.add_argument("--since-days", type=int, default=None)
    ap.add_argument("--gap-days", type=int, default=None)
    ap.add_argument("--skip-ai", action="store_true")
    args = ap.parse_args()

    # Diagnostic banner — makes failure causes visible in the workflow log.
    log.info("=" * 70)
    log.info("FSIS URL Guardian starting")
    log.info("Python:   %s", sys.version.replace("\n", " "))
    log.info("CWD:      %s", Path.cwd())
    log.info("xlsx arg: %s", args.xlsx)
    xp = Path(args.xlsx)
    log.info("xlsx exists: %s  (size=%s bytes)",
             xp.exists(),
             xp.stat().st_size if xp.exists() else "n/a")
    log.info("OPENAI_API_KEY:    %s", "set" if os.getenv("OPENAI_API_KEY") else "NOT SET")
    log.info("ANTHROPIC_API_KEY: %s", "set" if os.getenv("ANTHROPIC_API_KEY") else "NOT SET")
    log.info("=" * 70)

    if not xp.exists():
        log.error("FATAL: xlsx file not found at %s", xp.absolute())
        log.error("CWD contents: %s", sorted(p.name for p in Path.cwd().iterdir())[:20])
        sys.exit(1)

    try:
        result = guardian_run(args.xlsx, args.json,
                              args.since_days, args.gap_days, args.skip_ai)
        if not isinstance(result, dict) or not result.get("ok"):
            log.error("Guardian run returned failure: %s", result)
            sys.exit(1)
        log.info("Guardian completed successfully.")
        sys.exit(0)
    except Exception as exc:
        log.error("FATAL uncaught exception in guardian_run: %s: %s",
                  type(exc).__name__, exc)
        traceback.print_exc()
        sys.exit(1)
