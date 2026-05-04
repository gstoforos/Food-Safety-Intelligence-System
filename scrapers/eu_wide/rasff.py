"""
RASFF Window — EU Rapid Alert System for Food and Feed.

Architecture (as of 2026-05-04, replacing the 2026-04-29 disabled stub):

  RASFF Window's web UI exposes two XLS/CSV download buttons on the search/
  list page that return the FULL notification dataset (~30K rows) as a
  structured spreadsheet. This bypasses the SPA hydration problem entirely
  — we don't need to scrape the rendered DOM, we just download the export.

  Endpoint format (audit 2026-05-04, sourced from XLS button HAR capture):
    https://webgate.ec.europa.eu/rasff-window/services/excel
    https://webgate.ec.europa.eu/rasff-window/services/csv
    Both accept the same query parameters as the /screen/list URL.

  The bulk download contains 14 columns:
    reference, category, type, subject, date, notifying_country,
    classification, risk_decision, distribution, forAttention, forFollowUp,
    operator, origin, hazards

  We filter:
    - type == 'food'                     (drop feed, food contact materials)
    - classification in {alert, border}  (configurable; user-selected scope)
    - date >= today - since_days
    - hazards/subject contain a pathogen we monitor (PATHOGEN_RULES)

  RASFF rows are written following the existing 15-row schema in the xlsx:
    Source   = 'RASFF'
    Company  = 'Origin: X | Notifying: Y | Distributed: Z'   (per pipeline)
    Brand    = '—' (RASFF doesn't publish brand info publicly)
    Country  = origin country (canonical)
    URL      = https://webgate.ec.europa.eu/rasff-window/screen/notification/{reference}

  Note on URL format: existing manually-added RASFF rows in our corpus use
  integer IDs (e.g. /notification/838838) which are RASFF's internal DB
  primary keys. Those IDs are not exposed in the public XLS export. We use
  the public REFERENCE form (e.g. /notification/2026.3863) instead.
  pipeline.run_all._RASFF_NOTIFICATION_URL_RE is updated in lockstep to
  accept both forms.

Resilience:

  If RASFF removes or changes the XLS endpoint, this scraper will throw
  HTTP 404/410 and log a clear error. The stub fallback path returns
  zero rows, matching the disabled-state behavior. Country scrapers
  continue to provide consumer-facing recall coverage independently.

  We do NOT use Playwright. The XLS export approach is simpler, faster
  (~3s download vs 30-60s SPA hydration), and avoids adding chromium
  to the GitHub Actions runner image.
"""
from __future__ import annotations

import io
import logging
import re
from datetime import datetime, timedelta, timezone
from typing import List, Optional, Tuple

import openpyxl

from scrapers._base import BaseScraper, fetch
from scrapers._models import Recall, normalize_pathogen, normalize_country, infer_region

log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Tunables
# ---------------------------------------------------------------------------

# Public XLS download endpoint discovered 2026-05-04. The /screen/list page's
# "XLS" button posts to this URL with current filter state. With no filters,
# the response is the full corpus (~30K rows, ~3MB compressed).
XLS_URL = "https://webgate.ec.europa.eu/rasff-window/services/excel"

# Some servers also accept a CSV alias. Kept as fallback.
CSV_URL = "https://webgate.ec.europa.eu/rasff-window/services/csv"

# Categories we monitor — all "alert notification" + all "border rejection"
# notifications (alert + IFA distribution variants are picked up by national
# scrapers; border rejections are RASFF-exclusive).
MONITORED_CLASSIFICATIONS = (
    "alert notification",
    "border rejection",
    # Information notifications often fail the pathogen filter anyway, but
    # we keep them in scope so we capture them when they DO match a pathogen
    # (e.g. the 2026.3853 Salmonella in Polish chicken row).
    "information notification for attention",
    "information notification for follow-up",
)

# We monitor only food (drop feed, food contact materials, environmental).
MONITORED_TYPES = ("food",)

# Lookback window. Default 7 days, matching other scrapers.
DEFAULT_SINCE_DAYS = 7

# Per-row cap on Reason text length (matches existing RASFF row style).
REASON_MAX_LEN = 500


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _parse_rasff_date(s: str) -> Optional[datetime]:
    """RASFF XLS date format is 'DD-MM-YYYY HH:MM:SS' — return UTC datetime."""
    if not s:
        return None
    try:
        return datetime.strptime(str(s)[:19], "%d-%m-%Y %H:%M:%S").replace(
            tzinfo=timezone.utc
        )
    except (ValueError, TypeError):
        try:
            # Fallback: date-only format
            return datetime.strptime(str(s)[:10], "%d-%m-%Y").replace(
                tzinfo=timezone.utc
            )
        except (ValueError, TypeError):
            return None


def _extract_pathogen(subject: str, hazards: str) -> str:
    """
    Run normalize_pathogen across the subject + hazards strings. Returns
    canonical name if any pathogen is matched, else "" (row is rejected).

    The hazards field has the format:
        'cereulide  - {natural toxins (other)}'
        'Aflatoxin B1   - {mycotoxins},aflatoxin total  - {mycotoxins}'
        'Salmonella Enteritidis  - {pathogenic micro-organisms}'

    We test the hazards field FIRST (more reliable signal), then fall back
    to subject (catches RASFF rows where hazards is empty but subject mentions
    the contaminant — e.g. row 2026.3863 "Cerulide in infant formula").

    RASFF-side pre-normalisation: the public XLS export contains documented
    typos (e.g. "Cerulide" missing one 'e') that miss the strict regex in
    PATHOGEN_RULES. We pre-correct known typos BEFORE feeding to the canonical
    normalizer, so the global vocabulary stays clean.
    """
    for text in (hazards or "", subject or ""):
        if not text:
            continue
        # RASFF typo corrections (subject-only, not hazards)
        corrected = _correct_rasff_typos(text)
        canon = normalize_pathogen(corrected)
        if canon:
            return canon
    return ""


# Known typos seen in the RASFF Window XLS export (subject field).
# Audit 2026-05-04: "Cerulide" appears in #2026.3863 alongside the correctly-
# spelled "Cereulide" in #2026.3862 — they are sister notifications, so the
# typo is in RASFF's source data, not ours. Add new entries here when we
# spot them in production.
_RASFF_TYPO_FIXES = (
    (r"\bcerulide\b", "cereulide"),       # missing 'e'
    (r"\bcereuli?de\b", "cereulide"),     # belt-and-braces; matches both
    (r"\baflotoxin\w*\b", "aflatoxin"),   # 'o' for 'a'
)


def _correct_rasff_typos(text: str) -> str:
    """Apply RASFF-side typo corrections (case-insensitive)."""
    for pat, repl in _RASFF_TYPO_FIXES:
        text = re.sub(pat, repl, text, flags=re.IGNORECASE)
    return text


def _build_company_field(origin: str, notifying: str, distribution: str) -> str:
    """
    Build the 'Origin: X | Notifying: Y | Distributed: Z' company string per
    RASFF schema. All three fields can be comma-separated multi-country
    strings. Empty/None becomes 'unknown'.
    """
    def fmt(val: str) -> str:
        v = (val or "").strip()
        return v if v else "unknown"

    return (
        f"Origin: {fmt(origin)} | "
        f"Notifying: {fmt(notifying)} | "
        f"Distributed: {fmt(distribution)}"
    )


def _classify(classification: str) -> str:
    """Map RASFF classification → existing Class field convention."""
    c = (classification or "").lower().strip()
    if "alert" in c:
        return "Alert"
    if "border rejection" in c:
        return "Border Rejection"
    if "information" in c:
        return "Information"
    return "Recall"


def _primary_country(origin: str) -> str:
    """
    Return the first country listed in origin field (canonical form).
    RASFF origin can be 'Ireland,Poland' (multi-country production chain) —
    we use the first as Country, everything else lives in the Company field.
    """
    if not origin:
        return ""
    first = origin.split(",")[0].strip()
    return normalize_country(first) or first


def _reason_text(subject: str, hazards: str, distribution: str,
                 classification: str, decision: str) -> str:
    """
    Build the human-readable Reason field by concatenating the strongest
    signals. Truncated to REASON_MAX_LEN to match existing row style.
    """
    parts = []
    if subject:
        parts.append(subject.strip())
    if hazards:
        # Strip the '{category}' annotations for prose — keep the substance names
        clean_haz = re.sub(r"\s*-\s*\{[^}]*\}", "", hazards).strip()
        if clean_haz and clean_haz.lower() not in (subject or "").lower():
            parts.append(f"hazards: {clean_haz}")
    if decision and decision.lower() != "potentially serious":
        parts.append(f"risk: {decision}")
    if distribution:
        parts.append(f"distributed: {distribution}")
    text = "; ".join(parts)
    return text[:REASON_MAX_LEN]


# ---------------------------------------------------------------------------
# Scraper class
# ---------------------------------------------------------------------------

class RASFFScraper(BaseScraper):
    """
    EU Rapid Alert System for Food and Feed — uses the public XLS export
    endpoint to bypass RASFF Window's SPA hydration requirement.

    Scrapes notifications matching:
      - type == 'food'
      - classification in {alert, border rejection, information for attention/follow-up}
      - date >= today - since_days
      - hazards or subject contain a pathogen in PATHOGEN_RULES
    """
    AGENCY = "RASFF (EU)"
    COUNTRY = ""    # RASFF rows use per-row origin, not a fixed agency country
    LANGUAGE = "en"

    def scrape(self, since_days: int = DEFAULT_SINCE_DAYS) -> List[Recall]:
        cutoff = datetime.now(timezone.utc) - timedelta(days=since_days)

        # Step 1 — fetch the XLS export
        log.info("RASFF XLS export: fetching %s (since_days=%d)", XLS_URL, since_days)
        resp = fetch(self.session, XLS_URL, timeout=60)
        if resp is None or resp.status_code != 200:
            status = resp.status_code if resp is not None else "no-response"
            log.warning("RASFF XLS unavailable (status=%s) — falling back to CSV", status)
            resp = fetch(self.session, CSV_URL, timeout=60)
            if resp is None or resp.status_code != 200:
                log.warning(
                    "RASFF: both XLS and CSV endpoints failed; returning 0 rows. "
                    "Country scrapers continue to provide EU coverage."
                )
                return []
            return self._parse_csv(resp.content, cutoff)

        return self._parse_xlsx(resp.content, cutoff)

    # -----------------------------------------------------------------------
    # Parsers
    # -----------------------------------------------------------------------

    def _parse_xlsx(self, blob: bytes, cutoff: datetime) -> List[Recall]:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
        except Exception as exc:
            log.error("RASFF XLS parse failed: %s", exc)
            return []

        # The export uses a single sheet (default name "RASFF_window_results"
        # but we don't depend on the name).
        ws = wb[wb.sheetnames[0]]
        hdr_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not hdr_row:
            log.error("RASFF XLS empty / no header row")
            return []

        headers = [str(h or "").strip() for h in hdr_row]

        def col(name: str) -> int:
            try:
                return headers.index(name)
            except ValueError:
                return -1

        idx = {
            "reference":  col("reference"),
            "category":   col("category"),
            "type":       col("type"),
            "subject":    col("subject"),
            "date":       col("date"),
            "notifying":  col("notifying_country"),
            "classification": col("classification"),
            "decision":   col("risk_decision"),
            "distribution": col("distribution"),
            "origin":     col("origin"),
            "hazards":    col("hazards"),
        }
        # Verify the required columns are present — if RASFF renames, fail loud.
        missing = [k for k, v in idx.items() if v < 0]
        if missing:
            log.error("RASFF XLS missing columns: %s — schema may have changed; aborting",
                      missing)
            return []

        return self._iter_rows(ws.iter_rows(min_row=2, values_only=True), idx, cutoff)

    def _parse_csv(self, blob: bytes, cutoff: datetime) -> List[Recall]:
        """CSV fallback path (same column shape, comma-separated)."""
        import csv
        try:
            text = blob.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = blob.decode("latin-1")

        reader = csv.reader(io.StringIO(text))
        try:
            headers = [h.strip() for h in next(reader)]
        except StopIteration:
            log.error("RASFF CSV empty")
            return []

        def col(name: str) -> int:
            try: return headers.index(name)
            except ValueError: return -1

        idx = {
            "reference": col("reference"), "category": col("category"),
            "type": col("type"), "subject": col("subject"),
            "date": col("date"), "notifying": col("notifying_country"),
            "classification": col("classification"), "decision": col("risk_decision"),
            "distribution": col("distribution"), "origin": col("origin"),
            "hazards": col("hazards"),
        }
        missing = [k for k, v in idx.items() if v < 0]
        if missing:
            log.error("RASFF CSV missing columns: %s — aborting", missing)
            return []

        return self._iter_rows(reader, idx, cutoff)

    # -----------------------------------------------------------------------
    # Row parser (shared between XLSX + CSV)
    # -----------------------------------------------------------------------

    def _iter_rows(self, rows_iter, idx, cutoff: datetime) -> List[Recall]:
        out: List[Recall] = []
        n_total = 0
        n_dropped_date = n_dropped_type = n_dropped_class = n_dropped_pathogen = 0
        n_dropped_dup = 0
        seen_refs = set()

        for r in rows_iter:
            if not r:
                continue
            n_total += 1

            def get(key):
                i = idx[key]
                if i < 0 or i >= len(r):
                    return ""
                v = r[i]
                return "" if v is None else str(v).strip()

            # Date filter
            d = _parse_rasff_date(get("date"))
            if d is None or d < cutoff:
                n_dropped_date += 1
                continue

            # Type filter (food only)
            if get("type").lower() not in MONITORED_TYPES:
                n_dropped_type += 1
                continue

            # Classification filter
            classif = get("classification").lower()
            if not any(mc in classif for mc in MONITORED_CLASSIFICATIONS):
                n_dropped_class += 1
                continue

            # Pathogen filter
            pathogen = _extract_pathogen(get("subject"), get("hazards"))
            if not pathogen:
                n_dropped_pathogen += 1
                continue

            # Dedup by reference (just in case the export has duplicates)
            ref = get("reference")
            if ref in seen_refs:
                n_dropped_dup += 1
                continue
            seen_refs.add(ref)

            # Build the Recall row
            origin = get("origin")
            country = _primary_country(origin) or _primary_country(get("notifying"))
            row_class = _classify(classif)
            url = (
                f"https://webgate.ec.europa.eu/rasff-window/screen/notification/{ref}"
                if ref else ""
            )
            company = _build_company_field(
                origin, get("notifying"), get("distribution")
            )
            reason = _reason_text(
                get("subject"), get("hazards"), get("distribution"),
                classif, get("decision"),
            )

            recall = self._new_recall(
                Date=d.strftime("%Y-%m-%d"),
                Company=company,
                Brand="—",
                Product=get("subject") or get("category"),
                Pathogen=pathogen,
                Reason=reason,
                Class=row_class,
                URL=url,
                Outbreak=0,  # RASFF doesn't expose illness counts in the public export
                Notes=f"[RASFF #{ref}; classification: {classif}; "
                      f"category: {get('category')}]",
            )

            # Override Country/Region — _new_recall uses self.COUNTRY which is
            # empty for RASFF; we set per-row from origin instead.
            if country:
                recall.Country = country
                recall.Region = infer_region(country) or recall.Region

            out.append(recall)

        log.info(
            "RASFF: %d rows in export → %d kept (dropped: %d date, %d type, "
            "%d classification, %d pathogen, %d duplicate)",
            n_total, len(out),
            n_dropped_date, n_dropped_type, n_dropped_class,
            n_dropped_pathogen, n_dropped_dup,
        )
        return out
