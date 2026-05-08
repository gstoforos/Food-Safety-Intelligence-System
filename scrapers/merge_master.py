"""
Master merge logic (Pending-sheet architecture).

recalls.xlsx holds THREE sheets:
  - Recalls  : approved, published data (consumed by the weekly report)
  - Pending  : freshly scraped rows awaiting validation + review
  - NEWS     : unrelated news-feed sheet, preserved as-is

Daily pipeline flow:
  1. Scrapers write to Pending (via append_to_pending)
  2. Enrichment + URL validation + AI review run against Pending
  3. promote_approved() moves rows that pass all checks into Recalls
  4. Rejected rows stay in Pending with a rejection reason stored in Notes
     (prefixed "REJECTED: <reason> | <original notes>") so a human can triage.

Dedup:
  - Primary key: URL (lowercased, stripped)
  - Fallback:    date + company + pathogen
  - Dedup applies within Pending and across Pending->Recalls promotion.
"""
from __future__ import annotations
import json
import logging
import re
import unicodedata
from datetime import datetime, timezone
from pathlib import Path
from typing import List, Dict, Any, Tuple, Optional
from urllib.parse import urlparse
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill

from scrapers._models import Recall

log = logging.getLogger(__name__)

SCHEMA = ["Date", "Source", "Company", "Brand", "Product", "Pathogen", "Reason",
          "Class", "Country", "Region", "Tier", "Outbreak", "URL", "Notes"]

# ── Internal-only tracking columns (Recalls sheet only) ──────────────────
# These columns are appended to the Recalls sheet for internal bookkeeping
# and are EXCLUDED from the public-facing recalls.json that feeds the
# dashboard. See mirror_json_from_xlsx() for the filtering.
#
# DateAdded   — date the row was first promoted to Recalls (set once,
#               never changed). Used to distinguish original publication
#               date (Date) from when FSIS captured it.
# LastUpdated — date the row was last modified (any field changed). Set
#               by promote_approved on insert and by audit/fix code paths
#               that touch existing rows.
# LastChecked — date a URL gate (Gemini grounded check or url_guardian
#               reachability check) last validated this row's URL. Used
#               by url_guardian to skip rows checked recently and avoid
#               redundant Gemini grounded calls.
RECALLS_INTERNAL_COLUMNS = ["DateAdded", "LastUpdated", "LastChecked"]
RECALLS_SCHEMA = SCHEMA + RECALLS_INTERNAL_COLUMNS

# Pending sheet has the same columns plus three tracking columns.
# Audit 2026-05-05 — added RejectedBy column to track which reviewers have
# rejected this row. Stored as a comma-separated set of reviewer names
# (e.g. "claude-check,gemini-url-gate"). Used by mark_rejected_with_counter
# to physically delete rows once 2+ DIFFERENT reviewers have rejected.
PENDING_SCHEMA = SCHEMA + ["ScrapedAt", "Status", "RejectedBy"]

# ── Rejected sheet (audit 2026-05-08, per operator decision) ─────────────
# Rows that BOTH morning-pair (Gemini URL gate + Claude check) AND/OR
# evening-pair (Gemini URL gate + OpenRouter check) reject get archived
# to the Rejected sheet for human audit. Previously these rows were just
# physically deleted by the 2-reviewer counter — that loses evidence and
# makes it impossible to spot reviewer bias or systemic false-rejections.
# RejectedAt = UTC timestamp at which the 2nd-reviewer rejection landed
# RejectedBy = comma-separated set of reviewer names (already in PENDING_SCHEMA)
REJECTED_SCHEMA = SCHEMA + ["ScrapedAt", "Status", "RejectedBy", "RejectedAt"]

NEWS_HEADERS = ["Published (UTC)", "Pathogen", "Event", "Source", "Title",
                "Link", "Retrieved (UTC)"]

# Status values used in the Pending sheet
STATUS_PENDING = "pending"
STATUS_APPROVED = "approved"   # transient — promoted rows are removed from Pending
STATUS_REJECTED = "rejected"

# ── Gap-finder gating state machine (audit 2026-04-29) ──────────────────
# Gap-finder rows (Tavily/Exa/Gemini/Claude/OpenAI search-based recall
# discovery) are LESS trustworthy than scraper rows because they come from
# search-engine indexes, not from authoritative regulator pages. Per
# operator policy, they must NOT be auto-promoted to Recalls. Instead,
# they sit in Pending under one of these gating states until they pass:
#   1. Two independent Gemini URL grounding checks (different runs,
#      different model invocations — non-determinism is the point), then
#   2. One Claude page-content verification.
# State transitions (see promote_gap_rows.py / url_gate_gemini.py /
# claude_check.py for the implementations):
#   pending_gap     -- written by gap-finders; first url_gate run will
#                      either advance to pending_gap_v1 or reject.
#   pending_gap_v1  -- one Gemini URL pass. Second url_gate run advances
#                      to pending_gap_v2 or rejects.
#   pending_gap_v2  -- two Gemini URL passes. claude_check advances to
#                      pending (eligible for normal merge) or rejects.
# promote_approved skips ANY pending_gap* row — they never reach Recalls
# until claude_check has flipped them back to plain "pending".
STATUS_PENDING_GAP    = "pending_gap"
STATUS_PENDING_GAP_V1 = "pending_gap_v1"
STATUS_PENDING_GAP_V2 = "pending_gap_v2"

GAP_GATING_STATUSES = frozenset({
    STATUS_PENDING_GAP, STATUS_PENDING_GAP_V1, STATUS_PENDING_GAP_V2,
})


# ---------------------------------------------------------------------------
# Dedup
# ---------------------------------------------------------------------------
def _normalize_url_for_dedup(url: str) -> str:
    """Normalize a URL for dedup comparison.

    Audit 2026-05-06: previously the dedup key was just url.strip().lower().
    That missed http:// vs https:// duplicates — production showed a Tavily-
    sourced http://www.fsis.usda.gov/... slipping past the gate even though
    the https:// version was already in Recalls.

    Normalizations applied (all lowercase):
      • strip protocol (http:// → '', https:// → '')
      • strip leading 'www.'
      • strip trailing '/'
      • strip URL fragment '#...'
      • strip non-canonical query strings (utm_*, ref, etc.) but PRESERVE
        identifier-style query params (permalink, id, fiche, ref, recall_id)
    """
    if not url:
        return ""
    s = url.strip().lower()
    s = s.split("#", 1)[0]
    if s.startswith("https://"):
        s = s[8:]
    elif s.startswith("http://"):
        s = s[7:]
    if s.startswith("www."):
        s = s[4:]
    if "?" in s:
        path, _, query = s.partition("?")
        keepers = []
        for kv in query.split("&"):
            k = kv.split("=", 1)[0]
            if k in ("permalink", "id", "fiche", "ref", "recall_id"):
                keepers.append(kv)
        s = path + (("?" + "&".join(keepers)) if keepers else "")
    if s.endswith("/"):
        s = s[:-1]
    return s


def _dedup_key(row: Dict[str, Any]) -> str:
    """URL primary (normalized), fallback to date+company+pathogen."""
    url = _normalize_url_for_dedup(row.get("URL") or "")
    if url:
        return url
    co = unicodedata.normalize("NFD", row.get("Company") or "").encode("ascii", "ignore").decode().lower()
    co = re.sub(r"[^a-z0-9]", "", co)[:30]
    return f"{row.get('Date','')}|{co}|{(row.get('Pathogen','') or '')[:30]}"


# ---------------------------------------------------------------------------
# Date-consistency check (audit 2026-05-06: defense-in-depth)
# ---------------------------------------------------------------------------
# Production failure 2026-05-06: two USDA FSIS recalls from 2018 ("Oct. 19,
# 2018" Envolve Foods Listeria + "March 29, 2018" Target Corp Listeria)
# were promoted from Pending to Recalls and surfaced on the public dashboard
# with Date=2026-05-06 (today's date stamped by the Tavily date-extractor
# fallback). Three independent gates failed:
#   1. Tavily date-extractor regex didn't accept "Oct." (period after abbrev)
#   2. URL gate Gemini returned prose; parse-failure path defaulted to PASS
#   3. Pending → Recalls promotion had NO date-sanity check at all
#
# This function is the third-line defense. It compares the Date field
# against any date pattern found in Notes. If Notes mentions a year that's
# more than 1 year older than Date's year, the promotion is rejected and
# the row stays in Pending for manual review.

# Date-extraction regex used at the promotion gate. Matches:
#   "Oct. 19, 2018"   "March 29, 2018"   "2018-10-19"   "19 March 2018"
_PROMOTION_OLD_DATE_RX = re.compile(
    r"\b(?:"
    r"(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec|"
    r"january|february|march|april|june|july|august|"
    r"september|october|november|december)\.?\s+\d{1,2}[,\s]+(\d{4})"
    r"|"
    r"(\d{4})-(?:0[1-9]|1[0-2])-(?:0[1-9]|[12]\d|3[01])"
    r"|"
    r"\d{1,2}\s+(?:jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec|"
    r"january|february|march|april|june|july|august|"
    r"september|october|november|december)\.?\s+(\d{4})"
    r")\b",
    re.IGNORECASE,
)


def _check_date_consistency(row: Dict[str, Any]) -> Optional[str]:
    """Reject promotion when Notes mentions a year far older than Date field.

    Returns None if the row is consistent (or no date found in Notes).
    Returns a rejection-reason string if inconsistent.

    Heuristic: scan Notes for date patterns; if the OLDEST year mentioned
    is more than 1 calendar year older than the Date field's year, the
    Date field is almost certainly a date-extractor fallback (today's
    date stamped on an old archived page).

    Conservative — only rejects when the gap is unambiguous (>1 year).
    Rows where Notes mentions e.g. "in 2025" while Date is 2026 will pass.
    """
    date_field = str(row.get("Date") or "")[:10]
    notes = str(row.get("Notes") or "")
    if not date_field or not notes:
        return None
    try:
        date_year = int(date_field[:4])
    except ValueError:
        return None

    # Find all 4-digit years mentioned in date contexts in Notes
    years_found = []
    for m in _PROMOTION_OLD_DATE_RX.finditer(notes):
        for g in m.groups():
            if g and g.isdigit() and len(g) == 4:
                yr = int(g)
                if 2000 <= yr <= date_year + 1:
                    years_found.append(yr)
                break

    if not years_found:
        return None

    oldest = min(years_found)
    if oldest <= date_year - 2:
        return (f"date_inconsistent: Date field is {date_field} but Notes "
                f"mentions {oldest} (likely date-extractor fallback on an "
                f"archived page)")
    return None


# ---------------------------------------------------------------------------
# Near-duplicate detection (catches same-recall-different-URL cases)
# ---------------------------------------------------------------------------
# Why this exists: regulators sometimes publish the same recall under two
# URL formats (FDA's canonical company-slug URL vs. their share-link
# "?permalink=<hash>" wrapper). The OpenAI/search-based gap finders find
# the wrapper URL while the direct scraper finds the canonical URL —
# string dedup can't catch this because the URLs are entirely different
# paths. Even the date+company+pathogen fallback fails when the gap-finder
# discovers the recall N days after the scraper did (different Date field).
#
# The near-dup index keys on (source, normalized_company, pathogen) and
# stores a list of dates. A new row is rejected if there's already an
# entry with the same key dated within NEAR_DUP_WINDOW_DAYS days. This
# blocks rediscovery duplicates without blocking legitimate same-company
# recurring recalls (which are usually months apart).
NEAR_DUP_WINDOW_DAYS = 30


def _near_dup_key(row: Dict[str, Any]) -> str:
    """Normalized (source, company, pathogen) tuple — date-independent."""
    src = (row.get("Source") or "").strip().lower()
    co = unicodedata.normalize("NFD", row.get("Company") or "").encode("ascii", "ignore").decode().lower()
    co = re.sub(r"[^a-z0-9]", "", co)[:30]
    pa = (row.get("Pathogen") or "").strip().lower()[:50]
    if not (src and co and pa):
        return ""  # missing any of the three — can't make a meaningful match
    return f"{src}|{co}|{pa}"


def _build_near_dup_index(rows: List[Dict[str, Any]]) -> Dict[str, List[str]]:
    """Return {near_dup_key: [date_str, ...]} for all rows with valid keys."""
    idx: Dict[str, List[str]] = {}
    for r in rows:
        k = _near_dup_key(r)
        d = str(r.get("Date") or "")[:10]
        if k and d:
            idx.setdefault(k, []).append(d)
    return idx


def _is_near_duplicate(
    row: Dict[str, Any], near_dup_index: Dict[str, List[str]],
) -> Tuple[bool, str]:
    """Check if `row` is a near-dup of anything in the index. Returns (is_dup, match_date)."""
    k = _near_dup_key(row)
    new_date_str = str(row.get("Date") or "")[:10]
    if not (k and new_date_str and k in near_dup_index):
        return False, ""
    try:
        new_date = datetime.strptime(new_date_str, "%Y-%m-%d").date()
    except ValueError:
        return False, ""
    for old_date_str in near_dup_index[k]:
        try:
            old_date = datetime.strptime(old_date_str, "%Y-%m-%d").date()
            if abs((new_date - old_date).days) <= NEAR_DUP_WINDOW_DAYS:
                return True, old_date_str
        except ValueError:
            continue
    return False, ""


# ---------------------------------------------------------------------------
# Pending-row validation gate
# ---------------------------------------------------------------------------
# This is the SINGLE chokepoint that blocks garbage from EVERY source
# (scrapers, gap finders, manual injects). Every row added to Pending must
# pass validate_pending_row() — see PIPELINE_FIX_SPEC.md.

# Generic / non-detail URL patterns we never want in Pending. These are
# regulator landing/listing/transparency pages, not specific recall fiches.
_GENERIC_URL_PATTERNS = (
    r"vertexaisearch\.cloud\.google",        # Gemini grounding redirect
    r"fsai\.ie/news-alerts/food\?page=",      # FSAI paginated listing
    r"rasff-window/screen/list\?",            # RASFF list page
    r"quebec\.ca/.*/listeriosis",             # Quebec disease info
    r"quebec\.ca/.*/animal-disease",          # Quebec animal disease info
    r"quebec\.ca/.*/food-recalls$",           # Quebec generic recalls page
    r"regulatory-transparency-and-openness",  # CFIA transparency pages
    r"food-safety-investigations/$",          # CFIA investigation index
    r"/categorie/[\d/]+/?$",                  # RappelConso category index
    r"/rubrik/[^/]+/?$",                      # produktwarnung.eu rubrik
    r"/news-and-alerts/food-alerts/?$",       # FSAI alerts root
    r"/safety/recalls-market-withdrawals-safety-alerts/?$",  # FDA root
    r"/animal-veterinary/news-events/outbreaks-and-advisories/?$",  # FDA pet root
    # CFIA recalls landing page (any locale path or bare host). The CFIA
    # scraper finds specific recall slugs at recalls-rappels.canada.ca/<lang>/<slug>;
    # the bare /fr or /en URL is the listing page itself, never a recall.
    # Triggered by the audit 2026-04-28 leak where the French landing page
    # entered Recalls with the page H1 ("Trouvez des rappels...") as Company.
    r"recalls-rappels\.canada\.ca/(?:fr|en)/?$",
    r"recalls-rappels\.canada\.ca/?$",
    # FDA share-link wrapper format used by their "voluntary-recall" template.
    # Functionally a SPA route — same recall is also published at the
    # canonical /safety/recalls-market-withdrawals-safety-alerts/<slug> URL,
    # which the FDA scraper always finds. Reject the wrapper to prevent
    # duplicates from search-based gap finders (Tavily/Exa/OpenAI) that
    # return whichever URL Google indexed first.
    r"/safety/recalls-market-withdrawals-safety-alerts/voluntary-recall\?permalink=",
    # ── 2026-05-05 audit additions (gap-finder garbage patterns) ───────────
    # Generic full-text search query pages — these aren't recalls, they're
    # search-result lists. CFIA, RappelConso, EFSA, FSANZ all have these.
    r"/search/site",                                # CFIA "advanced search"
    r"/search\?",                                   # generic search query
    r"/recherche\?",                                # French generic search
    r"/recherche/",                                 # French search path
    r"/buscador",                                   # Spanish search
    r"/suche\?",                                    # German search
    # Pagination pages — never recall-specific
    r"/page/\d+/?(?:$|\?)",                         # /page/50/, /page/2/?...
    r"\bpage=\d+",                                  # ?page=50
    # Notification circulars / regulatory bulletins — these are notices ABOUT
    # things to come, not recalls themselves. FSANZ, ANSES, USDA-FSIS all have
    # circular indexes that should never appear in our recall feed.
    r"/notification-circulars?/?$",
    r"/notification-circulars?/index",
    r"/circulars?/notification-circular-",          # FSANZ specific circular ID
    r"/bulletins?/?$",
    r"/news-circulars?/?$",
)

# Company-field strings that the scraper has clearly bungled (they're page
# titles, section headers, or page text — never legitimate company names).
#
# IMPORTANT — what does NOT belong in this set:
#   • "Various brands" / "Various producers" / "Multiple brands"  → legit
#     descriptor when one recall covers many SKUs from different producers
#     (e.g. BLV Salmonellen-Weichkäse, RASFF multi-country alerts).
#   • "Unbranded" / "—" / "sans marque" / "No brand"              → legit
#     descriptor for RappelConso "sans marque" entries, generic raw products,
#     bulk commodity recalls.
#   • "Consult Food …", "Various Foods Ltd", etc.                 → real
#     company names that happen to start with normally-suspect words.
# Company-field cleanup beyond clear scraper bugs is the URL gate's job +
# downstream Claude review's job — not this gate's job.
_GARBAGE_COMPANIES = {
    "list of",                                       # FSAI/CFIA listing-page H1
    "food alerts",                                   # FSAI navigation
    "food alert",                                    # FSAI navigation
    "listeriosis",                                   # disease name as company
    "animals can catch and transmit salmonellosis",  # CFIA page text
    "food safety investigation:",                    # CFIA section header
    "timeline of events:",                           # CFIA section header
    "recall of",                                     # FSAI page-title leak (prefix)
}

# Hard cutoff: nothing dated before this enters Pending.
_MIN_VALID_DATE = "2026-01-01"


# News-outlet hosts. Any URL whose host (or parent domain) matches lands in
# the NEWS sheet via scrapers/news.py — never in Recalls. Gap-finders
# (Tavily/Exa/Gemini) sometimes surface news-article URLs while searching for
# recall content; without this blocklist they would slip into Pending and
# get promoted to Recalls with the article <title> tag scraped as Company.
# Triggered by the audit 2026-04-28 leak (foodsafetynews.com articles
# appearing in Recalls).
# ── RASFF (EU) URL pattern (audit 2026-04-29) ──────────────────────────────
# RASFF rows are accepted only when the URL is a specific notification
# detail page. The Window app at /screen/search and /screen/consumers is
# a Vue/Angular SPA shell with no recall content. The notification deep-
# link route /screen/notification/<id> IS rendered server-side enough to
# carry the recall record. validate_pending_row enforces this constraint
# only when Source starts with "RASFF". See _missing_required in
# run_all.py for the equivalent check during scraper run.
_RASFF_NOTIFICATION_URL_RE = re.compile(
    r"^https://webgate\.ec\.europa\.eu/rasff-window/screen/notification/\d+/?$",
    re.IGNORECASE,
)

_NEWS_HOSTS = frozenset({
    "foodsafetynews.com",
    "foodpoisonjournal.com",
    "foodpoisoningbulletin.com",
    "outbreaknewstoday.com",
    "cidrap.umn.edu",
    "food-safety.com",
    "barfblog.com",
    "foodbusinessnews.net",
    "foodnavigator.com",
    "foodnavigator-usa.com",
    "just-food.com",
    "foodmanufacture.co.uk",
    "foodprocessing.com",
    "foodengineeringmag.com",
    "fooddive.com",
    "reuters.com",
    "apnews.com",
    "bbc.com",
    "bbc.co.uk",
    "bloomberg.com",
    "theguardian.com",
    "nytimes.com",
    "washingtonpost.com",
    "medicalxpress.com",
    "sciencedaily.com",
    "yahoo.com",
    "msn.com",
    "news.google.com",
})


def _host_is_news_outlet(url: str) -> bool:
    """True if the URL host (or any parent of it) is in _NEWS_HOSTS."""
    if not url:
        return False
    try:
        host = urlparse(url).netloc.lower()
    except Exception:
        return False
    if host.startswith("www."):
        host = host[4:]
    if host in _NEWS_HOSTS:
        return True
    # Subdomain match (e.g. recalls.reuters.com endswith .reuters.com)
    for h in _NEWS_HOSTS:
        if host.endswith("." + h):
            return True
    return False


def validate_pending_row(
    row: Dict[str, Any],
    existing_urls: set,
) -> Tuple[bool, str]:
    """
    Return (is_valid, rejection_reason). Reject garbage before it enters
    Pending. Called by append_to_pending() for every candidate row, no
    matter the source (scrapers, gap finders, manual injects).

    Rules — see PIPELINE_FIX_SPEC.md for rationale:
      • REJECT vertexaisearch redirect URLs (Gemini grounding artifacts)
      • REJECT generic / category / paginated-listing pages
      • REJECT URLs that aren't http/https
      • REJECT garbage Company fields ("List of", "Food Alerts", etc.)
      • REJECT duplicate URLs already in Recalls or Pending
      • REJECT dates before 2026-01-01

    `existing_urls` is a set of already-seen lowercased URLs (Recalls +
    current Pending). Pass an empty set to skip the dedup check.

    NOTE on near-duplicate detection: helpers _near_dup_key /
    _build_near_dup_index / _is_near_duplicate exist above for use by
    standalone audit scripts. They are NOT wired into this gate because
    European regulators (especially RappelConso) routinely publish one
    fiche per SKU — same company, same pathogen, same day, different
    fiche IDs. A naive same-source/company/pathogen+30d gate would
    falsely reject those legitimate per-SKU rows. Use the helpers only
    when you've verified the URL host context makes near-dup detection
    safe (e.g. for FDA where wrapper URLs duplicate canonical URLs).
    """
    url = str(row.get("URL", "") or "").strip()
    company = str(row.get("Company", "") or "").strip()
    date_str = str(row.get("Date", "") or "")[:10]
    source = str(row.get("Source", "") or "").strip()

    # ── RASFF (EU) schema awareness (audit 2026-04-29) ──────────────────
    # RASFF rows don't publish company names — they publish origin and
    # distributed countries instead. The Company field on a valid RASFF
    # row holds the formatted string "Origin: <X> | Distributed: <Y>".
    # That string contains a pipe, which would trip the news-article-
    # title leak check below. Detect RASFF up front so we can skip
    # company-garbage heuristics for these rows AND impose the inverse
    # constraint: RASFF URL must point to a /screen/notification/<id>
    # page, not a search/landing/consumer shell.
    is_rasff = source.upper().startswith("RASFF")

    # ── Final gate (locked 2026-04-30) — modules imported at function ──
    # Local imports to avoid circular deps (merge_master is imported by
    # both url_gate_gemini and claude_check, which import these too).
    try:
        from pipeline._url_year import is_year_mismatch
        from pipeline._pathogen_scope import is_in_scope as _is_tier1_pathogen
        from pipeline._news_mirror_blocklist import is_news_mirror as _is_news_mirror
    except ImportError:
        is_year_mismatch = None
        _is_tier1_pathogen = None
        _is_news_mirror = None

    if _is_news_mirror is not None and _is_news_mirror(url):
        return False, "news_mirror_domain (locked 2026-04-30)"

    pathogen_str = str(row.get("Pathogen", "") or "")
    if _is_tier1_pathogen is not None and not _is_tier1_pathogen(pathogen_str):
        # Exception: enrichment-placeholder rows have empty Pathogen by
        # design. Two flavours:
        #
        #   1. HTML-fallback rows (FSAI/CFIA/FDA-press) — the scraper
        #      hit a listing page when the structured feed was dead and
        #      emitted only URL+title slug. Detail-page extraction is
        #      deferred to claude-check.
        #   2. Structured-feed partial-extract rows (RappelConso bulk
        #      JSON) — the scraper got a record but the API didn't
        #      carry Company / Brand / Pathogen for that fiche. Same
        #      deferral: claude-check fetches the fiche page and
        #      extracts the missing field.
        #
        # In both cases the scraper stamps Notes with a sentinel string
        # asking for AI enrichment. If we hard-reject these at the
        # pathogen scope gate, they never reach claude-check at all —
        # bug observed in 2026-05-08 run (10 FSAI rows + 9 RappelConso
        # rows silently dropped).
        #
        # MECHANISM (not "promote-time re-validation" — that's wrong;
        # validate_pending_row is only called by append_to_pending,
        # never by promote_approved): claude-check fetches the detail
        # page on its scheduled cycle, populates Pathogen (and Company/
        # Brand/Product as applicable), and writes a pass/fail/fix
        # verdict. promote_approved consumes that verdict via
        # rejected_flags — non-Tier-1 hazards get rejected at promote
        # time through claude-check's verdict, NOT by re-running this
        # gate. If claude-check never runs, the row sits in Pending
        # indefinitely with empty Pathogen — visible to the operator
        # rather than auto-promoted.
        #
        # The base token list is imported from enrichment/enrich_rows.py
        # so HTML-fallback semantics stay in lockstep with the AI-skip
        # rule there (single source of truth — no drift). The local
        # _PENDING_BYPASS_TOKENS extension adds tokens emitted only by
        # structured-feed scrapers like RappelConso, which enrich_rows
        # doesn't need to know about.
        try:
            from enrichment.enrich_rows import (
                _HTML_FALLBACK_NOTES_TOKENS as _UPSTREAM_TOKENS,
            )
        except ImportError:
            # Defensive fallback. Keep in sync with enrichment/enrich_rows.py
            # if you update either side.
            _UPSTREAM_TOKENS = (
                "claude-check needs to enrich",
                "html listing fallback",
                "html fallback",
            )
        _PENDING_BYPASS_TOKENS = _UPSTREAM_TOKENS + (
            # RappelConso bulk-JSON partial-extract sentinel — emitted
            # by scrapers/europe_eu/rappelconso.py when the structured
            # feed lacks Company/Brand. Same enrichment intent as
            # HTML-fallback rows. Adding here (not to enrich_rows)
            # because RappelConso rows DO have a usable detail page,
            # so Gemini enrichment is fine — only the scope-gate bypass
            # at this layer is needed.
            "claude-check please extract",
        )
        notes_lc = str(row.get("Notes", "") or "").lower()
        is_enrichment_placeholder = (
            not pathogen_str.strip()
            and any(tok in notes_lc for tok in _PENDING_BYPASS_TOKENS)
        )
        if not is_enrichment_placeholder:
            return False, f"pathogen_out_of_scope: {pathogen_str!r}"
        # else: fall through; claude-check verdict is the eventual gate.

    if is_year_mismatch is not None:
        try:
            row_d = (datetime.fromisoformat(date_str).date()
                     if date_str else None)
        except (TypeError, ValueError):
            row_d = None
        ym_reason = is_year_mismatch(row_d, url)
        if ym_reason:
            return False, f"url_year_mismatch: {ym_reason}"

    # Extraction garbage
    company_lc = company.lower()
    brand_lc = str(row.get("Brand", "") or "").strip().lower()
    GARBAGE = {"home","index","page","recalls","alerts","alert","recall","welcome","main"}
    if company_lc and company_lc == brand_lc and company_lc in GARBAGE:
        return False, f"extraction_garbage: Company=Brand={company_lc!r}"
    if re.search(r"/(home|index|main|welcome)/?$", url.lower()):
        return False, "extraction_garbage: URL is landing page"

    # ── REJECT: vertexaisearch redirect URLs (Gemini grounding artifacts) ──
    if "vertexaisearch.cloud.google" in url:
        return False, "Gemini grounding redirect URL, not a real recall"

    # ── REJECT: URL host is a news outlet, not a regulator ──────────────
    # News articles belong in the NEWS sheet (populated by scrapers/news.py).
    # Gap-finders sometimes surface news-article URLs while searching for
    # recall content; reject them at the gate before they reach Recalls.
    if _host_is_news_outlet(url):
        return False, f"News outlet URL — belongs in NEWS sheet, not Recalls: {url[:60]}"

    # ── RASFF URL gate (audit 2026-04-29) ───────────────────────────────
    # RASFF rows must point to a specific notification page. Anything
    # else (the search SPA shell at /screen/search, the consumer portal
    # at /screen/consumers, or a bare /rasff-window root) is rejected
    # here. The notification page is the only URL that contains the
    # actual recall record. The gap-finder URL filter rejects the same
    # landing pages, but this is the structural gate at promotion time.
    if is_rasff:
        if not _RASFF_NOTIFICATION_URL_RE.match(url):
            return False, (f"RASFF row URL must be /screen/notification/<id>, "
                           f"got: {url[:80]}")

    # ── REJECT: generic / informational / listing pages ─────────────────
    for pat in _GENERIC_URL_PATTERNS:
        if re.search(pat, url, re.IGNORECASE):
            return False, f"Generic/info page URL: matches {pat}"

    # ── REJECT: URL is not http/https ───────────────────────────────────
    if url and not url.lower().startswith(("http://", "https://")):
        return False, f"Invalid URL scheme: {url[:30]}"

    # ── REJECT: company field is clearly a page title, not a company ────
    # Skipped for RASFF rows — their Company field legitimately contains
    # the "Origin: <X> | Distributed: <Y>" pattern.
    if not is_rasff:
        co_low = company.lower()
        if co_low in _GARBAGE_COMPANIES:
            return False, f'Company field is not a company: "{company}"'
        # Substring check for "Recall of …" / "List of …" leakage
        for bad in _GARBAGE_COMPANIES:
            if co_low.startswith(bad + " ") or co_low.startswith(bad + ":"):
                return False, f'Company field starts with garbage prefix "{bad}"'
        # Article-title leak: scraped <title> tags from news pages contain a
        # pipe + outlet name (e.g. "Salmonella outbreak ... | Food Safety News")
        # or HTML/JS fragments. Real company names never contain these.
        if " | " in company and re.search(
            r"\|\s*(food\s*safety\s*news|food\s*poison|outbreak\s*news|cidrap|"
            r"reuters|bbc|bloomberg|guardian)\b", company, re.I):
            return False, f'Company field is a news article <title> tag: "{company[:60]}"'
        if re.search(r"window\.\w+|document\.querySelector|<\s*script\b|"
                     r"\{socials\b|addEventListener\(", company, re.I):
            return False, f'Company field contains HTML/JS fragment: "{company[:60]}"'

    # ── REJECT: duplicate URL already in Recalls or Pending ─────────────
    url_norm = url.rstrip("/").lower()
    if url_norm and url_norm in existing_urls:
        return False, "Duplicate URL already exists"

    # ── REJECT: date is before 2026-01-01 ───────────────────────────────
    if date_str and date_str < _MIN_VALID_DATE:
        return False, f"Date before {_MIN_VALID_DATE}: {date_str}"

    return True, "OK"


# ---------------------------------------------------------------------------
# Load
# ---------------------------------------------------------------------------
def _load_sheet(xlsx_path: Path, sheet: str, schema: List[str]) -> List[Dict[str, Any]]:
    if not xlsx_path.exists():
        return []
    wb = load_workbook(xlsx_path)
    if sheet not in wb.sheetnames:
        return []
    ws = wb[sheet]
    headers = [c.value for c in ws[1]]
    out = []
    # Defensive: openpyxl may return Date cells as datetime objects when a
    # cell was manually edited and Excel auto-typed it. Every downstream
    # consumer (gate, sort, dedup, JSON mirror) expects YYYY-MM-DD strings,
    # so coerce here at the single source of truth.
    from datetime import datetime as _dt, date as _dt_date
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v in (None, "") for v in row):
            continue
        rec = {h: (v if v is not None else "") for h, v in zip(headers, row)}
        # Normalise Date column
        d = rec.get("Date")
        if isinstance(d, (_dt, _dt_date)):
            try:
                rec["Date"] = d.strftime("%Y-%m-%d")
            except (TypeError, ValueError):
                rec["Date"] = ""
        elif d not in (None, "") and not isinstance(d, str):
            # Excel serial number or other unexpected type
            rec["Date"] = str(d)[:10]
        # Backfill any schema cols missing from the sheet (schema evolution safety)
        for col in schema:
            rec.setdefault(col, "" if col not in ("Tier", "Outbreak") else 0)
        out.append(rec)
    return out


def load_existing(xlsx_path: Path) -> List[Dict[str, Any]]:
    """Read approved Recalls sheet -> list of dicts."""
    out = _load_sheet(xlsx_path, "Recalls", RECALLS_SCHEMA)
    log.info("Loaded %d approved rows from Recalls", len(out))
    return out


def load_pending(xlsx_path: Path) -> List[Dict[str, Any]]:
    """Read Pending sheet -> list of dicts. Empty if sheet doesn't exist yet."""
    out = _load_sheet(xlsx_path, "Pending", PENDING_SCHEMA)
    log.info("Loaded %d rows from Pending", len(out))
    return out


# ---------------------------------------------------------------------------
# Merge scraped rows into Pending
# ---------------------------------------------------------------------------
def append_to_pending(
    existing_pending: List[Dict[str, Any]],
    approved: List[Dict[str, Any]],
    new_recalls: List[Recall],
    scraped_at: str,
) -> List[Dict[str, Any]]:
    """
    Take new scraped+enriched Recall objects and append them to the pending list.

    Dedup rules:
      - If the key is already approved in Recalls  -> skip silently
      - If the key is currently in Pending with Status='pending'    -> skip (waiting)
      - If the key is currently in Pending with Status='rejected'   -> DELETE the
        old rejected row and insert the freshly scraped row for re-validation.
        This gives the source a chance to fix broken links / fill missing fields
        before the next run, and prevents rejected rows from being silently
        re-skipped forever.
      - Otherwise (brand new key) -> insert as Status='pending'.
    """
    keys_in_approved = {_dedup_key(r) for r in approved}

    # ── Near-duplicate index (audit 2026-04-29) ──────────────────────────
    # Catches hallucinated-URL gap-finder duplicates that string dedup
    # can't see. Example: scraper has Listeria/LES ATELIERS DE SEBASTIEN/
    # 2026-04-28 at /fiche-rappel/22142/Interne (real). Gemini gap-finder
    # then hallucinates the same recall at /fiche-rappel/22185/Interne
    # (fake). _dedup_key() compares URLs → no match → both land in
    # Pending → URL gate runs once a day at 07:00 → in the meantime
    # merge_master promotes the hallucinated one to Recalls.
    #
    # _is_near_duplicate keys on (source, normalized_company, pathogen)
    # within a 30-day window — so the same (source, company, pathogen)
    # appearing at a different URL within 30 days is rejected as a near-
    # dup before it ever lands in Pending. This catches the leak at
    # ingest time, not promotion time.
    near_dup_index = _build_near_dup_index(approved + existing_pending)

    # Build set of all URLs already present (Recalls + current Pending) for
    # the validation gate's dedup check. Lowercased + trailing-slash-stripped
    # to match validate_pending_row()'s normalisation.
    #
    # IMPORTANT (audit 2026-04-29): exclude URLs of rows currently in
    # Status="rejected" from existing_urls. The retry path further down
    # promises that a freshly-scraped row matching a rejected row will
    # DELETE the old row and re-queue the new one. That promise was dead
    # because validate_pending_row's dup-URL check fired first (the URL
    # was in existing_urls regardless of Status), bouncing the retry
    # before the retry logic ever ran.
    existing_urls: set = set()
    for r in approved:
        u = str(r.get("URL", "") or "").strip().rstrip("/").lower()
        if u:
            existing_urls.add(u)
    for r in existing_pending:
        if (r.get("Status") or "").lower() == STATUS_REJECTED:
            continue  # let the retry path handle this URL
        u = str(r.get("URL", "") or "").strip().rstrip("/").lower()
        if u:
            existing_urls.add(u)

    # Index existing pending by key so we can drop rejected duplicates in place.
    # Multiple rows with the same key shouldn't happen, but if they do keep them
    # all (one will match; the others are untouched).
    pending_by_key: Dict[str, List[int]] = {}
    for i, r in enumerate(existing_pending):
        pending_by_key.setdefault(_dedup_key(r), []).append(i)

    # Decide which existing-pending rows to drop (rejected rows being re-scraped).
    indices_to_drop: set = set()
    fresh_rows: List[Dict[str, Any]] = []
    retried = 0
    appended = 0
    already_pending = 0
    already_approved = 0
    rejected_by_gate = 0

    for r in new_recalls:
        d = r.to_dict() if isinstance(r, Recall) else dict(r)
        for col in SCHEMA:
            d.setdefault(col, "" if col not in ("Tier", "Outbreak") else 0)

        # ── HARD GATE: validate before any other logic. Blocks garbage from
        # ── ALL sources (scrapers, gap-finders, manual injects).
        ok, why = validate_pending_row(d, existing_urls)
        if not ok:
            log.warning(
                "Pending gate REJECT: %s | url=%s | company=%s",
                why, str(d.get("URL", ""))[:100], str(d.get("Company", ""))[:50],
            )
            rejected_by_gate += 1
            continue

        k = _dedup_key(d)

        if k in keys_in_approved:
            already_approved += 1
            continue

        # ── Near-duplicate check (audit 2026-04-29) ─────────────────────
        # Reject if a recall with the same (source, normalized_company,
        # pathogen) was already approved or pended within the last 30
        # days. This is the gate that stops gap-finder URL hallucinations
        # from ever entering Pending — even when their hallucinated URL
        # is structurally valid (5-digit fiche ID in the right range).
        is_near, match_date = _is_near_duplicate(d, near_dup_index)
        if is_near:
            log.warning(
                "Pending near-dup REJECT: same (source, company, pathogen) "
                "already exists dated %s | new url=%s | company=%s",
                match_date, str(d.get("URL", ""))[:100],
                str(d.get("Company", ""))[:50],
            )
            rejected_by_gate += 1
            continue

        if k in pending_by_key:
            # Look at the FIRST matching row's status (practically there's only one)
            existing_idx = pending_by_key[k][0]
            existing_status = (existing_pending[existing_idx].get("Status") or "").lower()
            if existing_status == STATUS_REJECTED:
                # Drop the old rejected row, re-queue the fresh scrape
                indices_to_drop.add(existing_idx)
                d["ScrapedAt"] = scraped_at
                d["Status"] = STATUS_PENDING
                fresh_rows.append(d)
                retried += 1
            else:
                # Still pending from a prior run — leave it alone
                already_pending += 1
            continue

        # Brand new key
        d["ScrapedAt"] = scraped_at
        d["Status"] = STATUS_PENDING
        fresh_rows.append(d)
        appended += 1

    # Assemble output: existing pending minus dropped + new/retried
    kept = [r for i, r in enumerate(existing_pending) if i not in indices_to_drop]
    out = kept + fresh_rows

    log.info(
        "Pending: kept %d (dropped %d rejected for retry), +%d new, +%d retried "
        "(skipped: %d already-pending, %d already-approved, %d gate-rejected) = %d total",
        len(kept), len(indices_to_drop), appended, retried,
        already_pending, already_approved, rejected_by_gate, len(out),
    )
    return out


# ---------------------------------------------------------------------------
# Promotion: Pending -> Recalls
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# Rejection counter (audit 2026-05-05)
# ---------------------------------------------------------------------------
# Two-strikes-and-out: a Pending row is physically deleted once 2 DIFFERENT
# reviewers have rejected it. Single-reviewer rejections sit in Pending with
# Status=rejected, available for retry on the next scrape (the existing
# append_to_pending re-queue path).
#
# Reviewer name is parsed from the rejection reason via a `<reviewer>: <text>`
# convention. claude_check stamps "Claude check: <reason>", url_gate_gemini
# stamps "Gemini gate: <reason>", openrouter_check stamps "OpenRouter:
# <reason>", and merge_master itself uses "Pending gate: <reason>" for
# structural rejects (broken URLs, missing fields).
#
# Returns (should_delete, updated_row). Caller decides whether to drop the
# row from kept_in_pending based on should_delete.

def _reviewer_from_reason(reason: str) -> str:
    """Extract canonical reviewer name from rejection reason string.

    Maps the leading 'reviewer: ...' prefix to one of the canonical names:
      claude-check, gemini-url-gate, openrouter-check, pending-gate, manual.
    Returns 'unknown' if no recognised prefix is found — these are still
    counted (so the counter advances) but logged for investigation.
    """
    r = (reason or "").strip().lower()
    if r.startswith("claude check") or r.startswith("claude-check"):
        return "claude-check"
    if r.startswith("gemini gate") or r.startswith("gemini-url-gate") or \
            r.startswith("gemini url gate"):
        return "gemini-url-gate"
    if r.startswith("openrouter") or r.startswith("openrouter-check"):
        return "openrouter-check"
    if r.startswith("pending gate") or r.startswith("pending-gate"):
        return "pending-gate"
    if r.startswith("url validator") or r.startswith("url-validator"):
        return "url-validator"
    if r.startswith("manual"):
        return "manual"
    return "unknown"


def mark_rejected_with_counter(row: Dict[str, Any], reason: str
                               ) -> Tuple[bool, Dict[str, Any]]:
    """Apply a rejection to a Pending row, tracking which reviewers reject it.

    Args:
        row    : the Pending row dict (mutated in place)
        reason : human-readable rejection reason, ideally prefixed with
                 reviewer name (e.g. "Claude check: company mismatch")

    Returns:
        (should_delete, updated_row)

        should_delete is True when 2+ DIFFERENT reviewers have rejected this
        row — caller is expected to drop the row from the Pending list.

        Repeat rejections by the SAME reviewer are idempotent: Notes is
        updated with the latest reason but the counter doesn't double-count.
    """
    reviewer = _reviewer_from_reason(reason)

    # Parse existing reviewer set from the RejectedBy column
    raw = (row.get("RejectedBy") or "").strip()
    rejected_by = set(filter(None, (s.strip() for s in raw.split(","))))

    # Same-reviewer repeat is idempotent (still update Notes for visibility)
    is_new_reviewer = reviewer not in rejected_by
    rejected_by.add(reviewer)

    row["RejectedBy"] = ",".join(sorted(rejected_by))
    row["Status"] = STATUS_REJECTED

    # Stamp reason into Notes (preserve any prior REJECTED: prefix history
    # by appending instead of overwriting if a previous reviewer already
    # tagged it).
    orig_notes = (row.get("Notes") or "").strip()
    new_stamp = f"REJECTED: {reason}"
    if orig_notes.startswith("REJECTED:"):
        # Already has a rejection stamp — append the new reviewer's reason
        # so the audit trail is preserved.
        if reason not in orig_notes:
            row["Notes"] = f"{orig_notes} || {reviewer}: {reason}"
    else:
        row["Notes"] = new_stamp + (f" | {orig_notes}" if orig_notes else "")

    # Two-strikes-and-out: only DIFFERENT reviewers count toward delete.
    should_delete = len(rejected_by) >= 2

    if should_delete:
        log.info(
            "Pending DELETE (2+ reviewers rejected): RejectedBy=%s | url=%s",
            row["RejectedBy"], str(row.get("URL", ""))[:100],
        )

    return should_delete, row


def cleanup_orphan_rejected(pending: List[Dict[str, Any]],
                            min_age_hours: int = 24
                            ) -> Tuple[List[Dict[str, Any]], int]:
    """Physically delete already-rejected Pending rows older than min_age_hours.

    Catches orphan rejections from before the RejectedBy counter was added —
    rows that have Status=rejected and a REJECTED: prefix in Notes but no
    RejectedBy column. After 24h these are unlikely to be re-validated, so
    we delete them to keep the Pending sheet clean.

    Returns (filtered_pending, n_deleted).
    """
    from datetime import datetime, timezone, timedelta
    cutoff = datetime.now(timezone.utc) - timedelta(hours=min_age_hours)

    filtered: List[Dict[str, Any]] = []
    n_deleted = 0
    for r in pending:
        if (r.get("Status") or "").lower() != STATUS_REJECTED:
            filtered.append(r)
            continue

        # If RejectedBy already tracks 2+ reviewers, mark_rejected_with_counter
        # would have deleted this row already — keep it as-is for the caller.
        rejected_by = set(filter(None, (
            s.strip() for s in (r.get("RejectedBy") or "").split(","))))
        if len(rejected_by) >= 2:
            n_deleted += 1
            continue  # delete (already past the counter threshold)

        # Otherwise, age-based delete: anything rejected >24h ago goes away.
        scraped_at = (r.get("ScrapedAt") or "").strip()
        try:
            sa = datetime.fromisoformat(scraped_at.replace("Z", "+00:00"))
            if sa.tzinfo is None:
                sa = sa.replace(tzinfo=timezone.utc)
        except (ValueError, TypeError):
            sa = None

        if sa is None or sa < cutoff:
            n_deleted += 1
            log.info(
                "Pending DELETE (orphan rejected, age>=%dh): url=%s",
                min_age_hours, str(r.get("URL", ""))[:100],
            )
            continue

        filtered.append(r)

    return filtered, n_deleted


# ---------------------------------------------------------------------------
def promote_approved(
    pending: List[Dict[str, Any]],
    approved_existing: List[Dict[str, Any]],
    rejected_flags: Dict[int, str],
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]]]:
    """
    Split the Pending list into
        (new_approved_rows_for_Recalls,
         rows_to_keep_in_Pending,
         rows_to_archive_to_Rejected).

    - `rejected_flags` maps pending-row-index -> rejection reason string.
      First-time rejections: stamp Status='rejected', keep in Pending
      with reason appended to Notes.
      Second-DIFFERENT-reviewer rejection: row is REMOVED from Pending and
      added to the archive list (caller writes it to the Rejected sheet).
    - Rows NOT in rejected_flags (and whose current Status is 'pending') are
      treated as approved and moved to Recalls, deduped against approved_existing.
    - Rows already marked 'rejected' in a prior run stay in Pending untouched.

    Audit 2026-05-08: third return value (archive list) added per operator
    decision — previously second-rejector deletes were silent and untraceable.
    Now they accumulate in a Rejected sheet for human audit.
    """
    approved_keys = {_dedup_key(r) for r in approved_existing}

    new_approved: List[Dict[str, Any]] = []
    kept_in_pending: List[Dict[str, Any]] = []
    archived_rejected: List[Dict[str, Any]] = []

    archive_ts = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    def _archive(clean_row: Dict[str, Any]) -> None:
        """Stamp RejectedAt and append to archive list."""
        clean_row["RejectedAt"] = archive_ts
        archived_rejected.append(clean_row)

    for idx, row in enumerate(pending):
        # Strip runtime-only fields (e.g. _url_check) before persisting
        clean = {k: v for k, v in row.items() if not k.startswith("_")}

        # Previously-rejected rows: leave alone, don't re-promote, don't re-reject
        if clean.get("Status") == STATUS_REJECTED and idx not in rejected_flags:
            kept_in_pending.append(clean)
            continue

        # ── Gap-finder gating lock (audit 2026-04-29) ─────────────────
        # pending_gap / pending_gap_v1 / pending_gap_v2 rows must NOT
        # reach Recalls until claude_check flips them to "pending".
        # Failures are still routed through rejected_flags above by the
        # gate scripts. This is the structural guarantee that the FSANZ
        # "Australian food" landing-page incident cannot recur.
        if clean.get("Status") in GAP_GATING_STATUSES and idx not in rejected_flags:
            kept_in_pending.append(clean)
            continue

        if idx in rejected_flags:
            reason = rejected_flags[idx]
            # Use the counter — returns should_delete=True after 2+ different
            # reviewers reject this row. Otherwise just stamp Status=rejected
            # and append rejection reason to Notes.
            should_delete, _ = mark_rejected_with_counter(clean, reason)
            if should_delete:
                # 2 reviewers agree it's bad — archive to Rejected sheet.
                _archive(clean)
                continue
            kept_in_pending.append(clean)
            continue

        # ── Date-consistency gate (audit 2026-05-06: defense-in-depth) ──
        # Last-line defense against date-extractor fallback bugs. If
        # Notes mentions a year far older than the Date field, the
        # extractor probably stamped today's date on an archived page.
        # Refuse to promote; row stays in Pending for manual review.
        date_problem = _check_date_consistency(clean)
        if date_problem:
            should_delete, _ = mark_rejected_with_counter(clean, date_problem)
            if should_delete:
                _archive(clean)
                continue
            kept_in_pending.append(clean)
            continue

        # Approved row: dedup against existing Recalls
        k = _dedup_key(clean)
        if k in approved_keys:
            # Already published — drop silently from Pending
            continue
        approved_keys.add(k)

        # Strip pending-only tracking columns before inserting into Recalls.
        # Fill RECALLS_SCHEMA, including the internal tracking columns:
        #   DateAdded   = today (when row first promoted to Recalls)
        #   LastUpdated = today (initial insert counts as an update)
        #   LastChecked = "" (no URL re-validation has happened yet —
        #                     url_guardian/url_gate will fill this later)
        from datetime import date as _today_fn
        _today = _today_fn.today().isoformat()
        approved_row = {col: clean.get(col, "" if col not in ("Tier", "Outbreak") else 0)
                        for col in SCHEMA}
        approved_row["DateAdded"] = _today
        approved_row["LastUpdated"] = _today
        approved_row["LastChecked"] = ""
        new_approved.append(approved_row)

    rejected_kept = sum(1 for r in kept_in_pending if r.get("Status") == STATUS_REJECTED)
    log.info("Promotion: %d approved -> Recalls, %d kept in Pending (%d rejected, "
             "%d archived to Rejected sheet)",
             len(new_approved), len(kept_in_pending), rejected_kept,
             len(archived_rejected))
    return new_approved, kept_in_pending, archived_rejected


# ---------------------------------------------------------------------------
# Sort / Save
# ---------------------------------------------------------------------------
def sort_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Sort newest first by Date string (YYYY-MM-DD sorts lexically).

    Defensive: rows can land here with Date as a datetime/date object when
    a cell in the xlsx was manually edited and Excel auto-typed it as a
    date. Coerce every key to a YYYY-MM-DD string before comparing so the
    sort never crashes on mixed types.
    """
    def _key(r: Dict[str, Any]) -> str:
        d = r.get("Date")
        if d is None or d == "":
            return ""
        # datetime / date object → ISO string
        if hasattr(d, "strftime"):
            try:
                return d.strftime("%Y-%m-%d")
            except (TypeError, ValueError):
                return ""
        # Anything else → string (truncate to first 10 chars to drop time)
        return str(d)[:10]
    return sorted(rows, key=_key, reverse=True)


def _write_sheet(wb: Workbook,
                 sheet_name: str,
                 schema: List[str],
                 rows: List[Dict[str, Any]],
                 header_fill: PatternFill = None) -> None:
    """(Re)create a sheet with given schema + rows.

    Defensive: Date cells are forced to YYYY-MM-DD strings with General
    number_format so an upstream datetime never gets written back as a
    typed-date cell (which would re-introduce the Excel-serial-leak bug).
    """
    from datetime import datetime as _dt, date as _dt_date
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    for i, h in enumerate(schema, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True)
        if header_fill is not None:
            c.fill = header_fill
    for r_idx, row in enumerate(rows, 2):
        for c_idx, col in enumerate(schema, 1):
            v = row.get(col, "")
            if col in ("Tier", "Outbreak"):
                try:
                    v = int(v) if v not in ("", None) else 0
                except (ValueError, TypeError):
                    v = 0
            elif col == "Date":
                # Force every Date cell to a string + General format
                if isinstance(v, (_dt, _dt_date)):
                    try:
                        v = v.strftime("%Y-%m-%d")
                    except (TypeError, ValueError):
                        v = ""
                elif v not in (None, "") and not isinstance(v, str):
                    v = str(v)[:10]
            cell = ws.cell(row=r_idx, column=c_idx, value=v)
            if col == "Date":
                cell.number_format = "General"
    ws.freeze_panes = "A2"


def save_xlsx_with_pending(
    approved_rows: List[Dict[str, Any]],
    pending_rows: List[Dict[str, Any]],
    xlsx_path: Path,
    newly_rejected_rows: Optional[List[Dict[str, Any]]] = None,
) -> None:
    """
    Save BOTH sheets (Recalls + Pending), preserving NEWS sheet if present.
    Sheet order: Recalls (0), Pending (1), Rejected (2), NEWS (last).

    Audit 2026-05-08: optional `newly_rejected_rows` arg. When supplied, these
    rows are APPENDED to the Rejected sheet (audit archive — never overwritten).
    Pre-existing rows in the Rejected sheet are preserved across runs so
    operators can inspect the full history of what 2-reviewer-rejection killed.
    """
    if xlsx_path.exists():
        wb = load_workbook(xlsx_path)
    else:
        wb = Workbook()
        if wb.active and wb.active.max_row == 1 and wb.active.max_column == 1:
            wb.remove(wb.active)

    # Write Recalls (approved published data)
    _write_sheet(wb, "Recalls", RECALLS_SCHEMA, approved_rows)

    # Write Pending (amber-ish header fill to make the tab visually distinct)
    pending_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    _write_sheet(wb, "Pending", PENDING_SCHEMA, pending_rows, header_fill=pending_fill)

    # ── Rejected (audit archive — append-only) ─────────────────────────
    # Read any existing rows first; append new rejections; rewrite full sheet.
    if newly_rejected_rows:
        existing_rejected: List[Dict[str, Any]] = []
        if "Rejected" in wb.sheetnames:
            ws_rej = wb["Rejected"]
            headers_rej = [c.value for c in ws_rej[1]]
            for r in ws_rej.iter_rows(min_row=2, values_only=True):
                if all(v in (None, "") for v in r):
                    continue
                rec = {h: (v if v is not None else "") for h, v in zip(headers_rej, r)}
                # Backfill missing columns
                for col in REJECTED_SCHEMA:
                    rec.setdefault(col, "" if col not in ("Tier", "Outbreak") else 0)
                existing_rejected.append(rec)
        # Defensive: don't double-append a row whose URL+RejectedAt already
        # appears in existing_rejected (re-runs of save with same archive list).
        seen = {(r.get("URL", ""), r.get("RejectedAt", "")) for r in existing_rejected}
        deduped_new = [r for r in newly_rejected_rows
                       if (r.get("URL", ""), r.get("RejectedAt", "")) not in seen]
        all_rejected = existing_rejected + deduped_new
        rejected_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA",
                                    fill_type="solid")  # rose
        _write_sheet(wb, "Rejected", REJECTED_SCHEMA, all_rejected,
                     header_fill=rejected_fill)
        if deduped_new:
            log.info("Rejected sheet: appended %d row(s) (total now %d)",
                     len(deduped_new), len(all_rejected))

    # Ensure NEWS sheet exists (empty if it wasn't there before)
    if "NEWS" not in wb.sheetnames:
        news = wb.create_sheet("NEWS")
        for i, h in enumerate(NEWS_HEADERS, 1):
            c = news.cell(row=1, column=i, value=h)
            c.font = Font(bold=True)
        news.freeze_panes = "A2"

    # Reorder: Recalls, Pending, Rejected, (others), NEWS last
    ordered = ["Recalls", "Pending"]
    if "Rejected" in wb.sheetnames:
        ordered.append("Rejected")
    others = [s for s in wb.sheetnames
              if s not in ("Recalls", "Pending", "Rejected", "NEWS")]
    ordered += others
    if "NEWS" in wb.sheetnames:
        ordered.append("NEWS")
    wb._sheets = [wb[s] for s in ordered]

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
    log.info("Saved xlsx: Recalls=%d, Pending=%d -> %s",
             len(approved_rows), len(pending_rows), xlsx_path)


def save_json(rows: List[Dict[str, Any]], json_path: Path) -> None:
    """
    DEPRECATED — writes recalls.json from an in-memory list of rows.

    Using this is an architectural violation: recalls.json MUST mirror what's
    on the Recalls sheet of recalls.xlsx, not an arbitrary in-memory list.
    If the xlsx write fails or gets interrupted, the json would diverge from
    the file that's actually committed.

    Use `mirror_json_from_xlsx(xlsx_path, json_path)` instead. Kept here only
    so legacy callers don't crash outright during the transition.
    """
    log.warning("save_json (in-memory) is deprecated; "
                "use mirror_json_from_xlsx for guaranteed xlsx->json sync")
    json_path.parent.mkdir(parents=True, exist_ok=True)
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(rows, f, ensure_ascii=False, indent=1, default=str)
    log.info("Saved %d approved rows to %s", len(rows), json_path)


def mirror_json_from_xlsx(xlsx_path: Path, json_path: Path) -> int:
    """
    Write recalls.json as a strict mirror of the Recalls sheet in recalls.xlsx.

    This is the ONLY sanctioned way to produce recalls.json. It guarantees
    that json can never drift from xlsx: we read the file that was just
    committed to disk, normalise types (dates to ISO strings), and serialise.

    INTERNAL columns (DateAdded, LastUpdated, LastChecked) are STRIPPED
    here so they don't leak into the public-facing dashboard. Only the
    14 SCHEMA columns make it to recalls.json.

    Returns the number of rows written.
    """
    rows = load_existing(xlsx_path)
    out = []
    for r in rows:
        rec = {}
        for k, v in r.items():
            # Skip internal-only tracking columns — public consumers never see these
            if k in RECALLS_INTERNAL_COLUMNS:
                continue
            if hasattr(v, "isoformat"):
                rec[k] = v.isoformat()[:10]
            elif v is None:
                rec[k] = ""
            else:
                rec[k] = v
        out.append(rec)
    json_path.parent.mkdir(parents=True, exist_ok=True)
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=1, default=str)
    log.info("Mirrored %d rows from xlsx -> %s", len(out), json_path)
    return len(out)


# ---------------------------------------------------------------------------
# Daily-brief rebuild helper (used by url_gate, claude_check, merge_master CLI)
# ---------------------------------------------------------------------------
def rebuild_daily_briefs_for_promoted(
    new_approved: List[Dict[str, Any]],
    full_approved: List[Dict[str, Any]],
) -> Tuple[List[str], List[str]]:
    """
    After Pending → Recalls promotion, rebuild the per-date daily-brief HTML
    for every date that gained at least one new row. Without this, the
    dashboard's DAILY tab and rolling 7-day display stay stale until the
    next scheduled daily-recall-search run.

    Args:
        new_approved : rows just promoted to Recalls this run
        full_approved: the FULL Recalls sheet AFTER promotion (so we render
                       the complete day, not just the newly-added rows)

    Returns:
        (rebuilt_brief_paths, files_to_commit) — caller is responsible for
        adding `files_to_commit` to its git_commit_and_push call. Both lists
        are empty when nothing was promoted or the brief renderer module
        is unavailable.
    """
    files_to_commit: List[str] = []
    rebuilt_briefs: List[str] = []
    if not new_approved:
        return rebuilt_briefs, files_to_commit

    try:
        from pipeline.daily_recall_search import (  # noqa: WPS433
            render_daily_html, update_daily_index,
        )
        from scrapers._models import Recall as _Recall  # noqa: WPS433
    except ImportError as ie:
        log.warning("Cannot import brief renderer (%s) — skipping daily "
                    "brief rebuild", ie)
        return rebuilt_briefs, files_to_commit

    from collections import defaultdict
    from datetime import date as _date

    # Group newly-promoted rows by Date
    by_date: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for r in new_approved:
        d = str(r.get("Date") or "").strip()[:10]
        if d:
            by_date[d].append(r)

    # Fast-lookup of ALL Recalls by date so we render the full day, not
    # only the newly-promoted rows.
    full_by_date: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for r in full_approved:
        d = str(r.get("Date") or "").strip()[:10]
        if d:
            full_by_date[d].append(r)

    for date_str in sorted(by_date.keys(), reverse=True):
        try:
            y, m, d = date_str.split("-")
            target = _date(int(y), int(m), int(d))
        except (ValueError, AttributeError):
            log.warning("Skip brief rebuild — bad date '%s'", date_str)
            continue

        day_rows = full_by_date.get(date_str, [])
        recalls_objs = []
        for row in day_rows:
            try:
                recalls_objs.append(_Recall(**{
                    k: (v if v is not None else "")
                    for k, v in row.items()
                    if k in _Recall.__annotations__
                }))
            except Exception as cce:  # noqa: BLE001
                log.debug("skip row coerce: %s", cce)

        try:
            render_daily_html(target, recalls_objs)
            update_daily_index(target, recalls_objs)
            brief_path = f"docs/daily/{date_str}.html"
            rebuilt_briefs.append(brief_path)
            files_to_commit.append(brief_path)
            log.info("Rebuilt daily brief for %s (%d rows)",
                     date_str, len(recalls_objs))
        except Exception as rerr:  # noqa: BLE001
            log.warning("Brief rebuild failed for %s: %s", date_str, rerr)

    if rebuilt_briefs:
        files_to_commit.append("docs/daily-index.json")

    return rebuilt_briefs, files_to_commit


# ---------------------------------------------------------------------------
# Back-compat shims (kept so legacy call sites don't break)
# ---------------------------------------------------------------------------
def save_xlsx(rows: List[Dict[str, Any]], xlsx_path: Path) -> None:
    """DEPRECATED: single-sheet save. Kept for any legacy caller."""
    log.warning("save_xlsx (single-sheet) is deprecated — use save_xlsx_with_pending")
    existing_pending = load_pending(xlsx_path)
    save_xlsx_with_pending(rows, existing_pending, xlsx_path)


def merge_new(existing: List[Dict[str, Any]], new_recalls: List[Recall]) -> List[Dict[str, Any]]:
    """
    DEPRECATED: direct merge into Recalls (pre-Pending-sheet behavior).
    Kept for any back-compat call; new code should use append_to_pending +
    promote_approved instead.
    """
    existing_keys = {_dedup_key(r) for r in existing}
    merged = list(existing)
    appended = 0
    for r in new_recalls:
        d = r.to_dict() if isinstance(r, Recall) else dict(r)
        for col in SCHEMA:
            d.setdefault(col, "" if col not in ("Tier", "Outbreak") else 0)
        k = _dedup_key(d)
        if k in existing_keys:
            continue
        existing_keys.add(k)
        merged.append(d)
        appended += 1
    log.info("merge_new (legacy): %d existing + %d new = %d total",
             len(existing), appended, len(merged))
    return merged


# ---------------------------------------------------------------------------
# NEWS sheet merge (for RSS news feed scrapers)
# ---------------------------------------------------------------------------
def _news_dedup_key(row: Dict[str, Any]) -> str:
    """Dedup key for a NEWS row: link URL (lowered, stripped)."""
    link = (row.get("Link") or row.get("link") or "").strip().lower()
    if link:
        return link
    title = (row.get("Title") or row.get("title") or "").strip().lower()[:80]
    return f"{row.get('Published (UTC)', '')}|{title}"


def load_news(xlsx_path: Path) -> List[Dict[str, str]]:
    """Load existing NEWS rows from the xlsx."""
    if not xlsx_path.exists():
        return []
    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        if "NEWS" not in wb.sheetnames:
            wb.close()
            return []
        ws = wb["NEWS"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            d = {}
            for i, v in enumerate(row):
                if i < len(headers) and headers[i]:
                    d[headers[i]] = str(v) if v is not None else ""
            if d.get("Title") or d.get("Link"):
                rows.append(d)
        wb.close()
        return rows
    except Exception as e:
        log.warning("Failed to load NEWS sheet: %s", e)
        return []


def append_news_to_xlsx(
    xlsx_path: Path,
    new_items: List[Dict[str, str]],
) -> int:
    """
    Append new NEWS items to the NEWS sheet, deduped against existing rows.
    Returns count of actually-appended items.
    """
    if not new_items:
        return 0

    existing = load_news(xlsx_path)
    seen_keys = {_news_dedup_key(r) for r in existing}

    to_add = []
    for item in new_items:
        k = _news_dedup_key(item)
        if k in seen_keys:
            continue
        seen_keys.add(k)
        to_add.append(item)

    if not to_add:
        log.info("NEWS merge: 0 new items (all duplicates)")
        return 0

    # Open the workbook and append rows to the NEWS sheet
    wb = load_workbook(xlsx_path)
    if "NEWS" not in wb.sheetnames:
        ws = wb.create_sheet("NEWS")
        for i, h in enumerate(NEWS_HEADERS, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = Font(bold=True)
        ws.freeze_panes = "A2"
    else:
        ws = wb["NEWS"]

    for item in to_add:
        row_vals = [item.get(h, "") for h in NEWS_HEADERS]
        ws.append(row_vals)

    wb.save(xlsx_path)
    log.info("NEWS merge: appended %d new items (total now %d)",
             len(to_add), len(existing) + len(to_add))
    return len(to_add)



# =========================================================================
# CLI entry point — run by the hourly merge-master workflow.
#
# AUDIT 2026-04-29 — promotion semantics tightened:
#   The hourly CLI used to call promote_approved() with rejected_flags driven
#   only by review/url_validator (HTTP HEAD/GET reachability). That meant
#   ANY row with a 200-returning URL was promoted to Recalls — including
#   hallucinated RappelConso fiches (e.g. /fiche-rappel/22180/Interne) that
#   render a soft-200 page even when the fiche ID doesn't exist. This let
#   ~7 hallucinated Gemini gap-finder rows leak from the 07:34 Exa run into
#   Recalls before the once-daily 07:00 Gemini URL gate could catch them.
#
#   New rule (per George 2026-04-29):
#     "Only data from URL Gemini followed by Claude check must be allowed
#      1 or two times per day."
#
#   The hourly CLI is now a JANITOR ONLY — it cleans malformed URLs from
#   Pending and removes Pending rows whose URL has been confirmed dead,
#   but it NEVER promotes to Recalls. Promotion happens exclusively via:
#     1. pipeline/url_gate_gemini.py     (07:00 Athens — Gemini URL gate)
#     2. pipeline/claude_check.py        (07:45 Athens — Claude content check)
#
#   To override (e.g. backfill, manual catch-up), set MERGE_MASTER_PROMOTE=1.
# =========================================================================
if __name__ == "__main__":
    import sys, os
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    )
    from pathlib import Path
    ROOT = Path(__file__).resolve().parent.parent
    XLSX = ROOT / "docs" / "data" / "recalls.xlsx"

    if not XLSX.exists():
        log.info("No recalls.xlsx found — nothing to merge.")
        sys.exit(0)

    approved = load_existing(XLSX)
    pending = load_pending(XLSX)
    log.info("State: %d approved, %d pending", len(approved), len(pending))

    if not pending:
        log.info("No Pending rows — nothing to do.")
        sys.exit(0)

    # ── URL reachability check (cheap janitor pass) ─────────────────────
    # We still validate URLs at every hourly run so dead URLs can be
    # marked rejected in Pending. We do NOT use the result to drive
    # promotion — that would re-create the leak this CLI is here to
    # prevent.
    from review.url_validator import validate_all
    log.info("Validating %d Pending URLs (janitor pass — no promotion)...",
             len(pending))
    validated = validate_all(pending, max_workers=5)

    # Mark broken-URL rows as rejected so they don't promote later.
    # Anything that's currently 'pending' AND has a confirmed bad URL
    # gets stamped REJECTED so the next gate pass skips it.
    rejected_flags = {}
    for idx, row in enumerate(validated):
        check = row.get("_url_check", {})
        # Only reject if the URL check says it's BROKEN. Reachable URLs
        # stay 'pending' until the URL gate validates them properly.
        if not check.get("ok", False):
            rejected_flags[idx] = check.get("reason", "URL check failed")

    # Strip the runtime _url_check field before persisting
    clean_pending = [{k: v for k, v in row.items() if k != "_url_check"}
                     for row in validated]

    # ── Cleanup orphan rejected rows (audit 2026-05-05) ─────────────────
    # Physically delete Pending rows that are already rejected and either:
    #   (a) have RejectedBy with 2+ different reviewers (counter triggered)
    #   (b) are older than 24h (orphaned from before the counter existed)
    # Runs every cycle so the Pending sheet doesn't accumulate stale
    # rejections forever.
    clean_pending, n_cleaned = cleanup_orphan_rejected(clean_pending)
    if n_cleaned > 0:
        log.info("Cleanup: physically deleted %d orphan rejected rows", n_cleaned)
        # Re-index rejected_flags after deletion: any flag pointing at a now-
        # deleted index is stale, but since we iterated by index BEFORE
        # the cleanup, the flags still match positions in `validated`. After
        # cleanup, `clean_pending` has fewer entries — rejected_flags must
        # be remapped or cleared. Simplest safe choice: clear it. The next
        # url-validator run will re-flag any remaining broken URLs.
        rejected_flags = {}

    # ── Promotion gate ──────────────────────────────────────────────────
    # OFF by default — only the once-daily Gemini URL gate (07:00) and
    # Claude check (07:45) workflows are permitted to promote rows to
    # Recalls. The hourly CLI just stamps rejections and exits.
    promote_enabled = os.environ.get("MERGE_MASTER_PROMOTE", "").strip() in (
        "1", "true", "yes")

    archived_rejected: List[Dict[str, Any]] = []
    if promote_enabled:
        log.info("MERGE_MASTER_PROMOTE=1 — promotion ENABLED for this run "
                 "(use only for backfill/manual catch-up)")
        new_approved, remaining, archived_rejected = promote_approved(
            clean_pending, approved, rejected_flags,
        )
        if new_approved:
            log.info("Promoted %d rows Pending → Recalls", len(new_approved))
            final_approved = sort_rows(approved + new_approved)
        else:
            log.info("No rows promoted this run.")
            final_approved = sort_rows(approved)
    else:
        log.info("Janitor mode (default): NOT promoting. Only the daily "
                 "Gemini URL gate + Claude check are allowed to advance "
                 "rows from Pending → Recalls.")
        # We still update Pending with rejection stamps for broken URLs.
        # Walk clean_pending; for any idx in rejected_flags, copy the
        # rejection note onto the row so the next gate pass skips it.
        # Audit 2026-05-08: when 2nd reviewer rejection lands here in
        # janitor mode, archive the row to Rejected sheet AND drop it
        # from remaining (was previously left orphaned for cleanup_orphan
        # to find on next run — now handled atomically).
        archive_ts = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
        survivors: List[Dict[str, Any]] = []
        for idx, row in enumerate(clean_pending):
            if idx in rejected_flags and row.get("Status") != STATUS_REJECTED:
                reason = rejected_flags[idx]
                should_delete, _ = mark_rejected_with_counter(row, reason)
                if should_delete:
                    row["RejectedAt"] = archive_ts
                    archived_rejected.append(row)
                    continue
            survivors.append(row)
        new_approved = []
        remaining = survivors
        final_approved = sort_rows(approved)
        if rejected_flags:
            log.info("Marked %d Pending row(s) as rejected; archived %d to "
                     "Rejected sheet (2nd-reviewer)",
                     len(rejected_flags), len(archived_rejected))

    save_xlsx_with_pending(final_approved, sort_rows(remaining), XLSX,
                           newly_rejected_rows=archived_rejected)
    mirror_json_from_xlsx(XLSX, ROOT / "docs" / "data" / "recalls.json")

    # Mirror promotions into the Weekly_Review sheet + refresh the JSON
    # slice consumed by the Apps Script Thursday-17:00 mailer. Captures
    # the (rare) backfill / MERGE_MASTER_PROMOTE=1 path too.
    if new_approved:
        try:
            from pipeline.weekly_review_capture import record_promotions  # noqa: E402
            n_wr = record_promotions(new_approved, xlsx_path=XLSX)
            if n_wr:
                log.info("Weekly_Review: appended %d row(s)", n_wr)
        except Exception as _wr_err:
            log.warning("Weekly_Review capture failed: %s", _wr_err)

    # Rebuild daily briefs for any date that gained newly-promoted rows.
    # Without this, the dashboard's rolling 7-day display stays stale.
    # In janitor mode (no promotion), new_approved is empty — this is a
    # cheap no-op.
    rebuilt_briefs, brief_files = rebuild_daily_briefs_for_promoted(
        new_approved, final_approved,
    )
    if rebuilt_briefs:
        log.info("Rebuilt %d daily brief(s)", len(rebuilt_briefs))

    log.info("Done. Recalls=%d, Pending=%d", len(final_approved), len(remaining))
