"""Pending-sheet cleanup — drop stale rows that will never promote.

WHAT THIS DOES
==============
The Pending sheet has accumulated rows that no automated process will
ever clear:

  * `pending_gap_v2` rows whose underlying recall is too old to fall
    inside any since_days window (typically Tavily / Exa / DDGS finds
    from years ago that the gate flagged as old).
  * `pending_gap` rows whose source content is not actually a recall
    (e.g. NVWA "no cereulide detected in samples" — a clearing notice,
    not a recall).
  * `rejected` rows that have already been processed by claude-check
    but kept around for retry. After 12+ hours they're noise.

These rows clutter Pending, slow down every merge_master run (they're
scanned on each gate-check pass), and confuse the dashboard's Pending-
count display.

This script identifies them by a deterministic rule set, prints what
it would do, and (with `--apply`) deletes them from the xlsx with a
backup written first.

URL TRUNCATION DETECTION
========================
While cleaning, the script also detects URL truncation mismatches
between Recalls and Pending: cases where a recall exists in both
sheets under DIFFERENT URLs because one was a manual paste with a
truncated slug and the other was the full scraper-discovered URL.
These are reported but NOT auto-fixed — the operator must decide
which URL is canonical. As of 2026-05-18 audit, one such mismatch
exists: Kyan Culture (Recalls has ...recalled-due-pathogenic-e,
Pending has ...recalled-due-e-coli).

USAGE
=====
From repo root:

    python pending_cleanup.py                  # dry-run, prints plan
    python pending_cleanup.py --apply          # commits changes
    python pending_cleanup.py --apply --no-backup

The xlsx path defaults to docs/data/recalls.xlsx. Override with --xlsx.

The script is idempotent: running it twice in --apply mode after a
clean run will do nothing the second time.
"""
from __future__ import annotations

import argparse
import shutil
import sys
from datetime import datetime, timezone
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed. `pip install openpyxl`")
    sys.exit(2)


DEFAULT_XLSX = "docs/data/recalls.xlsx"


# Decision rules for what to drop. Returns (drop: bool, reason: str).
# Operates on a dict-shaped row keyed by header names.
def evaluate_row(row: dict, now_utc: datetime) -> tuple:
    status = (row.get("Status") or "").strip()
    url = (row.get("URL") or "").strip()
    company = (row.get("Company") or "").strip()
    pathogen = (row.get("Pathogen") or "").strip()
    notes = (row.get("Notes") or "").strip()
    scraped_at = row.get("ScrapedAt")

    # Compute age in days, if possible
    age_days = None
    if scraped_at:
        try:
            sa_dt = datetime.fromisoformat(str(scraped_at).replace("Z", "+00:00"))
            age_days = (now_utc - sa_dt).total_seconds() / 86400.0
        except (ValueError, TypeError):
            pass

    # Rule 1: pending_gap_v2 older than 2 days. The gap-gate has already
    # had a v2 pass on it; if it still hasn't promoted, the content is
    # almost certainly out of window or not a real recall.
    if status == "pending_gap_v2" and age_days is not None and age_days > 2.0:
        return (True, f"pending_gap_v2 stale ({age_days:.1f}d old)")

    # Rule 2: pending_gap rows whose Notes text indicates a CLEARING /
    # NO-CONTAMINATION announcement rather than a recall. NVWA's
    # "first results: samples contain no measurable cereulide" is the
    # archetypal case — a press release about NOT recalling something.
    no_recall_phrases = (
        "geen meetbare",   # NL: "no measurable"
        "no measurable",   # EN
        "not a recall",
        "no recall",
        "clearance notice",
        "lifting recall",  # post-recall lift announcement
        "recall closed",
        "recall withdrawn",
        "no contamination found",
        "samples tested negative",
    )
    if status.startswith("pending_gap"):
        lo = notes.lower()
        if any(p in lo for p in no_recall_phrases):
            for p in no_recall_phrases:
                if p in lo:
                    return (True, f"pending_gap with no-recall phrase {p!r}")

    # Rule 3: rejected rows older than 12 hours. claude-check has had
    # its pass; further retention is pointless noise.
    if status == "rejected" and age_days is not None and age_days > 0.5:
        return (True, f"rejected, {age_days:.1f}d old (>0.5d retention cap)")

    return (False, "")


def _slug_of(url: str) -> str:
    """Extract the trailing slug from a recall URL (the path component after
    the last '/'), stripped of query string and lowercased."""
    if not url:
        return ""
    bare = str(url).split("?", 1)[0].rstrip("/")
    return bare.rsplit("/", 1)[-1].lower()


def find_url_truncation_mismatches(wb) -> list:
    """Find rows where Pending and Recalls share a recall but URLs differ
    by truncation or near-truncation. Compares at the slug-word level
    (slugs split by hyphens), NOT URL-string prefix — otherwise every pair
    of FDA URLs trivially matches on their identical /safety/recalls-... path.

    A pair is reported as a mismatch when, AFTER extracting and tokenising
    each URL's trailing slug:
      - The slugs differ (URLs are not identical), AND
      - The slugs share at least _MIN_COMMON_WORDS leading hyphen-separated
        words (so they describe the same underlying recall).

    The minimum-common-words threshold is high enough that unrelated recalls
    (which share at most 1-2 generic leading words like a company name
    prefix that happens to recur) don't generate false positives. For the
    Kyan Culture case the slugs share 11 leading words, well above threshold.
    """
    _MIN_COMMON_WORDS = 5

    ws_r = wb["Recalls"]
    ws_p = wb["Pending"]
    hdr_r = [c.value for c in ws_r[1]]
    hdr_p = [c.value for c in ws_p[1]]
    u_i_r = hdr_r.index("URL")
    u_i_p = hdr_p.index("URL")

    pending_urls = []
    for row in ws_p.iter_rows(min_row=2, values_only=True):
        u = row[u_i_p]
        if not u:
            continue
        pending_urls.append(str(u))

    mismatches = []
    seen_pairs = set()
    for row in ws_r.iter_rows(min_row=2, values_only=True):
        ru = row[u_i_r]
        if not ru:
            continue
        ru = str(ru)
        r_slug_words = [w for w in _slug_of(ru).split("-") if w]
        if len(r_slug_words) < _MIN_COMMON_WORDS:
            continue
        for pu in pending_urls:
            if pu == ru:
                continue
            p_slug_words = [w for w in _slug_of(pu).split("-") if w]
            if len(p_slug_words) < _MIN_COMMON_WORDS:
                continue
            # Count common leading words
            common = 0
            for a, b in zip(r_slug_words, p_slug_words):
                if a == b:
                    common += 1
                else:
                    break
            if common >= _MIN_COMMON_WORDS:
                key = (ru, pu)
                if key not in seen_pairs:
                    seen_pairs.add(key)
                    mismatches.append((ru, pu))
    return mismatches


def main():
    ap = argparse.ArgumentParser(description=__doc__.split("\n\n")[0])
    ap.add_argument("--xlsx", default=DEFAULT_XLSX, help="path to recalls.xlsx")
    ap.add_argument("--apply", action="store_true",
                    help="commit changes (default: dry-run)")
    ap.add_argument("--no-backup", action="store_true",
                    help="skip the .bak backup when applying")
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"ERROR: xlsx not found: {xlsx_path}")
        sys.exit(2)

    now_utc = datetime.now(timezone.utc)
    print(f"Pending cleanup — {now_utc.isoformat(timespec='seconds')}")
    print(f"  xlsx: {xlsx_path}")
    print(f"  mode: {'APPLY' if args.apply else 'DRY-RUN (no changes)'}")
    print()

    wb = load_workbook(xlsx_path)
    if "Pending" not in wb.sheetnames:
        print("ERROR: no 'Pending' sheet in workbook")
        sys.exit(2)
    ws = wb["Pending"]
    headers = [c.value for c in ws[1]]

    # Collect drop decisions
    drops = []   # list of (row_index, reason, summary)
    for ri, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        rec = dict(zip(headers, row))
        drop, reason = evaluate_row(rec, now_utc)
        if drop:
            summary = (
                f"status={rec.get('Status')!r:24} "
                f"company={(rec.get('Company') or '')[:36]!r:38} "
                f"url={(rec.get('URL') or '')[:60]!r}"
            )
            drops.append((ri, reason, summary))

    # Print drop plan
    print(f"=== Rows to drop ({len(drops)}) ===")
    if not drops:
        print("  (none — Pending is clean)")
    else:
        for ri, reason, summary in drops:
            print(f"  row {ri:3}: {reason}")
            print(f"           {summary}")
    print()

    # Print URL-truncation mismatches
    mismatches = find_url_truncation_mismatches(wb)
    print(f"=== URL-truncation mismatches ({len(mismatches)}) ===")
    if not mismatches:
        print("  (none — URLs in Pending and Recalls align)")
    else:
        for ru, pu in mismatches:
            print(f"  RECALLS:  {ru}")
            print(f"  PENDING:  {pu}")
            print(f"  ACTION:   Operator decide which URL is canonical and update")
            print(f"            the Recalls row manually. Then re-run cleanup.")
            print()

    if not args.apply:
        print()
        print("Dry-run complete. Re-run with --apply to commit changes.")
        return

    if not drops:
        print("No changes to apply.")
        return

    # Backup before modifying
    if not args.no_backup:
        backup_path = xlsx_path.with_suffix(
            f".bak-{now_utc.strftime('%Y%m%dT%H%M%S')}.xlsx"
        )
        shutil.copy2(xlsx_path, backup_path)
        print(f"Backup written: {backup_path}")

    # Delete in reverse order so earlier indices stay valid
    for ri, _, _ in sorted(drops, reverse=True):
        ws.delete_rows(ri, 1)

    wb.save(xlsx_path)
    print(f"Applied: dropped {len(drops)} row(s) from Pending in {xlsx_path}")
    print()
    print("Re-run claude-check or merge_master afterwards to refresh the")
    print("snapshot. The dropped rows will NOT come back unless their")
    print("source URL re-appears in a fresh scrape (and even then, only")
    print("if it passes the gates this time).")


if __name__ == "__main__":
    main()
