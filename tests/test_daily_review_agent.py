"""tests/test_daily_review_agent.py — coverage for the FSIS Daily Review Agent.

Exercises the two-lane safety model:
  • Lane A deterministic fixes fire and apply correctly.
  • Lane B detectors queue (never delete) the four duplicate classes, reject
    the same-source-distinct-ID false positive, and honour the ledger.
  • A --dry-run touches neither recalls.xlsx nor recalls.json.
"""
from __future__ import annotations

import json
from datetime import date

import pytest
from openpyxl import Workbook, load_workbook

from pipeline import daily_review_agent as A
from pipeline import merge_master as mm

TODAY = date(2026, 7, 16)


# ── Lane A ────────────────────────────────────────────────────────────────
def _base(**over):
    row = {c: "" for c in mm.RECALLS_SCHEMA}
    row.update({"Date": "2026-07-10", "Company": "ACME", "Product": "Cheese",
                "Pathogen": "Salmonella", "Tier": 1, "URL": "https://fda.gov/x",
                "DateAdded": "2026-07-10"})
    row.update(over)
    return row


def test_lane_a_tier1_enforce_detected_and_applied():
    rows = [_base(Pathogen="Listeria monocytogenes", Tier=3)]
    changes = A.compute_lane_a(rows, TODAY)
    assert any(c["fix"] == "tier1_enforce" for c in changes)
    A.apply_lane_a(rows, changes, TODAY)
    assert int(rows[0]["Tier"]) == 1
    assert "tier-guard" in rows[0]["Notes"]


def test_lane_a_stamps_blank_dateadded_with_athens_today():
    rows = [_base(DateAdded="")]
    changes = A.compute_lane_a(rows, TODAY)
    A.apply_lane_a(rows, changes, TODAY)
    assert rows[0]["DateAdded"] == TODAY.isoformat()


def test_lane_a_strips_homepage_artifact():
    rows = [_base(Product="lebensmittelwarnung.de - Homepage - Instant Noodles")]
    changes = A.compute_lane_a(rows, TODAY)
    A.apply_lane_a(rows, changes, TODAY)
    assert rows[0]["Product"] == "Instant Noodles"


def test_lane_a_strip_homepage_leaves_clean_product_untouched():
    rows = [_base(Product="Sliced Ham 200g")]
    assert not any(c["fix"] == "strip_homepage" for c in A.compute_lane_a(rows, TODAY))


def test_lane_a_recovers_blank_date_from_url_slug():
    rows = [_base(Date="", URL="https://reg.example/alert/2026-07-07/ecoli")]
    changes = A.compute_lane_a(rows, TODAY)
    A.apply_lane_a(rows, changes, TODAY)
    assert rows[0]["Date"] == "2026-07-07"
    assert rows[0]["report_week"] == mm.compute_report_week("2026-07-07")
    assert "date-fix" in rows[0]["Notes"]


# ── Lane B ────────────────────────────────────────────────────────────────
def _lb(**over):
    row = {c: "" for c in mm.RECALLS_SCHEMA}
    row.update({"Tier": 1, "Source": "FDA", "Country": "USA"})
    row.update(over)
    return row


def test_lane_b_duplicate_url_queues_non_survivor():
    rows = [
        _lb(Date="2026-07-01", Company="Dup", Pathogen="Listeria",
            URL="https://fda.gov/recall/999", Notes="official-feed"),
        _lb(Date="2026-07-01", Company="Dup", Pathogen="Listeria",
            URL="https://www.fda.gov/recall/999/"),
    ]
    proposed = A.detect_lane_b(rows, resolved=set())
    assert len(proposed) == 1
    assert proposed[0]["category"] == "duplicate_url"
    assert proposed[0]["duplicate_of"]["URL"] == "https://fda.gov/recall/999"


def test_lane_b_cfia_fr_en_pair_keeps_english():
    rows = [
        _lb(Date="2026-07-02", Company="Maple", Pathogen="Salmonella", Source="CFIA",
            Country="Canada", Notes="official-feed",
            URL="https://recalls-rappels.canada.ca/en/alert-recall/maple-tahini-salmonella"),
        _lb(Date="2026-07-02", Company="Maple", Pathogen="Salmonella", Source="CFIA",
            Country="Canada",
            URL="https://recalls-rappels.canada.ca/fr/rappel-avis/tahini-maple-salmonella"),
    ]
    proposed = A.detect_lane_b(rows, resolved=set())
    assert [p["category"] for p in proposed] == ["cfia_fr_en"]
    assert "/fr/" in proposed[0]["row"]["URL"]
    assert "/en/" in proposed[0]["duplicate_of"]["URL"]


def test_lane_b_cfs_foreign_repost_flagged_and_hk_kept():
    rows = [
        _lb(Date="2026-03-23", Company="MOMA", Pathogen="Listeria", Source="CFS HK",
            Country="United Kingdom",
            URL="https://www.cfs.gov.hk/english/rc/20260323_1.pdf"),
        _lb(Date="2026-03-24", Company="Local", Pathogen="Salmonella", Source="CFS HK",
            Country="Hong Kong",
            URL="https://www.cfs.gov.hk/english/rc/20260324_1.pdf"),
    ]
    proposed = A.detect_lane_b(rows, resolved=set())
    cats = [p["category"] for p in proposed]
    assert cats == ["cfs_foreign_repost"]
    assert proposed[0]["row"]["Country"] == "United Kingdom"


def test_lane_b_cross_source_dup_requires_two_sources():
    rows = [
        _lb(Date="2026-04-15", Company="Bruno", Pathogen="Salmonella",
            Source="RASFF (EU)", Country="Brazil",
            URL="https://webgate.ec.europa.eu/rasff/notification/111"),
        _lb(Date="2026-04-15", Company="Bruno", Pathogen="Salmonella",
            Source="RappelConso (FR)", Country="Brazil", Notes="official-feed",
            URL="https://rappel.conso.gouv.fr/fiche-rappel/222/Interne"),
    ]
    proposed = A.detect_lane_b(rows, resolved=set())
    assert [p["category"] for p in proposed] == ["cross_source_dup"]


def test_same_source_distinct_ids_not_flagged():
    """Regression: two DIFFERENT recalls from one regulator on the same day
    (distinct fiche IDs) must NOT be proposed as duplicates."""
    rows = [
        _lb(Date="2026-07-10", Company="Jc David", Pathogen="Listeria",
            Source="RappelConso (FR)", Country="France",
            URL="https://rappel.conso.gouv.fr/fiche-rappel/22794/Interne"),
        _lb(Date="2026-07-10", Company="Jc David", Pathogen="Listeria",
            Source="RappelConso (FR)", Country="France",
            URL="https://rappel.conso.gouv.fr/fiche-rappel/22798/Interne"),
    ]
    assert A.detect_lane_b(rows, resolved=set()) == []


def test_shared_nonspecific_url_not_proposed_for_deletion():
    """Two DIFFERENT recalls sharing one non-specific listing URL must be
    reported as an integrity defect, never queued as a duplicate to delete."""
    rows = [
        _lb(Date="2026-06-12", Company="Fresh Tropical", Pathogen="Hydrocyanic acid",
            Source="Salute (IT)", Country="Italy",
            URL="https://www.salute.gov.it/new/external_data/avvisi"),
        _lb(Date="2026-06-12", Company="Azienda Agricola", Pathogen="Listeria",
            Source="Salute (IT)", Country="Italy",
            URL="https://www.salute.gov.it/new/external_data/avvisi"),
    ]
    proposed = A.detect_lane_b(rows, resolved=set())
    assert all(p["category"] != "duplicate_url" for p in proposed)
    integ = A.build_integrity_report(rows, TODAY)
    assert len(integ["shared_nonspecific_urls"]) == 1
    assert integ["duplicate_urls"] == []


def test_cfia_slug_variant_duplicate_flagged():
    """Same CFIA recall republished under a drifted slug + company = one dup."""
    rows = [
        _lb(Date="2026-07-10", Company="Charlevoisienne", Pathogen="Listeria",
            Source="CFIA", Country="Canada", Notes="official-feed",
            URL="https://recalls-rappels.canada.ca/en/alert-recall/charlevoisienne-and-joe-smoked-meat-brand-meat-products-recalled-due-listeria"),
        _lb(Date="2026-07-10", Company="Boucherie Charcuterie Lyn Tremblay",
            Pathogen="Listeria", Source="CFIA", Country="Canada",
            URL="https://recalls-rappels.canada.ca/en/alert-recall/charcuterie-charlevoisienne-and-joe-smoked-meat-brand-meat-products-recalled-due"),
    ]
    proposed = A.detect_lane_b(rows, resolved=set())
    assert [p["category"] for p in proposed] == ["cfia_duplicate"]


def test_cfia_same_nid_duplicate_flagged():
    rows = [
        _lb(Date="2026-05-01", Company="Firm A", Pathogen="Salmonella", Source="CFIA",
            Country="Canada", Notes="CFIA open-data NID=112233",
            URL="https://recalls-rappels.canada.ca/en/alert-recall/firm-a-product-recalled"),
        _lb(Date="2026-06-20", Company="Firm A updated", Pathogen="Salmonella",
            Source="CFIA", Country="Canada", Notes="CFIA open-data NID=112233",
            URL="https://recalls-rappels.canada.ca/en/alert-recall/firm-a-product-recalled-salmonella-2"),
    ]
    proposed = A.detect_lane_b(rows, resolved=set())
    assert [p["category"] for p in proposed] == ["cfia_duplicate"]
    assert "NID=112233" in proposed[0]["reason"]


def test_cfia_distinct_outbreak_recalls_not_merged():
    """Two DIFFERENT firms' recalls in one outbreak must NOT be merged."""
    rows = [
        _lb(Date="2026-04-21", Company="Kunafa's", Pathogen="Salmonella", Source="CFIA",
            Country="Canada",
            URL="https://recalls-rappels.canada.ca/en/alert-recall/various-pistachios-and-pistachio-containing-products-made-kunafa-s-recalled-due"),
        _lb(Date="2026-04-21", Company="Various", Pathogen="Salmonella", Source="CFIA",
            Country="Canada",
            URL="https://recalls-rappels.canada.ca/en/alert-recall/various-pistachios-and-pistachio-containing-products-recalled-due-salmonella-6"),
    ]
    assert all(p["category"] != "cfia_duplicate"
               for p in A.detect_lane_b(rows, resolved=set()))


def test_asset_integrity_flags_orphaned_marketing_pdf(tmp_path, monkeypatch):
    docs = tmp_path / "docs"
    data = docs / "data"
    (docs / "marketing").mkdir(parents=True)
    data.mkdir(parents=True)
    (docs / "marketing" / "2026-M04-marketing.pdf").write_bytes(b"%PDF-1.4\n")
    (data / "monthly-index.json").write_text(json.dumps([
        {"filename": "2026-M03.html", "pdf_url": "https://x/2026-M03-marketing.pdf"},
        {"filename": "2026-M04.html", "pdf_url": None},
    ]), encoding="utf-8")
    monkeypatch.setattr(A, "DOCS", docs)
    monkeypatch.setattr(A, "DATA", data)
    issues = A.build_asset_integrity()
    kinds = {(i["month"], i["kind"]) for i in issues}
    assert ("M04", "orphaned_marketing_pdf") in kinds
    # M03 is linked and has no PDF on disk -> broken_pdf_link, not orphaned
    assert ("M03", "broken_pdf_link") in kinds


def test_lane_b_ledger_suppresses_resolved_item():
    rows = [
        _lb(Date="2026-07-01", Company="Dup", Pathogen="Listeria",
            URL="https://fda.gov/recall/999", Notes="official-feed"),
        _lb(Date="2026-07-01", Company="Dup", Pathogen="Listeria",
            URL="https://www.fda.gov/recall/999/"),
    ]
    first = A.detect_lane_b(rows, resolved=set())
    assert first
    again = A.detect_lane_b(rows, resolved={first[0]["id"]})
    assert again == []


def test_item_id_is_stable_across_runs():
    r = _lb(Date="2026-07-01", Company="Dup", Pathogen="Listeria",
            URL="https://www.fda.gov/recall/999/")
    assert A._item_id("duplicate_url", r) == A._item_id("duplicate_url", r)


# ── In-progress week ──────────────────────────────────────────────────────
def test_week_friday_matches_isocalendar():
    assert A._week_friday("W29") == date.fromisocalendar(2026, 29, 5)
    assert A._week_friday("W28") == date.fromisocalendar(2026, 28, 5)
    assert A._week_friday("garbage") is None


# ── Dry-run is non-destructive ────────────────────────────────────────────
def _write_min_xlsx(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Recalls"
    ws.append(mm.RECALLS_SCHEMA)
    ws.append(["2026-07-10", "FDA", "ACME", "", "Cheese", "Listeria", "recall",
               "", "USA", "", 3, 0, "https://fda.gov/x", "", "2026-07-10", "", "", "W28"])
    wb.create_sheet("Pending")
    wb["Pending"].append(mm.PENDING_SCHEMA)
    wb.save(path)


def test_dry_run_touches_nothing(tmp_path, monkeypatch):
    xlsx = tmp_path / "recalls.xlsx"
    rjson = tmp_path / "recalls.json"
    _write_min_xlsx(xlsx)
    before = xlsx.read_bytes()

    monkeypatch.setattr(A, "XLSX", xlsx)
    monkeypatch.setattr(A, "RECALLS_JSON", rjson)
    monkeypatch.setattr(A, "DATA", tmp_path)
    monkeypatch.setattr(A, "DOCS", tmp_path)
    monkeypatch.setattr(A, "REPORT_JSON", tmp_path / "daily_review_report.json")
    monkeypatch.setattr(A, "DIGEST_MD", tmp_path / "daily_review_digest.md")
    monkeypatch.setattr(A, "LEDGER_JSON", tmp_path / "review_ledger.json")

    report = A.run(dry_run=True)

    assert xlsx.read_bytes() == before          # xlsx untouched
    assert not rjson.exists()                    # json never mirrored
    assert (tmp_path / "daily_review_report.json").exists()
    assert (tmp_path / "daily_review_digest.md").exists()
    assert report["mode"] == "dry-run"
    assert report["lane_a"]["applied"] is False
    # Lane A still DETECTS the mis-tier even though it doesn't apply it.
    assert report["counts"]["lane_a_changes"] >= 1
