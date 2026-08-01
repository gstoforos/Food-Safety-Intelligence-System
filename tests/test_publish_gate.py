"""Regression tests for the deterministic publish gate.

Every case here is a row that actually reached publication. See
pipeline/_publish_gate.py for the incident write-up.

Run:  python -m pytest tests/test_publish_gate.py -v
"""
from __future__ import annotations
import sys
import unittest
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from pipeline._publish_gate import (  # noqa: E402
    publish_blockers,
    is_publishable,
    VALID_REGIONS,
)


def _good(**overrides):
    """A row that must pass cleanly, so each test changes exactly one thing."""
    row = {
        "Date": "2026-07-24",
        "Source": "RappelConso (FR)",
        "Company": "CRUSTA C",
        "Brand": "Unbranded",
        "Product": "crevettes cuites sauvages 40/60",
        "Pathogen": "Listeria monocytogenes",
        "Reason": "Presence of Listeria monocytogenes",
        "Class": "Voluntary",
        "Country": "France",
        "Region": "Europe",
        "Tier": 1,
        "Outbreak": 0,
        "URL": "https://rappel.conso.gouv.fr/fiche-rappel/22963/Interne",
        "Notes": "",
    }
    row.update(overrides)
    return row


class TestBaselinePasses(unittest.TestCase):
    def test_a_good_row_has_no_blockers(self):
        self.assertEqual([], publish_blockers(_good()))
        self.assertTrue(is_publishable(_good()))


class TestTheNccIncident(unittest.TestCase):
    """The 2024 recall published as 2026-07-27 and emailed to subscribers."""

    ROW = {
        "Date": "2026-07-27",
        "Source": "NCC",
        "Company": "BM Foods (a member of Sea Harvest Group)",
        "Brand": "Shoprite Checkers",
        "Product": "Deli Hummus range",
        "Pathogen": "Listeria Monocytogenes",
        "Reason": "Recall ID 842632",
        "Class": None,
        "Country": "South Africa",
        "Region": "Not specified in the article",
        "Tier": 1,
        "Outbreak": 0,
        "URL": "https://thencc.org.za/media-statement-deli-hummus-range-product-safety-recall/",
        "Notes": "Discovered via news: timeslive.co.za",
    }

    def test_row_is_blocked(self):
        self.assertFalse(is_publishable(self.ROW))

    def test_every_deterministic_defect_is_named(self):
        blockers = " | ".join(publish_blockers(self.ROW)).lower()
        self.assertIn("reference number", blockers)   # Reason 'Recall ID 842632'
        self.assertIn("region", blockers)             # free prose in a controlled field
        self.assertIn("class is empty", blockers)     # Class None

    def test_it_needs_no_model_and_no_network(self):
        """The point of the gate: three of the six defects are visible from the
        row alone, with the Gemini quota at zero."""
        self.assertGreaterEqual(len(publish_blockers(self.ROW)), 3)


class TestNonFoodConsumerProducts(unittest.TestCase):
    """RappelConso covers all consumer goods. A pathogen database must not."""

    def test_car_blocked(self):
        row = _good(Product="mg3, mg3 hybrid+ voiture de tourisme.",
                    Pathogen=None, Reason="Dans certaines conditions de collision…",
                    URL="https://rappel.conso.gouv.fr/fiche-rappel/49461/Interne")
        self.assertIn("Pathogen is empty", " | ".join(publish_blockers(row)))

    def test_bath_toy_blocked(self):
        row = _good(Product="jouet de baignoire à pulvérisation",
                    Pathogen="", Company=None)
        self.assertFalse(is_publishable(row))

    def test_lamp_oil_and_bottle_blocked(self):
        for product in ("petrole lampant", "mento sport bottle"):
            self.assertFalse(is_publishable(_good(Product=product, Pathogen="")))


class TestIdOnlyReason(unittest.TestCase):
    def test_leaked_prompt_example_blocked(self):
        for reason in ("Recall ID 842632", "recall id 842632",
                       "Recall ID 08072026", "ID: 12345", "allerta 842632"):
            self.assertFalse(is_publishable(_good(Reason=reason)), reason)

    def test_id_appended_to_real_hazard_text_is_fine(self):
        """The corrected prompt puts the ID after the hazard — that must pass."""
        row = _good(Reason="Salmonella spp. in tahini used for production "
                           "[Recall ID 842632]")
        self.assertEqual([], publish_blockers(row))


class TestControlledVocabularies(unittest.TestCase):
    def test_every_valid_region_passes(self):
        for region in VALID_REGIONS:
            self.assertEqual([], publish_blockers(_good(Region=region)), region)

    def test_prose_region_blocked(self):
        for region in ("Not specified in the article", "unclear from the source",
                       "Not stated"):
            self.assertFalse(is_publishable(_good(Region=region)), region)

    def test_typo_region_blocked(self):
        self.assertFalse(is_publishable(_good(Region="EU")))


class TestUrlRule(unittest.TestCase):
    """Narrow by design — a gate that cries wolf gets switched off."""

    def test_landing_pages_blocked(self):
        for url in (
            "https://thencc.org.za/product-recalls/",
            "https://www.efet.gr/index.php/el/enimerosi/deltia-typou",
            "https://www.aesan.gob.es/AECOSAN/web/home/aecosan_inicio.htm",
            "https://rappel.conso.gouv.fr/",
        ):
            self.assertFalse(is_publishable(_good(URL=url)), url)

    def test_real_notices_with_trailing_slash_pass(self):
        """The first draft of this rule flagged 20+ of these. It must not."""
        for url in (
            "https://thencc.org.za/product-safety-recall-nutricia-aptamil-nutribiotik-2-and-nutricia-aptajunior-nutribiotik-3/",
            "https://thencc.org.za/product-recall-nan-special-pro-ha-infant-formula-800g/",
            "https://thencc.org.za/national-consumer-commission-refers-hummus-supplier-to-the-tribunal-over-listeria-contamination/",
            "https://www.cdc.gov/listeria/outbreaks/soft-cheese-06-26/index.html",
            "https://www.fda.gov.ph/fda-advisory-no-2026-0030-voluntary-recall-of-nan-optipro/",
        ):
            self.assertEqual([], publish_blockers(_good(URL=url)), url)

    def test_relative_url_blocked(self):
        self.assertFalse(is_publishable(_good(URL="/fiche-rappel/22963/Interne")))


class TestHonestDisclosureIsNotADefect(unittest.TestCase):
    """An enricher admitting it does not know beats it inventing a name.
    Blocking that would push the pipeline back toward fabrication."""

    def test_company_not_specified_passes(self):
        self.assertEqual([], publish_blockers(
            _good(Company="(not specified in FAVV notice)")))

    def test_rasff_unknown_origin_passes(self):
        self.assertEqual([], publish_blockers(
            _good(Company="Origin: unknown origin | Notifying: Ireland")))


class TestSourceAndCountryConsistency(unittest.TestCase):
    """USA rows must speak with one voice.

    The established conventions in this workbook are Source 'USDA FSIS' (never
    bare 'FSIS') and Country 'United States' (never 'USA'). Country is a join
    key — it drives Region, the country counts in the weekly and monthly
    reports, and the per-country filters on subscriber alert rules — so two
    spellings of one country silently split all of it. Before 2026-08-01 the
    workbook held 82 'United States' and 15 'USA', and the RASFF rows were
    self-contradictory: 'Origin: United States' in Company while Country said
    'USA'.
    """

    CANONICAL_US = "United States"
    BANNED_COUNTRY = frozenset({"usa", "us", "u.s.", "u.s.a.",
                                "united states of america"})
    BANNED_SOURCE = frozenset({"fsis", "usda", "usda-fsis", "usda fsis (us)"})

    def _rows(self):
        try:
            import openpyxl
        except ImportError:                       # pragma: no cover
            self.skipTest("openpyxl not installed")
        xlsx = ROOT / "docs" / "data" / "recalls.xlsx"
        if not xlsx.exists():                     # pragma: no cover
            self.skipTest("recalls.xlsx not present")
        wb = openpyxl.load_workbook(xlsx, read_only=True)
        out = []
        for sheet in wb.sheetnames:
            rows = list(wb[sheet].values)
            if not rows:
                continue
            hdr = [str(h) for h in rows[0]]
            if "Country" not in hdr:
                continue
            for r in rows[1:]:
                if r:
                    out.append((sheet, dict(zip(hdr, r))))
        return out

    def test_no_row_spells_the_country_usa(self):
        bad = [(s, str(r.get("Date")), str(r.get("Country")))
               for s, r in self._rows()
               if str(r.get("Country") or "").strip().lower() in self.BANNED_COUNTRY]
        self.assertEqual([], bad, f"non-canonical US country spelling: {bad[:8]}")

    def test_usda_fsis_source_label_is_canonical(self):
        bad = [(s, str(r.get("Date")), str(r.get("Source")))
               for s, r in self._rows()
               if str(r.get("Source") or "").strip().lower() in self.BANNED_SOURCE]
        self.assertEqual([], bad, f"non-canonical USDA FSIS source label: {bad[:8]}")

    def test_every_usda_fsis_row_is_united_states(self):
        bad = [(str(r.get("Date")), str(r.get("Country")))
               for s, r in self._rows()
               if str(r.get("Source") or "").strip() == "USDA FSIS"
               and str(r.get("Country") or "").strip() != self.CANONICAL_US]
        self.assertEqual([], bad, f"USDA FSIS row not tagged United States: {bad}")

    def test_writer_canonicalises_usa(self):
        """The guard must survive a writer round-trip, not just the one-off
        data repair."""
        from openpyxl import Workbook
        from pipeline.merge_master import _write_sheet, SCHEMA
        wb = Workbook()
        rows = [{"Date": "2026-07-30", "Source": "USDA FSIS", "Country": "USA",
                 "Company": "Establishment P-39928", "Product": "CURRY CHICKEN SALAD",
                 "Pathogen": "Listeria monocytogenes", "Reason": "Possible contamination",
                 "Class": "Public Health Alert", "Region": "North America",
                 "Tier": 1, "Outbreak": 0,
                 "URL": "https://www.fsis.usda.gov/recalls-alerts/x"}]
        _write_sheet(wb, "Recalls", SCHEMA, rows)
        col = SCHEMA.index("Country") + 1
        self.assertEqual("United States", wb["Recalls"].cell(2, col).value)


class TestOutbreakEvidence(unittest.TestCase):
    """An outbreak is a cluster of HUMAN ILLNESS.

    url_gate_gemini only required the model's outbreak-evidence quote to be
    non-empty. On RASFF rows the model satisfied that by quoting "risk:
    serious" — the hazard SEVERITY classification carried by every RASFF
    notification — and flipped Outbreak 0 -> 1. Five rows were published as
    outbreaks on that basis, one of them "Aflatoxin B1 in USA pistachio
    kernels", which cannot be an outbreak in any epidemiological sense. A
    sixth row's own Notes said Outbreak=0 while the field said 1.

    The scraper is right (scrapers/eu_wide/rasff.py sets Outbreak=0 because
    the consolidated feed exposes no illness counts) — the gate overrode it.
    """

    ILLNESS_WORDS = ("sick", "ill", "illness", "case", "hospital", "death",
                     "died", "fatal", "infect", "outbreak of", "people",
                     "patient", "notification", "cluster", "onset")

    def _is_illness_evidence(self, quote):
        return any(w in quote.lower() for w in self.ILLNESS_WORDS)

    @staticmethod
    def _strip_boilerplate(quote):
        import re as _re
        q = _re.sub(r"(may|can|could|might)\s+(also\s+|even\s+|still\s+)?(cause|lead to|result in|become|make)"
                    r"[^.;]{0,140}", " ", quote, flags=_re.I)
        return _re.sub(r"(no|without)\s+(reported\s+|confirmed\s+)?"
                       r"(illness|case|infection)\w*", " ", q, flags=_re.I)

    def test_regulator_risk_boilerplate_is_not_outbreak_evidence(self):
        """Every Listeria notice carries "may cause severe illness in pregnant
        women...". The FSANZ Lux Ham row was flagged on exactly that paragraph
        while its own notice states no confirmed illnesses."""
        for quote in (
            "Listeria monocytogenes may cause severe illness in pregnant women, "
            "unborn babies, neonates, the elderly and immunocompromised individuals",
            "The general population can also become ill from consuming food "
            "contaminated with Listeria monocytogenes",
            "There have been no reported illnesses associated with the "
            "consumption of these products",
        ):
            self.assertFalse(
                self._is_illness_evidence(self._strip_boilerplate(quote)), quote)

    def test_boilerplate_stripping_keeps_real_reports(self):
        for quote in ("20 people have been made sick and four hospitalized",
                      "3 sick people in Illinois, Lm cluster",
                      "48 cases of S. Bovismorbificans across UK and Ireland"):
            self.assertTrue(
                self._is_illness_evidence(self._strip_boilerplate(quote)), quote)

    def test_boilerplate_guard_is_present_in_the_gate(self):
        from pipeline import url_gate_gemini as ug
        src = Path(ug.__file__).read_text(encoding="utf-8")
        self.assertIn("Regulator risk boilerplate", src)

    def test_rasff_risk_severity_is_not_outbreak_evidence(self):
        for quote in ("risk: serious", "risk: potentially serious",
                      "risk: not serious"):
            self.assertFalse(self._is_illness_evidence(quote), quote)

    def test_real_outbreak_quotes_are_accepted(self):
        for quote in (
            "The Canadian Food Inspection Agency says 20 people have been made "
            "sick and four hospitalized after an outbreak of E. coli",
            "9 illnesses confirmed in California, onset dates 2026-03-27 to ...",
            "48 cases of S. Bovismorbificans across UK + Ireland + Finland",
            "3 sick people in Illinois, Lm cluster",
            "We received 36 clinical notifications of children developing ...",
        ):
            self.assertTrue(self._is_illness_evidence(quote), quote)

    def test_guard_is_present_in_the_gate(self):
        from pipeline import url_gate_gemini as ug
        src = Path(ug.__file__).read_text(encoding="utf-8")
        self.assertIn("_ILLNESS_WORDS", src)
        self.assertIn("DISCARDED outbreak verdict", src)

    def test_no_published_outbreak_rests_on_risk_severity(self):
        try:
            import openpyxl
        except ImportError:                       # pragma: no cover
            self.skipTest("openpyxl not installed")
        xlsx = ROOT / "docs" / "data" / "recalls.xlsx"
        if not xlsx.exists():                     # pragma: no cover
            self.skipTest("recalls.xlsx not present")
        import re as _re
        wb = openpyxl.load_workbook(xlsx, read_only=True)
        rows = list(wb["Recalls"].values)
        hdr = [str(h) for h in rows[0]]
        bad = []
        for r in rows[1:]:
            if not r:
                continue
            row = dict(zip(hdr, r))
            if str(row.get("Outbreak")) not in ("1", "1.0"):
                continue
            notes = str(row.get("Notes") or "")
            m = _re.search(r"\[?outbreak[^\]]{0,300}", notes, _re.I)
            ev = m.group(0) if m else ""
            if _re.search(r"0\s*→\s*1\s*—\s*risk:\s*(potentially\s+)?serious\s*$",
                          ev.strip(), _re.I):
                bad.append((str(row.get("Date")), str(row.get("Pathogen"))))
        self.assertEqual([], bad,
                         f"outbreak flags resting on RASFF risk severity: {bad}")

    def test_no_row_contradicts_its_own_outbreak_note(self):
        try:
            import openpyxl
        except ImportError:                       # pragma: no cover
            self.skipTest("openpyxl not installed")
        xlsx = ROOT / "docs" / "data" / "recalls.xlsx"
        if not xlsx.exists():                     # pragma: no cover
            self.skipTest("recalls.xlsx not present")
        import re as _re
        wb = openpyxl.load_workbook(xlsx, read_only=True)
        rows = list(wb["Recalls"].values)
        hdr = [str(h) for h in rows[0]]
        bad = []
        for r in rows[1:]:
            if not r:
                continue
            row = dict(zip(hdr, r))
            notes = str(row.get("Notes") or "")
            if (str(row.get("Outbreak")) in ("1", "1.0")
                    and _re.search(r"Outbreak\s*=\s*0\b", notes)
                    and "outbreak-fix" not in notes):
                bad.append(str(row.get("Date")))
        self.assertEqual([], bad, f"Outbreak=1 while Notes say Outbreak=0: {bad}")


class TestWorkbookStaysClean(unittest.TestCase):
    """The six unpublishable rows were archived to Weekly_Rejected on
    2026-08-01. Thirteen remain: all are GENUINE food recalls whose fields are
    incomplete, not out-of-scope rows, so each needs its source read rather
    than a bulk edit. They are pinned here by URL so no NEW violation can slip
    in unnoticed. This list must only ever shrink.
    """

    KNOWN_REMAINING = frozenset({
        # ONE row remains, and it is honest rather than broken:
        # EFET recalled a German "High Protein" chocolate pudding sold at Lidl
        # on 2026-06-26 (Greek press: pronews, LiFO, protothema, Documento).
        # The recall is real and the fields are right; only the URL is the EFET
        # press-release LISTING page, because no per-notice efet.gr permalink
        # is discoverable for it. EFET permalinks run
        # /item/<id>-deltio-typou-..., and the neighbouring IDs are 5401
        # (06-19) and 5404 (07-08) — so the missing one is almost certainly
        # 5402 or 5403. GUESSING that id is exactly the fabrication this whole
        # gate exists to stop, so the listing URL stays until the real
        # permalink is confirmed.
        "efet.gr/index.php/el/enimerosi/deltia-typou",
        # ── Revised 2026-08-02 (second pass) ──────────────────────────────
        # The ten RappelConso rows previously pinned here have been REPAIRED
        # from the official DGCCRF open-data record and are gone from this
        # list. rappel.conso.gouv.fr is unreachable from the audit environment
        # (robots.txt fetch fails TLS verification), but the DGCCRF publishes
        # every fiche as open data and the fiche id in the URL is the primary
        # key, so the join is exact and needs no model:
        #
        #   https://tabular-api.data.gouv.fr/api/resources/
        #       5a4e7174-657c-4920-af1f-3440a996837c/data/?id__exact=<fiche>
        #
        # See pipeline/verify_rappelconso.py, which now reconciles the whole
        # French corpus against that record on a schedule.
        #
        # ONE remains. Fiche 22205 is ABSENT from the export — the ids jump
        # 22204 -> 22206 — so there is no authoritative record to repair it
        # from, and guessing is the failure this gate exists to prevent.
        "fiche-rappel/22205",   # Listeria mc <- aflatoxin / corn wafers (SK)
    })

    def test_no_NEW_row_violates_the_gate(self):
        try:
            import openpyxl
        except ImportError:                       # pragma: no cover
            self.skipTest("openpyxl not installed")
        xlsx = ROOT / "docs" / "data" / "recalls.xlsx"
        if not xlsx.exists():                     # pragma: no cover
            self.skipTest("recalls.xlsx not present")
        wb = openpyxl.load_workbook(xlsx, read_only=True)
        rows = list(wb["Recalls"].values)
        hdr = [str(h) for h in rows[0]]
        offenders = []
        for r in rows[1:]:
            if not r or not r[hdr.index("URL")]:
                continue
            row = dict(zip(hdr, r))
            blockers = publish_blockers(row)
            if not blockers:
                continue
            url = str(row.get("URL") or "")
            if any(known in url for known in self.KNOWN_REMAINING):
                continue
            offenders.append((str(row.get("Date")), str(row.get("Product"))[:40],
                              blockers[0][:70]))
        self.assertEqual(
            [], offenders,
            f"{len(offenders)} NEW publish-gate violation(s) reached Recalls: "
            f"{offenders[:8]}")


if __name__ == "__main__":
    unittest.main(verbosity=2)


class TestOutbreakRankingUsesBurden(unittest.TestCase):
    """Inside the outbreak phase, rank by published illness burden.

    The ranker used pathogen-INTRINSIC severity then date, and never looked at
    how large the outbreak actually was. In July 2026 that put the Cyclospora
    iceberg-lettuce outbreak (1,644 laboratory-confirmed cases, 94
    hospitalisations, 9 states) BELOW the Lamia Salmonella cluster of about 20
    people, purely because Cyclospora scores 99 on the intrinsic table while
    Salmonella scores 4. For confirmed outbreaks the genus is a poor proxy for
    public-health weight; the case count is the direct measure.
    """

    def _ranker(self):
        import importlib.util
        spec = importlib.util.spec_from_file_location(
            "_wb", ROOT / "docs" / "build_weekly_report_afts.py")
        mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
        return mod.rank_top_recalls

    JULY = [
        {"Date": "2026-07-22", "Pathogen": "Salmonella Enteritidis", "Tier": 1,
         "Outbreak": 1, "Country": "United States", "Company": "Midwest Poultry",
         "Reason": "Multistate outbreak: 98 illnesses across 17 states, "
                   "26 hospitalizations, 0 deaths."},
        {"Date": "2026-07-20", "Pathogen": "Salmonella", "Tier": 1,
         "Outbreak": 1, "Country": "Greece", "Company": "Multiple chicken suppliers",
         "Reason": "Lamia investigation. Approximately 20 people sickened."},
        {"Date": "2026-07-18", "Pathogen": "Cyclospora", "Tier": 1,
         "Outbreak": 1, "Country": "United States", "Company": "Taylor Fresh Foods",
         "Reason": "CDC counted 1,644 laboratory-confirmed cases and 94 "
                   "hospitalisations across 5 states."},
        {"Date": "2026-07-30", "Pathogen": "Listeria monocytogenes", "Tier": 1,
         "Outbreak": 0, "Country": "France", "Company": "La Ferme des Tertres",
         "Reason": "Listeria monocytogenes."},
    ]

    def test_largest_outbreak_ranks_first(self):
        ranked = self._ranker()(self.JULY, n=5)
        self.assertEqual("Cyclospora", ranked[0]["Pathogen"],
                         "the 1,644-case outbreak must outrank the 98-case and "
                         "20-case ones")
        self.assertEqual("Salmonella Enteritidis", ranked[1]["Pathogen"])
        self.assertEqual("Salmonella", ranked[2]["Pathogen"])

    def test_outbreaks_still_outrank_non_outbreaks(self):
        """Burden must not let a big outbreak be beaten by a severe pathogen,
        nor let a non-outbreak jump the phase."""
        ranked = self._ranker()(self.JULY, n=5)
        self.assertEqual(0, ranked[-1]["Outbreak"],
                         "a non-outbreak Listeria row must stay below every "
                         "confirmed outbreak")

    def test_unknown_burden_falls_back_to_intrinsic_severity(self):
        rows = [
            {"Date": "2026-07-10", "Pathogen": "Salmonella", "Tier": 1,
             "Outbreak": 1, "Country": "X", "Reason": "outbreak, no counts given"},
            {"Date": "2026-07-10", "Pathogen": "Listeria monocytogenes", "Tier": 1,
             "Outbreak": 1, "Country": "X", "Reason": "outbreak, no counts given"},
        ]
        ranked = self._ranker()(rows, n=2)
        self.assertEqual("Listeria monocytogenes", ranked[0]["Pathogen"],
                         "with no published counts, intrinsic severity still "
                         "decides")
