"""
tests/conftest.py — shared pytest fixtures for the AFTS FSIS test suite.

This module exposes fixtures that every test file can use without importing.
The pytest framework auto-discovers conftest.py in the tests directory.

Fixtures provided:
    tmp_xlsx_factory          → callable that builds a synthetic recalls.xlsx
                                with arbitrary sheets and rows in a tmp_path
    sample_recall_row         → a representative single Recalls dict
    sample_pending_row        → a representative single Pending dict
    sample_weekly_review_row  → a representative single Weekly_Review dict
    sample_weekly_rejected_row→ a representative single Weekly_Rejected dict
    sample_news_row           → a representative single NEWS row

Design notes
============
We use synthetic xlsx files rather than fixtures of the real recalls.xlsx
because:

  1. Real data has 588 rows and grows over time — golden-file tests against
     a moving target produce noise.
  2. Bugs we're hunting are about SHEET-LEVEL invariants (Weekly_Rejected
     surviving a merge), not row-content correctness. A 3-row synthetic
     xlsx exposes the same bugs as a 588-row one.
  3. CI runs faster on tiny fixtures.

Each fixture returns plain dicts that can be passed verbatim to the
pipeline functions under test (which all consume `Dict[str, Any]` rows).
"""
from __future__ import annotations

import sys
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional

import pytest
from openpyxl import Workbook


# Repository root — let `import pipeline.xlsx_merge` resolve when pytest
# is run from the repo root (which is the conventional invocation).
REPO_ROOT = Path(__file__).resolve().parent.parent
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))


# ---------------------------------------------------------------------------
# Sheet schemas — kept in sync with pipeline/merge_master.RECALLS_SCHEMA
# and pipeline/weekly_review_capture.SHEET_COLS at the time of writing
# (audit 2026-05-13). If those modules add columns, mirror them here.
# ---------------------------------------------------------------------------
RECALLS_COLS = [
    "Date", "Source", "Company", "Brand", "Product", "Pathogen", "Reason",
    "Class", "Country", "Region", "Tier", "Outbreak", "URL", "Notes",
    "DateAdded", "LastUpdated", "LastChecked",
]
PENDING_COLS = RECALLS_COLS + ["ScrapedAt", "Status", "RejectedBy"]
WEEKLY_REVIEW_COLS = RECALLS_COLS + ["Week_Added", "Reviewed"]
WEEKLY_REJECTED_COLS = RECALLS_COLS + [
    "Week_Added", "RejectedBy", "RejectionReason", "Reviewed",
]
NEWS_COLS = [
    "Published (UTC)", "Pathogen", "Event", "Source",
    "Title", "Link", "Retrieved (UTC)",
]


# ---------------------------------------------------------------------------
# Row factories
# ---------------------------------------------------------------------------
@pytest.fixture
def sample_recall_row() -> Dict[str, Any]:
    """A representative verified Recalls row. Override fields per-test."""
    return {
        "Date": "2026-05-10",
        "Source": "USDA FSIS",
        "Company": "Acme Foods Inc.",
        "Brand": "Acme",
        "Product": "Frozen chicken breast 2lb",
        "Pathogen": "Listeria monocytogenes",
        "Reason": "Positive sample",
        "Class": "Class I",
        "Country": "United States",
        "Region": "NorthAmerica",
        "Tier": 1,
        "Outbreak": 0,
        "URL": "https://www.fsis.usda.gov/recalls/acme-2026-W19",
        "Notes": "",
        "DateAdded": "2026-05-10T08:00:00Z",
        "LastUpdated": "2026-05-10T08:00:00Z",
        "LastChecked": "2026-05-10T08:00:00Z",
    }


@pytest.fixture
def sample_pending_row(sample_recall_row) -> Dict[str, Any]:
    """A Pending row — Recalls cols + ScrapedAt + Status + RejectedBy."""
    return {
        **sample_recall_row,
        "URL": "https://www.fsis.usda.gov/recalls/pending-acme-X",
        "ScrapedAt": "2026-05-10T04:00:00Z",
        "Status": "pending",
        "RejectedBy": "",
    }


@pytest.fixture
def sample_weekly_review_row(sample_recall_row) -> Dict[str, Any]:
    """A Weekly_Review row — Recalls cols + Week_Added + Reviewed."""
    return {
        **sample_recall_row,
        "Week_Added": "2026-05-14",   # the Thursday this row is reviewed on
        "Reviewed": "N",
    }


@pytest.fixture
def sample_weekly_rejected_row(sample_recall_row) -> Dict[str, Any]:
    """A Weekly_Rejected row — Recalls cols + Week_Added + RejectedBy +
    RejectionReason + Reviewed."""
    return {
        **sample_recall_row,
        "Week_Added": "2026-05-14",
        "RejectedBy": "url_gate_gemini, claude-check",
        "RejectionReason": "URL dead per both reviewers",
        "Reviewed": "N",
    }


@pytest.fixture
def sample_news_row() -> Dict[str, Any]:
    return {
        "Published (UTC)": "2026-05-10T12:00:00Z",
        "Pathogen": "Listeria",
        "Event": "outbreak",
        "Source": "Food Safety News",
        "Title": "Sample news headline",
        "Link": "https://foodsafetynews.example.com/2026/05/sample",
        "Retrieved (UTC)": "2026-05-10T13:00:00Z",
    }


# ---------------------------------------------------------------------------
# xlsx factory — builds a synthetic recalls.xlsx with named sheets
# ---------------------------------------------------------------------------
@pytest.fixture
def tmp_xlsx_factory(tmp_path) -> Callable[..., Path]:
    """
    Returns a callable: tmp_xlsx_factory(name, sheets) → Path

    `name`   = filename to save into tmp_path (e.g. "remote.xlsx")
    `sheets` = dict mapping sheet_name → (headers, list_of_row_dicts)

    Sheets are written in the order they appear in the dict, matching the
    canonical ordering merge_master uses (Recalls first, NEWS last).

    Returns the absolute path to the written xlsx file. The file lives
    in pytest's tmp_path and is auto-cleaned between tests.

    Example:
        path = tmp_xlsx_factory("mine.xlsx", {
            "Recalls": (RECALLS_COLS, [recall_row1, recall_row2]),
            "Pending": (PENDING_COLS, []),
        })
    """
    def _factory(name: str,
                 sheets: Dict[str, tuple]) -> Path:
        wb = Workbook()
        # Workbook always starts with one sheet; we rename it for the first.
        first_name = next(iter(sheets.keys()))
        first_headers, first_rows = sheets[first_name]
        ws = wb.active
        ws.title = first_name
        _write(ws, first_headers, first_rows)

        # Remaining sheets in dict order.
        first = True
        for sheet_name, (headers, rows) in sheets.items():
            if first:
                first = False
                continue
            ws = wb.create_sheet(sheet_name)
            _write(ws, headers, rows)

        out = tmp_path / name
        wb.save(out)
        return out
    return _factory


def _write(ws, headers: List[str], rows: List[Dict[str, Any]]):
    """Write headers in row 1 and dict rows in subsequent rows."""
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)
    for r_idx, row_dict in enumerate(rows, 2):
        for col_idx, h in enumerate(headers, 1):
            v = row_dict.get(h, "")
            ws.cell(row=r_idx, column=col_idx, value=v)
