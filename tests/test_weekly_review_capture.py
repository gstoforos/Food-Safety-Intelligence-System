"""
tests/test_weekly_review_capture.py
=====================================
Tests for pipeline.weekly_review_capture.

Two functions under test:

    review_thursday_for(now_utc) → date
        Returns the Thursday date a row promoted at `now_utc` belongs to.
        Cutoff: Thursday 17:00 Athens local time.
          - Promoted at Wed 23:59 → THIS Thursday
          - Promoted at Thu 16:59 → THIS Thursday (still in window)
          - Promoted at Thu 17:00 → NEXT Thursday (rollover, strict ≥)
          - Promoted at Fri 09:00 → NEXT Thursday
          - Promoted at Sun 12:00 → NEXT Thursday

    record_promotions(promoted_rows, xlsx_path, json_path) → int
        Appends each row to the Weekly_Review sheet, tagged with the
        appropriate Thursday review date. Returns count of newly-appended
        rows. Idempotent: calling twice with the same rows is a no-op.

The Thursday-17:00 cutoff matters because that's when the operator's
Thursday review email fires (sendThursdayManualReview). A row promoted
at 17:01 must NOT appear in that day's email — it goes to next week.
If the cutoff drifts (e.g. someone changes REVIEW_HOUR_LOCAL), the
Thursday email shows phantom rows that the operator didn't approve yet.

The idempotency invariant matters because weekly_review_capture is
called from inside merge_master.append_to_recalls — which runs hourly
via merge-master.yml. If the function weren't idempotent, every hourly
run would duplicate every row in Weekly_Review.

Uses freezegun to freeze time at specific UTC instants and observe the
function's response. The Athens timezone math is non-obvious (DST), so
the tests use UTC inputs and compare against specific expected Athens-
local interpretations.
"""
from __future__ import annotations

import json
from datetime import date, datetime, timezone
from pathlib import Path

import pytest
from freezegun import freeze_time
from openpyxl import load_workbook

from pipeline.weekly_review_capture import (
    review_thursday_for,
    record_promotions,
    SHEET_NAME,
    SHEET_COLS,
)
from tests.conftest import RECALLS_COLS, PENDING_COLS, NEWS_COLS


# ───────────────────────────────────────────────────────────────────────
# review_thursday_for() — cutoff math
# ───────────────────────────────────────────────────────────────────────
class TestReviewThursdayFor:
    """
    Cutoff rule: Thursday 17:00 Athens local. Time arithmetic uses
    Europe/Athens which switches between EEST (UTC+3, summer) and EET
    (UTC+2, winter). Test instants are in UTC; expected output is the
    Athens-local Thursday they should resolve to.

    Reference dates (May 2026 = EEST = UTC+3):
        Wed May 13, 2026  UTC 06:00 → Athens 09:00 Wed
        Thu May 14, 2026  UTC 13:59 → Athens 16:59 Thu  (still in window)
        Thu May 14, 2026  UTC 14:00 → Athens 17:00 Thu  (rollover)
        Thu May 14, 2026  UTC 14:30 → Athens 17:30 Thu  (rollover)
        Fri May 15, 2026  UTC 09:00 → Athens 12:00 Fri  (next week)
    """

    def test_wednesday_promotion_lands_in_this_thursday(self):
        # Wed May 13, 2026 09:00 Athens (UTC 06:00). Next Thursday is May 14.
        now = datetime(2026, 5, 13, 6, 0, tzinfo=timezone.utc)
        assert review_thursday_for(now) == date(2026, 5, 14)

    def test_thursday_before_1700_lands_in_today(self):
        # Thu May 14, 2026 16:59 Athens (UTC 13:59) — JUST inside the window.
        now = datetime(2026, 5, 14, 13, 59, tzinfo=timezone.utc)
        assert review_thursday_for(now) == date(2026, 5, 14)

    def test_thursday_at_1700_rolls_to_next_week(self):
        # Thu May 14, 2026 17:00 Athens (UTC 14:00) — rollover boundary,
        # STRICT ≥ (i.e. 17:00:00 already counts as past cutoff).
        now = datetime(2026, 5, 14, 14, 0, tzinfo=timezone.utc)
        assert review_thursday_for(now) == date(2026, 5, 21)

    def test_thursday_after_1700_rolls_to_next_week(self):
        # Thu May 14, 2026 17:30 Athens (UTC 14:30)
        now = datetime(2026, 5, 14, 14, 30, tzinfo=timezone.utc)
        assert review_thursday_for(now) == date(2026, 5, 21)

    def test_friday_lands_in_next_thursday(self):
        # Fri May 15, 2026 09:00 Athens (UTC 06:00)
        now = datetime(2026, 5, 15, 6, 0, tzinfo=timezone.utc)
        assert review_thursday_for(now) == date(2026, 5, 21)

    def test_saturday_lands_in_next_thursday(self):
        # Sat May 16, 2026 — next Thursday is May 21.
        now = datetime(2026, 5, 16, 12, 0, tzinfo=timezone.utc)
        assert review_thursday_for(now) == date(2026, 5, 21)

    def test_sunday_lands_in_next_thursday(self):
        # Sun May 17, 2026 — next Thursday is May 21.
        now = datetime(2026, 5, 17, 12, 0, tzinfo=timezone.utc)
        assert review_thursday_for(now) == date(2026, 5, 21)

    def test_monday_lands_in_this_thursday(self):
        # Mon May 18, 2026 — next Thursday is May 21.
        now = datetime(2026, 5, 18, 9, 0, tzinfo=timezone.utc)
        assert review_thursday_for(now) == date(2026, 5, 21)

    def test_winter_dst_cutoff_eet(self):
        """
        Winter (EET = UTC+2): Thu Feb 5, 2026.
          UTC 14:59 → Athens 16:59 → still in window
          UTC 15:00 → Athens 17:00 → rollover

        Tests that the cutoff respects DST — using EET offsets, not EEST.
        """
        # In window
        now = datetime(2026, 2, 5, 14, 59, tzinfo=timezone.utc)
        assert review_thursday_for(now) == date(2026, 2, 5)
        # Rolled over
        now = datetime(2026, 2, 5, 15, 0, tzinfo=timezone.utc)
        assert review_thursday_for(now) == date(2026, 2, 12)

    def test_default_now_utc_returns_some_thursday(self):
        """Calling with no argument uses current UTC. The result must
        be a Thursday in the future (or today if before 17:00 Athens
        and it's Thursday)."""
        result = review_thursday_for()
        # 3 = Thursday in Python's weekday() (Mon=0..Sun=6)
        assert result.weekday() == 3


# ───────────────────────────────────────────────────────────────────────
# record_promotions() — Weekly_Review append + idempotency
# ───────────────────────────────────────────────────────────────────────
pytestmark_slow = pytest.mark.slow


@pytestmark_slow
class TestRecordPromotions:
    """End-to-end: build a minimal xlsx, call record_promotions, inspect
    the resulting Weekly_Review sheet."""

    def test_creates_weekly_review_sheet_when_missing(
        self, tmp_xlsx_factory, tmp_path, sample_recall_row
    ):
        """First call against an xlsx WITHOUT a Weekly_Review sheet
        creates the sheet and appends rows."""
        xlsx = tmp_xlsx_factory("recalls.xlsx", {
            "Recalls": (RECALLS_COLS, []),
            "Pending": (PENDING_COLS, []),
            "NEWS":    (NEWS_COLS, []),
        })
        json_out = tmp_path / "weekly-review-latest.json"

        appended = record_promotions(
            [sample_recall_row], xlsx_path=xlsx, json_path=json_out)
        assert appended == 1

        wb = load_workbook(xlsx)
        assert SHEET_NAME in wb.sheetnames
        ws = wb[SHEET_NAME]
        # 1 header row + 1 data row
        assert ws.max_row == 2

    def test_idempotent_same_row_twice(
        self, tmp_xlsx_factory, tmp_path, sample_recall_row
    ):
        """Calling twice with the same row appends ONCE, not twice."""
        xlsx = tmp_xlsx_factory("recalls.xlsx", {
            "Recalls": (RECALLS_COLS, []),
            "Pending": (PENDING_COLS, []),
            "NEWS":    (NEWS_COLS, []),
        })
        json_out = tmp_path / "wr.json"

        first  = record_promotions([sample_recall_row], xlsx_path=xlsx,
                                   json_path=json_out)
        second = record_promotions([sample_recall_row], xlsx_path=xlsx,
                                   json_path=json_out)

        assert first == 1
        assert second == 0, "Idempotency broken: same row appended twice"

        wb = load_workbook(xlsx)
        ws = wb[SHEET_NAME]
        assert ws.max_row == 2  # header + 1 data row, NOT 3

    def test_week_added_column_set_correctly(
        self, tmp_xlsx_factory, tmp_path, sample_recall_row
    ):
        """The Week_Added column must be set to the Thursday review date,
        matching review_thursday_for(). Lock by freezing time."""
        xlsx = tmp_xlsx_factory("recalls.xlsx", {
            "Recalls": (RECALLS_COLS, []),
            "Pending": (PENDING_COLS, []),
            "NEWS":    (NEWS_COLS, []),
        })
        json_out = tmp_path / "wr.json"

        # Freeze at Wed May 13, 2026 06:00 UTC (= 09:00 Athens Wed).
        # Next Thursday is May 14, 2026.
        with freeze_time("2026-05-13 06:00:00"):
            record_promotions([sample_recall_row], xlsx_path=xlsx,
                              json_path=json_out)

        wb = load_workbook(xlsx)
        ws = wb[SHEET_NAME]
        headers = [c.value for c in ws[1]]
        week_added_idx = headers.index("Week_Added")
        appended_value = ws.cell(row=2, column=week_added_idx + 1).value
        assert appended_value == "2026-05-14", \
            f"Expected Week_Added=2026-05-14, got {appended_value}"

    def test_reviewed_column_starts_as_N(
        self, tmp_xlsx_factory, tmp_path, sample_recall_row
    ):
        """Every new row enters with Reviewed='N'. Manual stamp by
        operator later flips it to 'Y'."""
        xlsx = tmp_xlsx_factory("recalls.xlsx", {
            "Recalls": (RECALLS_COLS, []),
            "Pending": (PENDING_COLS, []),
            "NEWS":    (NEWS_COLS, []),
        })
        json_out = tmp_path / "wr.json"
        record_promotions([sample_recall_row], xlsx_path=xlsx,
                          json_path=json_out)

        wb = load_workbook(xlsx)
        ws = wb[SHEET_NAME]
        headers = [c.value for c in ws[1]]
        reviewed_idx = headers.index("Reviewed")
        assert ws.cell(row=2, column=reviewed_idx + 1).value == "N"

    def test_url_dedup_collapses_collisions(
        self, tmp_xlsx_factory, tmp_path, sample_recall_row
    ):
        """If the same URL+Date appears twice in one batch, it appears
        ONCE in the sheet (matches Weekly_Review dedup_key rule)."""
        xlsx = tmp_xlsx_factory("recalls.xlsx", {
            "Recalls": (RECALLS_COLS, []),
            "Pending": (PENDING_COLS, []),
            "NEWS":    (NEWS_COLS, []),
        })
        json_out = tmp_path / "wr.json"
        # Same URL, different Company — but URL+Date is the dedup key.
        row_a = dict(sample_recall_row); row_a["Company"] = "First"
        row_b = dict(sample_recall_row); row_b["Company"] = "Second"
        record_promotions([row_a, row_b], xlsx_path=xlsx, json_path=json_out)

        wb = load_workbook(xlsx)
        ws = wb[SHEET_NAME]
        assert ws.max_row == 2, "URL+Date collision should collapse to 1 row"

    def test_json_sidecar_created(
        self, tmp_xlsx_factory, tmp_path, sample_recall_row
    ):
        """record_promotions also refreshes the JSON sidecar that the
        Apps Script Thursday mailer reads."""
        xlsx = tmp_xlsx_factory("recalls.xlsx", {
            "Recalls": (RECALLS_COLS, []),
            "Pending": (PENDING_COLS, []),
            "NEWS":    (NEWS_COLS, []),
        })
        json_out = tmp_path / "wr.json"
        record_promotions([sample_recall_row], xlsx_path=xlsx,
                          json_path=json_out)

        assert json_out.exists(), "JSON sidecar not written"
        data = json.loads(json_out.read_text())
        # The shape can vary; lock the minimum invariants.
        assert isinstance(data, dict), "JSON sidecar must be a dict"

    def test_empty_promotions_list_noop(
        self, tmp_xlsx_factory, tmp_path
    ):
        """Empty list → no rows appended, but JSON sidecar still refreshed
        (defensive: stale JSON would mislead the mailer)."""
        xlsx = tmp_xlsx_factory("recalls.xlsx", {
            "Recalls": (RECALLS_COLS, []),
            "Pending": (PENDING_COLS, []),
            "NEWS":    (NEWS_COLS, []),
        })
        json_out = tmp_path / "wr.json"
        appended = record_promotions([], xlsx_path=xlsx, json_path=json_out)
        assert appended == 0

    def test_missing_xlsx_returns_zero(self, tmp_path, sample_recall_row):
        """Missing xlsx file → return 0, do NOT raise."""
        missing = tmp_path / "does-not-exist.xlsx"
        json_out = tmp_path / "wr.json"
        result = record_promotions([sample_recall_row],
                                    xlsx_path=missing,
                                    json_path=json_out)
        assert result == 0
