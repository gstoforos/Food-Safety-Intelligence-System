"""Regression tests for the 2026-08-01 Auxico subscriber-alert incident.

WHAT HAPPENED
=============
The 13:25 AFTS Alert email of 2026-08-01 carried this row:

    2026-07-29 | UPDATED 30.07.26 | Auxico (Perth) Pty Ltd
                 LGM HOT CHILLI OIL 275G
                 Listeria monocytogenes | Australia | TIER-1

Verified against the live FSANZ notice (both slugs still resolve):

    https://www.foodstandards.gov.au/food-recalls/recall-alert/
        auxico-perth-pty-ltd-lgm-hot-chilli-oil-275g
    https://www.foodstandards.gov.au/food-recalls/recall-alert/
        updated-300726-auxico-perth-pty-ltd-lgm-hot-chilli-oil-275g

  • Published 29 July 2026, amended 30 July 2026.
  • Problem: "The recall is due to the presence of an undeclared allergen
    (peanuts)."  Hazard: peanut allergen.
  • The word "Listeria" does not appear anywhere on either page — not in the
    body, not in navigation, not in a related-recalls list.
  • Company is "Auxico (Perth) Pty Ltd". "UPDATED 30.07.26 | " is the page's
    own status banner, prepended to the <h1>.
  • The correct row for this exact recall was ALREADY in Recalls, at Tier 2,
    with Pathogen "Undeclared allergen (peanuts)".

FOUR INDEPENDENT DEFECTS, each pinned by a class below:

  1. Dedup — FSANZ republishes an amended alert at a NEW slug and keeps both
     live, so the URL key minted a second row for one recall.
  2. Scrape — "Company - Product" title split folded the status banner into
     Company.
  3. Gate — the Pathogen/Reason hazard-class cross-check existed, and worked,
     but lived inside claude_check's clean-row shortcut, so it guarded exactly
     one of several promotion paths. A weaker reviewer bypassed it.
  4. Rejection stickiness — claude_check archived the row to Weekly_Rejected
     for "pathogen mismatch" at 21:10 on 07-31. Nothing consulted that sheet
     before promoting, so the Qwen review agent re-approved the same URL at
     14:16 the next day.

Run:  python -m pytest tests/test_auxico_regression.py -v
"""
from __future__ import annotations
import sys
import unittest
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from pipeline._publish_gate import (  # noqa: E402
    publish_blockers,
    is_publishable,
    classify_hazard,
    pathogen_reason_class_mismatch,
    strip_title_status_prefix,
)
from pipeline.merge_master import (  # noqa: E402
    _normalize_url_for_dedup,
    _dedup_key,
    promote_approved,
    load_rejected_urls,
    _write_sheet,
    SCHEMA,
)

FSANZ = "https://www.foodstandards.gov.au/food-recalls/recall-alert/"
ORIGINAL = FSANZ + "auxico-perth-pty-ltd-lgm-hot-chilli-oil-275g"
AMENDED = FSANZ + "updated-300726-auxico-perth-pty-ltd-lgm-hot-chilli-oil-275g"

# The row exactly as it was published and emailed.
BAD_ROW = {
    "Date": "2026-07-29",
    "Source": "FSANZ (AU)",
    "Company": "UPDATED 30.07.26 | Auxico (Perth) Pty Ltd",
    "Brand": "—",
    "Product": "LGM HOT CHILLI OIL 275G",
    "Pathogen": "Listeria monocytogenes",
    "Reason": "The recall is due to the presence of an undeclared allergen "
              "(peanuts).",
    "Class": "Recall",
    "Country": "Australia",
    "Region": "Oceania",
    "Tier": 1,
    "Outbreak": 0,
    "URL": AMENDED,
    "Notes": "",
}

# The row as the source actually supports it. NOTE: this row is CORRECT and
# still unpublishable — the hazard is an undeclared peanut allergen, and
# allergen-only recalls are out of AFTS scope (policy 2026-07-29). Both the
# fabricated row and the corrected one belong in Weekly_Rejected; they get
# there for different reasons.
GOOD_ROW = dict(BAD_ROW, Company="Auxico (Perth) Pty Ltd", Brand="LGM",
                Pathogen="Undeclared allergen (peanuts)", Tier=2,
                URL=ORIGINAL)

# An in-scope row used as the "nothing wrong here" baseline, so a test that
# means "the gate is not blocking everything" does not accidentally lean on a
# row the scope rule rightly rejects.
IN_SCOPE_ROW = {
    "Date": "2026-07-24", "Source": "RappelConso (FR)",
    "Company": "CRUSTA C", "Brand": "Unbranded",
    "Product": "crevettes cuites sauvages 40/60",
    "Pathogen": "Listeria monocytogenes",
    "Reason": "Presence of Listeria monocytogenes",
    "Class": "Voluntary", "Country": "France", "Region": "Europe",
    "Tier": 1, "Outbreak": 0,
    "URL": "https://rappel.conso.gouv.fr/fiche-rappel/22963/Interne",
    "Notes": "",
}


class TestDefect1_DedupAcrossRepublication(unittest.TestCase):

    def test_amended_slug_collapses_onto_the_original(self):
        self.assertEqual(_normalize_url_for_dedup(ORIGINAL),
                         _normalize_url_for_dedup(AMENDED),
                         "FSANZ republishes an amended alert at a new slug; "
                         "both addresses are the same recall")

    def test_dedup_key_matches_so_the_duplicate_cannot_be_minted(self):
        self.assertEqual(_dedup_key(GOOD_ROW), _dedup_key(BAD_ROW))

    def test_two_genuinely_different_fsanz_recalls_still_differ(self):
        a = FSANZ + "viet-meatballs-chinese-sausage-500g"
        b = FSANZ + "auxico-perth-pty-ltd-lgm-hot-chilli-oil-275g"
        self.assertNotEqual(_normalize_url_for_dedup(a),
                            _normalize_url_for_dedup(b))

    def test_other_regulators_are_untouched(self):
        """The rule is host-scoped. A slug elsewhere that happens to start
        with 'updated-' must survive intact."""
        u = "https://recalls-rappels.canada.ca/en/alert-recall/updated-123456-x"
        self.assertIn("updated-123456-x", _normalize_url_for_dedup(u))

    def test_a_product_named_updated_mid_slug_is_untouched(self):
        u = FSANZ + "acme-updated-140326-sauce"
        self.assertIn("acme-updated-140326-sauce", _normalize_url_for_dedup(u))


class TestDefect2_StatusBannerInCompany(unittest.TestCase):

    def test_gate_names_the_banner(self):
        blockers = " | ".join(publish_blockers(BAD_ROW)).lower()
        self.assertIn("status banner", blockers)

    def test_stripper_recovers_the_real_company(self):
        self.assertEqual("Auxico (Perth) Pty Ltd",
                         strip_title_status_prefix(BAD_ROW["Company"]))

    def test_stripper_is_identity_on_a_normal_company(self):
        for name in ("Auxico (Perth) Pty Ltd", "BM Foods (a member of Sea "
                     "Harvest Group)", "E.Leclerc Outreau", "Lidl",
                     "Origin: China | Notifying: Germany"):
            self.assertEqual(name, strip_title_status_prefix(name), name)

    def test_rasff_company_convention_survives(self):
        """Company legitimately contains a pipe on every RASFF row. The rule
        must key on the STATUS WORD, never on the pipe alone."""
        for name in ("Origin: Czechia | Notifying: Czechia",
                     "Origin: unknown origin | Notifying: Ireland",
                     "Origin: Belgium, Poland | Notifying: Belgium"):
            self.assertEqual(name, strip_title_status_prefix(name), name)
            self.assertEqual([], publish_blockers(dict(IN_SCOPE_ROW,
                                                       Company=name)), name)

    def test_writer_strips_it_on_every_sheet(self):
        from openpyxl import Workbook
        wb = Workbook()
        for sheet in ("Recalls", "Pending", "Weekly_Rejected"):
            rows = [dict(BAD_ROW)]
            _write_sheet(wb, sheet, SCHEMA, rows)
            ws = wb[sheet]
            col = SCHEMA.index("Company") + 1
            self.assertEqual("Auxico (Perth) Pty Ltd", ws.cell(2, col).value,
                             f"banner reached the {sheet} sheet")


class TestDefect3_HazardClassContradiction(unittest.TestCase):
    """The check that existed, worked, and guarded only one caller."""

    def test_the_published_row_is_blocked(self):
        self.assertFalse(is_publishable(BAD_ROW))

    def test_the_contradiction_is_named_explicitly(self):
        blockers = " | ".join(publish_blockers(BAD_ROW)).lower()
        self.assertIn("contradicts reason", blockers)

    def test_the_corrected_row_no_longer_contradicts_itself(self):
        """Correcting the Pathogen removes the contradiction — but the row is
        STILL blocked, now by the scope rule, because an undeclared peanut
        allergen is not a hazard this database covers."""
        blockers = " | ".join(publish_blockers(GOOD_ROW)).lower()
        self.assertNotIn("contradicts reason", blockers)
        self.assertIn("out of afts scope", blockers)

    def test_an_in_scope_row_passes_cleanly(self):
        self.assertEqual([], publish_blockers(IN_SCOPE_ROW))

    def test_classifier_reads_both_fields(self):
        self.assertEqual({"biological"},
                         classify_hazard("Listeria monocytogenes"))
        self.assertEqual({"allergen"}, classify_hazard(BAD_ROW["Reason"]))
        self.assertTrue(pathogen_reason_class_mismatch(
            BAD_ROW["Pathogen"], BAD_ROW["Reason"]))

    def test_the_second_confirmed_case(self):
        """Viet Meatballs, 2026-07-24 — same shape, same invented pathogen."""
        row = dict(GOOD_ROW,
                   Company="Viet Meatballs",
                   Product="Chinese Sausage 500g",
                   Pathogen="Listeria monocytogenes",
                   Reason="The recall is due to the presence of an "
                          "undeclared allergen (gluten).",
                   URL=FSANZ + "viet-meatballs-chinese-sausage-500g")
        self.assertFalse(is_publishable(row))

    def test_conservative_when_a_field_is_unclassifiable(self):
        """Vague text must NOT manufacture a mismatch."""
        self.assertFalse(pathogen_reason_class_mismatch(
            "Listeria monocytogenes", "Microbial contamination"))
        self.assertFalse(pathogen_reason_class_mismatch("", "anything"))
        self.assertFalse(pathogen_reason_class_mismatch("Undeclared milk", ""))

    def test_rasff_category_text_does_not_trip_it(self):
        """RASFF Reason carries 'category: milk and milk products' on genuine
        Listeria notifications. A bare food-name token would break every one
        of them — this is why the allergen list is framing-tokens only."""
        for reason in (
            "Listeria monocytogenes in smoked ribs raw materials originating "
            "from Germany; risk: serious; category: meat and meat products "
            "(other than poultry)",
            "Salmonella spp. in cured sausage from Spain; risk: serious; "
            "category: meat and meat products (other than poultry)",
            "Listeria monocytogenes in cheese; risk: serious; category: milk "
            "and milk products",
            "Clostridium botulinum in pork pâté with almonds from Czechia; "
            "risk: serious; category: meat and meat products",
        ):
            row = dict(IN_SCOPE_ROW, Pathogen=reason.split()[0], Reason=reason)
            self.assertEqual([], publish_blockers(row), reason[:60])

    def test_claude_check_shares_the_one_table(self):
        """No second copy may exist — divergence is how this bug survived."""
        src = (ROOT / "pipeline" / "claude_check.py").read_text(
            encoding="utf-8")
        self.assertIn("from pipeline._publish_gate import", src)
        self.assertNotIn('"biological": (', src,
                         "claude_check re-declared the hazard table; the two "
                         "copies will drift and one caller will go blind")


class TestDefect4_RejectionIsSticky(unittest.TestCase):
    """A URL in Weekly_Rejected does not get a second verdict."""

    def _pending(self, **over):
        row = dict(BAD_ROW, Status="pending", ScrapedAt="2026-08-01T01:10:25Z",
                   RejectedBy="")
        row.update(over)
        return row

    def test_previously_rejected_url_is_not_promoted(self):
        registry = {_normalize_url_for_dedup(AMENDED):
                    "claude-check: fail; pathogen mismatch"}
        new_approved, remaining, archived = promote_approved(
            pending=[self._pending()], approved_existing=[],
            rejected_flags={}, archive_immediately=True,
            previously_rejected=registry)
        self.assertEqual([], new_approved,
                         "a row claude-check already killed was promoted "
                         "again by a different reviewer")
        self.assertEqual(1, len(archived))

    def test_the_prior_verdict_is_quoted_in_the_trail(self):
        registry = {_normalize_url_for_dedup(AMENDED):
                    "claude-check: fail; pathogen mismatch"}
        _, _, archived = promote_approved(
            pending=[self._pending()], approved_existing=[],
            rejected_flags={}, archive_immediately=True,
            previously_rejected=registry)
        notes = str(archived[0].get("Notes") or "")
        self.assertIn("re-promotion blocked", notes)
        self.assertIn("pathogen mismatch", notes)
        self.assertEqual("repromotion-guard", archived[0].get("RejectedBy"))

    def test_the_guard_matches_across_the_republished_slug(self):
        """Rejected at the amended URL, resubmitted at the original — still
        the same recall, still blocked."""
        registry = {_normalize_url_for_dedup(AMENDED): "claude-check: fail"}
        new_approved, _, archived = promote_approved(
            pending=[self._pending(URL=ORIGINAL)], approved_existing=[],
            rejected_flags={}, archive_immediately=True,
            previously_rejected=registry)
        self.assertEqual([], new_approved)
        self.assertEqual(1, len(archived))

    def test_a_clean_url_still_promotes(self):
        """The guard must not become a blanket freeze."""
        new_approved, _, _ = promote_approved(
            pending=[self._pending(**IN_SCOPE_ROW)], approved_existing=[],
            rejected_flags={}, archive_immediately=True,
            previously_rejected={})
        self.assertEqual(1, len(new_approved))

    def test_registry_loads_from_the_live_workbook(self):
        xlsx = ROOT / "docs" / "data" / "recalls.xlsx"
        if not xlsx.exists():                     # pragma: no cover
            self.skipTest("recalls.xlsx not present")
        reg = load_rejected_urls(xlsx)
        self.assertTrue(reg, "Weekly_Rejected produced no URLs — the guard "
                             "would be silently inert")

    def test_missing_workbook_degrades_to_no_guard_not_a_crash(self):
        self.assertEqual({}, load_rejected_urls(ROOT / "nope" / "nope.xlsx"))


class TestTheWorkbookIsClean(unittest.TestCase):
    """The published sheet, after the repair."""

    def _recalls(self):
        try:
            import openpyxl
        except ImportError:                       # pragma: no cover
            self.skipTest("openpyxl not installed")
        xlsx = ROOT / "docs" / "data" / "recalls.xlsx"
        if not xlsx.exists():                     # pragma: no cover
            self.skipTest("recalls.xlsx not present")
        wb = openpyxl.load_workbook(xlsx, read_only=True)
        rows = list(wb["Recalls"].values)
        hdr = [str(h) for h in rows[0]]
        return [dict(zip(hdr, r)) for r in rows[1:] if r]

    def test_both_auxico_rows_are_gone(self):
        """The duplicate for being a duplicate with an invented pathogen; the
        original for being an allergen-only recall, which AFTS does not
        cover. Neither belongs in Recalls."""
        hits = [r for r in self._recalls()
                if "auxico" in str(r.get("Company") or "").lower()]
        self.assertEqual([], hits, f"{len(hits)} Auxico row(s) still published")

    def test_no_allergen_only_row_is_published(self):
        IN_SCOPE = {"biological", "biotoxin", "mycotoxin", "physical",
                    "chemical", "fermentation", "spoilage"}
        bad = []
        for r in self._recalls():
            classes = (classify_hazard(str(r.get("Pathogen") or ""))
                       | classify_hazard(str(r.get("Reason") or "")))
            if classes and not (classes & IN_SCOPE):
                bad.append((r["Date"], str(r.get("Company"))[:32],
                            str(r.get("Pathogen"))[:32]))
        self.assertEqual([], bad,
                         f"{len(bad)} allergen-only row(s) in Recalls — "
                         f"out of AFTS scope per policy 2026-07-29")

    def test_no_row_carries_a_status_banner(self):
        bad = [(r["Date"], r["Company"]) for r in self._recalls()
               if strip_title_status_prefix(r.get("Company")) != r.get("Company")]
        self.assertEqual([], bad)

    # ── Pre-existing contamination, now repaired from source ──────────────
    # Ten RappelConso rows carried a hazard-class contradiction from the
    # url-gate batch mis-attribution (tests/test_url_gate_identity.py). Nine
    # have been REPAIRED from the official DGCCRF open-data record — see
    # pipeline/verify_rappelconso.py. Fiche 22205 is absent from the export
    # (ids jump 22204 -> 22206), so it stays pinned; guessing is the failure
    # this whole gate exists to prevent. This set must only ever shrink.
    KNOWN_CONTAMINATED = frozenset({
        "fiche-rappel/22205",   # Listeria mc <- aflatoxin / corn wafers (SK)
    })

    def test_no_NEW_row_contradicts_itself_on_hazard_class(self):
        bad = [(r["Date"], str(r.get("Pathogen"))[:30], str(r.get("URL"))[:70])
               for r in self._recalls()
               if pathogen_reason_class_mismatch(str(r.get("Pathogen") or ""),
                                                 str(r.get("Reason") or ""))
               and not any(k in str(r.get("URL") or "")
                           for k in self.KNOWN_CONTAMINATED)]
        self.assertEqual([], bad, f"{len(bad)} NEW self-contradicting row(s)")

    def test_the_repaired_french_rows_stayed_repaired(self):
        """The nine fixed from the official DGCCRF record, by fiche id."""
        import re as _re
        repaired = {"21975", "21987", "22067", "22082", "22113",
                    "22157", "22186", "22206", "22208", "22184"}
        seen = {}
        for r in self._recalls():
            m = _re.search(r"fiche-rappel/(\d+)", str(r.get("URL") or ""))
            if m and m.group(1) in repaired:
                seen[m.group(1)] = r
        # Not one of them may still carry RASFF notification text.
        leaked = [f for f, r in seen.items()
                  if "; risk:" in str(r.get("Reason") or "")
                  or "category:" in str(r.get("Reason") or "")]
        self.assertEqual([], leaked,
                         f"RASFF text is back on French fiche(s): {leaked}")
        # Nor RASFF's Company convention on a French fiche.
        rasff_co = [f for f, r in seen.items()
                    if str(r.get("Company") or "").startswith("Origin:")]
        self.assertEqual([], rasff_co,
                         f"RASFF Company convention on French fiche(s): "
                         f"{rasff_co}")

    def test_no_text_field_holds_the_literal_zero(self):
        bad = [(r["Date"], k, str(r.get("URL"))[:60])
               for r in self._recalls()
               for k in ("Company", "Brand", "Product", "Pathogen", "Reason",
                         "Class", "Country")
               if str(r.get(k) or "").strip() == "0"]
        self.assertEqual([], bad)

    def test_no_two_rows_share_a_dedup_key(self):
        import collections
        counts = collections.Counter(_dedup_key(r) for r in self._recalls())
        dupes = [k for k, n in counts.items() if n > 1]
        self.assertEqual([], dupes)


if __name__ == "__main__":
    unittest.main(verbosity=2)


class TestUkropsFabricatedPathogen(unittest.TestCase):
    """The 2026-08-04 fabrication: 'Hepatitis A virus' on a metal recall.

    Three guards failed in three different ways, so all three are pinned.
    Verified at source: the FDA permalink ends 'due-possible-foreign-object',
    the hazard is aluminium slivers from the baking pans, and no pathogen is
    named anywhere on the notice.
    """

    REASON = ("Baked products have potential for presence of aluminum slivers "
              "from the pans that were used")
    ROW = {
        "Date": "2026-08-01", "Source": "FDA",
        "Company": "Ukrops Homestyle Foods", "Brand": "Ukrops Homestyle Foods",
        "Product": "Baked Spaghetti and Bread Pudding products",
        "Pathogen": "Hepatitis A virus", "Reason": REASON,
        "Class": "Recall", "Country": "United States", "Region": "North America",
        "Tier": 1, "Outbreak": 0,
        "URL": "https://www.fda.gov/safety/recalls-market-withdrawals-safety-"
               "alerts/ukrops-homestyle-foods-announces-voluntary-recall-due-"
               "possible-foreign-object",
        "Notes": "",
    }

    def test_gap_1_foreign_object_wording_now_classifies(self):
        """'aluminum slivers' matched nothing, so the Reason was
        unclassifiable and the contradiction rule failed safe."""
        self.assertIn("physical", classify_hazard(self.REASON))

    def test_gap_2_the_bare_hav_keyword_no_longer_matches_the_word_have(self):
        """The worse half: 'hav' was in the BIOLOGICAL list as the Hepatitis A
        abbreviation and matched 'HAVe', so the invented pathogen and the
        reason appeared to agree."""
        self.assertNotIn("biological", classify_hazard(self.REASON))
        self.assertNotIn("biological", classify_hazard(
            "We have received no reports of illness to date"))
        # …while the real abbreviation still classifies.
        self.assertIn("biological", classify_hazard("Hepatitis A virus (HAV)"))

    def test_the_contradiction_is_caught(self):
        self.assertTrue(pathogen_reason_class_mismatch(
            "Hepatitis A virus", self.REASON))
        blockers = " | ".join(publish_blockers(self.ROW)).lower()
        self.assertIn("contradicts reason", blockers)

    def test_gap_3_tier_guard_does_not_escalate_on_a_contradicted_pathogen(self):
        """It forced Tier 2 -> 1 on the strength of the fabrication, and
        stamped Notes so the row looked reviewed."""
        from pipeline._pathogen_scope import enforce_tier1
        row = dict(self.ROW, Tier=2, Notes="")
        enforce_tier1(row)
        self.assertEqual(2, row["Tier"],
                         "the tier-guard escalated on a pathogen the row's own "
                         "Reason contradicts")
        self.assertIn("escalation SKIPPED", str(row["Notes"]))

    def test_a_genuine_always_tier1_row_still_escalates(self):
        """The rule itself is correct and must keep working."""
        from pipeline._pathogen_scope import enforce_tier1
        row = {"Pathogen": "Hepatitis A virus",
               "Reason": "Hepatitis A virus detected in frozen berries",
               "Tier": 3, "Notes": ""}
        enforce_tier1(row)
        self.assertEqual(1, row["Tier"])

    def test_slivered_almonds_are_not_a_physical_hazard(self):
        """Why the vocabulary uses qualified shapes and never a bare
        'sliver' — 'slivered almonds' is an ingredient."""
        self.assertNotIn("physical", classify_hazard(
            "Aflatoxin B1 above the limit in slivered almonds from the USA"))

    def test_the_workbook_row_is_repaired(self):
        try:
            import openpyxl
        except ImportError:                       # pragma: no cover
            self.skipTest("openpyxl not installed")
        xlsx = ROOT / "docs" / "data" / "recalls.xlsx"
        if not xlsx.exists():                     # pragma: no cover
            self.skipTest("recalls.xlsx not present")
        wb = openpyxl.load_workbook(xlsx, read_only=True)
        rows = list(wb["Recalls"].values)
        hdr = [str(h) for h in rows[0]]
        hits = [dict(zip(hdr, r)) for r in rows[1:] if r
                and "ukrops" in str(dict(zip(hdr, r)).get("Company") or "").lower()]
        self.assertEqual(1, len(hits))
        row = hits[0]
        self.assertNotIn("hepatitis", str(row["Pathogen"]).lower())
        self.assertEqual("2", str(row["Tier"]))
        self.assertNotIn("[tier-guard: Hepatitis A virus", str(row["Notes"]))
