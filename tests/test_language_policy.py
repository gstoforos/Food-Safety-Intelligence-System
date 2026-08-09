"""The English-output policy, pinned.

    "everything in English except brand / or product name"   — operator, 2026-08-02

WHAT WENT WRONG
===============
The dashboard card for 2026-07-31 rendered like this:

    AESAN (ES)
    Alerta por presencia de Salmonella spp. en chips de fruta procedentes
    de Alemania. (Ref. ES2026/473)                       <- Company
    Aesan - Agencia Española de Seguridad Alimentaria y
    Nutrición # Alerta por presencia de Salmonella spp   <- Product
    Salmonella                                           TIER-1  Spain

Two separate failures in one card. The Spanish is the visible one. The worse
one is that neither field is a NAME: Company held the alert page title and
Product held the site banner, so the card names no company and no product.

Across the workbook: 157 rows carried a non-English Reason, 30 carried a
bilingual RASFF notification subject in Product, and 5 AESAN rows had the
alert title or the regulator's own name in Company/Brand.

THE LINE THIS FILE DEFENDS
==========================
Reason, Class, Pathogen, Country and Region are DESCRIPTION → must be English.
Company, Brand and Product are NAMES → stay exactly as the regulator published
them. "brie a l'ail", "saucisson à l'ail fumé", "Χούμους" and "Freshona Bio
Beerenmischung" are correct and must never be "fixed".

The distinction is testable, not a matter of taste: a bilingual regulator
subject SPLITS into two languages around a separator; a product name does not.

Run:  python -m pytest tests/test_language_policy.py -v
"""
from __future__ import annotations
import sys
import unittest
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from pipeline._language import (  # noqa: E402
    detect_language, looks_non_english, split_bilingual, to_english,
    englishify_reason, REASON_EN,
)

MUST_BE_ENGLISH = ("Reason", "Class", "Pathogen", "Country", "Region")
NAME_FIELDS = ("Company", "Brand", "Product")


class TestDetector(unittest.TestCase):

    def test_flags_the_published_offenders(self):
        for text, lang in (
            ("Presence of salmonelle dans le produit", "fr"),
            ("Rappel pour raison sanitaire", "fr"),
            ("Presencia de Salmonella spp.", "es"),
            ("Presenza di Salmonella Typhimurium", "it"),
            ("In tiefgekühlten Beeren wurden Noroviren nachgewiesen. Eine "
             "Gesundheitsgefährdung kann nicht ausgeschlossen werden.", "de"),
        ):
            self.assertEqual(lang, detect_language(text), text[:40])

    def test_leaves_english_alone(self):
        for text in (
            "Presence of Listeria monocytogenes",
            "Detection of Salmonella during own-check testing",
            "T-2 and HT-2 toxin levels above the regulatory limit",
            "Cadmium above the maximum permitted level",
            "Non-compliant bacteriological analysis — Escherichia coli (STEC) "
            "present. Sold at Intermarché Le Passage (RappelConso fiche 22664)",
            "Fromagerie P. Jacquin & Fils brand Valençay AOP \"fromage de "
            "chèvre au lait cru\" recalled due to generic E. coli",
        ):
            self.assertFalse(looks_non_english(text), text[:60])

    def test_a_quoted_greek_name_does_not_make_a_sentence_greek(self):
        """An English sentence citing a Greek laboratory is still English —
        the same principle that exempts Greek product names."""
        self.assertFalse(looks_non_english(
            "EFET Regional Directorate of Thessaly conducted sampling under "
            "the 2026 Official Microbiological Criteria Programme. Sample "
            "analyzed at the Thessaloniki Food Testing & Research Laboratory "
            "(Δ/νση Εργαστηριακών Δομών)."))

    def test_real_greek_text_is_flagged(self):
        self.assertEqual("el", detect_language(
            "Ανάκληση προϊόντος λόγω παρουσίας Listeria monocytogenes στο "
            "προϊόν από τον Ενιαίο Φορέα Ελέγχου Τροφίμων"))

    def test_short_strings_are_never_guessed(self):
        for text in ("Salmonella", "Listeria monocytogenes", "", None, "brie"):
            self.assertFalse(looks_non_english(text), repr(text))


class TestBilingualSplit(unittest.TestCase):
    """RASFF publishes native//English in EITHER order."""

    def test_native_first(self):
        self.assertEqual(
            "Presence of Salmonella spp. in cured sausage from Spain; "
            "risk: serious; category: meat and meat products (other than poultry)",
            split_bilingual(
                "Presencia de Salmonela spp en salchichón procedente de España "
                "// Presence of Salmonella spp. in cured sausage from Spain; "
                "risk: serious; category: meat and meat products (other than "
                "poultry)"))

    def test_english_first(self):
        """The bug a naive 'take the second half' would have introduced."""
        out = split_bilingual(
            "Listeria in foie gras from Bulgaria / Listeria en bloc de foie "
            "gras de pato procedente de Bulgaria; risk: serious; category: "
            "poultry meat and poultry meat products")
        self.assertTrue(out.startswith("Listeria in foie gras from Bulgaria"),
                        out)
        self.assertNotIn("procedente", out)

    def test_five_slash_separator(self):
        out = split_bilingual(
            "Aflatoxins above permissible limits in peanuts from Nigeria "
            "///// Aflatoxinas por encima de los límites permitidos en "
            "cacahuetes procedentes de Nigeria; risk: serious; category: nuts")
        self.assertIn("Aflatoxins above permissible limits", out)
        self.assertNotIn("cacahuetes", out)

    def test_the_risk_tail_survives(self):
        out = split_bilingual(
            "Fumonisina en Harina de Maiz BIO procedente de Italia // "
            "Fumonisin in organic maize flour from Italy; risk: serious; "
            "category: cereals and bakery products")
        self.assertIn("; risk: serious; category: cereals and bakery products",
                      out)

    def test_a_product_NAME_never_splits(self):
        """This is what protects the brand/product-name exemption."""
        for name in ("brie a l'ail", "charcuterie seche",
                     "saucisson a l'ail a l'ancienne courbe fume",
                     "Freshona Bio Beerenmischung tiefgefroren, 300g",
                     "Χούμους (Hummus)",
                     "Bacalao Natura en aceite de girasol",
                     "Valençay AOP \"fromage de chèvre au lait cru\""):
            self.assertIsNone(split_bilingual(name), name)


class TestTranslationsAreVerifiedNotGuessed(unittest.TestCase):

    def test_known_strings_translate(self):
        self.assertEqual("Presence of Salmonella in the product",
                         to_english("Presence of salmonelle dans le produit"))
        self.assertEqual("Recall on public-health grounds",
                         to_english("Rappel pour raison sanitaire"))

    def test_clinically_distinct_motifs_stay_distinct(self):
        """'suspicion de', 'mise en évidence de' and 'présence présumée' are
        not interchangeable in a food-safety brief, which is why this is a
        table and not a phrase engine."""
        suspected = to_english("Suspicion de présence de Listeria monocytogenes")
        shown = to_english("Mise en évidence de la présence de Listeria monocytogenes")
        presumed = to_english("Présence présumée de e.coli stec o26:h11")
        self.assertIn("Suspected", suspected)
        self.assertIn("demonstrated", shown)
        self.assertIn("Presumed", presumed)
        self.assertNotEqual(suspected, shown)

    def test_an_unknown_string_is_NOT_machine_translated(self):
        self.assertIsNone(to_english(
            "Une nouvelle formulation jamais vue auparavant dans ce corpus "
            "avec des mots differents"))
        text = "Un motif totalement inconnu de cette table de correspondance"
        out, changed = englishify_reason(text)
        self.assertFalse(changed)
        self.assertEqual(text, out, "an unknown motif must be left alone and "
                                    "reported, never guessed at")

    def test_english_input_is_untouched(self):
        text = "Presence of Listeria monocytogenes"
        out, changed = englishify_reason(text)
        self.assertFalse(changed)
        self.assertEqual(text, out)

    def test_the_table_has_no_empty_translations(self):
        for key, value in REASON_EN.items():
            self.assertTrue(value.strip(), key[:60])
            self.assertFalse(looks_non_english(value),
                             f"translation is still not English: {value[:70]!r}")


class TestWriterGuardIsWired(unittest.TestCase):

    def test_merge_master_runs_the_split_on_every_write(self):
        src = (ROOT / "pipeline" / "merge_master.py").read_text(encoding="utf-8")
        self.assertIn("from pipeline._language import", src)
        self.assertIn("split_bilingual as _split_bilingual", src)
        self.assertIn('for _col in ("Reason", "Product")', src)

    def test_the_writer_does_not_machine_translate(self):
        """The writer may only do the MECHANICAL split. Translating a whole
        sentence is a deliberate, auditable act, never a silent one."""
        src = (ROOT / "pipeline" / "merge_master.py").read_text(encoding="utf-8")
        self.assertNotIn("englishify_reason", src)


class TestPublishedWorkbook(unittest.TestCase):

    def _rows(self, sheet="Recalls"):
        try:
            import openpyxl
        except ImportError:                       # pragma: no cover
            self.skipTest("openpyxl not installed")
        xlsx = ROOT / "docs" / "data" / "recalls.xlsx"
        if not xlsx.exists():                     # pragma: no cover
            self.skipTest("recalls.xlsx not present")
        wb = openpyxl.load_workbook(xlsx, read_only=True)
        rows = list(wb[sheet].values)
        hdr = [str(h) for h in rows[0]]
        return [dict(zip(hdr, r)) for r in rows[1:] if r]

    def test_every_description_field_is_english(self):
        offenders = []
        for row in self._rows():
            for field in MUST_BE_ENGLISH:
                if looks_non_english(row.get(field)):
                    offenders.append((field, detect_language(row.get(field)),
                                      str(row.get("Source")),
                                      str(row.get(field))[:70]))
        self.assertEqual([], offenders,
                         f"{len(offenders)} non-English value(s) in fields that "
                         f"must be English: {offenders[:6]}")

    def test_no_bilingual_subject_is_left_in_product(self):
        """Product may be a foreign NAME. On RASFF rows it may not be a
        two-language regulator sentence.

        SCOPED TO RASFF (audit 2026-08-04, second pass). Asking this question
        of every source is what truncated an FDA product name at its
        semicolon and replaced a Greek one with a fragment of its packing
        address — see TestProductSplitIsScopedToRASFF. Only RASFF writes a
        bilingual notification subject into Product; a "/" or ";" anywhere
        else is just punctuation in a name.
        """
        offenders = [(str(r.get("Source")), str(r.get("Product"))[:70])
                     for r in self._rows()
                     if str(r.get("Source")) == "RASFF (EU)"
                     and looks_non_english(r.get("Product"))
                     and split_bilingual(str(r.get("Product") or ""))]
        self.assertEqual([], offenders, f"{len(offenders)} bilingual subject(s) "
                                        f"still in Product: {offenders[:5]}")

    def test_foreign_product_names_were_preserved(self):
        """The exemption is real: these must still be there, untranslated."""
        products = {str(r.get("Product") or "") for r in self._rows()}
        for name in ("brie a l'ail", "charcuterie seche"):
            self.assertIn(name, products,
                          f"{name!r} was translated — product NAMES are exempt")

    def test_no_regulator_name_is_published_as_a_company(self):
        banned = ("agencia española de seguridad alimentaria",
                  "aesan - agencia", "food standards agency",
                  "agence nationale", "bundesamt für verbraucherschutz")
        offenders = [(str(r.get("Date")), str(r.get("Company"))[:60])
                     for r in self._rows()
                     for b in banned
                     if b in str(r.get("Company") or "").lower()]
        self.assertEqual([], offenders,
                         f"the regulator is published as the recalling "
                         f"company: {offenders}")

    def test_no_alert_title_is_published_as_a_company(self):
        """'Alerta por presencia de …' and 'X recalls Y because of Z' are
        headlines, not company names."""
        offenders = []
        for r in self._rows():
            c = str(r.get("Company") or "")
            low = c.lower()
            if low.startswith(("alerta por", "alerta alimentaria",
                               "ampliación de la alerta", "rappel de",
                               "rückruf von")) or " recalls " in low:
                offenders.append((str(r.get("Date")), c[:70]))
        self.assertEqual([], offenders, f"alert headline(s) in Company: "
                                        f"{offenders}")


if __name__ == "__main__":
    unittest.main(verbosity=2)


class TestAudit20260804(unittest.TestCase):
    """Two rows reached Recalls on 2026-08-04 that both guards let through.

    Reported by the operator's second reviewer; both confirmed at source.
    """

    def test_short_french_with_a_single_common_marker_is_flagged(self):
        """'Contamination à la salmonellle' — RappelConso fiche 23075.

        Missed because the only marker it carried was 'la', one hit against a
        two-hit threshold, while 'contamination' reads as English. 'à' is now
        a marker in its own right.
        """
        self.assertEqual("fr", detect_language("Contamination à la salmonellle"))
        out, changed = englishify_reason("Contamination à la salmonellle")
        self.assertTrue(changed)
        self.assertEqual("Contamination with Salmonella", out)

    def test_the_five_it_uncovered(self):
        """Widening the detector surfaced five more, invisible for the same
        reason. Each is translated, none is guessed."""
        for french in (
            "Contamination potentielle à la Listeria monocytogenes",
            "Presence of listéria monocytogènes inférieur à 10/g",
            "Detection of listéria monocytogenes inférieure à 10 ufc/g",
            "Presence of Listeria (inférieure à 10)",
            "Présence d'ochratoxine A",
        ):
            out, changed = englishify_reason(french)
            self.assertTrue(changed, french)
            self.assertFalse(looks_non_english(out), out)

    def test_widening_did_not_break_english(self):
        """'a' and 'à' became French markers. English must still pass."""
        for text in (
            "Presence of Listeria monocytogenes",
            "Baked products have potential for presence of aluminum slivers "
            "from the pans that were used",
            "Isolated detection of Salmonella on a batch of Lardons Fumés "
            "200g+25%",
            "Analysis demonstrated the presence of Salmonella spp.",
            "Listeria monocytogenes present at below 10 CFU/g",
            "Recall of batch 11070 of 9-month bone-in dry-cured ham from "
            "Salaisons Limousines because of a risk of Listeria monocytogenes "
            "contamination",
        ):
            self.assertFalse(looks_non_english(text), text[:70])


class TestProductSplitIsScopedToRASFF(unittest.TestCase):
    """Product is a NAME field everywhere except RASFF.

    Audit 2026-08-04, second pass. Running the bilingual splitter on every
    source truncated two real products before the scope was added:

        FDA   "Hellas Meze Golden Smoked Whole Herring, vacuum-packaged,
               refrigerated; production date 4/12/2025, best before
               4/12/2026, lot L120425F54. Imported from Karagounis Bros
               (Greece)."
              -> everything after "refrigerated" was lost.

        EFET  'Σαλάτα "ΜΑΡΟΥΛΕΝΙΑ" — μαρούλι romaine, butterhead & escarole
               (Lettuce salad "Maroulenia"…). Lot L/261462 5; use-by
               03/06/2026. Packaged in Greece (ΒΙ.ΠΕ. / Industrial Area,
               Central Macedonia)'
              -> reduced to "Industrial Area, Central Macedonia)." The
                 product name vanished and a fragment of the packing
                 address survived in its place.

    The heuristic was not wrong about those strings containing a "/" or a
    ";" — it was wrong to be asked the question at all. Only RASFF writes a
    bilingual notification SUBJECT into Product.
    """

    def test_the_writer_guard_is_scoped(self):
        src = (ROOT / "pipeline" / "merge_master.py").read_text(encoding="utf-8")
        self.assertIn('if _col == "Product" and', src)
        self.assertIn('!= "RASFF (EU)"', src)

    def test_the_one_off_is_scoped(self):
        src = (ROOT / "pipeline" / "fix_product_bilingual.py").read_text(
            encoding="utf-8")
        self.assertIn('!= "RASFF (EU)"', src)

    def test_the_two_damaged_products_are_whole(self):
        try:
            import openpyxl
        except ImportError:                       # pragma: no cover
            self.skipTest("openpyxl not installed")
        xlsx = ROOT / "docs" / "data" / "recalls.xlsx"
        if not xlsx.exists():                     # pragma: no cover
            self.skipTest("recalls.xlsx not present")
        wb = openpyxl.load_workbook(xlsx, read_only=True)
        rows = list(wb["Recalls"].values)
        hdr = [str(h) for h in rows[0]]
        products = [str(dict(zip(hdr, r)).get("Product") or "")
                    for r in rows[1:] if r]
        hellas = [p for p in products if p.startswith("Hellas Meze")]
        self.assertTrue(hellas)
        self.assertIn("lot L120425F54", hellas[0],
                      "the FDA product lost its lot/date/importer detail")
        self.assertNotIn("Industrial Area, Central Macedonia).", products,
                         "an EFET product name was replaced by a fragment of "
                         "its packing address")
        greek = [p for p in products if "ΜΑΡΟΥΛΕΝΙΑ" in p]
        self.assertTrue(greek, "the Greek product name was not restored")

    def test_no_published_product_is_a_bare_address_fragment(self):
        try:
            import openpyxl
        except ImportError:                       # pragma: no cover
            self.skipTest("openpyxl not installed")
        xlsx = ROOT / "docs" / "data" / "recalls.xlsx"
        if not xlsx.exists():                     # pragma: no cover
            self.skipTest("recalls.xlsx not present")
        wb = openpyxl.load_workbook(xlsx, read_only=True)
        rows = list(wb["Recalls"].values)
        hdr = [str(h) for h in rows[0]]
        bad = [str(dict(zip(hdr, r)).get("Product") or "") for r in rows[1:]
               if r and str(dict(zip(hdr, r)).get("Product") or "")
               .rstrip().endswith(")")
               and str(dict(zip(hdr, r)).get("Product") or "").count("(") == 0]
        self.assertEqual([], bad,
                         f"Product ends with an unmatched ')' — the tell of a "
                         f"split through the middle of a parenthetical: {bad}")


class TestAudit20260809(unittest.TestCase):
    """Half-translated PATHOGEN NAMES, and the detector-first ordering bug.

    26 published RappelConso Reasons read "Presence of salmonelle" — verb in
    English, organism still French. detect_language() cannot see them: it
    requires two function-word hits and that string has none. So the writer
    guard scored them English and left them alone.

    The deeper defect was the ORDER inside englishify_reason(): the statistical
    detector gated the verified table. A table entry someone wrote down by hand
    is stronger evidence than a word-frequency heuristic, so the table is now
    consulted first. Restoring the old order re-hides 57 rows.
    """

    def test_the_published_offenders_translate(self):
        for src, want in (
            ("Presence of salmonelle", "Presence of Salmonella"),
            ("Detection of salmonelle", "Detection of Salmonella"),
            ("Salmonelle", "Salmonella"),
            ("Presence of salmonelle enteritidis",
             "Presence of Salmonella Enteritidis"),
            ("Presence salmonelle s. typhimurium",
             "Presence of Salmonella Typhimurium"),
        ):
            out, changed = englishify_reason(src)
            self.assertTrue(changed, f"{src!r} was left alone")
            self.assertEqual(want, out)

    def test_the_detector_still_cannot_see_them(self):
        """The premise. If this ever starts returning True the ordering fix is
        no longer what is doing the work, and this suite would pass for the
        wrong reason."""
        self.assertFalse(looks_non_english("Presence of salmonelle"))

    def test_the_table_is_consulted_before_the_detector(self):
        src = (ROOT / "pipeline" / "_language.py").read_text(encoding="utf-8")
        body = src.split("def englishify_reason", 1)[1]
        # CODE only — the explanatory comment above names both functions, and
        # matching on prose would make this test assert about paragraph order.
        body = "\n".join(l for l in body.splitlines()
                         if not l.lstrip().startswith("#"))
        first_to_english = body.find("out = to_english(s)")
        first_detector = body.find("if not looks_non_english(s)")
        self.assertNotEqual(-1, first_to_english)
        self.assertNotEqual(-1, first_detector)
        self.assertLess(first_to_english, first_detector,
                        "the statistical detector is gating the verified "
                        "table again — that is what hid 26 rows")

    def test_an_unknown_string_is_still_never_guessed(self):
        text = "Une formulation totalement inconnue de cette table"
        out, changed = englishify_reason(text)
        self.assertFalse(changed)
        self.assertEqual(text, out)

    def test_no_salmonelle_survives_in_a_published_reason(self):
        try:
            import openpyxl
        except ImportError:                            # pragma: no cover
            self.skipTest("openpyxl not installed")
        xlsx = ROOT / "docs" / "data" / "recalls.xlsx"
        if not xlsx.exists():                          # pragma: no cover
            self.skipTest("recalls.xlsx not present")
        import re
        wb = openpyxl.load_workbook(xlsx, read_only=True)
        rows = list(wb["Recalls"].values)
        hdr = [str(h) for h in rows[0]]
        bad = [str(dict(zip(hdr, r)).get("Reason"))
               for r in rows[1:] if r
               and re.search(r"\bsalmonelle\b",
                             str(dict(zip(hdr, r)).get("Reason") or ""), re.I)]
        self.assertEqual([], bad, f"{len(bad)} published Reason(s) still name "
                                  f"the organism in French: {bad[:3]}")


class TestDittoIsNotAProductName(unittest.TestCase):
    """RappelConso fiche 23108 published with Product = "idem".

    Not a scraper artifact: the DGCCRF open-data record carries
    modeles_ou_references = "idem" verbatim. The notifier typed a ditto mark.
    The pipeline captured it faithfully — and a ditto pointing at a field the
    reader cannot see still names no product.
    """

    def test_the_gate_rejects_a_ditto_product(self):
        from pipeline._publish_gate import publish_blockers
        row = {"Date": "2026-08-06", "Source": "RappelConso (FR)",
               "Company": "La Fumerie du Coin", "Brand": "La Fumerie du Coin",
               "Product": "idem", "Pathogen": "Listeria monocytogenes",
               "Reason": "Presence of Listeria monocytogenes",
               "Class": "Voluntary", "Country": "France", "Region": "Europe",
               "URL": "https://rappel.conso.gouv.fr/fiche-rappel/23108/Interne"}
        self.assertTrue(any("Product is empty" in b
                            for b in publish_blockers(row)),
                        "a ditto mark passed as a product name")

    def test_a_real_product_name_is_untouched(self):
        from pipeline._publish_gate import publish_blockers
        row = {"Date": "2026-08-06", "Source": "RappelConso (FR)",
               "Company": "La Fumerie du Coin", "Brand": "La Fumerie du Coin",
               "Product": "Saumon fumé 200 g", "Pathogen": "Listeria monocytogenes",
               "Reason": "Presence of Listeria monocytogenes",
               "Class": "Voluntary", "Country": "France", "Region": "Europe",
               "URL": "https://rappel.conso.gouv.fr/fiche-rappel/23108/Interne"}
        self.assertEqual([], publish_blockers(row))


class TestReviewAgentUsesTheCanonicalHazardTable(unittest.TestCase):
    """The FSANZ Key-Sun lozenge (2026-08-05) carried a fabricated
    "Listeria monocytogenes" against Reason "risk of the presence of foreign
    matter (metal)". _publish_gate blocked it. recall_review_agent did not,
    because it kept a PRIVATE copy of the hazard table that knew
    "foreign body" and "foreign material" but not "foreign matter".

    Third drift of a duplicated hazard table. The agent now delegates.
    """

    def test_foreign_matter_is_a_physical_hazard(self):
        from pipeline._publish_gate import classify_hazard
        self.assertIn("physical", classify_hazard(
            "There is a risk of the presence of foreign matter (metal)."))

    def test_the_gate_blocks_the_lozenge(self):
        from pipeline._publish_gate import pathogen_reason_class_mismatch
        self.assertTrue(pathogen_reason_class_mismatch(
            "Listeria monocytogenes",
            "There is a risk of the presence of foreign matter (metal)."))

    def test_the_agent_delegates_instead_of_copying(self):
        src = (ROOT / "pipeline" / "recall_review_agent.py").read_text(
            encoding="utf-8")
        body = src.split("def _pathogen_reason_contradiction", 1)[1]
        self.assertIn("pathogen_reason_class_mismatch", body,
                      "recall_review_agent is still walking its private "
                      "hazard list as the primary path")
