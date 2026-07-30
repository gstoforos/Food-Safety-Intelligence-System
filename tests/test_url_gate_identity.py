"""Regression tests for the url-gate batch identity guard and the
writer-level Class normalisation.

WHY THIS FILE EXISTS (audit 2026-07-30)
=======================================
A. CROSS-ROW CONTAMINATION.
   url_gate_gemini sends rows to Gemini in batches and maps each returned
   decision back with `real_idx = start + d["row_index"]`. Every check around
   that line validated whether the decision was PLAUSIBLE — confidence,
   date_match, brand_match, bare-domain, JS artifacts — and none validated
   whether it belonged to that row. `row_index` is self-reported by the
   model, so a shifted index passed all of them: the content is genuine, it
   just describes a different recall.

   Found in production: ten RappelConso rows carrying a neighbour's Reason.
   Each shares its exact Reason string with a row from another source whose
   Pathogen matches that Reason correctly, e.g.

       Recalls row 740, fiche 22205, Pathogen "Listeria monocytogenes"
         Reason "Aflatoxins in mini corn wafers from Slovakia, raw material
                 from Hungary.; risk: serious; category: ..."
       Recalls row 745, RASFF (EU), Pathogen "Aflatoxin"
         Reason  <identical string>

   That "; risk: serious; category: ..." shape is verbatim RASFF
   notification text; RappelConso never emits it.

B. CLASS LANGUAGE BYPASS.
   41 rows reached the published Recalls sheet holding raw French Class
   values ("volontaire (sans arrete prefectoral)", "impose par arrete
   prefectoral") despite Recall.__post_init__ and promote_approved both
   normalising. All 41 carry a [url-gate ...] note — they were updated in
   place after promotion, so neither earlier gate ran again.

Run:  python -m pytest tests/test_url_gate_identity.py -v
"""
from __future__ import annotations
import sys
import unittest
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from pipeline.url_gate_gemini import _norm_echo_url  # noqa: E402
from pipeline.merge_master import _write_sheet, SCHEMA  # noqa: E402
from scrapers._models import _normalize_class_language  # noqa: E402


FICHE = "https://rappel.conso.gouv.fr/fiche-rappel/{}/Interne"


class TestEchoUrlNormalisation(unittest.TestCase):
    """Loose enough to tolerate cosmetic drift, strict enough that two
    different recalls never compare equal."""

    def test_scheme_host_case_and_trailing_slash_tolerated(self):
        a = "https://WWW.Rappel.Conso.gouv.FR/fiche-rappel/23044/Interne/"
        b = "http://rappel.conso.gouv.fr/fiche-rappel/23044/Interne"
        self.assertEqual(_norm_echo_url(a), _norm_echo_url(b))

    def test_different_fiche_ids_never_collide(self):
        self.assertNotEqual(_norm_echo_url(FICHE.format(23044)),
                            _norm_echo_url(FICHE.format(23045)))

    def test_different_nature_slug_never_collides(self):
        """Interne vs Externe is a different page, not cosmetic drift."""
        self.assertNotEqual(
            _norm_echo_url("https://rappel.conso.gouv.fr/fiche-rappel/23008/Interne"),
            _norm_echo_url("https://rappel.conso.gouv.fr/fiche-rappel/23008/Externe"))

    def test_different_host_never_collides(self):
        self.assertNotEqual(
            _norm_echo_url("https://rappel.conso.gouv.fr/x/1"),
            _norm_echo_url("https://webgate.ec.europa.eu/x/1"))

    def test_path_case_is_significant(self):
        """Only the HOST is case-folded. Regulator paths are case-sensitive."""
        self.assertNotEqual(_norm_echo_url("https://h/A/b"),
                            _norm_echo_url("https://h/a/b"))

    def test_empty_and_none(self):
        self.assertEqual(_norm_echo_url(""), "")
        self.assertEqual(_norm_echo_url(None), "")


class TestIdentityAttributionLogic(unittest.TestCase):
    """The attribution rule itself, exercised on the production shape.

    Mirrors the guard in url_gate_gemini: attribute by echoed URL, re-home a
    mis-indexed decision to the row it actually describes, discard a decision
    that matches no row.
    """

    def setUp(self):
        # A batch shaped like the real one: adjacent FR fiches plus the RASFF
        # row whose Reason bled onto row 740.
        self.chunk = [
            {"URL": FICHE.format(22205)},                       # 0
            {"URL": FICHE.format(22206)},                       # 1
            {"URL": FICHE.format(22208)},                       # 2
            {"URL": "https://webgate.ec.europa.eu/rasff-window/"
                    "screen/notification/841107"},              # 3
        ]

    def _attribute(self, claimed_index, echoed_url):
        """Return the resolved index, or None for 'discard'."""
        echo = _norm_echo_url(echoed_url)
        claimed = _norm_echo_url(self.chunk[claimed_index].get("URL"))
        if not echo:
            return None
        if echo == claimed:
            return claimed_index
        return next((k for k, r in enumerate(self.chunk)
                     if _norm_echo_url(r.get("URL")) == echo), None)

    def test_correct_decision_attributed_unchanged(self):
        self.assertEqual(0, self._attribute(0, FICHE.format(22205)))

    def test_off_by_one_shift_is_rehomed_not_written_to_the_wrong_row(self):
        """THE production bug: the model claims index 0 but is describing the
        row at index 3. Pre-fix this wrote RASFF content onto fiche 22205."""
        resolved = self._attribute(0, self.chunk[3]["URL"])
        self.assertEqual(3, resolved)
        self.assertNotEqual(0, resolved)

    def test_whole_batch_shifted_by_one_is_fully_recovered(self):
        for claimed in range(len(self.chunk) - 1):
            actual = claimed + 1
            self.assertEqual(
                actual, self._attribute(claimed, self.chunk[actual]["URL"]),
                f"claimed={claimed}")

    def test_echo_matching_nothing_is_discarded(self):
        self.assertIsNone(
            self._attribute(0, FICHE.format(99999)),
            "a decision that matches no row in the batch must be discarded, "
            "never written to the row it claims")

    def test_missing_echo_is_discarded(self):
        self.assertIsNone(self._attribute(0, None))
        self.assertIsNone(self._attribute(0, ""))


class TestPromptDemandsTheEcho(unittest.TestCase):
    """The guard is useless if the model is never asked for the echo."""

    def test_prompt_contains_url_echo_contract(self):
        from pipeline import url_gate_gemini as ug
        src = Path(ug.__file__).read_text(encoding="utf-8")
        self.assertIn('"url_echo"', src)
        self.assertIn("IDENTITY RULE", src)

    def test_batch_view_sends_the_url(self):
        from pipeline import url_gate_gemini as ug
        src = Path(ug.__file__).read_text(encoding="utf-8")
        self.assertIn('"URL":      r.get("URL", "")', src)


class TestWriterNormalisesClass(unittest.TestCase):
    """B: the writer is the one gate that cannot be bypassed."""

    def test_normaliser_maps_the_production_values(self):
        self.assertEqual(
            "Voluntary",
            _normalize_class_language("volontaire (sans arrêté préfectoral)"))
        self.assertEqual(
            "Mandatory",
            _normalize_class_language("imposé par arrêté préfectoral"))

    def test_write_sheet_normalises_raw_french_class(self):
        from openpyxl import Workbook
        wb = Workbook()
        rows = [
            {"Date": "2026-07-28", "Source": "RappelConso (FR)",
             "Company": "E.Leclerc Outreau", "Product": "x",
             "Pathogen": "Listeria monocytogenes",
             "Class": "imposé par arrêté préfectoral",
             "URL": FICHE.format(23008), "Tier": 1, "Outbreak": 0},
            {"Date": "2026-07-28", "Source": "RappelConso (FR)",
             "Company": "Bienheureux", "Product": "y",
             "Pathogen": "Listeria monocytogenes",
             "Class": "volontaire (sans arrêté préfectoral)",
             "URL": FICHE.format(22980), "Tier": 1, "Outbreak": 0},
        ]
        _write_sheet(wb, "Recalls", SCHEMA, rows)
        ws = wb["Recalls"]
        col = SCHEMA.index("Class") + 1
        written = [ws.cell(r, col).value for r in (2, 3)]
        self.assertEqual(["Mandatory", "Voluntary"], written,
                         "raw French Class reached the sheet — the writer "
                         "guard is not running")

    def test_workbook_has_no_raw_french_class(self):
        try:
            import openpyxl
        except ImportError:                       # pragma: no cover
            self.skipTest("openpyxl not installed")
        xlsx = ROOT / "docs" / "data" / "recalls.xlsx"
        if not xlsx.exists():                     # pragma: no cover
            self.skipTest("recalls.xlsx not present")
        wb = openpyxl.load_workbook(xlsx, read_only=True)
        offenders = []
        for sheet in wb.sheetnames:
            rows = list(wb[sheet].values)
            if not rows or "Class" not in [str(h) for h in rows[0]]:
                continue
            ic = [str(h) for h in rows[0]].index("Class")
            for n, r in enumerate(rows[1:], 2):
                if not r:
                    continue
                raw = str(r[ic] or "")
                if raw and _normalize_class_language(raw) != raw:
                    offenders.append((sheet, n, raw))
        self.assertEqual([], offenders,
                         f"un-normalised Class values in the workbook: "
                         f"{offenders[:10]}")


if __name__ == "__main__":
    unittest.main(verbosity=2)
