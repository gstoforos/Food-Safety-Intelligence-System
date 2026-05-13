"""
tests/test_xlsx_merge.py
=========================
End-to-end tests for pipeline.xlsx_merge.merge_xlsx_with_remote.

This is the function that runs inside every push-retry block when a
GitHub Actions push hits a non-fast-forward. It takes 3 paths:

    remote_path  — the just-pulled remote recalls.xlsx (what we lost to)
    ours_path    — our local recalls.xlsx (what we tried to push)
    out_path     — destination for the merged result

And produces a workbook that is the UNION of both sides, with these
invariants:

    1. Recalls — UNION by dedup_key. Never shrinks. Remote wins on tie.
    2. Pending — UNION, then strip any row whose dedup_key already
                 appears in the merged Recalls (Pending is the staging
                 area, Recalls is the published area; if a row landed
                 in Recalls remotely, our local Pending row is stale).
    3. NEWS    — UNION by news dedup_key (Link or Published+Title).
    4. Weekly_Review — UNION by (URL, Date). Pre-2026-05-06 this sheet
                 was silently dropped on every push-retry. Audit fixed.
    5. Weekly_Rejected — UNION by (URL, Date). Pre-2026-05-12 this sheet
                 was silently dropped on every push-retry. Same bug class
                 as the 2026-05-06 Weekly_Review wipe, fixed 5 days late.

The Weekly_Rejected bug (2026-05-12) is the highest-blast-radius
regression I want this test file to PREVENT FOREVER:

  - claude_check writes 30 rejected rows to Weekly_Rejected in run A
  - 30 seconds later news-feed runs, hits a push conflict, falls back
    to xlsx_merge — and silently destroys the Weekly_Rejected sheet
  - Operator's Thursday review email has no rejected rows to show
  - Bug discovered only after the fact ("where did the rejects go?")

If the test_weekly_rejected_survives_merge test below ever fails, we
have a regression of that exact bug. CI must fail loudly.

Marked @pytest.mark.slow because these tests do filesystem I/O via
openpyxl (~50–100ms each). Use `pytest -m "not slow"` to skip during
fast TDD loops.
"""
from __future__ import annotations

import pytest
from openpyxl import load_workbook

from pipeline.xlsx_merge import merge_xlsx_with_remote

# Pull schemas from conftest — they match the live pipeline schemas.
from tests.conftest import (
    RECALLS_COLS,
    PENDING_COLS,
    WEEKLY_REVIEW_COLS,
    WEEKLY_REJECTED_COLS,
    NEWS_COLS,
)


# ───────────────────────────────────────────────────────────────────────
# Helpers
# ───────────────────────────────────────────────────────────────────────
def _read_sheet_rows(xlsx_path, sheet_name):
    """Return list of dicts representing rows in the given sheet."""
    wb = load_workbook(xlsx_path)
    if sheet_name not in wb.sheetnames:
        return None  # sheet doesn't exist (distinct from "exists but empty")
    ws = wb[sheet_name]
    if ws.max_row < 2:
        return []
    headers = [str(c.value or "") for c in ws[1]]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        rows.append({h: r[i] for i, h in enumerate(headers)})
    return rows


pytestmark = pytest.mark.slow


# ───────────────────────────────────────────────────────────────────────
# Recalls — UNION, never shrinks
# ───────────────────────────────────────────────────────────────────────
class TestRecallsUnion:
    """The main invariant: Recalls only grows, never shrinks."""

    def test_remote_only_rows_preserved(self, tmp_xlsx_factory, tmp_path,
                                         sample_recall_row):
        """Rows that exist only in remote must survive the merge."""
        remote_row = {**sample_recall_row, "URL": "https://example.com/r-only"}
        ours_row   = {**sample_recall_row, "URL": "https://example.com/o-only"}

        remote = tmp_xlsx_factory("remote.xlsx", {
            "Recalls": (RECALLS_COLS, [remote_row]),
            "Pending": (PENDING_COLS, []),
            "NEWS":    (NEWS_COLS, []),
        })
        ours = tmp_xlsx_factory("ours.xlsx", {
            "Recalls": (RECALLS_COLS, [ours_row]),
            "Pending": (PENDING_COLS, []),
            "NEWS":    (NEWS_COLS, []),
        })
        out = tmp_path / "out.xlsx"

        counts = merge_xlsx_with_remote(remote, ours, out)

        merged_recalls = _read_sheet_rows(out, "Recalls")
        urls = {r["URL"] for r in merged_recalls}
        assert "https://example.com/r-only" in urls
        assert "https://example.com/o-only" in urls
        assert counts["recalls_merged"] == 2

    def test_recalls_never_shrinks(self, tmp_xlsx_factory, tmp_path,
                                    sample_recall_row):
        """Even if ours has FEWER recalls than remote, the merge must
        preserve every remote row."""
        remote_rows = [
            {**sample_recall_row, "URL": f"https://example.com/r{i}"}
            for i in range(5)
        ]
        ours_rows = [
            {**sample_recall_row, "URL": "https://example.com/r0"},
        ]  # only one row, but it overlaps with remote
        remote = tmp_xlsx_factory("remote.xlsx", {
            "Recalls": (RECALLS_COLS, remote_rows),
            "Pending": (PENDING_COLS, []),
            "NEWS":    (NEWS_COLS, []),
        })
        ours = tmp_xlsx_factory("ours.xlsx", {
            "Recalls": (RECALLS_COLS, ours_rows),
            "Pending": (PENDING_COLS, []),
            "NEWS":    (NEWS_COLS, []),
        })
        out = tmp_path / "out.xlsx"
        merge_xlsx_with_remote(remote, ours, out)

        merged = _read_sheet_rows(out, "Recalls")
        # Must contain all 5 remote rows (the 1 overlap dedupes to 1).
        urls = {r["URL"] for r in merged}
        assert len(urls) == 5
        for i in range(5):
            assert f"https://example.com/r{i}" in urls

    def test_collision_remote_wins(self, tmp_xlsx_factory, tmp_path,
                                    sample_recall_row):
        """When both sides have the same dedup_key, REMOTE's data wins.
        This matters because remote already passed gating and was
        committed by another writer; ours may be a stale view."""
        url = "https://example.com/collision"
        remote_row = {**sample_recall_row, "URL": url,
                      "Company": "Remote Wins Co"}
        ours_row   = {**sample_recall_row, "URL": url,
                      "Company": "Ours Loses Co"}
        remote = tmp_xlsx_factory("remote.xlsx", {
            "Recalls": (RECALLS_COLS, [remote_row]),
            "Pending": (PENDING_COLS, []),
            "NEWS":    (NEWS_COLS, []),
        })
        ours = tmp_xlsx_factory("ours.xlsx", {
            "Recalls": (RECALLS_COLS, [ours_row]),
            "Pending": (PENDING_COLS, []),
            "NEWS":    (NEWS_COLS, []),
        })
        out = tmp_path / "out.xlsx"
        merge_xlsx_with_remote(remote, ours, out)

        merged = _read_sheet_rows(out, "Recalls")
        assert len(merged) == 1
        assert merged[0]["Company"] == "Remote Wins Co", \
            "Remote must win on Recalls collision"


# ───────────────────────────────────────────────────────────────────────
# Pending — UNION, but stripped of rows already in Recalls
# ───────────────────────────────────────────────────────────────────────
class TestPendingStrip:
    """Pending is staging; if a row landed in Recalls it must vanish
    from Pending. Tests the 'no shadow' invariant."""

    def test_pending_row_promoted_to_recalls_removed_from_pending(
        self, tmp_xlsx_factory, tmp_path, sample_recall_row, sample_pending_row
    ):
        """
        Scenario: remote promoted a row from Pending → Recalls. Our
        local copy still has the row in Pending. After merge, the
        Pending entry must be gone (it's been published, so it doesn't
        belong in staging anymore).
        """
        url = "https://example.com/promoted"
        remote_recall = {**sample_recall_row, "URL": url}
        ours_pending  = {**sample_pending_row, "URL": url,
                         "Status": "pending"}
        remote = tmp_xlsx_factory("remote.xlsx", {
            "Recalls": (RECALLS_COLS, [remote_recall]),
            "Pending": (PENDING_COLS, []),
            "NEWS":    (NEWS_COLS, []),
        })
        ours = tmp_xlsx_factory("ours.xlsx", {
            "Recalls": (RECALLS_COLS, []),
            "Pending": (PENDING_COLS, [ours_pending]),
            "NEWS":    (NEWS_COLS, []),
        })
        out = tmp_path / "out.xlsx"
        merge_xlsx_with_remote(remote, ours, out)

        merged_pending = _read_sheet_rows(out, "Pending")
        # The row that's now in Recalls must have been stripped from
        # Pending. Empty Pending sheet OR a Pending sheet not containing
        # this URL both satisfy.
        assert all(r.get("URL") != url for r in (merged_pending or []))

    def test_genuine_pending_rows_preserved(
        self, tmp_xlsx_factory, tmp_path, sample_pending_row
    ):
        """Rows that exist ONLY in Pending (not in Recalls) must survive."""
        ours_pending = {**sample_pending_row,
                        "URL": "https://example.com/still-pending"}
        remote = tmp_xlsx_factory("remote.xlsx", {
            "Recalls": (RECALLS_COLS, []),
            "Pending": (PENDING_COLS, []),
            "NEWS":    (NEWS_COLS, []),
        })
        ours = tmp_xlsx_factory("ours.xlsx", {
            "Recalls": (RECALLS_COLS, []),
            "Pending": (PENDING_COLS, [ours_pending]),
            "NEWS":    (NEWS_COLS, []),
        })
        out = tmp_path / "out.xlsx"
        merge_xlsx_with_remote(remote, ours, out)

        merged_pending = _read_sheet_rows(out, "Pending")
        urls = {r.get("URL") for r in (merged_pending or [])}
        assert "https://example.com/still-pending" in urls


# ───────────────────────────────────────────────────────────────────────
# Weekly_Review — UNION, MUST survive (May 6 bug)
# ───────────────────────────────────────────────────────────────────────
class TestWeeklyReviewSurvives:
    """
    THE MAY 6 BUG.

    Pre-2026-05-06, merge_xlsx_with_remote created a fresh workbook with
    only Recalls/Pending/NEWS sheets. Weekly_Review (if it existed on
    either side) was silently destroyed.

    Fix (2026-05-06): Weekly_Review is union'd by (URL, Date) and written
    back if either side had data. These tests lock that behavior.
    """

    def test_remote_weekly_review_survives(
        self, tmp_xlsx_factory, tmp_path, sample_weekly_review_row
    ):
        """A Weekly_Review row on the remote side must reach the output."""
        remote = tmp_xlsx_factory("remote.xlsx", {
            "Recalls":       (RECALLS_COLS, []),
            "Pending":       (PENDING_COLS, []),
            "Weekly_Review": (WEEKLY_REVIEW_COLS, [sample_weekly_review_row]),
            "NEWS":          (NEWS_COLS, []),
        })
        # Ours has NO Weekly_Review sheet at all — this is the bug scenario
        # (most non-mailer pipelines don't touch Weekly_Review).
        ours = tmp_xlsx_factory("ours.xlsx", {
            "Recalls": (RECALLS_COLS, []),
            "Pending": (PENDING_COLS, []),
            "NEWS":    (NEWS_COLS, []),
        })
        out = tmp_path / "out.xlsx"
        merge_xlsx_with_remote(remote, ours, out)

        merged = _read_sheet_rows(out, "Weekly_Review")
        assert merged is not None, \
            "BUG REGRESSION: Weekly_Review sheet was dropped on merge"
        assert len(merged) == 1
        assert merged[0]["URL"] == sample_weekly_review_row["URL"]

    def test_ours_weekly_review_survives(
        self, tmp_xlsx_factory, tmp_path, sample_weekly_review_row
    ):
        """A Weekly_Review row on OUR side must also survive (the merge
        was previously asymmetric — only remote's sheet was kept)."""
        remote = tmp_xlsx_factory("remote.xlsx", {
            "Recalls": (RECALLS_COLS, []),
            "Pending": (PENDING_COLS, []),
            "NEWS":    (NEWS_COLS, []),
        })
        ours = tmp_xlsx_factory("ours.xlsx", {
            "Recalls":       (RECALLS_COLS, []),
            "Pending":       (PENDING_COLS, []),
            "Weekly_Review": (WEEKLY_REVIEW_COLS, [sample_weekly_review_row]),
            "NEWS":          (NEWS_COLS, []),
        })
        out = tmp_path / "out.xlsx"
        merge_xlsx_with_remote(remote, ours, out)

        merged = _read_sheet_rows(out, "Weekly_Review")
        assert merged is not None
        assert len(merged) == 1

    def test_weekly_review_union(
        self, tmp_xlsx_factory, tmp_path, sample_weekly_review_row
    ):
        """Both sides contribute different rows → both must be in output."""
        remote_row = {**sample_weekly_review_row,
                      "URL": "https://example.com/wr-remote"}
        ours_row   = {**sample_weekly_review_row,
                      "URL": "https://example.com/wr-ours"}
        remote = tmp_xlsx_factory("remote.xlsx", {
            "Recalls":       (RECALLS_COLS, []),
            "Pending":       (PENDING_COLS, []),
            "Weekly_Review": (WEEKLY_REVIEW_COLS, [remote_row]),
            "NEWS":          (NEWS_COLS, []),
        })
        ours = tmp_xlsx_factory("ours.xlsx", {
            "Recalls":       (RECALLS_COLS, []),
            "Pending":       (PENDING_COLS, []),
            "Weekly_Review": (WEEKLY_REVIEW_COLS, [ours_row]),
            "NEWS":          (NEWS_COLS, []),
        })
        out = tmp_path / "out.xlsx"
        merge_xlsx_with_remote(remote, ours, out)

        merged = _read_sheet_rows(out, "Weekly_Review")
        urls = {r["URL"] for r in merged}
        assert "https://example.com/wr-remote" in urls
        assert "https://example.com/wr-ours" in urls


# ───────────────────────────────────────────────────────────────────────
# Weekly_Rejected — UNION, MUST survive (May 12 bug)
# ───────────────────────────────────────────────────────────────────────
class TestWeeklyRejectedSurvives:
    """
    THE MAY 12 BUG.

    Same bug class as Weekly_Review — Weekly_Rejected was silently
    dropped on every push-retry conflict for 5 days after Weekly_Rejected
    was added to the pipeline (audit 2026-05-09 introduced the sheet,
    audit 2026-05-12 fixed the wipe). Found in production when the
    operator's Thursday review email showed 0 rejected rows despite
    claude_check having written 30 rejections that morning.

    These tests are the regression guard. If any of them fail, the
    xlsx_merge has reverted to the pre-2026-05-12 behavior.
    """

    def test_remote_weekly_rejected_survives(
        self, tmp_xlsx_factory, tmp_path, sample_weekly_rejected_row
    ):
        """The BUG: 30 rejected rows from claude_check vanish on next
        news-feed push-retry. This test FAILS LOUDLY if it ever happens
        again."""
        remote = tmp_xlsx_factory("remote.xlsx", {
            "Recalls":          (RECALLS_COLS, []),
            "Pending":          (PENDING_COLS, []),
            "Weekly_Rejected":  (WEEKLY_REJECTED_COLS, [sample_weekly_rejected_row]),
            "NEWS":             (NEWS_COLS, []),
        })
        # Ours has NO Weekly_Rejected sheet — typical when our local pipeline
        # is news-feed (which doesn't touch claude_check rejections).
        ours = tmp_xlsx_factory("ours.xlsx", {
            "Recalls": (RECALLS_COLS, []),
            "Pending": (PENDING_COLS, []),
            "NEWS":    (NEWS_COLS, []),
        })
        out = tmp_path / "out.xlsx"
        merge_xlsx_with_remote(remote, ours, out)

        merged = _read_sheet_rows(out, "Weekly_Rejected")
        assert merged is not None, \
            "BUG REGRESSION (2026-05-12): Weekly_Rejected sheet dropped on merge"
        assert len(merged) == 1
        assert merged[0]["URL"] == sample_weekly_rejected_row["URL"]

    def test_ours_weekly_rejected_survives(
        self, tmp_xlsx_factory, tmp_path, sample_weekly_rejected_row
    ):
        """And the reverse: rows ON our side must also survive."""
        remote = tmp_xlsx_factory("remote.xlsx", {
            "Recalls": (RECALLS_COLS, []),
            "Pending": (PENDING_COLS, []),
            "NEWS":    (NEWS_COLS, []),
        })
        ours = tmp_xlsx_factory("ours.xlsx", {
            "Recalls":         (RECALLS_COLS, []),
            "Pending":         (PENDING_COLS, []),
            "Weekly_Rejected": (WEEKLY_REJECTED_COLS, [sample_weekly_rejected_row]),
            "NEWS":            (NEWS_COLS, []),
        })
        out = tmp_path / "out.xlsx"
        merge_xlsx_with_remote(remote, ours, out)

        merged = _read_sheet_rows(out, "Weekly_Rejected")
        assert merged is not None
        assert len(merged) == 1

    def test_weekly_rejected_union(
        self, tmp_xlsx_factory, tmp_path, sample_weekly_rejected_row
    ):
        """Both sides contribute → both rows survive."""
        remote_row = {**sample_weekly_rejected_row,
                      "URL": "https://example.com/rej-remote"}
        ours_row   = {**sample_weekly_rejected_row,
                      "URL": "https://example.com/rej-ours"}
        remote = tmp_xlsx_factory("remote.xlsx", {
            "Recalls":         (RECALLS_COLS, []),
            "Pending":         (PENDING_COLS, []),
            "Weekly_Rejected": (WEEKLY_REJECTED_COLS, [remote_row]),
            "NEWS":            (NEWS_COLS, []),
        })
        ours = tmp_xlsx_factory("ours.xlsx", {
            "Recalls":         (RECALLS_COLS, []),
            "Pending":         (PENDING_COLS, []),
            "Weekly_Rejected": (WEEKLY_REJECTED_COLS, [ours_row]),
            "NEWS":            (NEWS_COLS, []),
        })
        out = tmp_path / "out.xlsx"
        merge_xlsx_with_remote(remote, ours, out)

        merged = _read_sheet_rows(out, "Weekly_Rejected")
        urls = {r["URL"] for r in merged}
        assert "https://example.com/rej-remote" in urls
        assert "https://example.com/rej-ours" in urls


# ───────────────────────────────────────────────────────────────────────
# NEWS — UNION by Link
# ───────────────────────────────────────────────────────────────────────
class TestNewsUnion:
    """NEWS sheet uses Link (or Title+date) as the dedup key."""

    def test_news_union(self, tmp_xlsx_factory, tmp_path, sample_news_row):
        remote_row = {**sample_news_row, "Link": "https://news.example.com/a"}
        ours_row   = {**sample_news_row, "Link": "https://news.example.com/b"}
        remote = tmp_xlsx_factory("remote.xlsx", {
            "Recalls": (RECALLS_COLS, []),
            "Pending": (PENDING_COLS, []),
            "NEWS":    (NEWS_COLS, [remote_row]),
        })
        ours = tmp_xlsx_factory("ours.xlsx", {
            "Recalls": (RECALLS_COLS, []),
            "Pending": (PENDING_COLS, []),
            "NEWS":    (NEWS_COLS, [ours_row]),
        })
        out = tmp_path / "out.xlsx"
        merge_xlsx_with_remote(remote, ours, out)

        merged = _read_sheet_rows(out, "NEWS")
        links = {r["Link"] for r in merged}
        assert "https://news.example.com/a" in links
        assert "https://news.example.com/b" in links


# ───────────────────────────────────────────────────────────────────────
# Full counts — log-correctness sanity
# ───────────────────────────────────────────────────────────────────────
class TestMergeCountsReporting:
    """The counts dict is logged on every merge; it's also the operator's
    only visibility into the merge's behavior. Lock the shape."""

    def test_counts_dict_contains_all_expected_keys(
        self, tmp_xlsx_factory, tmp_path,
        sample_recall_row, sample_weekly_review_row, sample_weekly_rejected_row
    ):
        remote = tmp_xlsx_factory("remote.xlsx", {
            "Recalls":         (RECALLS_COLS, [sample_recall_row]),
            "Pending":         (PENDING_COLS, []),
            "Weekly_Review":   (WEEKLY_REVIEW_COLS, [sample_weekly_review_row]),
            "Weekly_Rejected": (WEEKLY_REJECTED_COLS, [sample_weekly_rejected_row]),
            "NEWS":            (NEWS_COLS, []),
        })
        ours = tmp_xlsx_factory("ours.xlsx", {
            "Recalls": (RECALLS_COLS, []),
            "Pending": (PENDING_COLS, []),
            "NEWS":    (NEWS_COLS, []),
        })
        out = tmp_path / "out.xlsx"
        counts = merge_xlsx_with_remote(remote, ours, out)

        expected_keys = {
            "recalls_remote", "recalls_ours", "recalls_merged",
            "pending_remote", "pending_ours", "pending_merged",
            "weekly_review_remote", "weekly_review_ours", "weekly_review_merged",
            "weekly_rejected_remote", "weekly_rejected_ours", "weekly_rejected_merged",
            "news_remote", "news_ours", "news_merged",
        }
        actual_keys = set(counts.keys())
        missing = expected_keys - actual_keys
        assert not missing, f"counts dict missing keys: {missing}"
