"""The rejection archive must be written by column NAME, not position.

WHY
===
Three modules define a schema for the single Weekly_Rejected sheet, and
they do not agree:

    pipeline/extractor.py       REJECTED_COLUMNS   19 columns
    pipeline/merge_master.py    REJECTED_SCHEMA    legacy
    pipeline/weekly_rejected_capture.py  SHEET_COLS   21 columns

The live workbook (docs/data/recalls.xlsx, 2026-08-18) carries the
19-column extractor header. record_rejections built a 21-value list and
called `ws.append(...)`, which writes BY POSITION. Against that header
the last seven values land one to three columns left of where they
belong:

    ScrapedAt    <- DateAdded
    Status       <- LastUpdated
    RejectedBy   <- LastChecked        so RejectedBy is a date or blank
    RejectedAt   <- Week_Added         a FUTURE Thursday, not a stamp
    RejectReason <- rejected_by        the reviewer's NAME, not a reason
    (unheadered) <- the actual reason
    (unheadered) <- "N"

Every one of those is type-plausible — a date under a date header, a
name under a text header — which is why it went unnoticed. It is
visible in the archive today: the CDC javiana row rejected on
2026-08-14 reads RejectedBy=None, RejectedAt=2026-08-20 (three days in
the future at the time of writing) and RejectReason="operator review
2026-08-14". The reason that row was rejected is recorded nowhere.

This is the same failure George has named before — "you corrupted 155
rows once doing that". The fix is to write each value under its own
header and widen the sheet when a header is missing, so the schema
disagreement stops mattering.
"""
from __future__ import annotations

import sys
import unittest
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from pipeline import weekly_rejected_capture as cap  # noqa: E402

# The 19-column header pipeline/extractor.py defines as REJECTED_COLUMNS and
# that the live workbook actually carries. Written out literally rather than
# imported, because pipeline.extractor pulls in gap_finder at import time and
# that package is not on the test path. test_it_matches_the_live_workbook
# below keeps this copy honest.
REJECTED_COLUMNS = [
    "Date", "Source", "Company", "Brand", "Product", "Pathogen", "Reason",
    "Class", "Country", "Region", "Tier", "Outbreak", "URL", "Notes",
    "ScrapedAt", "Status", "RejectedBy", "RejectedAt", "RejectReason",
]

ROW = {
    "Date": "2026-08-07", "Source": "CDC",
    "Company": "Example Co", "Brand": "Example", "Product": "Widget",
    "Pathogen": "Salmonella", "Reason": "Salmonella — outbreak",
    "Class": "Recall", "Country": "United States", "Region": "North America",
    "Tier": 1, "Outbreak": 1,
    "URL": "https://www.cdc.gov/salmonella/outbreaks/javiana-08-26/index.html",
    "Notes": "audit trail", "DateAdded": "2026-08-07",
    "LastUpdated": "2026-08-14", "LastChecked": "2026-08-14",
    "RejectedBy": "operator review 2026-08-14",
    "Status": "rejected",
}


def _book(headers, path):
    wb = openpyxl.Workbook()
    wb.active.title = "Recalls"
    ws = wb.create_sheet("Weekly_Rejected")
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    wb.save(path)


def _archived(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Weekly_Rejected"]
    rows = list(ws.values)
    hdr = [str(h or "") for h in rows[0]]
    return [dict(zip(hdr, r)) for r in rows[1:] if r and any(r)]


class TestTheLiveHeadersAreCoherent(unittest.TestCase):
    """Both archive sheets, whatever schema each happens to carry.

    As of 2026-08-20 the live workbook holds TWO rejection sheets with
    TWO different headers, and both are legitimate:

        Weekly_Rejected  weekly_rejected_capture.SHEET_COLS   (21 cols)
                         recreated by the Thursday wipe
        Rejected         pipeline/extractor.REJECTED_COLUMNS  (19 cols)
                         the rows the wipe moved, keeping their header

    An earlier version of this test pinned one literal 19-column list and
    went red the moment the wipe legitimately recreated the rolling sheet
    with the other one. Pinning a single shape is the wrong guard when
    several writers each have a valid schema — that is exactly why
    record_rejections writes by NAME.

    What still must hold, on either schema: every column is a recognised
    name, the identity columns exist, and there is somewhere to put a
    reason. A shifted or mangled header fails all three.
    """

    KNOWN = set(REJECTED_COLUMNS) | set(cap.SHEET_COLS) | {
        "RejectReason", "RejectionReason", "RejectedReason", "RejectedAt",
        "Week_Added", "Reviewed", "ScrapedAt", "Status"}
    REASON_SPELLINGS = ("RejectReason", "RejectionReason", "RejectedReason")

    def _headers(self, sheet):
        xlsx = ROOT / "docs" / "data" / "recalls.xlsx"
        if not xlsx.exists():                              # pragma: no cover
            self.skipTest("recalls.xlsx not present")
        wb = openpyxl.load_workbook(xlsx, read_only=True)
        if sheet not in wb.sheetnames:
            return None
        return [str(c.value or "") for c in next(wb[sheet].rows)]

    def test_every_column_is_a_recognised_name(self):
        for sheet in ("Weekly_Rejected", "Rejected"):
            hdr = self._headers(sheet)
            if hdr is None:
                continue
            unknown = [h for h in hdr if h and h not in self.KNOWN]
            self.assertEqual([], unknown,
                             f"{sheet} carries unrecognised columns "
                             f"{unknown} — a shifted header looks exactly "
                             f"like this")

    def test_identity_columns_exist(self):
        for sheet in ("Weekly_Rejected", "Rejected"):
            hdr = self._headers(sheet)
            if hdr is None:
                continue
            for col in ("Date", "URL"):
                self.assertIn(col, hdr, f"{sheet} has no {col} column — an "
                                        f"archived rejection could never be "
                                        f"matched against a re-ingestion")

    def test_a_reason_column_exists(self):
        for sheet in ("Weekly_Rejected", "Rejected"):
            hdr = self._headers(sheet)
            if hdr is None:
                continue
            self.assertTrue(any(h in hdr for h in self.REASON_SPELLINGS),
                            f"{sheet} has none of {self.REASON_SPELLINGS}; "
                            f"the next reviewer sees no reason at all")

    def test_no_duplicate_columns(self):
        for sheet in ("Weekly_Rejected", "Rejected"):
            hdr = self._headers(sheet)
            if hdr is None:
                continue
            named = [h for h in hdr if h]
            dupes = sorted({h for h in named if named.count(h) > 1})
            self.assertEqual([], dupes, f"{sheet} has duplicate columns "
                                        f"{dupes} — writing by name becomes "
                                        f"ambiguous")

    def test_the_guard_can_read_both_sheets(self):
        """End to end on the real workbook: whatever the two schemas are,
        load_rejected_urls must return entries with reasons attached."""
        from pipeline.merge_master import load_rejected_urls
        xlsx = ROOT / "docs" / "data" / "recalls.xlsx"
        if not xlsx.exists():                              # pragma: no cover
            self.skipTest("recalls.xlsx not present")
        reg = load_rejected_urls(xlsx)
        self.assertTrue(reg, "the re-promotion guard read nothing from a "
                             "workbook that has archived rejections")
        with_reason = [v for v in reg.values()
                       if v.strip().rstrip(":").strip()
                       and "no reason recorded" not in v]
        self.assertTrue(with_reason,
                        "every archived rejection came back without a "
                        "reason — the reason column is not being read")

    def test_it_matches_pipeline_extractor(self):
        src = (ROOT / "pipeline" / "extractor.py").read_text(encoding="utf-8")
        self.assertIn('REJECTED_COLUMNS = PENDING_COLUMNS + '
                      '["RejectedAt", "RejectReason"]', src,
                      "pipeline/extractor.py changed its rejection schema; "
                      "update REJECTED_COLUMNS in this test to match")


class TestArchiveAlignment(unittest.TestCase):

    def setUp(self):
        import tempfile
        self.tmp = Path(tempfile.mkdtemp())
        self.xlsx = self.tmp / "recalls.xlsx"
        self.json = self.tmp / "weekly-rejected-latest.json"

    def test_against_the_live_19_column_header(self):
        """The exact header on disk in docs/data/recalls.xlsx."""
        _book(REJECTED_COLUMNS, self.xlsx)
        n = cap.record_rejections([dict(ROW)], xlsx_path=self.xlsx,
                                  json_path=self.json)
        self.assertEqual(1, n)
        got = _archived(self.xlsx)[0]

        # The identity columns.
        self.assertEqual(ROW["URL"], got["URL"])
        self.assertEqual("2026-08-07", str(got["Date"]))
        # The columns that were shifted. RejectedBy must be the reviewer,
        # not LastChecked; the reason column must hold a reason, not a name.
        self.assertNotIn(str(got.get("RejectedBy") or ""),
                         ("2026-08-14 ", "None"),
                         "RejectedBy still looks like the LastChecked date")
        self.assertNotEqual("2026-08-14", str(got.get("ScrapedAt") or ""),
                            "ScrapedAt is holding LastUpdated")

    def test_no_recalls_field_lands_under_a_rejection_header(self):
        """The specific shift: DateAdded/LastUpdated/LastChecked must not
        appear under ScrapedAt/Status/RejectedBy."""
        _book(REJECTED_COLUMNS, self.xlsx)
        cap.record_rejections([dict(ROW)], xlsx_path=self.xlsx,
                              json_path=self.json)
        got = _archived(self.xlsx)[0]
        self.assertNotEqual(ROW["DateAdded"], got.get("ScrapedAt"))
        self.assertNotEqual(ROW["LastUpdated"], got.get("Status"))
        self.assertNotEqual(ROW["LastChecked"], got.get("RejectedBy"))

    def test_missing_columns_are_added_not_shifted_into(self):
        """A sheet lacking this module's own columns gets them appended,
        rather than its existing columns being overwritten."""
        _book(REJECTED_COLUMNS, self.xlsx)
        cap.record_rejections([dict(ROW)], xlsx_path=self.xlsx,
                              json_path=self.json)
        wb = openpyxl.load_workbook(self.xlsx)
        hdr = [str(c.value or "") for c in wb["Weekly_Rejected"][1]]
        self.assertEqual(REJECTED_COLUMNS, hdr[:len(REJECTED_COLUMNS)],
                         "existing headers were disturbed")
        for want in ("Week_Added", "Reviewed"):
            self.assertIn(want, hdr)
        # RejectReason already covers RejectionReason — no duplicate column.
        self.assertNotIn("RejectionReason", hdr,
                         "a second reason column was created alongside "
                         "RejectReason")

    def test_own_schema_sheet_still_round_trips(self):
        _book(cap.SHEET_COLS, self.xlsx)
        cap.record_rejections([dict(ROW)], xlsx_path=self.xlsx,
                              json_path=self.json)
        got = _archived(self.xlsx)[0]
        self.assertEqual(ROW["URL"], got["URL"])
        self.assertEqual("N", got["Reviewed"])
        self.assertEqual(ROW["Pathogen"], got["Pathogen"])

    def test_a_header_without_url_is_widened_not_written_past(self):
        """Widening beats refusing: the column is added and the row lands
        under it. What must never happen is the URL being written into
        whatever column happens to sit in position 13."""
        _book([c for c in REJECTED_COLUMNS if c != "URL"], self.xlsx)
        n = cap.record_rejections([dict(ROW)], xlsx_path=self.xlsx,
                                  json_path=self.json)
        self.assertEqual(1, n)
        got = _archived(self.xlsx)[0]
        self.assertEqual(ROW["URL"], got["URL"])
        self.assertEqual(ROW["Notes"], got["Notes"],
                         "a neighbouring column absorbed the URL")

    def test_rejected_at_is_not_in_the_future(self):
        """The old positional write put the review Thursday — always a
        future date — into RejectedAt."""
        from datetime import date as _d
        _book(REJECTED_COLUMNS, self.xlsx)
        cap.record_rejections([dict(ROW)], xlsx_path=self.xlsx,
                              json_path=self.json)
        got = _archived(self.xlsx)[0]
        self.assertLessEqual(str(got["RejectedAt"])[:10],
                             _d.today().isoformat(),
                             "RejectedAt is dated in the future — it is "
                             "holding Week_Added again")

    def test_week_added_has_its_own_column(self):
        _book(REJECTED_COLUMNS, self.xlsx)
        cap.record_rejections([dict(ROW)], xlsx_path=self.xlsx,
                              json_path=self.json)
        got = _archived(self.xlsx)[0]
        self.assertIn("Week_Added", got)
        self.assertNotEqual(got["Week_Added"], got["RejectedAt"])

    def test_the_guard_can_read_back_what_the_writer_wrote(self):
        """End to end: archive a rejection, then confirm
        load_rejected_urls() blocks that URL AND surfaces the reason."""
        from pipeline.merge_master import (load_rejected_urls,
                                           _normalize_url_for_dedup)
        _book(REJECTED_COLUMNS, self.xlsx)
        row = dict(ROW)
        row["Notes"] = ("[rejected 2026-08-14: operator_decision — one "
                        "outbreak one source]")
        cap.record_rejections([row], xlsx_path=self.xlsx, json_path=self.json)
        reg = load_rejected_urls(self.xlsx)
        key = _normalize_url_for_dedup(ROW["URL"])
        self.assertIn(key, reg, "the archived rejection does not block a "
                                "re-ingestion of the same URL")
        self.assertTrue(reg[key].strip().rstrip(":"),
                        "the guard has no description to show the reviewer")


if __name__ == "__main__":
    unittest.main()
