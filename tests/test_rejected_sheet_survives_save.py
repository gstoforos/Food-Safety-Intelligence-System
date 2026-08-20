"""save_xlsx_with_pending must not delete the permanent Rejected sheet.

WHY (2026-08-20)
================
Two designs, three months apart, were destroying each other:

  2026-05-11  save_xlsx_with_pending gained a "one-shot cleanup" that
              deleted any sheet named "Rejected", because at the time it
              really was duplicate audit infrastructure.
  2026-08-15  tools/wipe_weekly_rejected.py was changed to MOVE rows into
              a permanent "Rejected" sheet on the Thursday reset instead
              of deleting them — the rolling sheet was the pipeline's only
              memory of what had been rejected, and wiping it meant the
              gap-finders re-found every one of those rows within a cycle.
  2026-08-18  load_rejected_urls() started reading BOTH sheets so a
              rejection survives the Thursday reset.

So the Thursday wipe created the sheet and the very next pipeline save
removed it. Measured on the live workbook: the 2026-08-20 wipe archived
138 rows at 14:51; the next save_xlsx_with_pending dropped all 138 and
logged it at INFO as "Removed legacy Rejected sheet". Weekly_Rejected had
just been emptied, so that single call would have taken the entire
rejection history.

The sheets have distinct roles now and both must persist:

    Weekly_Rejected   rolling; emptied every Thursday after the email
    Rejected          permanent; append-only; never wiped

This test exists because the deletion was silent, idempotent and
well-commented — everything except correct.
"""
from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from pipeline.merge_master import (  # noqa: E402
    save_xlsx_with_pending, RECALLS_SCHEMA)

ROW = {"Date": "2026-08-01", "Source": "FDA", "Company": "Example Co",
       "Brand": "Ex", "Product": "Widget", "Pathogen": "Listeria",
       "Reason": "Listeria detected", "Class": "Recall",
       "Country": "United States", "Region": "North America",
       "Tier": 1, "Outbreak": 0, "URL": "https://example.invalid/1",
       "Notes": "", "DateAdded": "2026-08-01",
       "LastUpdated": "2026-08-01", "LastChecked": ""}


class TestRejectedSheetSurvives(unittest.TestCase):

    def setUp(self):
        self.tmp = Path(tempfile.mkdtemp())
        self.xlsx = self.tmp / "recalls.xlsx"
        wb = openpyxl.Workbook()
        wb.active.title = "Recalls"
        for i, h in enumerate(RECALLS_SCHEMA, 1):
            wb.active.cell(row=1, column=i, value=h)
        wb.create_sheet("Pending")
        rej = wb.create_sheet("Rejected")
        rej.append(["Date", "Source", "URL", "RejectReason"])
        for n in range(138):
            rej.append([f"2026-07-{(n % 28) + 1:02d}", "FDA",
                        f"https://example.invalid/rejected/{n}",
                        "operator review"])
        wb.create_sheet("NEWS")
        wb.save(self.xlsx)

    def _rejected_rows(self):
        wb = openpyxl.load_workbook(self.xlsx, read_only=True)
        if "Rejected" not in wb.sheetnames:
            return None
        return len(list(wb["Rejected"].values)) - 1

    def test_the_sheet_is_still_there_after_a_save(self):
        self.assertEqual(138, self._rejected_rows())
        save_xlsx_with_pending([dict(ROW)], [], self.xlsx)
        self.assertIsNotNone(self._rejected_rows(),
                             "save_xlsx_with_pending deleted the permanent "
                             "Rejected sheet")

    def test_every_row_survives(self):
        save_xlsx_with_pending([dict(ROW)], [], self.xlsx)
        self.assertEqual(138, self._rejected_rows(),
                         "rows were lost from the permanent archive")

    def test_repeated_saves_do_not_erode_it(self):
        for _ in range(3):
            save_xlsx_with_pending([dict(ROW)], [], self.xlsx)
        self.assertEqual(138, self._rejected_rows())

    def test_the_deletion_is_gone_from_the_source(self):
        """Pin the specific line, so a future 'cleanup' that reintroduces
        it fails here rather than in production on a Thursday night."""
        src = (ROOT / "pipeline" / "merge_master.py").read_text(
            encoding="utf-8")
        body = src.split("def save_xlsx_with_pending", 1)[1].split(
            "\ndef ", 1)[0]
        # Read CODE, not prose. The fix's own comment quotes the deleted
        # line to explain what was reversed, and a naive substring scan
        # matched that quotation — a guard failing on its own
        # documentation. Strip comment bodies first.
        code = "\n".join(ln.split("#", 1)[0] for ln in body.splitlines())
        self.assertNotIn('del wb["Rejected"]', code,
                         "save_xlsx_with_pending deletes the permanent "
                         "Rejected sheet again")

    def test_the_live_workbook_still_has_it(self):
        xlsx = ROOT / "docs" / "data" / "recalls.xlsx"
        if not xlsx.exists():                              # pragma: no cover
            self.skipTest("recalls.xlsx not present")
        wb = openpyxl.load_workbook(xlsx, read_only=True)
        self.assertIn("Rejected", wb.sheetnames,
                      "the permanent rejection archive is missing from the "
                      "published workbook")


if __name__ == "__main__":
    unittest.main()
