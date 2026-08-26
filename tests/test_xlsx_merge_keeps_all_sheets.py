"""xlsx_merge must not drop a sheet it was not told about.

WHY (traced 2026-08-26)
=======================
merge_xlsx_with_remote rebuilds the workbook with `Workbook()` and creates
exactly Recalls, Pending, Weekly_Review, Weekly_Rejected and NEWS. Any
other sheet in either input is simply absent from the output — no error,
no log line.

The permanent "Rejected" archive is such a sheet. It is created by the
Thursday wipe (tools/wipe_weekly_rejected.py) so a rejection survives the
reset, and read by merge_master.load_rejected_urls so the gap-finders stop
re-finding decisions a human already made.

On the live repo the archive was restored — 138 rows — at 2026-08-22 11:44
and was gone at the very next writer. The commits that killed it carry
"(retry 1, row-merged)": this code path. It had already destroyed the same
138 rows twice before, and each loss was invisible.

Two guards:
  * "Rejected" is merged explicitly, with the same no-shrink assertion the
    other audit sheets get — it is append-only, so a merge producing fewer
    rows than either input is always a bug.
  * anything else unrecognised is carried across and logged, so the next
    sheet somebody adds cannot disappear the same way.
"""
from __future__ import annotations

import shutil
import sys
import tempfile
import unittest
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from pipeline.xlsx_merge import merge_xlsx_with_remote  # noqa: E402
from pipeline.merge_master import RECALLS_SCHEMA        # noqa: E402


def _book(path, rejected_rows=12, extra=None):
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Recalls"
    for i, h in enumerate(RECALLS_SCHEMA, 1):
        ws.cell(row=1, column=i, value=h)
    ws.append(["2026-08-01", "FDA", "Co", "Br", "Prod", "Listeria",
               "Listeria detected", "Recall", "United States",
               "North America", 1, 0, "https://example.invalid/1", "",
               "2026-08-01", "2026-08-01", ""][:len(RECALLS_SCHEMA)])
    wb.create_sheet("Pending").append(["Date", "Source", "URL"])
    rj = wb.create_sheet("Rejected")
    rj.append(["Date", "Source", "URL", "RejectReason"])
    for n in range(rejected_rows):
        rj.append([f"2026-07-{(n % 28) + 1:02d}", "FDA",
                   f"https://example.invalid/rejected/{n}", "operator review"])
    if extra:
        e = wb.create_sheet(extra)
        e.append(["Date", "URL"]); e.append(["2026-08-01", "https://x.invalid/1"])
    wb.create_sheet("NEWS").append(["Published (UTC)", "Title", "Link"])
    wb.save(path)


class TestMergeKeepsSheets(unittest.TestCase):

    def setUp(self):
        self.tmp = Path(tempfile.mkdtemp())
        self.remote = self.tmp / "remote.xlsx"
        self.ours = self.tmp / "ours.xlsx"
        self.out = self.tmp / "out.xlsx"

    def _sheets(self):
        wb = openpyxl.load_workbook(self.out, read_only=True)
        return {s: len(list(wb[s].values)) - 1 for s in wb.sheetnames}

    def test_the_rejected_archive_survives(self):
        _book(self.remote, 12); _book(self.ours, 12)
        merge_xlsx_with_remote(self.remote, self.ours, self.out)
        got = self._sheets()
        self.assertIn("Rejected", got,
                      "the permanent rejection archive was dropped by the merge")
        self.assertEqual(12, got["Rejected"])

    def test_it_never_shrinks(self):
        """One side has more rows than the other — the union must win."""
        _book(self.remote, 20); _book(self.ours, 5)
        merge_xlsx_with_remote(self.remote, self.ours, self.out)
        self.assertGreaterEqual(self._sheets()["Rejected"], 20)

    def test_an_unrecognised_sheet_is_carried_over(self):
        """The next sheet somebody adds must not vanish the same way."""
        _book(self.remote, 3, extra="SomeNewAuditSheet")
        _book(self.ours, 3)
        merge_xlsx_with_remote(self.remote, self.ours, self.out)
        self.assertIn("SomeNewAuditSheet", self._sheets())

    def test_a_workbook_with_no_archive_is_unaffected(self):
        """Not every workbook has one — absence must not create an empty
        sheet or raise."""
        for p in (self.remote, self.ours):
            _book(p, 0)
            wb = openpyxl.load_workbook(p)
            del wb["Rejected"]; wb.save(p)
        merge_xlsx_with_remote(self.remote, self.ours, self.out)
        self.assertNotIn("Rejected", self._sheets())

    def test_the_output_sheet_list_is_not_hardcoded_alone(self):
        """Pin the backstop itself. Without it the next added sheet is
        silently lost, which is the whole defect."""
        src = (ROOT / "pipeline" / "xlsx_merge.py").read_text(encoding="utf-8")
        body = src.split("def merge_xlsx_with_remote", 1)[1]
        self.assertIn("carrying over unrecognised sheet", body)


if __name__ == "__main__":
    unittest.main()
