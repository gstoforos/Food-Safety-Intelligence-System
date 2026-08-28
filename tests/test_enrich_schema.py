"""The analytical columns must exist, be honest, and never leak.

The schema spent a week as a proposal, a prompt and two extractors while
the register still had eighteen columns. This module is the assembler that
joins them, and these tests hold it to three promises: it does not touch
what the pipeline already writes, it does not publish what it adds, and it
does not invent values.
"""
from __future__ import annotations

import json
from pathlib import Path

import pytest

from pipeline import enrich_schema as ES
from pipeline.merge_master import RECALLS_INTERNAL_COLUMNS

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "docs" / "data" / "recalls.xlsx"

PIPELINE_OWNED = ("Date", "Source", "Company", "Brand", "Product", "Pathogen",
                  "Reason", "Class", "Country", "Region", "Tier", "Outbreak",
                  "URL", "Notes", "DateAdded", "LastUpdated", "LastChecked",
                  "report_week")


def test_every_added_column_is_registered_internal():
    """The only thing keeping them out of recalls.json."""
    missing = [c for c in ES.COLUMNS if c not in RECALLS_INTERNAL_COLUMNS]
    assert not missing, (
        f"{missing} would be published to the dashboard — register them in "
        f"merge_master.RECALLS_INTERNAL_COLUMNS")


def test_the_assembler_never_writes_a_pipeline_column():
    overlap = set(ES.COLUMNS) & set(PIPELINE_OWNED)
    assert not overlap, overlap


@pytest.mark.skipif(not XLSX.exists(), reason="no workbook")
def test_nothing_added_reaches_recalls_json():
    p = ROOT / "docs" / "data" / "recalls.json"
    if not p.exists():
        pytest.skip("recalls.json not built")
    rec = json.loads(p.read_text(encoding="utf-8"))[0]
    leaked = set(ES.COLUMNS) & set(rec)
    assert not leaked, f"leaked into the public JSON: {leaked}"


def test_a_hard_field_is_never_reported_at_100_percent():
    """FoodCategory read 100% once because split_cfia_category was handed the
    row instead of the category string, and stringified the whole dict.

    A perfect score on a field known to be hard is the tell, not the win.
    """
    row = {"Product": "roti de boeuf", "Reason": "Presence of listeria",
           "Class": "Voluntary", "Pathogen": "Listeria monocytogenes",
           "Notes": "", "Source": "RappelConso (FR)"}
    vals, _tier = ES.derive(row)
    assert vals["FoodCategory"] == "unknown", (
        "a RappelConso row carries no regulator category; anything but "
        "unknown here means the category is being manufactured")


def test_unknown_is_a_permitted_answer_on_every_axis():
    empty = {"Product": "", "Reason": "", "Notes": "", "Class": "",
             "Pathogen": ""}
    vals, _t = ES.derive(empty)
    for k in ("FoodCategory", "ProcessType", "ConsumptionState",
              "StorageCondition", "PackagingType", "PackagingForm",
              "HazardGroup", "HazardCertainty", "NoticeType"):
        assert vals[k] in ("unknown", ""), (k, vals[k])


def test_hazard_group_does_not_call_a_process_deviation_a_pathogen():
    for label in ("None (organoleptic spoilage)",
                  "Unspecified hazard",
                  "Possible incomplete pasteurization (process deviation)",
                  "Unintended fermentation (microbiological quality)"):
        assert ES._hazard_group(label) == "unknown", label


def test_hazard_group_places_the_real_families():
    cases = {
        "Listeria monocytogenes": "pathogen-bacterial",
        "Aflatoxin": "mycotoxin",
        "Ochratoxin A": "mycotoxin",
        "Histamine / scombrotoxin": "biotoxin",
        "Norovirus": "pathogen-viral",
        "Hepatitis A virus": "pathogen-viral",
        "Foreign material (glass)": "foreign-material",
        "Physical/foreign-body contamination": "foreign-material",
        "Cadmium (heavy metal)": "heavy-metal",
        "PFOA / PFAS": "chemical",
        "Cyclospora": "pathogen-parasitic",
    }
    for label, want in cases.items():
        assert ES._hazard_group(label) == want, (label, ES._hazard_group(label))


def test_class_splits_into_two_variables():
    """Class holds an action and a severity. Reading it as one made
    'Class I' and 'Class 1' separate values of a categorical."""
    for cls, sev in (("Class I", "class-i"), ("Class 1", "class-i"),
                     ("Class II", "class-ii"), ("Class 2", "class-ii")):
        v, _t = ES.derive({"Class": cls, "Product": "", "Reason": "",
                           "Notes": "", "Pathogen": "Salmonella"})
        assert v["SeverityClass"] == sev, cls
    v, _t = ES.derive({"Class": "Border Rejection", "Product": "",
                       "Reason": "", "Notes": "", "Pathogen": "Salmonella"})
    assert v["NoticeType"] == "border-rejection"


def test_the_tier_is_recorded_and_regulator_beats_keyword():
    rasff = {"Product": "halloumi", "Class": "Alert", "Pathogen": "Listeria",
             "Reason": "Listeria in halloumi; risk: serious; "
                       "category: milk and milk products",
             "Notes": "[RASFF #2026.1; classification: alert notification]"}
    _v, tier = ES.derive(rasff)
    assert tier == "tier1-regulator"


def test_a_human_value_is_never_overwritten(tmp_path):
    import openpyxl
    p = tmp_path / "w.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Recalls"
    hdr = list(PIPELINE_OWNED) + ["EnrichedBy", "FoodCategory"]
    ws.append(hdr)
    ws.append(["2026-08-01", "RASFF (EU)", "c", "b", "halloumi", "Listeria",
               "r", "Alert", "Cyprus", "Europe", 1, 0, "http://x", "", "",
               "", "", "W31", "human", "dairy-soft-cheese"])
    wb.save(p)
    res = ES.run(p, write=True)
    assert res["skipped_human"] == 1
    got = openpyxl.load_workbook(p)["Recalls"]
    row2 = {h: got.cell(row=2, column=i + 1).value
            for i, h in enumerate([c.value for c in got[1]])}
    assert row2["FoodCategory"] == "dairy-soft-cheese"
    assert row2["EnrichedBy"] == "human"
