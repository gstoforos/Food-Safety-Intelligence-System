"""
AFTS Food Safety Intelligence System - Weekly Report Generator
Template: 2026-W16.html (gold standard)
Output: docs/YYYY-WW.html + data/weekly-summary-latest.json
Reads: docs/data/recalls.xlsx Recalls sheet ONLY.
Schedule: Friday 10:00 Athens time (dual-DST cron + TZ guard).
"""

import json, logging, os, re, requests, sys, argparse, html as html_mod
from datetime import datetime, timezone, timedelta, date
from pathlib import Path
from collections import Counter, OrderedDict
from typing import List, Dict, Any, Tuple
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)
CLAUDE_API_KEY = os.getenv("ANTHROPIC_API_KEY")

# Stamped on the front of the deterministic P4 paragraph by
# _process_authority_note() and stripped by the HTML renderer. See the
# "THE NOTE IDENTIFIES ITSELF" comment at the render site: detection used to
# be phrase-sniffing, which let the Listeria narrative wear this label.
PA_NOTE_MARKER = "[[PA-NOTE]] "

SEVERITY = OrderedDict([
    ("clostridium botulinum",1),("botulinum",1),
    ("listeria monocytogenes",2),("listeria",2),
    ("stec",3),("e. coli o157",3),("e. coli",3),
    ("salmonella",4),("cereulide",5),("bacillus cereus",5),
    ("norovirus",6),("hepatitis",7),
])


# ---------------------------------------------------------------------------
# Plural-aware count phrasing (added 2026-05-08).
# Replaces the legacy "{n} thing(s)" pattern with grammatical phrasing.
# Use as: _count_phrase(n, "incident") -> "1 incident" / "5 incidents".
# Pass zero= to override the n==0 case (e.g. "no confirmed outbreak events").
# ---------------------------------------------------------------------------
def _count_phrase(n: int, singular: str, *, plural: str = None,
                  zero: str = None) -> str:
    if n == 0 and zero is not None:
        return zero
    if n == 1:
        return f"{n} {singular}"
    return f"{n} {plural or singular + 's'}"


# ---------------------------------------------------------------------------
# Pathogen synonym consolidation (added 2026-04-27).
# Mirrors the table in build_monthly_report_afts.py — kept separate to
# preserve the surgical "no shared module" architecture you've used.
# Buckets are CHEMICALLY/EPIDEMIOLOGICALLY distinct: rodenticide is
# anticoagulant chemical poisoning (HiPP), distinct from rodent
# contamination (live pests). Salmonella serovars are deliberately
# preserved.
# ---------------------------------------------------------------------------
_PATHOGEN_SYNONYMS = {
    # Rodenticide chemical poisoning
    "rat poison":                                          "Rodenticide",
    "Rat poison":                                          "Rodenticide",
    "Rodenticide (rat poison)":                            "Rodenticide",
    "rodenticide (rat poison)":                            "Rodenticide",
    "Rodenticide poisoning":                               "Rodenticide",
    "Bromadiolone":                                        "Rodenticide",
    "bromadiolone":                                        "Rodenticide",
    # Rodent / pest contamination (different category)
    "Rodent contamination (physical/microbial hazard)":    "Rodent contamination",
    "Rodent contamination (physical/biological hazard)":   "Rodent contamination",
    "Mouse contamination (physical/biological hazard)":    "Rodent contamination",
    "Mouse contamination":                                 "Rodent contamination",
    # Aflatoxin
    "Aflatoxins":                                          "Aflatoxin",
    # Ochratoxin — merge bare "Ochratoxin" (rare untyped mention) and
    # "Ochratoxin A" (the specific congener that drives almost every recall)
    # into a single bucket per audit 2026-06-12. Splitting them across two
    # rows in the pathogen-distribution table fragments what is operationally
    # one hazard class.
    "Ochratoxin":                                          "Ochratoxin / Ochratoxin A",
    "Ochratoxin A":                                        "Ochratoxin / Ochratoxin A",
    "Ochratoxine A":                                       "Ochratoxin / Ochratoxin A",
    "OTA":                                                 "Ochratoxin / Ochratoxin A",
    # Bacillus cereus / cereulide
    "Bacillus cereus / cereulide":                         "Bacillus cereus / Cereulide",
    "Bacillus cereus (cereulide)":                         "Bacillus cereus / Cereulide",
    "Cereulide (B. cereus toxin)":                         "Bacillus cereus / Cereulide",
    "Cereulide":                                           "Bacillus cereus / Cereulide",
    # E. coli / STEC variants → single bucket.
    # Audit 2026-05-15 (fid 22275 W20): the source feed emits both
    # "Shiga toxin-producing E. coli (STEC)" and "Escherichia coli STEC"
    # for the same pathogen on the same producer, the same day. The dict
    # previously only covered the first; the second fell through and
    # rendered as a separate row in the distribution table. Now every
    # known surface form is mapped. The regex-based catch-all in
    # _consolidate_pathogen_label() backstops any future variant.
    "STEC (Shiga toxin-producing E. coli)":                "E. coli STEC",
    "Shiga toxin-producing E. coli (STEC)":                "E. coli STEC",
    "Shiga toxin-producing E. coli":                       "E. coli STEC",
    "E. coli STEC (Shiga toxin-producing)":                "E. coli STEC",
    "Shigatoxin producing Escherichia coli (STEC)":        "E. coli STEC",
    "Shigatoxin producing Escherichia coli":               "E. coli STEC",
    "Shigatoxin-producing Escherichia coli (STEC)":        "E. coli STEC",
    "Escherichia coli STEC":                               "E. coli STEC",
    "Escherichia coli shiga toxinogène (STEC)":            "E. coli STEC",
    "Escherichia coli shiga toxinogène":                   "E. coli STEC",
    "STEC / E. coli O157:H7":                              "E. coli STEC",
    "E. coli O157:H7":                                     "E. coli STEC",
    "E. coli O157":                                        "E. coli STEC",
    "E. coli O26 (STEC)":                                  "E. coli STEC",
    "E. coli O26":                                         "E. coli STEC",
    "E. coli O145 (STEC)":                                 "E. coli STEC",
    "E. coli O145":                                        "E. coli STEC",
    "E. coli O103 (STEC)":                                 "E. coli STEC",
    "E. coli O111 (STEC)":                                 "E. coli STEC",
    "E. coli O121 (STEC)":                                 "E. coli STEC",
    # Bare "E. coli" / "Escherichia coli" / "Escherichia coli (generic)"
    # are DELIBERATELY left out of this mapping — those forms can refer
    # to hygiene-indicator counts (non-pathogenic), and lumping them into
    # the STEC bucket would inflate STEC counts. Keep them as their own
    # bare labels; the consolidator's regex pass only triggers on
    # explicit STEC / shiga-toxin / VTEC tokens, so bare-E.coli is safe.
    # Marine biotoxins
    "Lipophilic biotoxins (DSP)":                          "Marine biotoxins",
    "Lipophilic biotoxins":                                "Marine biotoxins",
    "Paralytic shellfish toxins (PSP)":                    "Marine biotoxins",
    "Paralytic shellfish toxins":                          "Marine biotoxins",
    "Paralytic Shellfish Toxins (saxitoxins)":             "Marine biotoxins",
    "Phytoplankton biotoxins":                             "Marine biotoxins",
    # Salmonella — bare label only; serovars preserved
    "Salmonella":                                          "Salmonella spp.",
    # Histamine
    "Histamine":                                           "Histamine / scombrotoxin",
    "Histamine (biotoxine endogène)":                      "Histamine / scombrotoxin",
    "Scombrotoxin":                                        "Histamine / scombrotoxin",
}


def _consolidate_pathogen_label(label):
    """Map raw pathogen label to canonical bucket. Idempotent.

    Resolution order (audit 2026-05-15):
      1. Direct exact-string match against _PATHOGEN_SYNONYMS.
      2. Strip any trailing parenthesised qualifier and re-match.
         Catches "Listeria monocytogenes (note)" → "Listeria monocytogenes".
      3. STEC catch-all regex — any string mentioning STEC, shiga-toxin,
         shigatoxin, or VTEC collapses to canonical "E. coli STEC".
         Backstops future surface-form variants the dict hasn't seen.
      4. Salmonella catch-all — bare "Salmonella" or "Salmonella spp"
         folds to "Salmonella spp." (named serovars are preserved).
      5. Fall back to the paren-stripped base form (best-effort cleanup).

    Idempotency: a canonical output fed back in must return itself
    unchanged. The synonym keys never appear as values of OTHER keys,
    so direct match in step 1 is a fixed point.
    """
    if not label:
        return label
    # A compound multi-pathogen OUTBREAK category is already canonical and
    # must survive re-consolidation. Without this guard the STEC catch-all
    # in step 3 collapses "STEC + Salmonella Agona" back into
    # "E. coli STEC", and the Hazard Profile loses the separate row while
    # the register still shows the compound name — the two disagreeing
    # about the same event. See co_pathogen_label(). (Audit 2026-08-28.)
    if " + " in str(label):
        return str(label)
    s = str(label).strip()
    # 1) Direct exact match
    if s in _PATHOGEN_SYNONYMS:
        return _PATHOGEN_SYNONYMS[s]
    # 2) Paren-stripped match — catches "X (note)" forms
    base = s.split("(")[0].strip()
    if base != s and base in _PATHOGEN_SYNONYMS:
        return _PATHOGEN_SYNONYMS[base]
    # 3) STEC catch-all — any STEC / shiga-toxin / VTEC mention
    if re.search(r"\b(stec|shiga[\s\-]?toxin|shigatoxin|shigatoxinogène|vtec)\b",
                 s, re.I):
        return "E. coli STEC"
    # 4) Bare Salmonella collapse (preserves serovars like "Salmonella Stanley")
    if re.fullmatch(r"salmonella(\s+spp\.?)?", s, re.I):
        return "Salmonella spp."
    # 5) Best-effort: bare base form
    return base if base else s


_CO_PATHOGEN_RE = re.compile(
    r"\bSalmonella\s+([A-Z][a-z]+)\b|\b(Shiga toxin-producing|STEC)\b",
    re.I)


def co_pathogen_label(row):
    """Second hazard on a multi-pathogen OUTBREAK row, or None.

    Audit 2026-08-28. The W35 sprout event is officially a STEC AND
    Salmonella Agona outbreak: of 55 cases, 46 STEC only, 7 Salmonella
    Agona only, 2 coinfected. Displaying it under either genus alone is
    wrong, and splitting the row across both double-counts a register
    whose whole contract is one row per notice.

    So it gets its own mutually exclusive category. The second pathogen is
    read from the row's Reason — the operator-verified statement already in
    the corpus — NOT hardcoded here and NOT written into the Pathogen
    column, which is a controlled vocabulary the detector strata depend on.
    A compound value there would create a label of exactly one record,
    below every sparsity floor, and it would vanish from the analysis.

    Only OUTBREAK rows qualify. A routine recall mentioning two hazards in
    its Reason is not a coinfection event and must not gain a category.
    """
    if str(row.get("Outbreak") or "").strip() not in ("1", "1.0"):
        return None
    reason = str(row.get("Reason") or "")
    base = _consolidate_pathogen_label(str(row.get("Pathogen") or "").strip())
    has_stec = bool(re.search(r"shiga toxin-producing|\bSTEC\b", reason, re.I))
    m = re.search(r"\bSalmonella\s+([A-Z][a-z]{2,})\b", reason)
    if base.startswith("E. coli") and m:
        return "STEC + Salmonella {}".format(m.group(1))
    if base.startswith("Salmonella") and has_stec:
        return "STEC + {}".format(base)
    return None


def hazard_label(row):
    """The category a row is counted and displayed under."""
    return (co_pathogen_label(row)
            or _consolidate_pathogen_label(str(row.get("Pathogen") or "").strip()))


def _consolidate_counter(c):
    """Return a new Counter with synonymous keys merged."""
    out = Counter()
    for k, v in c.items():
        out[_consolidate_pathogen_label(k)] += v
    return out


def _dot_color(pathogen):
    p = (pathogen or "").lower()
    return "#9333ea" if ("botulinum" in p or "clostridium" in p) else "#dc2626"

def _severity_score(pathogen):
    p = (pathogen or "").lower()
    for k, v in SEVERITY.items():
        if k in p: return v
    return 99

COUNTRY_DISPLAY = {"USA":"United States","UK":"United Kingdom"}
AUTHORITY_DISPLAY = {
    "FDA":"FDA","USDA FSIS":"USDA FSIS","CFIA":"CFIA","MAPAQ QC":"MAPAQ QC",
    "RappelConso (FR)":"RappelConso (FR)","BVL (DE)":"BVL (DE)",
    "FSA (UK)":"FSA (UK)","MPI NZ":"MPI NZ","MPI (NZ)":"MPI NZ",
    "Min. Salute (IT)":"Min. Salute (IT)","EFET (GR)":"EFET (GR)",
    "AESAN (ES)":"AESAN (ES)","RASFF (EU)":"RASFF (EU)",
    "FSANZ (AU)":"FSANZ (AU)","BLV (CH)":"BLV (CH)",
    "AGES (AT)":"AGES (AT)","SZPI (CZ)":"SZPI (CZ)",
    "ANVISA (BR)":"ANVISA (BR)","ANMAT (AR)":"ANMAT (AR)",
    "COFEPRIS (MX)":"COFEPRIS (MX)","NCC (ZA)":"NCC (ZA)",
    "SFA (SG)":"SFA (SG)","CFS (HK)":"CFS (HK)","MFDS (KR)":"MFDS (KR)",
}
GEO_AUTHORITY = {
    "France":"RappelConso / DGCCRF","Germany":"BVL","Canada":"CFIA",
    "United States":"FDA / USDA FSIS","USA":"FDA / USDA FSIS",
    "New Zealand":"MPI / FSANZ","Italy":"Ministero della Salute",
    "United Kingdom":"FSA","UK":"FSA","Spain":"AESAN","Greece":"EFET",
    "Australia":"FSANZ","Switzerland":"BLV","Austria":"AGES",
    "Brazil":"ANVISA","South Africa":"NCC / NRCS","Hong Kong":"CFS",
    "Singapore":"SFA","Japan":"MHLW","South Korea":"MFDS",
    "Taiwan":"MOHW / TFDA","Mexico":"COFEPRIS","Argentina":"ANMAT",
    "Czech Republic":"SZPI","Slovakia":"\u0160VPS","Belgium":"AFSCA",
}

def load_recalls(xlsx_path):
    wb = load_workbook(xlsx_path, data_only=True)
    if "Recalls" not in wb.sheetnames:
        log.error("No Recalls sheet"); return []
    ws = wb["Recalls"]; hdr = [c.value for c in ws[1]]
    return [{h:(v if v is not None else "") for h,v in zip(hdr, row)}
            for row in ws.iter_rows(min_row=2, values_only=True)]

def filter_week(recalls, week_end):
    """Return rows belonging to the weekly report ending Friday `week_end`.

    Audit 2026-05-10: prefer the sticky `report_week` stamp set by
    merge_master.compute_report_week() at promote time. Fall back to
    legacy date-window math ONLY for rows that have no stamp yet — i.e.
    historical rows promoted before 2026-05-10 (everything dated
    before 2026-05-02 per migrate_recalls_report_week.py).

    Why prefer the stamp:
      • Late-arriving rows (scraped Wednesday for an event published the
        prior Thursday) get correctly attributed to LAST week's report
        instead of leaking into THIS week's date window.
      • Friday-dated rows are deferred to NEXT week's report (the AM
        ship-time scrape can't yet capture full Friday data); the legacy
        date math wrongly included them in the same-day report.

    Stamp rule (mirrored in merge_master.compute_report_week):
        report_week = "W{nn}", nn = ISO week of the smallest Friday > Date.

    Legacy fallback window (for rows with empty report_week):
        week_end - 6  ≤  Date  ≤  week_end   (Sat..Fri inclusive, 7 days).
        This is the OLD inclusive-Friday rule under which historical
        weekly reports (W01..W19 of 2026) were originally built. Keeping
        it as fallback preserves identical output for refresh_stale_weeks
        runs against pre-stamp data.
    """
    _exp_iso = week_end.isocalendar()
    expected_tag = "W{:02d}".format(_exp_iso[1])
    expected_year = _exp_iso[0]
    legacy_ws = week_end - timedelta(days=6)

    def _row_report_year(row):
        """ISO year of the smallest Friday > Date (same rule as the stamp).

        The `report_week` stamp is year-less ("W30"), so an old row from a
        previous year with the same week number would otherwise match the
        current report. Audit 2026-07-24: a 2025-07-25 FSAI row stamped W30
        leaked into 2026-W30. Deriving the row's own report-year and
        requiring it to match closes that hole, and handles the Dec/Jan
        boundary correctly (a late-December date can legitimately roll into
        W01 of the following ISO year).
        """
        d = row.get("Date", "")
        try:
            rd = datetime.strptime(str(d)[:10], "%Y-%m-%d").date()
        except (ValueError, TypeError):
            return expected_year          # unparseable -> don't drop the row
        days = (4 - rd.weekday()) % 7 or 7   # smallest Friday strictly after
        return (rd + timedelta(days=days)).isocalendar()[0]

    out = []
    for r in recalls:
        stamp = (r.get("report_week") or "").strip()
        if stamp:
            # New path: trust the sticky stamp set at promote time,
            # but confirm the row belongs to THIS report year.
            if stamp == expected_tag and _row_report_year(r) == expected_year:
                out.append(r)
            continue
        # Legacy fallback: date-window math (pre-stamp historical rows).
        d = r.get("Date", "")
        if not d:
            continue
        try:
            rd = datetime.strptime(str(d)[:10], "%Y-%m-%d").date()
        except ValueError:
            continue
        if legacy_ws <= rd <= week_end:
            out.append(r)
    return out


def _display_window(week_end, filtered_recalls):
    """Return (display_start, display_end) for the report period header.

    Audit 2026-05-12: the displayed period depends on whether the report
    was built under the OLD inclusive-Friday rule (Sat→Fri, 7 days) or
    the NEW Friday-strict rule (Fri→Thu, 7 days). A report is treated
    as new-rule iff at least one of its filtered rows carries a
    report_week stamp — that stamp is set only by compute_report_week()
    at promote time and only exists for rows dated ≥ 2026-05-02 per
    migrate_recalls_report_week.py.

    NEW rule (any row stamped):
        display_start = week_end − 7 (previous Friday — first data day)
        display_end   = week_end − 1 (Thursday — last data day)
        ship Friday itself is NOT in the data window (rows dated that
        day get stamped to NEXT week's report).

    LEGACY rule (no stamped rows — pre-migration historicals):
        display_start = week_end − 6 (Saturday)
        display_end   = week_end    (Friday — ship day INCLUDED)
        matches what W01..W18 of 2026 originally shipped, so historical
        reports keep their published display.

    Transitional W19 caveat: the migration intentionally left rows
    dated 2026-05-01 (the prior ship Friday under the OLD rule) blank
    rather than re-stamping them W19, so W19's filtered set starts at
    May 2. display_start therefore derives from week_end − 6 in the
    new-rule branch only if no row predates that — otherwise we let
    the actual min(date) anchor the display start, so the period header
    matches the data it summarises. Same applies to display_end vs
    week_end − 1: if the last stamped row is earlier, use that.
    """
    any_stamp = any((r.get("report_week") or "").strip() for r in filtered_recalls)
    if any_stamp:
        # New-rule window: Friday (week_end − 7) through Thursday (week_end − 1).
        rule_start = week_end - timedelta(days=7)
        rule_end = week_end - timedelta(days=1)
        # Tighten to actual data span — only for new-rule weeks. Handles the
        # transitional W19 case where the rule window starts at Fri May 1
        # but the migration left May 1 blank-stamped, so its earliest
        # stamped row is Sat May 2. Legacy weeks are NEVER tightened —
        # their period header must match what shipped at the time.
        parsed_dates = []
        for r in filtered_recalls:
            d = r.get("Date", "")
            if not d:
                continue
            try:
                parsed_dates.append(datetime.strptime(str(d)[:10], "%Y-%m-%d").date())
            except ValueError:
                continue
        if parsed_dates:
            ds = min(parsed_dates); de = max(parsed_dates)
            display_start = max(rule_start, ds)
            display_end = min(rule_end, de)
            if display_start > display_end:
                display_start, display_end = rule_start, rule_end
        else:
            display_start, display_end = rule_start, rule_end
    else:
        # Legacy rule: Saturday (week_end − 6) through Friday (week_end).
        # No tightening — preserves the published display for W01..W18 of
        # 2026 even if the dataset is later corrected (rows added/removed).
        display_start = week_end - timedelta(days=6)
        display_end = week_end
    return display_start, display_end


def _safe_int(v, default=0):
    try: return int(v)
    except (ValueError, TypeError): return default

def _country_display(c): return COUNTRY_DISPLAY.get(c, c)


# ───────────────────────────────────────────────────────────────────────
# Public helper API — used by docs/build_monthly_report_afts.py.
#
# Historical note: the monthly builder imports six helpers from this
# module — safe_int, severity_score, is_report_grade_url,
# pathogen_badge_color, rank_top_recalls, render_top5_row. Two were
# already implemented as private (_safe_int, _severity_score). The
# other four were authored inside build_html() in past iterations and
# never hoisted to module scope, so the monthly build broke with
# AttributeError until this section was added (2026-05-01).
#
# Exposed here at module scope so:
#   • the monthly builder can call weekly.<name>(…) without changes
#   • unit tests can import them in isolation
#   • the weekly builder itself can use them in build_html() for any
#     future consolidation of duplicated inline code
#
# All functions return primitives (str, int, list, tuple) — no side
# effects, no I/O, no global state.
# ───────────────────────────────────────────────────────────────────────

#: Return value type for `severity_score()` — kept as a 2-tuple so the
#: monthly builder's `_, canon = severity_score(name)` unpack works.
def safe_int(v, default=0):
    """Public alias for _safe_int. Convert v to int, returning default on
    failure. Accepts strings, floats, None, or anything that int() can
    coerce; never raises."""
    return _safe_int(v, default)


def severity_score(pathogen):
    """Return (score, canonical_name) for a pathogen string.

    Score: lower = more severe (1 = botulism, 2 = listeria, ..., 99 = unknown).
    Canonical name: synonym-collapsed display label using _PATHOGEN_SYNONYMS.

    Caller patterns:
      score = severity_score(name)[0]
      _, canon = severity_score(name)
    """
    score = _severity_score(pathogen)
    canonical = _PATHOGEN_SYNONYMS.get(pathogen or "", pathogen or "Unknown")
    return score, canonical


def pathogen_badge_color(canonical):
    """Hex colour used for a pathogen's badge / table dot / progress-bar.

    Uses canonical names (post-synonym-collapse) so synonyms map to the
    same colour. The palette tracks SEVERITY tiers — botulism is the
    most-saturated red, listeria deep amber, salmonella mid-amber, the
    less-acute hazards in muted greys / blues.
    """
    name = (canonical or "").lower()
    # Tier 1 — fatal-risk pathogens
    if "botulin" in name:                            return "#b91c1c"   # deep red
    if "listeria" in name:                           return "#dc2626"   # red
    if "stec" in name or "shiga" in name or "o157" in name:
                                                     return "#ea580c"   # orange-red
    # Tier 2 — common acute pathogens
    # Severe Vibrio species sit in the Tier-1 band, not the cyan Tier-2 one
    # (added 2026-08-14). This palette is documented as tracking SEVERITY,
    # so leaving V. vulnificus — ~1 in 5 fatal per CDC — the same cyan as
    # V. parahaemolyticus would make the colour say the opposite of the
    # tier. Must precede the generic "vibrio" branch below.
    if "vulnificus" in name:                         return "#dc2626"   # red
    if "cholera" in name and "non-o" not in name:    return "#dc2626"   # red
    if "salmonella" in name:                         return "#f59e0b"   # amber
    if "campylobacter" in name:                      return "#d97706"   # amber-dark
    if "e. coli" in name or "escherichia" in name:   return "#f97316"   # orange
    if "vibrio" in name:                             return "#0891b2"   # cyan
    if "cronobacter" in name:                        return "#0d9488"   # teal
    # Tier 3 — viral / parasitic
    if "norovirus" in name or "norwalk" in name:     return "#7c3aed"   # violet
    if "hepatitis" in name:                          return "#a855f7"   # purple
    if "cyclospora" in name or "crypto" in name or "giardia" in name:
                                                     return "#9333ea"   # purple-dark
    # Tier 4 — toxins / chemical / physical
    if "aflatox" in name or "ochra" in name or "fumon" in name or "patulin" in name:
                                                     return "#854d0e"   # brown
    if "histamine" in name or "scombro" in name:     return "#a16207"   # khaki
    if "rodenticide" in name or "rat poison" in name:
                                                     return "#1e293b"   # near-black
    if any(t in name for t in ("lead", "cadmium", "mercury", "arsenic")):
                                                     return "#475569"   # slate
    if "physical" in name or "glass" in name or "metal" in name:
                                                     return "#525252"   # neutral
    # Default
    return "#6b7280"   # grey


def is_report_grade_url(url):
    """True if the URL points at a primary regulator / official source.

    Used in the monthly summary JSON so subscribers see only links to
    authoritative sources (FDA, USDA, CFIA, RASFF, RappelConso, etc.)
    and not aggregator/news rewrites. Hostname-based — match-anything
    in a small allow-list, no live HTTP check.
    """
    if not url:
        return False
    u = str(url).lower()
    if not u.startswith(("http://", "https://")):
        return False
    # Strip protocol + leading "www." for substring matching
    host = u.split("://", 1)[1].split("/", 1)[0].lstrip("www.")
    REGULATOR_HOSTS = (
        # USA
        "fda.gov", "fsis.usda.gov", "cdc.gov", "epa.gov",
        # Canada
        "canada.ca", "inspection.canada.ca", "gnb.ca", "mapaq.gouv.qc.ca",
        "quebec.ca",
        # EU + member states
        "europa.eu", "efsa.europa.eu", "rasff.eu",
        "rappel.conso.gouv.fr", "economie.gouv.fr",
        "bvl.bund.de", "bund.de",
        "salute.gov.it",
        "aesan.gob.es", "mscbs.gob.es",
        "efet.gr", "minagric.gr",
        "voedselveiligheid.be", "favv.be", "afsca.be",
        "blv.admin.ch",
        "ages.at",
        "szpi.gov.cz",
        # UK
        "food.gov.uk", "fss.scot", "gov.uk",
        # Asia-Pacific
        "mpi.govt.nz", "foodstandards.gov.au", "foodstandards.govt.nz",
        "sfa.gov.sg", "cfs.gov.hk", "mfds.go.kr", "mhlw.go.jp",
        "fda.gov.tw",
        # Latin America + Africa
        "anvisa.gov.br", "anmat.gob.ar", "cofepris.gob.mx",
        "nrcs.gov.za", "nccsa.org.za",
    )
    return any(host == h or host.endswith("." + h) for h in REGULATOR_HOSTS)


def rank_top_recalls(recalls, n=10):
    """Return up to n recalls sorted by composite severity rank.

    Three-phase ordering (lower phase wins outright; ties broken inside
    the phase):

      Phase 0  — VERIFIED OUTBREAKS  (Outbreak truthy)
                 Outbreaks indicate active human illness, so they are
                 escalated above every non-outbreak recall regardless of
                 the underlying pathogen. Inside the phase: severity score
                 (lower = more severe) → tier → date desc → country.
                 Effect: a Listeria outbreak outranks a Salmonella outbreak
                 from the same day, but BOTH outrank a non-outbreak
                 Clostridium botulinum recall.

      Phase 1  — Clostridium botulinum (non-outbreak)
                 Pulled out of the general pool because botulism toxin
                 is fatal-risk even at single-recall scale and warrants
                 prominence on the marketing one-pager. Inside the phase:
                 tier → date desc → country.

      Phase 2  — EVERYTHING ELSE  (by severity → tier → date desc → country)

    Pre-2026-05-01 the ranker used a flat (severity, tier, outbreak, date)
    key. That collapsed the three Apr-2026 botulinum recalls (sev=1)
    to ranks 1-3 and pushed the Good4U Salmonella outbreak (sev=4) out
    of top-10 entirely — which is the wrong editorial signal for the
    public marketing PDF. Phase-based fix mirrors how George curates
    the headline-incidents list in the subscriber report.
    """
    OUTBREAK_TRUTHY = {True, 1, "1", "TRUE", "True", "true", "Y", "Yes"}

    # ── Illness burden (audit 2026-08-01) ─────────────────────────────────
    # Inside the outbreak phase the ranker used pathogen-INTRINSIC severity
    # and then date, and never looked at how big the outbreak actually was.
    # In July 2026 that put the Cyclospora iceberg-lettuce outbreak — 1,644
    # laboratory-confirmed cases, 94 hospitalisations, 9 states — BELOW the
    # Lamia Salmonella cluster of about 20 people, purely because Cyclospora
    # scores 99 on the intrinsic table while Salmonella scores 4.
    #
    # For events already confirmed as outbreaks, the pathogen genus is a poor
    # proxy for public-health weight; the case count is the direct measure.
    # Burden now leads inside phase 0, with intrinsic severity retained as
    # the tie-break for outbreaks whose counts are unknown or equal.
    #
    # Counts are read from the row's own Reason text — the same place a
    # reader sees them — so nothing is inferred that is not published.
    _BURDEN_RE = re.compile(
        r"(\d[\d,]{1,7})\s*(?:laboratory-confirmed\s+)?"
        r"(?:confirmed\s+)?(?:cases?|illnesses|ill\b|sickened|infections?)",
        re.IGNORECASE)
    _HOSP_RE = re.compile(r"(\d[\d,]{1,6})\s*hospitali[sz]", re.IGNORECASE)

    def _illness_burden(r):
        """Highest published case count for the row, 0 when none is stated."""
        text = " ".join(str(r.get(k) or "") for k in ("Reason", "Notes"))
        best = 0
        for rx in (_BURDEN_RE, _HOSP_RE):
            for m in rx.finditer(text):
                try:
                    best = max(best, int(m.group(1).replace(",", "")))
                except ValueError:
                    continue
        return best

    def _rank_key(r):
        pathogen = r.get("Pathogen") or ""
        sev_score, _ = severity_score(pathogen)
        tier        = _safe_int(r.get("Tier"), 99)
        is_outbreak = r.get("Outbreak") in OUTBREAK_TRUTHY
        is_botulin  = "botulin" in pathogen.lower()

        if is_outbreak:
            phase = 0
        elif is_botulin:
            phase = 1
        else:
            phase = 2

        # Negative ordinal → newer-first ordering
        try:
            d = datetime.strptime(str(r.get("Date") or "")[:10], "%Y-%m-%d")
            d_key = -d.toordinal()
        except (ValueError, TypeError):
            d_key = 0

        country_key = str(r.get("Country") or "").lower()

        # Burden only reorders CONFIRMED outbreaks (phase 0). Phases 1 and 2
        # keep their existing intrinsic-severity ordering untouched.
        burden_key = -_illness_burden(r) if phase == 0 else 0

        return (phase, burden_key, sev_score, tier, d_key, country_key)

    return sorted(recalls, key=_rank_key)[:n]


def brand_subline(company, brand):
    """The brand to print UNDER the company, or "" when there is nothing to add.

    REVIEW 2026-08-07 — the published W32 page exposed database concatenation
    in the Company / Brand column:

        U Bio U Bio
        Les Pensees Sauvages Les Pensees Sauvages
        Herens et Nautilus Collection Herens et Nautilus Collection
        Origin: Czechia | Notifying: Czechia—
        (not specified in RappelConso fiche 23090)Unbranded

    None of those is a data defect. A small producer legitimately IS its own
    brand, so Company == Brand is the correct record; RASFF genuinely has no
    brand, so "—" is correct; and "(not specified in fiche N)" + "Unbranded"
    is two honest disclosures rendered one after the other. The defect is the
    RENDERER printing a second line that carries no information.

    Section 02 already suppressed the exact-match case. Section 04 did not,
    and neither suppressed placeholders, case/whitespace variants, or a
    placeholder brand sitting under a placeholder company. One helper now
    serves both tables so they cannot drift apart again.

    A REAL brand under a "(not specified ...)" company is still printed — on
    fiche 23065 the fiche names no manufacturer but the brand is genuinely
    "U", and that is the most identifying thing the reader gets.
    """
    company = str(company or "").strip()
    brand = str(brand or "").strip()
    if not brand:
        return ""
    placeholders = {"—", "-", "–", "n/a", "na", "none", "unknown",
                    "unbranded", "sans marque", "aucune", "no brand"}
    b_low = brand.lower()
    if b_low in placeholders:
        return ""
    # "U Bio" vs "U Bio ", "SuperValu" vs "Supervalu" — same string, twice.
    if b_low == company.lower().strip():
        return ""
    # A company that is itself a disclosure gains nothing from a second
    # disclosure underneath it.
    if brand.startswith("(") and company.startswith("("):
        return ""
    return brand


def render_top5_row(rank, r):
    """Render ONE table row (HTML <tr>) for the top-N recalls table.

    The function name predates the rename to top-10 — kept for backward
    compatibility with existing call sites in the monthly builder.

    Layout (6 columns, must match `<thead>` in build_monthly_report_afts.py):
      1. # (rank)
      2. Date
      3. Pathogen (italic name + tier chip + outbreak chip)
      4. Company / Brand (company on top, brand muted below if present)
      5. Product (truncated to 110 chars)
      6. Jurisdiction & Source (country on top, agency below, link to source)

    Pre-2026-05-01 this emitted only 5 cells in a different column order,
    which silently shifted every td by one column under WeasyPrint and
    produced the broken Apr-2026 subscriber PDF. The 6-cell layout below
    aligns with the <thead> declaration in both §08 and Appendix A.
    """
    canon_name = severity_score(r.get("Pathogen") or "")[1]
    badge_color = pathogen_badge_color(canon_name)
    # Top 5 must name the same category as the register and the Hazard
    # Profile. The badge colour still keys off the canonical single
    # pathogen — a compound label has no colour of its own.
    canon_name = co_pathogen_label(r) or canon_name
    date_str = str(r.get("Date") or "")[:10] or "—"
    country  = _country_display(r.get("Country") or "—")
    source   = AUTHORITY_DISPLAY.get(
        r.get("Source") or r.get("Agency") or "",
        r.get("Source") or r.get("Agency") or "—"
    )
    company  = r.get("Company") or "—"
    brand    = (r.get("Brand") or "").strip()
    product  = r.get("Product") or r.get("Description") or "—"
    if len(product) > 110:
        product = product[:107] + "…"
    url      = (r.get("URL") or "").strip()

    # Tier / outbreak chips (best-effort; missing fields render nothing)
    chips = []
    tier = _safe_int(r.get("Tier"), 99)
    # Same three-value fix as the table renderer below: a Tier-3 row used to
    # append no chip at all here, so the card silently showed no tier.
    if   tier == 1: chips.append('<span class="chip chip-tier-1">Tier&nbsp;1</span>')
    elif tier == 2: chips.append('<span class="chip chip-tier-2">Tier&nbsp;2</span>')
    elif tier in (3, 4): chips.append(
        '<span class="chip chip-tier-3">Tier&nbsp;{}</span>'.format(tier))
    if r.get("Outbreak") in (True, 1, "1", "TRUE", "True", "true", "Y", "Yes"):
        chips.append('<span class="chip chip-outbreak">Outbreak</span>')
    chip_html = " ".join(chips)

    # Brand sub-line only when it adds information — see brand_subline().
    brand_html = ""
    _b = brand_subline(company, brand)
    if _b:
        brand_html = f'<div class="brand">{html_mod.escape(_b)}</div>'

    # Source cell: agency name as a link to URL (if present), else plain text
    if url:
        source_html = (
            f'<div class="country">{html_mod.escape(country)}</div>'
            f'<div class="agency"><a href="{html_mod.escape(url)}" target="_blank" '
            f'rel="noopener">{html_mod.escape(source)} &rarr;</a></div>'
        )
    else:
        source_html = (
            f'<div class="country">{html_mod.escape(country)}</div>'
            f'<div class="agency">{html_mod.escape(source)}</div>'
        )

    return (
        f'<tr>'
        f'<td class="num rank">#{rank}</td>'
        f'<td class="date">{html_mod.escape(date_str)}</td>'
        f'<td class="pathogen">'
        f'<span class="path-dot" style="background:{badge_color}"></span>'
        f'<em class="path-name">{html_mod.escape(canon_name)}</em>'
        f'{("&nbsp;" + chip_html) if chip_html else ""}'
        f'</td>'
        f'<td class="company">'
        f'<div class="company-name">{html_mod.escape(company)}</div>'
        f'{brand_html}'
        f'</td>'
        f'<td class="prod">{html_mod.escape(product)}</td>'
        f'<td class="src">{source_html}</td>'
        f'</tr>'
    )


def _incident_note(rows):
    """One sentence naming each collapsed cluster, or "" when there are none.

    Added 2026-08-20. The report previously showed a headline incident
    count and an appendix of notices with no statement of which notices
    had been grouped, so the two numbers looked like a contradiction
    rather than two units. A reader must be able to see WHERE the
    difference comes from, not just be told that it exists.
    """
    try:
        from pipeline._incident_id import group_sizes
    except Exception:                                          # noqa: BLE001
        return ""
    groups = group_sizes(rows)
    if not groups:
        return ""
    parts = []
    for gid, n in sorted(groups.items(), key=lambda kv: -kv[1]):
        label = gid.split(":")[-1].replace("-", " ")
        parts.append("{} notices from {}".format(n, label))
    return ("This week that applies to: " + "; ".join(parts) + ".")


def compute_stats(wr, pr):
    # ── Total counts INCIDENTS, not notices (audit 2026-08-15) ─────────
    # A single event can produce many regulator notices. E.Leclerc Dinan's
    # suspected refrigeration failure on 15 Aug generated TWENTY DGCCRF
    # fiches, one per supplier whose chilled stock was in the store, all
    # with the identical motif and the same distributeur. Counted as rows
    # that one broken chiller would have tripled the week and made
    # Listeria its dominant pathogen.
    #
    # pipeline/_incident_id only collapses rows a human has explicitly
    # tagged [incident:<id>]; anything untagged counts as itself. So
    # count_incidents(rows) == len(rows) for every week that predates the
    # tagging, and no historical figure moves.
    #
    # Fails OPEN to the row count: an inflated total is visible and
    # arguable, a crashed build ships nothing.
    try:
        from pipeline._incident_id import count_incidents, group_sizes
        total = count_incidents(wr)
        _groups = group_sizes(wr)
        if _groups:
            log.info("incident grouping: %s (%d notices -> %d incidents)",
                     _groups, len(wr), total)
    except Exception as _ie:                                  # noqa: BLE001
        log.warning("incident grouping unavailable (%s) — counting notices, "
                    "which OVERSTATES multi-notice events", _ie)
        total = len(wr)
    # ── Tier 1 must count the SAME UNIT as `total` (audit 2026-08-20) ───
    # `total` became an INCIDENT count above while this line kept counting
    # ROWS, and the two units met on the dashboard card as
    #     W34   total 43   Tier-1 51
    # — more critical items than items. The Leclerc Dinan cluster is 20
    # Listeria rows collapsing to one incident, so it added 1 to `total`
    # and 20 to `tier1`.
    #
    # An incident is Tier 1 if ANY of its notices is: the cluster's severity
    # is the worst hazard in it, not an average. Untagged rows are their own
    # incident, so this is identical to the row count for every week that
    # predates incident tagging — no historical figure moves, exactly as
    # with count_incidents.
    #
    # Fails OPEN to the row count for the same reason as above.
    try:
        from pipeline._incident_id import derive as _incident_of
        _t1_untagged = 0
        _t1_groups = set()
        for r in wr:
            if _safe_int(r.get("Tier")) != 1:
                continue
            iid = _incident_of(r)
            if iid:
                _t1_groups.add(iid)
            else:
                _t1_untagged += 1
        tier1 = len(_t1_groups) + _t1_untagged
    except Exception as _te:                                  # noqa: BLE001
        log.warning("incident grouping unavailable for tier1 (%s) — counting "
                    "notices, which OVERSTATES multi-notice events", _te)
        tier1 = sum(1 for r in wr if _safe_int(r.get("Tier")) == 1)

    # ── Outbreaks are counted as EVENTS, not rows (audit 2026-08-14) ────
    # `Outbreak` is a per-ROW boolean, but one outbreak produces several
    # rows: the implicated ingredient, each downstream manufacturer, the
    # agency investigation page, and any regulator's public-health alert.
    # W33 is the worked example — the Salmonella Javiana jalapeño event
    # appears as three separate recalls (FDA Taylor Fresh 9 Aug, FDA Whole
    # Foods 12 Aug, CFIA RA-82480 13 Aug), all naming the same open
    # investigation. Summing the flag printed one outbreak three times,
    # which is the inflation the operator removed by hand on 13 Aug.
    #
    # pipeline/_outbreak_id.count_events() has solved this since
    # 2026-08-11 — it merges rows that share a real investigation slug or
    # an explicit [outbreak:<id>] override, and pointedly REFUSES to merge
    # on the low-confidence pathogen+month key (that key was measured
    # collapsing four unrelated April events into one). It was simply
    # never called from here.
    #
    # Fails OPEN to the old row count: an over-count is visible and
    # arguable, a crashed build ships nothing.
    try:
        from pipeline._outbreak_id import count_events
        outbreaks = count_events(wr)
    except Exception as _oe:                                  # noqa: BLE001
        log.warning("outbreak event-count unavailable (%s) — falling back "
                    "to row count, which OVERSTATES multi-row events", _oe)
        outbreaks = sum(1 for r in wr if _safe_int(r.get("Outbreak")) == 1)
    pc = Counter(); cc = Counter()
    cs = {}  # country → set of Source labels — used by geographic-table
             # builder to choose the right "Authority" label per row. Audit
             # 2026-06-12: countries like Türkiye, India, Uganda, Egypt
             # appear only because RASFF (EU) notified about imports from
             # them — they should NOT be labeled "National Authority".
    # ── §03 MUST COUNT THE SAME UNIT THE KPIs COUNT (fix 2026-08-20) ────
    # `total`, `tier1` and `outbreaks` above are INCIDENT/EVENT counts.
    # This loop used to run over `wr` — the raw NOTICES — so §03 and §04
    # published notice counts against an incident denominator. In W34 that
    # printed "Listeria monocytogenes 36 — 84%": 36 is the notice count,
    # 84% is 36/43 where 43 is the incident total. The table was internally
    # impossible, because Listeria 36 + Salmonella 14 = 50 already exceeded
    # its own denominator of 43. An external reviewer caught it on exactly
    # that arithmetic.
    #
    # The cause is the E.Leclerc Dinan cluster: 20 DGCCRF fiches for one
    # suspected cold-chain failure, all coded Listeria. Counted as notices
    # they made Listeria look like 84% of the week; counted as incidents
    # Listeria is 17 of 43 (40%) and Salmonella 14 (33%).
    #
    # One representative row per tagged incident, every untagged row as
    # itself — the same collapse count_incidents() applies, so the table
    # and the KPI banner can no longer disagree.
    try:
        from pipeline._incident_id import derive as _incident_of
    except Exception:                                          # noqa: BLE001
        def _incident_of(_row):                                # type: ignore
            return None
    _dist_rows, _seen_incidents = [], set()
    for _r in wr:
        _iid = _incident_of(_r)
        if _iid:
            if _iid in _seen_incidents:
                continue
            _seen_incidents.add(_iid)
        _dist_rows.append(_r)
    if len(_dist_rows) != total:
        log.warning("distribution rows (%d) != incident total (%d) — §03 "
                    "percentages may not sum as expected",
                    len(_dist_rows), total)

    for r in _dist_rows:
        p = (r.get("Pathogen") or "").strip()
        # Audit 2026-05-15: previously this site did
        #   pc[p.split("(")[0].strip()] += 1
        # which stripped the "(STEC)" suffix BEFORE the synonym lookup
        # in _consolidate_counter — defeating it for every synonym key
        # that contains parens. Result: "Shiga toxin-producing E. coli
        # (STEC)" and "Escherichia coli STEC" rendered as TWO separate
        # rows in §03 Pathogen Profile. Now the consolidator owns all
        # paren-stripping logic and runs once, here.
        if p:
            # hazard_label() returns the compound category for a
            # multi-pathogen outbreak row and the ordinary consolidated
            # label otherwise, so the profile stays mutually exclusive and
            # the column still totals the notice count.
            pc[hazard_label(r)] += 1
        country = _country_display(r.get("Country","") or "Unknown")
        cc[country] += 1
        # Track every source label that reported this country in the window —
        # used by the geographic-table builder to disambiguate "issued by
        # national authority" from "RASFF EU origin country", etc.
        src = str(r.get("Source") or "").strip()
        if src:
            cs.setdefault(country, set()).add(src)
    pt = len(pr); delta = total - pt
    # Apply synonym consolidation BEFORE deriving top_pathogen so the KPI
    # banner and the distribution table never disagree.
    pc_consolidated = _consolidate_counter(pc)
    # Genus-level co-dominance check (reviewer note 2026-07-03): the KPI shows
    # the single top consolidated label, but a named serovar (e.g. "Salmonella
    # Stanley") is kept separate from "Salmonella spp." on purpose. Roll every
    # label up to its genus to see whether the leader is actually tied.
    def _genus(label):
        l = str(label).lower()
        if "listeria" in l: return "Listeria"
        if "salmonella" in l: return "Salmonella"
        if "coli" in l or "stec" in l: return "E. coli"
        return str(label)
    genus_counts = Counter()
    for k, v in pc_consolidated.items():
        genus_counts[_genus(k)] += v
    co_dominant = None
    if genus_counts:
        top_genus, top_gn = genus_counts.most_common(1)[0]
        tied = [g for g, n in genus_counts.items() if n == top_gn and g != top_genus]
        if tied:
            co_dominant = {"leader": top_genus, "count": top_gn,
                           "also": tied}
    top_pair = pc_consolidated.most_common(1)[0] if pc_consolidated else ("\u2014", 0)
    return {"total":total,"tier1":tier1,"outbreaks":outbreaks,
            "top_pathogen":top_pair,
            "co_dominant":co_dominant,
            "pathogen_counts":pc_consolidated.most_common(20),"country_counts":cc.most_common(20),
            "country_sources":cs,
            "prev_total":pt,"delta":delta,
            "delta_pct":round(delta/max(pt,1)*100) if pt else 0}

def _disambiguate_genus_labels(pathogen_counts):
    """Relabel a genus bucket that sits in the same table as one of its own
    serovars.

    REVIEW 2026-08-07 (W32). The Pathogen Profile listed:

        Salmonella spp.       15
        Salmonella Javiana     1

    Arithmetically fine — 15 + 1 = 16 and the table totals correctly. But
    Javiana IS a Salmonella serovar, so a reader cannot tell whether the
    genus burden for the week is 15 (32%) or 16 (34%), and the honest answer
    is 16. Splitting the named outbreak serovar out is deliberate and worth
    keeping: it is the row the outbreak hangs on. What was missing was a
    label saying the other row excludes it.

    So the counts are untouched — only the LABEL changes, and only when a
    named serovar of the same genus is actually present in the same table:

        Other Salmonella spp.   15
        Salmonella Javiana       1

    Rendering this instead of merging keeps the outbreak serovar visible in
    §03 while removing the taxonomic ambiguity. A week with no named serovar
    is left exactly as it was.
    """
    rows = list(pathogen_counts)
    genera = ("Salmonella", "Listeria", "Escherichia", "Clostridium",
              "Campylobacter", "Vibrio", "Bacillus", "Cronobacter")
    labels = [str(p) for p, _ in rows]

    out = []
    for path, cnt in rows:
        label = str(path)
        low = label.lower()
        for genus in genera:
            g = genus.lower()
            # This row is the GENERIC bucket for its genus ("Salmonella spp.",
            # "Salmonella", "Listeria spp.") — not a named species/serovar.
            rest = low.replace(g, "", 1).replace("spp.", "").replace("spp", "")
            rest = rest.replace(".", "").strip()
            if not low.startswith(g) or rest:
                continue
            # Is a named member of the same genus also in this table?
            named = [o for o in labels
                     if o != label and o.lower().startswith(g)]
            if named:
                label = "Other " + str(path)
            break
        out.append((label, cnt))
    return out


def _diversify_by_country(sorted_recalls, cap=2, window=5):
    """Reorder a severity-sorted list so no single country dominates the
    top `window` slots. Defers a country's (cap+1)th entry past the window
    when other countries are still waiting. Severity order is otherwise kept.
    """
    if not sorted_recalls:
        return sorted_recalls
    picked, deferred = [], []
    counts = {}
    for r in sorted_recalls:
        country = str(r.get("Country") or "").strip().lower()
        if len(picked) < window and counts.get(country, 0) >= cap:
            deferred.append(r)
        else:
            picked.append(r)
            counts[country] = counts.get(country, 0) + 1
    return picked + deferred


def sort_by_severity(recalls):
    def key(r):
        return (_severity_score(r.get("Pathogen","")),
                -_safe_int(r.get("Outbreak",0)), _safe_int(r.get("Tier",2)),
                -(datetime.strptime(str(r.get("Date","2000-01-01"))[:10],"%Y-%m-%d").toordinal()
                  if str(r.get("Date",""))[:10] else 0))
    return sorted(recalls, key=key)

def esc(s):
    if s is None: return ""
    return html_mod.escape(str(s))

def _fmt_date(d):
    try: return datetime.strptime(str(d)[:10],"%Y-%m-%d").strftime("%-d %b %Y")
    except Exception: return str(d)[:10]

def _fmt_date_short(d):
    try: return datetime.strptime(str(d)[:10],"%Y-%m-%d").strftime("%-d %b")
    except Exception: return str(d)[:10]


def generate_analysis_claude(stats, recalls):
    """Generate §01 Intelligence Analysis.
    P1-P3 come from Claude (or the fallback). P4 Process Authority Note is
    always deterministic — it cites specific CFR/EU/CFIA paragraphs and
    company names, which must be factually stable (not LLM-generated)."""
    tp, tc = stats["top_pathogen"]; t = stats["total"]
    pct = round(tc/max(t,1)*100)
    bot = [r for r in recalls if "botulinum" in (r.get("Pathogen") or "").lower()
           or "clostridium" in (r.get("Pathogen") or "").lower()]

    # Jurisdictions actually present this week — pass to Claude so P3 names them
    auths = _jurisdictions_from_recalls(recalls)
    auth_hint = ", ".join(auths[:5]) if auths else "multiple jurisdictions"

    # ── Commodity mix for the leading pathogen (audit 2026-08-14) ───────
    # The prompt used to receive pathogen counts and country counts and
    # NOTHING about what the products actually were, so the model had no
    # basis for a commodity statement and filled the gap from the canned
    # examples in the instruction itself. W33 is the worked case: the
    # published paragraph said this week's Salmonella signal spanned
    # "dairy and soft cheese" and cited "peanut butter, flour, infant
    # formula", while the dataset contained ZERO dairy or cheese
    # Salmonella rows. The real distribution was meat and poultry 15/27,
    # sesame and seed products 5/27. An external review caught it.
    # Passing the measured mix is what makes the paragraph describe this
    # week instead of Salmonella in general.
    commodity_mix = _commodity_mix(recalls, tp)

    prompt = """You are a food safety intelligence analyst for AFTS. Generate the Intelligence Analysis section.

DATA: Total={}, Tier-1={}, Outbreaks={} (DISTINCT EVENTS, not rows — several recalls can belong to one outbreak), Leading={} ({}, {}%)
Pathogens: {}
Countries: {}
Jurisdictions this week: {}
Commodity mix for {} THIS WEEK (measured from the actual rows): {}

Generate EXACTLY these three paragraphs (plain text, no HTML, no headers, no paragraph numbers):
1. Executive overview — total, tier-1, outbreaks, leading pathogen %, interpret as regulatory-pressure signal.
2. Pathogen-specific process-engineering analysis for {} — describe the product categories USING THE MEASURED COMMODITY MIX GIVEN ABOVE. Do not name a commodity class that does not appear in that mix, and do not substitute textbook examples for it; if the mix is spread across several classes, say so and read it as multiple independent supply-chain events rather than one commodity-level failure. Then give the specific failure modes (e.g. Zone 1 environmental harbourage, sanitation SOP drift, post-lethality recontamination for Listeria) and cite only BINDING regulatory frameworks (21 CFR 117 Preventive Controls including environmental monitoring, 21 CFR 113/114 thermal lethality, 9 CFR 430 for post-lethality-exposed RTE meat and poultry, FDA Produce Safety Rule 21 CFR 112, Reg. (EC) 2073/2005 as amended by Reg. (EU) 2024/2895, etc. as applicable). Do NOT cite FDA CPG 555.320: FDA's own page marks it "Draft - Not for Implementation" containing "non-binding recommendations", and it does not set a blanket RTE zero-tolerance.
3. Regulatory/geographic assessment — name the actual authorities active this week ({}). Close with AFTS recommendation to re-verify the single highest-leverage control for the commodity and confirm documentation packages are ready for rapid regulatory response.

Tone: professional, process-engineering voice, no emojis, no bullets, no colons at paragraph starts. 3-5 sentences each. Preserve the word 'AFTS' exactly where referenced. Do NOT write a Process Authority Note — that is appended separately.""".format(
        t, stats["tier1"], stats["outbreaks"], tp, tc, pct,
        dict(stats["pathogen_counts"]), dict(stats["country_counts"]), auth_hint,
        tp, commodity_mix,
        tp, auth_hint)

    claude_out = None
    try:
        resp = requests.post("https://api.anthropic.com/v1/messages",
            headers={"Content-Type":"application/json","x-api-key":CLAUDE_API_KEY},
            json={"model":"claude-sonnet-4-20250514","max_tokens":1200,
                  "messages":[{"role":"user","content":prompt}]}, timeout=60)
        if resp.status_code == 200:
            claude_out = resp.json()["content"][0]["text"]
        else:
            log.error("Claude %d", resp.status_code)
    except Exception as e:
        log.error("Claude error: %s", e)

    # Deterministic fallback for P1-P3 if Claude failed (tail call to _fallback
    # without bot so it produces only P1-P3).
    if claude_out is None:
        claude_out = _fallback_p1_to_p3(stats, recalls)

    # Append deterministic P4 Process Authority Note (multi-trigger).
    #
    # Marked, not sniffed. The renderer used to identify this paragraph by
    # looking for stock phrases such as "qualified process authority" — which
    # meant (a) the Listeria P2 narrative, which happens to contain that
    # phrase, was labelled "Process Authority Note", and (b) the moment the
    # 2026-08-07 review removed US process-authority language from the EU
    # branch of this note, the REAL note stopped matching and lost its label
    # to that impostor. The paragraph knows what it is; it now says so.
    pa = _process_authority_note(recalls, bot)
    if pa:
        claude_out = claude_out.rstrip() + "\n\n" + PA_NOTE_MARKER + pa
    return claude_out


def _fallback_p1_to_p3(stats, recalls):
    """P1-P3 only. The PA Note is always handled by _process_authority_note()."""
    tp, tc = stats["top_pathogen"]; t = stats["total"]
    pct = round(tc/max(t,1)*100)
    p1 = ("This week produced {} food-safety hazard recall incidents across the AFTS "
          "monitoring network, with {} classified as Tier-1 and {}. {} dominated the "
          "surveillance window, accounting for {} of {} "
          "incidents ({}%). Although total activity may vary week to week, the high "
          "Tier-1 share indicates that serious microbiological and chemical hazards "
          "continue to drive regulatory action.").format(t, stats["tier1"],
                                 _count_phrase(stats["outbreaks"], "confirmed outbreak event",
                                               zero="no confirmed outbreak events"),
                                 tp, tc, t, pct)
    p2 = (_pathogen_narrative(tp, pct, _commodity_mix(recalls, tp))
          + _framework_jurisdiction_qualifier(recalls))
    # Outbreak Watch: if any outbreak-flagged recall exists, surface it by name
    # so the outbreak KPI is reflected in the narrative even when the Top-5
    # headline threats are a different pathogen (reviewer note 2026-07-03).
    OUTBREAK_TRUTHY = {True, 1, "1", "TRUE", "True", "true", "Y", "Yes"}
    ob_recalls = [r for r in (recalls or [])
                  if r.get("Outbreak") in OUTBREAK_TRUTHY]

    # ── One line per EVENT, not per row (audit 2026-08-14) ─────────────
    # compute_stats() now counts distinct events, but this paragraph still
    # iterated rows, so the first build after that change printed
    # "3 confirmed cluster events tracked this week" directly beneath a
    # KPI reading "Active Outbreaks 1", and named the same jalapeño event
    # three times over ("linked to guacamoles; ... dips, salsa, guacamole
    # ...; ... finished products"). Fixing the counter without fixing the
    # prose would have produced a report that contradicted itself in
    # consecutive sentences — worse than the original error, because the
    # KPI now looks wrong instead of the narrative.
    #
    # Keep the EARLIEST-DATED row of each event. The comment here used to
    # claim that and the code did the opposite: it kept the first row of a
    # NEWEST-first list, i.e. the most recent one. For the jalapeno event
    # that surfaced the CFIA guacamole recall, so W33 described an outbreak
    # whose vehicle is jalapeno peppers as "Salmonella linked to
    # guacamoles" — naming a downstream retail product as the source
    # commodity. The earliest row of an investigation is normally closer to
    # the implicated ingredient.
    try:
        from pipeline._outbreak_id import derive as _derive_oid
        _seen_events, _deduped = {}, []
        for _r in sorted(ob_recalls, key=lambda x: str(x.get("Date") or "")):
            _oid, _conf, _ = _derive_oid(_r)
            _key = _oid if (_oid and _conf == "high") else id(_r)
            if _key in _seen_events:
                continue
            _seen_events[_key] = True
            _deduped.append(_r)
        ob_recalls = _deduped
    except Exception as _de:                                   # noqa: BLE001
        log.warning("outbreak narrative dedupe unavailable (%s) — naming "
                    "every flagged row", _de)
    p_outbreak = None
    if ob_recalls:
        descs = []
        # Was ob_recalls[:3] while n_ob counted ALL of them, so a 4-outbreak
        # week rendered "four confirmed cluster events ... " followed by only
        # three named events (reviewer note 2026-07-10). Name every event.
        for r in ob_recalls:
            path = str(r.get("Pathogen") or "").strip()
            prod = str(r.get("Product") or r.get("Company") or "").strip()

            # ── Audit 2026-08-28 ──────────────────────────────────────────
            # "{Pathogen} linked to {Product}" is a fair description of a
            # single-hazard recall and a poor one of an outbreak. The W35
            # sprout event involved three STEC serotypes AND Salmonella
            # Agona across 15 states; rendered from Pathogen alone it read
            # "Salmonella linked to alfalfa sprouts", which named the wrong
            # dominant pathogen and dropped the epidemiology entirely.
            #
            # The Pathogen column is a controlled vocabulary and must stay
            # one canonical value per row, so the multi-hazard description
            # cannot live there. Where an outbreak row's Reason is written
            # as a full outbreak description, use it verbatim: it is the
            # operator-verified statement, and it is in the corpus rather
            # than hardcoded here.
            _reason = str(r.get("Reason") or "").strip()
            if _reason.lower().lstrip().startswith(("multistate outbreak",
                                                    "multi-state outbreak",
                                                    "outbreak of")):
                descs.append(_reason.rstrip(" .") )
                continue
            # REVIEW 2026-08-07 — this was a hard prod[:60], which cut the
            # jalapeno product mid-word and published:
            #   "linked to Fresh jalapeno peppers grown in Sinaloa, Mexico
            #    — distribute."
            # Product strings in this database carry lot codes, pack sizes and
            # distribution after a dash or semicolon, so take the first clause
            # and, only if that is still long, stop on a word boundary.
            # Dropping the parenthetical is right when it holds pack sizes
            # or lot codes, and wrong when it holds the only description
            # there is: "Finished products (e.g. dips, salsa, guacamole and
            # more) containing jalapeno" collapses to "Finished products",
            # which names no food at all. Keep the fuller string when the
            # stripped head is generic or too short to identify anything.
            _head = prod.split("(")[0].strip()
            _generic = _head.lower().rstrip(" .,;:").rstrip("s") in (
                "finished product", "product", "various product",
                "prepared food", "food", "item", "various item",
                "assorted product", "multiple product",
            )
            prod_short = prod.strip() if (_generic or len(_head) < 18) else _head
            for sep in ("—", " - ", ";", "|"):
                if sep in prod_short:
                    prod_short = prod_short.split(sep)[0].strip()
            prod_short = prod_short.rstrip(" ,;:–—-")
            if len(prod_short) > 70:
                prod_short = prod_short[:70].rsplit(" ", 1)[0].rstrip(
                    " ,;:–—-") + "…"
            # "linked to Fresh jalapeno peppers" — Product is title-cased as
            # a heading but sits mid-sentence here. Lower the first letter,
            # unless the opening looks like an acronym or proper noun run
            # (two capitals: "IQF", "EU").
            if prod_short[:2].isalpha() and not prod_short[:2].isupper():
                prod_short = prod_short[0].lower() + prod_short[1:]
            descs.append(f"{path} linked to {prod_short}" if prod_short else path)
        joined = "; ".join(descs)
        n_ob = len(ob_recalls)
        p_outbreak = (
            f"Outbreak watch: {_count_phrase(n_ob, 'confirmed cluster event')} "
            f"tracked this week \u2014 {joined}. Although {tp} remained the leading "
            f"incident driver, {'this outbreak' if n_ob == 1 else 'these outbreaks'} "
            f"should remain under active monitoring. "
            # Audit 2026-08-28: this sentence used to close with "given
            # cross-border exposure potential". That asserted international
            # distribution for every outbreak, whatever the source said. The
            # W35 sprout outbreak was distributed directly in two US states
            # only; illnesses appearing in fifteen states is case
            # geography, not distribution geography, and the two are not
            # the same claim. Nothing here may state a distribution scope
            # the corpus does not carry.
            f"Distribution scope is as stated by the notifying authority; "
            f"case geography is not distribution geography.")
    auths = _jurisdictions_from_recalls(recalls or [])
    if auths:
        auth_clause = ("Regulatory activity this week spanned multiple jurisdictions "
                       "({}{}), signalling continued inspection intensity. ").format(
                           ", ".join(auths),
                           ", and national authorities" if len(auths) >= 3 else "")
    else:
        auth_clause = ("Regulatory activity this week spanned multiple jurisdictions, "
                       "signalling continued inspection intensity. ")
    p3 = (auth_clause +
          "AFTS recommends that food manufacturers use this briefing as a prompt to "
          "re-verify the single highest-leverage control for their commodity this week "
          "and to confirm documentation packages are ready for rapid regulatory response.")
    parts = [p1]
    if p_outbreak:
        parts.append(p_outbreak)
    parts += [p2, p3]
    return "\n\n".join(parts)

# ---------------------------------------------------------------------------
# Pathogen-specific process-engineering narratives (P2 dispatcher).
# Each entry returns the W16-style paragraph: product category → failure modes
# → relevant regulatory framework. Mirrors the W16 gold-standard specificity.
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# Jurisdiction qualifier for the fallback pathogen narratives.
#
# REVIEW 2026-08-07 (W32), second-order finding. The reviewer flagged the
# botulinum Process Authority Note for wearing US law under EU citations. The
# same defect is live one paragraph up: every _pathogen_narrative() below
# closes with "The relevant frameworks are <US CFR citations>", unconditionally
# — so a week that was 54% France read as though 21 CFR 117 governed French
# raw-milk cheese.
#
# The engineering content of those paragraphs (product classes, failure modes)
# is jurisdiction-neutral and correct, so it is kept. What is added is an
# explicit statement of WHOSE law is being cited and where the equivalent duty
# sits when the week is not US-dominated. Rewriting twelve narratives with
# per-jurisdiction legal text is a bigger change than this defect warrants,
# and each rewrite would need its own verification.
#
# This affects the DETERMINISTIC FALLBACK only. In CI the P2 paragraph is
# written by Claude with the week's jurisdictions in the prompt.
# ---------------------------------------------------------------------------
_EQUIVALENT_FRAMEWORK = {
    "EU": "Regulation (EC) No 852/2004 (hygiene and HACCP) and Regulation (EC) "
          "No 2073/2005 (microbiological criteria), with national "
          "competent-authority oversight",
    "UK": "retained Regulation (EC) No 2073/2005, the UK Food Hygiene "
          "Regulations and FSA guidance",
    "CA": "the CFIA Safe Food for Canadians Regulations and the Preventive "
          "Control Plan",
    "AU_NZ": "the FSANZ Food Standards Code (Standards 1.6.1, 3.2.1 and 3.2.2)",
    "JP": "the Japan Food Sanitation Act and MHLW standards",
    "KR": "the MFDS Food Sanitation Act and Food Code",
}

_JURISDICTION_BY_SOURCE = {
    "FDA": "US", "USDA FSIS": "US", "CDC": "US", "FDA PH": "US",
    "RASFF (EU)": "EU", "RappelConso (FR)": "EU", "AESAN (ES)": "EU",
    "BVL (DE)": "EU", "FSAI (IE)": "EU", "FAVV (BE)": "EU", "EFET (GR)": "EU",
    "Ministero della Salute": "EU", "AGES (AT)": "EU", "NVWA (NL)": "EU",
    "EFSA": "EU", "ECDC": "EU",
    "FSA (UK)": "UK", "FSS (UK)": "UK",
    "CFIA": "CA",
    "FSANZ (AU)": "AU_NZ", "MPI (NZ)": "AU_NZ",
}


def _dominant_jurisdiction(recalls):
    """The jurisdiction most of this week's incidents belong to, or None.

    Sources not in the map are ignored rather than guessed — an unmapped
    regulator must not silently vote for a legal framework.
    """
    counts = {}
    for r in (recalls or []):
        key = _JURISDICTION_BY_SOURCE.get(str(r.get("Source") or "").strip())
        if key:
            counts[key] = counts.get(key, 0) + 1
    if not counts:
        return None
    return max(counts.items(), key=lambda kv: (kv[1], kv[0]))[0]


def _framework_jurisdiction_qualifier(recalls):
    """The sentence that stops a US citation from reading as local law."""
    key = _dominant_jurisdiction(recalls)
    if not key or key == "US":
        return ""
    equivalent = _EQUIVALENT_FRAMEWORK.get(key)
    if not equivalent:
        return ""
    return (" The frameworks cited above are United States instruments, given "
            "here as the most fully specified statement of the control. The "
            "majority of this week's incidents fall outside that jurisdiction, "
            "where the equivalent duties sit in {}.".format(equivalent))


_COMMODITY_CLASSES = (
    # (label, regex) — ordered, first match wins. Deliberately coarse: this
    # exists to stop the narrative naming commodity classes that are not in
    # the data, not to be a taxonomy.
    ("meat and poultry",
     r"chicken|poultry|poulet|pollo|turkey|dinde|duck|pork|porc|beef|veau|"
     r"veal|\blamb\b|mutton|sausage|merguez|salami|pastrami|charcuter|\bham\b|"
     r"bacon|\bmeat\b"),
        # Label said "sprouts" while no sprout incident was in the set it
    # described (external review 2026-08-14). A class label reads as a
    # description of contents; it must not advertise a member that is
    # not there.
    ("seeds and sesame",
     r"sesame|tahini|sprout|\bseeds?\b|linseed|sunflower seed|chia"),
    ("nuts and nut butters",
     r"peanut|almond|cashew|pistachio|hazelnut|walnut|nut butter"),
    ("eggs and egg products", r"\begg"),
    ("cheese and dairy",
     r"cheese|fromage|brie|camembert|cheddar|milk|lait|dairy|yog|cream|"
     r"butter\b|queso"),
    # WORD BOUNDARIES ARE LOAD-BEARING HERE. Without \b on "salmon" this
    # class swallowed every Salmonella row that no earlier class matched,
    # because "SALMONella" contains "salmon". The first build printed
    # "fish and seafood 5/26 (19%)" for a set that was a mango, a spice
    # mix, a green powder and two jalapeño salsa recalls. Caught by
    # listing the rows behind the number instead of trusting it.
    ("fish and seafood",
     r"\bfish\b|poisson|\bsalmon\b|\btrout\b|truite|\btuna\b|shrimp|crevette|"
     r"\bprawn|oyster|mussel|\bclam\b|molluscs?|seafood|langoustine|\bcrab\b"),
    ("fresh produce",
     r"salad|lettuce|leafy|spinach|jalape|salsa|guacamole|pico de gallo|"
     r"tomato|mango|melon|berry|berries|fruit|vegetable|legume|mushroom|"
     r"onion|herb|cucumber|carrot|sprouted"),
    ("spices and dried herbs",
     r"spice|paprika|curry|cinnamon|pepper corn|oregano|basil|dried herb|"
     r"seasoning"),
    ("low-moisture and bakery",
     r"flour|cereal|granola|\boat|rice|pasta|noodle|powder|infant formula|"
     r"biscuit|cracker|bread|pastry|cake|chocolate|cocoa"),
)


def _commodity_mix(recalls, pathogen, top_n=5):
    """Measured commodity distribution for `pathogen` in THIS week's rows.

    Returns a human-readable string like
        "meat and poultry 15/27 (56%), seeds, sesame and sprouts 5/27 (19%)"
    or "" when there is nothing to measure.

    Added 2026-08-14. Before this, both the AI prompt and the deterministic
    fallback asserted a commodity pattern from a hardcoded list, with no
    reference to the rows being summarised. In W33 that published "dairy and
    soft cheese" and "peanut butter, flour, infant formula" for a Salmonella
    signal containing zero dairy, zero cheese and zero low-moisture rows.
    """
    import re as _re_c
    p = (pathogen or "").strip().lower()
    if not p:
        return ""
    rows = [r for r in (recalls or [])
            if p.split()[0] in str(r.get("Pathogen") or "").lower()]
    if not rows:
        return ""
    counts = Counter()
    for r in rows:
        blob = " ".join(str(r.get(k) or "") for k in
                        ("Product", "Reason", "Company", "Brand")).lower()
        for label, pattern in _COMMODITY_CLASSES:
            if _re_c.search(pattern, blob):
                counts[label] += 1
                break
        else:
            counts["other or unclassified"] += 1
    n = len(rows)
    # THE PARTS MUST SUM TO n. The first version returned
    # counts.most_common(top_n) and nothing else, so W33 published
    # "meat and poultry 15/26 (58%), seeds ... 5/26 (19%), fresh produce
    # 3/26 (12%), eggs 1/26 (4%)" — four classes totalling 24 of 26,
    # presented as if it were the whole breakdown. An external review did
    # the arithmetic and caught it. Silent truncation dressed as a
    # complete accounting is the same failure this codebase keeps hitting
    # elsewhere; a tail bucket makes it impossible.
    shown = counts.most_common(top_n)
    parts = [f"{label} {c}/{n} ({round(c / n * 100)}%)" for label, c in shown]
    remainder = n - sum(c for _, c in shown)
    if remainder > 0:
        parts.append(f"other categories {remainder}/{n} "
                     f"({round(remainder / n * 100)}%)")
    assert sum(c for _, c in shown) + max(remainder, 0) == n
    return ", ".join(parts)


def _pathogen_narrative(pathogen, pct, commodity_mix=""):
    p = (pathogen or "").lower()
    # The pct token gives us "at this prevalence" / "at this concentration" flavour
    intensity = "at this concentration" if pct >= 30 else "at this prevalence"

    if "listeria" in p:
        return ("Listeria monocytogenes {intensity} points to post-process recontamination "
                "in ready-to-eat deli, dairy, and cooked-meat lines rather than thermal "
                "underprocess. The likely failure modes are Zone 1 environmental harbourage, "
                "sanitation SOP drift, and post-lethality recontamination. The relevant "
                "frameworks for review are the environmental monitoring programme under "
                "21 CFR 117 and the thermal-lethality validation applicable to the product "
                "class (21 CFR 113 / 114 where applicable), supported by qualified process "
                "authority oversight.").format(intensity=intensity)

    if "salmonella" in p:
        # AUDIT 2026-08-14 \u2014 this branch used to assert a fixed commodity
        # list ("produce, meat and poultry, dairy and soft cheese, nuts and
        # seeds, spices and dried herbs, and low-moisture foods (peanut
        # butter, flour, infant formula)") no matter what the week actually
        # contained. In W33 it published "dairy and soft cheese" for a
        # 27-row Salmonella signal holding zero dairy and zero cheese rows,
        # and cited peanut butter, flour and infant formula with none of
        # them present. It now states the measured mix, and falls back to a
        # commodity-free sentence rather than inventing one.
        if commodity_mix:
            spread = ("This week's Salmonella signal was concentrated in "
                      "{mix}. That distribution is consistent with multiple "
                      "independent supply-chain contamination events rather "
                      "than a single commodity-level failure."
                      ).format(mix=commodity_mix)
        else:
            spread = ("Salmonella {intensity} is consistent with "
                      "raw-ingredient contamination entering through "
                      "supplier-side controls."
                      ).format(intensity=intensity)
        return (spread + " The common-cause pattern points to gaps in "
                "supplier-verification, post-process recontamination control, "
                "validated kill-step parameters, and raw-material handling "
                "hygiene. The relevant frameworks are 21 CFR 117 Preventive "
                "Controls, HACCP critical limits on any validated kill step, "
                "supplier-verification programmes, and the FDA Produce Safety "
                "Rule (21 CFR 112) where applicable.")

    if "stec" in p or "e. coli" in p or "escherichia" in p:
        return ("E. coli / STEC {intensity} is most consistent with raw beef, leafy greens, "
                "raw milk/dairy, sprouts, or unpasteurised juice. The likely failure modes "
                "are grinding-plant cross-contamination, irrigation-water contamination, "
                "post-harvest wash-water carry-over, and sprout-seed microbial load. The "
                "relevant frameworks are 21 CFR 117, the FDA Produce Safety Rule (21 CFR 112), "
                "and USDA FSIS zero-tolerance for E. coli O157:H7 in non-intact beef.").format(intensity=intensity)

    if "botulinum" in p or "clostridium" in p:
        return ("Clostridium botulinum {intensity} points directly to a scheduled-thermal-"
                "process failure in a shelf-stable low-acid canned food (LACF), an acidified "
                "food, or a reduced-oxygen-packaged product. The likely failure modes are an "
                "unfiled or outdated scheduled process, a process deviation resolved without "
                "qualified review, container / seal integrity loss, or a formulation change "
                "(pH, a_w, salt, preservative) introduced without re-evaluation. The relevant "
                "frameworks are 21 CFR 113 (LACF), 21 CFR 114 (acidified foods), and the "
                "scheduled-process filing requirements under 21 CFR 108.").format(intensity=intensity)

    if "cereulide" in p or "bacillus cereus" in p:
        return ("Cereulide / Bacillus cereus {intensity} indicates temperature abuse during "
                "cooling or holding of cooked rice, pasta, infant formula, or dairy-based "
                "products. The emetic toxin is heat-stable and is not inactivated by reheating, "
                "so controls must target the cooling-rate critical limit and the cooked-product "
                "hot-hold / chill-hold regime. The relevant frameworks are 21 CFR 117 time-"
                "temperature critical limits and FDA Food Code cooling provisions.").format(intensity=intensity)

    if "norovirus" in p or "hepatitis a" in p or "hav" in p:
        return ("Norovirus / Hepatitis A {intensity} is most consistent with contamination by "
                "an infected food handler or with contaminated raw molluscan shellfish or "
                "ready-to-eat soft fruit. The likely failure modes are sick-worker exclusion "
                "failure, hand-hygiene breakdown, and supplier controls on shellfish-harvest "
                "waters. The relevant frameworks are the FDA Food Code employee-health "
                "provisions and the National Shellfish Sanitation Program.").format(intensity=intensity)

    if "histamine" in p or "scombroid" in p:
        return ("Histamine (scombrotoxin) {intensity} indicates a chill-chain breach on "
                "histidine-rich species (tuna, mackerel, mahi-mahi, sardines, anchovies, "
                "bonito). Histamine is heat-stable and is not removed by cooking or canning, "
                "so the control point is time-temperature at harvest, landing, and the full "
                "cold chain. The relevant framework is 21 CFR 123 (Seafood HACCP) with FDA "
                "action levels of 50 ppm (decomposition) and 500 ppm (hazard).").format(intensity=intensity)

    if "rodenticide" in p or "rat poison" in p or "bromadiolone" in p:
        return ("Rodenticide {intensity} points to a supply-chain integrity or intentional-"
                "adulteration event — a criminal tampering profile rather than a process "
                "failure. The likely failure modes are a compromised inbound-ingredient chain, "
                "a breach in packaging integrity between producer and retailer, or insider "
                "tampering. The relevant frameworks are 21 CFR 121 (FSMA Intentional "
                "Adulteration Rule) for food-defense vulnerability assessment and the "
                "corresponding EU Commission Regulation on food fraud.").format(intensity=intensity)

    if "rodent" in p or "pest" in p:
        return ("Rodent / pest contamination {intensity} indicates a sanitation and GMP "
                "programme failure — pest-management, facility integrity, or raw-material "
                "storage hygiene. The likely failure modes are a lapsed integrated-pest-"
                "management contract, structural entry points, or un-segregated storage of "
                "open-package ingredients. The relevant frameworks are 21 CFR 117 subpart B "
                "(sanitation / pest control) and HACCP GMP prerequisites.").format(intensity=intensity)

    if any(k in p for k in ("glass", "metal", "plastic", "foreign")):
        return ("Foreign-material contamination {intensity} indicates an equipment-wear or "
                "packaging-integrity failure along the line. The likely failure modes are "
                "sieve / filter breakage, equipment fatigue on contact surfaces, or a lapse "
                "in metal-detection or X-ray inspection validation. The relevant frameworks "
                "are HACCP physical-hazard critical control points and the routine re-"
                "qualification of detection equipment at its declared sensitivity.").format(intensity=intensity)

    if any(k in p for k in ("lead", "cadmium", "arsenic", "mercury", "heavy metal")):
        return ("Heavy-metal contamination {intensity} is most consistent with raw-material "
                "sourcing from contaminated soils or waters, or with leaching from processing "
                "equipment and packaging. The likely failure modes are supplier-verification "
                "gaps on agricultural inputs, unmonitored legacy equipment, or incompatible "
                "food-contact materials. The relevant frameworks are FDA action / guidance "
                "levels (Closer to Zero for infant foods), EU Reg. 2023/915 contaminant "
                "maximum levels, and Codex Alimentarius MLs.").format(intensity=intensity)

    if any(k in p for k in ("aflatoxin", "ochratoxin", "patulin", "alternaria", "mycotoxin", "mould", "mold")):
        return ("Mycotoxin contamination {intensity} indicates a post-harvest moisture-control "
                "failure in grains, nuts, dried fruit, or spices, or a storage-humidity "
                "breach along the supply chain. The likely failure modes are inadequate "
                "drying, compromised storage-silo integrity, or supplier-verification gaps "
                "on high-risk commodities. The relevant frameworks are FDA action levels on "
                "aflatoxins, EU Reg. 2023/915 maximum levels, and HACCP supplier-approval "
                "programmes.").format(intensity=intensity)

    # Fallback — keep the old generic line if pathogen not recognised
    return ("{} {} warrants review of the kill step, post-kill recontamination controls, "
            "and supplier-verification programme for the implicated commodity. The relevant "
            "framework is 21 CFR 117 Preventive Controls with HACCP critical limits "
            "appropriate to the product class.").format(pathogen, intensity)


# ---------------------------------------------------------------------------
# P3 — data-driven jurisdiction paragraph.
# Maps Source labels present in the week's data to their authority names.
# ---------------------------------------------------------------------------
_SOURCE_TO_AUTHORITY = {
    "fda": "FDA", "usda fsis": "USDA FSIS", "usda": "USDA FSIS", "cdc": "CDC",
    "cfia": "CFIA", "mapaq": "MAPAQ",
    "rappelconso": "RappelConso", "dgccrf": "DGCCRF", "dgal": "DGAL",
    "fsa": "FSA", "fss": "FSS", "fsai": "FSAI",
    "rasff": "RASFF", "efsa": "EFSA", "dg sante": "DG SANTE",
    "aesan": "AESAN", "bvl": "BVL", "bfr": "BfR", "ages": "AGES",
    "min. salute": "Italian Ministry of Health", "nvwa": "NVWA",
    "favv": "FAVV", "fødevarestyrelsen": "Fødevarestyrelsen",
    "livsmedelsverket": "Livsmedelsverket", "mattilsynet": "Mattilsynet",
    "ruokavirasto": "Ruokavirasto", "gis": "GIS", "nebih": "NEBIH",
    "ansvsa": "ANSVSA", "bfsa": "BFSA", "szpi": "SZPI", "švps": "ŠVPS",
    "svps": "ŠVPS", "mast": "MAST", "blv": "BLV",
    "fsanz": "FSANZ", "mpi nz": "MPI NZ",
    "cfs": "CFS Hong Kong", "mfds": "MFDS", "mhlw": "MHLW",
    "samr": "SAMR", "sfa": "SFA", "fssai": "FSSAI",
    "anvisa": "ANVISA", "cofepris": "COFEPRIS", "invima": "INVIMA",
    "anmat": "ANMAT", "arcsa": "ARCSA", "digesa": "DIGESA", "isp": "ISP",
    "sfda": "SFDA", "moccae": "MOCCAE", "moh": "MOH", "moph": "MOPH",
    "nafdac": "NAFDAC", "kebs": "KEBS", "nfsa": "NFSA", "onssa": "ONSSA",
}

def _jurisdictions_from_recalls(recalls, max_list=5):
    """Return a short list of authority names present in this week's data."""
    seen = []
    for r in recalls:
        src = (r.get("Source") or "").lower().strip()
        if not src: continue
        # Match first token or exact map
        for key, auth in _SOURCE_TO_AUTHORITY.items():
            if key in src and auth not in seen:
                seen.append(auth); break
    return seen[:max_list]


# ---------------------------------------------------------------------------
# The botulinum duty clause, per jurisdiction.
#
# REVIEW 2026-08-07 (W32). The note published for W32 read:
#
#   "... confirm that the scheduled thermal process has been established,
#    FILED, and validated under the guidance of a QUALIFIED PROCESS
#    AUTHORITY — REQUIRED under Regulation (EC) No 852/2004 and Regulation
#    (EC) No 2073/2005 ..."
#
# That sentence is correct US law and incorrect EU law. "Filing" a scheduled
# process and using a designated "process authority" are creatures of the US
# low-acid / acidified framework — 21 CFR 108 (Form 2541/2541e), 113 and 114.
# Neither cited EU Regulation creates a general scheduled-process filing
# regime or requires a US-style process authority:
#
#   · Reg. 852/2004 requires HACCP-based controls — hazard analysis, critical
#     limits, monitoring, corrective action, verification, documentation — and
#     in Annex II Chapter XI requires that heat treatment of food in
#     hermetically sealed containers achieve the intended time/temperature
#     objective, prevent post-process contamination, be routinely controlled
#     on the relevant parameters (temperature, pressure, sealing,
#     microbiology), and conform to an internationally recognised standard.
#   · Reg. 2073/2005 sets microbiological criteria and the associated
#     validation/verification duties inside HACCP. It creates no filing regime.
#
# The defect was structural, not a typo: ONE sentence carried a US-shaped
# obligation and the citation was swapped per jurisdiction underneath it. So
# the obligation itself is now per-jurisdiction, and the citation follows it.
#
# Each entry is (duty, gaps, records):
#   duty    — what the operator must be able to demonstrate
#   gaps    — the failure modes behind recalls of this profile there
#   records — what an inspector asks to see afterwards
# ---------------------------------------------------------------------------
_BOTULINUM_DUTY = {
    "US": (
        "the scheduled thermal process has been established and validated by a "
        "qualified process authority and filed with the competent authority, "
        "and that critical factors are controlled and documented as specified "
        "in that process",
        "an unfiled or outdated scheduled process, a process deviation "
        "resolved without qualified process authority review, a container or "
        "seal-integrity lapse, or a formulation change (pH, a_w, salt, "
        "preservative) introduced without re-evaluation",
        "scheduled-process filings, deviation handling, container integrity, "
        "and process-authority records"),
    "EU": (
        "the thermal process and its critical factors have been scientifically "
        "established, validated, implemented and documented within the HACCP "
        "system. Heat treatment applied to food in hermetically sealed "
        "containers must achieve the intended time/temperature objective, "
        "prevent post-process contamination, and be routinely controlled "
        "through the relevant parameters — temperature, pressure, container "
        "sealing and microbiological checks — using a process conforming to an "
        "internationally recognised standard. Applicable microbiological "
        "criteria and the associated validation and verification duties should "
        "also be considered",
        "a thermal process never validated against the actual formulation, "
        "container and fill conditions, a deviation closed without documented "
        "technical review, a container or seal-integrity lapse, or a "
        "formulation change (pH, a_w, salt, preservative) introduced without "
        "re-evaluation of the hazard analysis",
        "HACCP validation records, heat-treatment monitoring and calibration "
        "records, container-seal integrity checks, and deviation handling"),
    "UK": (
        "the thermal process and its critical factors have been scientifically "
        "established, validated and documented within the HACCP system, with "
        "heat treatment of hermetically sealed containers controlled on "
        "temperature, pressure and seal integrity against a recognised process "
        "standard",
        "a process not validated for the actual formulation and container, a "
        "deviation closed without documented technical review, a seal-integrity "
        "lapse, or an unreviewed formulation change",
        "HACCP validation records, heat-treatment monitoring records, seal "
        "integrity checks, and deviation handling"),
    "CA": (
        "the thermal process has been scientifically established and validated "
        "and is documented within the Preventive Control Plan, with critical "
        "factors monitored and deviations handled under it",
        "a process not validated for the product as made, a deviation closed "
        "outside the PCP, a seal-integrity lapse, or an unreviewed formulation "
        "change",
        "the Preventive Control Plan, process validation evidence, container "
        "integrity checks, and deviation records"),
    "AU_NZ": (
        "the thermal process has been validated and is documented within the "
        "food safety programme, with critical limits monitored and deviations "
        "handled under it",
        "a process not validated for the product as made, a deviation closed "
        "outside the food safety programme, a seal-integrity lapse, or an "
        "unreviewed formulation change",
        "the food safety programme, process validation evidence, container "
        "integrity checks, and deviation records"),
    "JP": (
        "the thermal process meets the applicable standards for shelf-stable "
        "and canned foods and that its establishment, validation and routine "
        "control are documented",
        "a process not validated for the product as made, an undocumented "
        "deviation, a seal-integrity lapse, or an unreviewed formulation "
        "change",
        "process validation evidence, heat-treatment records, container "
        "integrity checks, and deviation records"),
    "KR": (
        "the thermal process has been validated against the applicable "
        "requirements and that its establishment and routine control are "
        "documented",
        "a process not validated for the product as made, an undocumented "
        "deviation, a seal-integrity lapse, or an unreviewed formulation "
        "change",
        "process validation evidence, heat-treatment records, container "
        "integrity checks, and deviation records"),
}


def _process_authority_note(recalls, bot):
    """Return the P4 Process Authority Note string, or '' if no trigger fires.

    Multi-trigger system. A PA Note fires whenever ANY of these hazard-classes
    appear in the week. Each trigger gets its own fully-specified
    process-authority paragraph with the correct GMPs, validation framework,
    and regulatory citations — keyed to the JURISDICTIONS actually present in
    the implicated recalls, not boilerplate FDA citations.

    Triggers (evaluated in severity order; first match wins):
      1. Clostridium / botulinum   - scheduled-thermal-process / LACF filing
      2. Listeria monocytogenes    - RTE sanitation / env-monitoring / validated lethality
      3. Salmonella in low-moisture - kill-step validation, wet-dry segregation
      4. Salmonella (other)        - kill-step + supplier verification
      5. E. coli STEC              - raw-material supplier verification, kill-step
      6. Multi-jurisdiction outbreak OR Tier-1 >=3 across >=3 countries - regulatory-response PA
    """
    if not recalls:
        return ""

    # Helper: recalls implicating a specific pathogen
    def _by_pathogen(*needles):
        out = []
        for r in recalls:
            p = (r.get("Pathogen") or "").lower()
            if any(n in p for n in needles):
                out.append(r)
        return out

    # Low-moisture food heuristic (Salmonella trigger needs both signals)
    def _is_low_moisture(r):
        text = ((r.get("Product") or "") + " " + (r.get("Reason") or "") + " " +
                (r.get("Company") or "")).lower()
        return any(k in text for k in (
            "peanut", "nut butter", "almond", "cashew", "pistachio", "hazelnut",
            "flour", "cereal", "granola", "oat", "rice", "pasta", "grain",
            "powder", "powdered", "infant formula", "formula", "milk powder",
            "spice", "herb", "seasoning", "tea", "dried", "chocolate", "cocoa",
            "seed", "tahini", "sesame",
        ))

    def _names(rows, limit=3):
        # Dedupe by company so multi-SKU / multi-fiche recalls from the same
        # producer (e.g. two RappelConso fiches) aren't cited twice.
        #
        # REVIEW 2026-08-07 — a RASFF row's Company is the database convention
        # "Origin: X | Notifying: Y", not a firm. Dropped verbatim into the
        # Process Authority Note it read:
        #
        #   "with Origin: Czechia | Notifying: Czechia (Czechia) cited for
        #    Clostridium botulinum"
        #
        # which is a schema leaking into a regulatory paragraph. RASFF genuinely
        # does not name the operator, so the honest rendering is the
        # notification itself.
        seen = []
        for r in rows:
            company = str(r.get("Company") or "").strip()
            country = _country_display(r.get("Country", ""))
            m = re.match(r"^\s*Origin:\s*(.+?)\s*\|\s*Notifying:\s*(.+?)\s*$",
                         company)
            if m:
                origin, notifying = m.group(1), m.group(2)
                label = ("a RASFF notification from {}".format(notifying)
                         if origin.lower() == notifying.lower() else
                         "a RASFF notification from {} concerning product of "
                         "{} origin".format(notifying, origin))
            elif company:
                label = "{} ({})".format(company, country)
            else:
                label = country
            if label not in seen:
                seen.append(label)
            if len(seen) >= limit:
                break
        return ", ".join(seen)

    # ------------------------------------------------------------------
    # Jurisdiction-aware framework picker.
    # Examines the country codes of the implicated rows and returns the
    # appropriate framework citations ordered by dominant jurisdiction.
    # Each hazard trigger calls this to build its regulatory-citation block.
    # ------------------------------------------------------------------
    def _country_buckets(rows):
        us = eu = uk = ca = au_nz = jp = kr = other = 0
        countries = []
        EU_MEMBERS = {
            "France","Germany","Italy","Spain","Netherlands","Belgium","Ireland",
            "Denmark","Sweden","Finland","Austria","Poland","Portugal","Greece",
            "Czech Republic","Slovakia","Hungary","Romania","Bulgaria","Croatia",
            "Slovenia","Lithuania","Latvia","Estonia","Luxembourg","Malta","Cyprus",
        }
        for r in rows:
            c = (r.get("Country") or "").strip()
            if c and c not in countries:
                countries.append(c)
            cl = c.lower()
            if cl in ("united states", "usa", "us"):        us += 1
            elif c in EU_MEMBERS or cl == "european union": eu += 1
            elif cl in ("united kingdom", "uk", "scotland", "england", "wales", "northern ireland"): uk += 1
            elif cl == "canada":                             ca += 1
            elif cl in ("australia", "new zealand"):         au_nz += 1
            elif cl == "japan":                              jp += 1
            elif cl == "south korea":                        kr += 1
            else:                                            other += 1
        return {"US":us,"EU":eu,"UK":uk,"CA":ca,"AU_NZ":au_nz,"JP":jp,"KR":kr,"OTHER":other,
                "countries":countries}

    def _regs_for(hazard, rows):
        """Return a dict of framework-citation strings keyed to the jurisdictions
        actually present. `hazard` is one of: botulinum, listeria, salmonella_lm,
        salmonella, stec, regulatory.
        The returned dict has keys: primary (main framework sentence),
        parallels (comma-joined list of parallel-jurisdiction cites),
        typical_regulators (space-separated authority names for the closing sentence)."""
        b = _country_buckets(rows)

        # Per-hazard framework text per jurisdiction
        FRAMEWORKS = {
            "botulinum": {
                "US": "FDA 21 CFR 113 (LACF) / 21 CFR 114 (acidified foods) with scheduled-process filing under 21 CFR 108 (Form 2541 / 2541e)",
                "EU": "Regulation (EC) No 852/2004 (food hygiene) and Regulation (EC) No 2073/2005 (microbiological criteria), with national competent-authority oversight of low-acid / shelf-stable production",
                "UK": "UK Food Safety Act 1990, Food Hygiene (England) Regulations 2013 / retained Reg. 852/2004 + 2073/2005, and FSA Food Standards Agency oversight",
                "CA": "CFIA Safe Food for Canadians Regulations (SFCR) with Preventive Control Plan (PCP) requirements for low-acid foods",
                "AU_NZ": "FSANZ Food Standards Code Chapter 3 (Standard 3.2.1 Food Safety Programs, 3.2.2 Practices and General Requirements)",
                "JP": "Japan Food Sanitation Act and MHLW standards for shelf-stable and canned foods",
                "KR": "MFDS Food Sanitation Act with thermal-process validation requirements",
            },
            "listeria": {
                # AUDIT 2026-08-14 — CPG 555.320 REMOVED.
                # It was cited here as an applicable US framework and
                # glossed as "L. monocytogenes zero-tolerance in RTE".
                # Both halves are wrong on FDA's own page, which labels the
                # document "Draft - Not for Implementation", states it
                # "Contains non-binding recommendations", and says FDA
                # guidance documents "do not establish legally enforceable
                # responsibilities". It also does NOT set a blanket
                # zero-tolerance: it separates RTE foods that SUPPORT
                # L. monocytogenes growth (detection triggers action) from
                # those that do not (action at >=100 CFU/g). Citing a draft
                # as binding, and mis-stating what the draft says, in a
                # briefing food manufacturers act on is the kind of error
                # that costs credibility. Replaced with the binding
                # instruments: 21 CFR 117 preventive controls and
                # environmental monitoring where an environmental pathogen
                # is a hazard requiring a preventive control, plus FSIS
                # Directive 10,240.4 for post-lethality-exposed RTE
                # meat/poultry, which is where a real zero-tolerance
                # style regime does apply.
                "US": "FDA 21 CFR 117 Subparts B and G (Preventive Controls), including environmental monitoring where contamination of an RTE food with an environmental pathogen is a hazard requiring a preventive control, and USDA FSIS Directive 10,240.4 where post-lethality-exposed RTE meat/poultry is implicated",
                # AUDIT 2026-08-14 — "national RASFF notification
                # obligations" removed. RASFF is the rapid alert network
                # operated BETWEEN competent authorities and the
                # Commission; food business operators do not notify RASFF.
                # Their duty runs to the competent authority, and the
                # authority then decides what enters RASFF.
                "EU": "Regulation (EC) No 2073/2005 as amended by Reg. (EU) 2024/2895 (applicable from 1 July 2026): for RTE foods able to support growth of Listeria monocytogenes, operators must demonstrate compliance throughout shelf life; where a business cannot demonstrate to the competent authority that levels remain ≤100 CFU/g, the applicable criterion is not detected in 25 g. Regulation (EC) No 852/2004 HACCP requirements apply, and operators must notify and cooperate with the national competent authority, which is the body that issues any RASFF notification",
                "UK": "retained EU Reg. 2073/2005, UK Food Hygiene Regulations, and FSA listeria-in-RTE control guidance",
                "CA": "CFIA Policy on Control of Listeria monocytogenes in Ready-to-Eat Foods and the SFCR Preventive Control Plan",
                "AU_NZ": "FSANZ Food Standards Code Standard 1.6.1 (microbiological limits) and industry Listeria management guidelines",
                "JP": "Japan MHLW microbial standards and ready-to-eat food hygiene guidance",
                "KR": "MFDS RTE pathogen standards",
            },
            "salmonella_lm": {
                "US": "FDA 21 CFR 117 Subpart C Preventive Controls with FDA Guidance for Industry: Measures to Address the Risk for Contamination by Salmonella in Low-Moisture Ready-to-Eat Human Foods",
                "EU": "EU Reg. 2073/2005 microbiological criteria (Salmonella absence in 25 g for RTE categories) and Reg. 852/2004 HACCP",
                "UK": "retained EU Reg. 2073/2005 and FSA Salmonella low-moisture food guidance",
                "CA": "CFIA SFCR Preventive Control Plan with validated kill-step requirements for low-moisture foods",
                "AU_NZ": "FSANZ Food Standards Code Standards 1.6.1 and 3.2.2 with the FSANZ Getting Your Food Regulation Right guidance",
                "JP": "Japan MHLW microbial standards",
                "KR": "MFDS low-moisture food hygiene standards",
                "CODEX": "Codex CAC/RCP 75-2015 Code of Hygienic Practice for Low-Moisture Foods",
            },
            "salmonella": {
                "US": "FDA 21 CFR 117 Preventive Controls and, for meat/poultry, USDA FSIS performance standards",
                "EU": "EU Reg. 2073/2005 (Salmonella absence in 25 g for most RTE categories) and Reg. 852/2004 HACCP",
                "UK": "retained EU Reg. 2073/2005 and FSA / Food Hygiene Regulations",
                "CA": "CFIA SFCR Preventive Control Plan",
                "AU_NZ": "FSANZ Food Standards Code Standard 3.2.1 Food Safety Programs and Standard 1.6.1 microbiological limits",
                "JP": "Japan MHLW microbial standards",
                "KR": "MFDS Food Code pathogen limits",
            },
            "stec": {
                "US": "USDA FSIS adulterant declarations for E. coli O157:H7 and the Big-Six non-O157 STECs in non-intact raw beef (9 CFR 318/381), FDA Produce Safety Rule (21 CFR 112), and FDA Juice HACCP rule (21 CFR 120)",
                "EU": "EU Reg. 2073/2005 (STEC limits for RTE sprouted seeds), Reg. 852/2004 HACCP, and national STEC surveillance under Directive 2003/99/EC",
                "UK": "retained EU Reg. 2073/2005 and FSA STEC guidance with sprout-specific controls",
                "CA": "CFIA SFCR Preventive Control Plan with specific E. coli O157:H7 controls for ground beef",
                "AU_NZ": "FSANZ Food Standards Code Standard 1.6.1 (STEC limits in raw unpasteurised milk and cheese) and Standard 4.2.1 primary production",
                "JP": "Japan MHLW STEC control guidance",
                "KR": "MFDS STEC pathogen standards",
            },
            "regulatory": {
                "US": "FSMA §204 traceability (21 CFR 1 Subpart S) with Class I recall effectiveness checks",
                "EU": "EU Reg. 178/2002 Art. 18 traceability one-up / one-back and RASFF INFOSAN rapid-alert obligations",
                "UK": "retained Reg. 178/2002 Art. 18 and FSA Food Law Enforcement Code of Practice",
                "CA": "CFIA SFCR traceability and the Canadian food-recall framework under the Safe Food for Canadians Act",
                "AU_NZ": "FSANZ Food Standards Code 3.2.2 traceability and the Australian Consumer Law mandatory-reporting obligation",
                "JP": "Japan Food Sanitation Act notification obligations",
                "KR": "MFDS recall and traceability provisions",
            },
        }[hazard]

        # Determine ordering. The bucket with the most incidents comes first.
        order = sorted(
            [k for k in ("US","EU","UK","CA","AU_NZ","JP","KR") if b[k] > 0],
            key=lambda k: -b[k])
        if not order:
            # All recalls in "other" countries — default to global framing
            order = ["EU"]  # RASFF is the most common fallback framework

        primary_key = order[0]
        primary_text = FRAMEWORKS[primary_key]

        parallel_keys = order[1:]
        parallels_text = ""
        if parallel_keys:
            parallels_text = ". Parallel frameworks apply under " + "; ".join(
                FRAMEWORKS[k] for k in parallel_keys)

        # Codex as a supra-national reference for low-moisture Salmonella
        if hazard == "salmonella_lm":
            if "CODEX" in FRAMEWORKS:
                if parallels_text:
                    parallels_text += "; supranational reference: " + FRAMEWORKS["CODEX"]
                else:
                    parallels_text = ". Supranational reference: " + FRAMEWORKS["CODEX"]

        # Regulator names for the closing sentence
        AUTHORITY_NAMES = {
            "US":"FDA and USDA FSIS", "EU":"the national competent authority (with RASFF notification)",
            "UK":"the FSA", "CA":"CFIA", "AU_NZ":"FSANZ",
            "JP":"Japan MHLW", "KR":"MFDS",
        }
        authorities = [AUTHORITY_NAMES[k] for k in order]
        if len(authorities) == 1:
            regulators = authorities[0]
        elif len(authorities) == 2:
            regulators = " and ".join(authorities)
        else:
            regulators = ", ".join(authorities[:-1]) + ", and " + authorities[-1]

        return {"primary": primary_text,
                "parallels": parallels_text,
                "regulators": regulators,
                "primary_key": primary_key,
                "buckets": b}

    # ------------------------------------------------------------------
    # Trigger 1 — Clostridium / botulinum (highest severity)
    # ------------------------------------------------------------------
    if bot:
        regs = _regs_for("botulinum", bot)
        duty, gaps, records = _BOTULINUM_DUTY.get(
            regs["primary_key"], _BOTULINUM_DUTY["EU"])
        return ("This window contains {ip} implicating Clostridium or "
                "botulinum toxin, with {co} cited for {path}. Any shelf-stable "
                "low-acid, acidified, aseptic/UHT, hot-filled, or reduced-oxygen-"
                "packaged product in the affected category should be reviewed to "
                "confirm that {duty} \u2014 {primary}{parallels}. The most "
                "common compliance gaps behind recalls of this profile are {gaps}. "
                "Tier-1 / Class-I classification on a product of this class may "
                "trigger targeted regulatory follow-up by {regulators}, including "
                "review of {records} on the subsequent inspection."
                ).format(ip=_count_phrase(len(bot), "incident"), co=_names(bot),
                         path=bot[0].get("Pathogen","Clostridium botulinum"),
                         duty=duty, gaps=gaps, records=records,
                         primary=regs["primary"], parallels=regs["parallels"],
                         regulators=regs["regulators"])

    # ------------------------------------------------------------------
    # Trigger 2 — Listeria monocytogenes
    # ------------------------------------------------------------------
    lst = _by_pathogen("listeria")
    if lst:
        regs = _regs_for("listeria", lst)
        # Count STRICT "Listeria monocytogenes" for the note so the figure
        # matches the KPI/analysis Lm count (2026-07-17 fix). The KPI counts
        # the EXACT pathogen string "Listeria monocytogenes"; _by_pathogen
        # ("listeria") also catches a bare generic "Listeria" record and any
        # mixed record such as "Listeria monocytogenes, Salmonella spp", which
        # would inflate the note above the KPI (25 or 24 vs 23). The trigger
        # still fires on ANY listeria presence; only the printed count is exact.
        lst_strict = [r for r in lst
                      if (r.get("Pathogen") or "").strip().lower()
                      == "listeria monocytogenes"]
        _lm_count = len(lst_strict) if lst_strict else len(lst)
        return ("This window contains {ip}, "
                "with {co} among those cited. Ready-to-eat (RTE) manufacturers "
                "\u2014 particularly deli, soft cheese and dairy (including raw-"
                "milk cheese), charcuterie and cured-meat, smoked and cured "
                "seafood, cooked-meat, refrigerated prepared salads, and cut "
                "produce \u2014 should review their environmental monitoring "
                "programme (EMP), Zone 1\u20134 sampling plan, and corrective-"
                "action triggers under {primary}{parallels}. The process-"
                "authority deliverable for this hazard class is a validated "
                "lethality at the kill step (where one exists) or a documented "
                "post-lethality control programme where "
                "the organism cannot be eliminated in-pack. Typical compliance "
                "gaps: incomplete Zone 1 sampling, sanitation SOPs not validated "
                "against worst-case soil load, equipment hollows harbouring "
                "persistent strains, and post-lethality recontamination pathways "
                "not mapped. Incidents of this severity, particularly "
                "confirmed contamination of ready-to-eat product, "
                "may result in targeted regulatory follow-up by {regulators}, "
                "including environmental monitoring (EMP) audits, Zone 1\u20134 "
                "sampling verification, and review of post-lethality control "
                "programmes on the subsequent inspection."
                ).format(ip=_count_phrase(_lm_count, "Listeria monocytogenes incident"),
                         co=_names(lst_strict or lst),
                         primary=regs["primary"], parallels=regs["parallels"],
                         regulators=regs["regulators"])

    # ------------------------------------------------------------------
    # Trigger 3 — Salmonella in low-moisture foods
    # ------------------------------------------------------------------
    sal = _by_pathogen("salmonella")
    sal_lm = [r for r in sal if _is_low_moisture(r)]
    if sal_lm:
        regs = _regs_for("salmonella_lm", sal_lm)
        return ("This window contains {ip} in low-moisture "
                "food categories, with {co} among those cited. Low-moisture "
                "foods \u2014 peanut and nut butters, flour and grain, milk "
                "powder and infant formula, spices and dried herbs, chocolate "
                "and cocoa, seeds and tahini \u2014 require a validated kill "
                "step because Salmonella survives for months to years at low "
                "water activity (a_w) and is not reliably reduced by ambient "
                "handling. The process-authority deliverable is a kill-step "
                "validation study demonstrating \u22655-log Salmonella reduction "
                "under worst-case product composition (fat, a_w, particulate "
                "size) and worst-case equipment conditions, per {primary}"
                "{parallels}. Typical compliance gaps: unvalidated wet-dry zone "
                "segregation, raw-ingredient supplier programme without "
                "certificate-of-analysis (COA) verification, insufficient "
                "sanitary design at post-kill-step transitions, and the absence "
                "of a documented environmental-monitoring programme for the "
                "dry side. Incidents of this severity in this product class "
                "may result in targeted regulatory follow-up by {regulators}, "
                "including kill-step revalidation requests and review of "
                "supplier-verification programmes on the subsequent inspection."
                ).format(ip=_count_phrase(len(sal_lm), "Salmonella incident"),
                         co=_names(sal_lm),
                         primary=regs["primary"], parallels=regs["parallels"],
                         regulators=regs["regulators"])

    # Trigger 4 — Plain Salmonella (no low-moisture)
    if sal:
        regs = _regs_for("salmonella", sal)
        return ("This window contains {ip}, with {co} "
                "among those cited. The process-authority deliverable for "
                "Salmonella-implicated product is the validated kill step for "
                "the commodity (thermal, high-pressure, or equivalent), the "
                "supplier-verification programme on raw inputs, and the "
                "sanitary separation between raw-handling and ready-to-eat "
                "zones under {primary}{parallels}. Typical compliance gaps: "
                "supplier COAs accepted without independent verification "
                "sampling, shared equipment and tooling across raw-RTE "
                "boundaries, a validated kill-step not re-qualified after "
                "formulation or line changes, and time-temperature CCPs "
                "monitored without recording-rigour sufficient for inspector "
                # AUDIT 2026-08-14 — was "Tier-1 classification here
                # triggers HACCP-plan reassessment and formal enforcement
                # action by {regulators}". The strongest version of the
                # same error: AFTS Tier-1 is an internal severity label
                # with no legal standing and triggers nothing. Stating
                # that it compels formal enforcement is a claim about
                # regulators that a manufacturer could act on.
                "reconstruction. Findings of this kind ordinarily warrant "
                "HACCP-plan reassessment, and confirmed contamination of "
                "this product class may attract formal enforcement action "
                "by {regulators}."
                ).format(ip=_count_phrase(len(sal), "Salmonella incident"),
                         co=_names(sal),
                         primary=regs["primary"], parallels=regs["parallels"],
                         regulators=regs["regulators"])

    # ------------------------------------------------------------------
    # Trigger 5 — E. coli STEC / O-serogroup
    # ------------------------------------------------------------------
    stec = _by_pathogen("stec", "e. coli", "escherichia")
    if stec:
        regs = _regs_for("stec", stec)
        return ("This window contains {ip}, with {co} among those cited. The "
                "pathogen has an infectious dose as low as ~10 organisms, and "
                "the process-authority deliverable is a validated kill step "
                "demonstrating \u22655-log STEC reduction for non-intact-beef "
                "processors, or a pre-harvest + post-harvest control "
                "programme for leafy-green, sprout, raw-milk cheese, and "
                "unpasteurised-juice producers, per {primary}{parallels}. "
                "Typical compliance gaps: grinding-plant raw-material cross-"
                "contamination without batch segregation, irrigation-water "
                "microbial monitoring not meeting applicable generic E. coli "
                "numeric limits, post-harvest wash-water chemistry (free "
                "chlorine, pH, ORP) not validated to prevent carry-over, "
                "time-temperature CCPs on hot-hold and cook-chill programmes "
                "insufficiently monitored, and sprout-seed treatment "
                "protocols unvalidated. Confirmed STEC contamination in these "
                "product categories may result in targeted regulatory follow-up "
                "by {regulators}, including verification of cook-chill "
                "temperature controls, post-harvest wash-water chemistry, and "
                "sprout-seed treatment protocols on the subsequent inspection."
                ).format(ip=_count_phrase(len(stec), "E. coli (STEC / shiga-toxin-producing) incident"),
                         co=_names(stec),
                         primary=regs["primary"], parallels=regs["parallels"],
                         regulators=regs["regulators"])

    # ------------------------------------------------------------------
    # Trigger 6 — Regulatory-response PA note.
    # Fires on weeks with either: (a) an outbreak spanning multiple
    # jurisdictions, or (b) ≥3 Tier-1 recalls across ≥3 countries.
    # ------------------------------------------------------------------
    outbreaks = [r for r in recalls if _safe_int(r.get("Outbreak", 0)) == 1]
    outbreak_countries = set(
        _country_display(r.get("Country", "")) for r in outbreaks
        if (r.get("Country") or "").strip())
    tier1 = [r for r in recalls if _safe_int(r.get("Tier", 2)) == 1]
    tier1_countries = set(
        _country_display(r.get("Country", "")) for r in tier1
        if (r.get("Country") or "").strip())

    multi_jurisdiction_outbreak = outbreaks and len(outbreak_countries) >= 2
    high_tier1_pressure         = len(tier1) >= 3 and len(tier1_countries) >= 3

    if multi_jurisdiction_outbreak or high_tier1_pressure:
        if multi_jurisdiction_outbreak:
            anchor = ("a multi-jurisdiction outbreak pattern ({ep} across "
                      "{n_co} countries: {countries})"
                      ).format(ep=_count_phrase(len(outbreaks), "confirmed outbreak event"),
                               n_co=len(outbreak_countries),
                               countries=", ".join(sorted(outbreak_countries)))
            basis_rows = outbreaks
        else:
            anchor = ("elevated Tier-1 regulatory pressure ({n_t1} Tier-1 "
                      "recalls across {n_co} countries: {countries})").format(
                          n_t1=len(tier1), n_co=len(tier1_countries),
                          countries=", ".join(sorted(tier1_countries)[:6]))
            basis_rows = tier1
        regs = _regs_for("regulatory", basis_rows)
        return ("This window shows {anchor}, signalling either co-ordinated "
                "enforcement under mutual-recognition arrangements (RASFF "
                "INFOSAN, WHO IHR notifications, CFIA-FDA trilateral co-"
                "operation) or independent convergence on the same commodity "
                "class. The process-authority deliverable for affected "
                "manufacturers is a 24-hour regulatory-response readiness "
                "package: scheduled-process documentation, HACCP plan with "
                "all CCP records for the implicated lot window, environmental-"
                "monitoring programme output, supplier COAs, and traceability "
                "records one-up / one-back under {primary}{parallels}. Typical "
                "compliance gaps exposed under cross-jurisdiction scrutiny: "
                "inconsistent batch-coding between markets, labelling mis-"
                "alignment, a recall-effectiveness check not completed within "
                "the inspector-expected window, and a press / retailer "
                "notification lag exceeding 24 hours. Multi-jurisdiction "
                "Tier-1 / Class-I events of this profile routinely escalate "
                "to a full-facility regulatory inspection by {regulators} "
                "and can support import-alert listing, detention-without-"
                "physical-examination, or licence-suspension action depending "
                "on the jurisdiction."
                ).format(anchor=anchor, primary=regs["primary"],
                         parallels=regs["parallels"],
                         regulators=regs["regulators"])

    # No trigger fired — return empty string (no P4 appended)
    return ""

def review_with_claude(text):
    """Optional grammar polish via Claude Haiku 4.5. Returns text unchanged on failure
    or when ANTHROPIC_API_KEY is not set. (Replaced OpenAI gpt-4o-mini, Apr 2026.)"""
    if not CLAUDE_API_KEY: return text
    try:
        r = requests.post("https://api.anthropic.com/v1/messages",
            headers={"x-api-key": CLAUDE_API_KEY, "anthropic-version": "2023-06-01",
                     "Content-Type": "application/json"},
            json={"model": "claude-haiku-4-5-20251001", "max_tokens": 1200,
                  "messages": [{"role": "user",
                    "content": "Review this food safety analysis. Fix grammar. Keep structure/facts. Return polished version only.\n\n" + text}]},
            timeout=60)
        if r.status_code == 200:
            return r.json()["content"][0]["text"]
    except Exception as e: log.warning("Claude polish: %s", e)
    return text


def _recall_row(rank, r, top_n=5):
    pathogen = r.get("Pathogen","") or "Unknown"
    tier = _safe_int(r.get("Tier",2)); ob = _safe_int(r.get("Outbreak",0))
    company = r.get("Company","") or "\u2014"
    brand = r.get("Brand","") or "\u2014"
    product = r.get("Product","") or "\u2014"
    country = _country_display(r.get("Country","") or "Unknown")
    source = r.get("Source","") or ""; url = r.get("URL","") or ""
    dt = _fmt_date(r.get("Date",""))
    dot = _dot_color(pathogen)
    # Multi-pathogen outbreak rows display their compound category, so the
    # register and the Hazard Profile name the same thing (audit 2026-08-28).
    ps = co_pathogen_label(r) or pathogen.split("(")[0].strip()
    # AUDIT 2026-07-30 — this was a binary if/else over a THREE-valued field:
    #     'T1' if tier==1 else 'T2'
    # so every Tier-3 recall was labelled "T2" in every weekly report. 45 rows
    # across W14-W29 were mislabelled (mycotoxins, histamine, physical
    # hazards, heavy metals, allergens — all correctly Tier 3 in the
    # workbook). It read as stale HTML, but a full rebuild reproduced it
    # exactly: the reports were never stale, the renderer could not express
    # Tier 3. Anything that is not 1 or 2 now renders its real tier.
    if tier == 1:
        chip = '<span class="chip-tier1">T1</span>'
    elif tier == 2:
        chip = '<span class="chip-tier2">T2</span>'
    else:
        chip = '<span class="chip-tier3">T{}</span>'.format(_safe_int(tier, 3))
    obc = ' <span class="chip-outbreak">OUTBREAK</span>' if ob else ""
    rc = "rank-num" if rank<=top_n else "rank-num rank-num--multi"
    sd = AUTHORITY_DISPLAY.get(source, source)
    if url and url.strip():
        lk = '<div class="juris-link"><a class="src-link" href="{}" target="_blank" rel="noopener">View source &rarr;</a></div>'.format(esc(url))
    else:
        lk = '<div class="juris-link"><span class="src-na" title="No verified specific-recall URL available">unverified</span></div>'
    # Brand sub-line: printed only when it is not a repeat of the company and
    # not a placeholder. This used to print "\u2014" or a verbatim copy of the
    # company underneath itself \u2014 see brand_subline() for the W32 review.
    _b = brand_subline(company, brand)
    bs = ('<div class="brand-sub">{}</div>'.format(esc(_b))) if _b else ""
    return """    <tr>
      <td class="{rc}" data-label="#">{rank}</td>
      <td class="date-cell" data-label="Date">{dt}</td>
      <td data-label="Pathogen">
        <span class="path-dot" style="background:{dot}"></span>
        <span class="path-name">{ps}</span>
        {chip}{obc}
      </td>
      <td class="co-cell" data-label="Company"><strong>{co}</strong>{bs}</td>
      <td class="prod-cell" data-label="Product">{prod}</td>
      <td class="juris-cell" data-label="Jurisdiction">
        <div class="juris-country">{ctry}</div>
        <div class="src-sub">{src}</div>
        {lk}
      </td>
    </tr>""".format(rc=rc, rank=rank, dt=esc(dt), dot=dot, ps=esc(ps),
                    chip=chip, obc=obc, co=esc(company), bs=bs,
                    prod=esc(product), ctry=esc(country), src=esc(sd), lk=lk)


W16_CSS = """
:root {
  --black:#0a0e1a; --orange:#E8601A;
  --ink:#111827; --body:#1f2937; --muted:#6b7280; --dim:#9ca3af;
  --bg:#ffffff; --s1:#f9fafb; --s2:#f3f4f6; --brd:#e5e7eb;
  --red:#dc2626; --amber:#f59e0b; --violet:#9333ea; --green:#059669;
}
* { box-sizing:border-box; }
html, body { margin:0; padding:0; background:var(--bg); }
body {
  font-family:'DM Sans', -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
  color:var(--body); font-size:14px; line-height:1.65;
  max-width:1180px; margin:0 auto; padding:0 40px 60px;
}
a { color:var(--orange); text-decoration:none; }
a:hover { text-decoration:underline; }

.masthead {
  border-top:6px solid var(--black);
  padding:28px 0 22px;
  display:flex; justify-content:space-between; align-items:flex-start;
  border-bottom:1px solid var(--brd);
  margin-bottom:32px;
}
.brand-block .brand {
  font-family:'Syne', sans-serif; font-weight:800; font-size:24px;
  color:var(--black); letter-spacing:-0.01em; text-transform:uppercase;
  line-height:1.1;
}
.brand-block .brand em { color:var(--orange); font-style:normal; font-weight:800; }
.brand-block .tagline {
  font-family:'DM Mono', monospace; font-size:10px; font-weight:600;
  color:var(--muted); text-transform:uppercase; letter-spacing:0.14em;
  margin-top:8px;
}
.mast-right { text-align:right; }
.report-label {
  display:inline-block; background:var(--black); color:#fff;
  font-family:'DM Mono', monospace; font-size:10px; font-weight:700;
  padding:5px 11px; letter-spacing:0.12em; text-transform:uppercase;
  margin-bottom:10px;
}
.report-meta {
  font-family:'DM Mono', monospace; font-size:11px;
  color:var(--muted); line-height:1.8;
}
.report-meta strong { color:var(--ink); font-weight:700; }

.r-title {
  font-family:'Syne', sans-serif; font-weight:800; font-size:38px;
  color:var(--black); letter-spacing:-0.02em; line-height:1.15;
  margin:2px 0 10px;
}
.r-title .accent { color:var(--orange); }
.r-kicker {
  font-family:'Syne', sans-serif; font-weight:800; font-size:13px;
  color:var(--black); letter-spacing:0.08em; text-transform:uppercase;
  margin:8px 0 6px;
}
.r-kicker-dot { color:var(--orange); font-style:normal; margin:0 2px; }
.r-sub {
  color:var(--muted); font-size:14px; margin-bottom:16px;
}
.r-sub strong { color:var(--ink); font-weight:600; }
.r-authority {
  display:flex; align-items:center; gap:10px; flex-wrap:wrap;
  padding:10px 14px; background:var(--s1); border-left:3px solid var(--orange);
  font-family:'DM Mono', monospace; font-size:11px; color:var(--ink);
  margin-bottom:30px;
}
.auth-label {
  font-size:9px; font-weight:700; color:var(--orange);
  text-transform:uppercase; letter-spacing:0.14em;
  border-right:1px solid var(--brd); padding-right:10px;
}

.kpi-strip {
  display:grid; grid-template-columns:repeat(4, 1fr);
  gap:1px; background:var(--brd); border:1px solid var(--brd);
  margin-bottom:32px;
}
.kpi { background:#fff; padding:22px 20px; }
.kpi-label {
  font-family:'DM Mono', monospace; font-size:10px; font-weight:700;
  color:var(--muted); text-transform:uppercase; letter-spacing:0.1em;
  margin-bottom:8px;
}
.kpi-value {
  font-family:'Syne', sans-serif; font-weight:800; font-size:42px;
  color:var(--black); line-height:1; letter-spacing:-0.02em;
}
.kpi-value.red { color:var(--red); }
.kpi-value.violet { color:var(--violet); }
.kpi-value.orange { color:var(--orange); font-size:20px; line-height:1.2; font-style:italic; }
.kpi-value a { color:inherit; text-decoration:none; border-bottom:2px solid var(--orange); padding-bottom:1px; }
.kpi-value a:hover { opacity:0.8; text-decoration:none; }
.kpi-delta {
  font-family:'DM Mono', monospace; font-size:10px; font-weight:700;
  margin-top:10px; letter-spacing:0.04em;
}
.kpi-top { font-size:11px; color:var(--muted); margin-top:10px; font-style:italic; }

.sec-head {
  display:flex; align-items:baseline; gap:14px;
  margin:40px 0 16px;
}
.sec-num {
  font-family:'DM Mono', monospace; font-size:11px; font-weight:700;
  color:var(--orange); letter-spacing:0.12em;
}
.sec-title {
  font-family:'Syne', sans-serif; font-weight:800; font-size:22px;
  color:var(--black); letter-spacing:-0.01em;
}
.sec-rule { flex:1; height:1px; background:var(--brd); }
.sec-caption { color:var(--muted); font-size:13px; margin:-4px 0 14px; }
.sec-caption em { color:var(--ink); font-style:italic; }

.analysis {
  background:var(--s1); border-left:4px solid var(--orange);
  padding:26px 30px; margin-bottom:10px;
}
.analysis p { margin:0 0 14px; font-size:14.5px; line-height:1.75; }
.analysis p:last-child { margin-bottom:0; }
.analysis p.pa-note {
  margin:18px -30px 0 -30px; padding:18px 30px 2px 30px;
  background:#fff; border-top:1px solid var(--brd);
  font-size:13.5px; line-height:1.7;
}
.analysis p.pa-note .pa-label {
  display:inline; font-family:'DM Mono', monospace; font-weight:700;
  letter-spacing:0.08em; text-transform:uppercase;
  color:var(--red); font-size:10px; margin-right:8px;
}

table.data {
  width:100%; border-collapse:collapse; margin:0 0 10px;
  background:#fff; border:1px solid var(--brd);
  font-size:13px;
}
table.data th {
  background:var(--black); color:#fff;
  font-family:'DM Mono', monospace; font-size:10px; font-weight:700;
  text-transform:uppercase; letter-spacing:0.1em;
  padding:12px 12px; text-align:left; border-bottom:2px solid var(--orange);
}
table.data td {
  padding:14px 12px; border-bottom:1px solid var(--brd);
  vertical-align:top;
}
table.data tr:last-child td { border-bottom:none; }
table.data tr:nth-child(even) td { background:#fafbfc; }
table.data td.num {
  font-family:'DM Mono', monospace; font-weight:600; text-align:right;
  white-space:nowrap;
}
table.data td.empty {
  text-align:center; color:var(--muted); padding:28px; font-style:italic;
}

/* Top 5 column sizing - keeps table within A4 and desktop viewport */
table.top5 { table-layout:fixed; width:100%; }
table.top5 th:nth-child(1), table.top5 td:nth-child(1) { width:5%;  }  /* # */
table.top5 th:nth-child(2), table.top5 td:nth-child(2) { width:9%;  }  /* Date */
table.top5 th:nth-child(3), table.top5 td:nth-child(3) { width:19%; }  /* Pathogen */
table.top5 th:nth-child(4), table.top5 td:nth-child(4) { width:18%; }  /* Company */
table.top5 th:nth-child(5), table.top5 td:nth-child(5) { width:30%; }  /* Product */
table.top5 th:nth-child(6), table.top5 td:nth-child(6) { width:19%; }  /* Jurisdiction+Source */
table.top5 td { word-wrap:break-word; overflow-wrap:break-word; }

.rank-num {
  font-family:'Syne', sans-serif; font-weight:800; font-size:22px;
  color:var(--orange); text-align:center;
  white-space:nowrap; font-variant-numeric:tabular-nums;
  letter-spacing:-0.02em;
}
.rank-num.rank-num--multi { font-size:18px; }
.date-cell {
  font-family:'DM Mono', monospace; font-size:11px; color:var(--muted);
  white-space:nowrap;
}
.path-dot {
  display:inline-block; width:9px; height:9px; border-radius:50%;
  margin-right:7px; vertical-align:middle;
}
.path-name { font-weight:600; color:var(--ink); font-style:italic; }
.co-cell strong { color:var(--black); font-weight:700; display:block; }
.brand-sub { font-size:11px; color:var(--muted); margin-top:2px; font-style:italic; }
.prod-cell { color:var(--body); }
.juris-country { font-weight:600; color:var(--ink); }
.src-sub {
  font-family:'DM Mono', monospace; font-size:10px;
  color:var(--muted); margin-top:3px;
}
.juris-link { margin-top:6px; }
.chip-tier1 {
  display:inline-block; background:var(--red); color:#fff;
  font-family:'DM Mono', monospace; font-size:9px; font-weight:700;
  padding:2px 6px; border-radius:2px; margin-left:6px; letter-spacing:0.06em;
}
.chip-tier2 {
  display:inline-block; background:var(--amber); color:#fff;
  font-family:'DM Mono', monospace; font-size:9px; font-weight:700;
  padding:2px 6px; border-radius:2px; margin-left:6px; letter-spacing:0.06em;
}
.chip-tier3 {
  display:inline-block; background:var(--dim); color:#fff;
  font-family:'DM Mono', monospace; font-size:9px; font-weight:700;
  padding:2px 6px; border-radius:2px; margin-left:6px; letter-spacing:0.06em;
}
.chip-outbreak {
  display:inline-block; background:var(--violet); color:#fff;
  font-family:'DM Mono', monospace; font-size:9px; font-weight:700;
  padding:2px 6px; border-radius:2px; margin-left:4px; letter-spacing:0.06em;
}
.src-link {
  font-family:'DM Mono', monospace; font-size:11px; font-weight:700;
  color:var(--orange); letter-spacing:0.02em;
}
.src-na { color:var(--dim); font-family:'DM Mono', monospace; font-size:10px; font-style:italic; }

.dist-grid {
  display:grid; grid-template-columns:1fr 1fr; gap:24px; margin-bottom:10px;
}
.dist-grid h3 {
  font-family:'DM Mono', monospace; font-size:11px; color:var(--muted);
  text-transform:uppercase; letter-spacing:0.1em; margin:0 0 10px;
}
.bar-track {
  width:100%; height:8px; background:var(--s2);
  border-radius:1px; overflow:hidden;
}
.bar-fill { height:100%; }

.cta-box {
  margin:40px 0 30px;
  padding:26px 30px;
  background:var(--black); color:#fff;
  display:flex; justify-content:space-between; align-items:center;
  flex-wrap:wrap; gap:18px;
}
.cta-text { flex:1; min-width:280px; }
.cta-text h3 {
  font-family:'Syne', sans-serif; font-weight:800; font-size:20px;
  margin:0 0 6px; color:#fff; letter-spacing:-0.01em;
}
.cta-text p { margin:0; color:#d1d5db; font-size:13px; }
.cta-btn {
  background:var(--orange); color:#fff; font-family:'DM Mono', monospace;
  font-size:11px; font-weight:700; padding:14px 22px;
  text-transform:uppercase; letter-spacing:0.1em;
  border:none; cursor:pointer; white-space:nowrap;
}
.cta-btn:hover { background:#d35416; text-decoration:none; color:#fff; }

.meth {
  background:var(--s1); border:1px solid var(--brd);
  padding:22px 26px; margin-bottom:24px; font-size:13px;
  color:var(--body);
}
.meth strong { color:var(--black); }
.meth p { margin:0 0 10px; }
.meth p:last-child { margin-bottom:0; }

.footer {
  margin-top:50px; padding-top:26px; border-top:2px solid var(--black);
  display:flex; justify-content:space-between; align-items:flex-start;
  flex-wrap:wrap; gap:20px; font-size:12px;
}
.foot-brand {
  font-family:'Syne', sans-serif; font-weight:800; font-size:15px;
  color:var(--black); text-transform:uppercase; letter-spacing:0.02em;
}
.foot-brand em { color:var(--orange); font-style:normal; }
.foot-meta {
  font-family:'DM Mono', monospace; font-size:10px;
  color:var(--muted); line-height:1.8; margin-top:6px;
}
.foot-legal {
  font-size:11px; color:var(--muted); max-width:440px;
  text-align:right; line-height:1.6;
}

@media print {
  /* Running footer on every printed page: process-authority attribution
     anchors the AFTS differentiator visually throughout the document. */
  @page {
    size: A4;
    margin: 14mm 14mm 18mm 14mm;
    @bottom-left {
      content: "AFTS · Food Safety Validation Intelligence";
      font-family: 'DM Mono', monospace; font-size: 8pt; color: #6b7280;
      letter-spacing: 0.04em;
    }
    @bottom-right {
      content: "Page " counter(page) " / " counter(pages);
      font-family: 'DM Mono', monospace; font-size: 8pt; color: #6b7280;
      letter-spacing: 0.04em;
    }
  }
  body { max-width:none; padding:0; margin:0; font-family:'Times New Roman', Times, serif; font-size:12pt; }
  .cta-box { display:none; }

  /* Lock print-mode layout: even if the browser's print page is narrow,
     these must not collapse into mobile responsive layouts. */
  .masthead { flex-direction:row !important; }
  .mast-right { text-align:right !important; }
  .kpi-strip { grid-template-columns:repeat(4, 1fr) !important; }
  .dist-grid { display:block !important; grid-template-columns:1fr !important; gap:0 !important; }
  .dist-grid > div { width:100% !important; display:block !important; }
  .dist-grid > div:nth-child(2) { margin-top:18px !important; }
  .dist-grid > div { page-break-inside:avoid; break-inside:avoid; }

  /* Page 1 compression: tighten the above-the-fold so the first Intelligence
     Analysis paragraph opens on page 1 rather than orphaning the heading. */
  .masthead { border-top-width:4px; padding:18px 0 12px; margin-bottom:22px; }
  .brand-block .brand { font-size:18px; }
  .brand-block .tagline { font-size:10px; margin-top:5px; letter-spacing:0.12em; }
  .report-label { font-size:9px; padding:4px 10px; margin-bottom:8px; }
  .report-meta { font-size:10px; line-height:1.7; }
  .r-kicker { font-size:12px; margin:6px 0 5px; letter-spacing:0.07em; }
  .r-title { font-size:26px; margin:2px 0 8px; }
  .r-sub { font-size:13px; margin-bottom:12px; line-height:1.55; }
  .r-authority { padding:9px 12px; font-size:11px; margin-bottom:22px; }
  .auth-label { font-size:8px; padding-right:9px; }
  .kpi-strip { margin-bottom:24px; }
  .kpi { padding:16px 14px; }
  .kpi-label { font-size:9px; margin-bottom:6px; }
  .kpi-value { font-size:28px; }
  .kpi-value.orange { font-size:18px; }
  .kpi-delta { font-size:9px; margin-top:7px; }
  .kpi-top { font-size:10px; margin-top:7px; }
  .sec-head { margin:28px 0 12px; page-break-after:avoid; break-after:avoid; }
  .sec-num { font-size:10px; }
  .sec-title { font-size:20px; white-space:nowrap; }
  .analysis { padding:22px 26px; }
  .analysis p { font-size:13px; margin:0 0 12px; line-height:1.7; }
  .analysis p.pa-note { margin:14px -26px 0 -26px; padding:14px 26px 2px 26px; font-size:12px; line-height:1.65; }
  .analysis p.pa-note .pa-label { font-size:9px; }

  table.data th { background:var(--black) !important; color:#fff !important; -webkit-print-color-adjust:exact; print-color-adjust:exact; }
  /* Prevent any table row from splitting across a page break */
  table.data tr { page-break-inside:avoid; break-inside:avoid; }
  .analysis { border-left-width:3px; }
  /* Top-5 print tightening - fit all 6 columns on A4 */
  table.top5 { font-size:9px; page-break-inside:avoid; }
  table.top5 th { padding:6px 5px; font-size:8px; }
  table.top5 td { padding:6px 5px; line-height:1.35; }
  table.top5 tr { page-break-inside:avoid; }
  table.top5 .rank-num { font-size:14px; }
  table.top5 .path-name { font-size:9px; }
  table.top5 .date-cell { font-size:8px; }
  table.top5 .prod-cell { font-size:9px; line-height:1.35; }
  table.top5 .co-cell strong { font-size:9px; }
  table.top5 .juris-country { font-size:9px; }
  table.top5 .brand-sub, table.top5 .src-sub { font-size:8px; margin-top:1px; }
  table.top5 .chip-tier1, table.top5 .chip-tier2, table.top5 .chip-outbreak { font-size:7px; padding:1px 3px; margin-left:3px; }
  table.top5 .src-link { font-size:8px; }
  table.top5 .juris-link { margin-top:3px; }
  /* Tighten the caption above the Top 5 so more room for rows */
  .sec-caption { font-size:10px; margin:-2px 0 8px; }
  /* Force section boundaries on page breaks for clean 4-page distribution:
     P1 = masthead + KPI + § 01 Analysis
     P2 = § 02 Top 5
     P3 = § 03 Distribution
     P4 = § 04 Methodology + Footer */
  section.page-break, div.page-break { page-break-before:always; }
  .sec-head.break-before { page-break-before:always; break-before:page; }

  /* Footer: switch from flex to a clean vertical stack for print.
     WeasyPrint and some browser print engines overlap the two halves
     when flex wraps at narrow widths - block layout avoids it entirely.
     page-break-inside: avoid keeps brand block + disclaimer together on one page. */
  .footer {
    display:block !important;
    margin-top:26px;
    page-break-inside:avoid;
    break-inside:avoid;
  }
  .footer > div { display:block !important; width:auto !important; }
  .footer > div:first-child { margin-bottom:12px; }
  .foot-legal {
    text-align:left !important;
    max-width:none !important;
    padding-top:10px;
    border-top:1px solid var(--brd);
  }
  /* Keep the methodology section with its adjacent section intact */
  .meth { page-break-inside:avoid; break-inside:avoid; }
}

@media screen and (max-width:900px) {
  body { padding:0 20px 40px; }
  .kpi-strip { grid-template-columns:repeat(2,1fr); }
  .dist-grid { grid-template-columns:1fr; }
  .masthead { flex-direction:column; gap:16px; }
  .mast-right { text-align:left; }
  .r-title { font-size:28px; }
}

/* Mobile Top-5: switch from a 6-column table to stacked cards.
   On phones, a horizontal table would either scroll sideways (bad UX) or
   compress columns into unreadable widths. Instead, each row becomes a
   card with labeled fields - all data visible, no horizontal scroll. */
@media screen and (max-width:700px) {
  table.top5, table.top5 thead, table.top5 tbody, table.top5 tr, table.top5 td {
    display:block; width:auto !important;
  }
  /* Kill all fixed column widths - they would make card-mode cells unreadably narrow */
  table.top5 th:nth-child(1), table.top5 td:nth-child(1),
  table.top5 th:nth-child(2), table.top5 td:nth-child(2),
  table.top5 th:nth-child(3), table.top5 td:nth-child(3),
  table.top5 th:nth-child(4), table.top5 td:nth-child(4),
  table.top5 th:nth-child(5), table.top5 td:nth-child(5),
  table.top5 th:nth-child(6), table.top5 td:nth-child(6) {
    width:auto !important;
  }
  table.top5 { border:none; table-layout:auto !important; }
  table.top5 thead { display:none; }
  table.top5 tr {
    border:1px solid var(--brd); border-left:4px solid var(--orange);
    background:#fff; margin-bottom:12px; padding:8px 4px;
    position:relative;
  }
  table.top5 tr:nth-child(even) td { background:transparent; }
  table.top5 td {
    border:none !important; padding:7px 14px 7px 108px !important;
    position:relative; min-height:28px;
    word-wrap:normal; overflow-wrap:normal;
  }
  table.top5 td::before {
    content:attr(data-label);
    position:absolute; left:14px; top:7px; width:88px;
    font-family:'DM Mono', monospace; font-size:9px; font-weight:700;
    color:var(--muted); text-transform:uppercase; letter-spacing:0.08em;
  }
  /* Rank number sits in top-right corner as an orange badge */
  table.top5 .rank-num {
    position:absolute; top:8px; right:14px; padding:0 !important;
    font-size:28px; min-height:0; text-align:right;
  }
  table.top5 .rank-num::before { display:none; }
  table.top5 .date-cell { font-size:11px; }
  table.top5 .path-name { font-size:13px; }
  table.top5 .co-cell strong { font-size:13px; }
  table.top5 .prod-cell { line-height:1.45; font-size:13px; }
  table.top5 .juris-country { font-size:13px; }
  table.top5 .juris-link { margin-top:6px; }
}

@media screen and (max-width:480px) {
  body { padding:0 14px 30px; }
  .kpi-strip { grid-template-columns:1fr 1fr; }
  .kpi { padding:16px 14px; }
  .kpi-value { font-size:28px; }
  .r-title { font-size:24px; }
  .analysis { padding:18px 20px; }
  .analysis p { font-size:13px; }
  .analysis p.pa-note { margin:14px -20px 0 -20px; padding:14px 20px 2px 20px; }
}
"""


HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>AFTS Food Safety Hazard &amp; Pathogen Intelligence Briefing &middot; Week {wnum}, {year}</title>
<link href="https://fonts.googleapis.com/css2?family=Syne:wght@700;800&family=DM+Sans:wght@400;500;600;700&family=DM+Mono:wght@400;500;700&display=swap" rel="stylesheet">
<!-- Cloudflare Web Analytics — see docs/index.html for setup instructions. -->
<script defer src="https://static.cloudflareinsights.com/beacon.min.js"
  data-cf-beacon='{{"token": "__CF_BEACON_TOKEN__"}}'></script>
<style>
__CSS_PLACEHOLDER__
</style>
</head>
<body>

<header class="masthead">
  <div class="brand-block">
    <div class="brand">Advanced Food-Tech Solutions <em>&middot;</em> AFTS</div>
    <div class="tagline">Food Safety Intelligence System &middot; Weekly Briefing</div>
  </div>
  <div class="mast-right">
    <div class="report-label">Subscribers Edition</div>
    <div class="report-meta">
      <strong>ISSUE</strong> &middot; Week {wnum}, {year}<br>
      <strong>PERIOD</strong> &middot; {period}<br>
      <strong>{published_label}</strong> &middot; {published}
    </div>
  </div>
</header>

<div class="r-kicker">AFTS <span class="r-kicker-dot">&middot;</span> Food Safety Validation Intelligence</div>
<h1 class="r-title">Food Safety Hazard &amp; Pathogen Surveillance <span class="accent">&middot;</span> Week {wnum}</h1>
<p class="r-sub">
  AI-powered analysis of <strong>{total}</strong> regulatory food-safety incidents across
  <strong>{n_jurisdictions}</strong> jurisdictions, aggregated from 66 primary sources
  monitored continuously by the AFTS intelligence platform.
</p>
<p class="r-sub" style="margin-top:6px">
  <!-- Scope statement added 2026-08-14. An external review of W33 asked
       whether the register covers human food only or the wider food and
       feed chain, having found a canine-food recall and a live-poultry
       notification in the week. It is human food only; those two rows
       were archived and the pet-food filter that should have caught the
       first was repaired. Stating the boundary here means the question
       is answered in the deliverable rather than in a mailbox. -->
  <strong>Scope.</strong> Human food only. Pet food, animal feed and live
  animals are outside this register, as are allergen-only, labelling and
  quality or spoilage recalls; monitored hazards are pathogens, biotoxins,
  mycotoxins, foreign material, pest and chemical contamination.
  Outbreak figures count <strong>distinct events</strong>, not rows &mdash;
  several recalls arising from one investigation are counted once.
</p>

<div class="kpi-strip">
  <div class="kpi">
    <div class="kpi-label">Total Incidents</div>
    <div class="kpi-value"><a href="#all-recalls">{total}</a></div>
    {delta_html}
  </div>
  <div class="kpi">
    <div class="kpi-label">Tier-1 Critical</div>
    <div class="kpi-value red">{tier1}</div>
    <div class="kpi-delta" style="color:var(--muted)">Immediate public-health risk</div>
  </div>
  <div class="kpi">
    <div class="kpi-label">New Outbreak Events</div>
    <div class="kpi-value violet">{outbreaks}</div>
    <div class="kpi-delta" style="color:var(--muted)">Distinct events in this week's register</div>
  </div>
  <div class="kpi">
    <div class="kpi-label">Leading Pathogen</div>
    <div class="kpi-value orange">{top_pathogen_name}</div>
    <div class="kpi-top">{top_cnt} recall incidents &middot; {top_pct}% of total{co_dom_note}</div>
  </div>
</div>

<div class="sec-head">
  <span class="sec-num">&sect; 01</span>
  <h2 class="sec-title">Intelligence Analysis</h2>
  <span class="sec-rule"></span>
</div>
<div class="analysis">
{analysis_html}
</div>

<div class="sec-head">
  <span class="sec-num">&sect; 02</span>
  <h2 class="sec-title">Top 5 Critical Threats</h2>
  <span class="sec-rule"></span>
</div>
<p class="sec-caption">
  Ranked by outbreak status first &mdash; confirmed human-illness clusters are escalated above every
  non-outbreak recall &mdash; then by pathogen severity (<em>C. botulinum</em> &rarr; <em>Listeria</em> &rarr; STEC &rarr; <em>Salmonella</em>) and tier classification.
  Each row links to the originating regulatory notice.
</p>
<table class="data top5">
  <thead><tr><th>#</th><th>Date</th><th>Pathogen</th><th>Company / Brand</th><th>Product</th><th>Jurisdiction &amp; Source</th></tr></thead>
  <tbody>
    {top5_rows}
  </tbody>
</table>

<div class="sec-head">
  <span class="sec-num">&sect; 03</span>
  <h2 class="sec-title">Distribution Analysis</h2>
  <span class="sec-rule"></span>
</div>
<div class="dist-grid">
  <div>
    <h3>Hazard Profile</h3>
    <table class="data">
      <thead><tr><th>Hazard</th><th class="num">Incidents</th><th class="num">%</th><th>Share</th></tr></thead>
      <tbody>
{pathogen_rows}
      </tbody>
    </table>
  </div>
  <div>
    <h3>Geographic &middot; Regulatory</h3>
    <table class="data">
      <thead><tr><th>Origin country</th><th>Notifying authority</th><th class="num">Incidents</th><th class="num">%</th></tr></thead>
      <tbody>
{country_rows}
      </tbody>
    </table>
  </div>
</div>

<div class="cta-box">
  <div class="cta-text">
    <h3>Live Dashboard &middot; Full Dataset Access</h3>
    <p>Filter by pathogen, country, tier, and source. Access the cumulative XLSX dataset. Set custom alerts.</p>
  </div>
  <a class="cta-btn" href="https://www.advfood.tech/fsis-recalls" target="_blank" rel="noopener">Access Portal &rarr;</a>
</div>

<div id="all-recalls" class="sec-head">
  <span class="sec-num">&sect; 04</span>
  <h2 class="sec-title">{n_notices} Regulatory Notices &middot; {total} Distinct Incidents &middot; {period}</h2>
  <span class="sec-rule"></span>
</div>
<p class="sec-caption">
  <strong>This table is the regulatory-notice register: every notice is listed individually.</strong>
  The KPIs and the distributions in &sect; 03 count DISTINCT INCIDENTS, which is why
  {n_notices} rows appear below against a headline figure of {total}. Where one event produced
  several notices &mdash; a single cold-chain failure recalling many suppliers&rsquo; products from one
  store, for example &mdash; each notice stays individually searchable here and is counted once there.
  {incident_note}
  Sorted by pathogen severity, outbreak status, and tier classification. Each row links to the
  originating regulatory notice. A RASFF notification is a border or market-control notification
  between competent authorities and does not necessarily correspond to a consumer recall.
</p>
<table class="data top5">
  <thead><tr><th>#</th><th>Date</th><th>Pathogen</th><th>Company / Brand</th><th>Product</th><th>Jurisdiction &amp; Source</th></tr></thead>
  <tbody>
    {all_rows}
  </tbody>
</table>

<div class="sec-head">
  <span class="sec-num">&sect; 05</span>
  <h2 class="sec-title">Methodology &amp; Sources</h2>
  <span class="sec-rule"></span>
</div>
<div class="meth">
  <p><strong>Process authority.</strong> Analytical frameworks, severity rubrics, pathogen classification, and the engineering interpretation of each recall are developed by the AFTS process-authority practice, drawing on in-house expertise in food process engineering, thermal processing, and regulatory compliance. Every view is grounded in validated process engineering: thermal processing (21 CFR 113/114), pasteurisation (PMO), aseptic and UHT, hold-tube and F-value lethality, and HACCP. This is what the AFTS platform brings that pure data feeds do not &mdash; interpretation under engineering authority.</p>
  <p><strong>Data &amp; AI pipeline.</strong> The system aggregates regulatory recall notices from 66 primary sources across 60+ countries (FDA, USDA FSIS, RASFF, FSA, FSANZ, CFIA, RappelConso, BVL, AESAN, EFET, and national authorities) and processes each record through Gemini (extraction), OpenAI GPT (normalisation), and Claude (Tier-1 validation). Records are de-duplicated and harmonised into the cumulative dataset.</p>
  <p><strong>This briefing.</strong> Statistical analysis filters the cumulative dataset to the reporting week ({period}). AI-generated narrative is produced against AFTS process-authority prompts and edited for publication. Figures and pathogen names are preserved verbatim from source data.</p>
</div>

<footer class="footer">
  <div>
    <div class="foot-brand">Advanced Food-Tech Solutions <em>&middot;</em> AFTS</div>
    <div class="foot-meta">Food Safety Validation Intelligence<br>advfood.tech &middot; info@advfood.tech &middot; Athens, Greece<br>&copy; {year} Advanced Food-Tech Solutions</div>
  </div>
  <div class="foot-legal">This briefing is provided for informational purposes only and does not constitute regulatory, legal, or medical advice. Subscribers should verify recall status with the originating regulatory authority before taking action. Next issue: Friday, {next_issue}.</div>
</footer>

</body>
</html>"""


def build_html(week_end, recalls, prev_week, original_published=None):
    """Build a weekly report HTML.

    original_published:  None  → fresh publish; header reads
                                 "PUBLISHED · {week_end + 1}".
                         str   → rebuild; header reads
                                 "UPDATED · {today}". The string itself
                                 (returned by _extract_published_from_html)
                                 is just the "this is a rebuild" signal —
                                 its content is no longer used in render.

    Audit 2026-05-09 — added original_published with label-flip semantics.
    Pre-this change, every rebuild silently overwrote the PUBLISHED date
    with the new build's week_end+1, which on Wednesday rebuilds is
    identical to the original Saturday publish — so the user couldn't see
    that the report had been revised. Now rebuilds carry an explicit
    UPDATED label with today's date.
    """
    stats = compute_stats(recalls, prev_week)
    # Display window (audit 2026-05-12): under the new-rule (any row in
    # `recalls` carries a report_week stamp), the period runs Fri→Thu
    # (week_end−7 .. week_end−1) — the ship Friday is NOT part of the
    # data window. Legacy reports (no stamped rows) keep Sat→Fri
    # (week_end−6 .. week_end). See _display_window() for full reasoning.
    ws, display_end_date = _display_window(week_end, recalls)
    wnum = week_end.isocalendar()[1]; year = week_end.year
    total = stats["total"]
    sr = sort_by_severity(recalls)

    raw = generate_analysis_claude(stats, recalls)
    final = review_with_claude(raw)

    paras = [p.strip() for p in final.strip().split("\n\n") if p.strip()]

    # Sanitize light markdown that Claude sometimes returns despite system
    # instructions for clean HTML output (audit 2026-06-12):
    #   - strip leading "# ..." H1 lines (e.g. "# Food Safety Analysis – Weekly Report"
    #     leaked into W24 page 1)
    #   - convert *italic* runs (e.g. "*Listeria monocytogenes*", "*Listeria*")
    #     to <em>...</em>
    # The sanitizer runs BEFORE the PA-note classifier so heading lines don't
    # accidentally land in the regulatory paragraph; the <em> conversion runs
    # AFTER esc() so the angle brackets pass through HTML-escape untouched.
    def _strip_md_heading(p):
        # Remove a leading "# ..." H1 line; keep the rest of the paragraph.
        return re.sub(r"^\s*#{1,6}\s+[^\n]*\n?", "", p, count=1)

    def _md_em(s):
        # *text* → <em>text</em>. Only single asterisks; ignore ** (bold).
        # Negative lookarounds reject literal asterisks inside words and
        # leading/trailing whitespace inside the run.
        return re.sub(r"(?<!\*)\*([^\s*][^*\n]*?[^\s*]|[^\s*])\*(?!\*)",
                      r"<em>\1</em>", s)

    paras = [_strip_md_heading(p) for p in paras]
    paras = [p.strip() for p in paras if p.strip()]   # re-filter empties

    def _pa(text):
        return ('<p class="pa-note"><span class="pa-label">Process Authority '
                'Note:</span> ' + _md_em(esc(text.strip())) + '</p>')

    # THE NOTE IDENTIFIES ITSELF (review 2026-08-07). A marked paragraph wins
    # outright; the phrase heuristic below only runs when there is none.
    marked = [p for p in paras if p.startswith(PA_NOTE_MARKER)]
    pa_html = _pa(marked[-1][len(PA_NOTE_MARKER):]) if marked else ""
    reg = []
    for p in paras:
        if p.startswith(PA_NOTE_MARKER):
            continue
        pl = p.lower()
        # PA Note detection — any paragraph using the process-authority idiom
        # and citing a regulatory-enforcement framework. Jurisdiction-neutral
        # since the primary regulator block varies by week.
        # The note IDENTIFIES ITSELF. _process_authority_note() is the only
        # thing that produces this paragraph, and it stamps PA_NOTE_MARKER on
        # the front. The phrase heuristic below is kept ONLY as a fallback for
        # a Claude response that writes a note despite the prompt forbidding
        # it \u2014 and it can no longer steal the label, because a marked
        # paragraph wins outright.
        is_pa = (not pa_html) and (
            ("process-authority deliverable" in pl) or
            ("qualified process authority" in pl) or
            ("regulatory-response readiness" in pl) or
            ("environmental monitoring programme (emp)" in pl) or
            ("kill-step validation study" in pl) or
            ("\u22655-log" in p) or
            ("multi-jurisdiction" in pl and ("escalate" in pl or "import-alert" in pl))
        )
        if is_pa:
            pa_html = _pa(p)
        else:
            reg.append(p)
    analysis = "\n".join("  <p>{}</p>".format(_md_em(esc(p))) for p in reg)
    if pa_html: analysis += "\n" + pa_html

    # Top 5 uses the PHASE-BASED ranker (outbreaks escalated above every
    # non-outbreak recall), identical to the monthly/marketing one-pager.
    # sort_by_severity() ranks pathogen severity FIRST and treats Outbreak as
    # a tiebreak only within the same pathogen, so a 23-row Listeria week
    # buried every confirmed outbreak below the fold (reviewer note 2026-07-10).
    sr_top = _diversify_by_country(rank_top_recalls(recalls, n=len(recalls)),
                                   cap=2, window=5)
    t5rows = "\n".join(_recall_row(i+1, r, 5) for i,r in enumerate(sr_top[:5]))
    allrows = "\n".join(_recall_row(i+1, r, 5) for i,r in enumerate(sr))
    if not t5rows: t5rows = '<tr><td class="empty" colspan="6">No recalls this week</td></tr>'
    if not allrows: allrows = t5rows

    d = stats["delta"]; dp = stats["delta_pct"]
    if d < 0:
        dh = '<div class="kpi-delta" style="color:#059669">&#9660; {} ({}%) vs prior week</div>'.format(d, dp)
    elif d > 0:
        dh = '<div class="kpi-delta" style="color:#dc2626">&#9650; +{} (+{}%) vs prior week</div>'.format(d, dp)
    else:
        dh = '<div class="kpi-delta" style="color:var(--muted)">No change vs prior week</div>'

    tp, tc = stats["top_pathogen"]
    tpct = round(tc/max(total,1)*100) if total else 0
    _co = stats.get("co_dominant")
    if _co and _co.get("also"):
        co_dom_note = (f' &middot; {_co["leader"]} group also {_co["count"]} '
                       f'when {"/".join(_co["also"])} serovar-specific counts are included')
    else:
        co_dom_note = ""

    prows = ""
    for path, cnt in _disambiguate_genus_labels(stats["pathogen_counts"]):
        pct = round(cnt/max(total,1)*100); dot = _dot_color(path)
        prows += '        <tr>\n          <td><span class="path-dot" style="background:{}"></span><em class="path-name">{}</em></td>\n          <td class="num">{}</td>\n          <td class="num">{}%</td>\n          <td><div class="bar-track"><div class="bar-fill" style="width:{}%;background:{}"></div></div></td>\n        </tr>\n'.format(dot, esc(path), cnt, pct, pct, dot)
    if not prows: prows = '        <tr><td class="empty" colspan="4">No pathogen data</td></tr>\n'

    crows = ""
    country_sources = stats.get("country_sources") or {}
    for country, cnt in stats["country_counts"]:
        pct = round(cnt/max(total,1)*100)
        # Authority-column logic (audit 2026-06-12 — auditor flagged that
        # third-country origins like Türkiye, India, Uganda, Egypt were
        # mislabeled "National Authority" when the actual notifier was
        # RASFF (EU) reporting on imports from those countries):
        #
        #   1. Country IS in GEO_AUTHORITY → use that label (the named
        #      national/regional authority is the one that issued the
        #      recall or notice for this country's product).
        #   2. Country is NOT in GEO_AUTHORITY → look up which Source(s)
        #      reported this country in the window:
        #        - If ANY source mentions "rasff" → "RASFF (EU) — origin"
        #          (the EU's Rapid Alert System for Food and Feed notified
        #          about an import from this third country)
        #        - Else if a single recognisable source dominates → use it
        #          (e.g. CDC for a multi-state outbreak attributed to a
        #          single origin country)
        #        - Else "—" (em-dash; honest "no single notifying authority")
        srcs = country_sources.get(country, set())
        if country in GEO_AUTHORITY:
            auth = GEO_AUTHORITY[country]
        elif any("rasff" in s.lower() for s in srcs):
            auth = "RASFF (EU) \u2014 origin"
        elif len(srcs) == 1:
            auth = next(iter(srcs))
        else:
            auth = "\u2014"
        crows += '        <tr>\n          <td>{}</td>\n          <td>{}</td>\n          <td class="num">{}</td>\n          <td class="num">{}%</td>\n        </tr>\n'.format(esc(country), esc(auth), cnt, pct)
    if not crows: crows = '        <tr><td class="empty" colspan="4">No geographic data</td></tr>\n'

    wsd = _fmt_date_short(ws); wed = _fmt_date(display_end_date)
    period = "{} &ndash; {}".format(wsd, wed)
    # Header label flips between two states (audit 2026-05-09, revised
    # 2026-05-14 evening):
    #   First publish (original_published is None) →
    #     "PUBLISHED · {today}" — the actual build date.
    #   Rebuild (original_published is set) →
    #     "UPDATED · {today}" — the actual rebuild date.
    # The label change is the visible signal that the user is looking at
    # a revised version of a previously-published weekly briefing.
    #
    # PRIOR BEHAVIOR (fixed 2026-05-14 evening):
    # The PUBLISHED branch used `(week_end + timedelta(days=1))` — the
    # "Saturday-after-week-close" the author called "formulaic" in the
    # original comment. That was correct ONLY when the build ran on that
    # specific Saturday. When the build was triggered any other day
    # (manual rebuild, cron drift, ad-hoc regeneration), the header
    # displayed a FUTURE date, which was confusing and factually wrong:
    # a viewer reading the page on Thursday 14 May saw "PUBLISHED 16
    # May 2026" for a file that obviously could not have been published
    # in the future. Both branches now use today's date — the visible
    # difference is the label (PUBLISHED vs UPDATED), not a synthetic
    # future date.
    today_str = datetime.now(timezone.utc).strftime("%-d %b %Y")
    # Operator rule (2026-06-29): "PUBLISHED" marks ONLY the genuine first
    # issuance of a weekly briefing. Any later regeneration, correction, or
    # back-build is a REVIEW, not a publish — the masthead must read
    # "REVIEWED · {date}", never re-stamp "PUBLISHED" on a report that was
    # not actually first issued today. A rebuild is signalled by
    # original_published being set.
    # Operator rule (restated 2026-07-13): a report already published before
    # and now rebuilt MUST read "UPDATED · {today}" — never "PUBLISHED",
    # never "REVIEWED". First issuance reads "PUBLISHED".
    if original_published:
        published_label = "UPDATED"
    else:
        published_label = "PUBLISHED"
    pub = today_str
    nf = _fmt_date(week_end + timedelta(days=7))

    html = HTML_TEMPLATE.format(
        wnum=wnum, year=year, period=period,
        published=pub, published_label=published_label, total=total,
        n_jurisdictions=len(stats["country_counts"]), delta_html=dh,
        n_notices=len(recalls), incident_note=_incident_note(recalls),
        tier1=stats["tier1"], outbreaks=stats["outbreaks"],
        top_pathogen_name=esc(tp), top_cnt=tc, top_pct=tpct, co_dom_note=co_dom_note,
        analysis_html=analysis, top5_rows=t5rows, pathogen_rows=prows,
        country_rows=crows, all_rows=allrows, next_issue=nf,
    )
    html = html.replace("__CSS_PLACEHOLDER__", W16_CSS)
    # Cloudflare Web Analytics — substitute __CF_BEACON_TOKEN__ from env var
    # if set, otherwise strip the script tag so the deployed page does NOT
    # ship with a literal placeholder string (audit 2026-06-05: published W23
    # report still had unresolved __CF_BEACON_TOKEN__ in the HTML source).
    import os as _os, re as _re
    _cf_token = (_os.environ.get("CF_BEACON_TOKEN") or "").strip()
    if _cf_token:
        html = html.replace("__CF_BEACON_TOKEN__", _cf_token)
    else:
        # Drop the entire CF Web Analytics block — comment line + script tag
        # (script tag's data-cf-beacon attribute spans two lines, so use
        # DOTALL and lazy quantifier to match the whole element).
        html = _re.sub(
            r"\n?<!-- Cloudflare Web Analytics.*?</script>",
            "",
            html,
            flags=_re.DOTALL,
        )
    return html, stats

def update_dashboard_data(week_end, stats, all_recalls=None):
    """Write docs/data/weekly-index.json — the dashboard's loadReports()
    fetches this on tab open. Always writes the FULL accumulated history
    (every Friday-ending week that has at least one recall in the dataset),
    sorted newest-first. Dashboard handles slicing to 4 rich cards + per-
    year archive client-side, so we never have to touch index.html.

    Audit 2026-04-29 — switched from regex-mutating index.html (fragile,
    silently produced stale "outbreaks=3" cards when the dataset was
    corrected to outbreaks=1) to writing a JSON the dashboard re-fetches
    on every load. Single source of truth: docs/data/recalls.xlsx Recalls
    sheet → docs/data/weekly-index.json → dashboard cards.
    """
    out = ROOT / "docs" / "data" / "weekly-index.json"

    def _make_entry(we, st, wr_for_display=None):
        wn = we.isocalendar()[1]; yr = we.year
        # Display window via _display_window so JSON matches HTML period.
        # New-rule (stamped) reports: ws=we−7 (Fri), we_display=we−1 (Thu).
        # Legacy reports: ws=we−6 (Sat), we_display=we (Fri).
        ws, we_display = _display_window(we, wr_for_display or [])
        tp_name = st["top_pathogen"][0] if st["top_pathogen"] else "Mixed"
        return {"filename":"{}-W{:02d}.html".format(yr,wn),"week_num":wn,"year":yr,
                "week_start":ws.strftime("%Y-%m-%d"),"week_end":we_display.strftime("%Y-%m-%d"),
                "generated":datetime.now(timezone.utc).isoformat(),
                "total":st["total"],"tier1":st["tier1"],"outbreaks":st["outbreaks"],
                "top_pathogen":tp_name,
                "summary":"Week {} saw {} pathogen recalls with {} as primary concern.".format(
                    wn, st["total"], tp_name)}

    entries = []

    if all_recalls:
        # Find every distinct ISO week represented in the recall set, snap
        # each to its Friday end, build an entry. Only keep weeks whose
        # Friday is on or before today (no future weeks).
        # ── THE WEEK BEING BUILT IS NEVER A "FUTURE" WEEK (2026-08-20) ──
        # This guard exists to keep weeks with no data yet out of the index.
        # It read `if fri <= today`, and that quietly broke the operator's
        # actual cadence: the AFTS week closes THURSDAY and the report is
        # built with --week-end set to the FOLLOWING Friday. Building W34 on
        # Thursday 20 Aug with --week-end 2026-08-21 therefore produced
        #     docs/2026-W34.html                     written
        #     docs/data/weekly-summary-latest.json   points at W34
        #     docs/data/weekly-index.json            newest entry W33
        # — the report existed, the pointer named it, and the dashboard's
        # own list did not contain it. Every week, for one day, and it
        # self-heals only if something rebuilds the index on the Friday.
        #
        # The week this run was invoked for is by definition not a future
        # week: it has data, it has just been rendered. Everything after it
        # is still excluded.
        today = max(date.today(), week_end)
        week_ends_seen = set()
        for r in all_recalls:
            d = r.get("Date", "")
            if not d: continue
            try: rd = datetime.strptime(str(d)[:10], "%Y-%m-%d").date()
            except ValueError: continue
            # Snap rd's ISO week to that week's Friday (ISO weekday 5 = Fri)
            iso_year, iso_week, iso_dow = rd.isocalendar()
            # Friday of the same ISO week:
            mon = rd - timedelta(days=iso_dow - 1)
            fri = mon + timedelta(days=4)
            if fri <= today:
                week_ends_seen.add(fri)

        for we in sorted(week_ends_seen, reverse=True):
            prev_we = we - timedelta(days=7)
            wr = filter_week(all_recalls, we)
            if not wr: continue
            pr = filter_week(all_recalls, prev_we)
            st = compute_stats(wr, pr)
            entries.append(_make_entry(we, st, wr_for_display=wr))
    else:
        # Fallback: just the current week
        entries = [_make_entry(week_end, stats)]

    try:
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text(json.dumps(entries, indent=2, ensure_ascii=False),
                        encoding="utf-8")
        wk_list = ", ".join("W{}".format(e["week_num"]) for e in entries[:6])
        suffix = "" if len(entries) <= 6 else f" (+{len(entries)-6} more)"
        log.info("Wrote %s — %d entries: %s%s",
                 out, len(entries), wk_list, suffix)
    except Exception as e:
        log.error("weekly-index.json write failed: %s", e)

def write_weekly_summary_json(week_end, recalls, stats, data_dir):
    wnum = week_end.isocalendar()[1]; year = week_end.year
    # Audit 2026-05-12: week_end in this JSON is the DISPLAY end (Thu for
    # new-rule, Fri for legacy) so the dashboard formats it correctly.
    # The HTML filename and isocalendar-derived wnum still anchor to the
    # ship Friday — only the displayed window shifts.
    ws, we_display = _display_window(week_end, recalls)
    tp = stats.get("top_pathogen")
    leading = {"name":tp[0],"cases":tp[1],"pct":round(tp[1]/max(stats["total"],1)*100)} if tp and len(tp)>=2 else {"name":"Mixed","cases":0,"pct":0}
    # Same phase-based ranking as the report's Top 5 so the email, the
    # summary JSON and the HTML never disagree about the headline threats.
    threats = []
    sr = _diversify_by_country(rank_top_recalls(recalls, n=len(recalls)),
                               cap=2, window=5)
    for i,r in enumerate(sr[:5],1):
        threats.append({"rank":i,"date":str(r.get("Date",""))[:10],
            "pathogen":str(r.get("Pathogen","")),"pathogen_raw":str(r.get("Pathogen","")),
            "tier":_safe_int(r.get("Tier",2)),"outbreak":bool(_safe_int(r.get("Outbreak",0))),
            "company":str(r.get("Company","")),"brand":str(r.get("Brand","\u2014")),
            "product":str(r.get("Product","")),"country":str(r.get("Country","")),
            "source":str(r.get("Source","")),"url":str(r.get("URL",""))})
    summary = {"filename":"{}-W{:02d}.html".format(year,wnum),
        "report_url":"https://fsis.advfood.tech/{}-W{:02d}.html".format(year,wnum),
        "dashboard_url":"https://www.advfood.tech/fsis-recalls",
        "week_num":wnum,"year":year,"week_start":ws.isoformat(),"week_end":we_display.isoformat(),
        "week_start_display":ws.strftime("%-d %b"),"week_end_display":we_display.strftime("%-d %b %Y"),
        "generated_utc":datetime.now(timezone.utc).isoformat(),
        "stats":{"total":stats["total"],"tier1":stats["tier1"],"outbreaks":stats["outbreaks"],
                 "delta":stats.get("delta",0),"delta_pct":stats.get("delta_pct",0)},
        "leading_pathogen":leading,"ai_lead_paragraph":"","top_threats":threats,
        "country_count":len(set(str(r.get("Country","")) for r in recalls if r.get("Country")))}
    out = data_dir / "weekly-summary-latest.json"

    # ---------------------------------------------------------------------
    # "LATEST" MEANS LATEST (incident 2026-08-07).
    #
    # WHAT HAPPENED
    # Subscribers received a Week 27 briefing on Friday 7 August. Week 27
    # ended 2 July — five weeks stale — and had already been sent once.
    #
    # The mailer was NOT at fault. It fetches this file, checks that
    # generated_utc is fresh, and sends whatever week the file names. This
    # file said Week 27 and had been written four hours earlier, so the
    # staleness guard passed: the JSON was fresh, its CONTENT was ancient.
    # Nothing in the chain compared the week number to the calendar.
    #
    # HOW THIS FILE CAME TO SAY W27
    #   1. A Clostridium botulinum recall dated 2026-07-02 was promoted on
    #      2026-08-06. compute_report_week correctly stamped it W27.
    #   2. daily-review-agent.yml noticed W27's count had drifted and
    #      rebuilt it. Its loop sorts stale weeks ascending, with the
    #      comment "build W28/current last so latest pointer stays correct"
    #      — i.e. it ASSUMES the current week is always also stale, so the
    #      current week is always built last and reclaims the pointer.
    #   3. That day only W27 had drifted. The only week built was therefore
    #      also the last week built. It took the pointer.
    #
    # An assumption held in a comment in a YAML file is not a guard. The
    # guard belongs here, at the single place the pointer is written, where
    # it covers every caller — the daily review agent, the gap-filler, a
    # manual backfill, and whatever gets written next year.
    #
    # RULE: a build of a week that is NOT the newest week on record writes
    # its own HTML and its weekly-index.json row and leaves this pointer
    # alone. Retro-rebuilds of closed weeks are exactly what
    # weekly-updates-pending.json and the Wednesday notification exist for.
    # ---------------------------------------------------------------------
    if out.exists():
        try:
            prior = json.loads(out.read_text(encoding="utf-8"))
            prior_key = (int(prior.get("year") or 0),
                         int(prior.get("week_num") or 0))
            if prior_key > (year, wnum):
                log.warning(
                    "REFUSING to move the weekly latest-pointer backwards: "
                    "%s already points at %04d-W%02d and this build is "
                    "%04d-W%02d. Retro-rebuild of a closed week — the HTML "
                    "and weekly-index.json are updated, the subscriber "
                    "pointer is not. (Incident 2026-08-07: a W27 rebuild "
                    "took this pointer and a five-week-old briefing went to "
                    "every subscriber.)",
                    out.name, prior_key[0], prior_key[1], year, wnum)
                return
        except Exception as exc:                       # pragma: no cover
            # A corrupt pointer must not block a legitimate build — but say so.
            log.warning("weekly latest-pointer unreadable (%s: %s) — writing "
                        "%04d-W%02d", type(exc).__name__, exc, year, wnum)

    out.write_text(json.dumps(summary,indent=2,ensure_ascii=False),encoding="utf-8")
    log.info("Wrote %s",out)

def _extract_total_from_html(path):
    """Read an existing report HTML and extract the total-recalls KPI.
    Returns int or None if the file doesn't exist or can't be parsed."""
    try:
        html = Path(path).read_text(encoding="utf-8")
        m = re.search(r'<a href="#all-recalls">(\d+)</a>', html)
        if m: return int(m.group(1))
        # Fallback: the §04 heading. Accept BOTH wordings — the heading was
        # "All N Recalls" until the 2026-08-07 review renamed it to "All N
        # Regulatory Incidents", and every report published before that date
        # still carries the old text. A regex that only matched the new
        # wording would silently return None for the whole back catalogue,
        # which is how the week-on-week delta gets computed against nothing.
        m2 = re.search(r'All (\d+) (?:Regulatory Incidents|Recalls)', html)
        if m2: return int(m2.group(1))
    except Exception:
        pass
    return None


def _extract_published_from_html(path):
    """Read an existing report HTML and detect whether it has a header
    date marker (either "PUBLISHED" or "UPDATED").

    Returns the verbatim date string (e.g. "9 May 2026") or None.

    The CONTENT of the returned string is no longer used by build_html
    after the 2026-05-09 label-flip change — but the FACT that it
    returned non-None is the signal "this is a rebuild, flip the label
    to UPDATED and use today's date." Successive rebuilds keep flipping
    the date forward as expected.

    Two patterns supported (an already-rebuilt report carries UPDATED):
      • Fresh publish:  "PUBLISHED</strong> &middot; 9 May 2026"
      • After rebuild:  "UPDATED</strong> &middot; 13 May 2026"
    """
    try:
        html = Path(path).read_text(encoding="utf-8")
        m = re.search(
            r'<strong>(?:PUBLISHED|UPDATED|REVIEWED)</strong>\s*&middot;\s*([^<]+?)<',
            html,
        )
        if m:
            return m.group(1).strip()
    except Exception:
        pass
    return None


def _extract_label_from_html(path):
    """Read an existing report HTML and return its header label —
    either "PUBLISHED" or "UPDATED" — or None.

    Companion to _extract_published_from_html. Used by the Wednesday
    weekly-updates check (audit 2026-05-09) to detect the case where
    a report's count matches the current dataset BUT the label is
    still "PUBLISHED" — typically because a previous rebuild fired
    before this label-flip code was deployed and overwrote the file
    with a stale "PUBLISHED" header. The Wednesday flow uses this
    signal to force a one-shot rebuild that flips the label to
    "UPDATED" without waiting for a count change.
    """
    try:
        html = Path(path).read_text(encoding="utf-8")
        m = re.search(
            r'<strong>(PUBLISHED|UPDATED|REVIEWED)</strong>\s*&middot;\s*[^<]+<',
            html,
        )
        if m:
            return m.group(1)
    except Exception:
        pass
    return None


def refresh_stale_weeks(all_recalls, current_week_end, n_previous=1):
    """Check up to n_previous weeks before current_week_end.
    If the recall count in the dataset differs from what's baked into
    the existing HTML report, rebuild that week's report.
    Returns list of rebuilt week-end dates."""
    rebuilt = []
    for offset in range(1, n_previous + 1):
        prev_end = current_week_end - timedelta(days=7 * offset)
        prev_prev_end = prev_end - timedelta(days=7)
        wnum = prev_end.isocalendar()[1]
        year = prev_end.year
        report_path = ROOT / "docs" / "{}-W{:02d}.html".format(year, wnum)

        dataset_recalls = filter_week(all_recalls, prev_end)
        dataset_total = len(dataset_recalls)

        existing_total = _extract_total_from_html(report_path)

        if existing_total is None:
            log.info("W%02d: no existing report at %s — building fresh", wnum, report_path)
        elif existing_total == dataset_total:
            log.info("W%02d: report matches dataset (%d recalls) — no refresh needed", wnum, dataset_total)
            continue
        else:
            log.info("W%02d: STALE — report has %d recalls, dataset has %d — rebuilding",
                      wnum, existing_total, dataset_total)

        prev_week_recalls = filter_week(all_recalls, prev_prev_end)
        # Preserve the original publish date so the rebuilt HTML shows
        # "PUBLISHED · <orig> (updated <today>)" instead of overwriting
        # with today's date. Falls back to None (fresh build) if the
        # existing HTML doesn't carry a parseable PUBLISHED marker.
        orig_pub = _extract_published_from_html(report_path)
        html, stats = build_html(prev_end, dataset_recalls, prev_week_recalls,
                                 original_published=orig_pub)
        report_path.parent.mkdir(parents=True, exist_ok=True)
        report_path.write_text(html, encoding="utf-8")
        log.info("W%02d refreshed -> %s (%d bytes, %d recalls, original_pub=%r)",
                  wnum, report_path, len(html), dataset_total, orig_pub)
        rebuilt.append(prev_end)

    return rebuilt


def main():
    ap = argparse.ArgumentParser(description="Build AFTS weekly report")
    ap.add_argument("--week-end", required=True, help="Friday YYYY-MM-DD")
    ap.add_argument("--xlsx", default=str(ROOT/"docs"/"data"/"recalls.xlsx"))
    ap.add_argument("--output", default=None)
    ap.add_argument("--refresh-previous", type=int, default=1, metavar="N",
                    help="Check N previous weeks for stale data and rebuild if needed (default: 1)")
    ap.add_argument("--no-refresh", action="store_true",
                    help="Skip stale-week refresh (build current week only)")
    args = ap.parse_args()
    week_end = datetime.strptime(args.week_end,"%Y-%m-%d").date()
    log.info("Building report for %s", week_end)
    all_r = load_recalls(Path(args.xlsx))
    log.info("Loaded %d recalls", len(all_r))

    # --- Refresh stale previous weeks ---
    if not args.no_refresh and args.refresh_previous > 0:
        log.info("Checking %d previous week(s) for stale data...", args.refresh_previous)
        rebuilt = refresh_stale_weeks(all_r, week_end, args.refresh_previous)
        if rebuilt:
            log.info("Refreshed %d stale report(s): %s",
                      len(rebuilt), ", ".join(d.strftime("W%V") for d in rebuilt))

    # --- Build current week ---
    wr = filter_week(all_r, week_end)
    pr = filter_week(all_r, week_end - timedelta(days=7))
    log.info("This week: %d  Prev: %d", len(wr), len(pr))
    wnum = week_end.isocalendar()[1]
    out = Path(args.output) if args.output else ROOT/"docs"/"{}-W{:02d}.html".format(week_end.year, wnum)
    # Existing report on disk => REBUILD: preserve original publish date and
    # flip masthead to "UPDATED · {today}". Without this, manual/CLI rebuilds
    # silently re-stamped "PUBLISHED" (operator bug report 2026-07-13).
    orig_pub = _extract_published_from_html(out)
    html, stats = build_html(week_end, wr, pr, original_published=orig_pub)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(html, encoding="utf-8")
    log.info("Report -> %s (%d bytes)", out, len(html))
    update_dashboard_data(week_end, stats, all_r)
    write_weekly_summary_json(week_end, wr, stats, Path(args.xlsx).parent)
    return 0

if __name__ == "__main__":
    sys.exit(main())
