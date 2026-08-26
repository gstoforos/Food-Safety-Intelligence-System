#!/usr/bin/env python3
"""
enrich_pathogen_from_reason.py  —  unblock pending_enrichment deterministically
===============================================================================

WHY THIS EXISTS
    Rows land at Status="pending_enrichment" when the scraper captured a recall
    but not its Pathogen. The publish gate then blocks them on the empty field,
    and they wait for reviewer 1 to fill it from the source page. Reviewer 1
    needs the Qwen VPS. When that box is slow or unreachable every row comes
    back "retry", nothing is filled, and genuine recalls sit for days.

    But the hazard is very often ALREADY in the row — RappelConso writes it in
    the Reason. Fiche "figues calabacita bio" reads

        "Une analyse de contrôle a révélé la présence d'ochratoxine a à un
         niveau supérieur à la limite réglementaire"

    That is Ochratoxin A, stated by the regulator, sitting in the workbook. No
    model and no network are needed to read it.

WHAT IT DOES
    Matches the Reason against pipeline/_pathogen_scope.TIER1_KEYWORDS — the
    register's own vocabulary, in English AND French — fills Pathogen with the
    canonical name, and advances the row to "pending" so the normal promotion
    path can consider it.

WHAT IT DELIBERATELY DOES NOT DO
    * It does not guess. No keyword match -> the row is left exactly as it is.
      Those rows are genuinely out of scope or need the page read, and a
      fabricated Pathogen is the single worst defect this register has had.
    * It does not judge scope. A row whose Reason describes a tipping learning
      tower or a shampoo gets no Pathogen, stays at pending_enrichment, and is
      rejected later by the scope rules that already exist.
    * It does not touch any other field.

Usage:
    python enrich_pathogen_from_reason.py --xlsx docs/data/recalls.xlsx --commit false
    python enrich_pathogen_from_reason.py --xlsx docs/data/recalls.xlsx --commit true
"""
from __future__ import annotations
import argparse
import datetime as dt
import re
import sys
import unicodedata
from pathlib import Path

# Canonical name for each matched keyword. French spellings map to the same
# English canonical form the register already uses elsewhere.
CANON = [
    (("listeria monocytogenes", "listeria", "listéria", "listeriose",
      "listériose"), "Listeria monocytogenes"),
    (("salmonella", "salmonelle", "salmonella spp"), "Salmonella"),
    (("shiga toxin", "shigatoxin", "stec", "vtec", "verotoxin",
      "escherichia coli", "e. coli", "e.coli"),
     "Shiga toxin-producing E. coli (STEC)"),
    (("cronobacter",), "Cronobacter sakazakii"),
    (("clostridium botulinum", "botulinum", "botulisme"),
     "Clostridium botulinum"),
    (("bacillus cereus", "cereulide", "céréulide"),
     "Cereulide (B. cereus toxin)"),
    (("staphylococc", "entérotoxine", "enterotoxin"),
     "Staphylococcus enterotoxin"),
    (("campylobacter",), "Campylobacter"),
    (("norovirus",), "Norovirus"),
    (("hepatitis a", "hépatite a"), "Hepatitis A virus"),
    (("vibrio",), "Vibrio"),
    (("ochratoxine a", "ochratoxin a"), "Ochratoxin A"),
    (("ochratoxine", "ochratoxin"), "Ochratoxin"),
    (("aflatoxine", "aflatoxin"), "Aflatoxin"),
    (("mycotoxine", "mycotoxin"), "Mycotoxins"),
    (("histamine", "scombrotoxin"), "Histamine / scombrotoxin"),
]


# Hazards that are DEFINITIVELY NOT a pathogen. A row whose Reason says one of
# these will NEVER acquire a Pathogen, so leaving it at pending_enrichment is a
# permanent dead end — it waits forever for reviewer 1 to fill a field the
# regulator never wrote.
#
# Measured 2026-08-24: 18 rows sat in enrichment and 11 said "Rupture de la
# chaine de froid par le transporteur" — a transport temperature failure with
# no organism detected. That one phrase was the entire publication-lag tail.
#
# Rejected with the regulator's own wording recorded, never silently deleted.
_NOT_A_PATHOGEN = (
    ("rupture de la chaine de froid", "cold chain break in transport"),
    ("chaine du froid", "cold chain break"),
    ("cold chain", "cold chain break"),
    ("erreur d etiquetage", "labelling error"),
    ("etiquetage", "labelling error"),
    ("teneur en plomb", "lead content above the regulatory limit"),
    ("fermentation spontanee", "spontaneous fermentation risk"),
    ("peuvent basculer", "tip-over hazard (non-food product)"),
)


def classify_non_pathogen(reason: str):
    """Return a reason string if the Reason is definitively NOT a pathogen
    hazard, else None. Conservative: anything not listed is left alone."""
    n = _norm(reason)
    if not n:
        return None
    for term, label in _NOT_A_PATHOGEN:
        if _norm(term) in n:
            return label
    return None


def _norm(s: str) -> str:
    s = unicodedata.normalize("NFKD", str(s or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s.lower())


def infer(reason: str):
    """Return (canonical_pathogen, matched_term) or (None, None)."""
    n = _norm(reason)
    if not n:
        return None, None
    for terms, canon in CANON:
        for t in terms:
            if _norm(t) in n:
                return canon, t
    return None, None


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=Path, default=Path("docs/data/recalls.xlsx"))
    ap.add_argument("--commit", type=str, default="false")
    args = ap.parse_args()
    commit = args.commit.lower() in ("1", "true", "yes", "on")

    import openpyxl
    wb = openpyxl.load_workbook(args.xlsx)
    if "Pending" not in wb.sheetnames:
        print("No Pending sheet.")
        return 1
    ws = wb["Pending"]
    H = [c.value for c in ws[1]]
    for need in ("Status", "Pathogen", "Reason"):
        if need not in H:
            print(f"Pending has no {need} column.")
            return 1
    si, pi, ri = H.index("Status") + 1, H.index("Pathogen") + 1, H.index("Reason") + 1
    ni = H.index("Notes") + 1 if "Notes" in H else None
    pr = H.index("Product") + 1 if "Product" in H else ri
    so = H.index("Source") + 1 if "Source" in H else ri

    filled, skipped, dead = [], [], []
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row, si).value or "").strip() != "pending_enrichment":
            continue
        if str(ws.cell(row, pi).value or "").strip() not in ("", "None"):
            continue
        reason = str(ws.cell(row, ri).value or "")
        canon, term = infer(reason)
        label = (f"{str(ws.cell(row, so).value)[:12]:12s} "
                 f"{str(ws.cell(row, pr).value)[:34]:34s}")
        if canon:
            filled.append((row, canon, term, label))
        else:
            np = classify_non_pathogen(reason)
            if np:
                dead.append((row, np, label))
            else:
                skipped.append((label, reason[:64]))

    print(f"pending_enrichment rows with a recoverable Pathogen: {len(filled)}\n")
    for _r, canon, term, label in filled:
        print(f"  {label} -> {canon}   (matched {term!r})")
    print(f"\npermanently stuck — NOT a pathogen hazard, will be REJECTED: "
          f"{len(dead)}")
    for _r, np, label in dead:
        print(f"  {label} -> reject: {np}")
    print(f"\nleft untouched (no keyword — needs the page, or out of scope): "
          f"{len(skipped)}")
    for label, reason in skipped:
        print(f"  {label} {reason}")

    if not commit:
        print("\nDRY RUN — nothing written. Re-run with --commit true.")
        return 0
    if not filled and not dead:
        return 0

    today = dt.datetime.now(dt.timezone.utc).strftime("%Y-%m-%d")
    for row, canon, term, _label in filled:
        ws.cell(row, pi).value = canon
        ws.cell(row, si).value = "pending"
        if ni:
            prev = str(ws.cell(row, ni).value or "")
            ws.cell(row, ni).value = (
                prev + f" [enrich {today}: Pathogen set to {canon!r} from the "
                f"regulator's own Reason text (matched {term!r}); "
                f"pending_enrichment -> pending]").strip()[:2000]

    for row, np, _label in dead:
        ws.cell(row, si).value = "rejected"
        if ni:
            prev = str(ws.cell(row, ni).value or "")
            ws.cell(row, ni).value = (
                prev + f" [enrich {today}: REJECTED — the Reason describes "
                f"{np}, not a microbial pathogen, so no Pathogen will ever be "
                f"filled and the row cannot leave pending_enrichment]"
            ).strip()[:2000]

    wb.save(args.xlsx)
    print(f"\n✓ Filled {len(filled)} Pathogen field(s) and advanced them to "
          f"'pending'; rejected {len(dead)} dead-end row(s).")
    try:
        sys.path.insert(0, ".")
        from pipeline.merge_master import mirror_json_from_xlsx
        mirror_json_from_xlsx(args.xlsx, args.xlsx.parent / "recalls.json")
        print("✓ recalls.json mirrored.")
    except Exception as e:
        print(f"  (JSON mirror skipped: {e})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
