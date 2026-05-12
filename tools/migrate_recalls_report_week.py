#!/usr/bin/env python3
"""
migrate_recalls_report_week.py — one-time migration

Adds a new internal column `report_week` to the Recalls sheet of
docs/data/recalls.xlsx and stamps existing rows according to the new
weekly-window rule:

    report_week = "W{nn}", where nn is the ISO week number of the
    SMALLEST Friday F such that F > Date.

In plain English: each weekly report ships Friday morning. A row dated
Friday F itself is NOT yet captured by F's AM scrape, so it lands in
the NEXT week's report. Concretely:

    Fri  May 1   →  next Friday > May 1 = May 8   → W19  (left BLANK by this script — pre-cutoff)
    Sat  May 2   →  next Friday > May 2 = May 8   → W19
    Thu  May 7   →  next Friday > May 7 = May 8   → W19
    Fri  May 8   →  next Friday > May 8 = May 15  → W20
    Thu  May 14  →  next Friday > May 14 = May 15 → W20

Per operator instruction (chat 2026-05-10):

  • Do NOT touch rows dated before 2026-05-02. Historical reports were
    built under the OLD inclusive-Friday rule and are already shipped;
    we are not retro-stamping them. The column stays blank for those
    rows and the weekly builder's date-math fallback handles them.

  • Stamp every row dated 2026-05-02 or later using the rule above.
    This stamps May 2-7 as W19 (the report just shipped Fri May 8 AM)
    and May 8 as W20 (the first row of next week's report — caught by
    the new rule because Friday-dated rows go to the following week).

After this migration:
  - merge_master.py promote_approved() will stamp every newly-promoted
    row at insert using the same rule (sticky — never overwritten).
  - docs/build_weekly_report_afts.py filter_week() will prefer the
    stamp; rows with a blank stamp fall back to the legacy date math.

Usage:
    python -m tools.migrate_recalls_report_week
    # or:
    python tools/migrate_recalls_report_week.py

Idempotent: if `report_week` column already exists, the script reports
the current stamp distribution and exits without rewriting unless
--force is passed.
"""
from __future__ import annotations

import argparse
import sys
from datetime import datetime, date, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "docs" / "data" / "recalls.xlsx"

# Cutoff: rows dated >= this get stamped. Rows dated < this stay blank.
# 2026-05-02 = Saturday after the Fri-May-1 (W18) ship date — the first
# day captured by the new-rule W19 window.
CUTOFF = date(2026, 5, 2)

EXPECTED_BASE_HEADERS = [
    "Date", "Source", "Company", "Brand", "Product", "Pathogen", "Reason",
    "Class", "Country", "Region", "Tier", "Outbreak", "URL", "Notes",
    "DateAdded", "LastUpdated", "LastChecked",
]
NEW_COL_NAME = "report_week"


def compute_report_week(date_str: str) -> str:
    """Return 'W{nn}' for the next Friday strictly after the row date.

    Returns '' if the date can't be parsed.
    """
    if not date_str:
        return ""
    try:
        d = datetime.strptime(str(date_str)[:10], "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return ""
    # Day-of-week: Mon=0, Tue=1, Wed=2, Thu=3, Fri=4, Sat=5, Sun=6
    days_until_next_friday = (4 - d.weekday()) % 7
    if days_until_next_friday == 0:
        # d itself is a Friday — STRICT next Friday is +7
        days_until_next_friday = 7
    next_friday = d + timedelta(days=days_until_next_friday)
    iso_week = next_friday.isocalendar()[1]
    return f"W{iso_week:02d}"


def parse_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, (datetime, date)):
        return v.date() if isinstance(v, datetime) else v
    try:
        return datetime.strptime(str(v)[:10], "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return None


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--xlsx", default=str(XLSX),
                    help="Path to recalls.xlsx (default: docs/data/recalls.xlsx)")
    ap.add_argument("--cutoff", default=CUTOFF.isoformat(),
                    help=f"Earliest date to stamp (default: {CUTOFF.isoformat()}). "
                         "Rows dated before this stay blank.")
    ap.add_argument("--dry-run", action="store_true",
                    help="Print what would change but don't save.")
    ap.add_argument("--force", action="store_true",
                    help="Re-stamp even if the column already exists. "
                         "Existing non-empty stamps are preserved unless --overwrite.")
    ap.add_argument("--overwrite", action="store_true",
                    help="With --force: also overwrite existing non-empty stamps. "
                         "Use with caution.")
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"ERROR: {xlsx_path} does not exist", file=sys.stderr)
        sys.exit(1)

    cutoff = datetime.strptime(args.cutoff, "%Y-%m-%d").date()

    print(f"Opening {xlsx_path}")
    wb = load_workbook(xlsx_path)
    if "Recalls" not in wb.sheetnames:
        print("ERROR: no Recalls sheet", file=sys.stderr)
        sys.exit(1)

    ws = wb["Recalls"]

    # Inspect current header row
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    print(f"Current headers ({len(headers)}): {headers}")

    if NEW_COL_NAME in headers:
        col_idx = headers.index(NEW_COL_NAME) + 1
        print(f"Column '{NEW_COL_NAME}' already exists at position {col_idx}.")
        if not args.force:
            # Show stamp distribution and exit
            from collections import Counter
            stamps = Counter()
            for r in range(2, ws.max_row + 1):
                v = ws.cell(row=r, column=col_idx).value
                stamps[v or "<blank>"] += 1
            print("Current stamp distribution:")
            for k, n in sorted(stamps.items(), key=lambda x: (x[0] != "<blank>", x[0])):
                print(f"  {k!r:>10}  {n}")
            print("\nUse --force to re-run (preserving existing non-empty stamps).")
            print("Use --force --overwrite to also overwrite existing stamps.")
            sys.exit(0)
    else:
        # Add a new column at the end
        col_idx = ws.max_column + 1
        c = ws.cell(row=1, column=col_idx, value=NEW_COL_NAME)
        c.font = Font(bold=True)
        print(f"Added new header '{NEW_COL_NAME}' at position {col_idx}")

    # Stamp every row dated >= cutoff
    stamped = 0
    preserved = 0
    skipped_old = 0
    skipped_no_date = 0
    overwritten = 0

    # Date is column 1
    for r in range(2, ws.max_row + 1):
        date_cell = ws.cell(row=r, column=1).value
        d = parse_date(date_cell)
        if d is None:
            skipped_no_date += 1
            continue
        if d < cutoff:
            skipped_old += 1
            continue

        existing = ws.cell(row=r, column=col_idx).value
        new_stamp = compute_report_week(d.isoformat())

        if existing and existing.strip():
            if args.overwrite:
                if existing != new_stamp:
                    ws.cell(row=r, column=col_idx, value=new_stamp)
                    overwritten += 1
                else:
                    preserved += 1
            else:
                preserved += 1
        else:
            ws.cell(row=r, column=col_idx, value=new_stamp)
            stamped += 1

    print()
    print(f"Cutoff: {cutoff.isoformat()} (rows dated before this stay blank)")
    print(f"Total data rows: {ws.max_row - 1}")
    print(f"  Newly stamped:       {stamped}")
    print(f"  Preserved (already): {preserved}")
    print(f"  Overwritten:         {overwritten}")
    print(f"  Skipped (pre-cutoff): {skipped_old}")
    print(f"  Skipped (no date):   {skipped_no_date}")

    # Show the new stamp distribution (rows >= cutoff only)
    from collections import Counter
    stamps = Counter()
    for r in range(2, ws.max_row + 1):
        d = parse_date(ws.cell(row=r, column=1).value)
        if d is None or d < cutoff:
            continue
        v = ws.cell(row=r, column=col_idx).value
        stamps[v or "<blank>"] += 1
    print(f"\nStamp distribution for rows dated >= {cutoff.isoformat()}:")
    for k, n in sorted(stamps.items()):
        print(f"  {k!r:>10}  {n}")

    if args.dry_run:
        print("\n--dry-run set; not saving.")
        return

    wb.save(xlsx_path)
    print(f"\nSaved {xlsx_path}")


if __name__ == "__main__":
    main()
