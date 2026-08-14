"""tools/wipe_weekly_rejected.py — empty the Weekly_Rejected sheet after
the Thursday review email goes out.

WHY THIS EXISTS (audit 2026-05-09)
==================================
Architectural twin of tools/wipe_weekly_review.py. Per operator spec:
every claude-check / openrouter-check rejection mirrors into
Weekly_Rejected (Thu 17:00 Athens cutoff window, just like
Weekly_Review). The Apps Script Thursday-17:00 mailer reads
docs/data/weekly-rejected-latest.json and includes the rejection list
alongside the promotions in the operator-only review email.

After that email goes out, the Weekly_Rejected sheet must be empty so
the new Thu→Thu window starts fresh — same lifecycle as Weekly_Review.

This script does the wipe. It runs from the existing
.github/workflows/weekly-review-wipe.yml workflow at Thursday 17:30
Athens — 30 minutes after the email send to make sure the mailer has
finished. Adding to the existing workflow rather than creating a new
one keeps the wipe semantics atomic: both review sheets reset together.

WHAT IT DOES
============
1. Loads docs/data/recalls.xlsx
2. Clears all data rows from the Weekly_Rejected sheet (header preserved)
3. Saves the xlsx
4. Regenerates docs/data/weekly-rejected-latest.json (it'll be empty,
   reflecting the new state — protection against stale data being read
   between wipe time and the next rejection).

The Recalls sheet is NOT touched.
The Pending sheet is NOT touched.
The Weekly_Review sheet is NOT touched (separate wipe step).
Only Weekly_Rejected (the rolling Thu→Thu rejection queue) gets emptied.

USAGE
=====
    python -m tools.wipe_weekly_rejected                # interactive (y/N)
    python -m tools.wipe_weekly_rejected --yes          # non-interactive
    python -m tools.wipe_weekly_rejected --dry-run      # report only

EXIT CODES
==========
    0 = wipe completed (or sheet already empty)
    1 = error (missing xlsx, can't write, etc.)
    2 = aborted by user
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

XLSX = ROOT / "docs" / "data" / "recalls.xlsx"
JSON = ROOT / "docs" / "data" / "weekly-rejected-latest.json"


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--yes", action="store_true",
                   help="Skip interactive confirmation")
    p.add_argument("--dry-run", action="store_true",
                   help="Print what would change; do not write")
    args = p.parse_args()

    if not XLSX.exists():
        print(f"ERROR: {XLSX} does not exist", file=sys.stderr)
        return 1

    # Lazy imports — keeps script startup fast
    from openpyxl import load_workbook  # noqa: E402
    from pipeline.weekly_rejected_capture import (  # noqa: E402
        SHEET_NAME, SHEET_COLS, export_week_slice,
    )

    wb = load_workbook(XLSX)
    if SHEET_NAME not in wb.sheetnames:
        print(f"Weekly_Rejected sheet does not exist in {XLSX}. "
              f"Nothing to wipe.")
        return 0

    ws = wb[SHEET_NAME]
    n_data_rows = max(0, ws.max_row - 1)  # subtract header

    if n_data_rows == 0:
        print(f"Weekly_Rejected sheet already empty (only header present).")
        return 0

    print(f"Weekly_Rejected currently has {n_data_rows} data row(s).")

    if args.dry_run:
        print(f"[dry-run] Would wipe {n_data_rows} row(s), keeping header.")
        return 0

    if not args.yes:
        resp = input(f"Wipe {n_data_rows} row(s) from Weekly_Rejected? [y/N] "
                     ).strip().lower()
        if resp not in ("y", "yes"):
            print("Aborted.")
            return 2

    # ── MOVE, DON'T DELETE (audit 2026-08-14) ──────────────────────────
    # This step used to delete the rows outright. The workflow that calls
    # it has always described a permanent archive —
    #     "Does NOT touch Rejected sheet (permanent audit archive —
    #      separate from the rolling Weekly_Rejected)"
    # — but no "Rejected" sheet has ever existed in recalls.xlsx. The
    # sheets are Recalls, Pending, Weekly_Rejected, NEWS. The archive the
    # design assumed was doing the remembering was never there, so every
    # Thursday at 17:30 Athens the reasons were destroyed.
    #
    # MEASURED DAMAGE: Weekly_Rejected went from 263 rows to 12 in one
    # wipe. Two FSANZ allergen rows removed on 2026-07-29 for a documented
    # reason lost their archive record entirely, which is what turned
    # tests/test_hazard_class_guard.py::test_they_were_archived_not_deleted
    # red on main — the only reason anyone noticed. They had to be
    # reconstructed by hand on 2026-08-14.
    #
    # It also erases the re-promotion memory: load_rejected_urls() reads
    # this sheet, so after a wipe the pipeline forgets which URLs a human
    # already turned down and re-ingests them. The USDA jalapeno public
    # health alert came back exactly this way after being archived on
    # 2026-08-13.
    #
    # The operator rule is standing and explicit: removed rows are
    # archived with a reason, NEVER silently deleted. A scheduled job is
    # not an exception to it.
    #
    # Rows now MOVE to a permanent "Rejected" sheet first. The rolling
    # Thu->Thu window still resets, which is the entire point of the wipe,
    # but the reasons survive. Dedup on (URL, Date) keeps it idempotent.
    if ws.max_row >= 2:
        archive = (wb["Rejected"] if "Rejected" in wb.sheetnames
                   else wb.create_sheet("Rejected"))
        hdr = [c.value for c in ws[1]]
        if archive.max_row < 1 or all(c.value in (None, "")
                                      for c in archive[1]):
            for i, h in enumerate(hdr, 1):
                archive.cell(row=1, column=i, value=h)
            arch_hdr = list(hdr)
        else:
            arch_hdr = [c.value for c in archive[1]]

        try:
            iu, idt = arch_hdr.index("URL"), arch_hdr.index("Date")
        except ValueError:
            iu = idt = None
        seen = set()
        if iu is not None:
            for t in archive.iter_rows(min_row=2, values_only=True):
                if t and not all(v in (None, "") for v in t):
                    seen.add((str(t[iu] or "").strip().lower(),
                              str(t[idt] or "")[:10]))

        moved = already = 0
        for t in ws.iter_rows(min_row=2, values_only=True):
            if all(v in (None, "") for v in t):
                continue
            row_map = dict(zip(hdr, t))
            if iu is not None:
                key = (str(row_map.get("URL") or "").strip().lower(),
                       str(row_map.get("Date") or "")[:10])
                if key in seen:
                    already += 1
                    continue
                seen.add(key)
            archive.append([row_map.get(h, "") for h in arch_hdr])
            moved += 1

        print(f"  ✓ Archived {moved} row(s) to the permanent 'Rejected' "
              f"sheet ({already} already present); archive now holds "
              f"{archive.max_row - 1}.")

        ws.delete_rows(2, ws.max_row - 1)

    # Defensive: if header is somehow missing or wrong, restore it.
    expected_headers = list(SHEET_COLS)
    actual_headers = [c.value for c in ws[1]]
    if actual_headers[:len(expected_headers)] != expected_headers:
        print("  WARN: header row was missing or different — restoring.")
        # Clear row 1 and rewrite
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=1, column=col_idx).value = None
        for i, h in enumerate(expected_headers, 1):
            ws.cell(row=1, column=i, value=h)

    XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(XLSX)
    print(f"  ✓ Wiped {n_data_rows} row(s) from Weekly_Rejected.")

    # Regenerate the JSON snapshot to reflect the empty state.
    # This protects against the Apps Script mailer reading stale rows
    # if it happens to fire between this wipe and the next rejection.
    try:
        result = export_week_slice(xlsx_path=XLSX, json_path=JSON)
        print(f"  ✓ Regenerated {JSON.name}: {result['row_count']} rows "
              f"for week ending {result['week_end']}")
    except Exception as e:
        print(f"  WARN: JSON regenerate failed: {e}", file=sys.stderr)
        # Don't fail the script — the xlsx wipe is the canonical action.
        # Apps Script can re-derive on next email.

    print("Done.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
