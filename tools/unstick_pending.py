#!/usr/bin/env python3
"""
tools/unstick_pending.py — drop stuck Pending rows so they can be re-scraped.

USAGE:
    python -m tools.unstick_pending [--dry-run] [--fids FIDS] [--source SRC]

Audit 2026-05-08 — written to unstick the 5 RappelConso rows that piled
up in Pending with Status=rejected because of the V1→V2 API field rename
+ the FIX-path-was-date-only bug. After this script runs, the next
orchestrator run captures these URLs fresh with the V2-patched scraper
(non-empty Company; sans marque → Unbranded) and the row enters Pending
with correct data.

Default behavior (no --fids or --source given) is to delete the specific
fids hard-coded in DEFAULT_TARGETS. Pass --fids "22196,22197,..." to
override; pass --source "RappelConso (FR)" to delete every Pending row
from a specific source.

Run with --dry-run first to confirm what would be deleted.

Files modified:
  docs/data/recalls.xlsx (Pending sheet)
  docs/data/recalls.json (mirror — rebuilt from xlsx after delete)
"""
from __future__ import annotations

import argparse
import json
import logging
import sys
from pathlib import Path
from typing import List, Set

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

XLSX_PATH = ROOT / "docs" / "data" / "recalls.xlsx"
JSON_PATH = ROOT / "docs" / "data" / "recalls.json"

# The 5 stuck French rows from the 2026-05-07/08 incident.
# All have Status=rejected, empty Company, Brand='—', Pathogen='Salmonella'.
# Page content names a real recalling firm (Pierre Sajous / Les Ateliers /
# SAJOUS PIERRE) but the V1 scraper missed it and the FIX path couldn't
# write it in. After this delete + next orchestrator run, the V2 scraper
# will re-capture them with non-empty Company.
DEFAULT_TARGETS = {
    "22196",  # Pierre Sajous (Origine: Sajous Pierre)
    "22197",  # SAJOUS PIERRE — sans marque retail loose meat
    "22198",  # SAJOUS PIERRE — sans marque retail loose meat
    "22199",  # SAJOUS PIERRE — sans marque retail loose meat
    "22211",  # Les Ateliers
}


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
log = logging.getLogger("unstick-pending")


def _parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__.split("\n\n")[0])
    p.add_argument("--dry-run", action="store_true",
                   help="Print what would be deleted without modifying files")
    p.add_argument("--fids", default="",
                   help=f"Comma-separated fids (e.g. '22196,22197'). "
                        f"Default: {','.join(sorted(DEFAULT_TARGETS))}")
    p.add_argument("--source", default="",
                   help="Delete all Pending rows from this source "
                        "(e.g. 'RappelConso (FR)'). Overrides --fids.")
    p.add_argument("--xlsx", default=str(XLSX_PATH),
                   help=f"Path to recalls.xlsx (default: {XLSX_PATH})")
    p.add_argument("--json", default=str(JSON_PATH),
                   help=f"Path to recalls.json (default: {JSON_PATH})")
    return p.parse_args()


def _row_matches_fid(url: str, target_fids: Set[str]) -> str:
    """Return the matching fid if URL contains /fiche-rappel/<fid>/, else ''."""
    import re
    m = re.search(r"/fiche-rappel/(\d+)/", url or "")
    if not m:
        return ""
    fid = m.group(1)
    return fid if fid in target_fids else ""


def _row_matches_source(row_source: str, target: str) -> bool:
    if not target:
        return False
    return (row_source or "").strip().lower() == target.strip().lower()


def main() -> int:
    args = _parse_args()
    xlsx_path = Path(args.xlsx)
    json_path = Path(args.json)

    if not xlsx_path.exists():
        log.error("xlsx not found at %s", xlsx_path)
        return 1

    # Decide target set
    if args.source:
        log.info("Mode: delete all Pending rows where Source=%r", args.source)
        target_fids: Set[str] = set()
        target_source = args.source
    else:
        target_fids = (set(s.strip() for s in args.fids.split(",") if s.strip())
                       if args.fids else set(DEFAULT_TARGETS))
        target_source = ""
        log.info("Mode: delete Pending rows with fid ∈ {%s}",
                 ", ".join(sorted(target_fids)))

    # Load workbook
    wb = openpyxl.load_workbook(xlsx_path)
    if "Pending" not in wb.sheetnames:
        log.error("xlsx has no Pending sheet — nothing to do")
        return 1
    ws = wb["Pending"]

    headers = [c.value for c in ws[1]]
    try:
        url_idx = headers.index("URL")
        src_idx = headers.index("Source")
    except ValueError as e:
        log.error("Pending sheet missing required column: %s", e)
        return 1

    # Find rows to delete (1-indexed for openpyxl row numbers)
    rows_to_delete: List[int] = []
    for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True),
                                 start=2):
        url = str(row[url_idx] or "")
        src = str(row[src_idx] or "")
        matched = False
        if target_source:
            if _row_matches_source(src, target_source):
                matched = True
        else:
            fid_match = _row_matches_fid(url, target_fids)
            if fid_match:
                matched = True
        if matched:
            rows_to_delete.append(r_idx)
            company = row[headers.index("Company")] if "Company" in headers else "?"
            pathogen = row[headers.index("Pathogen")] if "Pathogen" in headers else "?"
            log.info("  MATCH row %d  Source=%r  URL=%s  Company=%r  Pathogen=%r",
                     r_idx, src, url[-50:], company, pathogen)

    if not rows_to_delete:
        log.info("No matching rows in Pending — nothing to delete")
        return 0

    log.info("Will delete %d row(s) from Pending sheet", len(rows_to_delete))

    if args.dry_run:
        log.info("--dry-run: not modifying files")
        return 0

    # Delete in REVERSE order so row numbers don't shift mid-loop
    for r_idx in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(r_idx, 1)

    wb.save(xlsx_path)
    log.info("Saved xlsx after delete: Pending=%d rows (was %d)",
             ws.max_row - 1, ws.max_row - 1 + len(rows_to_delete))

    # Mirror to JSON. recalls.json is the public mirror of the Recalls
    # sheet only — Pending is xlsx-only — so the json doesn't actually
    # need rewriting for a Pending delete. But if a json mirror module
    # is available, call it for consistency. Otherwise skip.
    try:
        from pipeline.merge_master import mirror_json_from_xlsx
        mirror_json_from_xlsx(xlsx_path, json_path)
        log.info("Mirrored Recalls → %s", json_path)
    except Exception as e:
        log.info("Skipped json mirror (%s) — not required for Pending-only "
                 "delete; recalls.json reflects Recalls sheet only", e)

    log.info("Done. Run the orchestrator (or wait for the next scheduled "
             "scrape) and the deleted URLs will be re-captured with the "
             "V2-patched scraper.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
