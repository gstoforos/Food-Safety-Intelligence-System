"""
Simulation pass — apply the new Layer B/C/D validators to the EXISTING
Recalls sheet of docs/data/recalls.xlsx and report which rows would have
been rejected, WITHOUT writing anything.

Three categories reported:
  • Layer B  : URL-gate offline rejects
                 - news-host URLs (foodsafetynews.com etc.)
                 - bare landing pages (recalls-rappels.canada.ca/fr etc.)
                 - structural reject (categorie/, rubrik/, search?, no-path)
  • Layer C  : EXTRACTION_GARBAGE deterministic checks
                 - Company == Brand byte-for-byte AND >5 words
                 - Product is a bare domain ("canada.ca", "fda.gov")
                 - JS/HTML artifact in any of Company/Brand/Product
  • Layer D  : HALLUCINATED_PATHOGEN OFFLINE pre-flight
                 - Pathogen value (or any multilingual alias) NOT found
                   in the row's Reason+Notes text. Substring/case-insensitive.
                 - This is a STRICT subset of full Layer D; full Layer D
                   ALSO fetches the source URL and checks the page body.
                 - The offline check is what AI-review (Layer C
                   HALLUCINATED_PATHOGEN code) would catch. Layer D adds
                   the page-body check on top.

Read-only. No writes. Idempotent. Run from repo root:
    python tools/simulate_validators.py
"""
from __future__ import annotations

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook  # noqa: E402

from pipeline.url_gate_gemini import (  # noqa: E402
    _is_news_host, _has_js_html_artifact, _is_bare_domain, _is_structurally_bad,
)
from pipeline.verify_pathogen_in_source import _pathogen_in_text  # noqa: E402

XLSX_PATH = ROOT / "docs" / "data" / "recalls.xlsx"


def main() -> int:
    if not XLSX_PATH.exists():
        print(f"ERROR: {XLSX_PATH} not found", file=sys.stderr)
        return 1

    wb = load_workbook(XLSX_PATH, data_only=True)
    if "Recalls" not in wb.sheetnames:
        print("ERROR: Recalls sheet missing", file=sys.stderr)
        return 1
    ws = wb["Recalls"]
    headers = [c.value for c in ws[1]]
    rows = []
    for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v in (None, "") for v in row):
            continue
        rec = dict(zip(headers, row))
        rec["__row__"] = ridx
        rows.append(rec)

    print(f"Loaded {len(rows)} Recalls rows from {XLSX_PATH}")
    print("=" * 80)

    # --- Layer B simulation ---
    layerB_news = []
    layerB_landing = []
    layerB_structural = []
    for r in rows:
        url = str(r.get("URL") or "").strip()
        if not url:
            continue
        if _is_news_host(url):
            layerB_news.append(r)
            continue
        why = _is_structurally_bad(url)
        if why:
            if "domain only" in why or "no path" in why:
                layerB_landing.append((r, why))
            else:
                layerB_structural.append((r, why))

    # --- Layer C deterministic EXTRACTION_GARBAGE ---
    layerC_co_eq_br = []
    layerC_bare_domain = []
    layerC_js_artifact = []
    for r in rows:
        co = str(r.get("Company") or "").strip()
        br = str(r.get("Brand") or "").strip()
        pr = str(r.get("Product") or "").strip()
        if co and co == br and len(co.split()) > 5:
            layerC_co_eq_br.append(r)
        if _is_bare_domain(pr):
            layerC_bare_domain.append(r)
        for label, val in (("Company", co), ("Brand", br), ("Product", pr)):
            if _has_js_html_artifact(val):
                layerC_js_artifact.append((r, label, val))
                break

    # --- Layer D OFFLINE pre-flight (HALLUCINATED_PATHOGEN against Reason+Notes) ---
    layerD_offline = []
    for r in rows:
        path = str(r.get("Pathogen") or "").strip()
        reason = str(r.get("Reason") or "")
        notes = str(r.get("Notes") or "")
        haystack = (reason + " " + notes).strip()
        if not path or len(haystack) < 20:
            continue   # skip per Layer C rule
        if not _pathogen_in_text(path, haystack):
            layerD_offline.append((r, haystack[:150]))

    # ---- Report ----
    def hdr(s):
        print()
        print("─" * 80)
        print(s)
        print("─" * 80)

    hdr(f"Layer B — News-outlet URLs in Recalls: {len(layerB_news)}")
    for r in layerB_news[:50]:
        print(f"  row{r['__row__']:>4}  {r.get('Date')}  {(r.get('Source') or '')[:18]:18s}  {(r.get('URL') or '')[:90]}")

    hdr(f"Layer B — Landing-page / no-path URLs: {len(layerB_landing)}")
    for r, why in layerB_landing[:50]:
        print(f"  row{r['__row__']:>4}  {r.get('Date')}  {why}  {(r.get('URL') or '')[:80]}")

    hdr(f"Layer B — Structural reject (category/listing): {len(layerB_structural)}")
    for r, why in layerB_structural[:50]:
        print(f"  row{r['__row__']:>4}  {r.get('Date')}  {why}  {(r.get('URL') or '')[:80]}")

    hdr(f"Layer C — Company == Brand AND >5 words: {len(layerC_co_eq_br)}")
    for r in layerC_co_eq_br[:50]:
        print(f"  row{r['__row__']:>4}  {r.get('Date')}  Co=Br={(r.get('Company') or '')[:80]!r}")

    hdr(f"Layer C — Product is bare domain: {len(layerC_bare_domain)}")
    for r in layerC_bare_domain[:50]:
        print(f"  row{r['__row__']:>4}  {r.get('Date')}  Product={r.get('Product')!r}  URL={(r.get('URL') or '')[:60]}")

    hdr(f"Layer C — JS/HTML artifact in field: {len(layerC_js_artifact)}")
    for r, fld, val in layerC_js_artifact[:50]:
        print(f"  row{r['__row__']:>4}  {r.get('Date')}  {fld}={val[:90]!r}")

    hdr(f"Layer D OFFLINE — Pathogen NOT in Reason+Notes (multilingual): {len(layerD_offline)}")
    for r, hay in layerD_offline[:50]:
        print(f"  row{r['__row__']:>4}  {r.get('Date')}  {(r.get('Source') or '')[:15]:15s}  "
              f"Path={r.get('Pathogen')!r}")
        print(f"          Reason+Notes={hay[:120]!r}")
        print(f"          URL={(r.get('URL') or '')[:90]}")

    # ---- Summary ----
    flagged_rownums = set()
    for grp in (layerB_news, layerC_co_eq_br, layerC_bare_domain):
        for r in grp:
            flagged_rownums.add(r["__row__"])
    for grp in (layerB_landing, layerB_structural, layerC_js_artifact, layerD_offline):
        for r, *_ in grp:
            flagged_rownums.add(r["__row__"])

    print()
    print("=" * 80)
    print(f"SUMMARY: {len(flagged_rownums)} unique rows flagged (out of {len(rows)})")
    print("=" * 80)
    return 0


if __name__ == "__main__":
    sys.exit(main())
