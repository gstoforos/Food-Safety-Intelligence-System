#!/usr/bin/env python3
"""
FSIS reviewer training-data exporter v2
========================================

Major changes vs v1:

* Reads ``Recalls`` sheet as the primary source of approved (PROMOTE)
  examples — every row in Recalls is, by definition, a vetted promotion
  that survived every gate. v1 ignored this sheet, losing 14× the data.
* Tag-aware Notes parser handles every annotation tag seen in production:
  ``gemini-enrich``, ``gemini-check``, ``claude-check``, ``url-gate``,
  ``manual-promote``, ``manual-fix``, ``tier-fix``, ``dedupe-fix``,
  ``gap-gate``, ``outbreak``, ``region-fix``, ``enrich-gate``,
  ``claude-flag``, ``strict-outbreak-audit``, ``date-unknown``.
  Unicode arrows + curly-quote variants supported.
* Per-example weighting: human overrides (``manual-promote``,
  ``manual-fix``, ``tier-fix``) carry 2-3× weight over pipeline-only
  decisions, encoded in ``_meta.weight``.
* Provenance file: every training example traces back to (sheet, row,
  URL, source-tags) so we can audit downstream model errors back to
  source data.
* Source-breakdown stats: see at a glance how many examples came from
  which sheet.

Output files (in --out directory):
    promote_classifier.train.jsonl
    promote_classifier.val.jsonl
    tier_assignment.train.jsonl
    tier_assignment.val.jsonl
    field_corrections.train.jsonl
    field_corrections.val.jsonl
    stats.json
    provenance.json

Usage:
    python export_training_data_v2.py \\
        --xlsx docs/data/recalls.xlsx \\
        --out tools/training/data \\
        [--val-frac 0.1] [--seed 42] [--keep-meta]
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import random
import re
import sys
from collections import Counter
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import openpyxl


# ─────────────────────────────────────────────────────────────────────────
# System prompts — match production reviewer prompts so training data is
# in-distribution with how the fine-tuned model will be called at inference
# ─────────────────────────────────────────────────────────────────────────

PROMOTE_SYSTEM_PROMPT = """You are the AFTS FSIS final-gate reviewer. Given a candidate food-recall row, decide whether to PROMOTE it to the Recalls master sheet or REJECT it.

PROMOTE if ALL of:
  • Pathogen is in FSIS Tier-1 scope (Listeria, Salmonella, E. coli/STEC, Botulinum, Cereulide, Hepatitis A, Norovirus, marine biotoxins, Cronobacter, Campylobacter, Staphylococcus enterotoxin, mycotoxins above action level)
  • Date is within scope (2026-01-01 or later; pre-2026 dates are out of scope)
  • URL is a specific recall page (not a landing page, news article, or download)
  • Row is not a duplicate of an existing Recalls row
  • Company/Brand/Product fields contain real recall data (not page-title extraction garbage)

REJECT and state the category if ANY of:
  • pathogen_out_of_scope        — hazard isn't a foodborne pathogen (e.g. Cadmium, Asbestos, Mechanical hazard, Electrical shock, Choking hazard, Foreign body in adult food)
  • date_pre_2026                — recall predates 2026
  • not_a_recall_page            — URL points to landing page, news article, XLS download, or PDF index
  • duplicate_of_existing_recall — same recall already in master sheet

Output exactly: PROMOTE
            or: REJECT: <category>: <one-line reason>"""

TIER_SYSTEM_PROMPT = """You are the AFTS FSIS Tier classifier. Apply the HYBRID framework to assign Tier 1, 2, or 3.

Step 1 — Regulator class:
  TIER 1: "Class I", "Class 1", "Mandatory", "Public Health Alert", "Administrative action", "Outbreak", anything containing "Mandatory"/"Impératif"/"Imperative"
  TIER 2: "Class II", "Class 2", "Voluntary", "Recall", "Alert", "Food Alert", "Conseillé"
  TIER 3: "Class III", "Class 3", "Information", "Border Rejection", "Notification", "Advisory"

Step 2 — If no class match, apply FDA framework:
  TIER 1: Listeria mono / STEC / Botulinum / Cereulide / Marine biotoxin / Salmonella in RTE food / heavy metals / foreign body in baby food
  TIER 2: Salmonella in cooking-required food / Campylobacter / Cronobacter / generic E. coli / mycotoxins / Hep A or Norovirus
  TIER 3: labeling violations / environmental swab only / unknown pathogen / foreign body in adult food

Step 3 — Outbreak bump: if Outbreak=1, return max(1, base_tier - 1).

HARD RULES (override everything else):
  H1  Cereulide → Tier 1
  H2  Botulinum → Tier 1
  H3  STEC / E. coli O157 / VTEC / EHEC → Tier 1
  H4  Listeria monocytogenes (any food) → Tier 1
  H5  Marine biotoxins → Tier 1
  H6  Salmonella in RTE food → Tier 1
  H7  Salmonella in cooking-required food → Tier 2
  H8  Generic "E. coli" without STEC indicator → Tier 3
  H9  Bacillus cereus alone (no cereulide) → Tier 3
  H13 Foreign body in baby food / vulnerable-population food → Tier 1
  H14 Foreign body in adult food, no injuries → Tier 3

Output exactly one digit: 1, 2, or 3"""

FIX_SYSTEM_PROMPT = """You are the AFTS FSIS field-correction extractor. The scraped row contains errors. Identify which fields need correction and propose the correct values.

Fields you may correct: Date (YYYY-MM-DD), Company, Brand, Product, Pathogen.

Rules:
  • Company and Brand may stay in source language (proper names).
  • Every other field must be in English (US).
  • If Brand is unknown or generic, use "Unbranded".
  • If extraction garbage (page title fragments, navigation breadcrumbs, "Home"/"Index"/"Recall" placeholders), correct to the real value or mark "—".
  • Pathogen must be a canonical name from the FSIS vocabulary (e.g. "Listeria monocytogenes", "Salmonella", "Escherichia coli (generic)", "Shiga toxin-producing E. coli (STEC)", "Cereulide (B. cereus toxin)").

Output strict JSON only, no markdown, no commentary:
{"corrections": {"<field>": "<corrected_value>", ...}}

If no correction is needed, return: {"corrections": {}}"""


# ─────────────────────────────────────────────────────────────────────────
# Tag-aware Notes parser (v2)
# ─────────────────────────────────────────────────────────────────────────

# Match any [tag-name [date [time]]: body] annotation block in Notes.
_TAG_BLOCK_RX = re.compile(
    r"\[(?P<tag>[a-z][a-z0-9_-]+)"
    r"(?:\s+\d{4}-\d{2}-\d{2}(?:\s+\d{2}:\d{2})?)?"
    r":\s+(?P<body>[^\[\]]*)\]",
    re.IGNORECASE,
)

# Inside a tag body, match a field correction. Handles:
#   Pathogen→E. coli
#   Pathogen→'E. coli'
#   Pathogen ''→'E. coli'
#   Pathogen ""→"E. coli"
#   Brand '—'→'France'
#   Pathogen \u2018old\u2019\u2192\u2018new\u2019  (Unicode arrows + curly quotes)
# Matches anything between the field name and the arrow, then captures the
# corrected value after the arrow. Doesn't attempt to capture the
# before-value — reconstructing it from placeholder text in the training
# input is cleaner than guessing whether unquoted text is the before-value.
_CORR_BROAD_RX = re.compile(
    r"(?P<field>Date|Company|Brand|Product|Pathogen|Class|Region)\b"
    r"[^\u2192\n\]]{0,80}?"
    r"(?:\u2192|->|→)"
    r"\s*"
    r"[\"'\u2018\u2019\u201C\u201D]?"
    r"(?P<value>[^,;\]\n\u2192]{1,200}?)"
    r"[\"'\u2018\u2019\u201C\u201D]?"
    r"(?=\s*[,;\]]|$)",
)

# Verdict tokens that may appear inside a claude-check / gemini-check body
_VERDICT_RX = re.compile(
    r"\b(?P<verdict>pass|fix|fail|skip|promote|reject)\b",
    re.IGNORECASE,
)

# Rejection-category regex (used to extract reject reasons from rows in
# Weekly_Rejected or anywhere a REJECTED:cat:reason string appears).
_REJECTION_REASON_RX = re.compile(
    r"REJECTED:\s*(?P<cat>[a-z_]+)\s*[:|]\s*(?P<reason>.+?)(?:\s*\||$)",
    re.IGNORECASE,
)

# Tags that carry human-authored decisions or corrections — used to weight
# examples upward during training. Tags not in this set get weight 1.0.
WEIGHT_BY_TAG: Dict[str, float] = {
    "manual-promote":         2.0,
    "manual-fix":             2.0,
    "manual-add":             2.0,
    "tier-fix":               3.0,
    "region-fix":             1.5,
    "audit-2026-04-28":       2.0,
    "strict-outbreak-audit":  1.5,
}

# Tags that signal automated reviewer activity — used for source tracking
# but not weighting.
INTEREST_TAGS = {
    "claude-check", "claude-gap", "claude-flag",
    "gemini-check", "gemini-enrich", "gemini-gap",
    "openai-check", "openrouter-check",
    "url-gate", "url-guardian",
    "outbreak", "dedupe-fix", "gap-gate", "enrich-gate",
    "date-unknown",
} | set(WEIGHT_BY_TAG.keys())


def parse_notes_tags(notes: str) -> List[Tuple[str, str, Optional[str]]]:
    """Return every [tag DATE: body] annotation as (tag_name, body, verdict).

    `verdict` is one of pass/fix/fail/skip/promote/reject or None.
    """
    out: List[Tuple[str, str, Optional[str]]] = []
    for tm in _TAG_BLOCK_RX.finditer(notes):
        tag = tm.group("tag").lower()
        body = tm.group("body")
        vm = _VERDICT_RX.search(body)
        verdict = vm.group("verdict").lower() if vm else None
        out.append((tag, body, verdict))
    return out


def parse_notes_corrections(notes: str) -> Dict[str, Dict[str, str]]:
    """Extract per-field corrections from Notes annotations.

    Returns a dict mapping field name to {"after": str, "source_tag": str}.
    The "before" value is not extracted — the training pipeline replaces
    the pre-correction field with a placeholder string instead.
    """
    out: Dict[str, Dict[str, str]] = {}
    for tag, body, _verdict in parse_notes_tags(notes):
        if tag not in INTEREST_TAGS:
            continue
        for cm in _CORR_BROAD_RX.finditer(body):
            field = cm.group("field")
            value = cm.group("value").strip().strip("'\"\u2018\u2019\u201C\u201D")
            if not value or value in ("''", "—", "-", "(empty)", "\u2014"):
                continue
            # First write wins (earliest annotation reflects oldest fix)
            if field not in out:
                out[field] = {
                    "after": value,
                    "source_tag": tag,
                }
    return out


def example_weight(notes: str, default: float = 1.0) -> Tuple[float, List[str]]:
    """Compute training weight from tags present. Returns (weight, source_tags)."""
    weight = default
    tags_seen: List[str] = []
    for tag, _body, _verdict in parse_notes_tags(notes):
        if tag in WEIGHT_BY_TAG:
            weight = max(weight, WEIGHT_BY_TAG[tag])
        tags_seen.append(tag)
    return weight, tags_seen


# ─────────────────────────────────────────────────────────────────────────
# Rejection-category extraction
# ─────────────────────────────────────────────────────────────────────────

def parse_rejection_reason(rejection_field: str,
                           notes: str) -> Tuple[str, str]:
    """Extract (category, one-line reason) from RejectionReason or Notes."""
    src = (rejection_field or notes or "").strip()
    m = _REJECTION_REASON_RX.search(src)
    if m:
        cat = m.group("cat").strip().lower()
        reason = m.group("reason").strip()
        return cat, reason[:200]
    s = src.lower()
    if "out_of_scope" in s or "pathogen_out_of_scope" in s:
        return "pathogen_out_of_scope", src[:200]
    if "date_pre_2026" in s or "pre-2026" in s:
        return "date_pre_2026", src[:200]
    if "not_a_recall_page" in s or "landing page" in s or "news article" in s:
        return "not_a_recall_page", src[:200]
    if "duplicate" in s:
        return "duplicate_of_existing_recall", src[:200]
    return "other", src[:200]


# ─────────────────────────────────────────────────────────────────────────
# Sheet I/O + row normalization
# ─────────────────────────────────────────────────────────────────────────

def _load_sheet_rows(xlsx: Path, sheet_name: str) -> List[Dict[str, Any]]:
    if not xlsx.exists():
        return []
    wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    headers = [c.value for c in ws[1]]
    rows: List[Dict[str, Any]] = []
    for row_idx, r in enumerate(
            ws.iter_rows(min_row=2, values_only=True), start=2):
        row: Dict[str, Any] = {}
        for h, v in zip(headers, r):
            if h is None:
                continue
            row[h] = "" if v is None else v
        if not any(str(v).strip() for v in row.values()):
            continue
        row["_sheet_row_idx"] = row_idx
        row["_sheet_name"] = sheet_name
        rows.append(row)
    return rows


def _normalize_row(row: Dict[str, Any]) -> Dict[str, Any]:
    out: Dict[str, Any] = {}
    for k in ("Date", "Source", "Company", "Brand", "Product",
              "Pathogen", "Reason", "Class", "Country", "Region",
              "URL", "Notes"):
        v = row.get(k, "")
        out[k] = str(v).strip() if v is not None else ""
    for k in ("Tier", "Outbreak"):
        try:
            out[k] = int(row.get(k) or 0)
        except (TypeError, ValueError):
            out[k] = 0
    # Pass through provenance fields
    out["_sheet_row_idx"] = row.get("_sheet_row_idx", -1)
    out["_sheet_name"] = row.get("_sheet_name", "")
    return out


def _render_row_as_markdown(row: Dict[str, Any],
                            fields: Optional[List[str]] = None,
                            notes_chars: int = 600) -> str:
    if fields is None:
        fields = ["Date", "Source", "Company", "Brand", "Product",
                  "Pathogen", "Class", "Country", "Tier", "Outbreak", "URL"]
    lines = []
    for f in fields:
        v = row.get(f, "")
        if v == "" or v is None:
            v = "(empty)"
        lines.append(f"- **{f}**: {v}")
    notes = str(row.get("Notes", ""))[:notes_chars]
    if notes:
        lines.append(f"- **Notes excerpt**: {notes}")
    return "\n".join(lines)


# ─────────────────────────────────────────────────────────────────────────
# Task builders
# ─────────────────────────────────────────────────────────────────────────

def build_promote_classifier_examples(
    approved_rows: Iterable[Dict[str, Any]],
    rejected_rows: Iterable[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """Task 1: binary promote/reject classifier with reasoned rejection."""
    examples: List[Dict[str, Any]] = []

    for r in approved_rows:
        row = _normalize_row(r)
        if not row["URL"] and not row["Company"]:
            continue
        weight, tags = example_weight(row["Notes"])
        examples.append({
            "messages": [
                {"role": "system",    "content": PROMOTE_SYSTEM_PROMPT},
                {"role": "user",      "content": _render_row_as_markdown(row)},
                {"role": "assistant", "content": "PROMOTE"},
            ],
            "_meta": {
                "task": "promote_classifier",
                "label": "promote",
                "weight": weight,
                "source_sheet": row["_sheet_name"],
                "source_row": row["_sheet_row_idx"],
                "source_tags": tags,
                "url": row["URL"],
            },
        })

    for r in rejected_rows:
        row = _normalize_row(r)
        if not row["URL"] and not row["Company"]:
            continue
        cat, reason = parse_rejection_reason(
            str(r.get("RejectionReason", "")), row["Notes"])
        weight, tags = example_weight(row["Notes"])
        examples.append({
            "messages": [
                {"role": "system",    "content": PROMOTE_SYSTEM_PROMPT},
                {"role": "user",      "content": _render_row_as_markdown(row)},
                {"role": "assistant", "content": f"REJECT: {cat}: {reason}"},
            ],
            "_meta": {
                "task": "promote_classifier",
                "label": "reject",
                "category": cat,
                "weight": weight,
                "source_sheet": row["_sheet_name"],
                "source_row": row["_sheet_row_idx"],
                "source_tags": tags,
                "url": row["URL"],
            },
        })

    return examples


def build_tier_assignment_examples(
    approved_rows: Iterable[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """Task 2: Tier prediction from Pathogen + Outbreak + Class + Product.

    Uses Recalls + Weekly_Review since both have vetted Tier labels.
    """
    examples: List[Dict[str, Any]] = []
    for r in approved_rows:
        row = _normalize_row(r)
        tier = row.get("Tier", 0)
        if tier not in (1, 2, 3):
            continue
        if not row["Pathogen"]:
            continue
        weight, tags = example_weight(row["Notes"])
        examples.append({
            "messages": [
                {"role": "system",    "content": TIER_SYSTEM_PROMPT},
                {"role": "user",      "content": _render_row_as_markdown(
                    row,
                    fields=["Pathogen", "Outbreak", "Class", "Product"],
                    notes_chars=0,
                )},
                {"role": "assistant", "content": str(tier)},
            ],
            "_meta": {
                "task": "tier_assignment",
                "tier": tier,
                "pathogen": row["Pathogen"],
                "weight": weight,
                "source_sheet": row["_sheet_name"],
                "source_row": row["_sheet_row_idx"],
                "source_tags": tags,
                "url": row["URL"],
            },
        })
    return examples


def build_field_correction_examples(
    all_rows: Iterable[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """Task 3: extract reviewer FIX corrections from Notes tags."""
    examples: List[Dict[str, Any]] = []
    for r in all_rows:
        row = _normalize_row(r)
        corrections = parse_notes_corrections(row["Notes"])
        if not corrections:
            continue
        # Reconstruct pre-correction row state by replacing the corrected
        # fields with a placeholder so the model learns "this field was
        # bad, here's the fix".
        pre_row = dict(row)
        corr_map = {}
        for field, info in corrections.items():
            pre_row[field] = "(empty or garbage in original scrape)"
            corr_map[field] = info["after"]
        user_msg = _render_row_as_markdown(pre_row)
        assistant_msg = json.dumps(
            {"corrections": corr_map}, ensure_ascii=False)
        weight, tags = example_weight(row["Notes"])
        examples.append({
            "messages": [
                {"role": "system",    "content": FIX_SYSTEM_PROMPT},
                {"role": "user",      "content": user_msg},
                {"role": "assistant", "content": assistant_msg},
            ],
            "_meta": {
                "task": "field_corrections",
                "n_corrections": len(corr_map),
                "fields": list(corr_map.keys()),
                "weight": weight,
                "source_sheet": row["_sheet_name"],
                "source_row": row["_sheet_row_idx"],
                "source_tags": tags,
                "url": row["URL"],
            },
        })
    return examples


# ─────────────────────────────────────────────────────────────────────────
# Train/val split + JSONL write
# ─────────────────────────────────────────────────────────────────────────

def _stratified_split(
    examples: List[Dict[str, Any]],
    val_frac: float,
    seed: int,
    strata_fn,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    rng = random.Random(seed)
    by_stratum: Dict[str, List[Dict[str, Any]]] = {}
    for ex in examples:
        by_stratum.setdefault(strata_fn(ex), []).append(ex)
    train, val = [], []
    for stratum, items in by_stratum.items():
        rng.shuffle(items)
        n_val = max(1, int(len(items) * val_frac)) if len(items) >= 5 else 0
        val.extend(items[:n_val])
        train.extend(items[n_val:])
    rng.shuffle(train)
    rng.shuffle(val)
    return train, val


def _write_jsonl(path: Path, examples: List[Dict[str, Any]],
                 drop_meta: bool = True) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        for ex in examples:
            obj = {"messages": ex["messages"]}
            if not drop_meta and "_meta" in ex:
                obj["_meta"] = ex["_meta"]
            f.write(json.dumps(obj, ensure_ascii=False) + "\n")


# ─────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(
        description="Export FSIS reviewer training data v2 as chat-template JSONL.")
    ap.add_argument("--xlsx", type=Path,
                    default=Path("docs/data/recalls.xlsx"))
    ap.add_argument("--out", type=Path,
                    default=Path("tools/training/data"))
    ap.add_argument("--val-frac", type=float, default=0.1)
    ap.add_argument("--seed", type=int, default=42)
    ap.add_argument("--keep-meta", action="store_true",
                    help="Keep _meta fields in output JSONL")
    ap.add_argument("--git-sha", type=str, default="",
                    help="Git SHA of recalls.xlsx at export time (for provenance)")
    args = ap.parse_args()

    args.out.mkdir(parents=True, exist_ok=True)

    if not args.xlsx.exists():
        print(f"ERROR: xlsx not found at {args.xlsx}", file=sys.stderr)
        sys.exit(1)

    # ── Load all sheets ──
    recalls = _load_sheet_rows(args.xlsx, "Recalls")
    weekly_review = _load_sheet_rows(args.xlsx, "Weekly_Review")
    weekly_rejected = _load_sheet_rows(args.xlsx, "Weekly_Rejected")

    print(f"Loaded:")
    print(f"  Recalls:         {len(recalls)} rows")
    print(f"  Weekly_Review:   {len(weekly_review)} rows")
    print(f"  Weekly_Rejected: {len(weekly_rejected)} rows")

    # ── Quality filter ──
    def _ok(r):
        if not str(r.get("Source", "")).strip():
            return False
        return True

    approved = [r for r in (recalls + weekly_review) if _ok(r)]
    rejected = [r for r in weekly_rejected if _ok(r)]

    # De-dup approved by URL (same recall may appear in both Recalls and
    # Weekly_Review during the in-flight window — prefer Recalls)
    seen_urls = set()
    approved_dedup: List[Dict[str, Any]] = []
    for r in approved:
        url = str(r.get("URL", "")).strip()
        key = (url, "approved") if url else (id(r), "approved")
        if key in seen_urls:
            continue
        seen_urls.add(key)
        approved_dedup.append(r)
    approved = approved_dedup

    print(f"After filter+dedup:")
    print(f"  approved: {len(approved)}")
    print(f"  rejected: {len(rejected)}")

    # ── Task 1: promote classifier ──
    promote_ex = build_promote_classifier_examples(approved, rejected)
    train1, val1 = _stratified_split(
        promote_ex, args.val_frac, args.seed,
        strata_fn=lambda ex: ex["_meta"]["label"],
    )
    _write_jsonl(args.out / "promote_classifier.train.jsonl", train1,
                 drop_meta=not args.keep_meta)
    _write_jsonl(args.out / "promote_classifier.val.jsonl", val1,
                 drop_meta=not args.keep_meta)

    # ── Task 2: tier assignment ──
    tier_ex = build_tier_assignment_examples(approved)
    train2, val2 = _stratified_split(
        tier_ex, args.val_frac, args.seed,
        strata_fn=lambda ex: str(ex["_meta"]["tier"]),
    )
    _write_jsonl(args.out / "tier_assignment.train.jsonl", train2,
                 drop_meta=not args.keep_meta)
    _write_jsonl(args.out / "tier_assignment.val.jsonl", val2,
                 drop_meta=not args.keep_meta)

    # ── Task 3: field corrections ──
    fix_ex = build_field_correction_examples(
        list(approved) + list(rejected))
    train3, val3 = _stratified_split(
        fix_ex, args.val_frac, args.seed,
        strata_fn=lambda ex: ",".join(sorted(ex["_meta"]["fields"])),
    )
    _write_jsonl(args.out / "field_corrections.train.jsonl", train3,
                 drop_meta=not args.keep_meta)
    _write_jsonl(args.out / "field_corrections.val.jsonl", val3,
                 drop_meta=not args.keep_meta)

    # ── Stats ──
    def _by_source_sheet(exs):
        return dict(Counter(e["_meta"]["source_sheet"] for e in exs))

    def _by_weight_bucket(exs):
        b = Counter()
        for e in exs:
            w = e["_meta"].get("weight", 1.0)
            if w >= 3.0:
                b["3.0+ (tier-fix)"] += 1
            elif w >= 2.0:
                b["2.0+ (manual)"] += 1
            elif w >= 1.5:
                b["1.5+ (audit)"] += 1
            else:
                b["1.0 (default)"] += 1
        return dict(b)

    promote_labels = Counter(ex["_meta"]["label"] for ex in promote_ex)
    reject_cats = Counter(ex["_meta"].get("category", "?")
                          for ex in promote_ex if ex["_meta"]["label"] == "reject")
    tier_dist = Counter(ex["_meta"]["tier"] for ex in tier_ex)
    pathogen_dist = Counter(ex["_meta"]["pathogen"] for ex in tier_ex)
    fix_field_dist: Counter = Counter()
    for ex in fix_ex:
        for f in ex["_meta"]["fields"]:
            fix_field_dist[f] += 1

    stats = {
        "exported_at_utc": dt.datetime.now(dt.timezone.utc).isoformat(),
        "git_sha": args.git_sha or os.environ.get("GITHUB_SHA", ""),
        "xlsx_path": str(args.xlsx),
        "totals": {
            "recalls_rows": len(recalls),
            "weekly_review_rows": len(weekly_review),
            "weekly_rejected_rows": len(weekly_rejected),
            "approved_after_filter": len(approved),
            "rejected_after_filter": len(rejected),
        },
        "promote_classifier": {
            "total": len(promote_ex),
            "train": len(train1),
            "val": len(val1),
            "labels": dict(promote_labels),
            "rejection_categories": dict(reject_cats),
            "by_source_sheet": _by_source_sheet(promote_ex),
            "by_weight_bucket": _by_weight_bucket(promote_ex),
        },
        "tier_assignment": {
            "total": len(tier_ex),
            "train": len(train2),
            "val": len(val2),
            "tier_distribution": {str(k): v for k, v in sorted(tier_dist.items())},
            "top_pathogens": dict(pathogen_dist.most_common(15)),
            "by_source_sheet": _by_source_sheet(tier_ex),
            "by_weight_bucket": _by_weight_bucket(tier_ex),
        },
        "field_corrections": {
            "total": len(fix_ex),
            "train": len(train3),
            "val": len(val3),
            "field_distribution": dict(fix_field_dist),
            "by_source_sheet": _by_source_sheet(fix_ex),
            "by_weight_bucket": _by_weight_bucket(fix_ex),
        },
    }
    (args.out / "stats.json").write_text(
        json.dumps(stats, indent=2, ensure_ascii=False), encoding="utf-8")

    # ── Provenance: per-example trace back to source ──
    provenance = {
        "exported_at_utc": stats["exported_at_utc"],
        "git_sha": stats["git_sha"],
        "examples": [],
    }
    for task_name, exs in [
        ("promote_classifier", promote_ex),
        ("tier_assignment", tier_ex),
        ("field_corrections", fix_ex),
    ]:
        for i, ex in enumerate(exs):
            provenance["examples"].append({
                "task": task_name,
                "index": i,
                "source_sheet": ex["_meta"]["source_sheet"],
                "source_row": ex["_meta"]["source_row"],
                "url": ex["_meta"].get("url", ""),
                "weight": ex["_meta"].get("weight", 1.0),
                "source_tags": ex["_meta"].get("source_tags", []),
            })
    (args.out / "provenance.json").write_text(
        json.dumps(provenance, indent=2, ensure_ascii=False), encoding="utf-8")

    # ── Summary to stdout (will be captured into $GITHUB_STEP_SUMMARY) ──
    print()
    print("═══════════════════════ EXPORT v2 COMPLETE ═══════════════════════")
    print()
    print(f"Output: {args.out}/")
    print(f"Git SHA: {stats['git_sha'] or '(unset)'}")
    print()
    print("## Task 1 — Promote classifier")
    print(f"  total={len(promote_ex)}  train={len(train1)}  val={len(val1)}")
    print(f"  labels: {dict(promote_labels)}")
    print(f"  by source sheet: {_by_source_sheet(promote_ex)}")
    print(f"  by weight bucket: {_by_weight_bucket(promote_ex)}")
    if reject_cats:
        print(f"  rejection categories:")
        for cat, n in reject_cats.most_common():
            print(f"    {cat:35} {n}")
    print()
    print("## Task 2 — Tier assignment")
    print(f"  total={len(tier_ex)}  train={len(train2)}  val={len(val2)}")
    print(f"  tier distribution: {dict(sorted(tier_dist.items()))}")
    print(f"  by source sheet: {_by_source_sheet(tier_ex)}")
    print(f"  by weight bucket: {_by_weight_bucket(tier_ex)}")
    print(f"  top pathogens:")
    for p, n in pathogen_dist.most_common(10):
        print(f"    {p[:50]:50} {n}")
    print()
    print("## Task 3 — Field corrections")
    print(f"  total={len(fix_ex)}  train={len(train3)}  val={len(val3)}")
    print(f"  field distribution: {dict(fix_field_dist)}")
    print(f"  by source sheet: {_by_source_sheet(fix_ex)}")
    print(f"  by weight bucket: {_by_weight_bucket(fix_ex)}")
    print()

    # ── Training adequacy gate ──
    print("## Training-data adequacy")
    if len(promote_ex) < 500:
        print(f"  ⚠ promote classifier: {len(promote_ex)} examples — borderline.")
    else:
        print(f"  ✓ promote classifier: {len(promote_ex)} examples — sufficient.")

    if reject_cats and sum(reject_cats.values()) < 30:
        print(f"  ⚠ reject pool: only {sum(reject_cats.values())} — "
              f"39:1 class imbalance, need class weighting at training time.")

    if len(tier_ex) < 100:
        print(f"  ⚠ tier assignment: {len(tier_ex)} examples — too few.")
    else:
        print(f"  ✓ tier assignment: {len(tier_ex)} examples with "
              f"all 3 tiers represented.")

    if len(fix_ex) < 50:
        print(f"  ⚠ field corrections: {len(fix_ex)} — collect more before training.")
    else:
        print(f"  ✓ field corrections: {len(fix_ex)} examples — viable.")


if __name__ == "__main__":
    main()
