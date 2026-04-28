#!/usr/bin/env python3
"""
One-off migration: backfill DateAdded / LastUpdated / LastChecked columns
for every existing row in the Recalls sheet of docs/data/recalls.xlsx.

Backfill rules:
  DateAdded   = the row's existing Date field (best approximation — we
                don't know exactly when the row was promoted, but Date is
                always ≤ DateAdded for legitimate rows).
  LastUpdated = today (we just touched it by adding the column).
  LastChecked = "" (empty — no URL has been re-validated under the new
                policy yet; url_guardian will populate on next run).

Run from repo root:
    python tools/migrate_recalls_internal_cols.py            # dry-run
    python tools/migrate_recalls_internal_cols.py --apply    # write

Idempotent: rows that already have DateAdded set are skipped.
"""
import argparse
import sys
from datetime import date
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = ROOT / "docs" / "data" / "recalls.xlsx"
INTERNAL_COLS = ["DateAdded", "LastUpdated", "LastChecked"]


def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument("--apply", action="store_true",
                   help="Write changes (default: dry-run)")
    args = p.parse_args()

    if not XLSX_PATH.exists():
        print(f"FAIL: {XLSX_PATH} not found")
        return 1

    wb = load_workbook(XLSX_PATH)
    if "Recalls" not in wb.sheetnames:
        print("FAIL: no Recalls sheet")
        return 1
    ws = wb["Recalls"]

    # Read header row
    headers = [c.value for c in ws[1]]
    today = date.today().isoformat()

    # Append the 3 new columns if not present
    new_cols = [c for c in INTERNAL_COLS if c not in headers]
    if new_cols:
        print(f"Will add columns to header: {new_cols}")
        for col_name in new_cols:
            headers.append(col_name)
            ws.cell(row=1, column=len(headers), value=col_name)

    # Find column indices
    date_idx = headers.index("Date") + 1
    date_added_idx = headers.index("DateAdded") + 1
    last_updated_idx = headers.index("LastUpdated") + 1
    last_checked_idx = headers.index("LastChecked") + 1

    # Backfill row by row
    backfilled = 0
    skipped = 0
    for row_num in range(2, ws.max_row + 1):
        existing_date_added = ws.cell(row=row_num, column=date_added_idx).value
        if existing_date_added:
            skipped += 1
            continue
        date_val = ws.cell(row=row_num, column=date_idx).value
        if hasattr(date_val, "isoformat"):
            date_str = date_val.isoformat()[:10]
        else:
            date_str = str(date_val or "")[:10]
        ws.cell(row=row_num, column=date_added_idx, value=date_str)
        ws.cell(row=row_num, column=last_updated_idx, value=today)
        ws.cell(row=row_num, column=last_checked_idx, value="")
        backfilled += 1

    print(f"Backfilled: {backfilled} rows")
    print(f"Skipped (already migrated): {skipped} rows")

    if not args.apply:
        print("\nDry run — re-run with --apply to write.")
        return 0

    wb.save(XLSX_PATH)
    print(f"\nSaved {XLSX_PATH}")
    print("Columns added: " + ", ".join(INTERNAL_COLS))
    print("Don't forget to commit + push.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
