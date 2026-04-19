#!/usr/bin/env python3
"""
One-off dataset cleanup — run after deploying the defensive-validation
changes in scrapers/_models.py.

Does three things:
  1. Drops rows whose Pathogen is garbage (paragraph leakage, "—", empty).
  2. Re-runs normalize_pathogen over every remaining row, so legacy
     non-canonical values like "Salmonella", "Salmonella spp",
     "Aflatoxin", "E. coli STEC" get rewritten to canonical names.
  3. Re-runs normalize_country + assign_tier over every row so the Tier
     column reflects the latest taxonomy (some legacy rows have Tier=2
     on hazards now classified as Tier=1).

Input/output paths mirror run_all.py: docs/data/recalls.xlsx is the
source of truth, docs/data/recalls.json is mirrored from it at the end.

Usage:
    python tools/cleanup_dataset.py             # dry run — prints diffs
    python tools/cleanup_dataset.py --apply     # write changes
    python tools/cleanup_dataset.py --apply --backup    # backup first
"""
from __future__ import annotations
import argparse
import json
import shutil
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from scrapers._models import (   # noqa: E402
    normalize_pathogen, normalize_country, assign_tier, infer_region,
)

DATA_DIR = ROOT / "docs" / "data"
XLSX = DATA_DIR / "recalls.xlsx"
JSON_PATH = DATA_DIR / "recalls.json"

# Rejection rules — rows matching any of these get dropped, not re-normalized.
PATHOGEN_MAX_LEN = 80
SENTINELS = {"", "—", "-", "unknown", "none", "n/a", "na"}


def should_drop(row: dict) -> tuple[bool, str]:
    """Return (drop?, reason). Mirrors the defensive logic in Recall.normalize."""
    p = (row.get("Pathogen") or "").strip()
    prod = (row.get("Product") or "").strip()
    if not p:
        return True, "empty Pathogen"
    if p.lower() in SENTINELS:
        return True, f"sentinel Pathogen ({p!r})"
    if len(p) > PATHOGEN_MAX_LEN:
        return True, f"paragraph Pathogen ({len(p)} chars)"
    if p == prod:
        return True, "Pathogen == Product (boilerplate leakage)"
    return False, ""


def clean_row(row: dict) -> tuple[dict, list[str]]:
    """Apply canonicalization. Returns (new_row, changes_applied)."""
    out = dict(row)
    changes = []

    p_old = (out.get("Pathogen") or "").strip()
    p_new = normalize_pathogen(p_old)
    if p_new != p_old:
        out["Pathogen"] = p_new
        changes.append(f"Pathogen: {p_old!r} -> {p_new!r}")

    c_old = (out.get("Country") or "").strip()
    if c_old:
        c_new = normalize_country(c_old)
        if c_new != c_old:
            out["Country"] = c_new
            changes.append(f"Country: {c_old!r} -> {c_new!r}")

    # Region backfill if missing but Country known
    if not (out.get("Region") or "").strip() and out.get("Country"):
        region = infer_region(out["Country"])
        if region:
            out["Region"] = region
            changes.append(f"Region: backfilled to {region!r}")

    # Tier recompute under current taxonomy
    try:
        outbreak = int(out.get("Outbreak", 0) or 0)
    except (TypeError, ValueError):
        outbreak = 0
    t_old = out.get("Tier", 2)
    try:
        t_old_int = int(t_old)
    except (TypeError, ValueError):
        t_old_int = 2
    t_new = assign_tier(out.get("Pathogen", ""), outbreak)
    if t_new != t_old_int:
        out["Tier"] = t_new
        changes.append(f"Tier: {t_old_int} -> {t_new}")

    return out, changes


def load_xlsx() -> tuple[list[dict], list[dict], list[dict], list[str]]:
    """Return (recalls_rows, pending_rows, news_rows, recalls_header_order)."""
    import openpyxl
    wb = openpyxl.load_workbook(XLSX)

    def _sheet_as_dicts(ws):
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            rows.append({h: v for h, v in zip(header, r)})
        return rows, header

    rec_rows, rec_header = _sheet_as_dicts(wb["Recalls"])
    pen_rows, _ = _sheet_as_dicts(wb["Pending"]) if "Pending" in wb.sheetnames else ([], [])
    nws_rows, _ = _sheet_as_dicts(wb["NEWS"]) if "NEWS" in wb.sheetnames else ([], [])
    return rec_rows, pen_rows, nws_rows, rec_header


def save_xlsx(recalls: list[dict], pending: list[dict], news: list[dict],
              recalls_header: list[str]) -> None:
    """Write back using the original header orders so downstream readers are happy."""
    import openpyxl
    wb = openpyxl.load_workbook(XLSX)

    def _write_sheet(ws_name, rows, header):
        if ws_name not in wb.sheetnames:
            return
        ws = wb[ws_name]
        # Clear existing data rows (preserve header)
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row)
        for r in rows:
            ws.append([r.get(h) for h in header])

    _write_sheet("Recalls", recalls, recalls_header)
    # Rewrite Pending with its own header too
    if "Pending" in wb.sheetnames and pending:
        pen_header = [c.value for c in next(wb["Pending"].iter_rows(min_row=1, max_row=1))]
        _write_sheet("Pending", pending, pen_header)
    if "NEWS" in wb.sheetnames and news:
        nws_header = [c.value for c in next(wb["NEWS"].iter_rows(min_row=1, max_row=1))]
        _write_sheet("NEWS", news, nws_header)
    wb.save(XLSX)


def mirror_json() -> None:
    import openpyxl
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Recalls"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        rows.append({h: (v if v is not None else "") for h, v in zip(header, r)})
    JSON_PATH.write_text(json.dumps(rows, ensure_ascii=False, indent=1))


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--apply", action="store_true",
                    help="Actually write changes. Without this flag, prints a dry-run diff.")
    ap.add_argument("--backup", action="store_true",
                    help="With --apply: copy recalls.xlsx to recalls.xlsx.bak-YYYYMMDD first.")
    ap.add_argument("--verbose", "-v", action="store_true",
                    help="Print every per-row change, not just the summary.")
    args = ap.parse_args()

    if not XLSX.exists():
        print(f"ERROR: {XLSX} not found", file=sys.stderr)
        return 1

    rec_rows, pen_rows, nws_rows, rec_header = load_xlsx()
    print(f"Loaded: {len(rec_rows)} Recalls, {len(pen_rows)} Pending, "
          f"{len(nws_rows)} NEWS rows")

    # Pass 1: drop garbage rows
    kept: list[dict] = []
    dropped: list[tuple[dict, str]] = []
    for row in rec_rows:
        drop, reason = should_drop(row)
        if drop:
            dropped.append((row, reason))
        else:
            kept.append(row)

    # Pass 2: canonicalize survivors
    changes_by_row: list[list[str]] = []
    cleaned: list[dict] = []
    for row in kept:
        new_row, changes = clean_row(row)
        cleaned.append(new_row)
        changes_by_row.append(changes)

    # Summary
    print(f"\n=== DROPPED {len(dropped)} rows ===")
    for row, reason in dropped:
        p = (row.get("Pathogen") or "")[:60]
        print(f"  {row.get('Date','?')} | {row.get('Source','?'):24s} | "
              f"{reason:40s} | Pathogen={p!r}")

    total_changes = sum(len(c) for c in changes_by_row)
    rows_changed = sum(1 for c in changes_by_row if c)
    print(f"\n=== NORMALIZED {rows_changed} rows ({total_changes} field edits) ===")

    # Pathogen value histogram before/after
    before = Counter((r.get("Pathogen") or "").strip() for r in rec_rows)
    after  = Counter((r.get("Pathogen") or "").strip() for r in cleaned)
    print(f"Distinct Pathogen values: {len(before)} -> {len(after)}")

    # Top changes (most common Pathogen renames)
    if args.verbose:
        print("\n--- per-row changes ---")
        for i, (row, changes) in enumerate(zip(cleaned, changes_by_row)):
            if changes:
                print(f"  {row.get('Date','?')} | {row.get('Source','?'):24s}")
                for c in changes:
                    print(f"    {c}")
    else:
        rename_pairs = Counter()
        for row_orig, row_new, changes in zip(kept, cleaned, changes_by_row):
            for c in changes:
                if c.startswith("Pathogen:"):
                    rename_pairs[c] += 1
        print("\nMost common Pathogen renames:")
        for change, n in rename_pairs.most_common(20):
            print(f"  [{n:3d}x] {change}")

    print(f"\nFinal row count: {len(cleaned)} (was {len(rec_rows)}, "
          f"-{len(rec_rows) - len(cleaned)})")

    if not args.apply:
        print("\nDRY RUN — no changes written. Re-run with --apply to commit.")
        return 0

    if args.backup:
        ts = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
        bak = XLSX.with_suffix(f".xlsx.bak-{ts}")
        shutil.copy2(XLSX, bak)
        print(f"\nBackup: {bak}")

    save_xlsx(cleaned, pen_rows, nws_rows, rec_header)
    mirror_json()
    print(f"\nWrote: {XLSX}  (+{JSON_PATH})")
    return 0


if __name__ == "__main__":
    sys.exit(main())
