"""tools/wipe_weekly_review.py — empty the Weekly_Review sheet after the
Thursday review email goes out.

WHY THIS EXISTS (audit 2026-05-08)
==================================
Per operator spec: every promotion mirrors into Weekly_Review (Thu 17:00
Athens cutoff window). The Apps Script `sendThursdayManualReview` mailer
reads `docs/data/weekly-review-latest.json` and sends the operator-only
review email at Thursday 17:00 Athens. After that email goes out, the
Weekly_Review sheet must be empty so the new Thu→Thu window starts fresh.

This script does the wipe. It runs from a separate workflow
(`.github/workflows/weekly-review-wipe.yml`) at Thursday 17:30 Athens —
30 minutes after the email send to make sure the mailer has finished.

WHAT IT DOES
============
1. Loads `docs/data/recalls.xlsx`
2. Clears all data rows from the Weekly_Review sheet (header preserved)
3. Saves the xlsx
4. Regenerates `docs/data/weekly-review-latest.json` (it'll be empty,
   reflecting the new state — protection against stale data being read
   between wipe time and the next promotion).

The Recalls sheet is NOT touched. Promoted rows persist there forever.
The Rejected sheet is NOT touched. The audit archive persists forever.
Only Weekly_Review (the rolling Thu→Thu review queue) gets emptied.

USAGE
=====
    python -m tools.wipe_weekly_review                # interactive (y/N)
    python -m tools.wipe_weekly_review --yes          # non-interactive
    python -m tools.wipe_weekly_review --dry-run      # report only

EXIT CODES
==========
    0 = wipe completed (or sheet already empty)
    1 = error (missing xlsx, can't write, etc.)
    2 = aborted by user
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

XLSX = ROOT / "docs" / "data" / "recalls.xlsx"
JSON = ROOT / "docs" / "data" / "weekly-review-latest.json"


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--yes", action="store_true",
                   help="Skip interactive confirmation")
    p.add_argument("--dry-run", action="store_true",
                   help="Print what would change; do not write")
    args = p.parse_args()

    if not XLSX.exists():
        print(f"ERROR: {XLSX} does not exist", file=sys.stderr)
        return 1

    # Lazy imports — keeps script startup fast
    from openpyxl import load_workbook  # noqa: E402
    from pipeline.weekly_review_capture import (  # noqa: E402
        SHEET_NAME, SHEET_COLS, export_week_slice,
    )

    wb = load_workbook(XLSX)
    if SHEET_NAME not in wb.sheetnames:
        print(f"Weekly_Review sheet does not exist in {XLSX}. "
              f"Nothing to wipe.")
        return 0

    ws = wb[SHEET_NAME]
    n_data_rows = max(0, ws.max_row - 1)  # subtract header

    if n_data_rows == 0:
        print(f"Weekly_Review sheet already empty (only header present).")
        return 0

    print(f"Weekly_Review currently has {n_data_rows} data row(s).")

    if args.dry_run:
        print(f"[dry-run] Would wipe {n_data_rows} row(s), keeping header.")
        return 0

    if not args.yes:
        resp = input(f"Wipe {n_data_rows} row(s) from Weekly_Review? [y/N] "
                     ).strip().lower()
        if resp not in ("y", "yes"):
            print("Aborted.")
            return 2

    # Wipe data rows: delete every row from row 2 to the end.
    # openpyxl's delete_rows(idx, amount) takes the FIRST row index and
    # number of rows. Row 1 = header, so we delete rows 2..max_row.
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    # Defensive: if header is somehow missing or wrong, restore it.
    expected_headers = list(SHEET_COLS)
    actual_headers = [c.value for c in ws[1]]
    if actual_headers[:len(expected_headers)] != expected_headers:
        print("  WARN: header row was missing or different — restoring.")
        # Clear row 1 and rewrite
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=1, column=col_idx).value = None
        for i, h in enumerate(expected_headers, 1):
            ws.cell(row=1, column=i, value=h)

    XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(XLSX)
    print(f"  ✓ Wiped {n_data_rows} row(s) from Weekly_Review.")

    # Regenerate the JSON snapshot to reflect the empty state.
    # This protects against the Apps Script mailer reading stale rows
    # if it happens to fire between this wipe and the next promotion.
    try:
        result = export_week_slice(xlsx_path=XLSX, json_path=JSON)
        print(f"  ✓ Regenerated {JSON.name}: {result['row_count']} rows "
              f"for week ending {result['week_end']}")
    except Exception as e:
        print(f"  WARN: JSON regenerate failed: {e}", file=sys.stderr)
        # Don't fail the script — the xlsx wipe is the canonical action.
        # Apps Script can re-derive on next email.

    print("Done.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
