#!/usr/bin/env python3
"""
Audit 2026-04-28 — DATA FIXES (v2: URL-matched, row-shift-safe)

This is a rewrite of fix_audit_2026_04_28.py that finds target rows by
URL or content fragment instead of row number. Row indices in xlsx shift
constantly as the hourly Pending->Recalls merge promotes new rows; an
old version of this script using fixed row numbers (5, 234, 250, etc.)
clobbered legitimate data when run after row indices had shifted.

WHAT THIS DOES (idempotent — safe to re-run):

  1. Row with RappelConso fiche 22129 + Company "—":
       Set Company and Brand to "Sans marque" (RappelConso convention
       for unbranded farmhouse products, used in 30+ other rows).

  2. Row with briefly.co.za URL (NAN Special Pro HA):
       Replace URL with the official NCC press release URL.

  3. Row with xataka.com.mx URL (NAN Alfamino):
       Replace URL with the official COFEPRIS PDF URL.

  4. Five rows still using news-aggregator URLs (iol.co.za,
     businesstech.co.za, htxt.co.za, timelessnews.co.za, unotv.com):
       Append [audit-2026-04-28] flag to Notes so url_resurrect.yml
       picks them up on its next run. URL itself is preserved.

  5. Append one new row: RappelConso fiche 20978 (FIGUES SECHEES LERIDA,
     16/01/2026, AREV, Aflatoxin + tenuazonic acid). This French
     mycotoxin recall was missed by the old hazard whitelist before the
     mycotoxin patch deployed today. Idempotent — skipped if already
     present.

  Each modified row gets LastUpdated set to today.

Run from repo root:
    python tools/fix_audit_2026_04_28.py            # dry-run
    python tools/fix_audit_2026_04_28.py --apply    # write
"""
import argparse
import sys
from datetime import date
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = ROOT / "docs" / "data" / "recalls.xlsx"

# ──────────────────────────────────────────────────────────────────────────
# Each fix specifies: how to FIND the row, then what to UPDATE.
# Match logic finds the FIRST row where every condition in `match` holds.
# Conditions are (column, substring) pairs — case-sensitive.
# Idempotency: if the row already has the target value, the fix is a no-op.
# ──────────────────────────────────────────────────────────────────────────
ROW_FIXES = [
    {
        "label": "RappelConso fiche 22129 — set Company/Brand to 'Sans marque'",
        "match": [("URL", "fiche-rappel/22129"), ("Company", "—")],
        "updates": {"Company": "Sans marque", "Brand": "Sans marque"},
    },
    {
        "label": "NAN Special Pro HA — briefly.co.za → official NCC press release",
        "match": [("URL", "briefly.co.za"), ("Brand", "NAN Special Pro HA")],
        "updates": {"URL": "https://thencc.org.za/product-recall-nan-special-pro-ha-infant-formula-800g/"},
    },
    {
        "label": "NAN Alfamino — xataka.com.mx → official COFEPRIS PDF",
        "match": [("URL", "xataka.com.mx"), ("Brand", "NAN Alfamino")],
        "updates": {"URL": "https://www.gob.mx/cms/uploads/attachment/file/1047052/Alerta_Sanitaria__Nestl__07012026.pdf"},
    },
]

# ──────────────────────────────────────────────────────────────────────────
# Rows to DELETE entirely. Use sparingly — only for rows where the data is
# unsalvageable (page-navigation extracted as Company, dead URL with no real
# recall behind it, etc.).
# ──────────────────────────────────────────────────────────────────────────
ROW_DELETIONS = [
    {
        "label": "CFIA garbage row — page nav extracted as Company/Brand/Product",
        # The /fr/ URL plus "Trouvez des rappels" in Company is unique to this
        # broken row — no legitimate row has both. URL alone is not enough
        # because url_resurrect could later create a clean row with same URL.
        "match": [
            ("URL", "/fr/avis-rappel/certains-produits-fromagers"),
            ("Company", "Trouvez des rappels"),
        ],
    },
]

# News-aggregator URLs to flag for url_resurrect (URL kept; Notes annotated)
URL_RESURRECT_FLAGS = [
    {"label": "Made for Tots Corn Puffs (NCC ZA)", "match": [("URL", "iol.co.za"), ("Brand", "Made for Tots")]},
    {"label": "Aptamil Nutribiotik 2 / businesstech (NCC ZA)", "match": [("URL", "businesstech.co.za"), ("Source", "NCC (ZA)")]},
    {"label": "RCL Foods pet food / htxt (NCC ZA)", "match": [("URL", "htxt.co.za"), ("Source", "NCC (ZA)")]},
    {"label": "Aptamil Nutribiotik 2 / timelessnews (NCC ZA)", "match": [("URL", "timelessnews.co.za"), ("Source", "NCC (ZA)")]},
    {"label": "ALULA Gold / unotv (COFEPRIS MX)", "match": [("URL", "unotv.com"), ("Source", "COFEPRIS (MX)")]},
]

# New rows to append (idempotency: skip if URL already in sheet)
NEW_ROWS_TO_APPEND = [
    {
        "Date": "2026-01-16",
        "Source": "RappelConso (FR)",
        "Company": "AREV",
        "Brand": "AREV",
        "Product": "FIGUES SECHEES LERIDA 500g",
        "Pathogen": "Aflatoxin B1, total aflatoxins, tenuazonic acid (Alternaria)",
        "Reason": "Dépassement des limites maximales pour les mycotoxines: Aflatoxine B1, aflatoxines totales et acide ténuazonique",
        "Class": "Class I",
        "Country": "France",
        "Region": "Europe",
        "Tier": 1,
        "Outbreak": 0,
        "URL": "https://rappel.conso.gouv.fr/fiche-rappel/20978/Interne",
        "Notes": "[audit-2026-04-28: backfilled — missed by old hazard whitelist before mycotoxin patch (Alternaria/tenuazonic added to scope today)]",
    },
]

AUDIT_FLAG_TEXT = "[audit-2026-04-28: news-aggregator URL — needs official agency URL via url_resurrect]"
AUDIT_FLAG_MARKER = "audit-2026-04-28"  # idempotency check


def find_row(ws, headers, match_conditions):
    """Find the first row index (1-based) matching all conditions, or None."""
    col_idx = {h: i + 1 for i, h in enumerate(headers)}
    for row_num in range(2, ws.max_row + 1):
        match = True
        for col, needle in match_conditions:
            if col not in col_idx:
                match = False
                break
            cell_val = ws.cell(row=row_num, column=col_idx[col]).value
            if cell_val is None:
                match = False
                break
            if needle not in str(cell_val):
                match = False
                break
        if match:
            return row_num
    return None


def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument("--apply", action="store_true",
                   help="Write changes (default: dry-run)")
    args = p.parse_args()

    if not XLSX_PATH.exists():
        print(f"FAIL: {XLSX_PATH} not found")
        return 1

    wb = load_workbook(XLSX_PATH)
    ws = wb["Recalls"]
    headers = [c.value for c in ws[1]]
    col_idx = {h: i + 1 for i, h in enumerate(headers)}
    today = date.today().isoformat()

    print(f"Loaded {ws.max_row - 1} rows from {XLSX_PATH}")
    print()

    changes = 0
    last_updated_idx = col_idx.get("LastUpdated")

    # 0. Delete unsalvageable rows FIRST so subsequent steps work on clean data
    print("=== ROW DELETIONS ===")
    rows_to_delete = []  # collect 1-based row numbers, delete in reverse order at the end
    for deletion in ROW_DELETIONS:
        row_num = find_row(ws, headers, deletion["match"])
        if row_num is None:
            print(f"  [{deletion['label']}]: SKIP — no row matches (already deleted?)")
            continue
        # Show what we're about to delete so user can sanity-check
        co = ws.cell(row=row_num, column=col_idx.get("Company", 1)).value
        url = ws.cell(row=row_num, column=col_idx.get("URL", 1)).value
        print(f"  [{deletion['label']}] — row {row_num}:")
        print(f"    Company: {str(co)[:80]!r}")
        print(f"    URL:     {str(url)[:120]!r}")
        print(f"    DELETE")
        rows_to_delete.append(row_num)
        changes += 1

    # Defer actual deletion until after all the fixes/flags below, otherwise
    # row numbers shift mid-script. We do it AFTER everything else.

    # 1. Apply per-row content fixes
    print("=== ROW FIXES (URL-matched) ===")
    for fix in ROW_FIXES:
        row_num = find_row(ws, headers, fix["match"])
        if row_num is None:
            print(f"  [{fix['label']}]: SKIP — no row matches conditions {fix['match']}")
            continue

        # Idempotency check: are all updates already applied?
        already_applied = all(
            str(ws.cell(row=row_num, column=col_idx[col]).value) == new_val
            for col, new_val in fix["updates"].items()
            if col in col_idx
        )
        if already_applied:
            print(f"  [{fix['label']}]: row {row_num} — already applied (skip)")
            continue

        print(f"  [{fix['label']}] — row {row_num}:")
        for col, new_val in fix["updates"].items():
            if col not in col_idx:
                continue
            cell = ws.cell(row=row_num, column=col_idx[col])
            old_val = cell.value
            print(f"    {col}: {str(old_val)[:80]!r}  →  {new_val!r}")
            cell.value = new_val
            changes += 1
        if last_updated_idx:
            ws.cell(row=row_num, column=last_updated_idx, value=today)

    # 2. Flag rows for url_resurrect (idempotent via marker check)
    print()
    print("=== URL_RESURRECT FLAGS ===")
    notes_idx = col_idx.get("Notes")
    if not notes_idx:
        print("  WARN: no Notes column — skipping flag step")
    else:
        for flag in URL_RESURRECT_FLAGS:
            row_num = find_row(ws, headers, flag["match"])
            if row_num is None:
                print(f"  [{flag['label']}]: SKIP — no match")
                continue
            existing = str(ws.cell(row=row_num, column=notes_idx).value or "")
            if AUDIT_FLAG_MARKER in existing:
                print(f"  [{flag['label']}]: row {row_num} — already flagged (skip)")
                continue
            new_notes = (existing + " " + AUDIT_FLAG_TEXT).strip()[:500]
            ws.cell(row=row_num, column=notes_idx, value=new_notes)
            if last_updated_idx:
                ws.cell(row=row_num, column=last_updated_idx, value=today)
            print(f"  [{flag['label']}]: row {row_num} — flagged")
            changes += 1

    # 3. Append new rows (idempotent via URL check)
    print()
    print("=== APPEND NEW ROWS ===")
    url_idx = col_idx.get("URL")
    existing_urls = set()
    if url_idx:
        for row_num in range(2, ws.max_row + 1):
            v = ws.cell(row=row_num, column=url_idx).value
            if v:
                existing_urls.add(str(v).strip())

    appended = 0
    date_added_idx = col_idx.get("DateAdded")
    for new_row in NEW_ROWS_TO_APPEND:
        target_url = new_row.get("URL", "").strip()
        if target_url and target_url in existing_urls:
            print(f"  SKIP (URL already present): {new_row.get('Product', '')[:50]}")
            continue
        new_row_num = ws.max_row + 1
        for col_name, val in new_row.items():
            if col_name in col_idx:
                ws.cell(row=new_row_num, column=col_idx[col_name], value=val)
        if date_added_idx:
            ws.cell(row=new_row_num, column=date_added_idx, value=today)
        if last_updated_idx:
            ws.cell(row=new_row_num, column=last_updated_idx, value=today)
        print(f"  APPENDED row {new_row_num}: {new_row.get('Source')} | {new_row.get('Product', '')[:50]} | {new_row.get('Pathogen', '')[:40]}")
        appended += 1
        changes += 1

    print()
    print(f"Grand total: {changes} changes ({appended} new rows appended)")

    # Execute deferred deletions (in reverse order so earlier indices stay valid)
    if rows_to_delete:
        print()
        print(f"=== EXECUTING {len(rows_to_delete)} DELETION(S) ===")
        for row_num in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(row_num, 1)
            print(f"  Deleted row {row_num}")

    if not args.apply:
        print("\nDry run — re-run with --apply to write.")
        return 0

    wb.save(XLSX_PATH)
    print(f"\nSaved {XLSX_PATH}")
    print("Don't forget to commit + push.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
