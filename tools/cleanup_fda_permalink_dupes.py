#!/usr/bin/env python3
"""
One-off cleanup: remove the FDA "voluntary-recall?permalink=" duplicate
row from docs/data/recalls.xlsx.

Targets only rows whose URL matches /voluntary-recall?permalink=… AND
whose (source, company, pathogen) appears elsewhere in Recalls — i.e.
confirmed wrapper duplicates of canonical recalls already in the sheet.

Run from repo root:
    python tools/cleanup_fda_permalink_dupes.py --apply

Without --apply, prints the rows that WOULD be removed (dry-run).
"""
import argparse
import sys
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
from pipeline.merge_master import _near_dup_key  # noqa: E402

XLSX_PATH = ROOT / "docs" / "data" / "recalls.xlsx"
WRAPPER_MARKER = "/safety/recalls-market-withdrawals-safety-alerts/voluntary-recall?permalink="


def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument("--apply", action="store_true",
                   help="Actually delete rows (default: dry-run)")
    args = p.parse_args()

    if not XLSX_PATH.exists():
        print(f"FAIL: {XLSX_PATH} not found")
        return 1

    wb = load_workbook(XLSX_PATH)
    if "Recalls" not in wb.sheetnames:
        print("FAIL: no Recalls sheet")
        return 1
    ws = wb["Recalls"]
    hdr = [c.value for c in ws[1]]
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        rows.append((i, dict(zip(hdr, row))))

    # Build a set of near_dup_keys with at least one CANONICAL (non-wrapper) URL
    has_canonical = set()
    for _, r in rows:
        url = str(r.get("URL", "") or "")
        if WRAPPER_MARKER in url:
            continue
        k = _near_dup_key(r)
        if k:
            has_canonical.add(k)

    # Find wrapper rows whose key has a canonical sibling
    to_delete = []
    for excel_idx, r in rows:
        url = str(r.get("URL", "") or "")
        if WRAPPER_MARKER not in url:
            continue
        k = _near_dup_key(r)
        if k in has_canonical:
            to_delete.append((excel_idx, r))

    if not to_delete:
        print("No wrapper duplicates found. Nothing to do.")
        return 0

    print(f"Found {len(to_delete)} wrapper duplicate row(s):")
    for excel_idx, r in to_delete:
        row_url = str(r.get("URL", "") or "")
        print(f"  Row {excel_idx}: {r.get('Date')} | {r.get('Company')} | "
              f"{r.get('Pathogen')} | URL={row_url[:90]}")

    if not args.apply:
        print("\nDry run — re-run with --apply to actually delete.")
        return 0

    # Delete from bottom-up to preserve indices
    for excel_idx, _ in sorted(to_delete, key=lambda x: -x[0]):
        ws.delete_rows(excel_idx)

    wb.save(XLSX_PATH)
    print(f"\nDeleted {len(to_delete)} row(s). Saved {XLSX_PATH}.")
    print("Don't forget to commit + push.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
