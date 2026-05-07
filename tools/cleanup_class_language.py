"""Backfill English-only Class values into docs/data/recalls.xlsx.

The scrapers were patched 2026-05-07 to translate non-English Class values
to canonical English short forms before emit. This tool retrofits the same
mapping to rows already promoted into Recalls (and any sitting in Pending),
so the xlsx matches the operator rule that every field except Company and
Brand is English (US).

Idempotent: re-running the script after a clean pass is a no-op.

Usage
-----
    # Dry run — show what would change, write nothing
    python -m tools.cleanup_class_language --dry-run

    # Apply in place (default path: docs/data/recalls.xlsx)
    python -m tools.cleanup_class_language --apply

    # Apply against a specific file
    python -m tools.cleanup_class_language --apply --path /tmp/recalls.xlsx
"""
from __future__ import annotations

import argparse
import sys
from collections import Counter
from pathlib import Path

import openpyxl

# Reuse the canonical translator from the model layer so this tool can
# never drift from the live scraper output.
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from scrapers._models import _normalize_class_language  # noqa: E402

DEFAULT_PATH = Path(__file__).resolve().parents[1] / "docs" / "data" / "recalls.xlsx"
SHEETS_TO_CLEAN = ("Recalls", "Pending")


def _find_class_col(header_row) -> int:
    """Return the 1-indexed column number for the 'Class' header."""
    for cell in header_row:
        if (cell.value or "").strip().lower() == "class":
            return cell.column
    raise RuntimeError("'Class' column not found in header row")


def cleanup(path: Path, apply: bool) -> int:
    if not path.exists():
        print(f"ERROR: file not found: {path}", file=sys.stderr)
        return 2

    wb = openpyxl.load_workbook(path)
    total_changes = 0
    summary: Counter = Counter()

    for sheet_name in SHEETS_TO_CLEAN:
        if sheet_name not in wb.sheetnames:
            print(f"  [{sheet_name}] sheet not present — skipping")
            continue

        ws = wb[sheet_name]
        class_col = _find_class_col(next(ws.iter_rows(min_row=1, max_row=1)))

        sheet_changes = 0
        for row in ws.iter_rows(min_row=2):
            cell = row[class_col - 1]
            old = cell.value
            if not isinstance(old, str) or not old.strip():
                continue
            new = _normalize_class_language(old)
            if new != old:
                summary[(old.strip(), new)] += 1
                sheet_changes += 1
                if apply:
                    cell.value = new

        total_changes += sheet_changes
        print(f"  [{sheet_name}] {sheet_changes} row(s) need translation")

    print()
    if summary:
        print("Translation summary:")
        for (old, new), n in sorted(summary.items(), key=lambda kv: -kv[1]):
            print(f"  {n:4d}   {old!r:50s} -> {new!r}")
    else:
        print("Nothing to translate — file already clean.")

    if apply and total_changes:
        wb.save(path)
        print(f"\nWrote: {path}")
    elif not apply and total_changes:
        print(f"\n(dry-run — re-run with --apply to write changes to {path})")

    return 0


def main(argv=None) -> int:
    p = argparse.ArgumentParser(description=__doc__.split("\n", 1)[0])
    g = p.add_mutually_exclusive_group(required=True)
    g.add_argument("--dry-run", action="store_true", help="report only")
    g.add_argument("--apply", action="store_true", help="write changes")
    p.add_argument("--path", default=str(DEFAULT_PATH),
                   help=f"xlsx path (default: {DEFAULT_PATH})")
    args = p.parse_args(argv)
    return cleanup(Path(args.path), apply=args.apply)


if __name__ == "__main__":
    raise SystemExit(main())
