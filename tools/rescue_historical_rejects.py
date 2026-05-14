#!/usr/bin/env python3
"""
rescue_historical_rejects.py
============================

Walk the git history of docs/data/recalls.xlsx and extract every row
that ever appeared in the Weekly_Rejected sheet, even rows that have
since been wiped by the weekly-review-wipe.yml workflow.

How it works
------------
1. List every commit that ever touched docs/data/recalls.xlsx
   (newest → oldest, capped at --max-commits to bound runtime).
2. For each commit, run ``git show <sha>:<xlsx_path>`` to extract the
   xlsx blob at that commit, open it in-memory with openpyxl, read the
   Weekly_Rejected sheet.
3. Deduplicate rows across commits using URL as the primary key
   (composite fallback when URL is empty).
4. For each unique reject, track first_seen / last_seen commit + date,
   so we can audit the rescue.
5. Write all unique rescued rejects to a new xlsx file with the same
   column structure as the live Weekly_Rejected sheet, plus 4
   provenance columns.

Output xlsx is consumed by the v2.2 training exporter as an additional
source of REJECT examples (alongside the live Weekly_Rejected sheet).

Usage:
    python rescue_historical_rejects.py \\
        --xlsx-path docs/data/recalls.xlsx \\
        --out tools/training/data/historical_rejects.xlsx \\
        [--max-commits 500] [--sheet Weekly_Rejected]
"""
from __future__ import annotations

import argparse
import io
import subprocess
import sys
from collections import Counter
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl


# ─── Git plumbing ─────────────────────────────────────────────────────────

def run_git(args: List[str], capture_bytes: bool = False) -> Any:
    """Run a git command. Returns stdout. Raises on nonzero exit unless
    the command is git-show (which fails for paths that didn't exist at
    that commit — we treat that as "no data, skip")."""
    cmd = ["git"] + args
    if capture_bytes:
        r = subprocess.run(cmd, capture_output=True, check=False)
        return r.stdout if r.returncode == 0 else None
    r = subprocess.run(cmd, capture_output=True, text=True, check=False)
    if r.returncode != 0:
        raise RuntimeError(f"git {' '.join(args)} failed: {r.stderr[:500]}")
    return r.stdout


def list_commits_touching(xlsx_path: str,
                          max_commits: int) -> List[Tuple[str, str]]:
    """Return list of (sha, ISO date) for every commit that touched xlsx_path,
    newest first. Capped at max_commits."""
    out = run_git([
        "log",
        f"--max-count={max_commits}",
        "--format=%H|%cI",
        "--", xlsx_path,
    ])
    commits = []
    for line in out.strip().splitlines():
        if "|" not in line:
            continue
        sha, date = line.split("|", 1)
        commits.append((sha, date))
    return commits


def read_xlsx_blob_at_commit(commit_sha: str,
                             xlsx_path: str,
                             sheet_name: str
                             ) -> Tuple[Optional[List[str]],
                                        Optional[List[Dict[str, Any]]]]:
    """Read a sheet from xlsx_path at a specific commit. Returns
    (headers, rows) or (None, None) if the file or sheet doesn't exist
    at that commit."""
    blob = run_git(["show", f"{commit_sha}:{xlsx_path}"], capture_bytes=True)
    if not blob:
        return None, None
    try:
        wb = openpyxl.load_workbook(io.BytesIO(blob),
                                     read_only=True, data_only=True)
    except Exception:
        return None, None
    if sheet_name not in wb.sheetnames:
        return None, None
    ws = wb[sheet_name]
    headers = [c.value for c in ws[1]]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        row = {h: ("" if v is None else v)
               for h, v in zip(headers, r) if h}
        if any(str(v).strip() for v in row.values()):
            rows.append(row)
    return headers, rows


# ─── Row identity / dedup ─────────────────────────────────────────────────

def row_key(row: Dict[str, Any]) -> tuple:
    """Stable identity key for dedup. URL is the primary anchor; if URL
    is empty, fall back to a composite of source + company + date + product."""
    url = str(row.get("URL", "")).strip()
    if url:
        return ("url", url)
    return (
        "composite",
        str(row.get("Source", ""))[:60].strip(),
        str(row.get("Company", ""))[:60].strip(),
        str(row.get("Date", ""))[:20].strip(),
        str(row.get("Product", ""))[:80].strip(),
    )


# ─── Main ─────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx-path", default="docs/data/recalls.xlsx",
                    help="Path to xlsx within the repo (not absolute)")
    ap.add_argument("--out", type=Path, required=True,
                    help="Output xlsx path")
    ap.add_argument("--max-commits", type=int, default=500,
                    help="Cap on commits to walk (newest first)")
    ap.add_argument("--sheet", default="Weekly_Rejected",
                    help="Sheet name to rescue")
    args = ap.parse_args()

    # ── List commits ──
    print(f"Listing commits that touched {args.xlsx_path}...")
    commits = list_commits_touching(args.xlsx_path, args.max_commits)
    print(f"Found {len(commits)} commits (newest first, capped at {args.max_commits})")
    if not commits:
        print("ERROR: no commits found. Ensure the workflow used fetch-depth: 0.")
        sys.exit(1)

    # ── Walk history, dedupe ──
    rescued: Dict[tuple, Dict[str, Any]] = {}
    unified_headers: Optional[List[str]] = None
    n_with_sheet = 0
    n_rows_seen = 0

    for i, (sha, date) in enumerate(commits, 1):
        if i % 25 == 0:
            print(f"  [{i}/{len(commits)}] commits processed | "
                  f"{n_rows_seen} row-occurrences | "
                  f"{len(rescued)} unique rejects")
        try:
            headers, rows = read_xlsx_blob_at_commit(
                sha, args.xlsx_path, args.sheet)
        except Exception as e:
            print(f"  [warn] failed to read {sha[:8]}: {e}")
            continue
        if headers is None:
            continue
        n_with_sheet += 1
        if unified_headers is None or len(headers) > len(unified_headers):
            unified_headers = headers

        for row in rows:
            n_rows_seen += 1
            key = row_key(row)
            if key in rescued:
                rec = rescued[key]
                # date is ISO 8601 — string comparison works
                if date < rec["first_seen_date"]:
                    rec["first_seen_date"] = date
                    rec["first_seen_sha"] = sha
                if date > rec["last_seen_date"]:
                    rec["last_seen_date"] = date
                    rec["last_seen_sha"] = sha
                    # Prefer the newer snapshot of the row (most recent
                    # values for any fields that may have been updated)
                    rec["row"] = row
            else:
                rescued[key] = {
                    "row": row,
                    "first_seen_date": date,
                    "first_seen_sha": sha,
                    "last_seen_date": date,
                    "last_seen_sha": sha,
                }

    print()
    print(f"════════════════════════════════════════════")
    print(f"Commits with {args.sheet} sheet present: {n_with_sheet}/{len(commits)}")
    print(f"Total row-occurrences seen across history:  {n_rows_seen}")
    print(f"Unique rescued rejects after dedup:         {len(rescued)}")

    if not rescued:
        print("Nothing to write.")
        return

    # ── Compare to live Weekly_Rejected ──
    live_xlsx = Path(args.xlsx_path)
    n_already_live = 0
    n_only_historical = 0
    if live_xlsx.exists():
        try:
            live_wb = openpyxl.load_workbook(live_xlsx,
                                              read_only=True, data_only=True)
            if args.sheet in live_wb.sheetnames:
                live_ws = live_wb[args.sheet]
                live_headers = [c.value for c in live_ws[1]]
                live_keys = set()
                for r in live_ws.iter_rows(min_row=2, values_only=True):
                    row = {h: ("" if v is None else v)
                           for h, v in zip(live_headers, r) if h}
                    if any(str(v).strip() for v in row.values()):
                        live_keys.add(row_key(row))
                n_already_live = sum(1 for k in rescued if k in live_keys)
                n_only_historical = len(rescued) - n_already_live
                print(f"Of which already in live {args.sheet}:     {n_already_live}")
                print(f"NEW rejects only in git history:            {n_only_historical}")
        except Exception as e:
            print(f"(could not compare to live sheet: {e})")

    print(f"════════════════════════════════════════════")
    print()

    # ── Build output xlsx ──
    args.out.parent.mkdir(parents=True, exist_ok=True)
    out_wb = openpyxl.Workbook()
    ws = out_wb.active
    ws.title = "Historical_Rejected"

    extra_cols = [
        "_first_seen_commit", "_first_seen_date",
        "_last_seen_commit", "_last_seen_date",
    ]
    out_headers = list(unified_headers or []) + extra_cols
    ws.append(out_headers)

    # Write rows sorted by last_seen_date desc so newest are at top
    sorted_recs = sorted(
        rescued.values(),
        key=lambda r: r["last_seen_date"],
        reverse=True,
    )
    for rec in sorted_recs:
        row_values = [rec["row"].get(h, "") for h in (unified_headers or [])]
        row_values += [
            rec["first_seen_sha"][:8],
            rec["first_seen_date"],
            rec["last_seen_sha"][:8],
            rec["last_seen_date"],
        ]
        ws.append(row_values)

    out_wb.save(args.out)
    print(f"Wrote {args.out} ({len(rescued)} rows)")

    # ── Breakdown by year-month ──
    by_month: Counter = Counter()
    for rec in rescued.values():
        ym = rec["first_seen_date"][:7]
        by_month[ym] += 1
    print()
    print("Rescued rejects by month of first appearance:")
    for ym, n in sorted(by_month.items()):
        print(f"  {ym}: {n}")


if __name__ == "__main__":
    main()
