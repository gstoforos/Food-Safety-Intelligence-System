#!/usr/bin/env python3
"""Reconstruct the Weekly_Review sheet from the Recalls sheet.

WHY THIS IS NEEDED
------------------
Weekly_Review mirrors Pending -> Recalls promotions so the Thursday 17:00
mailer has something to send. On 2026-08-27 the sheet was found absent from
docs/data/recalls.xlsx and docs/data/weekly-review-latest.json was frozen at
2026-08-13T14:40:19Z with row_count 0, while 85 rows had been promoted in the
14-20 Aug window alone. The Thursday email consequently reported "0 recalls
added" for a week that had 33.

The mirror is recoverable because `DateAdded` on the Recalls sheet records
when each row was promoted. This tool rebuilds the sheet for one Thursday
window from that column and refreshes the JSON slice the mailer reads.

CAVEAT, STATED PLAINLY
----------------------
`DateAdded` is a DATE, with no time of day. The live cutoff is Thursday
17:00 Athens. A row promoted between 17:00 and midnight on a Thursday
belongs to the following week's email, and this tool cannot tell it from one
promoted that morning. Such rows are assigned to the window whose Thursday
they fall on and are flagged in the output. Expect a handful per rebuild;
check them by eye rather than trusting the reconstruction.

This is a repair tool. It does not replace weekly_review_capture, which
remains the only correct path for live promotions.

    python -m tools.rebuild_weekly_review --week-end 2026-08-27
    python -m tools.rebuild_weekly_review --week-end 2026-08-27 --write
"""
from __future__ import annotations

import argparse
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from pipeline.weekly_review_capture import (  # noqa: E402
    RECALLS_COLS, SHEET_COLS, SHEET_NAME, _ensure_sheet, export_week_slice,
)

XLSX = ROOT / "docs" / "data" / "recalls.xlsx"


def window_for(week_end: date) -> tuple[date, date]:
    """Promotion dates belonging to the email closing on `week_end`.

    The live rule is a 17:00 Athens cutoff on consecutive Thursdays. With
    date-only DateAdded the best available approximation is the closed
    interval (previous Thursday, this Thursday].
    """
    if week_end.weekday() != 3:
        raise SystemExit(f"--week-end must be a Thursday; {week_end} is a "
                         f"{week_end.strftime('%A')}")
    return week_end - timedelta(days=6), week_end


def collect(xlsx: Path, lo: date, hi: date) -> tuple[list[dict], list[dict]]:
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb["Recalls"]
    headers = [str(c.value or "") for c in ws[1]]
    try:
        i_added = headers.index("DateAdded")
    except ValueError:
        raise SystemExit("Recalls sheet has no DateAdded column; the mirror "
                         "cannot be reconstructed from it")
    rows, boundary = [], []
    for r in ws.iter_rows(min_row=2, values_only=True):
        v = r[i_added]
        if v is None:
            continue
        d = v.date() if isinstance(v, datetime) else v
        if isinstance(d, str):
            try:
                d = date.fromisoformat(d[:10])
            except ValueError:
                continue
        if not isinstance(d, date):
            continue
        if lo < d <= hi:
            rec = {h: r[k] for k, h in enumerate(headers) if h in RECALLS_COLS}
            rows.append(rec)
            # A Thursday promotion straddles the 17:00 cutoff and cannot be
            # placed from a date alone.
            if d.weekday() == 3:
                boundary.append(rec)
    return rows, boundary


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--week-end", required=True,
                    help="ISO date of the closing Thursday, e.g. 2026-08-27")
    ap.add_argument("--xlsx", default=str(XLSX))
    ap.add_argument("--write", action="store_true",
                    help="write the sheet and refresh the JSON slice "
                         "(default: report only)")
    a = ap.parse_args(argv)

    week_end = date.fromisoformat(a.week_end)
    lo, hi = window_for(week_end)
    xlsx = Path(a.xlsx)
    rows, boundary = collect(xlsx, lo, hi)

    print(f"window  {lo.isoformat()} (exclusive) -> {hi.isoformat()} (inclusive)")
    print(f"rows    {len(rows)} promotions found on the Recalls sheet")
    if boundary:
        print(f"WARNING {len(boundary)} row(s) promoted on a Thursday — the "
              f"17:00 cutoff cannot be applied from a date alone. Check these "
              f"by eye:")
        for r in boundary[:10]:
            print(f"          {r.get('Date')}  {str(r.get('Product'))[:56]}")

    wb = openpyxl.load_workbook(xlsx)
    existed = SHEET_NAME in wb.sheetnames
    print(f"sheet   {SHEET_NAME} {'exists' if existed else 'ABSENT — will be created'}")

    if not a.write:
        print("\n(dry run — pass --write to apply)")
        return 0

    ws = _ensure_sheet(wb)
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)
    for r in rows:
        ws.append([r.get(c, "") for c in RECALLS_COLS] + [week_end.isoformat(), "N"])

    ordered = [s for s in ("Recalls", "Pending") if s in wb.sheetnames]
    ordered += [s for s in wb.sheetnames if s not in ordered and s != "NEWS"]
    if "NEWS" in wb.sheetnames:
        ordered.append("NEWS")
    wb._sheets = [wb[s] for s in ordered]
    wb.save(xlsx)

    meta = export_week_slice(xlsx_path=xlsx, week_end=week_end.isoformat())
    print(f"\nwrote   {SHEET_NAME}: {len(rows)} rows")
    print(f"        weekly-review-latest.json: row_count="
          f"{meta.get('row_count') if isinstance(meta, dict) else '?'}")
    print("\nCommit BOTH docs/data/recalls.xlsx AND "
          "docs/data/weekly-review-latest.json — the mailer reads the JSON, "
          "not the sheet.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
